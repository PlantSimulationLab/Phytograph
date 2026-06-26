from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime
from pydantic import BaseModel, model_validator
import pandas as pd
import numpy as np
import io
import json
from typing import Optional, List, Dict, Any, Literal, Sequence, Union, Iterable, Tuple
from scipy.optimize import curve_fit, minimize
import re
import math
import os
import sys
import time
import subprocess
from pathlib import Path
from pytexit import py2tex

# ==================== PyHelios source submodule ====================
#
# PyHelios is vendored as a git submodule at <repo>/pyhelios (with its nested
# helios-core C++ submodule) so we can co-develop it. The native library
# (libhelios) is compiled from source by scripts/build-pyhelios.mjs and lives at
# pyhelios/pyhelios_build/build/lib/. There is no pip wheel fallback.
#
# This block runs at import time, before any endpoint lazily imports pyhelios:
#   1. put the submodule on sys.path so `import pyhelios` resolves to the source;
#   2. auto-rebuild the native library if it's missing or stale (a C++/header
#      file under helios-core/ or native/ is newer than the compiled lib).
# In packaged (PyInstaller) builds the submodule and build script aren't present;
# pyhelios is bundled via collect-all instead, so we skip the whole block.
_PYHELIOS_SRC = Path(__file__).resolve().parent.parent / "pyhelios"

if (_PYHELIOS_SRC / "pyhelios" / "__init__.py").exists():
    if str(_PYHELIOS_SRC) not in sys.path:
        sys.path.insert(0, str(_PYHELIOS_SRC))

    if sys.platform == "darwin":
        _lib_name = "libhelios.dylib"
    elif sys.platform == "win32":
        _lib_name = "libhelios.dll"
    else:
        _lib_name = "libhelios.so"
    _lib_path = _PYHELIOS_SRC / "pyhelios_build" / "build" / "lib" / _lib_name

    # A failed/required rebuild is FATAL by default, not a warning. Falling back to
    # a stale/absent libhelios was a long-lived silent footgun: after a helios-core
    # bump the auto-rebuild would fail (e.g. a missing include path) yet the backend
    # kept running native code older than the source for days, because the failure
    # was a single WARNING line and the eager import below then loaded the stale lib
    # and printed a reassuring "loaded at startup". We refuse to start instead. The
    # escape hatch (PHYTOGRAPH_ALLOW_STALE_PYHELIOS=1) downgrades fatal→loud-warning
    # for the rare case of knowingly working on non-pyhelios code with a broken build.
    _allow_stale = os.environ.get("PHYTOGRAPH_ALLOW_STALE_PYHELIOS") == "1"

    def _pyhelios_fatal(title: str, detail: str) -> None:
        bar = "=" * 74
        waived = " (waived by PHYTOGRAPH_ALLOW_STALE_PYHELIOS=1)" if _allow_stale else ""
        print(
            f"\n{bar}\n"
            f"[pyhelios] {'STALE NATIVE LIB' if _allow_stale else 'FATAL'}: {title}{waived}\n"
            f"{bar}\n{detail}\n{bar}\n"
            + ("Continuing on the EXISTING (stale/broken) libhelios because the escape\n"
               "hatch is set. Native C++ behavior may not match the current source.\n"
               if _allow_stale else
               "Refusing to start on stale/broken native code — your results would not\n"
               "reflect the current C++ source. Fix the build above, then restart. To\n"
               "run on the existing lib anyway, set PHYTOGRAPH_ALLOW_STALE_PYHELIOS=1.\n")
            + bar,
            flush=True,
        )
        if not _allow_stale:
            raise SystemExit(1)

    _needs_build = False
    if not _lib_path.exists():
        print(f"[pyhelios] native library not found at {_lib_path}", flush=True)
        _needs_build = True
    else:
        _lib_mtime = _lib_path.stat().st_mtime
        _newest_source = 0.0
        for _src_dir in (_PYHELIOS_SRC / "helios-core", _PYHELIOS_SRC / "native"):
            if _src_dir.exists():
                for _pattern in ("*.cpp", "*.hpp", "*.h"):
                    for _f in _src_dir.rglob(_pattern):
                        _newest_source = max(_newest_source, _f.stat().st_mtime)
        if _newest_source > _lib_mtime:
            print("[pyhelios] native library is stale (C++ sources are newer); rebuilding", flush=True)
            _needs_build = True

    if _needs_build:
        _build_script = _PYHELIOS_SRC.parent / "scripts" / "build-pyhelios.mjs"
        if not _build_script.exists():
            _pyhelios_fatal(
                "native rebuild required but build script is missing",
                f"Expected build-pyhelios.mjs at {_build_script}.",
            )
        else:
            print("[pyhelios] building from source (this may take a few minutes)...", flush=True)
            # Output is NOT captured: the compiler error streams live to this log so
            # it's visible above the fatal banner (and so a slow build shows progress).
            _result = subprocess.run(["node", str(_build_script)], cwd=str(_build_script.parent.parent), timeout=1800)
            if _result.returncode == 0 and _lib_path.exists():
                print("[pyhelios] build complete", flush=True)
            else:
                _pyhelios_fatal(
                    "native rebuild FAILED",
                    f"build-pyhelios.mjs exited {_result.returncode}; libhelios "
                    f"{'exists but may be stale' if _lib_path.exists() else 'is MISSING'} "
                    f"at {_lib_path}.\nThe compiler error is in the build output above.",
                )

    # Eagerly load libhelios NOW, at module import, before any endpoint imports
    # open3d. libhelios links the Homebrew OpenMP runtime; open3d (and torch /
    # sklearn) ship their own libomp.dylib, and on macOS whichever loads first
    # wins the two-level-namespace binding. If open3d loads first, libhelios
    # binds to open3d's libomp and dies on a missing symbol
    # (e.g. ___kmpc_dispatch_deinit). Importing pyhelios here makes libhelios +
    # its correct libomp bind first, so subsequent open3d use is harmless.
    #
    # A failure here is also fatal (this block is dev-only — packaged builds skip
    # it entirely since the submodule is absent). It typically means the compiled
    # lib is incompatible with the current pyhelios Python wrapper — e.g. a missing
    # FFI symbol because the lib is stale even though its mtime looked fresh — which
    # is exactly the stale-lib state we don't want to serve silently.
    try:
        import pyhelios as _pyhelios_preload  # noqa: F401
        print("[pyhelios] native library loaded at startup", flush=True)
    except Exception as _e:  # noqa: BLE001
        _pyhelios_fatal(
            "native library failed to load",
            f"import pyhelios raised: {_e!r}\n"
            "Likely a libhelios/wrapper mismatch (stale lib missing an FFI symbol).",
        )


# ── Refuse to serve PyHelios "mock mode" ──────────────────────────────────────
# `import pyhelios` succeeding is NOT sufficient: the loader and each native-
# function wrapper independently fall back to no-op MOCK stubs when the compiled
# libhelios is absent, fails to dlopen, OR is missing an FFI symbol the (newer)
# Python wrapper expects (a stale-lib mismatch). In mock mode every LiDAR / plant
# call raises at use-time or silently yields nothing — which in practice meant
# Helios scans/triangulation/LAD returned ZERO points while the app looked like it
# "worked". That is never acceptable outside unit tests: a packaged build, the dev
# backend, and ESPECIALLY E2E must fail loudly, not serve mock results.
#
# This check is unconditional (runs in dev AND packaged builds, unlike the dev-only
# auto-rebuild block above) and fails hard. The ONLY escape is PYHELIOS_ALLOW_MOCK=1,
# which the pytest harness sets for backend unit tests that don't need native code.
def _assert_pyhelios_native() -> None:
    if os.environ.get("PYHELIOS_ALLOW_MOCK") == "1":
        return  # explicit opt-in — unit tests only

    bar = "=" * 74

    def _fatal(detail: str) -> None:
        print(
            f"\n{bar}\n[pyhelios] FATAL: running in MOCK mode — native LiDAR/plant "
            f"library not usable\n{bar}\n{detail}\n{bar}\n"
            "Helios scans, triangulation, LAD, and plant generation would silently "
            "produce NO results.\nRefusing to start. Rebuild the native library:\n"
            "    node scripts/build-pyhelios.mjs        # recompile libhelios\n"
            "    npm run build:backend                  # repackage the bundle (for E2E/installers)\n"
            "If you are running backend UNIT TESTS that don't need native code, set "
            "PYHELIOS_ALLOW_MOCK=1.\n" + bar,
            flush=True,
        )
        raise SystemExit(1)

    try:
        from pyhelios.plugins import get_library_info
        from pyhelios.wrappers import ULiDARWrapper, UPlantArchitectureWrapper
    except Exception as _e:  # noqa: BLE001
        _fatal(f"could not import pyhelios native wrappers: {_e!r}")
        return

    info = get_library_info()
    if info.get("is_mock"):
        _fatal(f"library loader reports mock mode (no native lib found). info={info!r}")

    # The library can load while an individual wrapper still mocked itself because
    # the lib was missing one of that wrapper's FFI symbols (the exact stale-lib
    # failure mode). Check the wrappers the app actually depends on.
    if not getattr(ULiDARWrapper, "_LIDAR_FUNCTIONS_AVAILABLE", False):
        _fatal(
            "the LiDAR wrapper fell back to mock — libhelios is missing one or more "
            "LiDAR FFI symbols the Python wrapper expects (stale/incomplete build)."
        )
    if not getattr(UPlantArchitectureWrapper, "_PLANTARCHITECTURE_FUNCTIONS_AVAILABLE", False):
        _fatal(
            "the PlantArchitecture wrapper fell back to mock — libhelios is missing "
            "one or more plant-architecture FFI symbols (stale/incomplete build)."
        )

    print("[pyhelios] native LiDAR + plant-architecture wrappers verified (not mock)", flush=True)


_assert_pyhelios_native()

# Unicode subscript to ASCII conversion for phytorch compatibility
SUBSCRIPT_TO_ASCII = {
    '₀': '0', '₁': '1', '₂': '2', '₃': '3', '₄': '4',
    '₅': '5', '₆': '6', '₇': '7', '₈': '8', '₉': '9',
}

def unicode_to_ascii(s: str) -> str:
    """Convert unicode subscripts to ASCII numbers for phytorch compatibility."""
    result = s
    for subscript, ascii_char in SUBSCRIPT_TO_ASCII.items():
        result = result.replace(subscript, ascii_char)
    return result

# Vendored libraries live under backend-api/vendor/. Put it on sys.path so
# `from treeiso.treeiso_core import ...` resolves in dev and in the PyInstaller
# bundle (where vendor/ travels with main.py via build-backend.mjs).
_VENDOR_DIR = Path(__file__).resolve().parent / "vendor"
if str(_VENDOR_DIR) not in sys.path:
    sys.path.insert(0, str(_VENDOR_DIR))

# Backend version - bump this when making backend changes that require restart
BACKEND_VERSION = "0.43.0"

import logging
logger = logging.getLogger("phytograph")


app = FastAPI(title="Phytograph API", version="0.1.0")

# Configure CORS. The renderer's dev-server origin is now a *dynamic* port
# (scripts/dev.mjs picks a free one per session), so a fixed allowlist of
# localhost:<port> entries no longer matches — it broke the splash, which
# fetches /version cross-origin from http://localhost:<dynamic>. Allow any
# loopback origin via regex instead. This is safe: the backend only ever binds
# 127.0.0.1, so only processes on this machine can reach it regardless of CORS,
# and CORS is a browser-enforced policy, not a network boundary. The packaged
# app loads from app:///file:// (covered by allow_origins below); dev and any
# localhost tooling are covered by the regex.
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "tauri://localhost",
        "app://localhost",
        "file://",
    ],
    # http(s)://localhost:<anyport> and http(s)://127.0.0.1:<anyport>
    allow_origin_regex=r"^https?://(localhost|127\.0\.0\.1)(:\d+)?$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Centralized error logging. The ~38 per-endpoint try/except blocks already call
# traceback.print_exc() and raise HTTPException(500, str(e)); this handler is the
# safety net for anything that DOESN'T (future endpoints, errors outside a try).
# It logs one structured line + traceback through the "phytograph" logger — which
# backend_wrapper.py routes to the rotating file — so unhandled 500s are captured
# in the session log a bug report can attach. The response shape is unchanged:
# clients still receive {"detail": "..."} with status 500, same as before.
from fastapi.requests import Request
from fastapi.responses import JSONResponse
from fastapi.exception_handlers import http_exception_handler
from starlette.exceptions import HTTPException as StarletteHTTPException


@app.exception_handler(StarletteHTTPException)
async def _log_http_exception(request: Request, exc: StarletteHTTPException):
    # 4xx (validation / client) at WARNING, 5xx (our bug) at ERROR.
    level = logging.ERROR if exc.status_code >= 500 else logging.WARNING
    logger.log(level, "%s %s -> %s: %s",
               request.method, request.url.path, exc.status_code, exc.detail)
    return await http_exception_handler(request, exc)


@app.exception_handler(Exception)
async def _log_unhandled_exception(request: Request, exc: Exception):
    logger.error("Unhandled error in %s %s",
                 request.method, request.url.path, exc_info=exc)
    return JSONResponse(status_code=500, content={"detail": str(exc)})


@app.get("/")
async def root():
    return {"message": "Phytograph API is running", "version": BACKEND_VERSION}


@app.get("/health")
def health_check():
    """Health check endpoint for backend status"""
    return {"status": "healthy", "service": "Phytograph Backend", "version": BACKEND_VERSION}


@app.get("/version")
def get_version():
    """Version endpoint for Tauri app to check backend compatibility"""
    return {"version": BACKEND_VERSION}


def _gpu_name() -> "str | None":
    """Best-effort human-readable GPU name via nvidia-smi (None if unavailable)."""
    import subprocess
    try:
        r = subprocess.run(
            ["nvidia-smi", "--query-gpu=name", "--format=csv,noheader"],
            capture_output=True, text=True, timeout=5,
        )
        if r.returncode == 0:
            first = r.stdout.strip().split("\n")[0].strip()
            return first or None
    except Exception:
        pass
    return None


@app.get("/api/device-info")
def device_info():
    """Report whether synthetic-scan ray tracing runs on GPU or CPU.

    The packaged Windows/Linux builds always compile the CUDA ray-tracing path
    (the release CI fails the build otherwise), and cudart is linked statically
    so a GPU build still runs on a machine with no driver — Helios's
    cudaGetDeviceCount() returns 0 and it falls back to CPU/OpenMP. macOS builds
    are always CPU-only (no CUDA on Apple hardware). So the effective path is
    decided entirely by a runtime probe for a usable NVIDIA GPU
    (pyhelios.runtime.get_gpu_runtime_info, primarily via nvidia-smi): GPU when
    one is present on a non-macOS build, CPU otherwise.
    """
    import platform as _platform
    is_macos = _platform.system().lower() == "darwin"

    try:
        from pyhelios.runtime import get_gpu_runtime_info
        info = get_gpu_runtime_info()
    except Exception:
        info = {}

    gpu_present = bool(info.get("cuda_runtime_available"))
    gpu_count = int(info.get("cuda_device_count") or 0)
    driver_version = info.get("cuda_version")

    if is_macos:
        path, reason = "cpu", "macOS builds are CPU-only (no CUDA on Apple hardware)."
    elif gpu_present:
        path, reason = "gpu", "GPU acceleration active."
    else:
        path, reason = "cpu", "No compatible NVIDIA GPU detected; scans run on CPU/OpenMP."

    return {
        "gpu_present": gpu_present,
        "gpu_count": gpu_count,
        "gpu_name": _gpu_name() if gpu_present else None,
        "driver_version": driver_version,
        "effective_path": path,
        "reason": reason,
    }


# Model mapping
def get_model(category: str, model_type: str):
    """Get the appropriate phytorch model based on category and type"""
    from phytorch.models.stomatal import MED2011, BWB1987, BBL1995
    from phytorch.models.hydraulics import Sigmoidal, SJB2018
    from phytorch.models.generic import Linear, RectangularHyperbola, NonrectangularHyperbola

    models = {
        "stomatal": {
            "med2011": MED2011,
            "bwb1987": BWB1987,
            "bbl1995": BBL1995,
        },
        "hydraulics": {
            "sigmoidal": Sigmoidal,
            "sjb2018": SJB2018,
        },
        "generic": {
            "linear": Linear,
            "rectangular_hyperbola": RectangularHyperbola,
            "nonrectangular_hyperbola": NonrectangularHyperbola,
        }
    }

    if category not in models:
        raise ValueError(f"Unknown category: {category}")
    if model_type not in models[category]:
        raise ValueError(f"Unknown model type: {model_type}")

    return models[category][model_type]()


@app.post("/api/fit")
async def fit_model(
    file: UploadFile = File(...),
    model_category: str = Form(...),
    model_type: str = Form(...),
    method: Optional[str] = Form("auto"),
    max_iterations: Optional[int] = Form(1000)
):
    """
    Fit a phytorch model to uploaded data.
    """
    try:
        # Read the uploaded file
        contents = await file.read()

        # Parse CSV
        if file.filename.endswith('.csv') or file.filename.endswith('.txt'):
            df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")

        # Convert to dict of numpy arrays
        data = {col: df[col].values for col in df.columns}

        # Get the model
        model = get_model(model_category, model_type)

        # Fit options
        options = {
            "method": method if method else "auto",
            "max_iterations": max_iterations,
            "verbose": False
        }

        # Import and run fit
        from phytorch import fit
        result = fit(model, data, options)

        # Format response
        response = {
            "success": True,
            "parameters": {k: float(v) if isinstance(v, (np.floating, float)) else v
                         for k, v in result.parameters.items()},
            "r_squared": float(result.r_squared) if hasattr(result, 'r_squared') else None,
            "loss": float(result.loss) if hasattr(result, 'loss') else None,
            "converged": bool(result.converged) if hasattr(result, 'converged') else True,
        }

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/models")
def list_models():
    """List available models and their required data fields"""
    return {
        "photosynthesis": {
            "fvcb": {
                "name": "FvCB",
                "required_data": ["A", "Ci", "Qin", "Tleaf"],
                "optional_data": ["CurveID"]
            }
        },
        "stomatal": {
            "med2011": {
                "name": "Medlyn 2011",
                "required_data": ["A", "VPD", "gs"],
                "optional_data": ["Ca", "Gamma"]
            },
            "bwb1987": {
                "name": "Ball-Woodrow-Berry",
                "required_data": ["A", "hs", "gs"],
                "optional_data": ["Ca"]
            },
            "bbl1995": {
                "name": "Ball-Berry-Leuning",
                "required_data": ["A", "VPD", "gs"],
                "optional_data": ["Ca", "Gamma", "D0"]
            }
        },
        "hydraulics": {
            "sigmoidal": {
                "name": "Vulnerability Curve",
                "required_data": ["psi", "K"]
            },
            "sjb2018": {
                "name": "P-V Curve",
                "required_data": ["w", "psi"]
            }
        },
        "generic": {
            "linear": {
                "name": "Linear",
                "required_data": ["x", "y"]
            },
            "rectangular_hyperbola": {
                "name": "Rectangular Hyperbola",
                "required_data": ["x", "y"]
            },
            "nonrectangular_hyperbola": {
                "name": "Non-rectangular Hyperbola",
                "required_data": ["x", "y"]
            }
        }
    }


# ==================== CUSTOM FIT ENDPOINT ====================

class ParameterDef(BaseModel):
    name: str
    symbol: str
    default: float
    boundsLow: float
    boundsHigh: float
    fixed: bool = False  # If True, use fixedValue instead of fitting
    fixedValue: Optional[float] = None  # Value to use when fixed


class ConstantDef(BaseModel):
    name: str
    symbol: str
    value: float


class InputDef(BaseModel):
    name: str
    symbol: str


class OutputDef(BaseModel):
    name: str
    symbol: str


class FitRequest(BaseModel):
    model_type: str  # 'builtin' or 'user'
    model_id: Optional[str] = None  # For built-in models
    equation: Optional[str] = None  # For user models
    equation_type: str = 'explicit'  # 'explicit' or 'implicit'
    parameters: List[ParameterDef]
    constants: Optional[List[ConstantDef]] = None
    inputs: List[InputDef]
    outputs: List[OutputDef]
    column_mappings: Dict[str, str]  # variable_key -> column_name
    data: Dict[str, List[float]]  # column_name -> values
    max_iterations: int = 1000


def sanitize_equation(equation: str) -> str:
    """Clean up equation for Python evaluation"""
    # Replace common mathematical notation with Python equivalents
    eq = equation

    # Handle Unicode symbols - map to ASCII parameter names
    # This will be handled by replacing symbols with their parameter names

    # Replace common math functions
    eq = re.sub(r'\bsqrt\b', 'np.sqrt', eq)
    eq = re.sub(r'\bexp\b', 'np.exp', eq)
    eq = re.sub(r'\blog\b', 'np.log', eq)
    eq = re.sub(r'\blog10\b', 'np.log10', eq)
    eq = re.sub(r'\bsin\b', 'np.sin', eq)
    eq = re.sub(r'\bcos\b', 'np.cos', eq)
    eq = re.sub(r'\btan\b', 'np.tan', eq)
    eq = re.sub(r'\babs\b', 'np.abs', eq)
    eq = re.sub(r'\bmin\b', 'np.minimum', eq)
    eq = re.sub(r'\bmax\b', 'np.maximum', eq)

    # Handle √ symbol - √X becomes np.sqrt(X) where X is a word or parenthesized expression
    # First handle √(expr) -> np.sqrt(expr)
    eq = re.sub(r'√\(([^)]+)\)', r'np.sqrt(\1)', eq)
    # Then handle √word -> np.sqrt(word)
    eq = re.sub(r'√(\w+)', r'np.sqrt(\1)', eq)

    # Handle × and ÷
    eq = eq.replace('×', '*')
    eq = eq.replace('÷', '/')

    # Handle ** for exponentiation (already Python compatible)
    # Handle ^ for exponentiation
    eq = eq.replace('^', '**')

    return eq


def create_model_function(equation: str, param_names: List[str], input_symbols: List[str],
                         constants: Dict[str, float], symbol_to_name: Dict[str, str]) -> callable:
    """Create a callable function from an equation string"""

    # Sanitize the equation
    eq = sanitize_equation(equation)

    # Replace symbols with parameter/input names that are valid Python identifiers
    # First replace constants
    for symbol, value in constants.items():
        eq = re.sub(rf'\b{re.escape(symbol)}\b', str(value), eq)

    # Replace symbols with their variable names
    for symbol, name in symbol_to_name.items():
        eq = re.sub(rf'\b{re.escape(symbol)}\b', name, eq)

    # Build the function
    def model_func(x_data: Dict[str, np.ndarray], *params):
        # Create local namespace with numpy functions
        local_ns = {
            'np': np,
            'sqrt': np.sqrt,
            'exp': np.exp,
            'log': np.log,
            'log10': np.log10,
            'sin': np.sin,
            'cos': np.cos,
            'tan': np.tan,
            'abs': np.abs,
            'min': np.minimum,
            'max': np.maximum,
            'pi': np.pi,
            'e': np.e,
        }

        # Add input data
        for inp_symbol, inp_data in x_data.items():
            local_ns[inp_symbol] = inp_data

        # Add parameters
        for name, value in zip(param_names, params):
            local_ns[name] = value

        # Add constants
        for symbol, value in constants.items():
            local_ns[symbol] = value

        try:
            result = eval(eq, {"__builtins__": {}, "np": np}, local_ns)
            return np.array(result, dtype=float)
        except Exception as e:
            raise ValueError(f"Error evaluating equation '{eq}': {str(e)}")

    return model_func


@app.post("/api/fit/custom")
async def fit_custom_model(request: FitRequest):
    """
    Fit a custom or built-in model to data.
    """
    try:
        # Extract data arrays
        data = {k: np.array(v) for k, v in request.data.items()}

        # Get input data based on column mappings
        input_data = {}
        for inp in request.inputs:
            key = f"input_{inp.symbol}"
            col_name = request.column_mappings.get(key)
            if not col_name or col_name not in data:
                raise HTTPException(status_code=400, detail=f"Missing data for input '{inp.name}'")
            input_data[inp.symbol] = data[col_name]

        # Get output data (y values)
        y_data = None
        output_symbol = None
        for out in request.outputs:
            key = f"output_{out.symbol}"
            col_name = request.column_mappings.get(key)
            if col_name and col_name in data:
                y_data = data[col_name]
                output_symbol = out.symbol
                break

        if y_data is None:
            raise HTTPException(status_code=400, detail="Missing output data column")

        # Build constants dict
        constants = {}
        if request.constants:
            for const in request.constants:
                constants[const.symbol] = const.value

        # Build symbol to name mapping
        symbol_to_name = {}
        for inp in request.inputs:
            symbol_to_name[inp.symbol] = inp.symbol  # Keep symbol as is for eval
        for param in request.parameters:
            symbol_to_name[param.symbol] = param.name

        # Get equation - handle built-in models
        equation = request.equation
        if request.model_type == 'builtin' and request.model_id:
            # For built-in models, we use the phytorch models directly if available
            # Otherwise fall back to equation fitting
            try:
                model = get_model_builtin(request.model_id)
                if model:
                    # Use phytorch fitting for built-in models
                    from phytorch import fit as phytorch_fit

                    # Map data to expected format
                    fit_data = {}
                    for inp in request.inputs:
                        key = f"input_{inp.symbol}"
                        col_name = request.column_mappings.get(key)
                        if col_name and col_name in data:
                            fit_data[inp.name] = data[col_name]

                    for out in request.outputs:
                        key = f"output_{out.symbol}"
                        col_name = request.column_mappings.get(key)
                        if col_name and col_name in data:
                            fit_data[out.name] = data[col_name]

                    # Build initial guess and bounds from request parameters
                    # Convert unicode subscripts to ASCII (gs₀ -> gs0) for phytorch compatibility
                    initial_guess = {}
                    bounds = {}
                    for param in request.parameters:
                        ascii_symbol = unicode_to_ascii(param.symbol)
                        # Handle fixed parameters by setting bounds to (value, value)
                        if param.fixed:
                            fixed_value = param.fixedValue if param.fixedValue is not None else param.default
                            initial_guess[ascii_symbol] = fixed_value
                            bounds[ascii_symbol] = (fixed_value, fixed_value)
                        else:
                            initial_guess[ascii_symbol] = param.default
                            low = param.boundsLow if np.isfinite(param.boundsLow) else None
                            high = param.boundsHigh if np.isfinite(param.boundsHigh) else None
                            if low is not None or high is not None:
                                bounds[ascii_symbol] = (low, high)

                    result = phytorch_fit(model, fit_data, {
                        "method": "auto",
                        "max_iterations": request.max_iterations,
                        "initial_guess": initial_guess,
                        "bounds": bounds if bounds else None,
                        "verbose": False
                    })

                    # Get n_points and y_data from the first output column
                    n_points = 0
                    y_data_measured = None
                    for out in request.outputs:
                        key = f"output_{out.symbol}"
                        col_name = request.column_mappings.get(key)
                        if col_name and col_name in data:
                            y_data_measured = np.array(data[col_name])
                            n_points = len(y_data_measured)
                            break

                    # Calculate fitted values using model.forward(data, parameters)
                    fitted_values = None
                    residuals = None

                    if hasattr(model, 'forward') and y_data_measured is not None:
                        try:
                            # Phytorch forward() requires both data and parameters
                            y_fitted_tensor = model.forward(fit_data, result.parameters)
                            # Convert tensor to numpy/list
                            if hasattr(y_fitted_tensor, 'detach'):
                                fitted_values = y_fitted_tensor.detach().numpy().flatten().tolist()
                            elif hasattr(y_fitted_tensor, 'tolist'):
                                fitted_values = y_fitted_tensor.tolist()
                            else:
                                fitted_values = list(y_fitted_tensor)
                            residuals = (y_data_measured - np.array(fitted_values)).tolist()
                        except Exception:
                            pass  # fitted_values will remain None if forward fails

                    return {
                        "success": True,
                        "parameters": {k: float(v) if isinstance(v, (np.floating, float)) else v
                                     for k, v in result.parameters.items()},
                        "r_squared": float(result.r_squared) if hasattr(result, 'r_squared') else None,
                        "rmse": float(np.sqrt(result.loss)) if hasattr(result, 'loss') else None,
                        "n_points": n_points,
                        "converged": bool(result.converged) if hasattr(result, 'converged') else True,
                        "fitted_values": fitted_values,
                        "residuals": residuals,
                    }
            except Exception as e:
                # Fall back to equation-based fitting
                print(f"Falling back to equation fitting: {e}")
                pass

        if not equation:
            raise HTTPException(status_code=400, detail="No equation provided")

        # Create model function
        param_names = [p.name for p in request.parameters]
        input_symbols = [inp.symbol for inp in request.inputs]
        model_func = create_model_function(equation, param_names, input_symbols, constants, symbol_to_name)

        # Separate fixed and free parameters
        all_params = request.parameters
        free_params = [p for p in all_params if not p.fixed]
        fixed_params = [p for p in all_params if p.fixed]

        # Build mapping of parameter index to fixed value
        fixed_values = {}
        for i, p in enumerate(all_params):
            if p.fixed:
                fixed_values[i] = p.fixedValue if p.fixedValue is not None else p.default

        # Initial parameter values and bounds (only for free parameters)
        p0 = [p.default for p in free_params]
        bounds_low = [p.boundsLow if np.isfinite(p.boundsLow) else -1e10 for p in free_params]
        bounds_high = [p.boundsHigh if np.isfinite(p.boundsHigh) else 1e10 for p in free_params]

        # Wrapper for curve_fit that injects fixed parameter values
        def fit_wrapper(*args):
            # args[0] is dummy x, rest are free params only
            free_param_values = args[1:]

            # Build full parameter list with fixed values injected
            full_params = []
            free_idx = 0
            for i in range(len(all_params)):
                if i in fixed_values:
                    full_params.append(fixed_values[i])
                else:
                    full_params.append(free_param_values[free_idx])
                    free_idx += 1

            return model_func(input_data, *full_params)

        # Use dummy x array for curve_fit interface
        x_dummy = np.arange(len(y_data))

        try:
            # Check if there are any free parameters to fit
            if len(free_params) == 0:
                # All parameters are fixed, just evaluate the model
                full_params = [fixed_values[i] for i in range(len(all_params))]
                y_fitted = model_func(input_data, *full_params)
                popt_free = []
                pcov = np.array([])
            else:
                # Try curve_fit with free parameters
                popt_free, pcov = curve_fit(
                    fit_wrapper,
                    x_dummy,
                    y_data,
                    p0=p0,
                    bounds=(bounds_low, bounds_high),
                    maxfev=request.max_iterations * 10
                )

                # Build full parameter values for fitted calculation
                full_popt = []
                free_idx = 0
                for i in range(len(all_params)):
                    if i in fixed_values:
                        full_popt.append(fixed_values[i])
                    else:
                        full_popt.append(popt_free[free_idx])
                        free_idx += 1

                # Calculate fitted values and statistics
                y_fitted = model_func(input_data, *full_popt)

            # R-squared
            ss_res = np.sum((y_data - y_fitted) ** 2)
            ss_tot = np.sum((y_data - np.mean(y_data)) ** 2)
            r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0

            # RMSE
            rmse = np.sqrt(np.mean((y_data - y_fitted) ** 2))

            # Standard errors from covariance matrix (only for free params)
            try:
                std_errors_free = np.sqrt(np.diag(pcov)) if pcov.size > 0 else []
            except:
                std_errors_free = [None] * len(popt_free)

            # Build result - include both fixed and fitted parameters
            param_results = {}
            free_idx = 0
            for i, p in enumerate(all_params):
                if p.fixed:
                    param_results[p.name] = {
                        "value": float(fixed_values[i]),
                        "std_error": None,  # No error for fixed params
                        "symbol": p.symbol,
                        "fixed": True,
                    }
                else:
                    std_err = std_errors_free[free_idx] if free_idx < len(std_errors_free) else None
                    param_results[p.name] = {
                        "value": float(popt_free[free_idx]),
                        "std_error": float(std_err) if std_err is not None and np.isfinite(std_err) else None,
                        "symbol": p.symbol,
                        "fixed": False,
                    }
                    free_idx += 1

            return {
                "success": True,
                "parameters": param_results,
                "r_squared": float(r_squared),
                "rmse": float(rmse),
                "n_points": int(len(y_data)),
                "converged": True,
                "fitted_values": y_fitted.tolist(),
                "residuals": (y_data - y_fitted).tolist(),
            }

        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Fitting failed: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


def get_model_builtin(model_id: str):
    """Get a built-in phytorch model instance"""
    try:
        from phytorch.models.stomatal import MED2011, BWB1987, BBL1995
        from phytorch.models.hydraulics import Sigmoidal, SJB2018
        from phytorch.models.generic import Linear, RectangularHyperbola, NonrectangularHyperbola

        models = {
            "med2011": MED2011,
            "bwb1987": BWB1987,
            "bbl1995": BBL1995,
            "sigmoidal": Sigmoidal,
            "vulnerability": Sigmoidal,
            "sjb2018": SJB2018,
            "pvcurve": SJB2018,
            "linear": Linear,
            "rectangular_hyperbola": RectangularHyperbola,
            "nonrectangular_hyperbola": NonrectangularHyperbola,
        }

        if model_id in models:
            return models[model_id]()
        return None
    except ImportError:
        return None


# ==================== LATEX CONVERSION ====================

class LatexRequest(BaseModel):
    equation: str
    output_symbol: Optional[str] = "y"
    equation_type: str = "explicit"  # 'explicit' or 'implicit'


def convert_symbol_to_latex(symbol: str) -> str:
    """
    Convert a symbol name to LaTeX, handling Greek letters and subscripts.
    """
    # Convert the symbol through pytexit to handle Greek letters and subscripts
    try:
        latex_result = py2tex(symbol, print_latex=False, print_formula=False)
        return latex_result.strip('$')
    except:
        return symbol


def python_to_latex(equation: str, output_symbol: str = "y", equation_type: str = "explicit") -> str:
    """
    Convert a Python/numpy equation string to LaTeX using pytexit.
    Preserves the original term order.
    """
    try:
        # Clean up the equation
        eq = equation.strip()

        # Only process the first line (before any 'where:' documentation)
        if '\n' in eq:
            eq = eq.split('\n')[0].strip()

        # Replace numpy prefixes
        eq = eq.replace('np.', '')

        # Convert using pytexit - preserves order
        latex_result = py2tex(eq, print_latex=False, print_formula=False)

        # py2tex returns $$...$$ wrapped, strip that
        latex_expr = latex_result.strip('$')

        # Post-process to handle lambda workarounds (since 'lambda' is a Python keyword)
        # Replace 'lam' with '\lambda' for wavelength notation
        import re
        latex_expr = re.sub(r'\blam\b', r'\\lambda', latex_expr)
        latex_expr = re.sub(r'\blambd\b', r'\\lambda', latex_expr)

        # Convert output symbol (handles Greek letters like psi -> \psi)
        latex_output = convert_symbol_to_latex(output_symbol)

        # Build full equation with output symbol
        if equation_type == "implicit":
            latex_full = f"0 = {latex_expr}"
        else:
            latex_full = f"{latex_output} = {latex_expr}"

        return latex_full

    except Exception as e:
        # Fallback: return a simple text-based representation
        print(f"LaTeX conversion error: {e}")
        prefix = "0" if equation_type == "implicit" else output_symbol
        return f"{prefix} = \\text{{{equation}}}"


@app.post("/api/latex")
async def convert_to_latex(request: LatexRequest):
    """
    Convert a Python equation to LaTeX format.
    """
    try:
        latex = python_to_latex(
            request.equation,
            request.output_symbol,
            request.equation_type
        )
        return {"success": True, "latex": latex}
    except Exception as e:
        return {"success": False, "latex": f"{request.output_symbol} = \\text{{{request.equation}}}", "error": str(e)}


@app.get("/api/latex")
async def convert_to_latex_get(equation: str, output_symbol: str = "y", equation_type: str = "explicit"):
    """
    Convert a Python equation to LaTeX format (GET version for easy testing).
    """
    try:
        latex = python_to_latex(equation, output_symbol, equation_type)
        return {"success": True, "latex": latex}
    except Exception as e:
        return {"success": False, "latex": f"{output_symbol} = \\text{{{equation}}}", "error": str(e)}


# ==================== EXCEL EXPORT ====================

class ExportParameter(BaseModel):
    name: str
    symbol: str
    value: float
    std_error: Optional[float] = None
    fixed: bool = False
    units: Optional[str] = None
    description: Optional[str] = None

class ExportRequest(BaseModel):
    model_name: str
    model_type: Optional[str] = None
    equation: Optional[str] = None
    parameters: Dict[str, Any]
    r_squared: Optional[float] = None
    rmse: Optional[float] = None
    n_points: Optional[int] = None
    fitted_values: Optional[List[float]] = None
    residuals: Optional[List[float]] = None
    input_data: Optional[Dict[str, List[float]]] = None
    y_measured: Optional[List[float]] = None
    y_column_name: Optional[str] = None
    file_name: Optional[str] = None
    max_iterations: Optional[int] = None
    model_tags: Optional[List[str]] = None


@app.post("/api/export")
async def export_fit_results(request: ExportRequest):
    """
    Export fit results to an Excel file with Parameters, Data, and Metadata sheets.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        # ==================== PARAMETERS SHEET ====================
        ws_params = wb.active
        ws_params.title = "Parameters"
        ws_params.append(["Parameter", "Value", "Std Error", "Status", "Symbol", "Units", "Description"])

        for param_name, param_data in request.parameters.items():
            if isinstance(param_data, dict):
                value = param_data.get("value", 0)
                std_error = param_data.get("std_error")
                fixed = param_data.get("fixed", False)
                symbol = param_data.get("symbol", param_name)
                units = param_data.get("units", "")
                description = param_data.get("description", "")
            else:
                value = param_data
                std_error = None
                fixed = False
                symbol = param_name
                units = ""
                description = ""

            status = "Fixed" if fixed else "Fitted"
            ws_params.append([param_name, value, std_error, status, symbol, units, description])

        # ==================== DATA SHEET ====================
        ws_data = wb.create_sheet("Data")

        # Build data columns
        data_columns = []
        data_values = []

        # Add input variables
        if request.input_data:
            for col_name, values in request.input_data.items():
                data_columns.append(col_name)
                data_values.append(values)

        # Add measured y values (use provided column name or default)
        if request.y_measured:
            y_col_name = request.y_column_name or "Y"
            data_columns.append(f"{y_col_name}_measured")
            data_values.append(request.y_measured)

        # Add modeled/fitted values
        if request.fitted_values:
            y_col_name = request.y_column_name or "Y"
            data_columns.append(f"{y_col_name}_modeled")
            data_values.append(request.fitted_values)

        # Add residuals
        if request.residuals:
            data_columns.append("Residual")
            data_values.append(request.residuals)

        # Write header
        ws_data.append(data_columns)

        # Write data rows
        if data_values:
            n_rows = len(data_values[0]) if data_values else 0
            for i in range(n_rows):
                row = [vals[i] if i < len(vals) else None for vals in data_values]
                ws_data.append(row)

        # ==================== METADATA SHEET ====================
        ws_meta = wb.create_sheet("Metadata")

        # Header
        ws_meta.append(["Phytograph Model Fit Results"])
        ws_meta.append([])

        # Data source section
        ws_meta.append(["DATA SOURCE"])
        ws_meta.append(["Model Name", request.model_name])
        ws_meta.append(["Model Type", request.model_type or "Custom"])
        ws_meta.append(["Source File", request.file_name or "Not specified"])
        ws_meta.append(["Fit DateTime", datetime.now().isoformat()])
        if request.model_tags:
            ws_meta.append(["Tags", ", ".join(request.model_tags)])
        ws_meta.append([])

        # Equation section
        if request.equation:
            ws_meta.append(["EQUATION"])
            ws_meta.append(["Equation", request.equation])
            ws_meta.append([])

        # Fit options section
        ws_meta.append(["FIT OPTIONS"])
        ws_meta.append(["Max Iterations", request.max_iterations or "Default"])
        ws_meta.append([])

        # Error metrics section
        ws_meta.append(["ERROR METRICS"])
        ws_meta.append(["R²", request.r_squared])
        ws_meta.append(["RMSE", request.rmse])
        ws_meta.append(["N Points", request.n_points])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        safe_model_name = re.sub(r'[^\w\-]', '_', request.model_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_model_name}_FitResults_{timestamp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ==================== PROSPECT-D SPECTRAL MODEL ====================

class ProspectFitRequest(BaseModel):
    """Request model for PROSPECT-D spectral fitting"""
    wavelengths: List[float]  # Wavelength values (nm)
    reflectance: List[float]  # Reflectance values (0-1)
    # Fitting options
    vis_weight: float = 6.0  # Weight for visible band
    n_starts: int = 6  # Number of multi-start attempts
    loss: str = "soft_l1"  # Loss function: linear, soft_l1, huber, cauchy, arctan
    f_scale: float = 0.08  # Scale for robust loss
    do_calib: bool = True  # Include affine calibration (a, b)
    stage1: bool = True  # Two-stage fitting (VIS first)
    # Gaussian window parameters
    green_center: float = 550.0
    rededge_center: float = 705.0
    violet_center: float = 410.0
    window_sigma: float = 20.0
    feature_boost: float = 2.0
    violet_boost: float = 2.0
    vis_min: float = 380.0
    vis_max: float = 750.0


def build_prosail_forward():
    """Build the PROSPECT-D forward model using prosail library"""
    try:
        import prosail
    except ImportError:
        raise RuntimeError("prosail library not installed. Run: pip install prosail")

    def call_forward(N, Cab, Car, Cbrown, Cw, Cm, Ant):
        """Call prosail PROSPECT-D forward model"""
        # prosail.run_prospect uses lowercase parameter names
        f = getattr(prosail, "run_prospect", None)
        if callable(f):
            # Primary call with correct lowercase parameters for prosail library
            return f(n=N, cab=Cab, car=Car, cbrown=Cbrown, cw=Cw, cm=Cm, ant=Ant, prospect_version="D")

        # Fallback: try alternative function names
        for fname in ["prospect_D", "prospect_d"]:
            f = getattr(prosail, fname, None)
            if callable(f):
                try:
                    return f(n=N, cab=Cab, car=Car, cbrown=Cbrown, cw=Cw, cm=Cm, ant=Ant)
                except TypeError:
                    return f(N=N, Cab=Cab, Car=Car, Cbrown=Cbrown, Cw=Cw, Cm=Cm, Ant=Ant)

        raise RuntimeError("No PROSPECT-D forward function found in prosail installation.")

    def normalize_output(out):
        """Normalize prosail output to (wavelength, reflectance) arrays"""
        wl = None
        R = None

        if isinstance(out, dict):
            for k in ["R", "refl", "reflectance"]:
                if k in out:
                    R = np.asarray(out[k], dtype=float)
                    break
            for k in ["wl", "wavelengths", "lambda", "lam"]:
                if k in out:
                    wl = np.asarray(out[k], dtype=float)
                    break
        elif hasattr(out, "__len__") and not isinstance(out, (bytes, str)):
            out_list = list(out)
            if len(out_list) == 1 and hasattr(out_list[0], "__len__"):
                R = np.asarray(out_list[0], dtype=float)
            else:
                # Try to identify wavelength array (monotonic, in nm range)
                for item in out_list:
                    arr = np.asarray(item, dtype=float)
                    if arr.ndim == 1 and arr.size > 10 and np.all(np.diff(arr) > 0) and 300 <= arr.min() <= 3000:
                        wl = arr
                        break
                # Take first non-wavelength array as reflectance
                for item in out_list:
                    arr = np.asarray(item, dtype=float)
                    if arr.ndim == 1 and arr.size > 10 and not (wl is not None and np.array_equal(arr, wl)):
                        R = arr
                        break
                if R is None and len(out_list) >= 1:
                    R = np.asarray(out_list[0], dtype=float)
        else:
            R = np.asarray(out, dtype=float)

        if R is None:
            raise RuntimeError("Could not identify reflectance array from prosail output.")
        if wl is None:
            wl = np.arange(400.0, 2501.0, 1.0)

        return wl, R

    def forward_reflectance(params):
        """Run forward model and return (wavelength, reflectance)"""
        N, Cab, Car, Cbrown, Cw, Cm, Ant = params[:7].tolist()
        out = call_forward(N, Cab, Car, Cbrown, Cw, Cm, Ant)
        return normalize_output(out)

    return forward_reflectance


def resample_spectrum(meas_wl: np.ndarray, meas_R: np.ndarray, model_wl: np.ndarray):
    """Resample measured spectrum to model wavelengths"""
    wl_min = max(np.min(meas_wl), np.min(model_wl))
    wl_max = min(np.max(meas_wl), np.max(model_wl))
    mask = (model_wl >= wl_min) & (model_wl <= wl_max)
    wl_common = model_wl[mask]
    R_interp = np.interp(wl_common, meas_wl, meas_R)
    return wl_common, R_interp


def fit_prospect_d(
    wavelengths: np.ndarray,
    reflectance: np.ndarray,
    vis_weight: float = 6.0,
    loss: str = "soft_l1",
    f_scale: float = 0.08,
    n_starts: int = 6,
    stage1: bool = True,
    do_calib: bool = True,
    green_center: float = 550.0,
    rededge_center: float = 705.0,
    window_sigma: float = 20.0,
    feature_boost: float = 2.0,
    vis_min: float = 380.0,
    vis_max: float = 750.0,
    violet_center: float = 410.0,
    violet_boost: float = 2.0,
    rng_seed: int = 42
) -> Dict[str, Any]:
    """
    Fit PROSPECT-D model to spectral reflectance data.

    Returns dict with:
        - success: bool
        - parameters: dict of fitted parameter values
        - rmse: float
        - r_squared: float
        - fitted_spectrum: list of fitted reflectance values
        - wavelengths: list of wavelengths for fitted spectrum
    """
    from scipy.optimize import least_squares

    forward = build_prosail_forward()

    # Parameter setup: 7 biophysical + optional 2 calibration
    names = ["N", "Cab", "Car", "Cbrown", "Cw", "Cm", "Ant"]
    lb = np.array([1.0, 0.0, 0.0, 0.0, 0.000, 0.000, 0.0], dtype=float)
    ub = np.array([3.0, 90.0, 30.0, 1.0, 0.040, 0.020, 5.0], dtype=float)
    p0 = np.array([1.6, 45.0, 8.0, 0.05, 0.015, 0.005, 0.6], dtype=float)

    if do_calib:
        names += ["a_scale", "b_offset"]
        lb = np.concatenate([lb, np.array([0.9, -0.03])])
        ub = np.concatenate([ub, np.array([1.1, 0.03])])
        p0 = np.concatenate([p0, np.array([1.0, 0.0])])

    # Prior means and sigmas for regularization
    mu = np.array([1.6, 45.0, 8.0, 0.05, 0.015, 0.005, 0.6], dtype=float)
    sigma = np.array([0.5, 30.0, 8.0, 0.2, 0.02, 0.01, 1.5], dtype=float)
    if do_calib:
        mu = np.concatenate([mu, np.array([1.0, 0.0])])
        sigma = np.concatenate([sigma, np.array([0.05, 0.02])])

    # Generate multi-start seeds
    rng = np.random.default_rng(rng_seed)
    seeds = [p0]
    for _ in range(max(0, n_starts - 1)):
        r = rng.uniform(0.85, 1.15, size=len(p0))
        seeds.append(np.clip(p0 * r, lb, ub))

    # Spectral weighting function
    def spectral_weights(wl):
        w = np.ones_like(wl, dtype=float)
        vis = (wl >= vis_min) & (wl <= vis_max)
        w[vis] *= vis_weight
        g = np.exp(-0.5 * ((wl - green_center) / window_sigma) ** 2)
        r = np.exp(-0.5 * ((wl - rededge_center) / window_sigma) ** 2)
        v = np.exp(-0.5 * ((wl - violet_center) / window_sigma) ** 2)
        w += feature_boost * (g + r) + violet_boost * v
        return w

    # Full-spectrum objective
    def objective_full(meas_wl, meas_R):
        def f(p):
            wl_model, R_model = forward(p[:7])
            wl_c, R_meas = resample_spectrum(meas_wl, meas_R, wl_model)
            R_fit = np.interp(wl_c, wl_model, R_model)
            if do_calib:
                R_fit = p[7] * R_fit + p[8]
            res_spec = spectral_weights(wl_c) * (R_fit - R_meas)
            res_prior = (p - mu) / sigma
            return np.concatenate([res_spec, 0.1 * res_prior])
        return f

    # VIS-only objective (stage 1)
    def objective_vis(meas_wl, meas_R):
        tight = sigma.copy()
        tight[4] = min(tight[4], 1e-6)  # Cw
        tight[5] = min(tight[5], 1e-6)  # Cm
        def f(p):
            wl_model, R_model = forward(p[:7])
            wl_c, R_meas = resample_spectrum(meas_wl, meas_R, wl_model)
            R_fit = np.interp(wl_c, wl_model, R_model)
            if do_calib:
                R_fit = p[7] * R_fit + p[8]
            res_spec = spectral_weights(wl_c) * (R_fit - R_meas)
            res_prior = (p - mu) / tight
            return np.concatenate([res_spec, 0.1 * res_prior])
        return f

    # Run optimization
    best = None
    meas_wl = np.array(wavelengths)
    meas_R = np.array(reflectance)

    if stage1:
        # Two-stage: VIS first, then full
        mask = (meas_wl >= 400) & (meas_wl <= 750)
        wl_vis = meas_wl[mask]
        R_vis = meas_R[mask]
        for s in seeds:
            try:
                res1 = least_squares(
                    objective_vis(wl_vis, R_vis), s, bounds=(lb, ub),
                    method="trf", x_scale="jac", loss=loss, f_scale=f_scale, max_nfev=4000
                )
                res2 = least_squares(
                    objective_full(meas_wl, meas_R), res1.x, bounds=(lb, ub),
                    method="trf", x_scale="jac", loss=loss, f_scale=f_scale, max_nfev=6000
                )
                if best is None or res2.cost < best.cost:
                    best = res2
            except Exception as e:
                print(f"Stage1 optimization failed: {e}")
                continue
    else:
        for s in seeds:
            try:
                res = least_squares(
                    objective_full(meas_wl, meas_R), s, bounds=(lb, ub),
                    method="trf", x_scale="jac", loss=loss, f_scale=f_scale, max_nfev=6000
                )
                if best is None or res.cost < best.cost:
                    best = res
            except Exception as e:
                print(f"Single-stage optimization failed: {e}")
                continue

    if best is None:
        return {
            "success": False,
            "converged": False,
            "error": "Optimization failed for all starting points",
            "parameters": {},
            "rmse": None,
            "r_squared": None,
            "n_points": len(meas_R),
        }

    # Calculate fitted spectrum
    wl_model, R_model = forward(best.x[:7])
    wl_c, R_meas = resample_spectrum(meas_wl, meas_R, wl_model)
    R_fit = np.interp(wl_c, wl_model, R_model)
    if do_calib:
        R_fit = best.x[7] * R_fit + best.x[8]

    # Calculate metrics
    residuals = R_fit - R_meas
    rmse = float(np.sqrt(np.mean(residuals ** 2)))
    ss_res = np.sum(residuals ** 2)
    ss_tot = np.sum((R_meas - np.mean(R_meas)) ** 2)
    r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0

    # Build parameter dict
    params = {}
    param_info = {
        "N": {"units": "-", "description": "Leaf structure parameter (number of compact layers)"},
        "Cab": {"units": "μg/cm²", "description": "Chlorophyll a+b content"},
        "Car": {"units": "μg/cm²", "description": "Carotenoid content"},
        "Cbrown": {"units": "-", "description": "Brown pigment content"},
        "Cw": {"units": "cm", "description": "Equivalent water thickness"},
        "Cm": {"units": "g/cm²", "description": "Dry matter content"},
        "Ant": {"units": "μmol/cm²", "description": "Anthocyanin content"},
        "a_scale": {"units": "-", "description": "Calibration scale factor"},
        "b_offset": {"units": "-", "description": "Calibration offset"},
    }

    for i, name in enumerate(names):
        info = param_info.get(name, {"units": "", "description": ""})
        params[name] = {
            "value": float(best.x[i]),
            "symbol": name,
            "units": info["units"],
            "description": info["description"],
            "fixed": False,
        }

    # For PROSPECT inversion, always report as converged if we got results
    # The quality is indicated by RMSE and R², not convergence status

    return {
        "success": True,
        "parameters": params,
        "rmse": rmse,
        "r_squared": float(r_squared),
        "n_points": len(R_meas),
        "converged": True,
        "fitted_spectrum": R_fit.tolist(),
        "measured_spectrum": R_meas.tolist(),
        "wavelengths": wl_c.tolist(),
        "residuals": residuals.tolist(),
        "message": best.message if hasattr(best, 'message') else "",
    }


@app.post("/api/fit/prospect")
async def fit_prospect_model(request: ProspectFitRequest):
    """
    Fit PROSPECT-D radiative transfer model to spectral reflectance data.

    Input:
        - wavelengths: array of wavelength values (nm)
        - reflectance: array of reflectance values (0-1 range)
        - Various fitting options (vis_weight, n_starts, etc.)

    Output:
        - Fitted biophysical parameters (N, Cab, Car, Cbrown, Cw, Cm, Ant)
        - Optional calibration parameters (a_scale, b_offset)
        - Fitted spectrum and fit statistics
    """
    try:
        result = fit_prospect_d(
            wavelengths=np.array(request.wavelengths),
            reflectance=np.array(request.reflectance),
            vis_weight=request.vis_weight,
            loss=request.loss,
            f_scale=request.f_scale,
            n_starts=request.n_starts,
            stage1=request.stage1,
            do_calib=request.do_calib,
            green_center=request.green_center,
            rededge_center=request.rededge_center,
            window_sigma=request.window_sigma,
            feature_boost=request.feature_boost,
            vis_min=request.vis_min,
            vis_max=request.vis_max,
            violet_center=request.violet_center,
            violet_boost=request.violet_boost,
        )
        return result
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PROSPECT fitting failed: {str(e)}")


# ==================== POINT CLOUD TRIANGULATION ====================

class PointSource(BaseModel):
    """Tell a downstream endpoint to read points from a live cloud SESSION
    (in-RAM, source of truth) or — only as a fallback — a file on disk, instead
    of an inline `points` array.

    Octree-backed clouds keep no positions in the renderer (the geometry lives
    only in the on-disk Potree octree, streamed to the GPU), so skeleton /
    triangulate / c2m / icp / export resolve their points here. When `session_id`
    is set the in-RAM session array is the point of truth (deletions honored, no
    file re-read); `source_path` is then provenance only and may be empty (e.g. a
    synthetic-scan session that never had a source file). When there is no session,
    `source_path` is read from disk.

    Resolved by `_read_points_from_source` (defined later, alongside the other
    point-cloud loaders it reuses).
    """
    # Empty/None is valid when `session_id` is set (the session is the source of
    # truth and source_path is provenance only). Required only for a file source.
    source_path: Optional[str] = None
    ascii_format: Optional[str] = None
    # Stride-downsample cap. None = full resolution. Stride (not reservoir)
    # preserves spatial uniformity, which skeleton/triangulation depend on.
    max_points: Optional[int] = None
    # [tx, ty, tz] ADDED to every point — matches the renderer's getDisplayData
    # (`positions[i*3] + tx`) and the in-RAM `_region_mask` translation path.
    translation: Optional[List[float]] = None
    want_colors: bool = False
    # Sky/miss points (`is_miss != 0`) are dropped by default for a session
    # source: a miss is a ray that hit nothing, projected ~1 km out, so the
    # surface-reconstruction consumers (triangulate, skeleton, segment, QSM) must
    # never see it — exactly as the hits-only octree excludes it. Set True to
    # keep misses (the only caller that does is export, which preserves whatever
    # the user imported). Has no effect on a file-path source (those recover
    # misses downstream via LAD's own readers, not this chokepoint).
    include_misses: bool = False
    # When set, points come from a live cloud session's in-RAM array (the
    # Family-1 source of truth) with its per-point deletions already applied —
    # NOT from `source_path` on disk. This is how downstream ops honor unbaked
    # deletions without a rebuild. `source_path` stays populated for provenance
    # but is not re-read when `session_id` is present. The compute consumers of
    # this path want positions only, so the session-source branch returns
    # positions and leaves colours/intensity as None (the session DOES hold them;
    # they're simply not surfaced here).
    session_id: Optional[str] = None


class TriangulationGrid(BaseModel):
    """A voxel grid to PIN a (ball-pivot / Open3D) triangulation to, so the mesh
    can later be re-used as the external triangulation for the leaf-area (LAD)
    inversion. Same shape as ``HeliosGrid`` (defined later in the file), declared
    separately so ``TriangulationRequest`` doesn't forward-reference it.

    When set on a TriangulationRequest the backend (a) crops points to the grid's
    world AABB before meshing — so only points inside the box are triangulated —
    and (b) bins each output triangle's centroid into the grid, returning a
    per-triangle cell id. The renderer needs both: the crop confines the mesh to
    the box, and the cell ids let the LAD reuse path drop any triangle whose
    centroid still falls outside the grid (belt-and-suspenders for "only
    in-grid triangles feed the inversion")."""
    center: List[float]  # [x, y, z] world coordinates
    size: List[float]    # [x, y, z] full extents
    nx: int = 1
    ny: int = 1
    nz: int = 1
    # Azimuthal rotation about +z (degrees). The renderer crops the ROTATED box
    # (via crop_box_rotation_deg), so points are confined correctly; the cell
    # binning here is axis-aligned, but the C++ inversion re-bins each triangle
    # with a rotation-aware containment test, so a rotated grid is still correct.
    rotation: float = 0.0


class TriangulationRequest(BaseModel):
    """Request model for point cloud triangulation"""
    # Inline points for flat clouds; octree clouds send `source` instead.
    points: Optional[List[List[float]]] = None
    # Read points from a file on disk (octree-backed clouds). The renderer sets
    # `source.max_points` from the global "triangulate max points" setting to
    # bound open3d's memory on huge clouds.
    source: Optional[PointSource] = None
    # MERGED multi-scan triangulation: read every source's points and fuse them
    # into one cloud before meshing. Octree-backed clouds can't be merged on the
    # client (their in-RAM points live in the backend session, and two source
    # descriptors can't be concatenated into one request.source), so the renderer
    # sends all contributing sources here and the backend vstacks them — keeping
    # the session array as the source of truth. Takes precedence over
    # `source` / `points` when non-empty. Each source's `max_points` still caps
    # per-source; the fused total is bounded by the renderer.
    sources: Optional[List[PointSource]] = None
    method: str = "ball_pivoting"  # "ball_pivoting", "poisson", "alpha_shape", "delaunay"
    # Optional crop-to-grid box, as [min_x, min_y, min_z, max_x, max_y, max_z] in
    # world coordinates. When set, points outside the (inclusive) box are dropped
    # before meshing — a numpy mask applied to the resolved point array,
    # regardless of source kind (inline / session / merged). Used by the
    # "Crop to grid" toggle on the Ball Pivot path so only points inside a voxel
    # box get triangulated. None = no crop (mesh every resolved point).
    crop_box: Optional[List[float]] = None
    # Azimuthal rotation of the crop box about +z, in DEGREES, about the box
    # center (= the AABB midpoint of crop_box). The voxel grid the user crops to
    # can be rotated (e.g. a Helios <grid> with <rotation>), and the renderer
    # renders it rotated; cropping by the axis-aligned crop_box alone keeps the
    # box's AABB corners, so a rotated grid leaks branches past its rotated walls.
    # Matches Helios's own grid-cell test (inverse-rotate the point into the box
    # frame, then AABB), so Ball Pivot and Helios crop the identical region.
    # None / ~0 = axis-aligned (the prior behavior).
    crop_box_rotation_deg: Optional[float] = None
    # PIN this triangulation to a voxel grid so the resulting mesh can later be
    # re-used as the external triangulation for the leaf-area (LAD) inversion.
    # When set, each output triangle's centroid is binned into the grid and the
    # per-triangle cell ids ride back in the response, so the LAD reuse path can
    # keep only in-grid triangles. The renderer also sets crop_box (+
    # crop_box_rotation_deg) from the SAME grid so points outside the box are
    # dropped before meshing — `grid` here is purely the binning/echo channel,
    # not a second crop. None = not pinned (mesh isn't LAD-reusable).
    grid: Optional[TriangulationGrid] = None
    # Ball pivoting parameters
    radii: Optional[List[float]] = None  # Ball radii for ball pivoting (auto if None)
    # Poisson parameters
    depth: int = 8  # Octree depth for Poisson reconstruction
    # Alpha shape parameters
    alpha: Optional[float] = None  # Alpha value (auto if None)
    # General parameters
    estimate_normals: bool = True  # Estimate point normals if not provided
    normal_radius: float = 0.1  # Radius for normal estimation
    normal_max_nn: int = 30  # Max neighbors for normal estimation


class TriangulationResponse(BaseModel):
    """Response model for triangulation results"""
    success: bool
    vertices: List[List[float]]  # [[x, y, z], ...]
    triangles: List[List[int]]  # [[i, j, k], ...] vertex indices
    normals: Optional[List[List[float]]] = None  # Vertex normals
    surface_area: Optional[float] = None
    num_triangles: int
    num_vertices: int
    method_used: str
    error: Optional[str] = None
    # Number of input points actually triangulated. With crop-to-grid this is the
    # in-grid count, which is naturally less than the whole cloud even when no
    # downsampling happened — so don't infer "downsampled" from it; use the flag.
    points_used: Optional[int] = None
    # True iff the max_points cap actually stride-downsampled the (post-crop)
    # points. The renderer keys its "downsampled" warning toast off THIS, not off
    # points_used < cloud size (which a crop alone also makes true).
    downsampled: Optional[bool] = None
    # Per-triangle grid cell (row-major i + nx*(j + ny*k)) when the request pinned
    # the mesh to a `grid`; -1 (packed as the uint32 sentinel 0xffffffff) means the
    # centroid fell outside every cell. Aligned 1:1 with `triangles`. Empty when no
    # grid was supplied. Lets the LAD reuse path keep only in-grid triangles.
    triangle_cell_ids: List[int] = []


def _do_open3d_triangulation(request: TriangulationRequest, progress=None) -> dict:
    """
    Triangulate a point cloud to create a mesh surface.

    Useful for reconstructing leaf surfaces from LiDAR point cloud data.

    Methods:
        - ball_pivoting: Ball Pivoting Algorithm - good for clean, uniformly sampled point clouds
        - poisson: Poisson Surface Reconstruction - creates watertight meshes, good for noisy data
        - alpha_shape: Alpha Shape - creates mesh based on alpha radius, good for concave shapes
        - delaunay: 2D Delaunay triangulation projected to 3D - fast, for roughly planar surfaces
    """
    def _report(fraction, message):
        if progress is not None:
            progress(fraction, message)

    def _ckpt():
        _cancel_checkpoint(progress)

    try:
        import open3d as o3d

        _ckpt()
        _report(0.05, "Reading points")
        if len(request.crop_box or []) not in (0, 6):
            return {"success": False, "method_used": request.method,
                    "num_triangles": 0, "num_vertices": 0, "points_used": 0,
                    "error": "crop_box must be [min_x, min_y, min_z, max_x, max_y, max_z]"}
        cropping = request.crop_box is not None

        # Resolve order with crop-to-grid: the per-source `max_points` cap (the
        # "Triangulate max points" setting) stride-downsamples inside
        # `_read_points_from_source`. Applied THERE (before the crop) it thins the
        # WHOLE cloud first, throwing away in-grid density for nothing — a grid
        # holding a few million points would still be served a 5M-strided cloud.
        # So we strip the cap from the resolve, CROP first, then cap the (usually
        # far smaller) cropped set. The cap is the max of the contributing
        # sources' caps; inline-only requests carry none.
        def _resolve(src: "PointSource") -> np.ndarray:
            return _read_points_from_source(src.model_copy(update={"max_points": None}))[0]

        cap: Optional[int] = None
        if request.sources:
            parts = [_resolve(s) for s in request.sources]
            if request.points:
                parts.append(np.array(request.points, dtype=np.float64))
            parts = [p for p in parts if len(p) > 0]
            points = np.vstack(parts) if parts else np.empty((0, 3), dtype=np.float64)
            caps = [s.max_points for s in request.sources if s.max_points]
            cap = max(caps) if caps else None
        elif request.source is not None:
            points = _resolve(request.source)
            cap = request.source.max_points
        else:
            points = np.array(request.points or [], dtype=np.float64)

        # Crop to grid: drop points outside the requested box before meshing.
        # A pure numpy mask on the already-resolved array — works the same for
        # octree/session, merged, and inline sources. Applied before the
        # min-3-points check so an over-tight box reports the right error.
        if cropping and len(points) > 0:
            cb = request.crop_box
            lo = np.array(cb[:3], dtype=np.float64)
            hi = np.array(cb[3:], dtype=np.float64)
            rot_deg = request.crop_box_rotation_deg or 0.0
            if abs(float(rot_deg)) > 1e-9:
                # Inverse-rotate each point about the box center into the box's
                # local axis-aligned frame, then AABB-test — identical to Helios's
                # calculateHitGridCell (rotatePointAboutLine by -rotation about z).
                # +rotation is CCW about +z (right-handed), matching the renderer's
                # three.js mesh rotation.z and the Helios <grid> convention.
                center = (lo + hi) / 2.0
                theta = -np.radians(float(rot_deg))  # inverse rotation
                cos_t, sin_t = np.cos(theta), np.sin(theta)
                dx = points[:, 0] - center[0]
                dy = points[:, 1] - center[1]
                local = np.empty_like(points)
                local[:, 0] = cos_t * dx - sin_t * dy + center[0]
                local[:, 1] = sin_t * dx + cos_t * dy + center[1]
                local[:, 2] = points[:, 2]
                mask = np.all((local >= lo) & (local <= hi), axis=1)
            else:
                mask = np.all((points >= lo) & (points <= hi), axis=1)
            points = points[mask]

        # Apply the cap AFTER the crop. `was_downsampled` is true only when the
        # cap actually fired — NOT merely when fewer points than the whole cloud
        # remain (cropping is not downsampling). The renderer keys its warning
        # toast off this, so a crop that stays under the cap won't falsely warn.
        was_downsampled = False
        if cap and cap > 0 and len(points) > cap:
            stride = int(math.ceil(len(points) / cap))
            points = points[::stride]
            was_downsampled = True

        points_used = int(len(points))

        if len(points) < 3:
            err = ("Need at least 3 points for triangulation — the crop-to-grid box "
                   "contains too few points. Enlarge the grid box or turn off "
                   "Crop to grid.") if request.crop_box is not None else \
                  "Need at least 3 points for triangulation"
            return {"success": False, "method_used": request.method,
                    "num_triangles": 0, "num_vertices": 0, "points_used": points_used,
                    "error": err}

        # Create Open3D point cloud
        _report(0.15, "Preparing point cloud")
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(points)

        # Estimate normals if needed.
        #
        # Only ball_pivoting and poisson consume point normals — alpha_shape
        # (Delaunay tetrahedralization) and delaunay (2D projection) never read
        # pcd.normals, so estimating here would be pure overhead. Skip them.
        needs_normals = request.method in ("ball_pivoting", "poisson")
        if request.estimate_normals and needs_normals:
            _report(0.25, "Estimating normals")
            pcd.estimate_normals(
                search_param=o3d.geometry.KDTreeSearchParamHybrid(
                    radius=request.normal_radius,
                    max_nn=request.normal_max_nn
                )
            )
            # Orient the normals into a usable field. Two strategies, by method:
            #
            #  - Poisson needs a GLOBALLY consistent inside/outside field to solve
            #    its watertight indicator function, so it keeps the MST-based
            #    orient_normals_consistent_tangent_plane. That pass is O(N log N)
            #    with a brutal constant (minutes on a few-million-point cloud) and
            #    is itself uninterruptible, so checkpoint a cancel on either side.
            #
            #  - Ball Pivoting is a LOCAL surface walk: it only needs each point's
            #    normal to agree with its immediate neighbors, not a global
            #    inside/outside. The MST is solving a far harder problem than BPA
            #    needs and costs ~270s where BPA itself costs ~15s. Orienting every
            #    normal toward the cloud centroid instead is O(N) (sub-second) and,
            #    measured on a 2M-point LiDAR tree, yields a BPA mesh within ~4% of
            #    the MST's triangle count — a ~17x speedup for the same surface.
            #    Centroid is used (not a scan origin) so single, merged, and
            #    origin-less clouds all take the identical cheap path; normals are
            #    flipped to face OUTWARD (orient_normals_towards_camera_location
            #    points them toward the reference) so BPA rolls on the right side.
            _ckpt()
            if request.method == "poisson":
                pcd.orient_normals_consistent_tangent_plane(k=15)
            else:
                centroid = np.asarray(pcd.points).mean(axis=0)
                pcd.orient_normals_towards_camera_location(centroid)
                pcd.normals = o3d.utility.Vector3dVector(-np.asarray(pcd.normals))
            _ckpt()

        mesh = None
        method_used = request.method
        # Last cheap exit before the monolithic Open3D/scipy meshing call (which
        # itself can't be interrupted mid-pass); a cancel here frees the cloud.
        _ckpt()
        _report(0.45, f"Meshing ({request.method.replace('_', ' ')})")

        if request.method == "ball_pivoting":
            # Ball Pivoting Algorithm
            if not pcd.has_normals():
                # Fallback for estimate_normals=False callers. Use the same cheap
                # centroid-facing orientation as the main path above (NOT the
                # minutes-long MST) — BPA only needs locally consistent normals.
                pcd.estimate_normals()
                centroid = np.asarray(pcd.points).mean(axis=0)
                pcd.orient_normals_towards_camera_location(centroid)
                pcd.normals = o3d.utility.Vector3dVector(-np.asarray(pcd.normals))

            # Auto-compute radii if not provided.
            #
            # Use the MEDIAN nearest-neighbor distance, not the mean: it is robust
            # to the handful of sparse outliers that survive miss exclusion, so a
            # few stray points can't inflate the ball radius (a too-large radius
            # makes BPA explode combinatorially). The median reflects the real
            # surface spacing.
            if request.radii is None:
                distances = pcd.compute_nearest_neighbor_distance()
                ref_dist = float(np.median(distances))
                radii = o3d.utility.DoubleVector([ref_dist, ref_dist * 2, ref_dist * 4])
            else:
                radii = o3d.utility.DoubleVector(request.radii)

            mesh = o3d.geometry.TriangleMesh.create_from_point_cloud_ball_pivoting(pcd, radii)

        elif request.method == "poisson":
            # Poisson Surface Reconstruction
            if not pcd.has_normals():
                pcd.estimate_normals()
                pcd.orient_normals_consistent_tangent_plane(k=15)

            mesh, densities = o3d.geometry.TriangleMesh.create_from_point_cloud_poisson(
                pcd, depth=request.depth
            )

            # Remove low-density vertices (artifacts)
            _report(0.7, "Filtering low-density vertices")
            densities = np.asarray(densities)
            density_threshold = np.quantile(densities, 0.1)
            vertices_to_remove = densities < density_threshold
            mesh.remove_vertices_by_mask(vertices_to_remove)

        elif request.method == "alpha_shape":
            # Alpha Shape.
            #
            # Open3D's alpha shape builds a Delaunay tetrahedralization (Qhull)
            # and keeps the faces of tetrahedra whose circumradius <= alpha. On
            # the near-planar surfaces typical of LiDAR'd leaves, Qhull produces
            # many exactly-coplanar (zero-volume) tetrahedra. Open3D skips each
            # one — logging "[CreateFromPointCloudAlphaShape] invalid tetra in
            # TetraMesh" — which both spams the console and drops faces that
            # should have bridged the surface, leaving a sparse, holey mesh.
            #
            # Two cleanups before handing the cloud to Open3D:
            #  1. Drop exact duplicate points (overlapping-scan coincidences),
            #     which are a separate source of degenerate tetrahedra.
            #  2. Add sub-micron jitter (scaled to the cloud's extent, well below
            #     any real measurement precision) to break exact coplanarity so
            #     Qhull yields non-degenerate tetrahedra. In practice this turns
            #     a handful of triangles into full surface coverage.
            pcd = pcd.remove_duplicated_points()
            jpts = np.asarray(pcd.points)
            if len(jpts) >= 4:
                extent = float(np.linalg.norm(jpts.max(axis=0) - jpts.min(axis=0)))
                if extent > 0:
                    rng = np.random.default_rng(0)  # deterministic across runs
                    jpts = jpts + rng.normal(0.0, extent * 1e-6, jpts.shape)
                    pcd.points = o3d.utility.Vector3dVector(jpts)

            if request.alpha is None:
                # Auto-compute alpha from point spacing
                distances = pcd.compute_nearest_neighbor_distance()
                alpha = np.mean(distances) * 2
            else:
                alpha = request.alpha

            # Suppress Open3D's per-tetra warnings; any genuinely degenerate
            # tetrahedra that survive jittering are correctly skipped, not an
            # error worth surfacing.
            with o3d.utility.VerbosityContextManager(o3d.utility.VerbosityLevel.Error):
                mesh = o3d.geometry.TriangleMesh.create_from_point_cloud_alpha_shape(pcd, alpha)

        elif request.method == "delaunay":
            # 2D Delaunay triangulation (project to XY plane, then lift back)
            from scipy.spatial import Delaunay

            # Project to 2D (XY plane)
            points_2d = points[:, :2]

            try:
                tri = Delaunay(points_2d)
                triangles = tri.simplices

                # Create mesh from triangulation
                mesh = o3d.geometry.TriangleMesh()
                mesh.vertices = o3d.utility.Vector3dVector(points)
                mesh.triangles = o3d.utility.Vector3iVector(triangles)
                mesh.compute_vertex_normals()
            except Exception as e:
                return {"success": False, "method_used": method_used,
                        "num_triangles": 0, "num_vertices": 0,
                        "error": f"Delaunay triangulation failed: {str(e)}"}
        else:
            return {"success": False, "method_used": method_used,
                    "num_triangles": 0, "num_vertices": 0,
                    "error": f"Unknown method: {request.method}. Use 'ball_pivoting', 'poisson', 'alpha_shape', or 'delaunay'"}

        if mesh is None or len(mesh.triangles) == 0:
            return {"success": False, "method_used": method_used,
                    "num_triangles": 0, "num_vertices": len(points),
                    "points_used": points_used,
                    "error": "Triangulation produced no triangles. Try adjusting parameters or using a different method."}

        _ckpt()
        # Clean up mesh
        _report(0.85, "Cleaning up mesh")
        mesh.remove_degenerate_triangles()
        mesh.remove_duplicated_triangles()
        mesh.remove_duplicated_vertices()
        mesh.remove_non_manifold_edges()

        # Compute normals if not present
        if not mesh.has_vertex_normals():
            mesh.compute_vertex_normals()

        # Calculate surface area
        _report(0.95, "Computing surface area")
        surface_area = mesh.get_surface_area()

        # Extract results as numpy (the endpoint packs them straight to binary).
        vertices = np.asarray(mesh.vertices)
        triangles = np.asarray(mesh.triangles)
        normals = np.asarray(mesh.vertex_normals) if mesh.has_vertex_normals() else None

        # Pin to grid: bin each triangle's centroid into the request grid so the
        # LAD reuse path can keep only in-grid triangles. The renderer already
        # cropped points to the grid (honoring rotation via crop_box_rotation_deg),
        # so almost every centroid lands inside; the few that straddle the boundary
        # get -1 and are dropped downstream. Mirrors the Helios triangulation's
        # per-triangle cell ids. Cell binning here is axis-aligned (rotation is
        # ignored for the bin); the C++ setExternalTriangulation re-bins each
        # centroid with getContainingGridCell, which DOES honor cell rotation, so
        # a rotated grid is still assigned correctly at inversion time.
        triangle_cell_ids: list = []
        if request.grid is not None and triangles.shape[0] > 0:
            _report(0.97, "Binning into grid")
            g = request.grid
            v = vertices[triangles]  # (T, 3, 3)
            centroids = v.mean(axis=1)  # (T, 3)
            cells = _bin_points_to_cells(
                centroids, g.center, g.size, g.nx, g.ny, g.nz
            )
            triangle_cell_ids = cells

        _report(1.0, "Finalizing")
        return {
            "success": True,
            "vertices": vertices,
            "triangles": triangles,
            "normals": normals,
            "surface_area": float(surface_area),
            "num_triangles": int(triangles.shape[0]),
            "num_vertices": int(vertices.shape[0]),
            "method_used": method_used,
            "points_used": points_used,
            "downsampled": was_downsampled,
            "triangle_cell_ids": triangle_cell_ids,
        }

    except ImportError:
        return {"success": False, "method_used": request.method,
                "num_triangles": 0, "num_vertices": 0,
                "error": "Open3D not installed. Run: pip install open3d"}
    except ScanCancelled:
        raise  # cancellation propagates to the streaming wrapper (memory freed)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "method_used": request.method,
                "num_triangles": 0, "num_vertices": 0,
                "error": f"Triangulation failed: {str(e)}"}


def _pack_mesh_frame(result: dict, *, index_key: str = "triangles") -> bytes:
    """Pack a mesh result dict (vertices + index array + optional normals/colors/
    uv) into a PHB1 binary frame. Non-array fields (counts, surface_area,
    method, textures, materials, error, …) go in meta. `index_key` is the dict
    key holding the triangle index array (open3d uses 'triangles', plant/QSM
    'indices'); it's always sent as the buffer named 'indices'."""
    if not result.get("success"):
        meta = {k: v for k, v in result.items()
                if k not in ("vertices", "triangles", "indices", "normals", "colors", "uv_coordinates")}
        return _bin_frame_bytes(meta, [])
    array_keys = {"vertices", "triangles", "indices", "normals", "colors", "uv_coordinates",
                  "triangle_cell_ids"}
    meta = {k: v for k, v in result.items() if k not in array_keys}
    buffers = [("vertices", result["vertices"], "f32"), ("indices", result[index_key], "u32")]
    for opt in ("normals", "colors", "uv_coordinates"):
        if result.get(opt) is not None:
            buffers.append((opt, result[opt], "f32"))
    # Per-triangle grid cell ids (ball-pivot LAD pin). Pack like the Helios
    # triangulation: cast -1 to the uint32 sentinel 0xffffffff the renderer
    # treats as "outside". Only emitted when present and non-empty.
    cell_ids = result.get("triangle_cell_ids")
    if cell_ids is not None and len(cell_ids) > 0:
        buffers.append(("triangle_cell_ids",
                        np.asarray(cell_ids).astype(np.int64) & 0xFFFFFFFF, "u32"))
    return _bin_frame_bytes(meta, buffers)


@app.post("/api/triangulate")
async def triangulate_point_cloud(request: TriangulationRequest, http_request: Request):
    """Triangulate a point cloud (Open3D). Returns a PHB1 binary frame."""
    run_id, cancel_event = _new_cancel_token()
    return _bin_frame_streaming_response(
        lambda progress: _pack_mesh_frame(_do_open3d_triangulation(request, progress=progress)),
        request=http_request, cancel_event=cancel_event, run_id=run_id)


# ==================== GROUND SEGMENTATION ====================

class GroundSegmentationRequest(BaseModel):
    """Request model for ground/non-ground segmentation via CSF.

    Provide either inline `points` (flat clouds) or a `source` descriptor
    (octree-backed clouds — the backend re-reads the file). Unlike skeleton /
    triangulate, segmentation must NOT downsample: the returned `labels` align
    1:1 with the resolved point order so the renderer can attach them as a
    per-point scalar. Callers leave `source.max_points` as None.
    """
    points: Optional[List[List[float]]] = None
    source: Optional[PointSource] = None
    # CSF parameters (defaults tuned for close-range plant scans on flat-ish
    # ground, not airborne LiDAR — see segment_ground()).
    cloth_resolution: float = 0.05
    rigidness: int = 3
    class_threshold: float = 0.02
    iterations: int = 500
    slope_smooth: bool = False


class GroundSegmentationResponse(BaseModel):
    """Per-point ground/plant labels aligned to the resolved point order."""
    success: bool
    labels: List[int] = []          # 1=ground, 2=plant
    num_ground: int = 0
    num_plant: int = 0
    num_points: int = 0
    error: Optional[str] = None


def _resolve_segmentation_points(request: GroundSegmentationRequest) -> np.ndarray:
    """Resolve a GroundSegmentationRequest to an Nx3 float64 array, reading from
    the source file (full resolution) when `source` is set."""
    if request.source is not None:
        # Force full resolution: labels must align 1:1 with the cloud's points.
        src = request.source.model_copy(update={"max_points": None})
        points, _, _ = _read_points_from_source(src)
        return points
    if request.points is not None:
        return np.array(request.points, dtype=np.float64)
    raise HTTPException(status_code=400, detail="Provide either `points` or `source`.")


@app.post("/api/segment/ground", response_model=GroundSegmentationResponse)
async def segment_ground_points(request: GroundSegmentationRequest):
    """Classify a point cloud into ground (1) and plant (2) points using the
    Cloth Simulation Filter. Returns per-point labels aligned to input order;
    persisting the result onto an octree-backed cloud is done by
    `/api/cloud/session/{session_id}/segment_ground`."""
    try:
        points = _resolve_segmentation_points(request)

        if len(points) < 10:
            return GroundSegmentationResponse(
                success=False,
                num_points=len(points),
                error="Need at least 10 points for ground segmentation",
            )

        try:
            labels = segment_ground(
                points,
                cloth_resolution=request.cloth_resolution,
                rigidness=request.rigidness,
                class_threshold=request.class_threshold,
                iterations=request.iterations,
                slope_smooth=request.slope_smooth,
            )
        except ImportError:
            return GroundSegmentationResponse(
                success=False,
                num_points=len(points),
                error="CSF (cloth-simulation-filter) not installed. Run: pip install cloth-simulation-filter",
            )

        num_ground = int(np.count_nonzero(labels == GROUND_CLASS_GROUND))
        return GroundSegmentationResponse(
            success=True,
            labels=labels.tolist(),
            num_ground=num_ground,
            num_plant=len(labels) - num_ground,
            num_points=len(labels),
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ground segmentation failed: {str(e)}")


# ==================== DIGITAL ELEVATION MODEL (DEM) ====================
# Build a bare-earth terrain surface from a point cloud. Logically downstream of
# ground segmentation: when the cloud carries a `ground_class` column (or the
# caller passes ground labels), only ground points are gridded; otherwise CSF is
# run on the fly (auto) or, as a last resort, the lowest returns of all points
# are used. The elevation surface is built by TIN / Delaunay-linear interpolation
# (the las2dem / PDAL approach) onto a regular cell-centred grid, with a per-cell
# low-percentile pre-bin for outlier robustness and convex-hull void masking (no
# extrapolation beyond measured data unless `fill_voids` is set). Output is a
# heightmap surface mesh (same PHB1 transport as /api/triangulate) plus the
# underlying regular grid, which the renderer round-trips to /api/dem/export-raster
# for ESRI ASCII (.asc) / GeoTIFF (.tif) export.

# Default DEM cell-size seeding bounds (metres). Mirrors groundSegmentDefaults:
# scale-dependent, so seed from extent when the caller doesn't specify.
_DEM_CELL_MIN = 0.01
_DEM_CELL_MAX = 5.0
# Hard cap on grid cells (nx*ny). A too-fine cell on a huge extent would allocate
# an enormous grid; reject with a clear error instead (mirrors the TreeIso cap).
_DEM_MAX_CELLS = 4_000_000


class DemRequest(BaseModel):
    """Generate a DEM from a flat cloud's inline `points` or a `source` descriptor.

    Ground-aware: pass `ground_labels` (1=ground, 2=plant, aligned 1:1 with the
    resolved points) to grid only ground points; else CSF is run when
    `auto_segment_ground` is set, else all points are used (lowest-return surface).
    """
    points: Optional[List[List[float]]] = None
    source: Optional[PointSource] = None
    # Per-point ground labels (1=ground) aligned to the resolved point order. When
    # present, the DEM is built from ground points only (the "column" path).
    ground_labels: Optional[List[int]] = None
    # Run CSF to derive ground when no labels are supplied. CSF tuning mirrors
    # GroundSegmentationRequest's close-range defaults.
    auto_segment_ground: bool = True
    cloth_resolution: float = 0.05
    rigidness: int = 3
    class_threshold: float = 0.02
    # Grid + algorithm
    cell_size: Optional[float] = None        # None ⇒ seed from extent
    bbox: Optional[List[float]] = None        # [minx, miny, maxx, maxy]; None ⇒ ground AABB
    method: str = "tin"                        # "tin"/"linear", "nearest", "idw"
    ground_percentile: float = 5.0            # per-cell z percentile (robust near-min)
    fill_voids: bool = False                   # nearest-neighbour extrapolate into gaps
    # Also return a per-point height-above-ground buffer (point z − gap-free
    # ground), aligned to the resolved points. Used by the flat-cloud CHM path.
    compute_height_above_ground: bool = False


def _resolve_dem_points(request: "DemRequest") -> np.ndarray:
    """Resolve a DemRequest to an Nx3 float64 array (full resolution)."""
    if request.source is not None:
        src = request.source.model_copy(update={"max_points": None})
        points, _, _ = _read_points_from_source(src)
        return points
    if request.points is not None:
        return np.array(request.points, dtype=np.float64)
    raise HTTPException(status_code=400, detail="Provide either `points` or `source`.")



def _compute_dem(
    points: np.ndarray,
    *,
    cell_size: Optional[float] = None,
    bbox: Optional[List[float]] = None,
    method: str = "tin",
    ground_percentile: float = 5.0,
    fill_voids: bool = False,
    sample_xy: Optional[np.ndarray] = None,
    progress=None,
) -> dict:
    """Grid ground elevation into a DEM and build its heightmap surface mesh.

    `points` are the points to grid (already ground-filtered by the caller, or
    all points). Returns a result dict shaped for `_pack_dem_frame`: success,
    vertices/triangles/normals, surface_area, counts, method_used, plus the
    regular grid (`grid_z` row-major (ny*nx), `grid_nx`, `grid_ny`, `grid_cell`,
    `grid_origin` = [minx, miny] lower-left corner) and `voids`. Raises
    ValueError when the requested cell size would exceed the grid-cell cap."""
    def _report(frac, msg):
        if progress is not None:
            progress(frac, msg)

    from scipy.interpolate import griddata

    pts = np.asarray(points, dtype=np.float64)
    if pts.ndim != 2 or pts.shape[1] < 3 or len(pts) < 3:
        return {"success": False, "method_used": f"dem_{method}", "num_triangles": 0,
                "num_vertices": 0, "error": "Need at least 3 points to build a DEM."}

    xy = pts[:, :2]
    z = pts[:, 2]
    if bbox is not None and len(bbox) == 4:
        minx, miny, maxx, maxy = (float(v) for v in bbox)
    else:
        minx, miny = float(xy[:, 0].min()), float(xy[:, 1].min())
        maxx, maxy = float(xy[:, 0].max()), float(xy[:, 1].max())
    ext_x, ext_y = maxx - minx, maxy - miny
    if not (ext_x > 0 and ext_y > 0):
        return {"success": False, "method_used": f"dem_{method}", "num_triangles": 0,
                "num_vertices": 0, "error": "Point cloud has a degenerate XY extent."}

    if cell_size is None or cell_size <= 0:
        cell = float(np.clip(max(ext_x, ext_y) / 256.0, _DEM_CELL_MIN, _DEM_CELL_MAX))
    else:
        cell = float(cell_size)
    nx = max(int(np.ceil(ext_x / cell)), 1)
    ny = max(int(np.ceil(ext_y / cell)), 1)
    if nx * ny > _DEM_MAX_CELLS:
        raise ValueError(
            f"Cell size {cell:g} m is too fine for this extent "
            f"({nx:,}×{ny:,} = {nx*ny:,} cells; cap {_DEM_MAX_CELLS:,}). "
            "Use a larger cell size.")

    _report(0.2, "Binning ground points")
    # --- Per-cell low-percentile pre-bin (robust near-minimum representative) ---
    # Clip points to the grid; bin into flat cell indices; per populated cell take
    # the Nth-percentile z. This kills high outliers (residual low vegetation),
    # is steadier than a hard per-cell minimum, AND bounds the triangulation input
    # to <= nx*ny points regardless of cloud size.
    ci = np.clip(((xy[:, 0] - minx) / cell).astype(np.int64), 0, nx - 1)
    cj = np.clip(((xy[:, 1] - miny) / cell).astype(np.int64), 0, ny - 1)
    flat = cj * nx + ci
    order = np.lexsort((z, flat))      # sort by cell, then z ascending
    flat_s = flat[order]
    z_s = z[order]
    uniq, start = np.unique(flat_s, return_index=True)
    counts = np.diff(np.append(start, len(flat_s)))
    p = float(np.clip(ground_percentile, 0.0, 100.0)) / 100.0
    pick = start + np.floor((counts - 1) * p).astype(np.int64)
    rep_z = z_s[pick]
    rep_ci = uniq % nx
    rep_cj = uniq // nx
    rep_xy = np.column_stack([minx + (rep_ci + 0.5) * cell,
                              miny + (rep_cj + 0.5) * cell])

    if len(uniq) < 3:
        return {"success": False, "method_used": f"dem_{method}", "num_triangles": 0,
                "num_vertices": 0,
                "error": "Too few populated cells to build a DEM; use a finer cell size."}

    _report(0.5, "Interpolating elevation grid")
    # Target: every cell center.
    gx = minx + (np.arange(nx) + 0.5) * cell
    gy = miny + (np.arange(ny) + 0.5) * cell
    GX, GY = np.meshgrid(gx, gy)        # (ny, nx)
    grid_xy = np.column_stack([GX.ravel(), GY.ravel()])

    m = (method or "tin").lower()
    if m in ("tin", "linear"):
        zi = griddata(rep_xy, rep_z, grid_xy, method="linear")
    elif m == "nearest":
        zi = griddata(rep_xy, rep_z, grid_xy, method="nearest")
    elif m == "idw":
        from scipy.spatial import cKDTree
        tree = cKDTree(rep_xy)
        k = int(min(12, len(rep_xy)))
        dist, idx = tree.query(grid_xy, k=k)
        if k == 1:
            dist = dist[:, None]; idx = idx[:, None]
        with np.errstate(divide="ignore"):
            w = 1.0 / np.maximum(dist, 1e-12) ** 2
        zi = np.sum(w * rep_z[idx], axis=1) / np.sum(w, axis=1)
        zi[dist[:, 0] == 0] = rep_z[idx[dist[:, 0] == 0, 0]]
    else:
        return {"success": False, "method_used": f"dem_{method}", "num_triangles": 0,
                "num_vertices": 0, "error": f"Unknown DEM method {method!r}."}
    grid_z = zi.reshape(ny, nx).astype(np.float64)

    if fill_voids:
        nanmask = ~np.isfinite(grid_z)
        if nanmask.any():
            znear = griddata(rep_xy, rep_z, grid_xy, method="nearest").reshape(ny, nx)
            grid_z[nanmask] = znear[nanmask]

    voids = int((~np.isfinite(grid_z)).sum())

    _report(0.75, "Building surface mesh")
    # --- Heightmap mesh: vertices at cell centres, 2 triangles per quad whose
    # four corner cells are all finite (clean holes at voids). ---
    finite = np.isfinite(grid_z)
    verts_full = np.column_stack([GX.ravel(), GY.ravel(), grid_z.ravel()])
    ii, jj = np.meshgrid(np.arange(nx - 1), np.arange(ny - 1))
    v00 = (jj * nx + ii).ravel()
    v10 = (jj * nx + ii + 1).ravel()
    v01 = ((jj + 1) * nx + ii).ravel()
    v11 = ((jj + 1) * nx + ii + 1).ravel()
    ok = (finite[jj, ii] & finite[jj, ii + 1] &
          finite[jj + 1, ii] & finite[jj + 1, ii + 1]).ravel()
    v00, v10, v01, v11 = v00[ok], v10[ok], v01[ok], v11[ok]
    if len(v00) == 0:
        return {"success": False, "method_used": f"dem_{method}", "num_triangles": 0,
                "num_vertices": 0,
                "error": "DEM produced no surface (too few ground points or all cells empty)."}
    # CCW winding so face normals point up (+z).
    tris_full = np.vstack([np.column_stack([v00, v10, v11]),
                           np.column_stack([v00, v11, v01])])
    used = np.unique(tris_full)
    remap = np.full(len(verts_full), -1, dtype=np.int64)
    remap[used] = np.arange(len(used))
    verts = verts_full[used].astype(np.float64)
    tris = remap[tris_full].astype(np.int64)

    # Normals + surface area via open3d (consistent with other meshes).
    normals = None
    surface_area = None
    try:
        import open3d as o3d
        mesh = o3d.geometry.TriangleMesh()
        mesh.vertices = o3d.utility.Vector3dVector(verts)
        mesh.triangles = o3d.utility.Vector3iVector(tris.astype(np.int32))
        mesh.compute_vertex_normals()
        normals = np.asarray(mesh.vertex_normals, dtype=np.float32)
        surface_area = float(mesh.get_surface_area())
    except Exception:
        pass

    # Gap-free ground elevation at arbitrary XY (for height-above-ground). The
    # DEM mesh keeps honest voids, but HAG needs a ground value under EVERY point
    # — including canopy that overhangs past the ground's footprint or sits over a
    # gap with no ground returns — or those points would be left without a ground
    # reference. So sample the ground surface with linear interpolation inside the
    # ground points' convex hull and NEAREST-ground extrapolation outside it,
    # rather than the void-prone grid (whose NaNs would otherwise collapse those
    # points to height 0 and read as "at ground level").
    sample_ground_z = None
    if sample_xy is not None and len(sample_xy):
        sg = griddata(rep_xy, rep_z, sample_xy, method="linear")
        outside = ~np.isfinite(sg)
        if outside.any():
            sg[outside] = griddata(rep_xy, rep_z, sample_xy[outside], method="nearest")
        sample_ground_z = sg

    _report(0.95, "Packing DEM")
    return {
        "success": True,
        "vertices": verts.astype(np.float32),
        "triangles": tris.astype(np.int64),
        "normals": normals,
        "surface_area": surface_area,
        "num_vertices": int(len(verts)),
        "num_triangles": int(len(tris)),
        "method_used": f"dem_{m}",
        "points_used": int(len(uniq)),
        "grid_z": grid_z.astype(np.float32).ravel(),   # row-major, row 0 = min y
        "grid_nx": nx,
        "grid_ny": ny,
        "grid_cell": cell,
        "grid_origin": [minx, miny],                    # lower-left corner
        "voids": voids,
        # Per-sample ground z aligned to `sample_xy` (caller pops this — it's not
        # framed). None when sample_xy wasn't supplied.
        "sample_ground_z": sample_ground_z,
    }


def _pack_dem_frame(result: dict) -> bytes:
    """Pack a `_compute_dem` result into a PHB1 frame: vertices/indices/normals as
    buffers plus the regular `grid_z` grid (for raster export); everything else
    (counts, grid params, ground_source, world_shift, cache info) rides in meta."""
    array_keys = {"vertices", "triangles", "normals", "grid_z", "hag", "sample_ground_z"}
    if not result.get("success"):
        return _bin_frame_bytes({k: v for k, v in result.items() if k not in array_keys}, [])
    meta = {k: v for k, v in result.items() if k not in array_keys}
    buffers = [("vertices", result["vertices"], "f32"),
               ("indices", result["triangles"], "u32")]
    if result.get("normals") is not None:
        buffers.append(("normals", result["normals"], "f32"))
    if result.get("grid_z") is not None:
        buffers.append(("grid_z", result["grid_z"], "f32"))
    if result.get("hag") is not None:
        buffers.append(("hag", result["hag"], "f32"))
    return _bin_frame_bytes(meta, buffers)


def _auto_csf_params(points: np.ndarray) -> dict:
    """Extent-scaled CSF parameters for the DEM tool's automatic ground
    extraction. CSF's cloth resolution is an ABSOLUTE distance: a 5 cm cloth
    suits a ~1 m plant scan but is pathological on a field/ALS tile — a 186 m
    tile at 5 cm is a ~3700×3700 (~14 M-node) cloth simulated 500× and
    effectively hangs. The Ground Segmentation tool seeds these from the cloud's
    extent (see groundSegmentDefaults.ts); the DEM panel exposes no CSF controls,
    so its auto path must self-scale identically. Flat terrain → coarse stiff
    cloth ∝ extent; sloped (relief ratio ≥ 0.2) → finer, low-rigidness,
    slope-smoothed cloth that conforms instead of bridging."""
    xy = points[:, :2]
    ext = float(max(np.ptp(xy[:, 0]), np.ptp(xy[:, 1])))
    if not (ext > 0):
        ext = 1.5
    relief = float(np.ptp(points[:, 2]))
    ratio = relief / ext
    CLOTH_MIN, CLOTH_MAX, THR_MIN, THR_MAX = 0.05, 2.0, 0.02, 1.0
    if ratio >= 0.2:
        cloth = min(ext / 200.0, 1.0)
        return {"cloth_resolution": float(np.clip(cloth, CLOTH_MIN, CLOTH_MAX)),
                "class_threshold": 0.5, "rigidness": 1, "slope_smooth": True}
    scaled = ext / 100.0
    return {"cloth_resolution": float(np.clip(scaled, CLOTH_MIN, CLOTH_MAX)),
            "class_threshold": float(np.clip(scaled, THR_MIN, THR_MAX)),
            "rigidness": 3, "slope_smooth": False}


def _dem_ground_mask(points: np.ndarray, request: "DemRequest") -> "tuple[Optional[np.ndarray], str, Optional[str]]":
    """Resolve the ground mask for a stateless DEM request. Returns
    (mask | None, ground_source, warning). Mask None ⇒ use all points."""
    n = len(points)
    if request.ground_labels is not None and len(request.ground_labels) == n:
        mask = np.asarray(request.ground_labels, dtype=np.int64) == GROUND_CLASS_GROUND
        if int(mask.sum()) >= 3:
            return mask, "column", None
        return None, "all_points", "Ground class had too few points; used all points."
    if request.auto_segment_ground:
        try:
            csf = _auto_csf_params(points)   # extent-scaled — never the 5 cm plant default on a big tile
            labels = segment_ground(points, **csf)
            mask = labels == GROUND_CLASS_GROUND
            if int(mask.sum()) >= 3:
                return mask, "csf_auto", None
        except ImportError:
            return None, "all_points", "CSF not installed; used all points (lowest returns)."
    return None, "all_points", "No ground classification; used all points (lowest returns)."


def _do_dem(request: "DemRequest", progress=None) -> dict:
    """Stateless DEM: resolve points, derive a ground mask, grid the surface."""
    try:
        points = _resolve_dem_points(request)
        if len(points) < 3:
            return {"success": False, "method_used": f"dem_{request.method}",
                    "num_triangles": 0, "num_vertices": 0,
                    "error": "Need at least 3 points to build a DEM."}
        mask, ground_source, warning = _dem_ground_mask(points, request)
        ground_pts = points[mask] if mask is not None else points
        result = _compute_dem(ground_pts, cell_size=request.cell_size, bbox=request.bbox,
                              method=request.method, ground_percentile=request.ground_percentile,
                              fill_voids=request.fill_voids,
                              sample_xy=(points[:, :2] if request.compute_height_above_ground else None),
                              progress=progress)
        if result.get("success"):
            sample_ground_z = result.pop("sample_ground_z", None)
            result["ground_source"] = ground_source
            result["world_shift"] = [0.0, 0.0, 0.0]
            if warning:
                result["warning"] = warning
            # Per-point height-above-ground buffer (gap-free ground under every
            # point), framed for the flat-cloud CHM path.
            if request.compute_height_above_ground and sample_ground_z is not None:
                hag = points[:, 2] - sample_ground_z
                hag[~np.isfinite(hag)] = 0.0
                result["hag"] = hag.astype(np.float32)
        else:
            result.pop("sample_ground_z", None)
        return result
    except ValueError as e:
        return {"success": False, "method_used": f"dem_{request.method}",
                "num_triangles": 0, "num_vertices": 0, "error": str(e)}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "method_used": f"dem_{request.method}",
                "num_triangles": 0, "num_vertices": 0, "error": f"DEM generation failed: {e}"}


@app.post("/api/dem")
async def generate_dem(request: DemRequest, http_request: Request):
    """Generate a DEM from a flat cloud (inline points / source). Returns a PHB1
    binary frame (heightmap mesh + regular grid)."""
    run_id, cancel_event = _new_cancel_token()
    return _bin_frame_streaming_response(
        lambda progress: _pack_dem_frame(_do_dem(request, progress=progress)),
        request=http_request, cancel_event=cancel_event, run_id=run_id)


# ---- DEM raster export (.asc / GeoTIFF) ----
# The renderer round-trips the grid it received from /api/dem (or the session DEM
# endpoint) here, with voids encoded as `nodata` (JSON can't carry NaN) and the
# origin shifted back to true-world coordinates (world_shift re-added). The .asc
# path is pure text (no dependency); GeoTIFF uses tifffile (pure-Python — no
# GDAL) to write the raster plus the georeferencing tags.

class DemRasterExportRequest(BaseModel):
    format: str = "asc"                # "asc" | "tif"
    grid_z: List[float]                # row-major (ny*nx), row 0 = min y; voids = nodata
    nx: int
    ny: int
    cell_size: float
    origin: List[float]                # [minx, miny] true-world lower-left corner
    nodata: float = -9999.0
    crs_epsg: Optional[int] = None     # set georeferencing CRS when known


def _dem_asc_bytes(grid: np.ndarray, minx: float, miny: float, cell: float, nodata: float) -> bytes:
    """ESRI ASCII grid. Rows are written north (max y) to south. Cell-centred
    grid with the corner origin (xllcorner/yllcorner)."""
    ny, nx = grid.shape
    lines = [f"ncols {nx}", f"nrows {ny}", f"xllcorner {minx:.6f}",
             f"yllcorner {miny:.6f}", f"cellsize {cell:.6f}", f"NODATA_value {nodata:g}"]
    out = "\n".join(lines) + "\n"
    rows = []
    for j in range(ny - 1, -1, -1):    # north to south
        row = grid[j]
        rows.append(" ".join(f"{v:.4f}" if np.isfinite(v) else f"{nodata:g}" for v in row))
    return (out + "\n".join(rows) + "\n").encode("utf-8")


def _dem_geotiff_bytes(grid: np.ndarray, minx: float, miny: float, cell: float,
                       nodata: float, crs_epsg: Optional[int]) -> bytes:
    """GeoTIFF via tifffile (no GDAL). Writes the raster north-up with
    ModelPixelScale + ModelTiepoint, plus a GeoKeyDirectory when an EPSG is given."""
    import io
    import tifffile
    ny, nx = grid.shape
    data = np.where(np.isfinite(grid), grid, nodata).astype(np.float32)
    data = data[::-1]                  # raster row 0 = north (max y)
    maxy = miny + ny * cell
    extratags = [
        (33550, 12, 3, (float(cell), float(cell), 0.0), True),                 # ModelPixelScale
        (33922, 12, 6, (0.0, 0.0, 0.0, float(minx), float(maxy), 0.0), True),  # ModelTiepoint
        (42113, 2, 0, str(nodata), True),                                      # GDAL_NODATA
    ]
    if crs_epsg:
        try:
            import pyproj
            geographic = pyproj.CRS.from_epsg(int(crs_epsg)).is_geographic
        except Exception:
            geographic = False
        if geographic:
            geokeys = (1, 1, 0, 3, 1024, 0, 1, 2, 1025, 0, 1, 1, 2048, 0, 1, int(crs_epsg))
        else:
            geokeys = (1, 1, 0, 3, 1024, 0, 1, 1, 1025, 0, 1, 1, 3072, 0, 1, int(crs_epsg))
        extratags.append((34735, 3, len(geokeys), geokeys, True))              # GeoKeyDirectory
    buf = io.BytesIO()
    tifffile.imwrite(buf, data, extratags=extratags)
    return buf.getvalue()


@app.post("/api/dem/export-raster")
async def export_dem_raster(request: DemRasterExportRequest):
    """Write a DEM grid to ESRI ASCII (.asc) or GeoTIFF (.tif); returns base64."""
    import base64
    if request.nx <= 0 or request.ny <= 0 or len(request.grid_z) != request.nx * request.ny:
        raise HTTPException(status_code=400, detail="grid_z length must equal nx*ny.")
    grid = np.asarray(request.grid_z, dtype=np.float64).reshape(request.ny, request.nx)
    grid = np.where(grid == request.nodata, np.nan, grid)   # decode voids
    minx, miny = float(request.origin[0]), float(request.origin[1])
    fmt = request.format.lower()
    try:
        if fmt == "asc":
            data = _dem_asc_bytes(grid, minx, miny, request.cell_size, request.nodata)
            ext = "asc"
        elif fmt in ("tif", "tiff", "geotiff"):
            data = _dem_geotiff_bytes(grid, minx, miny, request.cell_size, request.nodata, request.crs_epsg)
            ext = "tif"
        else:
            raise HTTPException(status_code=400, detail=f"Unknown raster format {request.format!r}.")
    except ImportError:
        raise HTTPException(status_code=500, detail="tifffile not installed (required for GeoTIFF export).")
    return {"success": True, "format": ext, "data_base64": base64.b64encode(data).decode("ascii")}


# ==================== WOOD/LEAF SEGMENTATION ====================
# `wood_class` per-point scalar: 1=wood (trunk/branches), 2=leaf. Geometric,
# non-ML (verticality + low-sphericity); see segment_wood(). Mirrors the
# ground-segment endpoints (inline `points` or `source`; full-resolution; the
# returned `labels` align 1:1 so the renderer can attach them as a scalar).

class WoodSegmentationRequest(BaseModel):
    """Per-point wood/leaf segmentation from XYZ geometry.

    Provide inline `points` (flat clouds) or a `source` descriptor (octree-backed
    clouds — the backend re-reads the file at full resolution). Like ground
    segmentation the result must NOT be downsampled, so `labels` align 1:1 with
    the resolved point order. The tuning fields map onto segment_wood():
    `wood_bias` is the wood-vs-leaf sensitivity (lower → more wood), the `k_*`
    fields set the neighbourhood-scale search, `reg_iters` the smoothing
    strength, and `voxel_size` (>0) enables downsample-classify-propagate for
    very large clouds."""
    points: Optional[List[List[float]]] = None
    source: Optional[PointSource] = None
    # AGGREGATE: several pre-registered scans segmented TOGETHER (denser local
    # neighbourhoods → better wood/leaf geometry). Each source is read and its
    # points concatenated in order; `source_counts` in the response gives each
    # source's point count so the caller can scatter the labels back per scan.
    # `sources` takes precedence over `points`/`source`.
    sources: Optional[List[PointSource]] = None
    k_min: int = 10
    k_max: int = 100
    k_step: int = 10
    wood_bias: float = 0.6
    reg_k: int = 20
    reg_iters: int = 3
    min_speckle: int = 0
    branch_grow_sph: float = 0.02
    voxel_size: float = 0.0
    # REFLECTANCE ASSIST (optional). When the cloud carries a per-point
    # reflectance/intensity scalar (Riegl `Reflectance` dB, an ASCII
    # intensity/reflectance column, or a LAS extra-dim), it can supplement the
    # geometric score — but only proportionally to how separable wood/leaf
    # actually are in it (auto-weighted per cloud; ~0 on low-contrast species so
    # the result falls back to pure geometry). `reflectance` is the inline path
    # (aligned 1:1 with `points`); for `source`/`sources`/session clouds the
    # backend re-reads the scalar itself (`scalar_slug` picks which session
    # extra-dim, default 'reflectance' then 'intensity'). `reflectance_weight_max`
    # caps the blend weight (0 disables the assist entirely).
    reflectance: Optional[List[float]] = None
    scalar_slug: Optional[str] = None
    reflectance_weight_max: float = 0.4
    # METHOD.
    #  'sota' (default) = literature-faithful segment-wise classifier: skeleton
    #     branch-segments classified by cylinder-fit quality, recovering thin
    #     branches the point-wise methods drop without flooding leaf. Best on real
    #     TLS; REQUIRES ground removal (rooted skeleton).
    #  'connectivity' = geodesic-skeleton backbone recovery (the prior default).
    #  'geometric' = the original point-wise classifier (local PCA + GMM).
    # `backbone_support` (0 = auto) tunes the connectivity support floor.
    method: Literal["sota", "connectivity", "geometric"] = "sota"
    backbone_support: float = 0.0


class WoodSegmentationResponse(BaseModel):
    """Per-point wood/leaf labels aligned to the resolved point order."""
    success: bool
    labels: List[int] = []          # 1=wood, 2=leaf
    num_wood: int = 0
    num_leaf: int = 0
    num_points: int = 0
    # For an aggregate (multi-source) request: the point count of each source in
    # the order given, so the caller can slice `labels` back per scan. Empty for
    # a single-source / inline request.
    source_counts: List[int] = []
    # Non-fatal advisories from the connectivity method (e.g. the cloud's base
    # looks like un-removed ground). The result is still returned; the UI surfaces
    # these as a warning toast.
    warnings: List[str] = []
    error: Optional[str] = None


def _wood_segment_kwargs(request: "WoodSegmentationRequest") -> dict:
    """Extract segment_wood() tuning kwargs from a request (shared by the
    stateless and session endpoints)."""
    return dict(
        k_min=request.k_min,
        k_max=request.k_max,
        k_step=request.k_step,
        wood_bias=request.wood_bias,
        reg_k=request.reg_k,
        reg_iters=request.reg_iters,
        min_speckle=request.min_speckle,
        branch_grow_sph=request.branch_grow_sph,
        voxel_size=request.voxel_size,
        reflectance_weight_max=request.reflectance_weight_max,
        method=request.method,
        backbone_support=request.backbone_support,
    )


@app.post("/api/segment/wood", response_model=WoodSegmentationResponse)
async def segment_wood_points(request: WoodSegmentationRequest):
    """Classify a point cloud into wood (1) and leaf (2) points from geometry.
    Returns per-point labels aligned to input order; persisting the result onto
    an octree-backed cloud is done by
    `/api/cloud/session/{session_id}/segment_wood`."""
    try:
        # AGGREGATE: read each source full-resolution and concatenate IN ORDER so
        # the labels can be sliced back per source. The combined cloud is denser,
        # which is exactly the point — better local neighbourhoods. Assumes the
        # sources are pre-registered (a common coordinate frame).
        source_counts: List[int] = []
        # Optional per-point reflectance, kept aligned 1:1 with `points`. For the
        # source/sources paths the scalar is re-read from disk alongside XYZ (so
        # it survives the same full-resolution read); for inline `points` the
        # caller supplies it directly. None ⇒ pure-geometry (today's behaviour).
        reflectance: Optional[np.ndarray] = None
        if request.sources:
            parts = []
            refl_parts: List[Optional[np.ndarray]] = []
            for s in request.sources:
                full = s.model_copy(update={"max_points": None})  # full resolution
                pts, _, inten = _read_points_from_source(full)
                parts.append(np.asarray(pts, dtype=np.float64))
                refl_parts.append(
                    np.asarray(inten, dtype=np.float64) if inten is not None else None
                )
                source_counts.append(len(pts))
            points = np.concatenate(parts, axis=0) if parts else np.empty((0, 3))
            # Only use reflectance if EVERY source carried it (else alignment and
            # per-cloud separability are meaningless across mixed sources).
            if parts and all(r is not None for r in refl_parts):
                reflectance = np.concatenate(refl_parts, axis=0)
        elif request.source is not None:
            src = request.source.model_copy(update={"max_points": None})
            points, _, inten = _read_points_from_source(src)
            points = np.asarray(points, dtype=np.float64)
            reflectance = np.asarray(inten, dtype=np.float64) if inten is not None else None
        elif request.points is not None:
            points = np.array(request.points, dtype=np.float64)
            if request.reflectance is not None and len(request.reflectance) == len(points):
                reflectance = np.asarray(request.reflectance, dtype=np.float64)
        else:
            raise HTTPException(status_code=400, detail="Provide `points`, `source`, or `sources`.")

        if len(points) < 3:
            return WoodSegmentationResponse(
                success=False,
                num_points=len(points),
                error="Need at least 3 points for wood/leaf segmentation",
            )

        warns: List[str] = []
        labels = segment_wood(points, reflectance=reflectance, warnings=warns,
                              **_wood_segment_kwargs(request))
        num_wood = int(np.count_nonzero(labels == WOOD_CLASS_WOOD))
        return WoodSegmentationResponse(
            success=True,
            labels=labels.tolist(),
            num_wood=num_wood,
            num_leaf=len(labels) - num_wood,
            num_points=len(labels),
            source_counts=source_counts,
            warnings=warns,
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Wood/leaf segmentation failed: {str(e)}")


# ==================== TREE INSTANCE SEGMENTATION (TreeIso) ====================
# `tree_instance` is the per-point scalar field this writes: 0 = unassigned,
# 1..N = tree ids. Categorical coloring + legend live in the renderer's
# classification.ts. TreeIso isolates above-ground tree structure, so callers
# should remove ground first (see `ground_warning`). Mirrors the ground-segment
# endpoints above (inline `points` or `source`; full-resolution; per-point labels).
TREE_INSTANCE_SLUG = "tree_instance"
TREE_INSTANCE_LABEL = "Tree instance"

# TreeIso is CPU-only and O(n log n)+ with heavy constants; beyond a few million
# points it crawls. Cap so the endpoints fail fast with an actionable message
# instead of appearing to hang (the dense TLS benchmark plots are ~30M points).
# Override via env for power users with patience / a big machine.
try:
    _TREEISO_MAX_POINTS = int(os.environ.get("PHYTOGRAPH_TREEISO_MAX_POINTS", "5000000"))
except (ValueError, TypeError):
    _TREEISO_MAX_POINTS = 5_000_000

# BFS skeleton extraction builds an in-RAM neighbour graph and runs BFS/cluster
# passes over it; beyond a few million points the graph + Python passes get
# heavy. Cap so the endpoint fails fast with an actionable message instead of
# appearing to hang. Override via env. (The renderer already downsamples
# octree-backed clouds to ~20k via source.max_points; this guards the inline /
# uncapped paths.)
try:
    _SKELETON_MAX_POINTS = int(os.environ.get("PHYTOGRAPH_SKELETON_MAX_POINTS", "3000000"))
except (ValueError, TypeError):
    _SKELETON_MAX_POINTS = 3_000_000


class TreeSegmentationRequest(BaseModel):
    """Request model for individual-tree segmentation via TreeIso.

    Like ground segmentation, provide inline `points` or a `source` descriptor;
    the result must NOT be downsampled so per-point `labels` align 1:1. Optional
    `seed_points` ([[x,y,z], ...]) anchor trees for human-in-the-loop seeding —
    each seed yields exactly one tree id. The remaining fields are TreeIso
    parameters (defaults match Xi & Hopkinson 2022)."""
    points: Optional[List[List[float]]] = None
    source: Optional[PointSource] = None
    seed_points: Optional[List[List[float]]] = None
    # Optional per-point ground/plant labels (aligned 1:1 with `points`), as
    # produced by a prior ground segmentation that was kept but not deleted.
    # When given, ground points (== GROUND_CLASS_GROUND) are excluded from
    # TreeIso and returned as tree id 0. Ignored on the session endpoint, which
    # reads the persisted `ground_class` column directly.
    ground_class: Optional[List[float]] = None
    reg_strength1: float = 1.0
    min_nn1: int = 5
    decimate_res1: float = 0.05
    reg_strength2: float = 15.0
    min_nn2: int = 20
    decimate_res2: float = 0.1
    max_gap: float = 2.0
    rel_height_length_ratio: float = 0.5
    vertical_weight: float = 0.5
    min_nn3: int = 20
    score_candidate_thresh: float = 0.7
    init_stem_rel_length_thresh: float = 1.5
    max_outlier_gap: float = 3.0


class TreeSegmentationResponse(BaseModel):
    """Per-point tree instance ids aligned to the resolved point order."""
    success: bool
    labels: List[int] = []          # 0 = unassigned, 1..N = tree ids
    num_trees: int = 0
    num_points: int = 0
    ground_warning: bool = False
    error: Optional[str] = None


def _treeiso_params(request: "TreeSegmentationRequest"):
    """Build a vendored TreeIsoParams from the request's tuning fields."""
    from treeiso.treeiso_core import TreeIsoParams

    return TreeIsoParams(
        reg_strength1=request.reg_strength1,
        min_nn1=request.min_nn1,
        decimate_res1=request.decimate_res1,
        reg_strength2=request.reg_strength2,
        min_nn2=request.min_nn2,
        decimate_res2=request.decimate_res2,
        max_gap=request.max_gap,
        rel_height_length_ratio=request.rel_height_length_ratio,
        vertical_weight=request.vertical_weight,
        min_nn3=request.min_nn3,
        score_candidate_thresh=request.score_candidate_thresh,
        init_stem_rel_length_thresh=request.init_stem_rel_length_thresh,
        max_outlier_gap=request.max_outlier_gap,
    )


def _auto_treeiso_decimation(points: np.ndarray, p) -> None:
    """Raise TreeIso's stage-1/2 voxel sizes in place when they're finer than the
    cloud's actual point spacing — otherwise voxel decimation is a no-op and
    cut-pursuit runs over the full N (a 2.6 M-node ALS tile → 15-20 min hang).

    Mirrors `_auto_csf_params`: the renderer seeds these from extent
    (treeSegmentDefaults.ts) when the Segment Trees panel opens, but the inline /
    eval path rides request defaults, so the backend must self-scale or it can
    still hang un-seeded. We measure the real signal — median nearest-neighbor
    spacing (the `_do_spacing_check` cKDTree k=2 pattern, sampled to stay cheap on
    multi-M clouds) — rather than extent, since spacing is exactly what makes the
    decimation a no-op.

    We bump ONLY when `decimate_res` is still at-or-below the upstream paper
    default (0.05 / 0.1) AND too fine for the spacing. A coarser value (already
    seeded by the UI for a big tile, or chosen by a power user) is left alone, so
    this is idempotent with the frontend seed. Small clouds early-out and keep the
    paper defaults bit-for-bit."""
    from scipy.spatial import cKDTree

    n = len(points)
    if n < 50_000:
        # Small clouds decimate fine and finish fast; never probe, never bump.
        return
    pts = np.asarray(points, dtype=np.float64)[:, :3]
    finite = np.isfinite(pts).all(axis=1)
    pts = pts[finite]
    if len(pts) < 50_000:
        return
    # Median NN spacing on a sample (cKDTree k=2: [0] is the point itself,
    # [1] its nearest neighbor). Sampling keeps the probe O(sample) on big tiles.
    sample = pts
    if len(pts) > 200_000:
        idx = np.random.default_rng(0).choice(len(pts), 200_000, replace=False)
        sample = pts[idx]
    tree = cKDTree(pts)
    dist, _ = tree.query(sample, k=2, workers=-1)
    nn = dist[:, 1]
    nn = nn[np.isfinite(nn) & (nn > 0)]
    if nn.size == 0:
        return
    spacing = float(np.median(nn))

    # Aim for ~3× spacing so each stage-1 voxel pools a handful of points.
    target1 = 3.0 * spacing
    # Defensive: if 3× spacing would still leave a huge decimated cloud (very
    # dense tile near the point cap), coarsen further so the decimated graph stays
    # bounded (~≤1 M nodes). Voxel count scales ~ (spacing / res)³ at constant
    # density, so res ∝ spacing · (n / target_nodes)^(1/3).
    TARGET_DECIMATED_NODES = 1_000_000
    if len(pts) > TARGET_DECIMATED_NODES:
        target1 = max(target1, spacing * (len(pts) / TARGET_DECIMATED_NODES) ** (1.0 / 3.0))

    # Only raise, and only when riding the paper default (gate a hair above 0.05 /
    # 0.1 for float tolerance — a UI-coarsened value is > 0.051 and no-ops here).
    if p.decimate_res1 <= 0.051 and target1 > p.decimate_res1:
        p.decimate_res1 = round(target1, 3)
    target2 = 2.0 * p.decimate_res1
    if p.decimate_res2 <= 0.101 and target2 > p.decimate_res2:
        p.decimate_res2 = round(target2, 3)


def _looks_like_ground_present(points: np.ndarray) -> bool:
    """Cheap advisory heuristic: is there a broad, flat, dense layer at the bottom?

    TreeIso expects ground-removed input. We flag (never block) when the lowest
    vertical slab holds a large share of points spread across the full XY extent
    — the signature of an un-removed ground surface."""
    if len(points) < 100:
        return False
    z = points[:, 2]
    z0, z1 = float(np.min(z)), float(np.max(z))
    span = z1 - z0
    if span <= 1e-6:
        return False
    slab = points[z <= z0 + 0.05 * span]
    if len(slab) < 50:
        return False
    frac = len(slab) / len(points)
    xy_extent = np.ptp(points[:, :2], axis=0)
    slab_extent = np.ptp(slab[:, :2], axis=0)
    wide = bool(np.all(slab_extent > 0.6 * np.maximum(xy_extent, 1e-6)))
    return bool(frac > 0.12 and wide)


def segment_trees(
    points: np.ndarray,
    params=None,
    seeds: Optional[np.ndarray] = None,
) -> np.ndarray:
    """Assign each point a tree id (0 = unassigned, 1..N) via TreeIso.

    With `seeds`, every TreeIso segment is reassigned to the id of its nearest
    seed (each seed -> exactly one tree); otherwise TreeIso's own 1..N labels
    are returned, aligned 1:1 to the input order."""
    from treeiso.treeiso_core import segment_trees as _ti_segment

    labels = _ti_segment(points, params)
    if seeds is None or len(seeds) == 0:
        return labels.astype(np.int64)

    # Reconcile TreeIso's segments with user trunk seeds so each seed yields one
    # tree. Greedily give every seed its nearest still-unclaimed segment (so no
    # seed is dropped when there are at least as many segments as seeds), then
    # assign any leftover segments to their nearest seed.
    seeds = np.asarray(seeds, dtype=np.float64)[:, :3]
    seg_ids = np.unique(labels)
    centroids = np.array([points[labels == s, :3].mean(axis=0) for s in seg_ids])
    cost = np.linalg.norm(centroids[:, None, :] - seeds[None, :, :], axis=2)  # (segs, seeds)

    seg_to_seed: dict[int, int] = {}
    unclaimed = set(range(len(seg_ids)))
    # Seeds whose closest segment is nearest get first pick of that segment.
    for j in np.argsort(cost.min(axis=0)):
        if not unclaimed:
            break
        cand = min(unclaimed, key=lambda i, jj=int(j): cost[i, jj])
        seg_to_seed[int(seg_ids[cand])] = int(j) + 1
        unclaimed.discard(cand)
    for i in unclaimed:
        seg_to_seed[int(seg_ids[i])] = int(np.argmin(cost[i])) + 1

    return np.array([seg_to_seed[int(s)] for s in labels], dtype=np.int64)


@app.post("/api/segment/trees", response_model=TreeSegmentationResponse)
async def segment_trees_points(request: TreeSegmentationRequest, http_request: Request):
    """Segment a multi-tree cloud into per-point tree instance ids via TreeIso.

    Mirrors `/api/segment/ground`: inline `points` or a `source` descriptor in,
    per-point integer labels out (0 = unassigned, 1..N = trees), full resolution
    so labels align 1:1. Persisting onto an octree-backed cloud is done by
    `/api/cloud/session/{session_id}/segment_trees`.

    The TreeIso pipeline is CPU-bound and runs for tens of seconds on a large
    tile, so it executes off the event loop (`_run_blocking_until_disconnect`):
    other `/api/*` requests stay responsive, and if the client disconnects (panel
    closed, fetch timeout) this returns at once instead of holding the request
    open.

    Labels-only by design: this interactive path returns predicted instance ids
    and nothing else. If the source carries ground-truth fields (e.g. a
    benchmark PLY's `instance`/`semantic`), they are NOT echoed back here — only
    `/apply` carries source scalars through into the octree, and the eval
    harness (`scripts/eval_tree_segmentation.py`) reads GT straight from the
    file. There is no GT consumer on this endpoint."""
    try:
        points = _resolve_segmentation_points(
            GroundSegmentationRequest(points=request.points, source=request.source)
        )
        n_full = len(points)

        if n_full < 10:
            return TreeSegmentationResponse(
                success=False, num_points=n_full,
                error="Need at least 10 points to segment trees",
            )

        # The points TreeIso actually sees: finite coordinates (cKDTree chokes
        # on NaN/inf) AND not labelled ground. Ground points (from a prior
        # ground segmentation that was kept, not deleted) are excluded so they
        # aren't clustered into a "tree". Labels are scattered back to the full
        # input order with 0 (unassigned) for every excluded point, so `labels`
        # aligns 1:1 with the points the caller sent.
        eligible = np.isfinite(points).all(axis=1)
        ground_excluded = False
        if request.ground_class is not None and len(request.ground_class) == n_full:
            eligible &= np.asarray(request.ground_class) != GROUND_CLASS_GROUND
            ground_excluded = True

        pts = points[eligible]
        if len(pts) < 10:
            reason = "non-finite and ground-labelled" if ground_excluded else "non-finite"
            return TreeSegmentationResponse(
                success=False, num_points=n_full,
                error=f"Fewer than 10 points remain after dropping {reason} points.",
            )

        if len(pts) > _TREEISO_MAX_POINTS:
            return TreeSegmentationResponse(
                success=False, num_points=n_full,
                error=(f"{len(pts):,} points exceeds the {_TREEISO_MAX_POINTS:,}-point "
                       "limit for tree segmentation. Downsample or crop first."),
            )

        # Skip the advisory ground heuristic when the caller already handed us
        # ground labels to exclude — the ground is gone from `pts`.
        ground_warning = (not ground_excluded) and _looks_like_ground_present(pts)
        seeds = (
            np.asarray(request.seed_points, dtype=np.float64)
            if request.seed_points else None
        )
        def _segment():
            ti_params = _treeiso_params(request)
            _auto_treeiso_decimation(pts, ti_params)   # ~0.8s probe — keep off-loop too
            return segment_trees(pts, ti_params, seeds)

        try:
            sub_labels = await _run_blocking_until_disconnect(_segment, http_request)
        except ClientDisconnected:
            # Client gave up (panel closed / fetch timeout). Nothing to return.
            return TreeSegmentationResponse(
                success=False, num_points=n_full,
                error="Tree segmentation was cancelled (client disconnected).",
            )
        except ImportError as e:
            return TreeSegmentationResponse(
                success=False, num_points=n_full,
                error=f"TreeIso dependencies not installed ({e}). "
                      "Run: pip install -r backend-api/requirements.txt",
            )

        labels = np.zeros(n_full, dtype=np.int64)
        labels[eligible] = np.asarray(sub_labels)
        num_trees = int(len(np.unique(labels[labels > 0])))
        return TreeSegmentationResponse(
            success=True,
            labels=[int(x) for x in labels],
            num_trees=num_trees,
            num_points=n_full,
            ground_warning=ground_warning,
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        return TreeSegmentationResponse(
            success=False, num_points=0, error=str(e),
        )


# ==================== HELIOS TRIANGULATION ====================

class PoseSample(BaseModel):
    """One 6-DOF platform pose sample: time + position + Hamilton (body->world)
    quaternion (qx, qy, qz, qw, scalar last)."""
    t: float
    x: float
    y: float
    z: float
    qx: float
    qy: float
    qz: float
    qw: float


class FrameMeta(BaseModel):
    """Documented frame/CRS so a timestamp join is never silently wrong. `up_axis`
    is the world up convention ('z' for Phytograph); `body_convention`/`time_ref`
    are recorded for downstream interoperability (FLU/FRD, GPS-week vs relative)."""
    crs: Optional[str] = None
    up_axis: str = "z"
    body_convention: str = "FLU"
    time_ref: Optional[str] = None


class PoseStream(BaseModel):
    """A dense timestamped 6-DOF platform trajectory plus its calibration — the
    canonical moving-platform representation (see backend-api/trajectory.py).

    Quaternions are Hamilton body->world (qx,qy,qz,qw). `lever_arm` (body-frame
    scanner optical center, meters) and `boresight_rpy` (sensor misalignment
    roll/pitch/yaw, radians) calibrate the platform->scanner transform; the per-beam
    emission origin is pos(t) + R(quat(t))·lever_arm. The backend resolver joins
    this to each return's timestamp (SLERP attitude, linear position)."""
    poses: List[PoseSample]
    frame: FrameMeta = FrameMeta()
    lever_arm: List[float] = [0.0, 0.0, 0.0]      # body-frame [x, y, z] meters
    boresight_rpy: List[float] = [0.0, 0.0, 0.0]  # [roll, pitch, yaw] radians
    source_format: str = "pose_csv"


class HeliosScanEntry(BaseModel):
    """A single scan with point data (or file path) and scanner position.

    Provide either `file_path` (preferred for large scans) or `points`.
    When `file_path` is given, the backend reads the file directly via pyhelios,
    avoiding large JSON transfers.
    """
    file_path: Optional[str] = None     # Path to scan file on disk (preferred)
    ascii_format: Optional[str] = None  # Column format e.g. "x y z timestamp" (auto-detected if omitted)
    points: Optional[List[List[float]]] = None  # [[x, y, z], ...] fallback when no file_path
    colors: Optional[List[List[float]]] = None  # [[r, g, b], ...] (0-1 range)
    origin: List[float]                 # [x, y, z] scanner position
    # Per-scan acquisition geometry, carried from the scan's own ScanParameters.
    # Helios triangulates each scan in its scanner-angular (theta, phi) grid, so
    # these describe how that scan was actually sampled. When omitted, the
    # backend falls back to the request-level theta/phi and a count-based
    # estimate of the grid resolution (see _do_helios_computation).
    n_theta: Optional[int] = None       # Zenith samples (Ntheta)
    n_phi: Optional[int] = None         # Azimuth samples (Nphi)
    theta_min: Optional[float] = None   # Zenith angle min (degrees)
    theta_max: Optional[float] = None   # Zenith angle max (degrees)
    phi_min: Optional[float] = None     # Azimuth angle min (degrees)
    phi_max: Optional[float] = None     # Azimuth angle max (degrees)
    # Return type for leaf-area-density. "single" needs only x y z; "multi"
    # (full-waveform) needs per-pulse timestamp/target_index/target_count
    # columns so Helios can group beams and gap-fill misses. Ignored by
    # triangulation; consumed by _do_lad_computation.
    return_type: str = "single"         # "single" | "multi"
    beam_exit_diameter: Optional[float] = None   # meters (multi-return only)
    beam_divergence: Optional[float] = None      # milliradians (multi-return only)
    # LAD point-data source (consumed by _do_lad_computation, ignored by
    # triangulation). A session-backed cloud passes `session_id`: the LAD path
    # feeds its surviving in-RAM points — including the per-pulse multi-return
    # columns when present — via _session_to_lad_arrays, honoring unbaked
    # deletions and never re-reading the source. A synthetic (flat, in-RAM)
    # cloud has no file or session, so it passes `points` plus, for
    # multi-return, the aligned per-pulse columns in `scalar_columns` (a dict of
    # column-name -> per-point values, e.g. timestamp/target_index/target_count).
    session_id: Optional[str] = None
    scalar_columns: Optional[Dict[str, List[float]]] = None
    # Moving-platform trajectory. When present, this scan is a moving-platform
    # acquisition: `origin` is only a fallback anchor (it should equal the first
    # pose's position), and the LAD path reconstructs a PER-BEAM emission origin
    # for every return by joining its `timestamp` to this trajectory (see
    # backend-api/trajectory.py). Those per-beam origins drive the beam-based
    # (Gtheta) leaf-area inversion, which needs no triangulation. Static scans
    # leave this None and behave exactly as before. Ignored by triangulation.
    trajectory: Optional[PoseStream] = None

class HeliosGrid(BaseModel):
    """An explicit triangulation grid, derived from a voxel box in the UI.

    Helios's XML loader requires a <grid> block; it bounds the triangulation
    region. center/size are in world coordinates (the voxel box's transform);
    nx/ny/nz are its per-axis cell subdivisions."""
    center: List[float]  # [x, y, z]
    size: List[float]    # [x, y, z] full extents
    nx: int = 1
    ny: int = 1
    nz: int = 1
    # Azimuthal rotation of the box about +z, in degrees (the voxel box's
    # z-Euler angle). Helios's <grid> honors this and crops the ROTATED box;
    # without it a rotated UI grid would crop only its axis-aligned extent,
    # leaking points past the rotated walls. 0 / absent = axis-aligned.
    rotation: float = 0.0

class HeliosTriangulationRequest(BaseModel):
    """Request model for Helios triangulation"""
    scans: List[HeliosScanEntry]
    lmax: float = 0.5
    max_aspect_ratio: float = 4.0
    # Request-level angular fallbacks, used only for scans that don't carry
    # their own theta/phi (see HeliosScanEntry). Kept for backward compatibility.
    theta_min: float = 30.0   # Zenith angle min (degrees)
    theta_max: float = 130.0  # Zenith angle max (degrees)
    phi_min: float = 0.0      # Azimuth angle min (degrees)
    phi_max: float = 360.0    # Azimuth angle max (degrees)
    # Explicit grid from a user-created voxel box. When absent the backend
    # auto-creates a single-cell grid encompassing all scan points and flags
    # grid_warning on the response.
    grid: Optional[HeliosGrid] = None

class HeliosFilterEstimate(BaseModel):
    """Auto-estimated triangulation filter, derived from the candidate edge-length
    distribution (Otsu separability) + a merged-multi-scan-cloud guard."""
    lmax: Optional[float] = None      # suggested Lmax (m); None when too little spread
    eta: float = 0.0                  # separation confidence in [0,1]
    label: str = "n/a"                # High / Medium / Low / n/a
    sep_ratio: Optional[float] = None # upper-mode / lower-mode median edge ratio
    sep_label: str = "n/a"            # mode separation: High (≥4x) / Medium (≥2x) / Low
    merged: bool = False
    merged_message: Optional[str] = None

class HeliosTriangulationResponse(BaseModel):
    """Response model for Helios triangulation"""
    success: bool
    vertices: List[List[float]] = []
    triangles: List[List[int]] = []
    colors: Optional[List[List[float]]] = None
    normals: Optional[List[List[float]]] = None
    surface_area: Optional[float] = None
    num_triangles: int = 0
    num_vertices: int = 0
    method_used: str = "helios"
    error: Optional[str] = None
    # Source scan index (0-based, into request.scans) for each triangle, aligned
    # 1:1 with `triangles`. Helios triangulates each scan independently, so every
    # triangle belongs to exactly one scan; this lets the UI color triangles by
    # their originating scan.
    triangle_scan_ids: List[int] = []
    # Grid cell index (into the request grid, row-major i + nx*(j + ny*k)) that
    # each triangle's centroid falls in, aligned 1:1 with `triangles`. -1 means
    # the centroid fell outside every cell. With the auto 1x1x1 grid every
    # triangle lands in cell 0. Lets the UI split per-cell leaf-angle
    # distributions.
    triangle_cell_ids: List[int] = []
    # True when no explicit grid was supplied and the backend triangulated all
    # points within their auto-computed bounding box (assumes ground/trunk were
    # already segmented or cropped). Carries a human-readable companion message.
    grid_warning: bool = False
    grid_message: Optional[str] = None
    # Interactive-filter support. The returned mesh is the auto-estimated default
    # (small candidate sets are returned whole). The front-end recomputes the
    # per-triangle edge metrics from the returned geometry and filters within
    # these loosening limits; `estimate` seeds the default filter and reports the
    # separation confidence + merged-cloud guard. `candidate_count` is the total
    # number of Delaunay candidates before the payload cap.
    cap_lmax: float = 0.0      # max Lmax (m) the returned set supports without a re-run
    cap_aspect: float = 1.0e9  # max aspect the returned set supports without a re-run
    candidate_count: int = 0
    estimate: Optional["HeliosFilterEstimate"] = None


class SpacingCheckResponse(BaseModel):
    """Verdict from /api/triangulate/check-spacing — an opt-in cross-check of the
    auto-estimated Lmax against the actual point spacing.

    The auto-Lmax estimator (Otsu over candidate triangle-edge lengths) silently
    fails on sparse shells: a thin layer of surface points generates mostly
    *bridge* triangles spanning the cell interior, so the candidate-edge
    distribution looks bimodal (eta/sep_ratio read "Medium"+) even though its
    lower mode is still bridges, not surface. The chosen Lmax then far exceeds the
    true point spacing and the reconstructed normals — hence G(theta) — are
    garbage. The candidate edges can't self-diagnose this; an INDEPENDENT measure
    of the surface scale can. We compute the median nearest-neighbor spacing of
    the points strictly inside the grid cell(s) and compare it to `lmax`. This is
    O(N log N) (a KD-tree build + query) and can take tens of seconds on a
    tens-of-millions-of-points cloud, which is why it's a user-triggered button
    rather than part of the triangulation, and why it's only offered when the
    Otsu indicators aren't both High."""
    success: bool
    median_spacing: Optional[float] = None  # median NN distance (m) of in-cell points
    lmax: float = 0.0                       # the Lmax this verdict was checked against
    ratio: Optional[float] = None           # lmax / median_spacing
    n_points: int = 0                       # in-cell points the spacing was measured on
    # True when ratio >= _SPACING_BRIDGE_RATIO: Lmax is large enough relative to the
    # point spacing that the triangulation is likely bridging across the surface.
    likely_bridging: bool = False
    message: Optional[str] = None           # human-readable verdict
    error: Optional[str] = None


# ==================== LEAF AREA DENSITY (LAD) ====================
# Per-voxel leaf area density (m^2/m^3) via the PyHelios LiDAR plugin. LAD is
# NOT the sum of triangle areas: the triangulation only supplies the per-cell
# G-function (the leaf-projection coefficient for Beer's law); Helios then
# traces beam paths through the voxel grid and inverts Beer's law per voxel to
# recover LAD. Unlike triangulation, the voxel grid is REQUIRED — it is the
# basis of the calculation.

class DemRaster(BaseModel):
    """A regular axis-aligned DEM elevation raster, as produced by /api/dem.

    Used by terrain-following LAD to sample a ground height under each voxel
    column. ``grid_z`` is the elevation per cell, row-major with row 0 = min y
    (``grid_z[j*nx + i]``), matching the renderer's MeshEntry.demGrid. Void cells
    are NaN. ``origin`` is the lower-left corner [minx, miny]; ``cell`` is the cell
    size (m). Coordinates are in the SAME frame as the scan points / grid (the
    renderer round-trips world_shift so the raster aligns with the cloud)."""
    grid_z: List[float]
    nx: int
    ny: int
    cell: float
    origin: List[float]                    # [minx, miny] in WORLD coords, cell-corner of (0,0)
    # JSON can't carry NaN, so the renderer encodes DEM voids as this sentinel
    # (mirroring the GeoTIFF export). Cells equal to it are treated as voids and
    # nearest-filled. None => no sentinel (all grid_z finite).
    nodata: Optional[float] = None


class LADComputeRequest(BaseModel):
    """Request model for leaf area density computation.

    Reuses HeliosScanEntry for scans (each carrying its scanner origin, angular
    geometry, and return_type). The grid is REQUIRED (its nx/ny/nz are the LAD
    voxel divisions) — there is no meaningful "auto single-cell" LAD.
    """
    scans: List[HeliosScanEntry]
    grid: HeliosGrid                       # REQUIRED — the LAD voxel grid
    # Terrain following: when enabled, each voxel column is shifted vertically so
    # its bottom face rides a DEM surface (clear of the ground). `dem` is REQUIRED
    # when terrain_follow is true. `safety_fraction` is the clearance between the
    # DEM surface and the lowest cell, expressed as a fraction of one cell's height
    # (size.z/nz) so it auto-scales with voxel resolution.
    terrain_follow: bool = False
    safety_fraction: float = 0.5
    dem: Optional[DemRaster] = None
    lmax: float = 0.1                      # max triangle edge length (G-function)
    max_aspect_ratio: float = 4.0         # max triangle aspect ratio
    min_voxel_hits: Optional[int] = None  # min ray hits for a voxel to be solved
    # Characteristic vegetation element width (m); broadleaf ~0.05, conifer ~0.002.
    # Drives the Pimont et al. (2018) per-voxel sampling-uncertainty computation,
    # which runs on every LAD inversion. element_width <= 0 yields a sampling-only
    # variance (term a; the element-position term b is omitted).
    element_width: float = 0.05
    # Request-level angular fallbacks (degrees), used only for scans that don't
    # carry their own theta/phi (mirrors HeliosTriangulationRequest).
    theta_min: float = 30.0
    theta_max: float = 130.0
    phi_min: float = 0.0
    phi_max: float = 360.0
    # Mean leaf-projection coefficient G(theta), in (0, 1] (0.5 = spherical/random
    # leaf-angle distribution). Used in two cases:
    #   1. Moving-platform scans ALWAYS need it — their pulses don't lie on a fixed
    #      theta-phi grid, so they can't be triangulated to derive G(theta).
    #   2. Static scans when gtheta_override is True (see below).
    gtheta: Optional[float] = None
    # Static-scan G(theta) override. When True, a static (terrestrial) scan SKIPS
    # triangulation and inverts Beer's law with the supplied `gtheta` directly —
    # the same beam-based path moving scans use. This is the right choice for a
    # homogeneous canopy whose leaf-angle distribution is known (or assumed
    # spherical), and for aerial-style data where triangulating sparse overhead
    # returns recovers a biased G(theta). Ignored for moving scans (they always
    # use the supplied G(theta)). When False, static scans triangulate to estimate
    # G(theta) per cell as before.
    gtheta_override: bool = False

class LADCell(BaseModel):
    """A single voxel result."""
    index: int
    center: List[float]   # [x, y, z]
    size: List[float]     # [x, y, z]
    leaf_area: float      # m^2 within the voxel
    lad: float            # m^2/m^3 (leaf_area / voxel volume)
    gtheta: float         # G(theta) leaf-projection coefficient
    hit_count: int        # points falling inside the voxel (numpy-binned)
    # Pimont et al. (2018) per-voxel sampling uncertainty. None when undefined for
    # the cell (unsolved voxel, or the variance/CI fell outside its validity range).
    beam_count: Optional[int] = None             # N beams that entered the voxel
    relative_density_index: Optional[float] = None  # I_rdi, fraction intercepted
    mean_path_length: Optional[float] = None     # mean beam path length (m)
    lad_variance: Optional[float] = None         # sampling variance of LAD, (1/m)^2
    lad_std: Optional[float] = None              # sqrt(lad_variance)
    ci_valid: Optional[bool] = None              # single-voxel Pimont validity gate
    leaf_area_ci_lower: Optional[float] = None   # m^2 (only when ci_valid)
    leaf_area_ci_upper: Optional[float] = None   # m^2 (only when ci_valid)

class LADComputeResponse(BaseModel):
    """Response model for leaf area density computation."""
    success: bool
    cells: List[LADCell] = []
    nx: int = 1
    ny: int = 1
    nz: int = 1
    grid_center: List[float] = []
    grid_size: List[float] = []
    grid_rotation: float = 0.0        # azimuth about +z, degrees (rotated grids)
    bounds: List[List[float]] = []   # [[lo_x,lo_y,lo_z], [hi_x,hi_y,hi_z]]
    # Terrain following (echoed for display). dropped_columns counts (x,y) columns
    # whose center fell outside the DEM footprint and were excluded from `cells`.
    terrain_follow: bool = False
    dropped_columns: int = 0
    is_multi_return: bool = False
    return_mode: str = "single"      # "single" | "multi"
    total_leaf_area: float = 0.0
    method_used: str = "helios"
    # Group-scale LAD confidence interval (Pimont et al. 2018, Eq. 39) over all
    # solved voxels — the recommended aggregate (much tighter than single-voxel).
    # group_ci_valid=False => the interval fell outside the Pimont validity range
    # (or no voxel was solved) and lower/upper are omitted.
    group_ci_valid: Optional[bool] = None
    group_lad_mean: Optional[float] = None       # m^2/m^3
    group_lad_ci_lower: Optional[float] = None   # m^2/m^3
    group_lad_ci_upper: Optional[float] = None   # m^2/m^3
    confidence_level: Optional[float] = None     # e.g. 0.95
    element_width: Optional[float] = None         # echo of the width used (m)
    warnings: List[str] = []
    error: Optional[str] = None


def _detect_ascii_format(file_path: str) -> str:
    """Auto-detect the ASCII column format of a scan file."""
    with open(file_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
            cols = len(line.split())
            if cols <= 3:
                return "x y z"
            elif cols == 4:
                return "x y z timestamp"
            elif cols == 5:
                return "x y z timestamp intensity"
            elif cols >= 6:
                return "x y z timestamp intensity color"
            break
    return "x y z"


def _count_file_lines(file_path: str) -> int:
    """Fast line count for a text file."""
    count = 0
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 16), b''):
            count += chunk.count(b'\n')
    return count


def _xyz_column_indices(ascii_format: Optional[str]) -> tuple:
    """Return the (x, y, z) column indices for a Helios ASCII_format string.

    Helios scan files don't always lead with x y z — e.g. the lidar plugin's
    sphere fixture uses 'row column x y z r g b reflectance', putting the
    coordinates in columns 2-4. Tokenize the format to find them; fall back to
    columns 0-2 when no format is given (the common bare-XYZ case).
    """
    if ascii_format:
        tokens = _tokenize_ascii_format(ascii_format)
        try:
            return tokens.index('x'), tokens.index('y'), tokens.index('z')
        except ValueError:
            pass
    return 0, 1, 2


def _file_xyz_bounds(file_path: str, ascii_format: Optional[str] = None):
    """Stream an ASCII point file once and return (n_points, min_xyz, max_xyz).

    Used to build the auto-grid bounding box when no explicit grid is supplied.
    The x/y/z columns are located via the scan's ASCII_format (see
    _xyz_column_indices), so formats that don't lead with the coordinates are
    handled correctly. Non-numeric / comment lines are skipped. Returns
    (count, None, None) if no coordinates were found.
    """
    import math
    xi, yi, zi = _xyz_column_indices(ascii_format)
    need = max(xi, yi, zi) + 1
    n = 0
    lo = [math.inf, math.inf, math.inf]
    hi = [-math.inf, -math.inf, -math.inf]
    with open(file_path) as f:
        for line in f:
            line = line.strip()
            if not line or line[0] in '#/':
                continue
            cols = line.split()
            if len(cols) < need:
                continue
            try:
                x, y, z = float(cols[xi]), float(cols[yi]), float(cols[zi])
            except ValueError:
                continue
            if x < lo[0]: lo[0] = x
            if y < lo[1]: lo[1] = y
            if z < lo[2]: lo[2] = z
            if x > hi[0]: hi[0] = x
            if y > hi[1]: hi[1] = y
            if z > hi[2]: hi[2] = z
            n += 1
    if n == 0:
        return 0, None, None
    return n, lo, hi


def _grid_xml_block(center: list, size: list, nx: int = 1, ny: int = 1,
                    nz: int = 1, rotation_deg: float = 0.0) -> list:
    """Build the lines of a Helios ``<grid>`` block. ``rotation_deg`` (azimuth
    about +z) is emitted as a ``<rotation>`` tag only when meaningfully non-zero
    — the Helios parser treats it as optional, defaulting to 0. Shared by the
    triangulation config writer and the scan-export grid round-trip so both stay
    consistent."""
    lines = ['<grid>']
    lines.append(f'    <center>{center[0]} {center[1]} {center[2]}</center>')
    lines.append(f'    <size>{size[0]} {size[1]} {size[2]}</size>')
    lines.append(f'    <Nx>{nx}</Nx>')
    lines.append(f'    <Ny>{ny}</Ny>')
    lines.append(f'    <Nz>{nz}</Nz>')
    if abs(float(rotation_deg)) > 1e-9:
        lines.append(f'    <rotation>{rotation_deg}</rotation>')
    lines.append('</grid>')
    return lines


def _inject_grids_into_helios_xml(xml_path: str, grids: list) -> None:
    """Insert <grid> blocks (one per ScanExportGrid) into an existing Helios XML,
    immediately before the closing </helios>. Used to round-trip scene grids on
    scan export, since PyHelios exportScans() emits only <scan> blocks. No-op when
    ``grids`` is empty. Falls back to appending if no </helios> tag is found."""
    if not grids:
        return
    blocks = []
    for g in grids:
        blocks.append('')
        blocks.extend(_grid_xml_block(g.center, g.size, g.nx, g.ny, g.nz, g.rotation))
    addition = '\n'.join(blocks) + '\n'

    with open(xml_path, "r") as fh:
        text = fh.read()
    idx = text.rfind('</helios>')
    if idx == -1:
        text = text.rstrip() + '\n' + addition
    else:
        text = text[:idx] + addition + text[idx:]
    with open(xml_path, "w") as fh:
        fh.write(text)


def _generate_helios_xml(tmpdir: str, scans_info: list, grid_center: list,
                         grid_size: list, grid_nx: int = 1, grid_ny: int = 1,
                         grid_nz: int = 1, xml_name: str = "helios_config.xml",
                         grid_rotation_deg: float = 0.0) -> str:
    """Generate a pyhelios XML config file for scan triangulation.

    Each entry in ``scans_info`` carries its own per-scan acquisition geometry
    (``n_theta``/``n_phi`` and ``theta_min``/``theta_max``/``phi_min``/
    ``phi_max``), since Helios triangulates each scan in its own scanner-angular
    grid. ``grid_nx``/``grid_ny``/``grid_nz`` set the grid cell subdivisions
    (1×1×1 single cell by default). ``grid_rotation_deg`` is the grid box's
    azimuth about +z (degrees) — Helios crops the rotated box when non-zero.
    ``xml_name`` lets callers write one config per scan into the same temp dir
    without clobbering.
    """
    import os

    xml_lines = ['<?xml version="1.0"?>', '<helios>', '']

    for scan in scans_info:
        xml_lines.append('<scan>')
        xml_lines.append(f'    <filename>{scan["filepath"]}</filename>')
        xml_lines.append(f'    <ASCII_format>{scan["ascii_format"]}</ASCII_format>')
        xml_lines.append(f'    <origin>{scan["origin"][0]} {scan["origin"][1]} {scan["origin"][2]}</origin>')
        xml_lines.append(f'    <size>{scan["n_theta"]} {scan["n_phi"]}</size>')
        xml_lines.append(f'    <thetaMin>{scan["theta_min"]}</thetaMin>')
        xml_lines.append(f'    <thetaMax>{scan["theta_max"]}</thetaMax>')
        xml_lines.append(f'    <phiMin>{scan["phi_min"]}</phiMin>')
        xml_lines.append(f'    <phiMax>{scan["phi_max"]}</phiMax>')
        xml_lines.append('</scan>')
        xml_lines.append('')

    # Grid section is required for triangulation to work
    xml_lines.extend(_grid_xml_block(grid_center, grid_size, grid_nx, grid_ny,
                                     grid_nz, grid_rotation_deg))
    xml_lines.append('')
    xml_lines.append('</helios>')

    xml_path = os.path.join(tmpdir, xml_name)
    with open(xml_path, "w") as f:
        f.write('\n'.join(xml_lines))

    return xml_path


def _triangulation_zero_message(diag: dict, lmax: float, max_aspect: float) -> str:
    """Human-readable explanation for a 0-triangle result, derived from the
    filter breakdown so the user can tell a data problem from a filter problem.
    """
    cand = diag["candidates"]
    if cand == 0:
        return ("No triangles generated: the triangulation produced no candidate "
                "triangles at all. The selected points are too few, collinear, or "
                "fell outside the grid box — check that the scan has data inside "
                "the triangulation grid.")
    lm, asp = diag["dropped_lmax"], diag["dropped_aspect"]
    # Attribute to whichever filter removed the most candidates.
    if lm >= asp:
        return (f"No triangles generated: all {cand} candidate triangle(s) were "
                f"filtered, {lm} of them because an edge exceeded Lmax "
                f"({lmax:g} m). The point spacing is coarser than Lmax — increase "
                f"Lmax (or rescan at higher resolution).")
    return (f"No triangles generated: all {cand} candidate triangle(s) were "
            f"filtered, {asp} of them by the aspect-ratio limit "
            f"({max_aspect:g}). Raise the Max Aspect Ratio to keep elongated "
            f"triangles.")


def _triangulation_quality_warning(diag: dict):
    """Warn when a mesh was produced but almost nothing was filtered, which
    means Lmax/aspect are large enough to bridge real gaps — the mesh may
    contain spurious triangles spanning unsampled regions. Returns None when
    the result looks healthy.
    """
    cand = diag["candidates"]
    if cand == 0 or diag["kept"] == 0:
        return None
    dropped = diag["dropped_lmax"] + diag["dropped_aspect"] + diag["dropped_degenerate"]
    if dropped / cand < 0.01:
        return (f"Only {dropped} of {cand} candidate triangle(s) were filtered "
                f"(<1%). Lmax/Max Aspect Ratio may be large enough to bridge real "
                f"gaps in the point cloud — inspect the mesh for triangles spanning "
                f"unsampled regions, and lower Lmax if so.")
    return None


# A full-resolution TLS tree easily yields several million candidate triangles;
# the unfiltered set (every Delaunay candidate) can exceed V8's ~512 MB max
# string length when serialized to JSON, which the renderer then can't parse
# ("Unexpected end of JSON input"). So the triangulation response is bounded:
# small candidate sets are returned whole (the front-end can filter freely in
# both directions), while large ones are pre-filtered to the default aspect and
# the densest this-many triangles by edge length (the front-end can then tighten
# the filter, but loosening past the cap needs a re-run). The binary PHB1
# transport removes the JSON string-length ceiling, so this is a render/memory
# safety cap (12M ≈ a full-resolution tree) rather than a transport limit.
_HELIOS_MAX_RETURN_TRIANGLES = 12_000_000
# A scan looks like a merged multi-scan cloud when its median candidate edge is
# this many times its 5th-percentile edge (the finest true spacing).
_HELIOS_MERGED_RATIO = 10.0


def _session_hit_positions_for_triangulation(session_id: str) -> "np.ndarray":
    """Surviving HIT positions (N,3 float64, world coords) of a cloud session,
    for Helios triangulation. Honors unbaked deletions and ALWAYS excludes sky/
    miss points — a miss is a ray that hit nothing, projected ~1 km out, never a
    surface, so it must never be triangulated (it would mesh a phantom shell and,
    with the post-backfill miss count often exceeding the hits, balloon the run).

    Mirrors the miss exclusion + world-shift add-back of the open3d session feed
    (see the `src.session_id` branch of the point-source resolver). The session's
    backfilled-miss buffer lives OUTSIDE `positions`, so it is excluded for free;
    interleaved `is_miss` points (when present) are dropped via the extra.
    """
    import numpy as np
    sess = _cloud_sessions.get(session_id)
    if sess is None:
        raise ValueError(
            f"Cloud session not found: {session_id}. The backend may have "
            "restarted since import. Re-import the scan and try again.")
    with _cloud_session_lock:
        keep = ~sess.deleted
        if _MISS_SLUG in sess.extras:
            keep = keep & (sess.extras[_MISS_SLUG] == 0)
        positions = sess.positions[keep].copy()
        world_shift = sess.world_shift
    positions = positions.astype(np.float64, copy=False)
    if world_shift is not None:
        positions = positions + np.asarray(world_shift, dtype=np.float64)
    return np.ascontiguousarray(positions)
_HELIOS_MERGED_MESSAGE = (
    "These points look like a merged multi-scan cloud — candidate edges are far "
    "larger than the point spacing, which happens when a single scan actually "
    "combines several scanner positions. Helios triangulation assumes single-scan-"
    "position data, so it bridges surfaces seen from different origins. Triangulate "
    "each scan position separately for a clean result."
)


def _otsu_threshold_eta(log_vals, nbins: int = 256):
    """Otsu's threshold + separability eta on 1-D values (here log edge lengths).
    eta = max between-class variance / total variance, in [0, 1] (1 = cleanly
    bimodal, 0 = unimodal). Returns (threshold_log, eta); threshold_log is NaN
    when there's too little spread to threshold."""
    import numpy as np
    if len(log_vals) < 8 or np.allclose(log_vals, log_vals[0]):
        return float("nan"), 0.0
    hist, bin_edges = np.histogram(log_vals, bins=nbins)
    centers = 0.5 * (bin_edges[:-1] + bin_edges[1:])
    p = hist.astype(np.float64) / max(hist.sum(), 1)
    omega = np.cumsum(p)
    mu = np.cumsum(p * centers)
    mu_t = mu[-1]
    denom = omega * (1.0 - omega)
    with np.errstate(divide="ignore", invalid="ignore"):
        sigma_b2 = np.where(denom > 1e-12, (mu_t * omega - mu) ** 2 / denom, 0.0)
    idx = int(np.argmax(sigma_b2))
    total_var = float((p * (centers - mu_t) ** 2).sum())
    eta = float(sigma_b2[idx] / (total_var + 1e-12))
    return float(centers[idx]), max(0.0, min(1.0, eta))


def _helios_filter_estimate(edge_max, scan_ids) -> dict:
    """Auto-estimate the triangulation Lmax + separation confidence (+ merged-
    cloud guard) from the candidate per-triangle max-edge lengths. Returns
    {lmax, eta, label, merged, merged_message}; lmax is None when the spread is
    too small to threshold. Folds the (retired) /suggest logic into the main run
    so estimation is instant — no second triangulation."""
    import numpy as np
    e = np.asarray(edge_max, dtype=np.float64)
    e = e[e > 0]
    if e.size < 16:
        return {"lmax": None, "eta": 0.0, "label": "n/a",
                "sep_ratio": None, "sep_label": "n/a",
                "merged": False, "merged_message": None}
    thr_log, eta = _otsu_threshold_eta(np.log(e))
    lmax = float(np.exp(thr_log)) if math.isfinite(thr_log) else None
    # Merged-cloud guard, per scan (each is triangulated independently).
    merged = False
    sids = np.asarray(scan_ids, dtype=np.int64)
    pos = np.asarray(edge_max, dtype=np.float64) > 0
    sids = sids[pos]
    for s in np.unique(sids):
        es = e[sids == s]
        if es.size < 64:
            continue
        ratio = float(np.median(es)) / max(float(np.percentile(es, 5)), 1e-9)
        if ratio > _HELIOS_MERGED_RATIO:
            merged = True
            break
    label = "High" if eta >= 0.7 else ("Medium" if eta >= 0.5 else "Low")
    # Mode-separation ratio: median edge of the upper Otsu class / median edge of
    # the lower class. eta says the two classes separate *cleanly*; this says how
    # far *apart* they are. A genuine gap-bridge population sits many times above
    # the surface-triangle spacing, so the threshold trims bridges. Two families
    # of merely anisotropic but equally-valid surface triangles (e.g. a coarse
    # convex scan) sit close together (~1.5x), so a clean split (high eta) at a
    # small ratio is a sign the cut is slicing one surface rather than trimming
    # bridges. Reported next to eta so the user can spot that case.
    sep_ratio = None
    sep_label = "n/a"
    if lmax is not None and math.isfinite(thr_log):
        lower = e[e <= lmax]
        upper = e[e > lmax]
        if lower.size and upper.size:
            med_lo = float(np.median(lower))
            if med_lo > 0:
                sep_ratio = float(np.median(upper)) / med_lo
                sep_label = ("High" if sep_ratio >= 4.0
                             else "Medium" if sep_ratio >= 2.0 else "Low")
    return {"lmax": lmax, "eta": eta, "label": label,
            "sep_ratio": sep_ratio, "sep_label": sep_label,
            "merged": merged, "merged_message": _HELIOS_MERGED_MESSAGE if merged else None}


def _do_helios_computation(request: HeliosTriangulationRequest, edges_only: bool = False, progress=None) -> dict:
    """Run Helios triangulation synchronously. Returns a result dict.

    When `edges_only` is True (the Lmax-suggestion path), skip the vertex dedup,
    index/area/cell building, and Python-list serialization entirely — the
    suggestion only needs the per-triangle edge-length distribution, so we
    return the raw edge lengths + scan ids as numpy arrays straight from the
    triangulator. On large clouds the dedup + `.tolist()` round-trips dominate
    that path, so this is the bulk of the suggest cost, not the triangulation.

    Supports two modes:
    - **File-path mode** (preferred): scans provide `file_path` pointing to scan
      files on disk. PyHelios reads them directly — no JSON point transfer overhead.
    - **Points mode** (fallback): scans provide `points` arrays sent over JSON.
      Works for small point clouds but too slow for large scans (680K+ points).

    Output vertices are deduplicated and indexed to minimize response size.
    """
    import math
    import tempfile
    import shutil
    import os
    import numpy as np

    def _report(fraction, message):
        if progress is not None:
            progress(fraction, message)

    def _ckpt():
        _cancel_checkpoint(progress)

    tmpdir = None
    try:
        from pyhelios import LiDARCloud

        _report(0.05, "Reading scans")
        tmpdir = tempfile.mkdtemp(prefix="phytograph_helios_")

        scans_info = []

        # Per-scan angular geometry comes from the scan's own ScanParameters when
        # present; the request-level values are only a fallback for scans that
        # don't carry their own. Likewise the grid resolution (n_theta/n_phi) is
        # used as sent and only estimated from point count when absent.
        def _angles(scan_entry):
            return (
                scan_entry.theta_min if scan_entry.theta_min is not None else request.theta_min,
                scan_entry.theta_max if scan_entry.theta_max is not None else request.theta_max,
                scan_entry.phi_min if scan_entry.phi_min is not None else request.phi_min,
                scan_entry.phi_max if scan_entry.phi_max is not None else request.phi_max,
            )

        def _resolution(scan_entry, n_points, theta_span, phi_span):
            """Per-scan Ntheta/Nphi, preferring the values the scan carries."""
            if scan_entry.n_theta and scan_entry.n_phi:
                return int(scan_entry.n_theta), int(scan_entry.n_phi)
            # Fallback: back out a plausible grid from the point count and aspect.
            aspect = theta_span / max(phi_span, 1e-10)
            n_phi = max(int(math.sqrt(n_points / max(aspect, 0.01))), 10)
            n_theta = max(int(n_points / n_phi), 10)
            return n_theta, n_phi

        # Bounding box over all scan points, accumulated as scans are processed,
        # so the auto-grid (when no explicit grid is supplied) tightly encloses
        # every point regardless of which mode produced them.
        bb_lo = np.array([np.inf, np.inf, np.inf])
        bb_hi = np.array([-np.inf, -np.inf, -np.inf])

        # Resolve each scan's point SOURCE by the same priority the renderer +
        # LAD use, so the in-RAM arrays stay the source of truth and the file is
        # never re-read once a session exists:
        #   1. session_id — a session-backed (octree) cloud: triangulate its
        #      surviving in-RAM HIT positions (deletions honored, MISSES EXCLUDED).
        #      Written to a temp x-y-z file so Helios reconstructs theta/phi from
        #      (pos - origin) exactly as for any scan file.
        #   2. file_path — a file-backed cloud with no session: Helios reads it
        #      directly from disk (no huge JSON, original columns preserved).
        #   3. inline points — a flat in-RAM cloud (e.g. synthetic scan).
        # A session_id whose session is gone (backend restarted) falls back to
        # file_path when the entry carries one, else errors.
        for idx, scan_entry in enumerate(request.scans):
            origin = scan_entry.origin
            if len(origin) != 3:
                raise ValueError(f"Origin must have 3 elements, got {len(origin)}")

            theta_min, theta_max, phi_min, phi_max = _angles(scan_entry)

            session_id = scan_entry.session_id
            session_present = False
            if session_id is not None:
                with _cloud_session_lock:
                    session_present = session_id in _cloud_sessions

            if session_id is not None and session_present:
                # Session source of truth: hits only, misses excluded.
                xyz = _session_hit_positions_for_triangulation(session_id)
                if xyz.shape[0] == 0:
                    raise ValueError(
                        "The edited point cloud has no surviving hit points to "
                        "triangulate (all points deleted or only misses remain).")
                bb_lo = np.minimum(bb_lo, xyz.min(axis=0))
                bb_hi = np.maximum(bb_hi, xyz.max(axis=0))
                pts_path = os.path.join(tmpdir, f"scan_{idx}.txt")
                # %.8g, not %.6g: world coords can sit near z~100 m, where 6
                # sig-figs rounds to ~1 mm and quantizes the angular structure the
                # spherical-projection Delaunay depends on — coarse enough to
                # shatter leaf surfaces (most candidates then exceed Lmax). 8
                # sig-figs keeps ~µm precision at 100 m.
                np.savetxt(pts_path, xyz, fmt="%.8g", delimiter=" ")
                n_theta, n_phi = _resolution(
                    scan_entry, xyz.shape[0], theta_max - theta_min, phi_max - phi_min)
                fmt = "x y z"
                fp = pts_path
            elif scan_entry.file_path:
                # File-path mode: pyhelios reads the scan file directly from disk.
                fp = scan_entry.file_path
                if not fp or not os.path.isfile(fp):
                    raise ValueError(f"Scan file not found: {fp}")
                fmt = scan_entry.ascii_format or _detect_ascii_format(fp)
                n_points, lo, hi = _file_xyz_bounds(fp, fmt)
                if lo is not None:
                    bb_lo = np.minimum(bb_lo, lo)
                    bb_hi = np.maximum(bb_hi, hi)
                n_theta, n_phi = _resolution(
                    scan_entry, n_points, theta_max - theta_min, phi_max - phi_min)
            elif scan_entry.session_id is not None:
                # session_id given but the session is gone and no file fallback.
                raise ValueError(
                    f"Cloud session not found: {scan_entry.session_id}. The "
                    "backend may have restarted since import. Re-import the scan "
                    "and try again.")
            else:
                # Points mode (fallback): write inline points to a temp file.
                points = scan_entry.points
                if not points:
                    raise ValueError("Scan entry has no points, file_path, or session_id")
                pts_arr_scan = np.asarray(points, dtype=float)
                bb_lo = np.minimum(bb_lo, pts_arr_scan[:, :3].min(axis=0))
                bb_hi = np.maximum(bb_hi, pts_arr_scan[:, :3].max(axis=0))
                pts_path = os.path.join(tmpdir, f"scan_{idx}.txt")
                # %.8g for the same precision reason as the session branch above.
                np.savetxt(pts_path, pts_arr_scan[:, :3], fmt="%.8g", delimiter=" ")
                n_theta, n_phi = _resolution(
                    scan_entry, len(points), theta_max - theta_min, phi_max - phi_min)
                fmt = "x y z"
                fp = pts_path

            scans_info.append({
                "filepath": fp,
                "ascii_format": fmt,
                "origin": origin,
                "n_theta": n_theta,
                "n_phi": n_phi,
                "theta_min": theta_min,
                "theta_max": theta_max,
                "phi_min": phi_min,
                "phi_max": phi_max,
            })

        # Resolve the grid. An explicit grid (from a voxel box in the UI) is used
        # verbatim. Otherwise auto-create a single cell tightly enclosing all
        # points and flag the response so the UI can warn that every point is
        # being triangulated (assumes ground/trunk already segmented/cropped).
        _report(0.1, "Resolving grid")
        grid_warning = False
        grid_message = None
        if request.grid is not None:
            grid_center = list(request.grid.center)
            grid_size = list(request.grid.size)
            grid_nx, grid_ny, grid_nz = request.grid.nx, request.grid.ny, request.grid.nz
            grid_rotation_deg = float(request.grid.rotation)
        else:
            if not np.all(np.isfinite(bb_lo)) or not np.all(np.isfinite(bb_hi)):
                raise ValueError("Could not determine point bounds for auto-grid")
            grid_center = ((bb_lo + bb_hi) / 2).tolist()
            grid_size = (np.maximum(bb_hi - bb_lo, 0.01) * 1.1).tolist()
            grid_nx = grid_ny = grid_nz = 1
            grid_rotation_deg = 0.0
            grid_warning = True
            grid_message = (
                "No grid box was specified — triangulating all points within "
                "their bounding box. This assumes ground and trunk have already "
                "been segmented or cropped."
            )

        # Triangulate each scan independently and tag every output triangle with
        # its source scan index. Helios already triangulates per scan internally
        # (it groups hit points by scanID before the Delaunay pass), so running
        # one LiDARCloud per scan yields the same triangles as a single merged
        # cloud — it just makes the provenance explicit, which a merged run
        # discards at the Context boundary. Triangle vertices are pulled in ONE
        # bulk FFI call per scan (getTriangleVerticesAll) instead of a per-UUID
        # Context loop, then deduplicated globally in numpy so shared geometry
        # between scans still collapses.
        scan_vert_blocks = []   # list of (T_s*9,) float32 flat-vertex arrays
        scan_id_blocks = []     # list of (T_s,) per-triangle source scan index

        # Triangulation filter diagnostics, summed across the per-scan clouds.
        # Helios attributes each dropped candidate to ONE primary reason (edge
        # length > Lmax, then aspect ratio, then degenerate/NaN area), so
        # candidates == kept + dropped_lmax + dropped_aspect + dropped_degenerate.
        diag = {"candidates": 0, "kept": 0, "dropped_lmax": 0,
                "dropped_aspect": 0, "dropped_degenerate": 0}

        n_scans = len(scans_info)
        for scan_idx, scan_info in enumerate(scans_info):
            # Fresh LiDARCloud per scan ⇒ a cancel here unwinds cleanly before the
            # next (monolithic) triangulateHitPoints call.
            _ckpt()
            _report(0.1 + 0.7 * scan_idx / max(n_scans, 1),
                    f"Triangulating scan {scan_idx + 1} of {n_scans}"
                    if n_scans > 1 else "Triangulating")
            xml_path = _generate_helios_xml(
                tmpdir, [scan_info], grid_center, grid_size,
                grid_nx, grid_ny, grid_nz,
                xml_name=f"helios_config_{scan_idx}.xml",
                grid_rotation_deg=grid_rotation_deg,
            )

            cloud = LiDARCloud()
            cloud.disableMessages()
            cloud.loadXML(xml_path)
            cloud.triangulateHitPoints(request.lmax, request.max_aspect_ratio)
            tri_count = cloud.getTriangleCount()

            stats = cloud.getTriangulationStats()
            diag["candidates"] += stats["candidates"]
            diag["kept"] += tri_count
            diag["dropped_lmax"] += stats["dropped_lmax"]
            diag["dropped_aspect"] += stats["dropped_aspect"]
            diag["dropped_degenerate"] += stats["dropped_degenerate"]

            if tri_count == 0:
                continue

            flat, _ = cloud.getTriangleVerticesAll()
            scan_vert_blocks.append(flat)
            scan_id_blocks.append(np.full(tri_count, scan_idx, dtype=np.int64))

        if not scan_vert_blocks:
            # A genuine zero. Report success=False (the UI keys off this) and
            # surface the filter breakdown so the user can tell whether the data
            # was too coarse for Lmax (most/all candidates dropped by Lmax),
            # the shape filter was too aggressive (dropped by aspect), or there
            # was simply nothing to triangulate (no candidates at all).
            return {
                "success": False,
                "vertices": [],
                "triangles": [],
                "num_triangles": 0,
                "num_vertices": 0,
                "method_used": "helios",
                "error": _triangulation_zero_message(diag, request.lmax,
                                                     request.max_aspect_ratio),
                "diagnostics": diag,
                "grid_warning": grid_warning,
                "grid_message": grid_message,
            }

        # Edges-only fast path (Lmax suggestion): compute per-triangle max edge
        # length directly from the raw flat (T*9) vertex blocks — no dedup, no
        # index/area/cell building, no Python-list/JSON round-trip. Each block is
        # 9 floats per triangle (3 vertices x xyz), so reshape to (T,3,3).
        if edges_only:
            flat = np.concatenate(scan_vert_blocks).reshape(-1, 3, 3)
            scan_ids = np.concatenate(scan_id_blocks)
            a, b, c = flat[:, 0], flat[:, 1], flat[:, 2]
            edges = np.maximum.reduce([
                np.linalg.norm(a - b, axis=1),
                np.linalg.norm(b - c, axis=1),
                np.linalg.norm(c - a, axis=1),
            ])
            return {
                "success": True,
                "edges": edges,
                "scan_ids": scan_ids,
                "diagnostics": diag,
            }

        # Per-triangle filter metrics over ALL candidates, computed from the RAW
        # (pre-dedup) flat vertex blocks. Order matches the concatenation order
        # (and thus the per-scan ids). These drive the auto-estimate and the
        # server-side payload cap; they are NOT sent to the front-end (which
        # recomputes them from the returned geometry) — serializing two float
        # arrays per million triangles is what pushes a big mesh past the JSON
        # string limit.
        tri_v = np.concatenate(scan_vert_blocks).reshape(-1, 3, 3)
        _e = np.stack([
            np.linalg.norm(tri_v[:, 0] - tri_v[:, 1], axis=1),
            np.linalg.norm(tri_v[:, 1] - tri_v[:, 2], axis=1),
            np.linalg.norm(tri_v[:, 0] - tri_v[:, 2], axis=1),
        ], axis=1)
        tri_edge_max = _e.max(axis=1)
        tri_edge_min = _e.min(axis=1)
        with np.errstate(divide="ignore", invalid="ignore"):
            tri_aspect = np.where(tri_edge_min > 0, tri_edge_max / tri_edge_min, 1.0e9)
        scan_ids_all = np.concatenate(scan_id_blocks)

        # Auto-estimate Lmax + separation confidence from the full candidate
        # distribution (instant — no second triangulation), and flag merged
        # multi-scan clouds.
        estimate = _helios_filter_estimate(tri_edge_max, scan_ids_all)

        # Bound the returned payload. Small candidate sets go back whole (the
        # front-end can then filter in either direction); large ones are
        # pre-filtered to the default aspect and capped at the densest
        # _HELIOS_MAX_RETURN_TRIANGLES by edge length, so the JSON stays
        # parseable. cap_lmax / cap_aspect tell the front-end the loosening
        # limits of the returned set.
        n_cand = int(tri_edge_max.shape[0])
        if n_cand <= _HELIOS_MAX_RETURN_TRIANGLES:
            keep = np.ones(n_cand, dtype=bool)
            cap_lmax = float(tri_edge_max.max()) if n_cand else 0.0
            cap_aspect = 1.0e9
        else:
            cap_aspect = 4.0
            asp_ok = tri_aspect <= cap_aspect
            edges_ok = tri_edge_max[asp_ok]
            if edges_ok.size > _HELIOS_MAX_RETURN_TRIANGLES:
                # Lmax of the Nth-smallest aspect-passing edge → keeps ~N triangles.
                cap_lmax = float(np.partition(edges_ok, _HELIOS_MAX_RETURN_TRIANGLES)[_HELIOS_MAX_RETURN_TRIANGLES])
            else:
                cap_lmax = float(edges_ok.max()) if edges_ok.size else 0.0
            keep = asp_ok & (tri_edge_max <= cap_lmax)

        tri_v = tri_v[keep]
        kept_scan_ids = scan_ids_all[keep]
        if tri_v.shape[0] == 0:
            return {
                "success": False, "vertices": [], "triangles": [],
                "num_triangles": 0, "num_vertices": 0, "method_used": "helios",
                "error": "No triangles survived the payload cap.",
                "diagnostics": diag, "grid_warning": grid_warning,
                "grid_message": grid_message,
            }

        # Dedup vertices in numpy. Round to 5 dp first to match the old hash-dedup
        # (100 µm), so shared edges/vertices collapse to one index. np.unique
        # returns the inverse map, which becomes the (T,3) triangle index list.
        # The big arrays are kept as numpy (not .tolist()) — the endpoint packs
        # them straight into the binary frame, skipping the list/JSON round-trip.
        _report(0.85, "Deduplicating vertices")
        all_verts = tri_v.reshape(-1, 3)
        rounded = np.round(all_verts, 5)
        unique_arr, inverse = np.unique(rounded, axis=0, return_inverse=True)
        # ravel() guards against numpy versions that return inverse as (N,1).
        triangles_arr = inverse.ravel().reshape(-1, 3)

        # Surface area from deduplicated data.
        v0 = unique_arr[triangles_arr[:, 0]]
        v1 = unique_arr[triangles_arr[:, 1]]
        v2 = unique_arr[triangles_arr[:, 2]]
        total_area = float(0.5 * np.linalg.norm(np.cross(v1 - v0, v2 - v0), axis=1).sum())

        # Per-triangle grid cell: bin each triangle's centroid into the request
        # grid (the same grid Helios triangulated within). Lets the UI split the
        # leaf-angle distribution per voxel. With the auto 1x1x1 grid this is all
        # zeros. -1 (outside) is preserved; the binary packer casts it to the
        # uint32 sentinel 0xffffffff the renderer already treats as "outside".
        _report(0.93, "Binning into grid")
        centroids = (v0 + v1 + v2) / 3.0
        triangle_cell_ids = _bin_points_to_cells(
            centroids, grid_center, grid_size, grid_nx, grid_ny, grid_nz
        )

        _report(1.0, "Finalizing")
        return {
            "success": True,
            "vertices": unique_arr,
            "triangles": triangles_arr,
            "surface_area": total_area,
            "num_triangles": int(triangles_arr.shape[0]),
            "num_vertices": int(unique_arr.shape[0]),
            "method_used": "helios",
            "triangle_scan_ids": kept_scan_ids,
            "triangle_cell_ids": triangle_cell_ids,
            # Loosening limits of the returned set + the auto-estimate, so the
            # front-end can seed and bound its interactive filter.
            "cap_lmax": cap_lmax,
            "cap_aspect": cap_aspect,
            "candidate_count": n_cand,
            "estimate": estimate,
            "diagnostics": diag,
            "grid_warning": grid_warning,
            "grid_message": grid_message,
        }

    except ImportError as e:
        return {
            "success": False,
            "vertices": [],
            "triangles": [],
            "num_triangles": 0,
            "num_vertices": 0,
            "method_used": "helios",
            "error": f"PyHelios not installed: {str(e)}"
        }
    except ScanCancelled:
        raise  # cancellation propagates to the streaming wrapper (memory freed)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "vertices": [],
            "triangles": [],
            "num_triangles": 0,
            "num_vertices": 0,
            "method_used": "helios",
            "error": f"Helios triangulation failed: {str(e)}"
        }
    finally:
        if tmpdir and os.path.isdir(tmpdir):
            shutil.rmtree(tmpdir, ignore_errors=True)


def _pack_helios_triangulation(result: dict) -> bytes:
    """Pack a _do_helios_computation result dict into a PHB1 binary frame: the
    big arrays (vertices, triangles, per-triangle scan/cell ids) become buffers,
    everything else (counts, estimate, cap, diagnostics, warnings) goes in meta."""
    if not result.get("success"):
        return _bin_frame_bytes({
            "success": False,
            "error": result.get("error"),
            "diagnostics": result.get("diagnostics"),
            "grid_warning": result.get("grid_warning", False),
            "grid_message": result.get("grid_message"),
        }, [])

    meta = {
        "success": True,
        "method_used": result.get("method_used", "helios"),
        "num_triangles": result["num_triangles"],
        "num_vertices": result["num_vertices"],
        "surface_area": result.get("surface_area"),
        "cap_lmax": result.get("cap_lmax"),
        "cap_aspect": result.get("cap_aspect"),
        "candidate_count": result.get("candidate_count"),
        "estimate": result.get("estimate"),
        "diagnostics": result.get("diagnostics"),
        "grid_warning": result.get("grid_warning", False),
        "grid_message": result.get("grid_message"),
    }
    buffers = [
        ("vertices", result["vertices"], "f32"),
        ("triangles", result["triangles"], "u32"),
        # -1 (outside grid) wraps to the 0xffffffff sentinel the renderer expects.
        ("triangle_scan_ids", result["triangle_scan_ids"], "u32"),
        ("triangle_cell_ids", np.asarray(result["triangle_cell_ids"]).astype(np.int64) & 0xFFFFFFFF, "u32"),
    ]
    return _bin_frame_bytes(meta, buffers)


@app.post("/api/triangulate/helios")
async def helios_triangulate(request: HeliosTriangulationRequest, http_request: Request):
    """Triangulate point cloud data using PyHelios spherical Delaunay triangulation.

    Returns a PHB1 binary frame (see _bin_frame_bytes) so multi-million-triangle
    meshes transfer compactly and parse as zero-copy typed arrays. Keepalive
    chunks during the (long) computation keep WebKit's stall timeout at bay.
    """
    run_id, cancel_event = _new_cancel_token()
    return _bin_frame_streaming_response(
        lambda progress: _pack_helios_triangulation(_do_helios_computation(request, progress=progress)),
        request=http_request, cancel_event=cancel_event, run_id=run_id)


# Lmax / median-spacing ratio at or above which we call the triangulation "likely
# bridging". A clean surface triangulated near its point spacing sits around 1-1.5x
# (the self-test's 0.04 Lmax over ~0.03 spacing is ~1.3x); a bridge-contaminated
# run like the leafcube_multi auto-estimate is ~7x. 3x is the conservative middle.
_SPACING_BRIDGE_RATIO = 3.0


def _resolve_scan_positions(scan_entry, warnings: list) -> "np.ndarray":
    """Surviving (N,3) positions for one scan, by the same source priority as the
    triangulation/LAD paths: live session (honoring unbaked deletions, never
    re-reading the source) -> inline points -> source file. Positions only — the
    spacing check needs no ray directions or multi-return columns."""
    import numpy as np

    sess = None
    if scan_entry.session_id:
        with _cloud_session_lock:
            sess = _cloud_sessions.get(scan_entry.session_id)
        if sess is None and scan_entry.file_path:
            warnings.append(
                "The edited point-cloud session was no longer available (the "
                "backend likely restarted), so the spacing check used the source "
                "file on disk. Unbaked deletions were not applied."
            )
        elif sess is None:
            raise ValueError(
                f"Cloud session not found: {scan_entry.session_id}. The backend "
                "may have restarted since import. Re-import the scan and try again."
            )

    if sess is not None:
        with _cloud_session_lock:
            return np.ascontiguousarray(sess.positions[~sess.deleted], dtype=np.float64)
    if scan_entry.points:
        return np.asarray(scan_entry.points, dtype=np.float64).reshape(-1, 3)
    if scan_entry.file_path:
        if not os.path.isfile(scan_entry.file_path):
            raise ValueError(f"Scan file not found: {scan_entry.file_path}")
        xyz, _dirs, _labels, _vals, _flags = _file_to_lad_arrays(
            scan_entry.file_path, scan_entry.ascii_format, scan_entry.origin)
        return np.asarray(xyz, dtype=np.float64)
    raise ValueError("Scan entry has no points, file_path, or session_id")


def _points_inside_grid(xyz: "np.ndarray", grid_center, grid_size,
                        grid_rotation_deg: float = 0.0) -> "np.ndarray":
    """Boolean mask for points STRICTLY inside the grid's overall box.

    Distinct from `_cull_to_grid` (beam-intersects-AABB, which deliberately keeps
    far miss rays for Beer's law): for the spacing check we want only the points
    that actually lie on the reconstructed surface, so a far sky/miss point at
    max range — whose beam crosses the grid but whose position is nowhere near it
    — is excluded. Otherwise its nearest neighbor would be another distant miss,
    poisoning the spacing estimate. Uses the grid's full extent (all cells), which
    matches the region the triangulation tags with a gridCell — including the
    grid's azimuthal rotation, so the spacing is measured over the SAME rotated
    box the triangulation crops to (inverse-rotate into the box frame, then AABB,
    exactly as Helios's calculateHitGridCell does)."""
    import numpy as np

    c = np.asarray(grid_center, dtype=np.float64)
    half = np.asarray(grid_size, dtype=np.float64) / 2.0
    lo = c - half
    hi = c + half
    if abs(float(grid_rotation_deg)) > 1e-9:
        theta = -np.radians(float(grid_rotation_deg))  # inverse rotation about +z
        cos_t, sin_t = np.cos(theta), np.sin(theta)
        dx = xyz[:, 0] - c[0]
        dy = xyz[:, 1] - c[1]
        local = np.empty_like(xyz)
        local[:, 0] = cos_t * dx - sin_t * dy + c[0]
        local[:, 1] = sin_t * dx + cos_t * dy + c[1]
        local[:, 2] = xyz[:, 2]
        return np.all((local >= lo) & (local <= hi), axis=1)
    return np.all((xyz >= lo) & (xyz <= hi), axis=1)


def _do_spacing_check(request: HeliosTriangulationRequest) -> dict:
    """Measure in-cell point spacing and judge the request's Lmax against it.

    See SpacingCheckResponse for the why. Pools the surviving in-cell points
    across all scans, computes the median nearest-neighbor distance via a KD-tree,
    and flags the run as likely-bridging when Lmax exceeds that spacing by
    `_SPACING_BRIDGE_RATIO`."""
    import numpy as np
    from scipy.spatial import cKDTree

    warnings: list = []

    if request.grid is not None:
        grid_center = list(request.grid.center)
        grid_size = list(request.grid.size)
        grid_rotation_deg = float(request.grid.rotation)
    else:
        # No explicit grid: the triangulation auto-boxed all points, so the
        # spacing check measures the whole cloud (every point is "in-cell").
        grid_center = grid_size = None
        grid_rotation_deg = 0.0

    pooled = []
    for scan_entry in request.scans:
        xyz = _resolve_scan_positions(scan_entry, warnings)
        if xyz.size == 0:
            continue
        if grid_center is not None:
            xyz = xyz[_points_inside_grid(xyz, grid_center, grid_size, grid_rotation_deg)]
        if xyz.size:
            pooled.append(xyz)

    if not pooled:
        return {
            "success": False,
            "lmax": request.lmax,
            "n_points": 0,
            "error": "No points fall inside the grid — nothing to measure spacing on.",
        }

    pts = np.concatenate(pooled, axis=0)
    n = pts.shape[0]
    if n < 2:
        return {
            "success": False,
            "lmax": request.lmax,
            "n_points": int(n),
            "error": "Too few in-grid points to measure a nearest-neighbor spacing.",
        }

    tree = cKDTree(pts)
    # k=2: [0] is the point itself (distance 0), [1] is its nearest neighbor.
    dist, _ = tree.query(pts, k=2, workers=-1)
    nn = dist[:, 1]
    nn = nn[np.isfinite(nn) & (nn > 0)]
    if nn.size == 0:
        return {
            "success": False,
            "lmax": request.lmax,
            "n_points": int(n),
            "error": "Nearest-neighbor distances were all zero (duplicate points?).",
        }

    median_spacing = float(np.median(nn))
    ratio = request.lmax / median_spacing if median_spacing > 0 else None
    likely = ratio is not None and ratio >= _SPACING_BRIDGE_RATIO

    if likely:
        message = (
            f"Lmax ({request.lmax:g} m) is {ratio:.1f}x the median point spacing "
            f"({median_spacing:.3g} m) inside the grid. The triangulation is likely "
            f"bridging across a sparsely-sampled surface, which corrupts the leaf "
            f"normals and G(theta). Try lowering Lmax toward the point spacing."
        )
    elif ratio is not None:
        message = (
            f"Lmax ({request.lmax:g} m) is {ratio:.1f}x the median point spacing "
            f"({median_spacing:.3g} m) inside the grid — within the normal range, "
            f"so bridging is unlikely to be distorting the result."
        )
    else:
        message = f"Median point spacing inside the grid is {median_spacing:.3g} m."

    return {
        "success": True,
        "median_spacing": median_spacing,
        "lmax": request.lmax,
        "ratio": ratio,
        "n_points": int(n),
        "likely_bridging": likely,
        "message": message,
        "warnings": warnings,
    }


@app.post("/api/triangulate/check-spacing", response_model=SpacingCheckResponse)
async def triangulate_check_spacing(request: HeliosTriangulationRequest):
    """Cross-check the auto-estimated Lmax against the actual in-grid point spacing.

    An OPT-IN diagnostic (the renderer offers it as a button when the Otsu
    indicators aren't both High). Potentially expensive — a KD-tree over up to
    tens of millions of points — so it streams keepalive whitespace to survive
    WebKit's ~60s stall timeout, then yields the JSON
    verdict. Reuses HeliosTriangulationRequest so the renderer sends the exact
    scans + grid + lmax it triangulated with."""
    import asyncio

    def compute_and_serialize():
        return json.dumps(_do_spacing_check(request))

    async def stream_result():
        loop = asyncio.get_event_loop()
        future = loop.run_in_executor(None, compute_and_serialize)
        while not future.done():
            yield " "
            await asyncio.sleep(5)
        yield await future

    return StreamingResponse(stream_result(), media_type="application/json")



# Tokens the Helios ASCII loader stores into a hit's data map (the `else` branch
# of loadASCIIFile). Multi-return LAD needs these three present so Helios can
# group beams by pulse and gap-fill misses; without them isMultiReturnData() and
# gapfillMisses() cannot work.
_LAD_MULTI_RETURN_COLUMNS = ("timestamp", "target_index", "target_count")


def _directions_from_origin(xyz: "np.ndarray", origin) -> "np.ndarray":
    """Reconstruct per-hit ray directions as Helios's loadASCIIFile does when a
    scan has no zenith/azimuth columns: cart2sphere(xyz - origin).

    `origin` is either a single (3,) scanner position (static scan — every beam
    shares it) or a per-beam (N,3) array of emission origins (moving-platform
    scan — each return has its own origin, joined from the trajectory by
    timestamp). Numpy broadcasting handles both; with all origins equal the
    per-beam form is bit-identical to the single-origin form.

    Returns (N,3) float32 [radius, elevation, azimuth] matching helios
    cart2sphere (global.cpp): radius = ||d||, elevation = asin(dz/radius),
    azimuth = atan2_2pi(dx, dy) which equals np.arctan2(dx, dy) (x first, y
    second; signed, no 2*pi wrap). Verified against getHitRaydir to ~1e-7.
    """
    import numpy as np

    d = np.asarray(xyz, dtype=np.float64) - np.asarray(origin, dtype=np.float64)
    r = np.linalg.norm(d, axis=1)
    rs = np.where(r > 0, r, 1.0)
    elevation = np.arcsin(np.clip(d[:, 2] / rs, -1.0, 1.0))
    azimuth = np.arctan2(d[:, 0], d[:, 1])
    return np.column_stack([r, elevation, azimuth]).astype(np.float32)


def _cull_to_grid(xyz: "np.ndarray", origin, grid_center, grid_size,
                  expand: float = 0.05) -> "np.ndarray":
    """Boolean keep-mask for points whose beam (origin -> point) can pass through
    the grid's (expanded) AABB.

    LAD needs through-grid MISS rays for Beer's law, so cull by "segment
    origin->point intersects the grid AABB", NOT "point is inside the grid": a
    far miss whose beam grazes the small grid still contributes transmittance and
    must be kept. Slab test over t in [0, 1]; axis-parallel beams (d==0) are kept
    only if the origin already lies within that axis's slab. `expand` adds a
    small margin so beams grazing the grid face (finite footprint) aren't dropped.

    `origin` is either a single (3,) position (static) or per-beam (N,3) origins
    (moving platform); the slab math broadcasts identically for both, and the
    axis-parallel `origin_in` test becomes per-beam when origins are per-beam.
    """
    import numpy as np

    o = np.asarray(origin, dtype=np.float64)
    half = np.asarray(grid_size, dtype=np.float64) / 2.0 + expand
    lo = np.asarray(grid_center, dtype=np.float64) - half
    hi = np.asarray(grid_center, dtype=np.float64) + half

    d = np.asarray(xyz, dtype=np.float64) - o  # segment vector (length = range)
    with np.errstate(divide="ignore", invalid="ignore"):
        t1 = (lo - o) / d
        t2 = (hi - o) / d
    tmin = np.nanmax(np.minimum(t1, t2), axis=1)
    tmax = np.nanmin(np.maximum(t1, t2), axis=1)

    # Axis-parallel beams: d==0 on an axis -> only inside that slab if the origin
    # is within [lo, hi] on it. Otherwise the beam can never enter the box.
    par = d == 0.0
    origin_in = (o >= lo) & (o <= hi)
    bad_par = np.any(par & ~origin_in, axis=1)

    keep = (tmax >= np.maximum(tmin, 0.0)) & (tmin <= 1.0) & ~bad_par
    return keep


# Flags a LAD array-builder reports back about the resolved scan data, so
# `_do_lad_computation` can decide whether to gapfill (timestamp present, no
# misses yet), whether the cloud already carries miss points (skip gapfill), and
# whether to warn that the inversion will be inaccurate (no misses, no timestamp).
def _lad_flags(has_timestamp: bool, is_multi: bool, has_misses: bool,
               has_grid: bool = False) -> dict:
    return {"has_timestamp": has_timestamp, "multi": is_multi,
            "has_misses": has_misses, "has_grid": has_grid}


def _lad_labels_vals(column_getter, n: int):
    """Assemble the (labels, vals, flags) a LAD builder feeds Helios.

    `column_getter(slug)` returns the aligned (N,) float array for a slug or None.
    Always include `timestamp` when present (gapfillMisses() needs only it);
    include target_index/target_count too when ALL three exist (full multi-return
    path).

    `is_miss` is forwarded to Helios as a per-hit data value (0.0 = return,
    1.0 = miss) whenever the cloud carries the column AT ALL — not only when a
    miss is currently flagged. This is the canonical convention the C++
    `calculateLeafArea` fail-fast check reads to refuse a cloud with no sky/miss
    rays, so every return must arrive explicitly tagged 0.0 rather than relying on
    the label's absence. A cloud with no `is_miss` column (e.g. plain XYZ) does
    NOT get a synthesised one — those recover misses via gapfillMisses(), which
    sets the same flag C++-side. Returns labels (list[str]), vals ((N,k)|None),
    flags (see `_lad_flags`).
    """
    import numpy as np

    ts = column_getter('timestamp')
    has_timestamp = ts is not None
    is_multi = has_timestamp and all(
        column_getter(c) is not None for c in ('target_index', 'target_count'))
    miss = column_getter(_MISS_SLUG)
    has_miss_column = miss is not None
    has_misses = has_miss_column and bool(np.any(np.asarray(miss) != 0))

    labels: List[str] = []
    cols: list = []
    if is_multi:
        for c in _LAD_MULTI_RETURN_COLUMNS:
            labels.append(c)
            cols.append(np.asarray(column_getter(c), dtype=np.float64))
    elif has_timestamp:
        labels.append('timestamp')
        cols.append(np.asarray(ts, dtype=np.float64))
    # Forward is_miss for every hit whenever the column exists, so returns carry
    # an explicit 0.0 and the C++ check can count misses by flag, not geometry.
    if has_miss_column:
        labels.append(_MISS_SLUG)
        cols.append(np.asarray(miss, dtype=np.float64))
    # Forward the structured-grid indices when present so the C++ grid-based
    # direction recovery (for unplaceable misses) has the raster to interpolate
    # over. Carried only when the source provided them (E57 row/column index).
    has_grid = (column_getter('row_index') is not None
                and column_getter('column_index') is not None)
    for grid_slug in ('row_index', 'column_index'):
        gcol = column_getter(grid_slug)
        if gcol is not None:
            labels.append(grid_slug)
            cols.append(np.asarray(gcol, dtype=np.float64))

    vals = np.column_stack(cols).astype(np.float64) if cols else None
    return labels, vals, _lad_flags(has_timestamp, is_multi, has_misses, has_grid)


def _session_to_lad_arrays(sess: "CloudSession", origin, include_backfilled: bool = True):
    """Surviving session points as in-RAM arrays for the LAD path — no disk, no
    source-file read.

    Returns (xyz float64 (N,3), dirs float32 (N,3), labels list[str],
    vals float64 (N,k)|None, flags). Honors ~sess.deleted. `flags` (see
    `_lad_flags`) tells the caller whether to gapfill / warn.

    When `include_backfilled` is True (the default — the LAD read path) and the
    session has an explicit miss buffer (see CloudSession.backfilled_misses), its
    rows are APPENDED as extra hits tagged `is_miss`=1, and `has_misses` is set
    True. The backfill endpoint itself passes False so it gapfills over the raw
    hits only, never re-feeding already-recovered misses.
    """
    import numpy as np

    keep = ~sess.deleted
    xyz = np.ascontiguousarray(sess.positions[keep], dtype=np.float64)
    dirs = _directions_from_origin(xyz, origin)

    def _get(slug):
        # The per-point timestamp lives on the dedicated float64 `sess.timestamps`
        # field (not float32 `extras`) to preserve GPS-time precision for the
        # trajectory join — prefer it. The length guard tolerates any mutation site
        # that fails to re-slice it (degrade to the float32 extra rather than
        # misalign). Fall back to `extras` for the ASCII import path, which still
        # routes timestamp through the column plan.
        if (slug == 'timestamp' and sess.timestamps is not None
                and sess.timestamps.shape[0] == sess.positions.shape[0]):
            return np.asarray(sess.timestamps, dtype=np.float64)[keep]
        return sess.extras[slug][keep] if slug in sess.extras else None

    labels, vals, flags = _lad_labels_vals(_get, xyz.shape[0])

    # Per-pulse beam origins from LAS ExtraBytes (ground truth) — surfaced to the
    # LAD caller so it can use them directly and bypass the timestamp join. Subset
    # by `keep` to stay aligned with `xyz`. Carried on `flags` (which already flows
    # out) rather than widening the return tuple. None unless the LAS carried them.
    flags["beam_origins"] = (
        np.ascontiguousarray(sess.beam_origins[keep], dtype=np.float64)
        if getattr(sess, "beam_origins", None) is not None else None
    )

    backfilled = sess.backfilled_misses if include_backfilled else None
    if (backfilled is not None and backfilled.get("positions") is not None
            and np.asarray(backfilled["positions"]).shape[0] > 0):
        xyz, dirs, labels, vals, flags = _append_backfilled_misses(
            xyz, dirs, labels, vals, flags, backfilled)
        # The backfilled buffer was computed against the hits as they were at
        # backfill time. If a later crop removed hits, the buffer is stale (its
        # miss ratio no longer matches the surviving hits) — surface that so the
        # LAD endpoint can warn. Only meaningful when the buffer is actually used.
        flags["misses_stale"] = bool(getattr(sess, "backfilled_misses_stale", False))

    return xyz, dirs, labels, vals, flags


def _append_backfilled_misses(xyz, dirs, labels, vals, flags, backfilled):
    """Append a session's backfilled-miss buffer to assembled LAD hit arrays.

    Hits carry `is_miss`=0; the appended misses carry `is_miss`=1 (creating the
    column if the hits lacked it). Direction rows come from the buffer; timestamp/
    origin columns are filled where both the hit arrays and the buffer carry them
    (else 0 for the miss rows, which Helios ignores for sky points). Returns the
    extended (xyz, dirs, labels, vals, flags) with `has_misses` True.
    """
    import numpy as np

    m_xyz = np.ascontiguousarray(backfilled["positions"], dtype=np.float64)
    m_dirs = np.ascontiguousarray(
        backfilled.get("directions", _directions_from_origin(m_xyz, m_xyz)),
        dtype=np.float32)
    n_hits = xyz.shape[0]
    n_miss = m_xyz.shape[0]

    labels = list(labels)
    # Materialise hit vals as a (N,k) float64 matrix (or an empty (N,0) when the
    # hits had no columns) so we can stack miss rows column-aligned.
    hit_vals = (np.asarray(vals, dtype=np.float64) if vals is not None
                else np.empty((n_hits, 0), np.float64))
    if _MISS_SLUG not in labels:
        labels.append(_MISS_SLUG)
        hit_vals = np.column_stack([hit_vals, np.zeros(n_hits, np.float64)])

    miss_cols = []
    for slug in labels:
        if slug == _MISS_SLUG:
            miss_cols.append(np.ones(n_miss, np.float64))
        elif slug == 'timestamp' and backfilled.get("timestamp") is not None:
            miss_cols.append(np.asarray(backfilled["timestamp"], dtype=np.float64))
        elif slug in ('origin_x', 'origin_y', 'origin_z') and backfilled.get("origins") is not None:
            axis = {'origin_x': 0, 'origin_y': 1, 'origin_z': 2}[slug]
            miss_cols.append(np.asarray(backfilled["origins"], dtype=np.float64)[:, axis])
        else:
            miss_cols.append(np.zeros(n_miss, np.float64))
    miss_vals = np.column_stack(miss_cols) if miss_cols else np.empty((n_miss, 0), np.float64)

    xyz = np.vstack([xyz, m_xyz])
    dirs = np.vstack([dirs, m_dirs]).astype(np.float32)
    vals = np.vstack([hit_vals, miss_vals])
    flags = dict(flags)
    flags["has_misses"] = True
    return xyz, dirs, labels, vals, flags


def _inline_to_lad_arrays(points: list, scalar_columns: Optional[dict], origin):
    """Inline synthetic points (+ aligned scalar_columns) as LAD arrays. Same
    contract as _session_to_lad_arrays.
    """
    import numpy as np

    xyz = np.asarray(points, dtype=np.float64)
    if xyz.ndim != 2 or xyz.shape[1] < 3:
        raise ValueError("LAD points must be an (N, >=3) array")
    xyz = np.ascontiguousarray(xyz[:, :3])
    n = xyz.shape[0]
    dirs = _directions_from_origin(xyz, origin)

    def _get(slug):
        if scalar_columns and slug in scalar_columns and len(scalar_columns[slug]) == n:
            return np.asarray(scalar_columns[slug])
        return None

    labels, vals, flags = _lad_labels_vals(_get, n)
    return xyz, dirs, labels, vals, flags


def _file_to_lad_arrays(file_path: str, ascii_format: Optional[str], origin):
    """Legacy fallback: read an ASCII scan file once into LAD arrays (used only
    when a scan has neither a live session nor inline points — e.g. a stale
    session id that fell back to the source file). Locates x/y/z plus any
    timestamp/target/miss columns via the format string; never re-read after this.
    """
    import numpy as np

    fmt = ascii_format or _detect_ascii_format(file_path)
    tokens = fmt.split()
    xi, yi, zi = _xyz_column_indices(fmt)
    # Locate every per-pulse / miss column the format declares.
    wanted = list(_LAD_MULTI_RETURN_COLUMNS) + [_MISS_SLUG]
    col_idx = {c: tokens.index(c) for c in wanted if c in tokens}

    rows = []
    extra_rows: dict = {c: [] for c in col_idx}
    need = max([xi, yi, zi] + list(col_idx.values())) + 1
    with open(file_path) as f:
        for line in f:
            line = line.strip()
            if not line or line[0] in "#/":
                continue
            cols = line.split()
            if len(cols) < need:
                continue
            try:
                vals_xyz = (float(cols[xi]), float(cols[yi]), float(cols[zi]))
                parsed = {c: float(cols[col_idx[c]]) for c in col_idx}
            except (ValueError, IndexError):
                continue
            rows.append(vals_xyz)
            for c in col_idx:
                extra_rows[c].append(parsed[c])

    xyz = np.asarray(rows, dtype=np.float64) if rows else np.empty((0, 3), np.float64)
    dirs = _directions_from_origin(xyz, origin)

    def _get(slug):
        return np.asarray(extra_rows[slug], dtype=np.float64) if slug in col_idx else None

    labels, vals, flags = _lad_labels_vals(_get, xyz.shape[0])
    return xyz, dirs, labels, vals, flags


_GPS_WEEK_SECONDS = 604800.0
_GPS_STANDARD_OFFSET = 1e9  # Standard − Adjusted-Standard GPS seconds


def _validate_join_coverage(ts, traj_t, warnings=None):
    """Fail loudly when point timestamps don't line up with the trajectory's clock.

    `ts` (N,) and `traj_t` (M,) are float64 seconds in whatever clock each carries.
    The trajectory join (`origins_for_returns` → `resolve_pose_at`) CLAMPS query
    times to [t0, t1], which is correct for small legitimate edge overruns but
    catastrophic for a clock mismatch: every return clamps to one endpoint pose and
    silently collapses to a single static origin. This validates the join BEFORE
    that happens:
      - zero overlap                 → raise (almost always an epoch/unit mismatch)
      - systematic ~1e9 / N-week gap → raise explicitly (never auto-correct)
      - 0 < coverage < ~98%          → append a warning (partial; clamped tail)
      - NaN timestamps               → raise
    A single-pose trajectory (M == 1) is a degenerate "static" pose applied to all
    returns by design — skip the overlap check for it.
    """
    import numpy as np

    if ts.size == 0 or traj_t.size == 0:
        return
    if not np.all(np.isfinite(ts)):
        raise ValueError(
            "Moving-platform LAD: the per-point timestamp column contains NaN/inf "
            "values, so returns cannot be joined to the trajectory. Re-import the "
            "scan with a clean GPS-time column.")

    t0, t1 = float(traj_t[0]), float(traj_t[-1])
    if traj_t.size == 1 or t1 == t0:
        return  # single pose: applies to every return; no span to overlap

    ts_lo, ts_hi = float(np.min(ts)), float(np.max(ts))
    coverage = float(((ts >= t0) & (ts <= t1)).mean())

    # When the clocks don't line up, first try to NAME the specific GPS confusion
    # (a systematic ~1e9 s or integer-GPS-week offset between the medians) — a far
    # more actionable message than the generic "no overlap". Never auto-correct: a
    # wrong guess would silently shift every origin.
    offset = float(np.median(ts) - (t0 + t1) / 2.0)
    if abs(abs(offset) - _GPS_STANDARD_OFFSET) < 1.0:
        raise ValueError(
            f"Moving-platform LAD: the point timestamps lead the trajectory by "
            f"~{offset:.0f} s, the exact Standard − Adjusted-Standard GPS offset "
            f"(1e9 s). One side is Standard GPS time and the other Adjusted-Standard. "
            f"Re-export both on the same GPS-time encoding; the offset is not "
            f"auto-corrected.")
    if abs(offset) > _GPS_WEEK_SECONDS and (abs(offset) % _GPS_WEEK_SECONDS) < 1.0:
        weeks = round(abs(offset) / _GPS_WEEK_SECONDS)
        raise ValueError(
            f"Moving-platform LAD: the point timestamps and trajectory differ by "
            f"~{weeks} GPS week(s) ({offset:.0f} s). One side is likely GPS Week Time "
            f"and the other an absolute clock. Re-align them to a common epoch; the "
            f"offset is not auto-corrected.")

    if coverage == 0.0:
        raise ValueError(
            f"Moving-platform LAD: the point timestamps and the trajectory do not "
            f"overlap in time (points span [{ts_lo:.3f}, {ts_hi:.3f}] s, trajectory "
            f"spans [{t0:.3f}, {t1:.3f}] s). This usually means the two use different "
            f"clocks (e.g. GPS Week Time vs Adjusted-Standard, or a ~1e9 s epoch "
            f"offset). Re-export them on a common clock; the join is not auto-corrected.")

    if coverage < 0.98 and warnings is not None:
        warnings.append(
            f"Moving-platform LAD: only {coverage * 100:.1f}% of returns fall within "
            f"the trajectory's time span [{t0:.3f}, {t1:.3f}] s; the remaining returns "
            f"are clamped to the nearest trajectory endpoint and their per-beam "
            f"origins may be wrong. Check that the trajectory covers the whole scan.")


def _apply_trajectory_origins(xyz, dirs, labels, vals, trajectory, shift=None, warnings=None):
    """Convert a static LAD scan into a per-beam (moving-platform) one.

    Given a scan's resolved arrays plus a `trajectory` (Pydantic PoseStream) and an
    optional local-frame `shift` (3,), reconstruct a PER-BEAM emission origin for
    every return by joining its `timestamp` column to the trajectory (SLERP
    attitude, linear position; see trajectory.py), then:
      - recompute `dirs` as cart2sphere(xyz - per_beam_origin),
      - append origin_x/origin_y/origin_z to (labels, vals) so they land in each
        hit's data map — this is exactly what the C++ getHitOrigin() / beam-based
        (Gtheta) inversion reads.

    `xyz` and the trajectory must already be in the SAME frame; pass `shift` to move
    BOTH into a local frame for float32 precision (the inversion is frame-invariant).
    Returns (dirs, labels, vals, origins). Raises if no `timestamp` column exists —
    a moving scan cannot be reconstructed without the per-return join key.

    Before joining, the per-point timestamps are checked for OVERLAP with the
    trajectory's time span. With no overlap (e.g. a GPS epoch/unit mismatch) every
    return would clamp to a single endpoint pose and silently collapse to one static
    origin — so this RAISES instead. Partial coverage appends a warning to `warnings`
    (a list, when provided) rather than blocking. A systematic ~1e9 (Standard vs
    Adjusted-Standard) or integer-GPS-week offset is refused explicitly rather than
    auto-corrected (a wrong auto-correct silently corrupts every origin).
    """
    import numpy as np
    from trajectory import PoseStream as _TrajPoseStream, origins_for_returns

    if 'timestamp' not in labels:
        raise ValueError(
            "Moving-platform LAD requires a per-point 'timestamp' column to join "
            "returns to the trajectory, but the scan data has none.")
    ts = np.asarray(vals[:, labels.index('timestamp')], dtype=np.float64)

    poses = trajectory.poses
    stream = _TrajPoseStream.from_samples(
        t=[p.t for p in poses],
        pos=[[p.x, p.y, p.z] for p in poses],
        rot=[[p.qx, p.qy, p.qz, p.qw] for p in poses],
        rot_is_quaternion=True,
        lever_arm=list(trajectory.lever_arm),
        boresight_rpy=list(trajectory.boresight_rpy),
        frame_crs=trajectory.frame.crs,
    )
    if shift is not None:
        stream = stream.recentered(np.asarray(shift, dtype=np.float64))

    _validate_join_coverage(ts, np.asarray(stream.t, dtype=np.float64), warnings)

    origins = origins_for_returns(stream, ts)              # (N,3) float64, local frame
    dirs, labels, vals = _attach_origins(xyz, labels, vals, origins)
    return dirs, labels, vals, origins


def _attach_origins(xyz, labels, vals, origins):
    """Recompute per-beam ray directions from explicit emission `origins` and append
    origin_x/origin_y/origin_z to (labels, vals).

    Shared by the trajectory-join path (`_apply_trajectory_origins`) and the LAS
    ExtraBytes path (`_do_lad_computation`) so both produce the IDENTICAL wire shape
    the C++ getHitOrigin() / beam-based (Gtheta) inversion reads. `xyz` and `origins`
    must already be in the same frame. Returns (dirs, labels, vals).
    """
    import numpy as np

    dirs = _directions_from_origin(xyz, origins)           # per-beam directions
    labels = list(labels) + ['origin_x', 'origin_y', 'origin_z']
    origin_cols = np.asarray(origins, dtype=np.float64)
    vals = (np.column_stack([vals, origin_cols]) if vals is not None
            else origin_cols)
    return dirs, labels, vals


def _run_gapfill_extract(cloud):
    """Recover sky/miss points on a built PyHelios cloud and return the synthesised
    ones as (synth_xyz (M,3) float64, count int).

    gapfillMisses() (which auto-selects the row/column or timestamp path in C++)
    APPENDS the reconstructed misses to the cloud, each flagged as a miss. We slice
    them out via the BULK numpy getters — getHitMissArray() (the per-hit miss flag,
    1 == sky/miss) selects the rows, getHitsXYZRGBArrays() supplies their coords —
    never the per-hit getHitXYZ loop, which would be millions of FFI crossings at
    10M scale. Because we only call this on a cloud built from a scan with NO real
    misses (the endpoint guards on that), every miss-flagged row is a synthesised
    one. The real hits are left behind — the session already holds them.
    """
    import numpy as np

    cloud.gapfillMisses()
    miss_flag = np.asarray(cloud.getHitMissArray(), dtype=np.int32)
    xyz, _rgb = cloud.getHitsXYZRGBArrays()
    xyz = np.asarray(xyz, dtype=np.float64)
    if miss_flag.shape[0] != xyz.shape[0]:
        # Defensive: bulk getters must agree on hit count. If they don't, recover
        # nothing rather than mis-slice.
        return np.empty((0, 3), np.float64), 0
    mask = miss_flag != 0
    synth_xyz = np.ascontiguousarray(xyz[mask], dtype=np.float64)
    return synth_xyz, int(mask.sum())


def _sample_dem_columns(grid_center, grid_size, grid_nx, grid_ny, grid_nz,
                        grid_rotation_rad, dem: "DemRaster", safety_fraction: float):
    """Sample a DEM under each voxel column for terrain-following LAD.

    For each of the nx*ny columns of the (possibly rotated) grid, compute the
    column center's WORLD (x,y) — rotating the local lattice center by +rotation
    about grid_center, the forward of the inverse-rotation `_points_inside_grid` /
    `_count_points_per_cell` apply — then sample the axis-aligned DEM raster there.

    The per-column vertical offset shifts the regular lattice so the column's
    BOTTOM face sits at ``dem_z + safety_fraction*(size.z/nz)`` (clearance scales
    with one cell's height). Concretely the lattice z-origin is the grid bottom
    (grid_center.z - size.z/2), so the offset added to every cell in the column is
    ``dem_z + clearance - (grid_center.z - size.z/2)``.

    Void DEM cells (NaN) that lie INSIDE the raster footprint inherit the nearest
    valid DEM cell's elevation (so a hole doesn't drop the column). Columns whose
    center lies OUTSIDE the raster footprint are dropped (excluded from results).

    Returns (column_offsets, kept_mask, dropped_count):
      column_offsets : list[float] length nx*ny, row-major [j*nx + i]; dropped
                       columns get 0.0 (their cells are filtered from the response).
      kept_mask      : np.ndarray[bool] length nx*ny; False = dropped column.
      dropped_count  : int, number of dropped columns.
    """
    import numpy as np

    nx, ny, nz = int(grid_nx), int(grid_ny), int(grid_nz)
    cx, cy, cz = float(grid_center[0]), float(grid_center[1]), float(grid_center[2])
    sx, sy, sz = float(grid_size[0]), float(grid_size[1]), float(grid_size[2])

    # DEM raster as a (ny, nx)-ish 2D array, row-major min-y first.
    z2d = np.asarray(dem.grid_z, dtype=np.float64).reshape(dem.ny, dem.nx)
    # Decode the void sentinel (JSON-safe stand-in for NaN) back to NaN.
    if dem.nodata is not None:
        z2d = np.where(z2d == float(dem.nodata), np.nan, z2d)
    cell = float(dem.cell)
    ox, oy = float(dem.origin[0]), float(dem.origin[1])

    # Precompute a nearest-valid lookup over the DEM so in-footprint voids inherit
    # the closest measured elevation. distance_transform_edt returns, for each cell,
    # the index of the nearest non-void cell — one pass, vectorized.
    valid = np.isfinite(z2d)
    if valid.all():
        filled = z2d
    elif valid.any():
        from scipy import ndimage
        inv = ~valid
        # indices of nearest valid cell for every (void) cell
        nearest_idx = ndimage.distance_transform_edt(
            inv, return_distances=False, return_indices=True)
        filled = z2d[tuple(nearest_idx)]
    else:
        # DEM is entirely void — nothing to follow; treat all columns as dropped.
        filled = z2d

    cos_t, sin_t = math.cos(grid_rotation_rad), math.sin(grid_rotation_rad)
    grid_bottom_z = cz - 0.5 * sz
    clearance = float(safety_fraction) * (sz / nz)

    column_offsets = [0.0] * (nx * ny)
    kept_mask = np.zeros(nx * ny, dtype=bool)
    dropped = 0

    for j in range(ny):
        ly = -0.5 * sy + (j + 0.5) * (sy / ny)
        for i in range(nx):
            lx = -0.5 * sx + (i + 0.5) * (sx / nx)
            # forward-rotate the local lattice center about +z, then translate to world
            wx = cx + (lx * cos_t - ly * sin_t)
            wy = cy + (lx * sin_t + ly * cos_t)
            di = int(math.floor((wx - ox) / cell))
            dj = int(math.floor((wy - oy) / cell))
            col = j * nx + i
            if di < 0 or di >= dem.nx or dj < 0 or dj >= dem.ny:
                dropped += 1
                continue
            dem_z = filled[dj, di]
            if not np.isfinite(dem_z):
                # in-footprint but no valid neighbor anywhere (fully-void DEM)
                dropped += 1
                continue
            column_offsets[col] = float(dem_z) + clearance - grid_bottom_z
            kept_mask[col] = True

    return column_offsets, kept_mask, dropped


def _do_lad_computation(request: "LADComputeRequest", progress=None,
                        reuse_mesh: "Optional[Tuple[np.ndarray, np.ndarray, np.ndarray]]" = None) -> dict:
    """Compute per-voxel leaf area density via PyHelios. Returns a result dict.

    `reuse_mesh`, when supplied, is a (vertices, indices, scan_ids) triple from a
    previously-run Helios triangulation that the caller wants to REUSE instead of
    re-triangulating: vertices is (V, 3) float32 world coords, indices is (T, 3)
    uint32 triangle vertex indices, scan_ids is (T,) per-triangle source-scan
    index already remapped to this request's scan order. On the static path it is
    injected via cloud.setExternalTriangulation, skipping the (potentially
    minutes-long) Delaunay recompute. It is invalid for moving-platform scans
    (which can't be triangulated) and is rejected there.

    LAD is NOT a sum of triangle areas. The triangulation supplies the per-cell
    G-function (the leaf-projection coefficient for Beer's law); Helios then
    traces beam paths through the voxel grid and inverts Beer's law per voxel to
    recover leaf area density. Unlike triangulation, the grid is required.

    Point data is fed to Helios straight from the in-RAM session arrays via the
    bulk addHitPointsWithData FFI — no ASCII file, no disk round-trip — for both
    single- and multi-return (the data map carries timestamp/target_index/
    target_count so isMultiReturnData()/gapfillMisses() work). Points are first
    culled to the grid's beam frustum, then ray directions are reconstructed as
    cart2sphere(xyz - origin) to match the old loadASCIIFile path. All scans
    share ONE LiDARcloud and ONE grid, since beam tracing and Beer's-law
    inversion are per-voxel over the whole grid.
    """
    import math
    import os
    import numpy as np

    def _report(fraction, message):
        if progress is not None:
            progress(fraction, message)

    def _ckpt():
        _cancel_checkpoint(progress)

    warnings: List[str] = []
    try:
        from pyhelios import LiDARCloud, Context

        if request.grid is None:
            raise ValueError("A voxel grid is required for leaf area density.")

        # Per-scan angular geometry / resolution: prefer the scan's own values,
        # fall back to request-level (mirrors _do_helios_computation).
        def _angles(scan_entry):
            return (
                scan_entry.theta_min if scan_entry.theta_min is not None else request.theta_min,
                scan_entry.theta_max if scan_entry.theta_max is not None else request.theta_max,
                scan_entry.phi_min if scan_entry.phi_min is not None else request.phi_min,
                scan_entry.phi_max if scan_entry.phi_max is not None else request.phi_max,
            )

        def _resolution(scan_entry, n_points, theta_span, phi_span):
            if scan_entry.n_theta and scan_entry.n_phi:
                return int(scan_entry.n_theta), int(scan_entry.n_phi)
            aspect = theta_span / max(phi_span, 1e-10)
            n_phi = max(int(math.sqrt(n_points / max(aspect, 0.01))), 10)
            n_theta = max(int(n_points / n_phi), 10)
            return n_theta, n_phi

        grid_center = list(request.grid.center)
        grid_size = list(request.grid.size)
        grid_nx, grid_ny, grid_nz = request.grid.nx, request.grid.ny, request.grid.nz
        # Azimuthal rotation of the grid box about +z (degrees in the request;
        # Helios's addGrid takes radians). The rotation pivots about grid_center,
        # so the lad_shift recenter below (a pure translation) leaves it unchanged.
        grid_rotation_deg = float(request.grid.rotation)
        grid_rotation_rad = math.radians(grid_rotation_deg)

        # Moving-platform scans carry a 6-DOF trajectory; their per-beam origins are
        # reconstructed and fed to PyHelios's float32 geometry path. If the
        # trajectory is in a large (e.g. UTM) frame, float32 loses decimeters, so
        # recenter EVERYTHING — points, trajectory, grid, and (on the way out) the
        # result cell centers — by a single local shift. The shift is the floor of
        # the min corner over the trajectory samples and the grid box, so all
        # coordinates reaching Helios are small. Static-only requests skip this
        # entirely (lad_shift is None) and are byte-for-byte unchanged.
        # A scan is "moving" if it carries a trajectory OR its session holds
        # per-pulse beam origins (LAS ExtraBytes). Both reconstruct per-beam origins
        # in the float32 geometry path and so need the local-frame recenter when the
        # coordinates are large (UTM). Peek at sessions here (read-only) so an
        # ExtraBytes-only scan with NO trajectory still contributes to the shift floor.
        def _scan_beam_origins(s):
            if not s.session_id:
                return None
            with _cloud_session_lock:
                _s = _cloud_sessions.get(s.session_id)
            return getattr(_s, "beam_origins", None) if _s is not None else None

        scan_origin_mins = []
        for s in request.scans:
            bo = _scan_beam_origins(s)
            if bo is not None and len(bo):
                scan_origin_mins.append(np.min(np.asarray(bo, dtype=np.float64), axis=0).tolist())
        any_moving = (any(s.trajectory is not None for s in request.scans)
                      or bool(scan_origin_mins))
        lad_shift = None
        if any_moving:
            mins = list(scan_origin_mins)
            for s in request.scans:
                if s.trajectory is not None:
                    for p in s.trajectory.poses:
                        mins.append([p.x, p.y, p.z])
            gc = np.asarray(grid_center, dtype=np.float64)
            gh = np.asarray(grid_size, dtype=np.float64) / 2.0
            mins.append((gc - gh).tolist())
            shift = np.floor(np.min(np.asarray(mins, dtype=np.float64), axis=0))
            # Only bother when coordinates are actually large; a near-origin scene
            # keeps its coordinates so tests/fixtures read naturally.
            if np.any(np.abs(shift) > 1.0):
                lad_shift = shift
                grid_center = (gc - lad_shift).tolist()

        # Each scan resolves to in-RAM arrays (xyz + ray directions + an optional
        # multi-return data map). Multi-return is detected AUTHORITATIVELY from
        # whether the resolved data actually carries the three per-pulse columns —
        # never from the `return_type` label alone. Source priority:
        #   1. session_id  -> surviving in-RAM points (+ multi-return columns when
        #      the session holds them), honoring unbaked deletions, never
        #      re-reading the source file.
        #   2. points (+ scalar_columns) -> a flat in-RAM cloud (e.g. a synthetic
        #      full-waveform scan).
        #   3. file_path -> read the ASCII file from disk once (legacy fallback).
        # Points are then culled to the grid's beam frustum before ingest.
        scans_arrays = []
        scan_xyz_for_counts = []
        n_scans = max(len(request.scans), 1)
        _report(0.05, "Reading scans")
        for scan_idx, scan_entry in enumerate(request.scans):
            # Step 0.05 → 0.35 across scans so multi-scan ingest visibly advances.
            _ckpt()
            _report(0.05 + 0.30 * (scan_idx / n_scans),
                    f"Reading scan {scan_idx + 1} of {n_scans}")
            origin = scan_entry.origin
            if len(origin) != 3:
                raise ValueError(f"Origin must have 3 elements, got {len(origin)}")

            theta_min, theta_max, phi_min, phi_max = _angles(scan_entry)

            # Resolve the live session if one was requested. A session is an
            # in-memory object that does NOT survive a backend restart, so the
            # renderer's session id can be stale (e.g. after a dev reload or a
            # respawn). When that happens, fall back to the source file the
            # renderer also sent — losing only unbaked deletions, which we warn
            # about — instead of failing the whole computation.
            sess = None
            if scan_entry.session_id:
                with _cloud_session_lock:
                    sess = _cloud_sessions.get(scan_entry.session_id)
                if sess is None and scan_entry.file_path:
                    warnings.append(
                        "The edited point-cloud session was no longer available "
                        "(the backend likely restarted), so LAD used the source "
                        "file on disk. Any unbaked deletions were not applied — "
                        "re-apply them and recompute if needed."
                    )
                elif sess is None:
                    raise ValueError(
                        f"Cloud session not found: {scan_entry.session_id}. The "
                        "backend may have restarted since import. Re-import the "
                        "scan and try again."
                    )

            if sess is not None:
                with _cloud_session_lock:
                    xyz, dirs, labels, vals, scan_flags = _session_to_lad_arrays(sess, origin)
            elif scan_entry.points:
                xyz, dirs, labels, vals, scan_flags = _inline_to_lad_arrays(
                    scan_entry.points, scan_entry.scalar_columns, origin)
            elif scan_entry.file_path:
                if not os.path.isfile(scan_entry.file_path):
                    raise ValueError(f"Scan file not found: {scan_entry.file_path}")
                xyz, dirs, labels, vals, scan_flags = _file_to_lad_arrays(
                    scan_entry.file_path, scan_entry.ascii_format, origin)
            else:
                raise ValueError("Scan entry has no points, file_path, or session_id")
            scan_multi = scan_flags["multi"]

            # Resolve a per-beam emission origin so the beam-based (Gtheta) inversion
            # can read each beam's own origin (carried as origin_x/y/z, frustum-culled
            # via `cull_origin`). Precedence:
            #   1. LAS ExtraBytes per-pulse origins — GROUND TRUTH, no join needed.
            #   2. trajectory join — reconstruct origins by joining each return's
            #      timestamp to the 6-DOF trajectory.
            #   3. static origin — a single fixed scanner position.
            # Moving paths recenter points + origins by `lad_shift` into a small local
            # frame for float32 precision (the inversion is frame-invariant; the grid
            # + output cells are shifted to match below).
            beam_origins = scan_flags.get("beam_origins")
            if beam_origins is not None and beam_origins.shape[0] == xyz.shape[0]:
                if scan_entry.trajectory is not None:
                    warnings.append(
                        "Scan carries per-pulse beam-origin columns (LAS ExtraBytes) "
                        "AND a trajectory; the explicit origins were used and the "
                        "trajectory was ignored.")
                if lad_shift is not None:
                    xyz = xyz - lad_shift
                    beam_origins = beam_origins - lad_shift
                dirs, labels, vals = _attach_origins(xyz, labels, vals, beam_origins)
                cull_origin = beam_origins
                scan_moving = True
            elif scan_entry.trajectory is not None:
                # Moving-platform scan: reconstruct a PER-BEAM emission origin for
                # every return by joining its timestamp to the trajectory, recompute
                # the ray directions, and carry origin_x/y/z so the beam-based
                # (Gtheta) inversion reads each beam's own origin. Points and
                # trajectory are recentered by `lad_shift` for float32 precision.
                if lad_shift is not None:
                    xyz = xyz - lad_shift
                dirs, labels, vals, cull_origin = _apply_trajectory_origins(
                    xyz, dirs, labels, vals, scan_entry.trajectory, shift=lad_shift,
                    warnings=warnings)
                scan_moving = True
            else:
                cull_origin = origin
                scan_moving = False

            # Cull to the grid's beam frustum: keep only beams (origin -> point)
            # whose segment can pass through the grid AABB. This preserves
            # through-grid miss rays (Beer's law) while dropping far-field points
            # whose beams never touch the grid — the biggest win for a localized
            # grid in a large scene. For moving scans `cull_origin` is per-beam.
            n_before = xyz.shape[0]
            keep = _cull_to_grid(xyz, cull_origin, grid_center, grid_size)
            if not keep.all():
                xyz = xyz[keep]
                dirs = dirs[keep]
                if vals is not None:
                    vals = vals[keep]
            n_after = xyz.shape[0]
            if n_after < n_before:
                print(f"[lad] grid cull: kept {n_after}/{n_before} points "
                      f"({100.0 * n_after / max(n_before, 1):.1f}%)", flush=True)

            # A scan flagged multi-return whose resolved data lacks the per-pulse
            # columns can't run the full-waveform algorithm (gapfillMisses() would
            # hard-error). Surface that rather than silently mislabeling it.
            if (scan_entry.return_type or "single") == "multi" and not scan_multi:
                warnings.append(
                    "Scan marked multi-return but its point data lacks the per-pulse "
                    "timestamp/target_index/target_count columns; computed as "
                    "single-return. Re-import or re-scan preserving those columns "
                    "for full-waveform LAD."
                )

            # Miss accounting drives LAD accuracy. A scan is in good shape if it
            # already carries miss points (E57 / structured PLY) OR it has a
            # timestamp column we can gapfill from. With neither, the inversion
            # can't see the gaps and is likely inaccurate — warn loudly.
            scan_label = (getattr(scan_entry, 'label', None)
                          or os.path.basename(scan_entry.file_path or '')
                          or 'this scan')
            if not scan_flags["has_misses"] and not scan_flags["has_timestamp"]:
                warnings.append(
                    f"Scan '{scan_label}' has no sky/miss points and no timestamp "
                    "column, so LAD cannot account for beams that hit the sky and is "
                    "likely to be inaccurate. Re-import a scan that carries miss "
                    "points (E57 / structured PLY) or a timestamp column so misses "
                    "can be recovered by gapfilling."
                )
            # Backfilled misses that predate a later crop reflect the pre-crop
            # hits, so their ratio against the surviving hits is off and the
            # inversion may be inaccurate. (Set by delete_region; carried through
            # _session_to_lad_arrays.) Warn so the user re-runs Backfill Misses.
            if scan_flags.get("misses_stale"):
                warnings.append(
                    f"Scan '{scan_label}' has sky/miss points that were computed "
                    "before a later crop, so the leaf-area-density result may be "
                    "inaccurate. Re-run Backfill Misses on the cropped cloud to "
                    "recompute them against the current points."
                )

            # Ntheta/Nphi describe the scanner's angular raster, not the surviving
            # points — so estimate from the PRE-cull count when the scan doesn't
            # carry explicit values. The LAD inversion is very sensitive to this
            # resolution (it sets the beam raster the Beer's-law path lengths are
            # traced over), so culling must not shrink it.
            n_theta, n_phi = _resolution(
                scan_entry, n_before, theta_max - theta_min, phi_max - phi_min)

            scans_arrays.append({
                "origin": origin,
                "xyz": xyz,
                "dirs": dirs,
                "labels": labels,
                "vals": vals,
                "n_theta": n_theta,
                "n_phi": n_phi,
                "theta_min": theta_min,
                "theta_max": theta_max,
                "phi_min": phi_min,
                "phi_max": phi_max,
                "multi": scan_multi,
                "moving": scan_moving,
                "has_timestamp": scan_flags["has_timestamp"],
                "has_misses": scan_flags["has_misses"],
            })
            scan_xyz_for_counts.append(xyz)

        is_multi = any(s["multi"] for s in scans_arrays)
        return_mode = "multi" if is_multi else "single"
        # Misses must already be present — either retained by the source format
        # (E57 / structured PLY) or recovered up front via the explicit Backfill
        # Misses step (which persists them in the session and surfaces them through
        # _session_to_lad_arrays). LAD no longer gapfills silently; if a scan has
        # no misses the guard below stops here with an actionable error.
        any_has_misses = any(s["has_misses"] for s in scans_arrays)

        # Build the cloud entirely in RAM: one scan + bulk hit ingest per scan,
        # then the grid — no XML, no ASCII file. addScan takes radians; our
        # angles are degrees. Beam divergence is supplied in milliradians.
        _report(0.45, "Building voxel grid")
        cloud = LiDARCloud()
        cloud.disableMessages()
        for entry, scan_entry in zip(scans_arrays, request.scans):
            sid = cloud.addScan(
                origin=entry["origin"],
                Ntheta=entry["n_theta"],
                theta_range=(math.radians(entry["theta_min"]), math.radians(entry["theta_max"])),
                Nphi=entry["n_phi"],
                phi_range=(math.radians(entry["phi_min"]), math.radians(entry["phi_max"])),
                exit_diameter=(scan_entry.beam_exit_diameter or 0.0),
                beam_divergence=((scan_entry.beam_divergence or 0.0) / 1000.0),
            )
            if entry["xyz"].shape[0] > 0:
                cloud.addHitPointsWithData(
                    sid, entry["xyz"], entry["dirs"],
                    entry["labels"], entry["vals"])
        # Honor the grid box's azimuthal rotation. Helios stores the cells in the
        # rotated frame, so getCellCenter returns rotated centers and the beam
        # tracing / Beer's-law inversion run against the rotated voxels — matching
        # the rotated grid mesh the user laid out. Without this the LAD result grid
        # came back axis-aligned and visibly skewed off the original box.
        # NOTE: Helios's LiDARcloud::addGrid takes the rotation in DEGREES (it
        # multiplies by pi/180 internally), despite the PyHelios docstring saying
        # radians — pass degrees, matching the <grid><rotation> XML triangulation path.
        #
        # Terrain following: when enabled (with a DEM), shift each voxel column
        # vertically so its bottom rides the DEM surface. The offset is baked into
        # the Helios cell centers here so the beam tracing / inversion run against
        # the terrain-following voxels — a renderer-only shift would be wrong.
        column_offsets = None
        terrain_dropped = 0
        terrain_kept_mask = None
        if request.terrain_follow:
            if request.dem is None:
                raise ValueError("terrain_follow requires a DEM (none supplied).")
            # Sample the DEM in the COMPUTE frame: on a moving-platform scan the
            # points + grid were recentered by lad_shift, so shift the DEM origin by
            # the same (x,y) so raster, grid, and points share one frame.
            dem_for_sampling = request.dem
            if lad_shift is not None:
                dem_for_sampling = DemRaster(
                    grid_z=request.dem.grid_z, nx=request.dem.nx, ny=request.dem.ny,
                    cell=request.dem.cell,
                    origin=[request.dem.origin[0] - float(lad_shift[0]),
                            request.dem.origin[1] - float(lad_shift[1])],
                    nodata=request.dem.nodata)
            column_offsets, terrain_kept_mask, terrain_dropped = _sample_dem_columns(
                grid_center, grid_size, grid_nx, grid_ny, grid_nz,
                grid_rotation_rad, dem_for_sampling, request.safety_fraction)
            if not terrain_kept_mask.any():
                return {
                    "success": False,
                    "cells": [],
                    "method_used": "helios",
                    "error": "Terrain following dropped every voxel column — the grid "
                             "footprint lies entirely outside the DEM. Check that the "
                             "grid and DEM overlap.",
                    "warnings": warnings,
                }
            if terrain_dropped:
                warnings.append(
                    f"Terrain following dropped {terrain_dropped} voxel column(s) whose "
                    f"footprint fell outside the DEM; their voxels were excluded.")

        if column_offsets is not None:
            cloud.addGrid(center=grid_center, size=grid_size,
                          ndiv=[grid_nx, grid_ny, grid_nz],
                          rotation=grid_rotation_deg,
                          column_offsets=column_offsets)
        else:
            cloud.addGrid(center=grid_center, size=grid_size,
                          ndiv=[grid_nx, grid_ny, grid_nz],
                          rotation=grid_rotation_deg)

        # Two inversion paths:
        #   - Supplied-G(theta) (beam-based): no triangulation. Each return is
        #     traced from its own beam origin and the caller's mean G(theta) is
        #     used. Taken when ANY scan is moving (their pulses can't be
        #     triangulated) OR when gtheta_override is set for a static scan (a
        #     homogeneous/known-G canopy, or aerial-style data where triangulating
        #     sparse overhead returns biases G).
        #   - Triangulated: static scans estimate G(theta) per cell from a mesh.
        any_moving = any(s["moving"] for s in scans_arrays)
        use_supplied_gtheta = any_moving or bool(request.gtheta_override)
        if use_supplied_gtheta:
            # A reused triangulation is meaningless when we invert from a supplied
            # G(theta) (no mesh is used). Reject loudly rather than silently ignore.
            if reuse_mesh is not None:
                raise ValueError(
                    "A reused triangulation cannot be combined with a moving-platform "
                    "scan or a G(theta) override (both invert from a supplied G(theta), "
                    "not a mesh).")
            supplied_gtheta = request.gtheta if request.gtheta is not None else 0.5
            if request.gtheta is None:
                warnings.append(
                    "LAD with a G(theta) override needs a mean leaf-projection "
                    "coefficient; defaulted to 0.5 (spherical leaf-angle "
                    "distribution). Set it to match the canopy if known.")
            if not (0.0 < supplied_gtheta <= 1.0):
                raise ValueError("gtheta must be in (0, 1] (0.5 = spherical).")
        elif reuse_mesh is not None:
            # Reuse a previously-run triangulation: inject it instead of re-running
            # the Delaunay pass. The mesh's vertices are world coordinates; the
            # static path never applies lad_shift (only moving scans do), so they
            # already match the cloud's points — assert that invariant rather than
            # shifting. Expand the indexed mesh to the flat 9-floats-per-triangle
            # soup setExternalTriangulation expects (one vectorized gather).
            assert lad_shift is None, \
                "triangulation reuse is static-only; lad_shift must be None"
            _ckpt()
            _report(0.55, "Reusing triangulation")
            verts, tri_idx, scan_ids = reuse_mesh
            soup = verts[tri_idx].reshape(-1).astype(np.float32, copy=False)
            cloud.setExternalTriangulation(soup, scan_ids)
            if cloud.getTriangleCount() == 0:
                return {
                    "success": False,
                    "cells": [],
                    "method_used": "helios",
                    "error": "The reused triangulation produced no triangles inside the "
                             "grid — cannot compute the G-function. Check that the mesh "
                             "and grid overlap.",
                    "warnings": warnings,
                }
        else:
            _ckpt()
            _report(0.55, "Triangulating hit points")
            cloud.triangulateHitPoints(request.lmax, request.max_aspect_ratio)
            if cloud.getTriangleCount() == 0:
                return {
                    "success": False,
                    "cells": [],
                    "method_used": "helios",
                    "error": "No triangles generated — cannot compute the G-function. "
                             "Try increasing Lmax or adjusting max_aspect_ratio.",
                    "warnings": warnings,
                }

        # calculateLeafArea() needs miss points (transmitted beams) for the
        # Beer's-law denominator and fail-fasts without them. LAD no longer
        # synthesises misses on the fly — they must already be present (retained
        # by the source format, or recovered via the explicit Backfill Misses
        # step). If none are present, stop here with an actionable error rather
        # than letting the raw Helios exception bubble up through the generic
        # handler below. The C++ check still backstops.
        if not any_has_misses:
            return {
                "success": False,
                "cells": [],
                "method_used": "helios",
                "error": "This scan has no sky/miss points, so leaf area density "
                         "cannot be computed — the inversion needs beams that passed "
                         "through the canopy without returning. Run Backfill Misses "
                         "first to recover them (if the scan carries a timestamp or "
                         "row/column grid), or re-import a scan that retains misses "
                         "(E57 / structured PLY).",
                "warnings": warnings,
            }

        # Pimont et al. (2018) per-voxel sampling uncertainty is computed on every
        # inversion. It requires min_voxel_hits to be set; the renderer always
        # sends a value, but a bare API caller may not — floor to 5 and warn rather
        # than failing the whole run.
        conf_level = 0.95
        element_width = request.element_width
        min_hits = request.min_voxel_hits
        if min_hits is None:
            min_hits = 5
            warnings.append(
                "Leaf-area uncertainty (Pimont et al.) requires a minimum voxel-hits "
                "threshold; defaulted to 5. Set 'Min Voxel Hits' to control which "
                "voxels are solved."
            )
        _ckpt()
        _report(0.80, "Inverting Beer's law")
        with Context() as ctx:
            if use_supplied_gtheta:
                # Beam-based inversion: traverses each beam from its own origin
                # (getHitOrigin), uses the supplied G(theta), no triangulation.
                # Works for both moving-platform scans and static scans with a
                # G(theta) override.
                cloud.calculateLeafArea(ctx, min_hits, element_width,
                                        Gtheta=supplied_gtheta)
            else:
                cloud.calculateLeafArea(ctx, min_hits, element_width)

        def _clean(x):
            """NaN -> None so the value survives JSON serialization."""
            return None if (x != x) else x

        _report(0.92, "Collecting voxel results")
        # Per-cell hit counts: Helios exposes no getter, so bin the points into
        # the grid AABBs ourselves. Reads each scan file once (positions only).
        n_cells = cloud.getGridCellCount()
        cell_centers = [cloud.getCellCenter(i) for i in range(n_cells)]
        cell_sizes = [cloud.getCellSize(i) for i in range(n_cells)]
        hit_counts = _count_points_per_cell(scan_xyz_for_counts, cell_centers, cell_sizes,
                                            grid_rotation_rad,
                                            column_z_offsets=column_offsets,
                                            ndiv=(grid_nx, grid_ny, grid_nz))

        cells = []
        total_leaf_area = 0.0
        solved_indices = []  # voxels with a real LAD solution + defined variance
        # For terrain following, cells in dropped columns (outside the DEM footprint)
        # are excluded from the reported results. Helios still holds them (so the ray
        # tracing/lattice is intact); we just don't surface them. Column of cell i is
        # i % (nx*ny) under Helios's k-major ordering.
        ncols_terrain = grid_nx * grid_ny
        for i in range(n_cells):
            if (i & 0xFFF) == 0:  # every 4096 voxels — negligible overhead
                _ckpt()
            if terrain_kept_mask is not None and not terrain_kept_mask[i % ncols_terrain]:
                continue
            c = cell_centers[i]
            s = cell_sizes[i]
            la = float(cloud.getCellLeafArea(i))
            lad = float(cloud.getCellLeafAreaDensity(i))
            gt = float(cloud.getCellGtheta(i))
            lad_solved = lad == lad  # finite before the NaN->0 squash below
            # Helios returns NaN for unsolved cells; surface them as 0 so the UI
            # can treat them as empty rather than choking on NaN in JSON.
            if la != la:
                la = 0.0
            if lad != lad:
                lad = 0.0
            if gt != gt:
                gt = 0.0
            total_leaf_area += la

            # Pimont per-voxel uncertainty. Sentinels: beam_count -1, variance -1,
            # and NaN all map to None so the UI gates on presence. Single-voxel CI
            # bounds are only meaningful when the validity gate passes.
            var = float(cloud.getCellLADVariance(i))
            bc = int(cloud.getCellBeamCount(i))
            ci_valid, ci_lo, ci_hi = cloud.getCellLeafAreaConfidenceInterval(i, conf_level)
            # Cell centers come back in the (possibly recentered) compute frame;
            # add lad_shift back so results are reported in true world coordinates.
            center = ([c.x + lad_shift[0], c.y + lad_shift[1], c.z + lad_shift[2]]
                      if lad_shift is not None else [c.x, c.y, c.z])
            cell = {
                "index": i,
                "center": center,
                "size": [s.x, s.y, s.z],
                "leaf_area": la,
                "lad": lad,
                "gtheta": gt,
                "hit_count": int(hit_counts[i]),
                "beam_count": bc if bc >= 0 else None,
                "relative_density_index": _clean(float(cloud.getCellRelativeDensityIndex(i))),
                "mean_path_length": _clean(float(cloud.getCellMeanPathLength(i))),
                "lad_variance": var if var >= 0 else None,
                "lad_std": math.sqrt(var) if var >= 0 else None,
                "ci_valid": bool(ci_valid),
                "leaf_area_ci_lower": _clean(float(ci_lo)) if ci_valid else None,
                "leaf_area_ci_upper": _clean(float(ci_hi)) if ci_valid else None,
            }
            if lad_solved and var >= 0:
                solved_indices.append(i)
            cells.append(cell)

        # Group-scale CI (Pimont Eq. 39) over the solved voxels — the recommended,
        # much-tighter aggregate. Reported even when individual single-voxel CIs are
        # gated out. Zero solved voxels => not computable.
        group_ci_valid = False
        group_lad_mean = None
        group_lad_ci_lower = None
        group_lad_ci_upper = None
        if solved_indices:
            gvalid, gmean, glo, ghi = cloud.getGroupLADConfidenceInterval(
                solved_indices, conf_level)
            group_ci_valid = bool(gvalid)
            group_lad_mean = _clean(float(gmean))
            if gvalid:
                group_lad_ci_lower = _clean(float(glo))
                group_lad_ci_upper = _clean(float(ghi))
        else:
            warnings.append(
                "Leaf-area uncertainty was computed, but no voxel met the criteria "
                "for a solution, so no confidence interval could be reported."
            )

        # Report the grid center + bounds in true WORLD coordinates, consistent
        # with the per-cell `center` (which adds lad_shift back above). On a
        # moving-platform scan with large coordinates `grid_center` was recentered
        # by `lad_shift` into the local compute frame; undo that here so a consumer
        # can't read a grid_center/bounds that disagrees with the voxels by the
        # shift (hundreds of km for a UTM scene). Static scans leave it untouched.
        grid_center_world = (
            [grid_center[k] + lad_shift[k] for k in range(3)]
            if lad_shift is not None else list(grid_center))
        bb_lo = [grid_center_world[k] - grid_size[k] / 2 for k in range(3)]
        bb_hi = [grid_center_world[k] + grid_size[k] / 2 for k in range(3)]

        _report(1.0, "Done")
        return {
            "success": True,
            "cells": cells,
            "nx": grid_nx,
            "ny": grid_ny,
            "nz": grid_nz,
            "grid_center": grid_center_world,
            "grid_size": grid_size,
            # Azimuthal rotation of the grid box about +z (degrees). The cell
            # centers are already in the rotated lattice; the renderer needs this
            # to orient each voxel cube so the result grid aligns with the box.
            "grid_rotation": grid_rotation_deg,
            "bounds": [bb_lo, bb_hi],
            "terrain_follow": bool(request.terrain_follow),
            "dropped_columns": int(terrain_dropped),
            "is_multi_return": is_multi,
            "return_mode": return_mode,
            "total_leaf_area": total_leaf_area,
            "group_ci_valid": group_ci_valid,
            "group_lad_mean": group_lad_mean,
            "group_lad_ci_lower": group_lad_ci_lower,
            "group_lad_ci_upper": group_lad_ci_upper,
            "confidence_level": conf_level,
            "element_width": element_width,
            "gapfilled_misses": 0,  # LAD no longer gapfills; misses come from Backfill Misses / source
            "had_miss_points": any_has_misses,
            "method_used": "helios",
            "warnings": warnings,
        }

    except ImportError as e:
        return {
            "success": False,
            "cells": [],
            "method_used": "helios",
            "error": f"PyHelios not installed: {str(e)}",
            "warnings": warnings,
        }
    except ScanCancelled:
        raise  # cancellation propagates to the streaming wrapper (memory freed)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "cells": [],
            "method_used": "helios",
            "error": f"Leaf area density computation failed: {str(e)}",
            "warnings": warnings,
        }


def _bin_points_to_cells(points, grid_center, grid_size, nx: int, ny: int, nz: int):
    """Assign each point to a grid cell (row-major ``i + nx*(j + ny*k)``; outside
    -> -1). Thin wrapper over :func:`qsm.grid.bin_points_to_cells`, which is the
    single shared implementation used by both the LAD/triangulation pipeline and
    the QSM leaf-angle adjustment.
    """
    from qsm.grid import bin_points_to_cells
    return bin_points_to_cells(points, grid_center, grid_size, nx, ny, nz)


def _count_points_per_cell(scan_xyz_list: list, cell_centers: list, cell_sizes: list,
                           grid_rotation_rad: float = 0.0,
                           column_z_offsets: "Optional[list]" = None,
                           ndiv: "Optional[tuple]" = None):
    """Count how many scan points fall inside each voxel.

    Bins the in-RAM (N,3) position arrays (one per scan) by the grid's regular
    structure inferred from the cell centers/sizes. Used only to populate the
    per-voxel `hit_count` for the UI — not part of the LAD math.

    A rotated grid (``grid_rotation_rad`` != 0, azimuth about +z) is handled by
    inverse-rotating the POINTS about the grid center into the axis-aligned cell
    frame before binning — the same convention Helios uses internally
    (getContainingGridCell inverse-rotates the query point by -rotation). NOTE:
    Helios's getCellCenter returns the UNROTATED lattice center (the rotation is
    stored per-cell, not baked into the center), so the cell centers here are
    already axis-aligned and must NOT be rotated.

    Terrain following: when ``column_z_offsets`` (length nx*ny, row-major [j*nx+i])
    and ``ndiv``=(nx,ny,nz) are given, each voxel column was shifted vertically by
    its offset, so the cell z's are no longer a regular lattice. We restore the
    regular frame for binning by subtracting each cell's column offset from its
    center z, and each point's column offset from its z (looked up by the point's
    (i,j) column in the unrotated frame). Helios orders cells k-major (for k: for
    j: for i), so cell index ``idx`` has column ``idx % (nx*ny)``.
    """
    import numpy as np

    n_cells = len(cell_centers)
    counts = np.zeros(n_cells, dtype=np.int64)
    if n_cells == 0:
        return counts

    centers = np.array([[c.x, c.y, c.z] for c in cell_centers], dtype=np.float64)
    sizes = np.array([[s.x, s.y, s.z] for s in cell_sizes], dtype=np.float64)

    terrain = column_z_offsets is not None and ndiv is not None
    if terrain:
        nx, ny = int(ndiv[0]), int(ndiv[1])
        ncols = nx * ny
        col_off = np.asarray(column_z_offsets, dtype=np.float64)
        # Un-shift cell centers back to the regular lattice for axis-aligned z bins.
        cell_col = np.arange(n_cells) % ncols
        centers[:, 2] -= col_off[cell_col]

    # Pivot for inverse-rotating points: the grid center. The cell centers are the
    # UNROTATED lattice, so their mean is exactly the grid center. Rotation is
    # about +z, so z is untouched. (For terrain grids the x,y pivot is unaffected
    # by the z un-shift above.)
    pivot = centers.mean(axis=0)

    def _unrotate(p):
        if abs(grid_rotation_rad) <= 1e-9:
            return p
        theta = -grid_rotation_rad  # inverse rotation about +z
        cos_t, sin_t = np.cos(theta), np.sin(theta)
        d = p - pivot
        out = p.copy()
        out[:, 0] = pivot[0] + d[:, 0] * cos_t - d[:, 1] * sin_t
        out[:, 1] = pivot[1] + d[:, 0] * sin_t + d[:, 1] * cos_t
        return out

    # Grid lower corner and per-axis cell counts/steps from the (axis-aligned) cells.
    grid_lo = (centers - sizes / 2).min(axis=0)
    grid_hi = (centers + sizes / 2).max(axis=0)
    step = sizes[0]  # cells are uniform within a Helios grid
    safe_step = np.where(step > 0, step, 1)
    nper = np.maximum(np.round((grid_hi - grid_lo) / safe_step).astype(int), 1)

    # Map each cell's (i,j,k) to its index, matching however Helios ordered them.
    cell_ijk = np.clip(
        np.floor((centers - grid_lo) / safe_step).astype(int), 0, nper - 1)
    ijk_to_index = {
        (cell_ijk[idx, 0], cell_ijk[idx, 1], cell_ijk[idx, 2]): idx
        for idx in range(n_cells)
    }
    # Flat (i,j,k) -> cell index lookup table, vectorized over all points.
    lut = np.full((nper[0], nper[1], nper[2]), -1, dtype=np.int64)
    for (i, j, k), idx in ijk_to_index.items():
        lut[i, j, k] = idx

    for xyz in scan_xyz_list:
        xyz = np.asarray(xyz, dtype=np.float64)
        if xyz.size == 0:
            continue
        xyz = _unrotate(xyz)
        if terrain:
            # Subtract each point's column offset from its z so it bins into the
            # same regular lattice the (un-shifted) cells now occupy. Points whose
            # (i,j) column lies outside the grid get no offset and fail the bounds
            # check below anyway.
            xy_idx = np.floor((xyz[:, :2] - grid_lo[:2]) / safe_step[:2]).astype(int)
            in_xy = np.all((xy_idx >= 0) & (xy_idx < nper[:2]), axis=1)
            pi = np.clip(xy_idx[:, 0], 0, nx - 1)
            pj = np.clip(xy_idx[:, 1], 0, ny - 1)
            pt_off = np.where(in_xy, col_off[pj * nx + pi], 0.0)
            xyz = xyz.copy()
            xyz[:, 2] -= pt_off
        ijk = np.floor((xyz - grid_lo) / safe_step).astype(int)
        inside = np.all((ijk >= 0) & (ijk < nper), axis=1)
        ijk = ijk[inside]
        if ijk.size == 0:
            continue
        cell_idx = lut[ijk[:, 0], ijk[:, 1], ijk[:, 2]]
        cell_idx = cell_idx[cell_idx >= 0]
        np.add.at(counts, cell_idx, 1)

    return counts


def _decode_lad_request_frame(body: bytes) -> "Tuple[LADComputeRequest, Tuple[np.ndarray, np.ndarray, np.ndarray]]":
    """Decode a PHB1 LAD-reuse request frame into (request, (vertices, indices, scan_ids)).

    The renderer sends a reuse LAD request as a PHB1 binary frame (the same format
    _bin_frame_bytes produces for responses): the scalar LADComputeRequest fields
    live in the header `meta`, and the reused mesh rides as three buffers —
    `mesh_vertices` (f32, V*3), `mesh_indices` (u32, T*3), `mesh_scan_ids` (u32, T;
    non-negative scan indices already remapped to this request's scan order).
    Sending the mesh as raw bytes (not JSON numbers or base64) keeps a 1M+ triangle
    payload ~18 MB instead of 150+ MB of text that would trip V8's string ceiling.
    """
    import struct
    if body[:4] != _BIN_FRAME_MAGIC:
        raise ValueError("LAD binary request must start with a PHB1 frame magic")
    (header_len,) = struct.unpack_from("<I", body, 4)
    header = json.loads(body[8:8 + header_len].decode("utf-8"))
    meta = header.get("meta", {})
    descs = header.get("buffers", [])

    buffers: "dict[str, np.ndarray]" = {}
    off = 8 + header_len
    for d in descs:
        n = int(d["length"])
        np_dtype = np.float32 if d["dtype"] == "f32" else np.uint32
        nbytes = n * 4
        buffers[d["name"]] = np.frombuffer(body, dtype=np_dtype, count=n, offset=off).copy()
        off += nbytes

    request = LADComputeRequest(**meta)

    for key in ("mesh_vertices", "mesh_indices", "mesh_scan_ids"):
        if key not in buffers:
            raise ValueError(f"LAD reuse frame missing required buffer '{key}'")

    verts = buffers["mesh_vertices"].astype(np.float32, copy=False)
    if verts.size % 3 != 0:
        raise ValueError("mesh_vertices length must be a multiple of 3 (xyz per vertex)")
    verts = verts.reshape(-1, 3)
    V = verts.shape[0]

    tri_idx = buffers["mesh_indices"].astype(np.uint32, copy=False)
    if tri_idx.size % 3 != 0:
        raise ValueError("mesh_indices length must be a multiple of 3 (three per triangle)")
    tri_idx = tri_idx.reshape(-1, 3)
    T = tri_idx.shape[0]

    # scan_ids: non-negative indices into request.scans order; keep as int32 for
    # setExternalTriangulation (its C-ABI takes int*).
    scan_ids = buffers["mesh_scan_ids"].astype(np.int32, copy=False)

    if T == 0:
        raise ValueError("mesh_indices is empty — a reused triangulation needs at least one triangle")
    if scan_ids.shape[0] != T:
        raise ValueError(f"mesh_scan_ids has {scan_ids.shape[0]} entries, expected {T} (one per triangle)")
    if V > 0 and int(tri_idx.max()) >= V:
        raise ValueError("mesh_indices references a vertex beyond mesh_vertices")
    n_scans = len(request.scans)
    if int(scan_ids.min()) < 0 or int(scan_ids.max()) >= n_scans:
        raise ValueError(
            f"mesh_scan_ids must be in [0, {n_scans}) (the request's scan count); "
            f"got range [{int(scan_ids.min())}, {int(scan_ids.max())}]")

    return request, (verts, tri_idx, scan_ids)


@app.post("/api/lad/compute")
async def lad_compute(http_request: Request):
    """Compute per-voxel leaf area density via PyHelios.

    Accepts either a JSON LADComputeRequest body (the default / fresh-triangulation
    path) or — when reusing a previously-run Helios triangulation — a PHB1 binary
    frame carrying the request fields in its header plus the mesh as raw buffers
    (see _decode_lad_request_frame). The binary path lets a 1M+ triangle mesh ride
    back to the backend compactly so it can be injected via setExternalTriangulation
    instead of being re-triangulated from scratch.

    Either way the response streams PHP1 progress markers (see
    _bin_frame_streaming_response) ahead of the JSON result so the renderer shows a
    real per-stage progress bar and the keepalive survives WebKit's ~60s stall
    timeout. The renderer drains the markers and parses the JSON tail.
    """
    content_type = (http_request.headers.get("content-type") or "").lower()
    reuse_mesh = None
    if "application/json" in content_type or content_type == "":
        request = LADComputeRequest(**(await http_request.json()))
    else:
        body = await http_request.body()
        request, reuse_mesh = _decode_lad_request_frame(body)

    run_id, cancel_event = _new_cancel_token()
    return _bin_frame_streaming_response(
        lambda progress: json.dumps(
            _do_lad_computation(request, progress=progress, reuse_mesh=reuse_mesh)
        ).encode("utf-8"),
        request=http_request, cancel_event=cancel_event, run_id=run_id)


class TrajectoryParseRequest(BaseModel):
    """Parse a binary trajectory file into the canonical PoseStream wire shape.

    `path` is a server-readable file path (the renderer sends the picked path, same
    as the cloud-import endpoints). `format` is auto-detected from the extension when
    omitted. `smrmsg_path` optionally points at the SBET accuracy companion for a QC
    warning. `target_poses` caps the decimated pose count."""
    path: str
    format: Optional[str] = None
    smrmsg_path: Optional[str] = None
    target_poses: int = 3000


@app.post("/api/trajectory/parse")
async def trajectory_parse(request: TrajectoryParseRequest):
    """Parse a binary trajectory (currently SBET .sbet/.out) into the canonical
    PoseStream wire dict the renderer's poseStreamFromWire consumes. Text trajectories
    (.csv/.txt/.tsv/.traj) are parsed client-side and never hit this endpoint.

    Binary parsing belongs server-side: it needs pyproj (Python-only) for the
    geographic->UTM projection, and the renderer only does IPC text reads.
    """
    import sbet

    src = _Path(request.path).expanduser()
    if not src.is_file():
        raise HTTPException(status_code=404, detail=f"Trajectory file not found: {request.path}")

    fmt = (request.format or src.suffix.lstrip(".")).lower()
    if fmt in ("sbet", "out"):
        try:
            return sbet.parse_sbet(
                str(src), target_poses=request.target_poses,
                smrmsg_path=request.smrmsg_path)
        except sbet.SbetParseError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:  # pyproj / unexpected — surface a clean 400, not a 500
            raise HTTPException(status_code=400,
                                detail=f"Failed to parse SBET trajectory: {e}")
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported binary trajectory format '.{fmt}'. Supported: .sbet/.out "
               f"(SBET). Text trajectories (.csv/.txt/.tsv/.traj) are imported directly.")


# ==================== SCAN EXPORT (Helios XML + per-scan ASCII) ====================
# Export one or more scans to the native Helios scan format: an XML metadata file
# (scanner origin, angular sweep, beam optics, ASCII_format) plus ONE ASCII data
# file per scan (`<base>_<scanID>.xyz`). The per-scan split is load-bearing — the
# XML references each data file by scan, so a single merged file could not be
# re-associated with its scanner parameters. Re-loadable via PyHelios loadXML().
#
# This is the round-trip-faithful path: it preserves the `is_miss` flag (and any
# other per-hit scalar columns) by passing them through addScan(column_format=...)
# so exportScans() writes them and the loader can read them back. Edits are
# honored because the points are resolved through the same session/inline/file
# machinery the LAD path uses (a session-backed cloud yields its surviving,
# translation-applied points; deletions already applied, never re-reading disk).

class ScanExportGrid(BaseModel):
    """A voxel-box grid to write as a Helios <grid> block on scan XML export, so a
    bundle like sphere.xml round-trips its grid. center/size are world extents;
    nx/ny/nz the per-axis cell counts; rotation the azimuth about +z in degrees."""
    center: List[float]      # [x, y, z]
    size: List[float]        # [x, y, z] full extents
    nx: int = 1
    ny: int = 1
    nz: int = 1
    rotation: float = 0.0    # degrees about +z


class ScanExportEntry(BaseModel):
    """One scan to export. Point source is one of session_id / points / file_path
    (resolved in that precedence, mirroring the LAD path). Scanner geometry is
    written into the XML; translation is applied to the points on export."""
    origin: List[float]                          # [x, y, z] scanner position
    # "raster" (default) or "spinning_multibeam". Multibeam scans export via
    # beam_elevation_angles_deg; n_theta/theta_* are ignored for them.
    scan_pattern: Optional[str] = "raster"
    beam_elevation_angles_deg: Optional[List[float]] = None  # deg above horizon (multibeam)
    n_theta: Optional[int] = None                # Ntheta (zenith samples)
    n_phi: Optional[int] = None                  # Nphi (azimuth samples)
    theta_min: float = 0.0                       # degrees
    theta_max: float = 180.0
    phi_min: float = 0.0
    phi_max: float = 360.0
    beam_exit_diameter: Optional[float] = None   # meters
    beam_divergence: Optional[float] = None      # milliradians
    # Initial scanner heading (degrees). Round-trips into the XML via PyHelios
    # exportScans() (v0.1.23+), which writes a <scanAzimuthOffset> tag the renderer
    # reads back on import.
    scan_azimuth_offset: Optional[float] = None
    # Point source (one of):
    session_id: Optional[str] = None             # live edited session (honors deletions)
    points: Optional[List[List[float]]] = None   # inline flat cloud
    scalar_columns: Optional[Dict[str, List[float]]] = None  # aligned per-point columns
    file_path: Optional[str] = None              # source file fallback
    ascii_format: Optional[str] = None
    # World-space offset applied to the points (viewer translation), or null.
    translation: Optional[List[float]] = None
    # Ordered ASCII column slugs chosen in the export modal (includes x y z plus
    # any scalar columns the user kept, in their chosen order). When given, it
    # overrides the default column set/order; x/y/z are always written regardless.
    columns: Optional[List[str]] = None


class ScanExportRequest(BaseModel):
    """Export one or more scans to a Helios XML + per-scan ASCII bundle."""
    scans: List[ScanExportEntry]
    base_name: Optional[str] = None              # output base (→ <base>.xml, <base>_<id>.xyz)
    include_misses: bool = True                  # write miss points (+ is_miss column)
    # When True (default), write the Helios scan XML alongside the per-scan data
    # files (re-loadable bundle, data always .xyz). When False, write ONLY the
    # per-scan data files (no XML) in `data_format` — one file per scan.
    write_xml: bool = True
    # Data-only output format (write_xml=False). One of las/laz/ply/xyz/csv/txt/
    # obj/e57. Ignored when write_xml=True (the Helios bundle is always .xyz).
    data_format: str = "xyz"
    # Voxel-box grids to inject as <grid> blocks (write_xml=True only). None/empty
    # → no grid blocks written. Lets a scan bundle round-trip its grid(s).
    grids: Optional[List["ScanExportGrid"]] = None


# Standard scalar columns we try to preserve on export, in a stable order. Any of
# these present on the scan is written; absent ones are skipped. `is_miss` first so
# the miss flag is always adjacent to xyz when present.
_SCAN_EXPORT_SCALAR_COLUMNS = ['is_miss', 'timestamp', 'target_index',
                               'target_count', 'intensity']


def _resolve_scan_export_arrays(scan_entry, include_misses: bool):
    """Resolve a scan entry to (xyz float64 (N,3), labels list[str],
    vals float64 (N,k)|None) for native export — translation applied, edits
    honored, all standard scalar columns preserved.

    Mirrors the LAD resolver's source precedence (live session → inline points →
    source file), but surfaces the full _SCAN_EXPORT_SCALAR_COLUMNS set rather
    than only the LAD subset, and optionally drops miss rows.
    """
    import numpy as np

    origin = scan_entry.origin
    if len(origin) != 3:
        raise ValueError(f"Origin must have 3 elements, got {len(origin)}")

    # Build a slug -> (N,) getter for whichever source backs this scan, matching
    # the LAD path: session (honors deletions) → inline columns → source file.
    sess = None
    if scan_entry.session_id:
        with _cloud_session_lock:
            sess = _cloud_sessions.get(scan_entry.session_id)
        if sess is None and not (scan_entry.points or scan_entry.file_path):
            raise ValueError(
                f"Cloud session not found: {scan_entry.session_id}. The backend "
                "may have restarted since import. Re-import the scan and try again.")

    if sess is not None:
        with _cloud_session_lock:
            keep = ~sess.deleted
            xyz = np.ascontiguousarray(sess.positions[keep], dtype=np.float64)
            extras = {k: v[keep] for k, v in sess.extras.items()}
            # Prefer the dedicated float64 timestamps field over the float32 extra
            # so GPS-magnitude times export at full precision. Subset by the same
            # `keep`; guard on length so a stale/misaligned field falls back.
            ts64 = (np.asarray(sess.timestamps, dtype=np.float64)[keep]
                    if sess.timestamps is not None
                    and sess.timestamps.shape[0] == sess.positions.shape[0]
                    else None)
        def _get(slug):
            if slug == 'timestamp' and ts64 is not None:
                return ts64
            return np.asarray(extras[slug]) if slug in extras else None
    elif scan_entry.points:
        xyz = np.ascontiguousarray(
            np.asarray(scan_entry.points, dtype=np.float64)[:, :3])
        _inline_cols = scan_entry.scalar_columns or {}
        _inline_n = xyz.shape[0]
        def _get(slug):
            return (np.asarray(_inline_cols[slug])
                    if slug in _inline_cols and len(_inline_cols[slug]) == _inline_n
                    else None)
    elif scan_entry.file_path:
        if not os.path.isfile(scan_entry.file_path):
            raise ValueError(f"Scan file not found: {scan_entry.file_path}")
        xyz, _, _ = _load_pointcloud_arrays(
            scan_entry.file_path, scan_entry.ascii_format)
        xyz = np.ascontiguousarray(np.asarray(xyz, dtype=np.float64)[:, :3])
        col_map = _read_scan_columns_from_file(
            scan_entry.file_path, scan_entry.ascii_format)
        def _get(slug):
            return col_map.get(slug)
    else:
        raise ValueError("Scan entry has no points, file_path, or session_id")

    # Apply translation so the exported coordinates match what the user sees.
    t = getattr(scan_entry, 'translation', None)
    if t is not None:
        tt = np.asarray(t, dtype=np.float64)
        if tt.shape != (3,):
            raise ValueError(f"translation must be [tx, ty, tz]; got {t!r}")
        xyz = xyz + tt

    # Decide which scalar columns to emit, and in what order. When the caller
    # supplied an explicit `columns` list (from the export modal), honor it (minus
    # x/y/z, which are written as geometry); otherwise fall back to every standard
    # scalar column the scan carries, in canonical order.
    chosen = getattr(scan_entry, 'columns', None)
    if chosen:
        wanted = [s for s in chosen if s not in ('x', 'y', 'z')]
    else:
        wanted = list(_SCAN_EXPORT_SCALAR_COLUMNS)
    labels: List[str] = []
    cols: list = []
    for slug in wanted:
        col = _get(slug)
        if col is not None and len(col) == xyz.shape[0]:
            labels.append(slug)
            cols.append(np.asarray(col, dtype=np.float64))

    # Optionally drop miss rows (and then the now-uniform is_miss column).
    if not include_misses and 'is_miss' in labels:
        mi = labels.index('is_miss')
        keep_rows = np.asarray(cols[mi]) == 0
        xyz = xyz[keep_rows]
        cols = [c[keep_rows] for c in cols]
        # Drop the is_miss column entirely — every surviving row is a return.
        labels.pop(mi)
        cols.pop(mi)

    vals = np.column_stack(cols).astype(np.float64) if cols else None
    return xyz, labels, vals


# The geometry / colour slugs handled specially by the multi-format writer; all
# other requested slugs are treated as named scalar columns.
_DATA_GEOMETRY_SLUGS = ('x', 'y', 'z')
_DATA_COLOR_SLUGS = ('r', 'g', 'b')


def _resolve_scan_for_format(scan_entry, include_misses: bool):
    """Resolve a scan entry to the channels a per-format writer needs:
    {positions (N,3), colors (N,3 0-1)|None, intensity (N,)|None,
     scalars: {slug: (N,)}, ordered: [slugs]}. Translation applied, edits honored,
     miss rows optionally dropped (applied uniformly to every channel).

    `ordered` is the export column order (from the modal's `columns`, minus x/y/z),
    used by the ASCII writer; binary writers ignore it and use their fixed schema.
    """
    import numpy as np

    origin = scan_entry.origin
    if len(origin) != 3:
        raise ValueError(f"Origin must have 3 elements, got {len(origin)}")

    sess = None
    if scan_entry.session_id:
        with _cloud_session_lock:
            sess = _cloud_sessions.get(scan_entry.session_id)
        if sess is None and not (scan_entry.points or scan_entry.file_path):
            raise ValueError(
                f"Cloud session not found: {scan_entry.session_id}. The backend "
                "may have restarted since import. Re-import the scan and try again.")

    colors = None
    intensity = None
    if sess is not None:
        with _cloud_session_lock:
            keep = ~sess.deleted
            xyz = np.ascontiguousarray(sess.positions[keep], dtype=np.float64)
            extras = {k: np.asarray(v[keep]) for k, v in sess.extras.items()}
            if getattr(sess, 'colors', None) is not None:
                colors = np.ascontiguousarray(sess.colors[keep], dtype=np.float64)
            # Prefer the float64 timestamps field (see _resolve_scan_export_arrays).
            ts64 = (np.asarray(sess.timestamps, dtype=np.float64)[keep]
                    if sess.timestamps is not None
                    and sess.timestamps.shape[0] == sess.positions.shape[0]
                    else None)
        def _get(slug):
            if slug == 'timestamp' and ts64 is not None:
                return ts64
            return extras.get(slug)
    elif scan_entry.points:
        xyz = np.ascontiguousarray(np.asarray(scan_entry.points, dtype=np.float64)[:, :3])
        _inline = scan_entry.scalar_columns or {}
        _n = xyz.shape[0]
        def _get(slug):
            return (np.asarray(_inline[slug])
                    if slug in _inline and len(_inline[slug]) == _n else None)
    elif scan_entry.file_path:
        if not os.path.isfile(scan_entry.file_path):
            raise ValueError(f"Scan file not found: {scan_entry.file_path}")
        xyz, fcolors, fintensity = _load_pointcloud_arrays(
            scan_entry.file_path, scan_entry.ascii_format)
        xyz = np.ascontiguousarray(np.asarray(xyz, dtype=np.float64)[:, :3])
        colors = np.asarray(fcolors, dtype=np.float64) if fcolors is not None else None
        intensity = np.asarray(fintensity, dtype=np.float64) if fintensity is not None else None
        _col_map = _read_scan_columns_from_file(scan_entry.file_path, scan_entry.ascii_format)
        def _get(slug):
            return _col_map.get(slug)
    else:
        raise ValueError("Scan entry has no points, file_path, or session_id")

    # Translation.
    t = getattr(scan_entry, 'translation', None)
    if t is not None:
        tt = np.asarray(t, dtype=np.float64)
        if tt.shape != (3,):
            raise ValueError(f"translation must be [tx, ty, tz]; got {t!r}")
        xyz = xyz + tt

    n = xyz.shape[0]
    # Intensity: prefer the channel resolved above, else an 'intensity' column.
    if intensity is None:
        icol = _get('intensity')
        if icol is not None and len(icol) == n:
            intensity = np.asarray(icol, dtype=np.float64)
    # Colour: prefer the resolved channel, else r/g/b columns if the scan has them.
    if colors is None:
        rgb = [_get(s) for s in _DATA_COLOR_SLUGS]
        if all(c is not None and len(c) == n for c in rgb):
            colors = np.column_stack(rgb).astype(np.float64)

    # The ordered scalar columns the user kept (everything that isn't geometry/
    # colour/intensity handled above). Honor the chosen order when given.
    chosen = getattr(scan_entry, 'columns', None)
    wanted = ([s for s in chosen if s not in _DATA_GEOMETRY_SLUGS + _DATA_COLOR_SLUGS]
              if chosen else list(_SCAN_EXPORT_SCALAR_COLUMNS))
    scalars: dict = {}
    ordered: list = []
    for slug in wanted:
        if slug == 'intensity':
            continue  # intensity is its own channel
        col = _get(slug)
        if col is not None and len(col) == n:
            scalars[slug] = np.asarray(col, dtype=np.float64)
            ordered.append(slug)

    # Drop miss rows uniformly across every channel.
    if not include_misses:
        miss = _get('is_miss')
        if miss is not None and len(miss) == n:
            mask = np.asarray(miss) == 0
            xyz = xyz[mask]
            colors = colors[mask] if colors is not None else None
            intensity = intensity[mask] if intensity is not None else None
            scalars = {k: v[mask] for k, v in scalars.items() if k != 'is_miss'}
            ordered = [s for s in ordered if s != 'is_miss']

    return {"positions": xyz, "colors": colors, "intensity": intensity,
            "scalars": scalars, "ordered": ordered}


def _write_scan_to_bytes(resolved: dict, fmt: str, base: str) -> tuple:
    """Write one resolved scan (from _resolve_scan_for_format) to the requested
    format. Returns (filename, raw_bytes). ASCII formats honor the chosen column
    order; binary/structured formats use their fixed schema (xyz + colour +
    intensity + scalars where the format supports them).
    """
    import numpy as np
    import tempfile

    xyz = resolved["positions"]
    colors = resolved["colors"]          # (N,3) 0-1 or None
    intensity = resolved["intensity"]    # (N,) or None
    scalars = resolved["scalars"]        # {slug: (N,)}
    ordered = resolved["ordered"]        # ordered scalar slugs
    n = xyz.shape[0]
    fmt = fmt.lower()

    # ---- ASCII text (xyz / txt / csv) — full column control ----
    if fmt in ("xyz", "txt", "csv"):
        delim = "," if fmt == "csv" else " "
        prefix = "" if fmt == "csv" else "# "
        # Column order: x y z, then colour (if present), then the ordered scalars.
        cols = ["x", "y", "z"]
        if colors is not None:
            cols += ["r255", "g255", "b255"]
        if intensity is not None:
            cols += ["intensity"]
        cols += ordered
        rgb = (np.clip(np.rint(colors * 255.0), 0, 255).astype(int)
               if colors is not None else None)
        lines = [f"{prefix}{delim.join(cols)}"]
        for i in range(n):
            row = [f"{xyz[i,0]:.6f}", f"{xyz[i,1]:.6f}", f"{xyz[i,2]:.6f}"]
            if colors is not None:
                row += [str(rgb[i, 0]), str(rgb[i, 1]), str(rgb[i, 2])]
            if intensity is not None:
                row += [f"{float(intensity[i]):.4f}"]
            for s in ordered:
                row.append(str(scalars[s][i]))
            lines.append(delim.join(row))
        return f"{base}.{fmt}", ("\n".join(lines)).encode("utf-8")

    # ---- OBJ (vertices only) ----
    if fmt == "obj":
        lines = [f"# Scan exported from Phytograph", f"# {n} points"]
        for i in range(n):
            lines.append(f"v {xyz[i,0]:.6f} {xyz[i,1]:.6f} {xyz[i,2]:.6f}")
        return f"{base}.obj", ("\n".join(lines)).encode("utf-8")

    # ---- PLY (ascii; preserves colour + scalar fields) ----
    if fmt == "ply":
        header = ["ply", "format ascii 1.0", f"element vertex {n}",
                  "property float x", "property float y", "property float z"]
        if colors is not None:
            header += ["property uchar red", "property uchar green", "property uchar blue"]
        if intensity is not None:
            header += ["property float intensity"]
        for s in ordered:
            header += [f"property float {s}"]
        header.append("end_header")
        rgb = (np.clip(np.rint(colors * 255.0), 0, 255).astype(int)
               if colors is not None else None)
        lines = header
        for i in range(n):
            row = f"{xyz[i,0]:.6f} {xyz[i,1]:.6f} {xyz[i,2]:.6f}"
            if colors is not None:
                row += f" {rgb[i,0]} {rgb[i,1]} {rgb[i,2]}"
            if intensity is not None:
                row += f" {float(intensity[i]):.6f}"
            for s in ordered:
                row += f" {float(scalars[s][i]):.6f}"
            lines.append(row)
        return f"{base}.ply", ("\n".join(lines)).encode("utf-8")

    # ---- LAS / LAZ (laspy) ----
    if fmt in ("las", "laz"):
        import laspy
        point_format = 3 if (colors is not None) else 1  # 1/3 carry intensity; 3 adds RGB
        header = laspy.LasHeader(point_format=point_format, version="1.4")
        # Extra dimensions for any scalar columns (incl is_miss) so nothing is lost.
        for s in ordered:
            header.add_extra_dim(laspy.ExtraBytesParams(name=s, type=np.float32))
        las = laspy.LasData(header)
        las.header.offsets = np.floor(xyz.min(axis=0)) if n else [0, 0, 0]
        las.header.scales = [0.001, 0.001, 0.001]
        las.x, las.y, las.z = xyz[:, 0], xyz[:, 1], xyz[:, 2]
        if intensity is not None:
            iv = np.asarray(intensity, dtype=np.float64)
            rng = iv.max() - iv.min() if n else 0
            las.intensity = (((iv - iv.min()) / rng * 65535).astype(np.uint16)
                             if rng > 0 else np.zeros(n, dtype=np.uint16))
        if colors is not None:
            c = np.clip(colors, 0, 1)
            las.red = (c[:, 0] * 65535).astype(np.uint16)
            las.green = (c[:, 1] * 65535).astype(np.uint16)
            las.blue = (c[:, 2] * 65535).astype(np.uint16)
        for s in ordered:
            las[s] = np.asarray(scalars[s], dtype=np.float32)
        ext = ".laz" if fmt == "laz" else ".las"
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tpath = tmp.name
        try:
            las.write(tpath)
            with open(tpath, "rb") as fh:
                return f"{base}{ext}", fh.read()
        finally:
            os.remove(tpath)

    # ---- E57 (pye57) ----
    if fmt == "e57":
        import pye57
        with tempfile.NamedTemporaryFile(suffix=".e57", delete=False) as tmp:
            tpath = tmp.name
        try:
            e = pye57.E57(tpath, mode="w")
            data = {
                "cartesianX": np.ascontiguousarray(xyz[:, 0]),
                "cartesianY": np.ascontiguousarray(xyz[:, 1]),
                "cartesianZ": np.ascontiguousarray(xyz[:, 2]),
            }
            if intensity is not None:
                data["intensity"] = np.ascontiguousarray(intensity.astype(np.float64))
            if colors is not None:
                c = np.clip(colors, 0, 1) * 255.0
                data["colorRed"] = np.ascontiguousarray(c[:, 0])
                data["colorGreen"] = np.ascontiguousarray(c[:, 1])
                data["colorBlue"] = np.ascontiguousarray(c[:, 2])
            e.write_scan_raw(data, name=base)
            e.close()
            with open(tpath, "rb") as fh:
                return f"{base}.e57", fh.read()
        finally:
            os.remove(tpath)

    raise ValueError(f"Unsupported scan export format: {fmt}")


def _read_scan_columns_from_file(file_path: str, ascii_format: Optional[str]) -> dict:
    """Read every NON-geometry column the ASCII_format declares into a
    slug -> (N,) float array dict. The export column picker may request any
    column the file carries (reflectance, row, column, r/g/b, …), so we read all
    of them — not just the LAD subset — keyed by their format token."""
    import numpy as np

    fmt = ascii_format or _detect_ascii_format(file_path)
    tokens = fmt.split()
    # Read every token that names a real column (skip x/y/z geometry and the
    # 'skip' placeholder). First occurrence of a repeated token wins.
    col_idx: dict = {}
    for i, tok in enumerate(tokens):
        if tok in ('x', 'y', 'z', 'skip') or tok in col_idx:
            continue
        col_idx[tok] = i
    if not col_idx:
        return {}
    need = max(col_idx.values()) + 1
    rows: dict = {c: [] for c in col_idx}
    with open(file_path) as f:
        for line in f:
            line = line.strip()
            if not line or line[0] in "#/":
                continue
            parts = line.split()
            if len(parts) < need:
                continue
            try:
                parsed = {c: float(parts[col_idx[c]]) for c in col_idx}
            except (ValueError, IndexError):
                continue
            for c in col_idx:
                rows[c].append(parsed[c])
    return {c: np.asarray(v, dtype=np.float64) for c, v in rows.items()}


def _do_scan_export(request: "ScanExportRequest") -> dict:
    """Export the request's scans, one data file per scan. Two modes:

    * write_xml=True  → a Helios scan bundle: <base>.xml + <base>_<id>.xyz, via
      PyHelios exportScans() (re-loadable as scans; data always Helios ASCII).
    * write_xml=False → one <base>_<id>.<fmt> per scan in `data_format`
      (las/laz/ply/xyz/csv/txt/obj/e57), written directly — no XML, no PyHelios.

    Returns {"success", "files": [{"name", "data"(base64), "is_xml"}], ...}. The
    renderer writes every returned file into the user-chosen folder.
    """
    import base64

    if not request.scans:
        return {"success": False, "error": "No scans to export"}

    raw = request.base_name or "scans"
    base = os.path.splitext(os.path.basename(raw))[0] or "scans"

    try:
        if request.write_xml:
            return _do_scan_export_xml(request, base)
        return _do_scan_export_data(request, base)
    except Exception as e:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        return {"success": False, "error": f"Scan export failed: {e}"}


def _do_scan_export_xml(request: "ScanExportRequest", base: str) -> dict:
    """XML+data mode: Helios scan bundle via PyHelios exportScans()."""
    import base64
    import tempfile
    import math as _math

    try:
        from pyhelios import LiDARCloud
    except ImportError as e:
        return {"success": False, "error": f"PyHelios not installed: {e}"}

    cloud = LiDARCloud()
    cloud.disableMessages()
    total_points = 0
    for scan_entry in request.scans:
        xyz, labels, vals = _resolve_scan_export_arrays(
            scan_entry, request.include_misses)
        origin = scan_entry.origin
        theta_min, theta_max = scan_entry.theta_min, scan_entry.theta_max
        phi_min, phi_max = scan_entry.phi_min, scan_entry.phi_max
        if scan_entry.n_theta and scan_entry.n_phi:
            n_theta, n_phi = int(scan_entry.n_theta), int(scan_entry.n_phi)
        else:
            aspect = (theta_max - theta_min) / max(phi_max - phi_min, 1e-10)
            n_phi = max(int(_math.sqrt(xyz.shape[0] / max(aspect, 0.01))), 10)
            n_theta = max(int(xyz.shape[0] / n_phi), 10)
        column_format = ['x', 'y', 'z'] + labels
        # Initial scanner heading (degrees -> radians). For a raster scan PyHelios
        # exportScans() serializes this as a <scanAzimuthOffset> tag (v0.1.23+); for a
        # spinning scan addScanSpinning has no offset kwarg, so the heading is folded
        # into the trajectory orientation below. Either way it round-trips.
        azimuth_offset_rad = _math.radians(float(scan_entry.scan_azimuth_offset or 0.0))
        if scan_entry.scan_pattern == "spinning_multibeam":
            # Re-loadable multibeam scan. PyHelios v0.1.24 removed addScanMultibeam;
            # a spinning scan is now registered through addScanSpinning, which takes
            # per-channel ELEVATION angles (rad, above horizon — no zenith conversion),
            # an azimuth step (rad/firing-step), a PRF, and a trajectory. exportScans
            # still writes <scanPattern>spinning_multibeam</scanPattern> +
            # <beamElevationAngles>, and additionally a trajectory sidecar CSV +
            # <PRF>/<azimuthStep> tags, so the bundle reloads as a SPINNING scan.
            elevs = scan_entry.beam_elevation_angles_deg or []
            if not elevs:
                return {"success": False,
                        "error": "Spinning-multibeam scan has no beam elevation angles to export"}
            beam_elevation_angles = [_math.radians(float(e)) for e in elevs]
            # Azimuth resolution per firing step from the requested Nphi over the
            # azimuth sweep. Guard against a degenerate zero-width / zero-count span.
            phi_span = _math.radians(phi_max) - _math.radians(phi_min)
            azimuth_step = (phi_span / int(n_phi)) if (n_phi and phi_span > 0) else _math.radians(1.0)
            if azimuth_step <= 0:
                azimuth_step = _math.radians(1.0)
            # These points are already computed; the scan object is just a container
            # for serialization. addScanSpinning still requires a trajectory, so use
            # the documented "spin in place" idiom: two coincident poses one
            # acquisition-duration apart. PyHelios derives steps_per_rev = 2*pi /
            # azimuth_step and demands PRF * duration >= steps_per_rev (at least one
            # azimuth step fires). With a 1 s window, PRF = steps_per_rev gives
            # exactly one revolution. PRF is otherwise immaterial for a pre-supplied
            # cloud — the stored points are not re-traced.
            steps_per_rev = max(int(round((2.0 * _math.pi) / azimuth_step)), 1)
            pulse_rate_hz = float(steps_per_rev)  # over the 1 s trajectory below
            # The scanner heading (scan_azimuth_offset) has no addScanSpinning kwarg,
            # so fold it into the platform orientation as a right-hand yaw about +z:
            # q = (0, 0, sin(h/2), cos(h/2)). It then round-trips via the trajectory.
            half = 0.5 * azimuth_offset_rad
            heading_quat = [0.0, 0.0, _math.sin(half), _math.cos(half)]
            origin_xyz = [float(origin[0]), float(origin[1]), float(origin[2])]
            cloud.addScanSpinning(
                beam_elevation_angles=beam_elevation_angles,
                azimuth_step=azimuth_step,
                pulse_rate_hz=pulse_rate_hz,
                traj_t=[0.0, 1.0],
                traj_pos=[origin_xyz, origin_xyz],
                traj_rot=[heading_quat, heading_quat],
                rot_is_quaternion=True,
                exit_diameter=float(scan_entry.beam_exit_diameter or 0.0),
                beam_divergence=float(scan_entry.beam_divergence or 0.0) / 1000.0,
                column_format=column_format,
            )
        else:
            cloud.addScan(
                origin=[float(origin[0]), float(origin[1]), float(origin[2])],
                Ntheta=int(n_theta),
                theta_range=(_math.radians(theta_min), _math.radians(theta_max)),
                Nphi=int(n_phi),
                phi_range=(_math.radians(phi_min), _math.radians(phi_max)),
                exit_diameter=float(scan_entry.beam_exit_diameter or 0.0),
                beam_divergence=float(scan_entry.beam_divergence or 0.0) / 1000.0,
                column_format=column_format,
                scan_azimuth_offset=azimuth_offset_rad,
            )
        sid = cloud.getScanCount() - 1
        if xyz.shape[0] > 0:
            dirs = _directions_from_origin(xyz, origin)
            cloud.addHitPointsWithData(sid, xyz, dirs, labels, vals)
        total_points += int(xyz.shape[0])

    with tempfile.TemporaryDirectory() as tmpdir:
        xml_path = os.path.join(tmpdir, f"{base}.xml")
        cloud.exportScans(xml_path)
        # PyHelios exportScans() writes only <scan> blocks. To round-trip the
        # scene's grids (e.g. sphere.xml's voxel box), inject the requested
        # <grid> blocks just before the closing </helios>.
        if request.grids:
            _inject_grids_into_helios_xml(xml_path, request.grids)
        files = []
        for name in sorted(os.listdir(tmpdir)):
            with open(os.path.join(tmpdir, name), "rb") as fh:
                files.append({
                    "name": name,
                    "data": base64.b64encode(fh.read()).decode("ascii"),
                    "is_xml": name.lower().endswith(".xml"),
                })

    return {"success": True, "files": files, "point_count": total_points,
            "scan_count": len(request.scans)}


def _do_scan_export_data(request: "ScanExportRequest", base: str) -> dict:
    """Data-only mode: one <base>_<id>.<fmt> per scan in `data_format`."""
    import base64

    fmt = (request.data_format or "xyz").lower()
    if fmt not in ("las", "laz", "ply", "xyz", "csv", "txt", "obj", "e57"):
        return {"success": False, "error": f"Unsupported data format: {fmt}"}

    files = []
    total_points = 0
    for i, scan_entry in enumerate(request.scans):
        resolved = _resolve_scan_for_format(scan_entry, request.include_misses)
        total_points += int(resolved["positions"].shape[0])
        name, raw_bytes = _write_scan_to_bytes(resolved, fmt, f"{base}_{i}")
        files.append({
            "name": name,
            "data": base64.b64encode(raw_bytes).decode("ascii"),
            "is_xml": False,
        })

    return {"success": True, "files": files, "point_count": total_points,
            "scan_count": len(request.scans)}


@app.post("/api/scan/export-xml")
async def scan_export_xml(request: "ScanExportRequest"):
    """Export scans to a Helios XML + per-scan ASCII bundle (base64 files)."""
    return _do_scan_export(request)


# ==================== SYNTHETIC LIDAR SCANNING ====================
# True ray-traced synthetic scanning via the PyHelios `lidar` plugin. The scene
# geometry (plant + imported meshes, already world-space transformed by the
# renderer) is loaded into a Helios Context; each placed scanner becomes an
# addScan() with its ScanParameters; syntheticScan() ray-traces the scene and the
# resulting hit points are returned as a point cloud — respecting occlusion,
# scanner position, field of view, and resolution (unlike random surface sampling).

class LidarScanMaterial(BaseModel):
    """A textured material group on a scan mesh.

    ``texture_data`` is a base64-encoded image (PNG/JPG). When it carries an
    alpha channel, Helios uses that channel as a transparency mask during ray
    tracing — leaf textures are leaf-shaped cutouts on a transparent
    background, so rays only register hits where the leaf is opaque instead of
    on the full rectangular quad. ``triangle_indices`` are ordinals into the
    mesh's ``triangles`` array that use this material.
    """
    name: str
    texture_data: str  # base64 PNG/JPG
    has_alpha: bool = False
    triangle_indices: List[int]


class LidarScanMesh(BaseModel):
    """A single mesh to load into the scannable scene (world-space coordinates)."""
    vertices: List[List[float]]  # [[x, y, z], ...]
    triangles: List[List[int]]   # [[i, j, k], ...] vertex indices
    colors: Optional[List[List[float]]] = None  # per-vertex [[r, g, b], ...] (0-1)
    uv_coordinates: Optional[List[List[float]]] = None  # per-vertex [[u, v], ...]
    materials: Optional[List[LidarScanMaterial]] = None  # textured material groups
    # Optional per-triangle organ-type code (parallel to `triangles`). Sent only
    # when the user opts into organ carry; stamped onto primitives so the scan can
    # sample it per hit (see _load_scan_mesh / _ORGAN_LABEL_TO_CODE).
    organ_codes: Optional[List[int]] = None


class LidarScanScanner(BaseModel):
    """A single scanner position + acquisition geometry (mirrors ScanParameters)."""
    id: str                      # renderer scan id — results are returned keyed by this
    origin: List[float]          # [x, y, z] scanner position
    # Acquisition pattern. "raster" = uniform Ntheta x Nphi grid (default).
    # "spinning_multibeam" = one fixed laser channel per beam_elevation_angles_deg
    # entry (rotating multi-channel sensor); n_theta / theta_*_deg are ignored.
    scan_pattern: str = "raster"
    # Per-channel beam elevation angles, degrees above horizon (multibeam only).
    # Converted to zenith (zenith = 90 - elevation) before the pyhelios call.
    beam_elevation_angles_deg: Optional[List[float]] = None
    n_theta: int                 # zenith samples (Ntheta)
    n_phi: int                   # azimuth samples (Nphi)
    theta_min_deg: float         # zenith angle range (degrees, 0-180)
    theta_max_deg: float
    phi_min_deg: float           # azimuth angle range (degrees, 0-360)
    phi_max_deg: float
    # How many returns the pulse reports (an instrument property):
    #   "single" — at most one return per pulse (RETURN_MODE_SINGLE, maxReturns=1)
    #              chosen by return_selection (strongest/first/last).
    #   "multi"  — all detected returns up to max_returns (RETURN_MODE_MULTI).
    # For an idealized exact scan, send rays_per_pulse=1 (a run option), which the
    # engine reduces to one exact return per pulse for either mode.
    return_mode: str = "single"
    max_returns: int = 5                 # "multi" only — cap on returns per pulse
    return_selection: str = "strongest"  # "single" only — strongest | first | last
    exit_diameter_m: float = 0.0
    beam_divergence_mrad: float = 0.0

    @model_validator(mode="before")
    @classmethod
    def _migrate_return_type(cls, data):
        # Back-compat: an older client (or persisted request) sends `return_type`
        # ("single"/"multi") and no `return_mode`. Map it so the request still
        # validates. A present `return_mode` wins.
        if isinstance(data, dict) and "return_mode" not in data and "return_type" in data:
            legacy = data.get("return_type")
            data = {**data, "return_mode": "multi" if legacy == "multi" else "single"}
        return data
    # Measurement-error model (applies to both return types). Defaults of 0
    # disable each effect, preserving an ideal noise-free, perfectly-level scan.
    range_noise_m: float = 0.0       # Gaussian along-beam range noise stddev (meters)
    angle_noise_mrad: float = 0.0    # Gaussian beam-pointing jitter stddev (milliradians)
    # Residual scanner tilt away from plumb (a dual-axis inclinometer's two
    # angles). Roll is applied first, then pitch. Degrees here; converted to the
    # radians pyhelios expects below.
    tilt_roll_deg: float = 0.0
    tilt_pitch_deg: float = 0.0
    # Initial scanner heading in the horizontal plane (degrees). Applied by the
    # synthetic-scan generator via PyHelios addScan's scan_azimuth_offset (v0.1.23+)
    # at the static call site below (moving scans take heading from the trajectory).
    scan_azimuth_offset_deg: float = 0.0
    # Moving-platform trajectory. When present, this scanner is driven by
    # addScanMoving instead of the static addScan: the scanner
    # pose walks the trajectory over the sweep, `origin` is ignored (the
    # trajectory supplies position), and every hit/miss records its own per-pulse
    # emission origin + real timestamp. `pulse_rate_hz` sets the per-pulse spacing
    # in time (t = t0 + ordinal / pulse_rate_hz); the Ntheta*Nphi pulse budget
    # therefore determines how much of the flight the sweep spans. Static scanners
    # leave this None and behave exactly as before.
    trajectory: Optional[PoseStream] = None
    pulse_rate_hz: Optional[float] = None


class LidarScanRequest(BaseModel):
    """Request model for a synthetic LiDAR scan."""
    meshes: List[LidarScanMesh]
    scanners: List[LidarScanScanner]
    # Extra per-hit scalar fields to record, beyond the always-read standard set
    # (intensity, distance, timestamp, target_index, target_count). Two kinds:
    #   - engine-produced optionals ("deviation", "nRaysHit") — recorded by
    #     syntheticScan for multi-return scans; only need to be READ (added to
    #     fields_to_read). They must NOT go in column_format (there is no such
    #     primitive data to sample).
    #   - primitive extras (e.g. "reflectance" or any custom primitive label) —
    #     treated as a column-format label so syntheticScan samples that named
    #     primitive data from the struck primitive onto each hit.
    # See _ENGINE_OPTIONAL_FIELDS for the split.
    extra_fields: List[str] = []
    # Which of the standard fields to keep on the resulting cloud. Used only on
    # the misses-on (session) path to decide which standard scalars become octree
    # extra-dims, so the color-by list matches the flat-cloud path. None => keep
    # all standards (backward compatible for older callers / standalone launches).
    retained_standard_fields: Optional[List[str]] = None
    # Full-waveform tuning (used only when a scanner has return_type == "multi").
    rays_per_pulse: int = 100
    pulse_distance_threshold: float = 0.02
    # Synthetic-scan run options (Synthetic Scan Options popup):
    #   record_misses  — include sky/miss points (rays that hit nothing). When on,
    #                     the result is routed through a cloud session so the
    #                     existing miss overlay + LAD work.
    #   scan_grid_only — restrict ray-tracing to the cells of `grid`.
    #   grid           — voxel grid to crop to (required when scan_grid_only).
    record_misses: bool = False
    scan_grid_only: bool = False
    grid: Optional[HeliosGrid] = None
    # Soft cap (MB) on the transient ray-tracing scratch buffers, set via
    # LiDARCloud.setSyntheticScanMemoryBudget. Bounds peak RAM by chunking the
    # beam fan-out (not the output cloud) on large scans. None => leave Helios's
    # automatic path-dependent default (4 GiB CPU / 8 GiB GPU) untouched.
    synthetic_scan_memory_budget_mb: Optional[int] = None


class LidarScanResult(BaseModel):
    """Per-scanner scan result."""
    scanner_id: str
    points: List[List[float]] = []                # [[x, y, z], ...] hit points
    colors: Optional[List[List[float]]] = None    # [[r, g, b], ...] per-point (0-1)
    # Per-point scalar fields (name -> values aligned 1:1 with points). Includes
    # "intensity" plus distance/timestamp/target_index/target_count and any
    # requested extra_fields that the engine actually recorded.
    scalars: Dict[str, List[float]] = {}
    num_points: int = 0


class LidarScanResponse(BaseModel):
    """Response model for synthetic LiDAR scan results — one entry per scanner."""
    success: bool
    results: List[LidarScanResult] = []
    error: Optional[str] = None


# Standard per-hit scalar keys syntheticScan records (see helios-core lidar
# LiDAR.cpp). "intensity" is the beam/normal dot product (can be negative) scaled
# by reflectivity; we abs() it when surfacing so it reads as a 0..1-ish magnitude.
_LIDAR_STANDARD_HIT_FIELDS = ["intensity", "distance", "timestamp", "target_index", "target_count"]

# Optional per-hit fields syntheticScan computes itself (for multi-return scans)
# but does NOT record unless asked. They are READ (added to fields_to_read) when
# requested via extra_fields, but must never be passed in column_format — there
# is no primitive data by these names to sample. Anything in extra_fields that
# is NOT one of these is treated as a primitive-data label (e.g. "reflectance").
_ENGINE_OPTIONAL_FIELDS = frozenset({"deviation", "nRaysHit"})

# Per-hit fields whose magnitude (GPS Adjusted-Standard time ~3.5e8 s) needs
# float64; the float32 getHitDataArray() path quantizes them to ~10 s, which
# corrupts the moving-platform LAD trajectory join (timestamp is the join key).
# These are read additionally via the columnar getHitDataColumnArray() float64
# path and routed to the dedicated CloudSession.timestamps field — the float32
# copy stays in extras for octree color-by / backfill-UI display.
_LIDAR_FLOAT64_HIT_FIELDS = frozenset({"timestamp"})

# Soft ceiling on a moving scan's total pulse count. A real PRF over a long flight
# is genuinely millions of pulses (we fire them — physically faithful), but past
# this we surface a warning so a 60 s flight at 1 MHz (60M pulses) doesn't lock up
# unannounced. Not a hard cap; the scan still runs.
_MOVING_SCAN_PULSE_WARN = 20_000_000


def _derive_moving_scan_grid(n_theta: int, n_phi_per_rev: int, pulse_rate_hz: float,
                             traj_t: list) -> dict:
    """Derive the full-flight scan grid for a moving-platform scan.

    A real spinning sensor fires continuously at its pulse rate (PRF), spinning at
    a rate set by PRF and its per-revolution resolution, for the whole flight. The
    user specifies resolution PER REVOLUTION (n_theta channels/zenith rows ×
    n_phi_per_rev azimuth steps); the PRF is the instrument's fixed laser spec.
    Everything else follows from the trajectory's duration:

        pulses_per_rev = n_theta * n_phi_per_rev
        rotation_rate  = PRF / pulses_per_rev          (revolutions / second)
        duration       = traj_t[-1] - traj_t[0]        (seconds, from the path)
        n_revolutions  = rotation_rate * duration
        n_phi_total    = round(n_phi_per_rev * n_revolutions)
        phi_max        = phi_min + n_revolutions * 2pi  (the head spins that many
                                                         times across the flight)
        total_pulses   = n_theta * n_phi_total  (≈ PRF * duration)

    So addScanMoving fires ~PRF*duration pulses spanning [t0, t0+duration] — the
    cloud covers the entire path. Returns a dict with n_phi_total, phi_max_rad
    (relative to phi_min=0; caller offsets), rotation_rate_hz, n_revolutions,
    total_pulses, duration_s.
    """
    import math
    pulses_per_rev = max(int(n_theta) * int(n_phi_per_rev), 1)
    duration = float(traj_t[-1]) - float(traj_t[0]) if len(traj_t) > 1 else 0.0
    rotation_rate = float(pulse_rate_hz) / pulses_per_rev
    n_rev = rotation_rate * duration
    # A zero-duration (single-pose) trajectory or a degenerate rate falls back to a
    # single revolution so the scan still produces a sensible static-like sweep.
    if not (n_rev > 0) or not math.isfinite(n_rev):
        n_rev = 1.0
    n_phi_total = max(int(round(int(n_phi_per_rev) * n_rev)), 1)
    phi_max_rad = n_rev * 2.0 * math.pi
    return {
        "n_phi_total": n_phi_total,
        "phi_max_rad": phi_max_rad,
        "rotation_rate_hz": rotation_rate,
        "n_revolutions": n_rev,
        "total_pulses": int(n_theta) * n_phi_total,
        "duration_s": duration,
    }


def _load_scan_mesh(ctx, mesh: "LidarScanMesh", tmpdir: str) -> None:
    """Load one scan mesh into the Helios ``ctx``, honoring textured materials.

    Geometry arrives world-space transformed. Triangles belonging to a textured
    material (a material with ``texture_data`` and per-triangle indices) are
    loaded via ``addTrianglesFromArraysTextured`` so Helios ray-traces against
    the texture's alpha channel (leaf-shaped cutouts) instead of the full quad.
    Any remaining triangles — flat-colored organs, or the whole mesh when it has
    no textures — go through the color-only ``addTrianglesFromArrays`` path.

    ``tmpdir`` is a directory whose lifetime spans the Context; decoded texture
    images are written there because Helios requires real files on disk.
    """
    import base64

    verts = np.asarray(mesh.vertices, dtype=np.float32)
    tris = np.asarray(mesh.triangles, dtype=np.int32)
    if verts.ndim != 2 or verts.shape[1] != 3 or len(verts) < 3:
        return
    if tris.ndim != 2 or tris.shape[1] != 3 or len(tris) < 1:
        return

    colors = None
    if mesh.colors and len(mesh.colors) == len(verts):
        colors = np.asarray(mesh.colors, dtype=np.float32)

    # Optional per-triangle organ codes (sent only when the user opts in). When
    # present they're stamped onto the loaded primitives as "organ" int data so
    # the synthetic scan samples organ type per hit (LiDAR.cpp column_format).
    organ_arr = None
    if mesh.organ_codes is not None and len(mesh.organ_codes) == len(tris):
        organ_arr = np.asarray(mesh.organ_codes, dtype=np.int32)

    # Decide which triangles are textured. We need per-vertex UVs (one [u,v] per
    # vertex) and at least one material that carries a usable texture image. If
    # any precondition fails we fall back to the plain color path for the whole
    # mesh — never silently drop triangles.
    uvs = None
    if mesh.uv_coordinates and len(mesh.uv_coordinates) == len(verts):
        uvs = np.asarray(mesh.uv_coordinates, dtype=np.float32)

    textured_assignment = {}  # triangle ordinal -> material slot index
    texture_files = []        # slot index -> on-disk texture path
    if uvs is not None and mesh.materials:
        for mat in mesh.materials:
            if not mat.texture_data or not mat.triangle_indices:
                continue
            try:
                img_bytes = base64.b64decode(mat.texture_data)
            except Exception:
                continue
            # Decode to a real file; keep the original extension so Helios reads
            # the alpha channel correctly (PNG carries alpha; JPG does not).
            ext = ".png" if mat.has_alpha else ".jpg"
            slot = len(texture_files)
            tex_path = os.path.join(tmpdir, f"scan_tex_{id(mesh) & 0xffffff}_{slot}{ext}")
            try:
                with open(tex_path, "wb") as fh:
                    fh.write(img_bytes)
            except Exception:
                continue
            texture_files.append(tex_path)
            for ti in mat.triangle_indices:
                if 0 <= ti < len(tris):
                    textured_assignment[ti] = slot

    if not textured_assignment:
        # No usable textures — original color-only path for the whole mesh.
        uuids = ctx.addTrianglesFromArrays(verts, tris, colors=colors)
        if organ_arr is not None and uuids:
            # addTrianglesFromArrays returns UUIDs in input-triangle order.
            ctx.setPrimitiveDataInt(list(uuids), "organ", organ_arr.tolist())
        return

    textured_tri_idx = np.array(sorted(textured_assignment.keys()), dtype=np.int64)
    textured_mask = np.zeros(len(tris), dtype=bool)
    textured_mask[textured_tri_idx] = True

    # Textured triangles: feed per-triangle material_ids (the slot each one
    # uses) so multi-texture meshes (e.g. leaves + bark) map correctly. UVs are
    # per-vertex and shared across both sub-meshes.
    tex_tris = tris[textured_mask]
    material_ids = np.array(
        [textured_assignment[int(ti)] for ti in textured_tri_idx], dtype=np.uint32
    )
    if organ_arr is None:
        # Default path: one bulk multi-textured call (unchanged behavior).
        ctx.addTrianglesFromArraysTextured(
            verts, tex_tris, uvs, texture_files, material_ids=material_ids
        )
    else:
        # Organ carry on: load one material slot at a time. A single-texture call
        # returns UUIDs in input-triangle order on both the C++ and fallback
        # backends, whereas the bulk multi-textured C++ path regroups triangles by
        # material id — which would scramble a per-triangle organ array. Slotting
        # keeps the mapping correct regardless of which backend is compiled in.
        for slot in range(len(texture_files)):
            local = np.nonzero(material_ids == slot)[0]
            if len(local) == 0:
                continue
            slot_uuids = ctx.addTrianglesFromArraysTextured(
                verts, tex_tris[local], uvs, texture_files[slot], material_ids=None
            )
            if slot_uuids:
                slot_codes = [int(organ_arr[textured_tri_idx[p]]) for p in local]
                ctx.setPrimitiveDataInt(list(slot_uuids), "organ", slot_codes)

    # Everything else keeps the flat-color path (stems, flowers, untextured
    # organs that share the mesh).
    rest_tris = tris[~textured_mask]
    if len(rest_tris) > 0:
        rest_uuids = ctx.addTrianglesFromArrays(verts, rest_tris, colors=colors)
        if organ_arr is not None and rest_uuids:
            # rest_tris is tris[~textured_mask]; organ_arr[~textured_mask] is the
            # same ascending-index order addTrianglesFromArrays returns.
            ctx.setPrimitiveDataInt(list(rest_uuids), "organ",
                                    organ_arr[~textured_mask].tolist())


def _apply_return_mode(lidar, scan_id: int, s: "LidarScanScanner",
                       ReturnMode, selection_map: dict) -> None:
    """Set a freshly-added scan's stored return mode from the request.

    The stored mode is honored by syntheticScan() when called with rays_per_pulse
    and no explicit return_mode override:
      - "single": RETURN_MODE_SINGLE, maxReturns=1, selection per request.
      - "multi" : RETURN_MODE_MULTI, maxReturns = request max_returns.

    For an idealized exact scan the caller sends rays_per_pulse=1, which the engine
    reduces to one exact return per pulse regardless of the stored mode — so there
    is no separate "clean" mode to handle here.
    """
    if s.return_mode == "multi":
        lidar.setScanReturnMode(scan_id, ReturnMode.MULTI)
        lidar.setScanMaxReturns(scan_id, max(1, int(s.max_returns)))
    else:  # "single"
        lidar.setScanReturnMode(scan_id, ReturnMode.SINGLE)
        lidar.setScanMaxReturns(scan_id, 1)
        sel = selection_map.get(s.return_selection)
        if sel is not None:
            lidar.setScanSingleReturnSelection(scan_id, sel)


def _do_lidar_scan(request: LidarScanRequest, progress=None) -> dict:
    """
    Perform a true ray-traced synthetic LiDAR scan of the supplied geometry.

    The renderer sends:

    * ``meshes`` — plant / imported-mesh geometry (already world-space transformed)
    * ``scanners`` — one entry per visible scanner marker, each carrying its
      renderer ``id`` plus its ``ScanParameters`` (origin, angular field of view in
      degrees, resolution, return type, beam optics)

    All meshes load into one Helios ``Context``; every scanner becomes a
    ``LiDARCloud.addScan`` (added in request order, so the Helios scanID equals the
    scanner's request index); ``syntheticScan`` ray-traces the scene once. Hits are
    then partitioned back to their originating scanner via ``getHitScanID`` and
    returned **per scanner** — so the renderer can attach each scanner's points to
    its own scan, with intensity + scalar fields for color-by/filter.
    """
    def _report(fraction, message):
        if progress is not None:
            progress(fraction, message)

    def _ckpt():
        _cancel_checkpoint(progress)

    try:
        if not request.meshes:
            return {"success": False, "error": "No geometry to scan"}
        if not request.scanners:
            return {"success": False, "error": "No scanners defined"}

        # Total triangle count guards against accidentally feeding a huge mesh.
        total_tris = sum(len(m.triangles) for m in request.meshes)
        if total_tris < 1:
            return {"success": False, "error": "Geometry has no triangles"}

        from pyhelios import LiDARCloud, Context
        from pyhelios.LiDARCloud import ReturnMode, SingleReturnSelection

        # Shared cancel flag the C++ ray loop polls. The stream loop flips it to
        # 1 the moment this run is cancelled (disconnect or /api/cancel), so the
        # in-flight syntheticScan bails its per-scan + inner ray loops. Bind it to
        # the reporter so propagate_cancel() can mirror the Event into it.
        import ctypes as _ctypes
        _scan_cancel_flag = _ctypes.c_int(0)
        if progress is not None and getattr(progress, "bind_cancel_int", None):
            progress.bind_cancel_int(_scan_cancel_flag)

        _report(0.05, "Loading geometry")

        # Every scan samples the beam cone (fires rays_per_pulse sub-rays per pulse).
        # rays_per_pulse=1 collapses the cone to one exact ray (an idealized scan);
        # the return mode then decides how the resolved waveform is reduced to points.
        _SELECTION_MAP = {
            "strongest": SingleReturnSelection.STRONGEST,
            "first": SingleReturnSelection.FIRST,
            "last": SingleReturnSelection.LAST,
        }
        extra_fields = [f for f in request.extra_fields if f]
        # Split the requested extras: engine-produced optionals (deviation/nRaysHit)
        # only need to be READ, while primitive extras must be sampled via
        # column_format. column_format drives which custom primitive data the scan
        # samples onto hits; the standard keys are always recorded.
        engine_optionals = [f for f in extra_fields if f in _ENGINE_OPTIONAL_FIELDS]
        primitive_extras = [f for f in extra_fields if f not in _ENGINE_OPTIONAL_FIELDS]
        column_format = primitive_extras if primitive_extras else None

        # Phase timing — set PHYTOGRAPH_SCAN_PROFILE=1 to print where scan time
        # goes (mesh load / ray trace / hit extraction / post-processing). The
        # split matters: ray-trace cost scales with ray count, extraction with
        # hit count, so the profile says which knob (resolution vs. fields) helps.
        _prof_on = os.environ.get("PHYTOGRAPH_SCAN_PROFILE") == "1"
        _prof = {"start": time.perf_counter()}

        import tempfile

        # Decoded leaf/bark textures must live as real files on disk for the
        # whole Context lifetime (Helios reads them lazily during ray tracing),
        # so the temp dir wraps the Context and is cleaned up after it closes.
        _ckpt()
        with tempfile.TemporaryDirectory(prefix="phyto_scan_tex_") as _tex_dir, \
                Context() as ctx:
            # Load every mesh into the scannable scene (textured-aware).
            for mesh in request.meshes:
                _ckpt()
                _load_scan_mesh(ctx, mesh, _tex_dir)

            _prof["mesh_load"] = time.perf_counter()
            _report(0.15, "Configuring scanners")

            with LiDARCloud() as lidar:
                lidar.disableMessages()

                # Add scans in request order so Helios scanID == request index.
                # Tilt (deg→rad) and noise (range in m verbatim, angle mrad→rad)
                # are passed per-scan; 0 leaves them disabled.
                for s in request.scanners:
                    # Moving-platform scanner: the pose walks the trajectory over
                    # the sweep. addScanMoving takes the raster grid (Ntheta x Nphi)
                    # plus the trajectory + pulse rate; it fires Ntheta*Nphi pulses
                    # at t = t0 + ordinal/pulse_rate, interpolating the platform pose
                    # per pulse and recording each hit/miss's own origin + timestamp.
                    # (Spinning-multibeam motion is not yet a separate moving entry
                    # point in pyhelios; a moving scan uses the raster grid form.)
                    # A spinning-multibeam sensor rotates continuously, so it only
                    # makes sense as a moving (trajectory-driven) scan — there is no
                    # coherent "stationary spinning" capture without a time element.
                    # A truly stationary spinning capture is expressed as a trajectory
                    # with two coincident poses one revolution-duration apart. Reject
                    # a multibeam scanner that carries no trajectory.
                    if s.scan_pattern == "spinning_multibeam" and s.trajectory is None:
                        return {"success": False,
                                "error": f"Spinning-multibeam scanner '{s.id}' requires a "
                                         "trajectory: a rotating sensor is a moving-platform "
                                         "scan. For a stationary capture, use a trajectory "
                                         "with two poses at the same position one revolution "
                                         "apart."}
                    if s.trajectory is not None:
                        poses = s.trajectory.poses
                        if len(poses) == 0:
                            return {"success": False,
                                    "error": f"Moving scanner '{s.id}' has an empty trajectory"}
                        pulse_rate = s.pulse_rate_hz or 0.0
                        if pulse_rate <= 0.0:
                            return {"success": False,
                                    "error": f"Moving scanner '{s.id}' needs a positive pulse_rate_hz (PRF)"}
                        # Ntheta is the number of zenith rows fired per pulse-column:
                        # for a spinning-multibeam sensor that's the channel count
                        # (one row per beam elevation angle); for a raster sensor
                        # it's the zenith point count. Nphi from the request is the
                        # azimuth resolution PER REVOLUTION.
                        if s.scan_pattern == "spinning_multibeam":
                            elevs = s.beam_elevation_angles_deg or []
                            if not elevs:
                                return {"success": False,
                                        "error": f"Moving multibeam scanner '{s.id}' has no beam elevation angles"}
                            n_theta = len(elevs)
                            # Per-channel zenith range; addScanMoving uses a uniform
                            # theta grid, so span the channels' min..max elevation.
                            zeniths = [math.radians(90.0 - float(e)) for e in elevs]
                            theta_range = (min(zeniths), max(zeniths))
                        else:
                            n_theta = int(s.n_theta)
                            theta_range = (math.radians(s.theta_min_deg), math.radians(s.theta_max_deg))
                        traj_t = [float(p.t) for p in poses]
                        # Derive the full-flight grid: spin the head at the rate the
                        # PRF + per-rev resolution imply, for the whole flight, so the
                        # scan fires ~PRF*duration pulses spanning the trajectory.
                        grid = _derive_moving_scan_grid(
                            n_theta=n_theta, n_phi_per_rev=int(s.n_phi),
                            pulse_rate_hz=float(pulse_rate), traj_t=traj_t)
                        if grid["total_pulses"] > _MOVING_SCAN_PULSE_WARN:
                            print(f"[lidar] moving scanner '{s.id}': {grid['total_pulses']:,} "
                                  f"pulses ({grid['n_revolutions']:.0f} revolutions over "
                                  f"{grid['duration_s']:.1f}s) — large scan, may be slow",
                                  flush=True)
                        phi_min = math.radians(s.phi_min_deg)
                        sid = lidar.addScanMoving(
                            Ntheta=n_theta,
                            theta_range=theta_range,
                            Nphi=grid["n_phi_total"],
                            phi_range=(phi_min, phi_min + grid["phi_max_rad"]),
                            exit_diameter=float(s.exit_diameter_m),
                            beam_divergence=float(s.beam_divergence_mrad) * 1e-3,
                            traj_t=traj_t,
                            traj_pos=[[float(p.x), float(p.y), float(p.z)] for p in poses],
                            traj_rot=[[float(p.qx), float(p.qy), float(p.qz), float(p.qw)] for p in poses],
                            pulse_rate_hz=float(pulse_rate),
                            rot_is_quaternion=True,
                            lever_arm=[float(v) for v in s.trajectory.lever_arm],
                            boresight_rpy=[float(v) for v in s.trajectory.boresight_rpy],
                            column_format=column_format,
                            range_noise_stddev=float(s.range_noise_m),
                            angle_noise_stddev=float(s.angle_noise_mrad) * 1e-3,
                        )
                        _apply_return_mode(lidar, sid, s, ReturnMode, _SELECTION_MAP)
                    else:
                        # Static raster scan (a multibeam scanner without a
                        # trajectory was already rejected above; spinning multibeam
                        # is a moving-only pattern).
                        sid = lidar.addScan(
                            origin=[float(s.origin[0]), float(s.origin[1]), float(s.origin[2])],
                            Ntheta=int(s.n_theta),
                            theta_range=(math.radians(s.theta_min_deg), math.radians(s.theta_max_deg)),
                            Nphi=int(s.n_phi),
                            phi_range=(math.radians(s.phi_min_deg), math.radians(s.phi_max_deg)),
                            exit_diameter=float(s.exit_diameter_m),
                            beam_divergence=float(s.beam_divergence_mrad) * 1e-3,
                            column_format=column_format,
                            range_noise_stddev=float(s.range_noise_m),
                            angle_noise_stddev=float(s.angle_noise_mrad) * 1e-3,
                            scan_tilt_roll=math.radians(s.tilt_roll_deg),
                            scan_tilt_pitch=math.radians(s.tilt_pitch_deg),
                            scan_azimuth_offset=math.radians(s.scan_azimuth_offset_deg),
                        )
                        _apply_return_mode(lidar, sid, s, ReturnMode, _SELECTION_MAP)

                # Crop-to-grid: scan_grid_only restricts ray-tracing to the cells
                # of a grid, which must be defined on the cloud first via addGrid.
                scan_grid_only = bool(request.scan_grid_only and request.grid is not None)
                if scan_grid_only:
                    g = request.grid
                    lidar.addGrid(
                        center=[float(g.center[0]), float(g.center[1]), float(g.center[2])],
                        size=[float(g.size[0]), float(g.size[1]), float(g.size[2])],
                        ndiv=[int(g.nx), int(g.ny), int(g.nz)],
                    )

                # One ray pass for all scanners (one BVH build). append=False clears
                # once, then every scan contributes; hits carry their scanID.
                # record_misses is user-controlled now: when on, rays that hit
                # nothing are kept (far-field placeholder points flagged via
                # isHitMiss below) so the miss overlay + LAD can use them.
                #
                # rays_per_pulse is a single global arg: every scan fires that many
                # sub-rays across its beam cone, and each scan's stored return mode
                # (set above) reduces its waveform to single- or multi-return points.
                # rays_per_pulse=1 collapses the cone to one exact ray per pulse — the
                # idealized scan — for either mode.
                record_misses = bool(request.record_misses)
                # Optional user-set cap on the ray trace's transient buffers. Only
                # override when a positive value is supplied; otherwise leave Helios's
                # automatic path-dependent default in place.
                if request.synthetic_scan_memory_budget_mb is not None \
                        and request.synthetic_scan_memory_budget_mb > 0:
                    lidar.setSyntheticScanMemoryBudget(
                        int(request.synthetic_scan_memory_budget_mb) * 1024 * 1024)
                _prof["add_scans"] = time.perf_counter()
                _ckpt()
                # The ray trace's per-scan loop and inner ray loop honor a cancel
                # flag (see _scan_cancel_flag below), so a cancel mid-trace bails
                # the C++ pass.
                #
                # syntheticScan fires a progress callback at the start of each
                # scan's trace (serial, on this thread, outside its OpenMP regions),
                # passing scans-completed / scanCount in [0, 1]. With more than one
                # scanner that's genuine determinate progress across the long trace,
                # so map it into this op's ray-trace band [0.15, 0.85]. The callback
                # only ticks at per-scan boundaries, so a single-scanner trace can't
                # animate mid-scan — fall back to the indeterminate pulse (null
                # fraction) there so the bar moves rather than freezing. (_report
                # only queues a marker and never raises, so it is safe to invoke
                # from inside the native callback; cancellation is handled out of
                # band via _scan_cancel_flag, not by raising through C++.)
                n_scanners = len(request.scanners)
                if n_scanners > 1:
                    def _on_scan_progress(fraction, _message):
                        # fraction = scans-completed / scanCount; the scan being
                        # traced is that count + 1 (round, not int — the C++ float
                        # division lands a hair under the integer, e.g. 1/3*3≈0.9999,
                        # which int() would truncate to the wrong scan number).
                        frac = 0.0 if fraction < 0.0 else 1.0 if fraction > 1.0 else fraction
                        current = min(n_scanners, round(frac * n_scanners) + 1)
                        _report(0.15 + 0.70 * frac,
                                f"Ray-tracing scene (scan {current}/{n_scanners})")
                    lidar.setProgressCallback(_on_scan_progress)
                else:
                    _report(None, "Ray-tracing scene")
                lidar.syntheticScan(
                    ctx,
                    rays_per_pulse=int(request.rays_per_pulse),
                    pulse_distance_threshold=float(request.pulse_distance_threshold),
                    scan_grid_only=scan_grid_only,
                    record_misses=record_misses,
                    append=False,
                    cancel_flag=_scan_cancel_flag,
                )
                # Drop the native callback bridge promptly once the trace returns
                # (the LiDARCloud context manager would also release it on exit).
                if n_scanners > 1:
                    lidar.setProgressCallback(None)

                _prof["raytrace"] = time.perf_counter()
                # The C++ trace may have stopped early on a cancel — surface it as
                # ScanCancelled before we spend time/RAM extracting a partial cloud.
                _ckpt()
                _report(0.85, "Extracting hits")

                # ---- Bulk extraction. This was a per-hit Python loop doing ~13×N
                # FFI crossings (getHitScanID/getHitXYZ/getHitColor/isHitMiss +
                # doesHitDataExist/getHitData per field), which dominated scan time
                # on million-hit clouds. Pull every quantity for ALL hits in a
                # handful of FFI calls, then partition per scanner with numpy masks.
                fields_to_read = _LIDAR_STANDARD_HIT_FIELDS + engine_optionals + primitive_extras
                # A moving-platform scan records each beam's own emission origin and
                # firing index; pull them too so the result cloud carries the
                # per-beam geometry the leaf-area inversion needs (origin_x/y/z) —
                # static scans don't write these, so the columns come back all-NaN
                # and are dropped when the session is built.
                any_moving_scan = any(s.trajectory is not None for s in request.scanners)
                if any_moving_scan:
                    fields_to_read = fields_to_read + ["origin_x", "origin_y", "origin_z", "pulse_id"]
                n = lidar.getHitCount()
                if n > 0:
                    all_xyz, all_rgb = lidar.getHitsXYZRGBArrays()   # (n,3),(n,3) f32
                    all_scan_ids = lidar.getHitScanIDArray()         # (n,) int32
                    # isHitMiss is only meaningful when misses were recorded; skip
                    # the FFI pass otherwise (every hit is a real return).
                    all_miss = (lidar.getHitMissArray() if record_misses
                                else np.zeros(n, dtype=np.int32))    # (n,) int32, 1==miss
                    # Each field is (n,) f32, NaN where the label is absent for a hit.
                    all_fields = {f: lidar.getHitDataArray(f) for f in fields_to_read}
                    # Precision-sensitive fields (timestamp) read additionally at
                    # float64 via the columnar path; absent_value=np.nan matches the
                    # float32 path's NaN-where-absent semantics so the all-NaN drop
                    # and LAD NaN-guards stay consistent (a -9999 sentinel would
                    # survive those filters and poison the trajectory join).
                    fields_f64 = {f: lidar.getHitDataColumnArray(f, absent_value=np.nan)
                                  for f in fields_to_read if f in _LIDAR_FLOAT64_HIT_FIELDS}
                else:
                    all_xyz = np.empty((0, 3), np.float32)
                    all_rgb = np.empty((0, 3), np.float32)
                    all_scan_ids = np.empty((0,), np.int32)
                    all_miss = np.empty((0,), np.int32)
                    all_fields = {f: np.empty((0,), np.float32) for f in fields_to_read}
                    fields_f64 = {f: np.empty((0,), np.float64)
                                  for f in fields_to_read if f in _LIDAR_FLOAT64_HIT_FIELDS}

                # intensity is a signed beam·normal dot product; surface its
                # magnitude so it reads as a 0..1-ish value (matches the old loop).
                # abs in place — a fresh array here is a needless full-size (N f32)
                # allocation on a multi-million-hit cloud.
                if "intensity" in all_fields:
                    np.abs(all_fields["intensity"], out=all_fields["intensity"])

                # Partition per scanner by Helios scanID (== request index). Hits
                # with an out-of-range scanID match no mask and are dropped, exactly
                # as the old `sid < 0 or sid >= len` guard did.
                #
                # Boolean-mask indexing COPIES, so each scanner's slice duplicates
                # its share of the data; summed over scanners that reproduces the
                # whole hitlist a second time while the `all_*` originals are still
                # live. To cap peak RAM we (a) take a no-copy view when there is a
                # single scanner whose hits are the entire cloud, and (b) `del` the
                # `all_*` arrays the moment the partition is built so only one full
                # copy survives into the session-building phase below.
                results = []
                single_scanner = len(request.scanners) == 1
                for sid, s in enumerate(request.scanners):
                    _ckpt()
                    if single_scanner:
                        sel = slice(None)  # the whole cloud belongs to this scanner
                    else:
                        sel = (all_scan_ids == sid)
                    results.append({
                        "scanner_id": s.id,
                        "origin": [float(s.origin[0]), float(s.origin[1]), float(s.origin[2])],
                        "points": all_xyz[sel],                       # (m,3) f32
                        "colors": all_rgb[sel],                       # (m,3) f32
                        "scalars": {f: arr[sel] for f, arr in all_fields.items()},
                        # Float64 copies of the precision-sensitive fields, sliced
                        # with the SAME sel as scalars so they stay aligned. Routed
                        # to CloudSession.timestamps (not the float32 frame).
                        "timestamps_f64": {f: arr[sel] for f, arr in fields_f64.items()},
                        # is_miss stays compact (uint8); the session re-casts as it
                        # needs. A float32 flag is 4× the bytes for a 0/1 value.
                        "is_miss": all_miss[sel].astype(np.uint8),    # (m,) u8, 1==miss
                    })
                # Release the full-cloud staging arrays now that every scanner owns
                # its slice. Without this they stay referenced until the function
                # returns, so they coexist with the f64 session positions + the
                # laspy records built below — the peak that drove RAM to tens of GB.
                del all_xyz, all_rgb, all_scan_ids, all_miss, all_fields, fields_f64

        _prof["extract"] = time.perf_counter()
        _report(0.95, "Building point clouds")
        out = []
        # Pop from the front so each scanner's raw `r` (hits+misses, all fields)
        # is released as soon as its render arrays + session are built — otherwise
        # every scanner's full staging dict stays live alongside the sessions and
        # laspy records, re-stacking the peak we just trimmed above.
        while results:
            _ckpt()
            r = results.pop(0)
            # Route EVERY synthetic scan through a cloud session (octree-backed,
            # exactly like an imported cloud) so its in-RAM points are the
            # backend's source of truth for triangulation / LAD / edits. A
            # session_id keeps those requests tiny; the previous misses-off path
            # left the scan a FLAT in-RAM cloud, so Helios triangulation / LAD had
            # to serialise every point as an uncapped JSON `points` body — which
            # overflows the JS string limit ("Invalid string length") and OOMs the
            # pydantic parse on a multi-million-point scan, and rendered a large
            # flat cloud past V8's heap limit. The session is built from the FULL
            # `r` (hits + misses): its octree is hits-only, and any misses live in
            # the session's is_miss extra dim for the overlay + LAD.
            is_miss = r["is_miss"]  # (m,) u8, 1==miss
            has_misses = bool(record_misses and is_miss.size and np.any(is_miss != 0))
            session = None
            if r["points"].shape[0] > 0:
                try:
                    session = _create_lidar_scan_session(r, request.retained_standard_fields)
                except Exception:
                    import traceback
                    traceback.print_exc()
                    # A session failure must not lose the scan — fall back to a
                    # plain in-memory cloud below (built from hits only).

            # Strip miss rows from the in-memory render arrays. syntheticScan()
            # places each miss ~1 km out along its beam (LIDAR_RAYTRACE_MISS_T =
            # 1001 m); leaving them in the primary point array blows the cloud's
            # bounding box to ~2 km, so the camera auto-fit (distance = 2 * maxDim)
            # parks the view kilometres from a sub-metre target and the user can't
            # zoom back in. They'd also be drawn twice (here + the MissOverlay).
            # The misses are preserved in the session above (its octree is already
            # hits-only), so the renderer only needs the hits.
            if has_misses:
                keep = (is_miss == 0)
                points = r["points"][keep]
                colors = r["colors"][keep]
                raw_scalars = {k: v[keep] for k, v in r["scalars"].items()}
            else:
                points = r["points"]
                colors = r["colors"]
                raw_scalars = r["scalars"]

            npts = int(points.shape[0])
            # Drop scalar fields that never resolved (all-NaN) so the renderer
            # doesn't offer a dead color-by option.
            scalars = {
                k: v for k, v in raw_scalars.items()
                if v.size and not np.all(np.isnan(v))
            }
            entry = {
                "scanner_id": r["scanner_id"],
                "points": points,
                "colors": colors if npts else None,
                "scalars": scalars,
                "num_points": npts,
            }
            if session is not None:
                entry["session"] = session
            out.append(entry)
            # `entry` now owns the arrays it needs (the has_misses branch copied via
            # `keep`; the no-miss branch shares them — both keep `r`'s arrays alive
            # only through `entry`). Drop the staging dict so its hit+miss arrays for
            # this scanner free before the next iteration builds the next session.
            del r, points, colors, raw_scalars, scalars, entry

        _prof["post"] = time.perf_counter()
        if _prof_on:
            _total = _prof["post"] - _prof["start"]
            _hits = int(sum(e["num_points"] for e in out))
            _seg = lambda a, b: _prof[b] - _prof[a]
            print(
                "[scan-profile] "
                f"total={_total:.2f}s  "
                f"mesh_load={_seg('start','mesh_load'):.2f}s  "
                f"add_scans={_seg('mesh_load','add_scans'):.2f}s  "
                f"raytrace={_seg('add_scans','raytrace'):.2f}s  "
                f"extract={_seg('raytrace','extract'):.2f}s  "
                f"post={_seg('extract','post'):.2f}s  "
                f"(hits_kept={_hits:,}, record_misses={record_misses})",
                flush=True,
            )

        return {"success": True, "results": out}

    except ScanCancelled:
        # Cancellation is not a failure — let it propagate so the streaming
        # wrapper emits the cancelled marker (the `with` blocks already unwound,
        # freeing C++/numpy memory).
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": f"Synthetic LiDAR scan failed: {str(e)}"}


def _create_lidar_scan_session(r: dict, retained_standard_fields: Optional[List[str]] = None) -> dict:
    """Build a CloudSession from one synthetic scanner's in-memory hits and return
    the create_cloud_session-style metadata (session_id + octree cache + miss
    summary). Every synthetic scan is routed through a session (not just
    miss-recording ones) so triangulation / LAD / edits read its points from the
    backend by session_id, exactly like an imported cloud — no uncapped inline
    `points` body. The miss overlay rides the same session when misses exist.

    `r` carries `points` (list of [x,y,z]), optional `colors`, the per-hit
    `scalars` dict, an `is_miss` list (1==sky), and the scanner `origin`. The
    octree is built hits-only (far-field misses would poison its bbox); misses
    live in the session's is_miss extra dim for the overlay + LAD.

    `retained_standard_fields`, when not None, restricts which STANDARD scalars
    (see _LIDAR_STANDARD_HIT_FIELDS) become octree extra-dims, so the misses-on
    color-by list matches the flat-cloud path. Non-standard scalars (requested
    extras) and the is_miss flag are always kept. None => keep all standards.
    """
    import time
    import tempfile
    from pathlib import Path as _PathLocal

    positions = np.asarray(r["points"], dtype=np.float64)
    n = int(positions.shape[0])

    # Colors: scan colors are 0..1 floats; LAS sessions store uint16 (0..65535).
    # `r["colors"]` is a numpy (m,3) array — guard with `is not None`/size, not
    # truthiness (a numpy array has no unambiguous bool).
    colors = None
    rc = r.get("colors")
    if rc is not None and len(rc):
        c = np.asarray(rc, dtype=np.float64)
        if c.shape == (n, 3):
            colors = np.clip(c * 65535.0, 0, 65535).astype(np.uint16)

    # Extra dims: every per-hit scalar field, plus the canonical is_miss flag.
    # `intensity` is a STANDARD LAS dimension (it maps onto the session's
    # intensity field below), so it must NOT also be added as an extra dim — that
    # would make laspy's record dtype carry two `intensity` fields and blow up.
    intensity = None
    extras: Dict[str, np.ndarray] = {}
    extra_dims_meta: List[dict] = []
    # When a retained-standards set is given, a standard scalar not in it is
    # pruned (so it doesn't reappear in color-by). Non-standard scalars (requested
    # extras like deviation/reflectance) are always kept — requesting one IS the
    # retention choice.
    retained_set = set(retained_standard_fields) if retained_standard_fields is not None else None
    for name, vals in r.get("scalars", {}).items():
        arr = np.asarray(vals, dtype=np.float32)
        if arr.shape[0] != n:
            continue
        if name == "intensity":
            # Session intensity is uint16 (0..65535 LAS scale); the scan surfaces
            # intensity as a 0..1 magnitude. Scale and clamp.
            intensity = np.clip(arr * 65535.0, 0, 65535).astype(np.uint16)
            continue
        if (retained_set is not None and name in _LIDAR_STANDARD_HIT_FIELDS
                and name not in retained_set):
            continue
        extras[name] = arr
        extra_dims_meta.append({"slug": name, "label": name})
    miss = np.asarray(r["is_miss"], dtype=np.float32)
    extras[_MISS_SLUG] = miss
    extra_dims_meta.append({"slug": _MISS_SLUG, "label": _MISS_LABEL})

    # Per-point timestamp at float64 — the moving-platform LAD trajectory join key.
    # Routed to the dedicated CloudSession.timestamps field (NOT the float32 extras
    # above) so GPS-magnitude timestamps keep sub-second precision. The float32
    # copy stays in extras for color-by/backfill display. None when timestamp was
    # not recorded or is all-NaN (e.g. a static scan with no per-pulse time).
    ts_f64 = r.get("timestamps_f64", {}).get("timestamp")
    timestamps = None
    if ts_f64 is not None:
        arr64 = np.asarray(ts_f64, dtype=np.float64)
        if arr64.shape[0] == n and not np.all(np.isnan(arr64)):
            timestamps = arr64

    session_id = uuid.uuid4().hex[:8]
    sess = CloudSession(
        session_id=session_id,
        source_path="<synthetic-scan>",
        ascii_format=None,
        column_plan=None,
        positions=positions,
        colors=colors,
        intensity=intensity,
        extras=extras,
        extra_dims_meta=extra_dims_meta,
        timestamps=timestamps,
        world_shift=None,  # synthetic scans are authored near the origin
        deleted=np.zeros(n, dtype=bool),
        deleted_history=[],
        octree_cache_id=None,
        created_at=time.time(),
        last_accessed=time.time(),
    )

    # Register the session FIRST — it (positions + is_miss) is what powers the
    # miss overlay and LAD, neither of which reads the octree. The Potree octree
    # is a best-effort extra: a synthetic cloud renders from its in-memory points,
    # not the octree, so if PotreeConverter is unavailable we still return a
    # usable session (just without octree metadata).
    _sweep_cloud_sessions()
    with _cloud_session_lock:
        _cloud_sessions[session_id] = sess

    miss_count = int(np.count_nonzero(miss != 0))
    out = {
        "session_id": session_id,
        "point_count": n,
        "has_misses": miss_count > 0,
        "miss_slug": _MISS_SLUG,
        "miss_count": miss_count,
        "scan_origin": r.get("origin"),
        "miss_octree_cache_id": None,
    }

    # Build the octree from a hits-only LAS (drop ~20 km misses so the bbox /
    # camera framing stay tight), mirroring create_cloud_session. Best-effort.
    try:
        with tempfile.TemporaryDirectory() as _tmp:
            hits_las = _PathLocal(_tmp) / "octree_hits.las"
            _session_to_las(sess, hits_las, exclude_misses=True)
            cache_key, cache_dir, meta = _build_octree_from_las(hits_las, extra_dims_meta)
        sess.octree_cache_id = cache_key
        out["cache_id"] = cache_key
        out["cache_dir"] = str(cache_dir)
        out.update(meta)
    except Exception:
        import traceback
        traceback.print_exc()

    # Build the SECOND (projected-miss) octree so the synthetic scan's misses
    # stream + toggle exactly like an imported scan's. The scan origin is the
    # scanner head; synthetic clouds carry no world shift, so it's already in the
    # session frame. Best-effort + only when misses exist (returns None otherwise).
    if miss_count > 0:
        scan_origin = r.get("origin")
        sess.miss_octree_origin = list(scan_origin) if scan_origin is not None else None
        sess.miss_octree_cache_id = _build_miss_octree(sess, sess.miss_octree_origin)
        out["miss_octree_cache_id"] = sess.miss_octree_cache_id

    return out


# Per-hit scalar columns the RENDERER never displays — only the leaf-area
# inversion needs them, and it reads them from the session (or re-derives them
# from the trajectory), not from this frame. A moving scan attaches all of these,
# which on a multi-million-point cloud add tens of MB to the frame the renderer
# must transfer + decode + buffer (the "scan hangs" symptom). Excluded from the
# frame; still kept in the session for LAD / export / the miss overlay.
_LIDAR_FRAME_EXCLUDE_FIELDS = frozenset({"origin_x", "origin_y", "origin_z", "pulse_id"})


def _pack_lidar_scan(result: dict) -> bytes:
    """Pack a _do_lidar_scan result into a PHB1 frame: per-scanner points / colors
    / scalar fields become buffers (named s{i}.points, s{i}.colors, s{i}.scalar{j}),
    and meta carries the per-scanner descriptors (id, count, has_colors, the
    ordered scalar field names). Renderer-irrelevant per-beam columns
    (_LIDAR_FRAME_EXCLUDE_FIELDS) are dropped from the frame — they stay in the
    session for LAD/export."""
    if not result.get("success"):
        return _bin_frame_bytes({"success": False, "error": result.get("error")}, [])
    scanners_meta = []
    buffers = []
    for i, r in enumerate(result["results"]):
        npts = int(r["num_points"])
        fields = [f for f in r["scalars"].keys() if f not in _LIDAR_FRAME_EXCLUDE_FIELDS]
        has_colors = r["colors"] is not None and npts > 0
        entry_meta = {
            "scanner_id": r["scanner_id"], "num_points": npts,
            "has_colors": has_colors, "scalar_fields": fields,
        }
        # Forward the cloud session's metadata (created for every scan with hits)
        # so the renderer wires sessionId/hasMisses onto the cloud's octree.
        if r.get("session"):
            entry_meta["session"] = r["session"]
        scanners_meta.append(entry_meta)
        pts = np.asarray(r["points"], dtype=np.float32).reshape(-1) if npts else np.empty(0, np.float32)
        buffers.append((f"s{i}.points", pts, "f32"))
        if has_colors:
            buffers.append((f"s{i}.colors", np.asarray(r["colors"], dtype=np.float32).reshape(-1), "f32"))
        for j, f in enumerate(fields):
            buffers.append((f"s{i}.scalar{j}", np.asarray(r["scalars"][f], dtype=np.float32).reshape(-1), "f32"))
    return _bin_frame_bytes({"success": True, "scanners": scanners_meta}, buffers)


@app.post("/api/lidar/scan")
async def lidar_scan(request: LidarScanRequest, http_request: Request):
    """Ray-traced synthetic LiDAR scan. Returns a PHB1 binary frame (points +
    scalars per scanner can be millions of values)."""
    run_id, cancel_event = _new_cancel_token()
    return _bin_frame_streaming_response(
        lambda progress: _pack_lidar_scan(_do_lidar_scan(request, progress=progress)),
        request=http_request, cancel_event=cancel_event, run_id=run_id)


@app.post("/api/cancel/{run_id}")
async def cancel_run(run_id: str):
    """Cancel an in-flight streaming op (synthetic scan / triangulation / LAD).

    The streaming endpoints emit their run_id as the first PHP1 marker; the
    renderer POSTs it here to stop the work and free the C++/numpy memory without
    waiting on the (possibly huge) computation to finish. Idempotent: an unknown
    or already-finished run_id returns found=False rather than an error."""
    found = _cancel_run(run_id)
    return {"cancelled": found, "run_id": run_id}


# ==================== TREE SKELETON EXTRACTION (BFS Graph-Based Algorithm) ====================
# Based on: Li et al. 2017 "An Automatic Tree Skeleton Extracting Method Based on Point Cloud"

class SkeletonRequest(BaseModel):
    """Request model for tree skeleton extraction using BFS graph-based algorithm"""
    # Inline points for flat clouds. Optional because octree-backed clouds send
    # `source` instead (their renderer positions buffer is empty).
    points: Optional[List[List[float]]] = None
    # Read points from a file on disk (octree-backed clouds). Mutually exclusive
    # with `points`; renderer sets max_points=20000 here so the backend does the
    # downsample the renderer used to do in JS.
    source: Optional[PointSource] = None

    # Pre-processing options
    remove_outliers: bool = True  # Statistical outlier removal
    outlier_nb_neighbors: int = 20  # Neighbors for outlier detection
    outlier_std_ratio: float = 2.0  # Std deviation threshold for outliers

    # Graph building parameters
    search_radius: float = 0.05  # Radius for KD-tree neighbor search (S-R in paper)
    max_neighbors: int = 20  # Maximum neighbors per point

    # Root detection
    root_threshold: float = 0.02  # Height threshold τ for root set selection (meters)

    # Quantization parameters
    quantization_levels: int = 60  # Number of quantization intervals (Q-S in paper)
    use_nonlinear_quantization: bool = True  # Use sqrt scaling for better branch detail

    # Filtering parameters
    threshold_filter: int = 30  # Minimum points per block (T-F in paper)
    use_proportion_filter: bool = True  # Use parent/child ratio filter (P-F)
    proportion_threshold: float = 0.1  # Min ratio of child to parent block size

    # Smoothing
    smooth_skeleton: bool = True  # Apply Laplace smoothing
    smoothing_iterations: int = 2  # Number of smoothing passes

    # Legacy parameters (for API compatibility)
    dominant_axis: str = "z"  # Not used in BFS algorithm but kept for API compatibility


class SkeletonBlock(BaseModel):
    """Information about a skeleton block (cluster of points at same BFS level)"""
    block_id: int
    center: List[float]  # [x, y, z] centroid (skeleton node)
    quantized_level: int  # BFS level after quantization
    num_points: int  # Number of points in block
    parent_block_id: Optional[int] = None  # Parent block ID
    child_block_ids: List[int] = []  # Child block IDs


class SkeletonEdge(BaseModel):
    """Edge connecting two skeleton nodes"""
    from_block: int
    to_block: int
    length: float


class SkeletonResponse(BaseModel):
    """Response model for BFS-based skeleton extraction"""
    success: bool
    # Skeleton data - ordered from root to tips
    skeleton_points: List[List[float]]  # [[x, y, z], ...] skeleton nodes
    skeleton_edges: Optional[List[List[int]]] = None  # [[from_idx, to_idx], ...] connections
    branch_orders: Optional[List[int]] = None  # Branch order (Strahler number) for each node
    # Metrics
    total_length: Optional[float] = None  # Total skeleton length
    num_nodes: int  # Number of skeleton nodes
    num_edges: int = 0  # Number of skeleton edges
    num_branches: int = 0  # Number of branch points (nodes with >2 connections)
    max_branch_order: int = 0  # Maximum branch order in the skeleton
    # Block info (optional detailed output)
    blocks: Optional[List[SkeletonBlock]] = None
    # Processing info
    points_before_filtering: Optional[int] = None
    points_after_filtering: Optional[int] = None
    num_blocks_before_filter: Optional[int] = None
    num_blocks_after_filter: Optional[int] = None
    # Legacy fields for API compatibility
    dominant_axis: str = "z"
    slice_thickness: float = 0.0
    num_slices: int = 0  # Alias for num_nodes
    diameters: Optional[List[float]] = None
    error: Optional[str] = None


# ==================== GROUND SEGMENTATION HELPER (Cloth Simulation Filter) ====================

# Ground/non-ground class labels. Matches the convention of the Helios
# `object_label` column in our validation scans (1=ground, 2=plant), so a
# segmented cloud can be compared directly against ground-truth annotations.
GROUND_CLASS_GROUND = 1
GROUND_CLASS_PLANT = 2


def segment_ground(
    points: np.ndarray,
    cloth_resolution: float = 0.05,
    rigidness: int = 3,
    class_threshold: float = 0.02,
    iterations: int = 500,
    slope_smooth: bool = False,
    time_step: float = 0.65,
) -> np.ndarray:
    """Classify each point as ground (1) or plant (2) via the Cloth Simulation
    Filter (Zhang et al. 2016).

    CSF drapes an inverted cloth over the point cloud and labels points the
    cloth settles onto as ground. Defaults here are tuned for close-range plant
    scans on roughly-flat ground (cm-scale cloth resolution, high rigidness),
    NOT the airborne-LiDAR defaults the upstream docs assume.

    Returns an int array of length len(points), aligned to input order, with
    values GROUND_CLASS_GROUND / GROUND_CLASS_PLANT.

    Raises ImportError if the `CSF` extension is unavailable — the caller turns
    that into a clean error response rather than a 500.
    """
    import CSF  # SWIG C-extension; import here so a missing dep is catchable
    import os
    import tempfile

    n = len(points)
    csf = CSF.CSF()
    csf.params.bSloopSmooth = bool(slope_smooth)
    csf.params.cloth_resolution = float(cloth_resolution)
    csf.params.rigidness = int(rigidness)
    csf.params.class_threshold = float(class_threshold)
    csf.params.time_step = float(time_step)
    # NOTE: the CSF API misspells "iterations" as "interations".
    csf.params.interations = int(iterations)

    # CSF wants a contiguous float64 Nx3 array.
    csf.setPointCloud(np.ascontiguousarray(points[:, :3], dtype=np.float64))

    ground_idx, non_ground_idx = CSF.VecInt(), CSF.VecInt()
    # do_filtering() unconditionally dumps a debug `cloth_nodes.txt` into the
    # current working directory (hardcoded in the C++ lib). Run it from a temp
    # dir so the packaged app doesn't litter the user's filesystem, then drop
    # the artifact.
    prev_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmp:
        try:
            os.chdir(tmp)
            csf.do_filtering(ground_idx, non_ground_idx)
        finally:
            os.chdir(prev_cwd)

    labels = np.full(n, GROUND_CLASS_PLANT, dtype=np.int32)
    gi = np.fromiter(ground_idx, dtype=np.int64, count=len(ground_idx))
    if gi.size:
        labels[gi] = GROUND_CLASS_GROUND
    return labels


# ==================== WOOD/LEAF SEGMENTATION HELPER (geometric, non-ML) ====================

# Wood/leaf class labels. Mirrors the ground convention (small ints, comparable
# to a ground-truth annotation column): 1=wood (trunk/branches), 2=leaf.
WOOD_CLASS_WOOD = 1
WOOD_CLASS_LEAF = 2

WOOD_CLASS_SLUG = "wood_class"
WOOD_CLASS_LABEL = "Wood Class"

# Above this many points, segment_wood auto-downsamples (voxel) the geometry
# step and propagates labels back to full resolution — the per-point k-NN
# feature extraction is O(N·k_max) in memory and a multi-million-point cloud at
# full res can OOM the machine. ~1.5M keeps peak RAM ~3-4 GB. Env-overridable
# for big-memory machines; the result is always full-length regardless.
try:
    _WOOD_SEGMENT_MAX_POINTS = int(os.environ.get("PHYTOGRAPH_WOOD_MAX_POINTS", "1500000"))
except (ValueError, TypeError):
    _WOOD_SEGMENT_MAX_POINTS = 1_500_000


def _wood_local_pca_features(
    points: np.ndarray,
    k_min: int,
    k_max: int,
    k_step: int,
    chunk: int = 50_000,
) -> tuple[np.ndarray, np.ndarray]:
    """Per-point geometric features at an eigen-entropy-optimal neighbourhood
    scale (Demantke 2011 / Weinmann 2015 dimensionality).

    Returns (features, nbr_idx):
      features: (N,3) = [linearity, verticality, sphericity]
      nbr_idx:  (N, k_max) int32 sorted neighbour indices (reused by the
                regularisation pass — query the KD-tree only once).

    For each point we pick the k in [k_min..k_max] (step k_step) that minimises
    the eigen-entropy E = -Σ αᵢ ln αᵢ, αᵢ = λᵢ/Σλ, then derive features from the
    eigenvalues/eigenvectors at that winning scale. The wood/leaf discriminators
    are verticality (wood upright) and sphericity (foliage scatters in 3D);
    linearity is returned for completeness but is not used by segment_wood
    (branches are linear too). See segment_wood for the scoring rationale.
    """
    from scipy.spatial import cKDTree

    n = len(points)
    pts = np.ascontiguousarray(points[:, :3], dtype=np.float64)
    # k can't exceed the cloud size; the self-point is included at column 0.
    k_max_eff = int(min(k_max, n))
    k_candidates = [k for k in range(k_min, k_max_eff + 1, k_step)]
    if not k_candidates or k_candidates[-1] != k_max_eff:
        # always include the largest usable scale
        k_candidates.append(k_max_eff)
    k_candidates = sorted({k for k in k_candidates if k >= 3})
    if not k_candidates:
        k_candidates = [min(3, n)]

    tree = cKDTree(pts)
    # Query at the largest scale; smaller scales are prefixes of the sorted
    # neighbour list (cKDTree returns neighbours in increasing distance). Query
    # in CHUNKS and keep only int32 indices: a single full query allocates the
    # (N,k) float64 distances AND int64 indices at once (~10 GB at N=6M, k=100),
    # which OOMs a 16 GB machine. Chunking caps that transient to one block, and
    # we discard distances immediately.
    nbr_idx = np.empty((n, k_max_eff), dtype=np.int32)
    q_chunk = max(1, min(chunk, n))
    for start in range(0, n, q_chunk):
        end = min(start + q_chunk, n)
        _, idx_block = tree.query(pts[start:end], k=k_max_eff, workers=-1)
        if idx_block.ndim == 1:  # k==1 edge case
            idx_block = idx_block[:, None]
        nbr_idx[start:end] = idx_block.astype(np.int32, copy=False)
        del idx_block

    best_E = np.full(n, np.inf, dtype=np.float64)
    best_lin = np.zeros(n, dtype=np.float64)
    best_sph = np.zeros(n, dtype=np.float64)
    best_vert = np.zeros(n, dtype=np.float64)

    eps = 1e-12
    for start in range(0, n, chunk):
        end = min(start + chunk, n)
        idx_chunk = nbr_idx[start:end]  # (m, k_max_eff)
        m = end - start
        c_best_E = np.full(m, np.inf)
        c_lin = np.zeros(m)
        c_sph = np.zeros(m)
        c_vert = np.zeros(m)
        for k in k_candidates:
            nb = pts[idx_chunk[:, :k]]               # (m, k, 3)
            mu = nb.mean(axis=1, keepdims=True)
            d = nb - mu
            cov = np.einsum("mki,mkj->mij", d, d) / float(k - 1)  # (m,3,3)
            # eigh: ascending eigenvalues w0<=w1<=w2, columns are eigenvectors.
            w, v = np.linalg.eigh(cov)
            lam = np.clip(w[:, ::-1], eps, None)     # λ1>=λ2>=λ3
            s = lam.sum(axis=1)
            a = lam / s[:, None]
            E = -(a * np.log(a)).sum(axis=1)
            better = E < c_best_E
            if better.any():
                l1 = lam[:, 0]
                l2 = lam[:, 1]
                l3 = lam[:, 2]
                lin = (l1 - l2) / l1
                sph = l3 / l1
                # normal = eigenvector of smallest eigenvalue = column 0 of v.
                nz = np.abs(v[:, 2, 0])
                vert = 1.0 - nz
                c_best_E = np.where(better, E, c_best_E)
                c_lin = np.where(better, lin, c_lin)
                c_sph = np.where(better, sph, c_sph)
                c_vert = np.where(better, vert, c_vert)
        best_E[start:end] = c_best_E
        best_lin[start:end] = c_lin
        best_sph[start:end] = c_sph
        best_vert[start:end] = c_vert

    features = np.column_stack([best_lin, best_vert, best_sph]).astype(np.float64)
    features = np.nan_to_num(features, nan=0.0, posinf=0.0, neginf=0.0)
    return features, nbr_idx


def _wood_regularize(
    labels: np.ndarray,
    nbr_idx: np.ndarray,
    reg_k: int,
    iters: int,
    weights: Optional[np.ndarray] = None,
) -> np.ndarray:
    """LeWoS-style label smoothing: iterated k-NN majority vote over the graph
    already built for feature extraction. Cleans isolated misclassifications
    (a stray "wood" point inside a leaf cluster flips to leaf, and vice versa)
    without eroding genuine thin branches at low iteration counts.

    `weights` (optional, per-point in [0,1]) up-weights confident neighbours so
    high-certainty points dominate the vote near wood/leaf boundaries.
    """
    if iters <= 0:
        return labels.astype(np.int32, copy=True)
    k = int(min(reg_k, nbr_idx.shape[1]))
    if k < 2:
        return labels.astype(np.int32, copy=True)
    idx = nbr_idx[:, :k]
    w_is_wood = (labels == WOOD_CLASS_WOOD).astype(np.float64)
    nbr_w = weights[idx] if weights is not None else None
    denom = nbr_w.sum(axis=1) if nbr_w is not None else float(k)
    denom = np.where(np.asarray(denom) == 0, 1.0, denom)
    for _ in range(int(iters)):
        gathered = w_is_wood[idx]                      # (N, k)
        if nbr_w is not None:
            frac = (gathered * nbr_w).sum(axis=1) / denom
        else:
            frac = gathered.mean(axis=1)
        w_is_wood = (frac >= 0.5).astype(np.float64)
    out = np.where(w_is_wood >= 0.5, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF)
    return out.astype(np.int32)


def _wood_grow_branches(
    wood_seed: np.ndarray,
    sphericity: np.ndarray,
    nbr_idx: np.ndarray,
    sph_thresh: float,
    grow_k: int = 10,
    max_iters: int = 30,
) -> np.ndarray:
    """Region-grow the wood label along CONNECTED woody geometry.

    The verticality-weighted seed has high precision but misses HORIZONTAL
    scaffold branches: they're geometrically woody (compact, low sphericity) but
    not vertical, so the score drops them. They ARE connected to the trunk,
    though — so flood the wood label outward from the seed into any point that
    (a) has low sphericity (`sphericity < sph_thresh`, i.e. locally compact like
    a branch, not a scattered leaf) AND (b) touches a current wood point via its
    k-NN. Iterated to BFS along the branch network. Leaf clusters that are
    coincidentally low-sphericity (e.g. synthetic narrow leaves) are mostly NOT
    connected to the trunk, so they don't get swept in.

    Returns a boolean wood mask (seed ∪ grown).
    """
    if sph_thresh <= 0:
        return wood_seed.copy()
    woody = sphericity < float(sph_thresh)
    k = int(min(grow_k, nbr_idx.shape[1]))
    if k < 2:
        return wood_seed.copy()
    idx = nbr_idx[:, :k]
    wood = wood_seed.copy()
    prev = int(wood.sum())
    for _ in range(int(max_iters)):
        # A woody candidate flips to wood if any of its k-NN is already wood.
        wood = wood | (woody & wood[idx].any(axis=1))
        cur = int(wood.sum())
        if cur == prev:
            break
        prev = cur
    return wood


def _wood_prune_speckle(
    points: np.ndarray,
    wood_mask: np.ndarray,
    nbr_idx: np.ndarray,
    k: int = 8,
    min_size: int = 10,
) -> np.ndarray:
    """Drop tiny isolated wood connected-components (speckle): build a graph
    among candidate-wood points via their mutual k-NN, then flip components
    smaller than `min_size` back to leaf. Wood is a connected skeleton, so real
    branches survive; scattered single-point false-wood in foliage does not.
    Conservative by design (small `min_size`) so it never erodes thin branches.
    """
    from scipy.sparse import csr_matrix
    from scipy.sparse.csgraph import connected_components

    wi = np.where(wood_mask)[0]
    if len(wi) < max(3, min_size):
        return wood_mask
    kk = int(min(k, nbr_idx.shape[1]))
    sub_nbr = nbr_idx[wi, :kk]
    wood_set = np.zeros(len(points), dtype=bool)
    wood_set[wi] = True
    remap = -np.ones(len(points), dtype=np.int64)
    remap[wi] = np.arange(len(wi))
    keep_edge = wood_set[sub_nbr]                       # (m, kk) neighbour is wood
    rows = np.repeat(np.arange(len(wi)), kk)[keep_edge.ravel()]
    cols = remap[sub_nbr.ravel()[keep_edge.ravel()]]
    if rows.size == 0:
        return wood_mask
    adj = csr_matrix((np.ones(rows.size), (rows, cols)), shape=(len(wi), len(wi)))
    adj = adj + adj.T
    _, comp = connected_components(adj, directed=False)
    counts = np.bincount(comp)
    keep_local = counts[comp] >= min_size
    out = wood_mask.copy()
    out[wi] = keep_local
    return out


def _wood_reflectance_wood_seed(
    reflectance: np.ndarray,
    weight_max: float,
    tail_pctile_at_max: float = 88.0,
) -> np.ndarray:
    """Boolean mask of points to PROMOTE to wood from a one-sided high-reflectance
    cut. Returns all-False when reflectance is unusable.

    Empirically (Weiser oak/beech, Riegl 1550 nm), P(wood | reflectance) is
    flat-low across the bottom ~70-80 % of the range and only rises steeply in
    the TOP couple of deciles (to ~0.9-0.99) — the mid/low range overlaps too
    much to separate (Beland 2014 Fig 8). So the only safe radiometric move is to
    promote the very brightest returns the geometry missed. `weight_max` ∈ (0,1]
    sets how deep into the upper tail we cut: at the max (1.0) we take everything
    above `tail_pctile_at_max` (default 88th percentile), scaling toward the 99th
    percentile as weight_max → 0. Confining promotion to a thin top tail keeps it
    from flooding wood on low-contrast species (where that tail isn't specifically
    wood, but is small). Promotion only ADDS wood — it never demotes — so the
    assist can only recover missed wood, and the user toggle disables it entirely
    for species known to lack contrast.
    """
    finite = np.isfinite(reflectance)
    if finite.sum() < 50 or weight_max <= 0:
        return np.zeros(reflectance.shape[0], dtype=bool)

    rf = reflectance[finite].astype(np.float64)
    w = float(min(weight_max, 1.0))
    # weight_max → percentile: w=1 ⇒ tail_pctile_at_max; w→0 ⇒ ~99th.
    pctile = 99.0 - (99.0 - tail_pctile_at_max) * w
    thresh = float(np.percentile(rf, pctile))
    # Promote STRICTLY above the threshold. A constant / contrast-free reflectance
    # has thresh == max, so `> thresh` selects nothing — the assist is inert and
    # the result falls back to pure geometry (the harmless-on-low-contrast
    # guarantee). Using `>=` here would promote the entire flat distribution.
    if not np.isfinite(thresh) or thresh >= float(rf.max()):
        return np.zeros(reflectance.shape[0], dtype=bool)

    mask = np.zeros(reflectance.shape[0], dtype=bool)
    mask[finite] = rf > thresh
    return mask


def _wood_geometric_labels(
    pts: np.ndarray,
    refl: Optional[np.ndarray],
    *,
    k_min: int, k_max: int, k_step: int, wood_bias: float,
    reflectance_weight_max: float,
    precomputed: Optional[tuple] = None,
):
    """Shared geometric classification core: returns
    (raw_labels, seed_mask, sphericity, nbr_idx) on the given (already-downsampled)
    `pts`. `raw_labels` is the per-point GMM split + reflectance promotion (BEFORE
    branch-grow/speckle/regularise). `seed_mask` is the HIGH-PRECISION wood seed
    (strict GMM posterior) the connectivity method anchors its backbone on — the
    trunk and obvious branches, kept tight to avoid seeding the backbone from
    false-wood. Both the geometric and connectivity methods call this so they can
    never drift apart. No region-grow / speckle / regularise here (the callers add
    those). `precomputed=(features, nbr_idx)` reuses an existing feature pass (the
    SOTA path computes features once and shares them — the pass is ~25s on 480k)."""
    from sklearn.mixture import GaussianMixture

    n = len(pts)
    if precomputed is not None:
        features, nbr_idx = precomputed
    else:
        features, nbr_idx = _wood_local_pca_features(pts, k_min, k_max, k_step)
    verticality = features[:, 1]
    sphericity = features[:, 2]

    # Wood saliency = verticality + (1 - sphericity); see segment_wood docstring.
    sph_n = np.clip(sphericity / (np.quantile(sphericity, 0.99) + 1e-9), 0.0, 1.0)
    score = 0.5 * verticality + 0.5 * (1.0 - sph_n)

    raw = None
    seed = np.zeros(n, dtype=bool)
    if n >= 50:
        try:
            gmm = GaussianMixture(n_components=2, n_init=3, reg_covar=1e-6, random_state=0)
            s = score.reshape(-1, 1)
            gmm.fit(s)
            comp = gmm.predict(s)
            means = [score[comp == c].mean() if np.any(comp == c) else -np.inf
                     for c in (0, 1)]
            wood_comp = int(np.argmax(means))
            spread = float(np.std(score)) + 1e-9
            separation = abs(means[0] - means[1]) / spread
            if np.isfinite(separation) and separation >= 0.25:
                proba = gmm.predict_proba(s)[:, wood_comp]
                raw = np.where(proba >= wood_bias, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF)
                # High-precision seed: a STRICTER posterior than wood_bias, so the
                # connectivity backbone is anchored only on near-certain wood.
                seed = proba >= max(float(wood_bias), 0.85)
        except Exception:
            raw = None

    if raw is None:
        if float(np.mean(score)) >= 0.5:
            raw = np.full(n, WOOD_CLASS_WOOD, dtype=np.int32)
        else:
            raw = np.full(n, WOOD_CLASS_LEAF, dtype=np.int32)
    raw = raw.astype(np.int32)

    # Reflectance assist (one-sided high-reflectance promotion) — see
    # _wood_reflectance_wood_seed. Promoted points also strengthen the seed.
    if refl is not None and reflectance_weight_max and reflectance_weight_max > 0:
        promote = _wood_reflectance_wood_seed(refl, float(reflectance_weight_max))
        raw = np.where(promote, WOOD_CLASS_WOOD, raw).astype(np.int32)
        seed = seed | promote

    return raw, seed, sphericity, nbr_idx


def _wood_geometric_core(
    pts: np.ndarray,
    refl: Optional[np.ndarray],
    *,
    k_min: int, k_max: int, k_step: int, wood_bias: float,
    reg_k: int, reg_iters: int, min_speckle: int, branch_grow_sph: float,
    reflectance_weight_max: float,
    precomputed: Optional[tuple] = None,
) -> np.ndarray:
    """The original point-wise wood/leaf pipeline on a single (downsampled) cloud:
    geometric GMM + reflectance → branch-grow → speckle prune → regularise. Returns
    down-resolution int32 labels (byte-identical to the pre-refactor inline code).
    `precomputed=(features, nbr_idx)` shares an existing feature pass (SOTA path)."""
    raw, _seed, sphericity, nbr_idx = _wood_geometric_labels(
        pts, refl, k_min=k_min, k_max=k_max, k_step=k_step, wood_bias=wood_bias,
        reflectance_weight_max=reflectance_weight_max, precomputed=precomputed,
    )

    # Branch recovery: grow wood into connected low-sphericity points (recovers
    # horizontal scaffold branches the verticality score misses).
    if branch_grow_sph and branch_grow_sph > 0:
        grown = _wood_grow_branches(
            raw == WOOD_CLASS_WOOD, sphericity, nbr_idx, float(branch_grow_sph),
        )
        raw = np.where(grown, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF).astype(np.int32)

    if min_speckle and min_speckle > 1:
        wood_mask = _wood_prune_speckle(pts, raw == WOOD_CLASS_WOOD, nbr_idx,
                                        min_size=int(min_speckle))
        raw = np.where(wood_mask, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF).astype(np.int32)

    return _wood_regularize(raw, nbr_idx, reg_k, reg_iters)


def _wood_subtree_support(graph) -> np.ndarray:
    """Per-node count of cloud points hanging at or below the node in the rooted
    skeleton tree (the node's own point_count plus all descendants'). This is the
    path-funnel / betweenness proxy: the trunk funnels ALL descendants (== N at the
    root), a major fork funnels the crown fraction it feeds, a terminal leaf-cluster
    tip funnels only its own handful. Computed by a single reverse-`level` sweep:
    children always have a strictly higher level than their parent (skeleton is
    acyclic by construction), so processing high→low level and adding each node's
    accumulated total into its parent is correct and O(K)."""
    parent = np.asarray(graph.parent)
    level = np.asarray(graph.level)
    support = np.asarray(graph.point_count, dtype=np.int64).copy()
    # Process nodes from deepest level to shallowest; push each node's running
    # total up to its parent.
    order = np.argsort(-level, kind="stable")
    for i in order:
        p = int(parent[i])
        if p >= 0:
            support[p] += support[i]
    return support


def _wood_backbone_nodes(graph, support: np.ndarray, seed_node_mask: np.ndarray,
                         min_support: int) -> np.ndarray:
    """Boolean (K,) mask of skeleton nodes that are the woody BACKBONE: the union of
    (a) nodes whose subtree support ≥ `min_support` (they funnel enough descendants
    to be a trunk/branch, not a leaf tip), and (b) the ANCESTOR CLOSURE of the
    geometric seed nodes — every node on the path from a confirmed-wood node back to
    the root. (b) is the connectivity move: it guarantees a continuous woody chain
    from each seed to the base, recovering thin branches the geometry missed, and it
    bounds (a) (a high-support node not on any seed path won't drag in leaf)."""
    K = len(graph)
    parent = np.asarray(graph.parent)
    backbone = support >= int(min_support)
    # Ancestor closure of seeds: walk parent links to root, marking each node.
    seen = np.zeros(K, dtype=bool)
    for start in np.where(seed_node_mask)[0]:
        cur = int(start)
        while cur >= 0 and not seen[cur]:
            seen[cur] = True
            backbone[cur] = True
            cur = int(parent[cur])
    return backbone


def _segment_wood_connectivity(
    pts: np.ndarray,
    refl: Optional[np.ndarray],
    *,
    k_min: int, k_max: int, k_step: int, wood_bias: float,
    reg_k: int, reg_iters: int, min_speckle: int, branch_grow_sph: float,
    reflectance_weight_max: float, backbone_support: float,
    warnings: list,
) -> np.ndarray:
    """Connectivity-based wood/leaf classification on a single (downsampled) cloud.

    Fuses the geometric classifier with a rooted geodesic skeleton (reusing
    `qsm.skeleton.extract_skeleton`): geometry gives a high-precision wood seed,
    connectivity recovers the woody backbone (points on continuous paths to the
    trunk base) and prunes disconnected geometric false-wood. Falls back to the
    geometric result whenever the skeleton can't give a usable backbone (degenerate
    skeleton, no reachable points, etc.). Appends a ground-not-removed notice to
    `warnings` when the base looks like residual ground. Returns down-resolution
    int32 labels."""
    from qsm.skeleton import extract_skeleton, SkeletonOptions

    # 1. Geometric pass (shared core) → raw labels + high-precision seed + features.
    raw, seed, sphericity, nbr_idx = _wood_geometric_labels(
        pts, refl, k_min=k_min, k_max=k_max, k_step=k_step, wood_bias=wood_bias,
        reflectance_weight_max=reflectance_weight_max,
    )
    geom_wood = raw == WOOD_CLASS_WOOD

    def _finish(wood_mask):
        """Apply the shared post-processing (speckle prune + regularise) and return
        labels. branch-grow is skipped for connectivity — the backbone already does
        the recovery branch-grow approximates, and re-growing would re-introduce the
        leaf-bleed connectivity is meant to avoid."""
        lab = np.where(wood_mask, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF).astype(np.int32)
        if min_speckle and min_speckle > 1:
            wm = _wood_prune_speckle(pts, lab == WOOD_CLASS_WOOD, nbr_idx, min_size=int(min_speckle))
            lab = np.where(wm, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF).astype(np.int32)
        return _wood_regularize(lab, nbr_idx, reg_k, reg_iters)

    # 2. Build the rooted geodesic skeleton. On any failure, fall back to the full
    # geometric pipeline (branch-grow included) — never worse than geometric.
    def _geometric_fallback():
        if branch_grow_sph and branch_grow_sph > 0:
            grown = _wood_grow_branches(geom_wood, sphericity, nbr_idx, float(branch_grow_sph))
            return _finish(grown)
        return _finish(geom_wood)

    try:
        graph = extract_skeleton(pts, SkeletonOptions())
    except Exception:
        return _geometric_fallback()

    nop = graph.node_of_point
    if (nop is None or len(graph) < 2 or graph.meta.get("degenerate")
            or graph.meta.get("fallback") or not np.any(nop >= 0)):
        return _geometric_fallback()

    # 3. Ground guard: the geodesic roots at the lowest points, so if the base
    # level-set node spans a broad flat slab (much wider than a trunk) the cloud
    # likely still contains ground and the whole tree will funnel through it. Warn
    # but proceed (per product decision). Heuristic: base node horizontal extent vs
    # the cloud's vertical extent.
    try:
        base_pts = pts[nop == int(graph.root)]
        if len(base_pts) >= 3:
            horiz = float(np.hypot(*(base_pts[:, :2].max(axis=0) - base_pts[:, :2].min(axis=0))))
            height = float(pts[:, 2].max() - pts[:, 2].min()) + 1e-9
            if horiz > 0.5 * height and horiz > 0.5:
                warnings.append(
                    "Connectivity segmentation: the tree base spans a wide flat area, "
                    "which usually means the ground wasn't removed. Crop the ground "
                    "first for a correct result."
                )
    except Exception:
        pass

    # 4. Build the woody BACKBONE = the ancestor closure of the high-precision
    # geometric seed nodes (every node on the path from a confirmed-wood point back
    # to the trunk base). This is the connectivity signal: a continuous chain from
    # each seed to the root. We deliberately do NOT add a blanket subtree-support
    # floor — that over-keeps badly on dense crowns (validated: a global floor reads
    # 60-90% wood on tropical/oak). Support is used only to PRUNE (step 5b).
    n = len(pts)
    assigned = nop >= 0
    support = _wood_subtree_support(graph)
    seed_nodes = np.zeros(len(graph), dtype=bool)
    seed_nodes[nop[(assigned) & seed]] = True
    backbone_nodes = _wood_backbone_nodes(graph, support, seed_nodes,
                                          min_support=10**18)  # support floor OFF
    point_on_backbone = np.zeros(n, dtype=bool)
    point_on_backbone[assigned] = backbone_nodes[nop[assigned]]

    # 5. Geometry-PRIMARY fusion with two conservative connectivity corrections
    # (mirrors TLSeparation/CWLS: geometry classifies, connectivity refines):
    wood = geom_wood.copy()

    # 5a. RECOVER thin branches the geometry dropped: a LEAF point that sits on the
    # backbone path (between confirmed wood and the root) AND is locally compact
    # (not a 3D-scattered leaf) is a missed twig → wood. The sphericity gate keeps
    # this from dragging in leaf clusters that merely hang off a branch node.
    sph_gate = sphericity < max(2.0 * float(branch_grow_sph or 0.02), 0.04)
    recover = (~geom_wood) & point_on_backbone & sph_gate
    wood |= recover

    # 5b. PRUNE topologically-isolated false-wood: a WOOD point whose node is NOT on
    # the backbone and has tiny subtree support (a compact leaf clump that fooled the
    # local geometry) → leaf. `backbone_support` (>0) sets the support floor as a
    # fraction of N; auto uses a small absolute floor. Conservative: only nodes well
    # below the floor AND off the backbone are pruned.
    if backbone_support and backbone_support > 0:
        prune_floor = max(2, int(backbone_support * n))
    else:
        prune_floor = max(3, int(0.0005 * n))  # ~25 pts at 50k
    node_tiny = support < prune_floor
    point_tiny = np.zeros(n, dtype=bool)
    point_tiny[assigned] = node_tiny[nop[assigned]]
    prune = geom_wood & assigned & ~point_on_backbone & point_tiny
    wood &= ~prune

    # Unreachable / pruned-spur points (nop == -1) keep their geometric label — a
    # second un-bridged trunk must not be forced to leaf.
    wood[~assigned] = geom_wood[~assigned]

    return _finish(wood)


# ==================== SOTA SEGMENT-WISE WOOD/LEAF (multi-scale + cylinder gate) ====================
# Literature-faithful unsupervised recipe (LeWoS/Wang 2019/Wan 2021/CWLS 2025),
# validated in a 5-iteration Phase-0 PoC. The current point-wise single-scale
# classifier is the documented WEAK baseline; this segments first and classifies
# whole branch segments by CYLINDER-FIT QUALITY (a real branch wraps a tight,
# well-covered cylinder; a leaf-laden region does not). It recovers thin crown
# branches the geometric core drops (the diagnosed failure) without flooding leaf.

def _wood_skeleton_segments(pts: np.ndarray):
    """Group points into BRANCH-SIZED segments via the geodesic skeleton.

    Returns (graph, pt_seg, segments) where `pt_seg[i]` is point i's segment id
    (-1 if unassigned), `segments` is the qsm `_Segment` list (runs of skeleton
    nodes between forks). Returns (None, None, None) if the skeleton degenerates.

    Why skeleton-segments and not curvature/connected-components: smooth wood is
    all mutually connected, so a connected-component pass merges the entire tree
    into ONE blob (Phase-0 PoC v2). The skeleton's fork-to-fork node runs are
    naturally branch-sized — the unit a cylinder fit is meaningful on.
    """
    from qsm.skeleton import extract_skeleton, SkeletonOptions
    from qsm.segments import build_segments

    graph = extract_skeleton(pts, SkeletonOptions())
    nop = graph.node_of_point
    if (nop is None or len(graph) < 2 or graph.meta.get("degenerate")
            or graph.meta.get("fallback") or not np.any(nop >= 0)):
        return None, None, None
    segments = build_segments(graph)
    if not segments:
        return None, None, None
    # node -> segment lookup (vectorised); points map through node_of_point.
    node_seg = np.full(len(graph), -1, dtype=np.int64)
    for s in segments:
        node_seg[np.asarray(s.node_ids, dtype=np.int64)] = s.seg_id
    pt_seg = np.full(len(pts), -1, dtype=np.int64)
    assigned = nop >= 0
    pt_seg[assigned] = node_seg[nop[assigned]]
    return graph, pt_seg, segments


def _wood_classify_segments(
    pts: np.ndarray, graph, pt_seg: np.ndarray, segments,
    geom_wood: np.ndarray, *,
    min_seg: int = 15, surf_cov_min: float = 0.3, mad_frac_max: float = 0.5,
    shell_frac: float = 0.5,
) -> np.ndarray:
    """Cylinder-fit segment gate + tube-shell recovery (Phase-0 PoC v5).

    Starts from the geometric wood labels and ADDS recovered wood: for each
    branch-sized segment, fit a cylinder (qsm `fit_cylinder`); if it is a TIGHT,
    WELL-COVERED tube (`surf_cov >= surf_cov_min` and `mad <= mad_frac_max*radius`)
    it is a real branch — recover its TUBE SHELL (points within `shell_frac*radius`
    of the fitted surface) as wood. Leaves splay off the shell and keep their
    geometric label. `surf_cov_min`/`mad_frac_max`/`shell_frac` are the
    operating-point dials (leaf-off recall vs leaf-on precision). Returns a wood
    boolean mask (regularisation is applied by the caller).
    """
    from qsm.cylinders import fit_cylinder, CylinderFitOptions

    wood = geom_wood.copy()
    copts = CylinderFitOptions()
    nodes = graph.nodes
    for s in segments:
        m = pt_seg == s.seg_id
        if int(m.sum()) < min_seg:
            continue
        p = pts[m]
        seg_nodes = nodes[np.asarray(s.node_ids, dtype=np.int64)]
        seed_start, seed_end = seg_nodes[0], seg_nodes[-1]
        axis = seed_end - seed_start
        axlen = float(np.linalg.norm(axis))
        if axlen < 1e-6:
            continue
        axis = axis / axlen
        c = p - p.mean(axis=0)
        r_perp = np.linalg.norm(c - np.outer(c @ axis, axis), axis=1)
        seed_r = float(np.median(r_perp)) or 0.02
        fit = fit_cylinder(p, seed_start, seed_end, seed_r, copts)
        if fit is None or not fit.reliable:
            continue
        if not (fit.surf_cov >= surf_cov_min and fit.mad <= mad_frac_max * fit.radius):
            continue  # not a clean tube → leaf-laden segment, reject
        shell = np.abs(r_perp - fit.radius) <= max(shell_frac * fit.radius, 0.01)
        idx = np.where(m)[0]
        wood[idx[shell]] = True
    return wood


def _segment_wood_sota(
    pts: np.ndarray,
    refl: Optional[np.ndarray],
    *,
    k_min: int, k_max: int, k_step: int, wood_bias: float,
    reg_k: int, reg_iters: int, min_speckle: int, branch_grow_sph: float,
    reflectance_weight_max: float, backbone_support: float,
    warnings: list,
) -> np.ndarray:
    """SOTA segment-wise wood/leaf classification on a single (downsampled) cloud.

    Pipeline: geometric core (seed) → skeleton segments → cylinder-fit gate +
    tube-shell recovery → regularise. Falls back to the geometric core whenever
    the skeleton degenerates. Returns down-resolution int32 labels."""
    # Compute the per-point PCA features ONCE and share them with the geometric
    # core (the feature pass is ~25s on 480k points; computing it twice was the
    # bulk of the SOTA path's runtime). `nbr_idx` is reused for the final regularise.
    features, nbr_idx = _wood_local_pca_features(pts, k_min, k_max, k_step)
    geom_labels = _wood_geometric_core(
        pts, refl, k_min=k_min, k_max=k_max, k_step=k_step, wood_bias=wood_bias,
        reg_k=reg_k, reg_iters=reg_iters, min_speckle=min_speckle,
        branch_grow_sph=branch_grow_sph, reflectance_weight_max=reflectance_weight_max,
        precomputed=(features, nbr_idx),
    )
    graph, pt_seg, segments = _wood_skeleton_segments(pts)
    if graph is None:
        return geom_labels  # degenerate skeleton → geometric fallback

    geom_wood = geom_labels == WOOD_CLASS_WOOD
    wood = _wood_classify_segments(pts, graph, pt_seg, segments, geom_wood)
    labels = np.where(wood, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF).astype(np.int32)
    if min_speckle and min_speckle > 1:
        wm = _wood_prune_speckle(pts, labels == WOOD_CLASS_WOOD, nbr_idx, min_size=int(min_speckle))
        labels = np.where(wm, WOOD_CLASS_WOOD, WOOD_CLASS_LEAF).astype(np.int32)
    return _wood_regularize(labels, nbr_idx, reg_k, reg_iters)


def segment_wood(
    points: np.ndarray,
    k_min: int = 10,
    k_max: int = 100,
    k_step: int = 10,
    wood_bias: float = 0.6,
    reg_k: int = 20,
    reg_iters: int = 3,
    min_speckle: int = 0,
    branch_grow_sph: float = 0.02,
    voxel_size: float = 0.0,
    max_points: Optional[int] = None,
    reflectance: Optional[np.ndarray] = None,
    reflectance_weight_max: float = 0.4,
    method: str = "geometric",
    backbone_support: float = 0.0,
    warnings: Optional[list] = None,
) -> np.ndarray:
    """Classify each point as wood (1) or leaf (2) from geometry (+ optional
    reflectance assist), via one of two `method`s.

    `method="geometric"` (the original) is purely point-wise. `method="connectivity"`
    additionally roots a geodesic skeleton at the trunk base and recovers the woody
    backbone — the set of points on continuous paths back to the base — which the
    point-wise method can't see, so it recovers thin branches/twigs and prunes
    geometrically-compact-but-disconnected leaf clumps. It REQUIRES ground removal
    (the geodesic roots at the lowest points); if the base looks like residual
    ground it appends a warning to `warnings` (a caller-supplied list) but proceeds,
    and it falls back to geometry when the skeleton degenerates. See
    `_segment_wood_connectivity`. `backbone_support` (0 = auto) tunes how much
    subtree support a node needs to count as backbone (higher → only major scaffold).

    Pipeline (classical / non-ML, runs on CPU in seconds-to-minutes):
      1. Per-point local-PCA features at an eigen-entropy-optimal scale
         (verticality, sphericity — Demantke 2011 / Weinmann 2015).
      2. A wood saliency score = verticality + (1 − sphericity): wood is
         vertical (trunk/branches) and locally COMPACT (low sphericity), while
         foliage scatters the neighbourhood in 3D (high sphericity) and hangs at
         varied non-vertical angles. OPTIONAL reflectance assist: when a
         per-point `reflectance` scalar is supplied (and `reflectance_weight_max`
         > 0), a 1-D GMM on the reflectance contributes a wood-probability term
         to this score, weighted by how separable wood/leaf are in it — so it
         helps high-contrast species (oak/beech: wood reads higher reflectance)
         and is inert (weight ≈ 0) on low-contrast ones (almond/redbud) or when
         `reflectance` is None. See `_wood_blend_reflectance`.
         A 2-component 1-D Gaussian Mixture splits
         the score; the higher-mean component is wood and `wood_bias` is its
         posterior threshold (0.5 = argmax; higher → stricter / less wood; lower
         → more wood).
      3. Branch recovery (`branch_grow_sph` > 0): the verticality-weighted seed
         is high-precision but misses HORIZONTAL scaffold branches (low
         verticality → low score) even though they're woody (compact/low
         sphericity). Region-grow the wood label from the seed into CONNECTED
         low-sphericity points (`sphericity < branch_grow_sph`) to recover them
         (`_wood_grow_branches`). This is what makes thick crown branches read
         as wood rather than leaf.
      4. Optional speckle pruning (`min_speckle` > 1): flip tiny isolated wood
         components back to leaf.
      5. LeWoS-style graph regularisation (iterated k-NN majority vote).

    Linearity is deliberately NOT used: branches are linear (cylinders) and so
    are needles / narrow leaves, so linearity cannot separate the two. This was
    validated across two benchmark families — real TLS trees (Weiser et al.
    heiDATA: oak/beech/maple/pine/spruce, where neighbourhood sphericity carries
    the signal, mean OA ≈ 0.85) and synthetic almond scans (narrow flat leaves,
    where verticality carries it, mean OA ≈ 0.80). The additive blend handles
    both; the one weak case is densely-scattered-leaf forms (e.g. the synthetic
    central-leader archetype) where only linearity/planarity separate — a trade
    accepted because production input is real TLS.

    Very large clouds are handled automatically: above `max_points`
    (default `_WOOD_SEGMENT_MAX_POINTS` ≈ 1.5M, env-overridable) the geometry
    step runs on a voxel-downsampled subset and labels propagate back to full
    resolution by nearest neighbour — the per-point k-NN feature extraction is
    O(N·k_max) in memory and a multi-million-point cloud at full res can OOM the
    machine. Set `voxel_size` > 0 to choose the downsample resolution explicitly
    instead. Either way the returned labels are full-length.

    Returns an int32 array length len(points), aligned to input order, with
    values WOOD_CLASS_WOOD / WOOD_CLASS_LEAF.
    """
    # `warnings` is an optional caller-supplied list the connectivity method
    # appends to (e.g. a ground-not-removed notice); keep a local handle.
    _warnings = warnings if warnings is not None else []

    pts_full = np.ascontiguousarray(points[:, :3], dtype=np.float64)
    n_full = len(pts_full)
    if n_full == 0:
        return np.zeros(0, dtype=np.int32)

    # Reflectance must align 1:1 with the full cloud to survive the downsample
    # mapping below; a length mismatch means a caller bug — drop it rather than
    # silently misalign the assist.
    refl_full = None
    if reflectance is not None:
        refl_full = np.asarray(reflectance, dtype=np.float64).ravel()
        if refl_full.shape[0] != n_full or not np.isfinite(refl_full).any():
            refl_full = None

    cap = int(max_points) if max_points is not None else _WOOD_SEGMENT_MAX_POINTS

    # AUTO safety downsample: the per-point k-NN feature extraction allocates an
    # (N, k_max) int32 neighbour array plus transient query buffers, so a
    # multi-million-point cloud at full resolution can exhaust RAM (a 6.4M-point
    # tree peaked ~13 GB and hard-crashed a 16 GB machine). When the cloud is
    # over the cap and the caller hasn't picked a `voxel_size`, derive one that
    # targets ~`cap` points from the cloud's bounding volume, classify the
    # reduced set, and propagate labels back to full resolution by nearest
    # neighbour. Result stays full-length; only the heavy geometry step shrinks.
    if (not voxel_size or voxel_size <= 0) and cap > 0 and n_full > cap:
        span = pts_full.max(axis=0) - pts_full.min(axis=0)
        vol = float(np.prod(np.clip(span, 1e-6, None)))
        # voxel edge so that vol / edge^3 ≈ cap (one survivor per occupied voxel
        # is optimistic, so this overshoots the target — fine, it's a safety net).
        voxel_size = float((vol / cap) ** (1.0 / 3.0))

    # Voxel downsample (explicit or auto) → classify reduced set → propagate.
    if voxel_size and voxel_size > 0 and n_full > 0:
        import open3d as o3d
        from scipy.spatial import cKDTree

        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(pts_full)
        down = pcd.voxel_down_sample(float(voxel_size))
        pts = np.asarray(down.points, dtype=np.float64)
        if len(pts) < 3:
            pts = pts_full  # downsample collapsed the cloud; fall back
            refl = refl_full
        elif refl_full is not None:
            # voxel_down_sample averages/reorders points, so reflectance can't be
            # strided — assign each downsampled point its nearest full-res point's
            # reflectance (same NN mapping used to propagate labels back).
            _, nn_down = cKDTree(pts_full).query(pts, k=1, workers=-1)
            refl = refl_full[nn_down]
        else:
            refl = None
    else:
        pts = pts_full
        refl = refl_full

    n = len(pts)
    if n < 3:
        # Too few points to estimate local geometry — call everything leaf
        # (the conservative default; "wood removal" then keeps all points).
        labels_small = np.full(n_full, WOOD_CLASS_LEAF, dtype=np.int32)
        return labels_small

    # Per-cloud classification of the (possibly downsampled) `pts`. The geometric
    # method is the original point-wise pipeline; the connectivity method roots a
    # geodesic skeleton at the trunk base and recovers the woody backbone (it
    # falls back to geometry internally when the skeleton degenerates). Both
    # return down-resolution labels; the full-res propagation below is shared.
    if method == "sota":
        labels_down = _segment_wood_sota(
            pts, refl,
            k_min=k_min, k_max=k_max, k_step=k_step, wood_bias=wood_bias,
            reg_k=reg_k, reg_iters=reg_iters, min_speckle=min_speckle,
            branch_grow_sph=branch_grow_sph,
            reflectance_weight_max=reflectance_weight_max,
            backbone_support=backbone_support,
            warnings=_warnings,
        )
    elif method == "connectivity":
        labels_down = _segment_wood_connectivity(
            pts, refl,
            k_min=k_min, k_max=k_max, k_step=k_step, wood_bias=wood_bias,
            reg_k=reg_k, reg_iters=reg_iters, min_speckle=min_speckle,
            branch_grow_sph=branch_grow_sph,
            reflectance_weight_max=reflectance_weight_max,
            backbone_support=backbone_support,
            warnings=_warnings,
        )
    else:
        labels_down = _wood_geometric_core(
            pts, refl,
            k_min=k_min, k_max=k_max, k_step=k_step, wood_bias=wood_bias,
            reg_k=reg_k, reg_iters=reg_iters, min_speckle=min_speckle,
            branch_grow_sph=branch_grow_sph,
            reflectance_weight_max=reflectance_weight_max,
        )

    if n == n_full:
        return labels_down

    # Propagate downsampled labels back to full resolution by nearest neighbour,
    # then regularise once more at full resolution to smooth voxel seams.
    from scipy.spatial import cKDTree

    nn_tree = cKDTree(pts)
    _, nn = nn_tree.query(pts_full, k=1, workers=-1)
    labels_full = labels_down[nn].astype(np.int32)
    full_tree = cKDTree(pts_full)
    _, full_idx = full_tree.query(pts_full, k=int(min(reg_k, n_full)), workers=-1)
    if full_idx.ndim == 1:
        full_idx = full_idx[:, None]
    labels_full = _wood_regularize(labels_full, full_idx.astype(np.int32), reg_k, max(1, reg_iters))
    return labels_full


# ==================== BFS SKELETON HELPER FUNCTIONS (Li et al. 2017) ====================

def remove_statistical_outliers(points: np.ndarray, nb_neighbors: int = 20, std_ratio: float = 2.0) -> np.ndarray:
    """
    Remove statistical outliers using k-nearest neighbors.
    Points with mean distance > std_ratio * global std are removed.
    """
    try:
        import open3d as o3d
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(points)
        cl, ind = pcd.remove_statistical_outlier(nb_neighbors=nb_neighbors, std_ratio=std_ratio)
        return np.asarray(cl.points)
    except ImportError:
        # Fallback: manual implementation
        from scipy.spatial import KDTree
        tree = KDTree(points)
        distances, _ = tree.query(points, k=min(nb_neighbors + 1, len(points)))
        mean_distances = np.mean(distances[:, 1:], axis=1)  # Exclude self
        threshold = np.mean(mean_distances) + std_ratio * np.std(mean_distances)
        mask = mean_distances <= threshold
        return points[mask]


def build_neighbor_graph(points: np.ndarray, search_radius: float, max_neighbors: int = 20) -> dict:
    """
    Build an undirected graph connecting neighboring points using a KD-tree.

    Each point is linked to its up-to-`max_neighbors` nearest neighbours that
    lie within `search_radius`. Implemented as a single batched k-NN query
    (cKDTree, parallel workers) plus a radius mask — NOT a per-point
    `query_ball_point` Python loop, which on a multi-million-point TLS cloud
    materialises a Python list per point and runs for minutes. The batched
    query mirrors the wood-segmentation neighbour pass (`_wood_local_pca_features`).

    Because cKDTree returns neighbours sorted by increasing distance, taking the
    first `max_neighbors` after dropping self and applying the radius mask is
    exactly the "keep the closest within radius" semantics of the old code — a
    behaviour-preserving rewrite, not a quality change.

    Args:
        points: Nx3 array of point coordinates
        search_radius: Maximum distance to consider points as neighbors
        max_neighbors: Maximum number of neighbors per point

    Returns:
        dict with 'neighbors' (per-point int32 ndarray of neighbour indices, so
        `for j in neighbors[i]` still yields ints) and 'kdtree'
    """
    from scipy.spatial import cKDTree

    tree = cKDTree(points)
    n_points = len(points)
    if n_points == 0:
        return {'neighbors': [], 'kdtree': tree}

    # k+1: the nearest neighbour of a point is itself (distance 0); we drop it.
    k = min(max_neighbors + 1, n_points)

    # Query in chunks: a single full query allocates the (N,k) float64 distances
    # AND int64 indices at once, which OOMs on large clouds. Chunking caps the
    # transient to one block (same rationale as the wood-seg pass).
    neighbors: list = [None] * n_points
    chunk = 200_000
    for start in range(0, n_points, chunk):
        end = min(start + chunk, n_points)
        dist_block, idx_block = tree.query(points[start:end], k=k, workers=-1)
        if idx_block.ndim == 1:  # k == 1 edge case (n_points == 1)
            idx_block = idx_block[:, None]
            dist_block = dist_block[:, None]
        for r in range(end - start):
            gi = start + r          # global point index
            cols = idx_block[r]
            d = dist_block[r]
            keep = (cols != gi) & (d <= search_radius)  # drop self + outside radius
            sel = cols[keep]
            if sel.shape[0] > max_neighbors:
                sel = sel[:max_neighbors]  # already distance-sorted → closest
            neighbors[gi] = sel.astype(np.int32, copy=False)
        del dist_block, idx_block

    return {'neighbors': neighbors, 'kdtree': tree}


def select_root_set(points: np.ndarray, threshold: float = 0.02) -> list:
    """
    Select root points near the lowest point in z-direction.

    Args:
        points: Nx3 array of point coordinates
        threshold: Height threshold τ (meters) above lowest point

    Returns:
        List of indices of root points
    """
    z_min = np.min(points[:, 2])
    root_indices = np.where(points[:, 2] - z_min < threshold)[0].tolist()
    return root_indices


def bfs_label_points(neighbors: list, root_indices: list, n_points: int) -> np.ndarray:
    """
    Label all points with their BFS distance from root set.

    Args:
        neighbors: List of neighbor indices for each point
        root_indices: Indices of root points (labeled as 1)
        n_points: Total number of points

    Returns:
        Array of labels (distance from root, -1 for unreachable points)
    """
    from collections import deque

    labels = np.full(n_points, -1, dtype=np.int32)

    # Initialize root points with label 1
    queue = deque()
    for idx in root_indices:
        labels[idx] = 1
        queue.append(idx)

    # BFS traversal
    while queue:
        current = queue.popleft()
        current_label = labels[current]

        for neighbor in neighbors[current]:
            if labels[neighbor] == -1:  # Not yet visited
                labels[neighbor] = current_label + 1
                queue.append(neighbor)

    return labels


def quantize_labels(labels: np.ndarray, num_levels: int = 60, use_nonlinear: bool = True) -> np.ndarray:
    """
    Quantize BFS labels into discrete intervals.

    Args:
        labels: Raw BFS labels
        num_levels: Number of quantization levels (Q-S in paper)
        use_nonlinear: If True, use sqrt scaling for better branch detail

    Returns:
        Quantized labels (0 to num_levels)
    """
    valid_mask = labels > 0
    if not np.any(valid_mask):
        return np.zeros_like(labels)

    max_label = np.max(labels[valid_mask])
    quantized = np.zeros_like(labels)

    if use_nonlinear:
        # Nonlinear mapping: Val' = floor(sqrt(Val/Valmax) * num_levels)
        # This preserves more detail in branches far from root
        quantized[valid_mask] = np.floor(
            np.sqrt(labels[valid_mask] / max_label) * num_levels
        ).astype(np.int32)
    else:
        # Linear mapping
        quantized[valid_mask] = np.floor(
            (labels[valid_mask] / max_label) * num_levels
        ).astype(np.int32)

    return quantized


def cluster_blocks(quantized_labels: np.ndarray, neighbors: list) -> tuple:
    """
    Cluster connected points with the same quantized label into blocks.
    Uses DFS to find connected components within each quantization level.

    Args:
        quantized_labels: Quantized BFS labels
        neighbors: List of neighbor indices for each point

    Returns:
        (block_assignments, block_info) where:
        - block_assignments: Array mapping each point to its block ID (-1 if not assigned)
        - block_info: List of dicts with block metadata
    """
    n_points = len(quantized_labels)
    block_assignments = np.full(n_points, -1, dtype=np.int32)
    block_info = []
    current_block_id = 0

    visited = np.zeros(n_points, dtype=bool)

    for start_idx in range(n_points):
        if visited[start_idx] or quantized_labels[start_idx] < 0:
            continue

        # DFS to find all connected points with same quantized label
        target_level = quantized_labels[start_idx]
        block_points = []
        stack = [start_idx]

        while stack:
            idx = stack.pop()
            if visited[idx]:
                continue
            if quantized_labels[idx] != target_level:
                continue

            visited[idx] = True
            block_points.append(idx)
            block_assignments[idx] = current_block_id

            # Add unvisited neighbors with same label
            for neighbor in neighbors[idx]:
                if not visited[neighbor] and quantized_labels[neighbor] == target_level:
                    stack.append(neighbor)

        if len(block_points) > 0:
            block_info.append({
                'id': current_block_id,
                'level': target_level,
                'point_indices': block_points,
                'size': len(block_points)
            })
            current_block_id += 1

    return block_assignments, block_info


def find_block_connectivity(block_assignments: np.ndarray, block_info: list,
                           neighbors: list) -> dict:
    """
    Find which blocks are connected based on edge connectivity in the graph.

    Returns:
        Dict mapping block_id -> set of connected block_ids
    """
    connectivity = {b['id']: set() for b in block_info}

    n_points = len(block_assignments)
    for i in range(n_points):
        if block_assignments[i] < 0:
            continue

        my_block = block_assignments[i]
        for neighbor in neighbors[i]:
            neighbor_block = block_assignments[neighbor]
            if neighbor_block >= 0 and neighbor_block != my_block:
                connectivity[my_block].add(neighbor_block)
                connectivity[neighbor_block].add(my_block)

    return connectivity


def filter_blocks(block_info: list, block_assignments: np.ndarray,
                  connectivity: dict, threshold_filter: int = 30,
                  use_proportion_filter: bool = True,
                  proportion_threshold: float = 0.1) -> tuple:
    """
    Filter out small blocks (noise, leaves, small twigs).

    Filters:
    1. Threshold filter: Remove blocks with fewer than threshold_filter points
    2. Proportion filter: Remove blocks whose size is too small relative to parent

    Returns:
        (filtered_block_info, filtered_assignments, filtered_connectivity)
    """
    # Sort blocks by level (lower level = closer to root = parent)
    sorted_blocks = sorted(block_info, key=lambda b: b['level'])

    # Build parent-child relationships based on connectivity and level
    for block in sorted_blocks:
        block['parent_id'] = None
        block['children'] = []

    block_by_id = {b['id']: b for b in sorted_blocks}

    for block in sorted_blocks:
        # Find parent: connected block with lower level
        min_parent_level = block['level']
        parent_id = None
        for connected_id in connectivity.get(block['id'], []):
            connected_block = block_by_id.get(connected_id)
            if connected_block and connected_block['level'] < min_parent_level:
                min_parent_level = connected_block['level']
                parent_id = connected_id

        if parent_id is not None:
            block['parent_id'] = parent_id
            block_by_id[parent_id]['children'].append(block['id'])

    # Apply filters
    keep_block_ids = set()

    for block in sorted_blocks:
        # Threshold filter
        if block['size'] < threshold_filter:
            continue

        # Proportion filter
        if use_proportion_filter and block['parent_id'] is not None:
            parent = block_by_id[block['parent_id']]
            if block['size'] / parent['size'] < proportion_threshold:
                continue

        keep_block_ids.add(block['id'])

    # Also keep ancestors of kept blocks
    for block in sorted_blocks:
        if block['id'] in keep_block_ids:
            # Walk up to root, keeping all ancestors
            current = block
            while current['parent_id'] is not None:
                keep_block_ids.add(current['parent_id'])
                current = block_by_id[current['parent_id']]

    # Filter
    filtered_blocks = [b for b in block_info if b['id'] in keep_block_ids]

    # Update assignments
    filtered_assignments = block_assignments.copy()
    for i in range(len(filtered_assignments)):
        if filtered_assignments[i] not in keep_block_ids:
            filtered_assignments[i] = -1

    # Update connectivity
    filtered_connectivity = {
        bid: {c for c in conns if c in keep_block_ids}
        for bid, conns in connectivity.items()
        if bid in keep_block_ids
    }

    return filtered_blocks, filtered_assignments, filtered_connectivity


def compute_skeleton_nodes(points: np.ndarray, block_info: list) -> dict:
    """
    Compute skeleton nodes as centroids of each block.

    Returns:
        Dict mapping block_id -> centroid [x, y, z]
    """
    skeleton_nodes = {}

    for block in block_info:
        block_points = points[block['point_indices']]
        centroid = np.mean(block_points, axis=0)
        skeleton_nodes[block['id']] = centroid.tolist()

    return skeleton_nodes


def build_skeleton_tree(block_info: list, connectivity: dict,
                        skeleton_nodes: dict) -> tuple:
    """
    Build skeleton edges based on block connectivity.
    Uses a tree structure rooted at the lowest-level block.

    Returns:
        (edges, edge_lengths) where edges are pairs of block IDs
    """
    if not block_info:
        return [], []

    # Sort by level
    sorted_blocks = sorted(block_info, key=lambda b: b['level'])
    block_by_id = {b['id']: b for b in sorted_blocks}

    edges = []
    edge_lengths = []

    # Connect each block to its parent
    for block in sorted_blocks:
        if block.get('parent_id') is not None and block['parent_id'] in skeleton_nodes:
            parent_id = block['parent_id']
            child_id = block['id']

            if child_id in skeleton_nodes:
                parent_pos = np.array(skeleton_nodes[parent_id])
                child_pos = np.array(skeleton_nodes[child_id])
                length = float(np.linalg.norm(child_pos - parent_pos))

                # Skip very short edges (likely noise)
                if length > 0.001:
                    edges.append([parent_id, child_id])
                    edge_lengths.append(length)

    return edges, edge_lengths


def laplace_smooth_skeleton(skeleton_nodes: dict, edges: list,
                            iterations: int = 2) -> dict:
    """
    Smooth skeleton using Laplace smoothing.
    Each non-bifurcation node is moved to average of itself and neighbors.

    NEW_A = (A + B + C) / 3  where B is parent, C is child
    """
    if iterations <= 0 or not edges:
        return skeleton_nodes

    # Build adjacency from edges
    adjacency = {}
    for block_id in skeleton_nodes:
        adjacency[block_id] = []

    for parent_id, child_id in edges:
        if parent_id in adjacency and child_id in adjacency:
            adjacency[parent_id].append(child_id)
            adjacency[child_id].append(parent_id)

    # Find degree of each node
    degrees = {bid: len(adj) for bid, adj in adjacency.items()}

    smoothed = {bid: np.array(pos) for bid, pos in skeleton_nodes.items()}

    for _ in range(iterations):
        new_positions = {}

        for block_id, pos in smoothed.items():
            neighbors = adjacency.get(block_id, [])
            degree = degrees.get(block_id, 0)

            # Skip bifurcation points (degree > 2) and endpoints (degree <= 1)
            if degree != 2:
                new_positions[block_id] = pos
                continue

            # Average with neighbors
            neighbor_sum = np.zeros(3)
            for neighbor_id in neighbors:
                neighbor_sum += smoothed[neighbor_id]

            new_positions[block_id] = (pos + neighbor_sum) / 3

        smoothed = new_positions

    return {bid: pos.tolist() for bid, pos in smoothed.items()}


def calculate_skeleton_length(skeleton_nodes: dict, edges: list) -> float:
    """Calculate total length of skeleton by summing edge lengths."""
    total_length = 0.0

    for parent_id, child_id in edges:
        if parent_id in skeleton_nodes and child_id in skeleton_nodes:
            p1 = np.array(skeleton_nodes[parent_id])
            p2 = np.array(skeleton_nodes[child_id])
            total_length += np.linalg.norm(p2 - p1)

    return total_length


def count_branch_points(edges: list) -> int:
    """Count nodes with more than 2 connections (branch points)."""
    if not edges:
        return 0

    degree = {}
    for parent_id, child_id in edges:
        degree[parent_id] = degree.get(parent_id, 0) + 1
        degree[child_id] = degree.get(child_id, 0) + 1

    return sum(1 for d in degree.values() if d > 2)


def calculate_skeleton_length_from_edges(skeleton_nodes: dict, edges: list) -> float:
    """
    Calculate total skeleton length by summing all edge lengths.

    Args:
        skeleton_nodes: dict mapping block_id to [x, y, z] coordinates
        edges: list of (parent_id, child_id) tuples

    Returns:
        Total length of all edges
    """
    if not edges or not skeleton_nodes:
        return 0.0

    total_length = 0.0
    for parent_id, child_id in edges:
        if parent_id in skeleton_nodes and child_id in skeleton_nodes:
            p1 = np.array(skeleton_nodes[parent_id])
            p2 = np.array(skeleton_nodes[child_id])
            total_length += np.linalg.norm(p2 - p1)

    return float(total_length)


def calculate_branch_orders(skeleton_nodes: dict, edges: list, block_info: list) -> dict:
    """
    Calculate Strahler branch order for each node in the skeleton.

    Branch order (Strahler number) classification:
    - Order 1: terminal branches (tips/leaves)
    - When two branches of same order n meet, parent gets order n+1
    - When branches of different orders meet, parent gets the higher order

    Uses iterative post-order traversal to avoid Python recursion limits.

    Args:
        skeleton_nodes: dict mapping block_id to [x, y, z] coordinates
        edges: list of [parent_id, child_id] lists
        block_info: list of block dictionaries with 'id' and 'level' keys

    Returns:
        dict mapping block_id to branch order (int)
    """
    if not skeleton_nodes or not edges:
        print(f"[Branch Order] No skeleton nodes or edges")
        return {bid: 1 for bid in skeleton_nodes}

    # Build adjacency list (undirected)
    adjacency = {bid: set() for bid in skeleton_nodes}
    for edge in edges:
        parent_id, child_id = edge[0], edge[1]
        if parent_id in adjacency and child_id in adjacency:
            adjacency[parent_id].add(child_id)
            adjacency[child_id].add(parent_id)

    # Find leaves (nodes with degree 1)
    leaves = [bid for bid, neighbors in adjacency.items() if len(neighbors) <= 1]
    print(f"[Branch Order] Found {len(leaves)} leaves out of {len(skeleton_nodes)} nodes")

    # Initialize: all leaves get order 1
    branch_orders = {bid: 1 for bid in leaves}

    # Use iterative approach: process from leaves toward root
    # Track which children have been processed for each node
    processed = set(leaves)
    to_process = list(leaves)

    while to_process:
        current = to_process.pop(0)

        for neighbor in adjacency[current]:
            if neighbor in processed:
                continue

            # Check if all children of neighbor (except current's direction) are processed
            neighbor_children = [n for n in adjacency[neighbor] if n != current and n in processed]
            unprocessed_neighbors = [n for n in adjacency[neighbor] if n not in processed]

            # Only process this neighbor if all its other neighbors are processed
            # (current is the parent direction, all others should be children)
            if len(unprocessed_neighbors) <= 1:  # Only unprocessed one is current's direction (or none)
                # Get orders of all processed neighbors (children in tree sense)
                child_orders = [branch_orders[n] for n in adjacency[neighbor] if n in processed]

                if not child_orders:
                    # No children processed yet - this shouldn't happen, but handle it
                    branch_orders[neighbor] = 1
                else:
                    # Calculate Strahler order
                    max_order = max(child_orders)
                    count_max = child_orders.count(max_order)

                    if count_max >= 2:
                        branch_orders[neighbor] = max_order + 1
                    else:
                        branch_orders[neighbor] = max_order

                processed.add(neighbor)
                to_process.append(neighbor)

    # Handle any unprocessed nodes (disconnected components)
    disconnected = 0
    for bid in skeleton_nodes:
        if bid not in branch_orders:
            branch_orders[bid] = 1
            disconnected += 1

    # Debug output
    order_counts = {}
    for order in branch_orders.values():
        order_counts[order] = order_counts.get(order, 0) + 1
    max_order = max(branch_orders.values()) if branch_orders else 0
    print(f"[Branch Order] Order distribution: {order_counts}, max: {max_order}, disconnected: {disconnected}")

    return branch_orders


def order_skeleton_points(skeleton_nodes: dict, edges: list, block_info: list) -> list:
    """
    Order skeleton points from root to tips using BFS from lowest-level node.
    Returns list of [x, y, z] coordinates in order.
    """
    if not skeleton_nodes:
        return []

    if not edges:
        # No edges, just return nodes sorted by level
        block_by_id = {b['id']: b for b in block_info}
        sorted_ids = sorted(skeleton_nodes.keys(),
                           key=lambda bid: block_by_id.get(bid, {}).get('level', 0))
        return [skeleton_nodes[bid] for bid in sorted_ids]

    # Build adjacency
    adjacency = {bid: [] for bid in skeleton_nodes}
    for parent_id, child_id in edges:
        if parent_id in adjacency and child_id in adjacency:
            adjacency[parent_id].append(child_id)
            adjacency[child_id].append(parent_id)

    # Find root (lowest level node)
    block_by_id = {b['id']: b for b in block_info}
    root_id = min(skeleton_nodes.keys(),
                  key=lambda bid: block_by_id.get(bid, {}).get('level', float('inf')))

    # BFS from root
    from collections import deque
    ordered = []
    visited = set()
    queue = deque([root_id])

    while queue:
        current = queue.popleft()
        if current in visited:
            continue
        visited.add(current)
        ordered.append(skeleton_nodes[current])

        for neighbor in adjacency.get(current, []):
            if neighbor not in visited:
                queue.append(neighbor)

    return ordered


def order_skeleton_points_with_mapping(skeleton_nodes: dict, edges: list, block_info: list) -> tuple:
    """
    Convert skeleton nodes dict to ordered list with index mapping.
    Returns tuple of (list of [x, y, z] coordinates, dict mapping block_id to array index).

    Important: We must include ALL nodes and create a complete mapping so that
    edge indices can be correctly converted to array positions.
    """
    if not skeleton_nodes:
        return [], {}

    # Sort nodes by level (lowest first = root to tips)
    block_by_id = {b['id']: b for b in block_info}
    sorted_ids = sorted(skeleton_nodes.keys(),
                       key=lambda bid: block_by_id.get(bid, {}).get('level', 0))

    # Create ordered list and mapping
    ordered = [skeleton_nodes[bid] for bid in sorted_ids]
    id_to_idx = {bid: idx for idx, bid in enumerate(sorted_ids)}

    return ordered, id_to_idx


# Legacy function for API compatibility
def fit_circle_ransac(points_2d: np.ndarray, n_iterations: int = 100,
                       threshold_ratio: float = 0.02, min_inliers_ratio: float = 0.5) -> dict:
    """
    Robust circle fitting using RANSAC.

    Args:
        points_2d: Nx2 array of 2D points
        n_iterations: Number of RANSAC iterations
        threshold_ratio: Inlier threshold as fraction of estimated radius
        min_inliers_ratio: Minimum fraction of points that must be inliers

    Returns:
        dict with center, radius, inliers, rmse, confidence
    """
    n_points = len(points_2d)
    if n_points < 3:
        return {"success": False, "error": "Need at least 3 points"}

    # Initial estimate for threshold
    centroid = np.mean(points_2d, axis=0)
    est_radius = np.median(np.linalg.norm(points_2d - centroid, axis=1))
    threshold = threshold_ratio * est_radius

    best_inliers = []
    best_center = centroid
    best_radius = est_radius

    for _ in range(n_iterations):
        # Random sample of 3 points
        idx = np.random.choice(n_points, size=min(3, n_points), replace=False)
        sample = points_2d[idx]

        # Fit circle through 3 points
        try:
            center, radius = fit_circle_through_3_points(sample)
            if center is None or radius <= 0 or radius > est_radius * 5:
                continue
        except:
            continue

        # Count inliers
        distances = np.abs(np.linalg.norm(points_2d - center, axis=1) - radius)
        inliers = np.where(distances < threshold)[0]

        if len(inliers) > len(best_inliers):
            best_inliers = inliers
            best_center = center
            best_radius = radius

    # Refine with all inliers using least squares
    if len(best_inliers) >= 3:
        inlier_points = points_2d[best_inliers]
        refined = fit_circle_least_squares(inlier_points, best_center, best_radius)
        if refined["success"]:
            best_center = refined["center"]
            best_radius = refined["radius"]

            # Recalculate inliers with refined model
            distances = np.abs(np.linalg.norm(points_2d - best_center, axis=1) - best_radius)
            best_inliers = np.where(distances < threshold)[0]

    # Calculate RMSE on inliers
    if len(best_inliers) > 0:
        inlier_distances = np.linalg.norm(points_2d[best_inliers] - best_center, axis=1)
        rmse = np.sqrt(np.mean((inlier_distances - best_radius) ** 2))
        confidence = len(best_inliers) / n_points
    else:
        rmse = None
        confidence = 0.0

    return {
        "success": len(best_inliers) >= n_points * min_inliers_ratio,
        "center": best_center,
        "radius": best_radius,
        "inliers": best_inliers,
        "rmse": rmse,
        "confidence": confidence
    }


def fit_circle_through_3_points(points: np.ndarray) -> tuple:
    """
    Fit a circle through exactly 3 points.
    Returns (center, radius) or (None, None) if collinear.
    """
    if len(points) != 3:
        return None, None

    ax, ay = points[0]
    bx, by = points[1]
    cx, cy = points[2]

    d = 2 * (ax * (by - cy) + bx * (cy - ay) + cx * (ay - by))
    if abs(d) < 1e-10:  # Collinear
        return None, None

    ux = ((ax*ax + ay*ay) * (by - cy) + (bx*bx + by*by) * (cy - ay) + (cx*cx + cy*cy) * (ay - by)) / d
    uy = ((ax*ax + ay*ay) * (cx - bx) + (bx*bx + by*by) * (ax - cx) + (cx*cx + cy*cy) * (bx - ax)) / d

    center = np.array([ux, uy])
    radius = np.linalg.norm(points[0] - center)

    return center, radius


def fit_circle_least_squares(points_2d: np.ndarray, center_init: np.ndarray = None,
                              radius_init: float = None) -> dict:
    """
    Fit circle using least squares (for refinement).
    """
    from scipy.optimize import least_squares

    if len(points_2d) < 3:
        return {"success": False}

    if center_init is None:
        center_init = np.mean(points_2d, axis=0)
    if radius_init is None:
        radius_init = np.median(np.linalg.norm(points_2d - center_init, axis=1))

    def residuals(params):
        cx, cy, r = params
        center = np.array([cx, cy])
        distances = np.linalg.norm(points_2d - center, axis=1)
        return distances - r

    try:
        result = least_squares(residuals, [center_init[0], center_init[1], radius_init],
                              method='lm', max_nfev=100)
        cx, cy, r = result.x
        rmse = np.sqrt(np.mean(result.fun ** 2))
        return {
            "success": True,
            "center": np.array([cx, cy]),
            "radius": abs(r),
            "rmse": rmse
        }
    except:
        return {"success": False}


def adaptive_slice_extraction(
    points: np.ndarray,
    principal_direction: np.ndarray,
    num_slices: Optional[int] = None,
    slice_thickness: Optional[float] = None,
    min_points_per_slice: int = 10,
    use_local_pca: bool = True,
    fit_circles: bool = True,
    use_ransac: bool = True,
    ransac_iterations: int = 100,
    ransac_threshold: float = 0.02
) -> tuple:
    """
    Extract skeleton using adaptive slicing along principal direction.
    Uses local PCA to orient each slice perpendicular to local stem direction.

    Returns:
        slices: List of slice information dicts
        slice_thickness: The thickness used
    """
    # Project points onto principal axis
    centroid = np.mean(points, axis=0)
    projections = np.dot(points - centroid, principal_direction)

    # Get bounds along principal axis
    h_min = np.min(projections)
    h_max = np.max(projections)
    h_range = h_max - h_min

    if h_range <= 0:
        return [], 0

    # Determine slice thickness
    if slice_thickness is not None:
        thickness = slice_thickness
    elif num_slices is not None:
        thickness = h_range / num_slices
    else:
        # Auto: aim for ~50 slices, but ensure each has enough points
        target_slices = min(50, len(points) // (min_points_per_slice * 2))
        target_slices = max(10, target_slices)
        thickness = h_range / target_slices

    # Create slices
    slices = []
    current_h = h_min + thickness / 2

    # Local direction tracking for smoothness
    local_direction = principal_direction.copy()

    while current_h < h_max:
        # Get points in this slice (along principal direction)
        mask = (projections >= current_h - thickness/2) & (projections < current_h + thickness/2)
        slice_points = points[mask]

        if len(slice_points) >= min_points_per_slice:
            # Compute local direction using local PCA
            if use_local_pca and len(slice_points) >= 10:
                try:
                    local_dir, _, _ = compute_pca_direction(slice_points)
                    # Ensure consistent direction (don't flip 180 degrees)
                    if np.dot(local_dir, local_direction) < 0:
                        local_dir = -local_dir
                    # Smooth transition: blend with previous direction
                    local_direction = 0.7 * local_dir + 0.3 * local_direction
                    local_direction = local_direction / np.linalg.norm(local_direction)
                except:
                    pass  # Keep previous direction

            # Slice center point (along the axis)
            slice_center_point = centroid + current_h * principal_direction

            # Project slice points onto plane perpendicular to local direction
            points_2d, u1, u2 = project_to_plane(slice_points, local_direction, slice_center_point)

            # Circle fitting
            diameter = None
            fit_rmse = None
            num_inliers = None
            confidence = None
            center_3d = np.mean(slice_points, axis=0)  # Default to centroid

            if fit_circles and len(points_2d) >= 4:
                if use_ransac:
                    result = fit_circle_ransac(
                        points_2d,
                        n_iterations=ransac_iterations,
                        threshold_ratio=ransac_threshold
                    )
                else:
                    result = fit_circle_least_squares(points_2d)
                    if result["success"]:
                        result["confidence"] = 1.0
                        result["inliers"] = list(range(len(points_2d)))

                if result.get("success"):
                    center_2d = result["center"]
                    diameter = result["radius"] * 2
                    fit_rmse = result.get("rmse")
                    num_inliers = len(result.get("inliers", []))
                    confidence = result.get("confidence", 0.0)

                    # Convert 2D center back to 3D
                    center_3d = slice_center_point + center_2d[0] * u1 + center_2d[1] * u2

            slices.append({
                "center": center_3d.tolist(),
                "height": float(current_h),
                "diameter": diameter,
                "num_points": len(slice_points),
                "num_inliers": num_inliers,
                "fit_rmse": fit_rmse,
                "local_direction": local_direction.tolist(),
                "confidence": confidence
            })

        current_h += thickness

    return slices, thickness


def remove_outlier_skeleton_points(slices: list, threshold_factor: float = 2.5) -> list:
    """
    Remove skeleton points that deviate significantly from the local trajectory.
    Uses distance from a smoothed trajectory as criterion.
    """
    if len(slices) < 5:
        return slices

    centers = np.array([s["center"] for s in slices])

    # Compute local deviation: distance from line through neighbors
    deviations = []
    for i in range(len(centers)):
        if i == 0:
            # First point: check distance from line to next two points
            if len(centers) >= 3:
                v = centers[2] - centers[1]
                v = v / (np.linalg.norm(v) + 1e-10)
                proj = centers[1] + np.dot(centers[i] - centers[1], v) * v
                deviations.append(np.linalg.norm(centers[i] - proj))
            else:
                deviations.append(0)
        elif i == len(centers) - 1:
            # Last point: check distance from line through previous two points
            v = centers[i-1] - centers[i-2]
            v = v / (np.linalg.norm(v) + 1e-10)
            proj = centers[i-1] + np.dot(centers[i] - centers[i-1], v) * v
            deviations.append(np.linalg.norm(centers[i] - proj))
        else:
            # Middle points: distance from line through neighbors
            v = centers[i+1] - centers[i-1]
            v = v / (np.linalg.norm(v) + 1e-10)
            proj = centers[i-1] + np.dot(centers[i] - centers[i-1], v) * v
            deviations.append(np.linalg.norm(centers[i] - proj))

    deviations = np.array(deviations)
    threshold = np.median(deviations) + threshold_factor * np.std(deviations)

    # Filter slices
    filtered_slices = [s for s, d in zip(slices, deviations) if d <= threshold]

    return filtered_slices


def smooth_skeleton_robust(skeleton_points: np.ndarray, smoothing_factor: float = 0.0) -> np.ndarray:
    """
    Smooth skeleton using spline interpolation with curvature constraints.
    """
    from scipy.interpolate import splprep, splev

    if len(skeleton_points) < 4:
        return skeleton_points

    try:
        # Calculate cumulative arc length for parameterization
        diffs = np.diff(skeleton_points, axis=0)
        segment_lengths = np.linalg.norm(diffs, axis=1)
        cumulative_length = np.concatenate([[0], np.cumsum(segment_lengths)])
        total_length = cumulative_length[-1]

        if total_length <= 0:
            return skeleton_points

        # Normalize to [0, 1]
        u = cumulative_length / total_length

        # Fit spline
        tck, _ = splprep(
            [skeleton_points[:, 0], skeleton_points[:, 1], skeleton_points[:, 2]],
            u=u,
            s=smoothing_factor * total_length,  # Scale smoothing by length
            k=min(3, len(skeleton_points) - 1)
        )

        # Evaluate at same number of points
        u_new = np.linspace(0, 1, len(skeleton_points))
        smoothed = np.array(splev(u_new, tck)).T
        return smoothed
    except Exception as e:
        print(f"Smoothing failed: {e}")
        return skeleton_points


def smooth_diameters(diameters: list, window_size: int = 5) -> list:
    """
    Smooth diameter profile using median filter.
    Returns a list of floats with NO None values (interpolates missing values).
    Returns None if all diameters are None.
    """
    from scipy.ndimage import median_filter

    if not diameters:
        return None

    # Handle None values - check if we have any valid values
    valid_mask = [d is not None for d in diameters]
    if not any(valid_mask):
        return None  # All None, can't compute diameters

    # Convert to array, using NaN for None values
    arr = np.array([d if d is not None else np.nan for d in diameters], dtype=np.float64)

    # Interpolate NaN values using linear interpolation
    nans = np.isnan(arr)
    if np.any(nans):
        x = np.arange(len(arr))
        valid_x = x[~nans]
        valid_y = arr[~nans]
        arr[nans] = np.interp(x[nans], valid_x, valid_y)

    # Apply median filter if we have enough values
    if len(arr) >= window_size:
        smoothed = median_filter(arr, size=window_size)
    else:
        smoothed = arr

    return smoothed.tolist()


def calculate_skeleton_length(skeleton_points: np.ndarray) -> float:
    """Calculate total length of skeleton by summing segment lengths."""
    if len(skeleton_points) < 2:
        return 0.0

    diff = np.diff(skeleton_points, axis=0)
    segment_lengths = np.sqrt(np.sum(diff**2, axis=1))
    return float(np.sum(segment_lengths))


@app.post("/api/skeleton/extract", response_model=SkeletonResponse)
async def extract_stem_skeleton(request: SkeletonRequest):
    """
    Extract tree skeleton from point cloud using BFS graph-based algorithm.

    Based on Li et al. 2017 "An Automatic Tree Skeleton Extracting Method
    Based on Point Cloud of Terrestrial Laser Scanner"

    Algorithm stages:
    1. Pre-processing: Statistical outlier removal
    2. Build KD-tree neighbor graph
    3. Select root points near base (lowest z-coordinate)
    4. BFS label all points with distance from root
    5. Nonlinear quantization using sqrt scaling
    6. Cluster connected points with same quantized label into blocks
    7. Filter blocks (threshold filter + proportion filter)
    8. Compute skeleton nodes from block centroids
    9. Build skeleton edges based on block connectivity
    10. Apply Laplace smoothing

    Returns skeleton points, edges, and metrics.
    """
    try:
        if request.source is not None:
            # Octree-backed cloud: read (and downsample) from the source file.
            points, _, _ = _read_points_from_source(request.source)
        else:
            points = np.array(request.points or [], dtype=np.float64)
        points_original_count = len(points)

        if points_original_count > _SKELETON_MAX_POINTS:
            return SkeletonResponse(
                success=False,
                skeleton_points=[],
                num_nodes=0,
                dominant_axis=request.dominant_axis,
                points_before_filtering=points_original_count,
                points_after_filtering=points_original_count,
                error=(
                    f"{points_original_count:,} points exceeds the "
                    f"{_SKELETON_MAX_POINTS:,}-point limit for skeleton extraction. "
                    "Downsample or crop first."
                ),
            )

        if len(points) < 50:
            return SkeletonResponse(
                success=False,
                skeleton_points=[],
                num_nodes=0,
                dominant_axis=request.dominant_axis,
                points_before_filtering=points_original_count,
                points_after_filtering=len(points),
                error="Need at least 50 points for BFS skeleton extraction"
            )

        # Stage 1: Pre-processing - Statistical outlier removal
        if request.remove_outliers:
            points = remove_statistical_outliers(
                points,
                nb_neighbors=request.outlier_nb_neighbors,
                std_ratio=request.outlier_std_ratio
            )

            if len(points) < 50:
                return SkeletonResponse(
                    success=False,
                    skeleton_points=[],
                    num_nodes=0,
                    dominant_axis=request.dominant_axis,
                    points_before_filtering=points_original_count,
                    points_after_filtering=len(points),
                    error="Too few points remaining after outlier removal. Try increasing std_ratio."
                )

        points_after_filtering = len(points)
        print(f"[BFS Skeleton] After outlier removal: {points_after_filtering} points")

        # Auto-calculate search_radius from point density when it's left unset
        # (the UI sends 0 for "auto"). Done backend-side so octree clouds —
        # whose renderer has no positions to sample — get a sensible radius;
        # the flat path historically computed this in JS, but the backend KD-tree
        # estimate is identical and works for both paths.
        effective_search_radius = request.search_radius
        if effective_search_radius is None or effective_search_radius < 0.001:
            from scipy.spatial import KDTree as _KDTree
            sample_n = min(500, len(points))
            # Deterministic evenly-spaced sample (no RNG) — same density estimate
            # without per-call variation.
            sample_idx = np.linspace(0, len(points) - 1, sample_n).astype(np.int64)
            tree = _KDTree(points)
            # k=2: nearest is the point itself (dist 0), second is the true NN.
            dists, _ = tree.query(points[sample_idx], k=2)
            nn = dists[:, 1]
            nn = nn[nn > 0]
            avg_nn = float(nn.mean()) if nn.size > 0 else 0.05
            effective_search_radius = avg_nn * 2.5
            print(f"[BFS Skeleton] Auto search_radius={effective_search_radius:.4f} "
                  f"(avg NN dist {avg_nn:.4f})")

        # Stage 2: Build KD-tree neighbor graph
        print(f"[BFS Skeleton] Building neighbor graph with radius={effective_search_radius}")
        graph_info = build_neighbor_graph(
            points,
            search_radius=effective_search_radius,
            max_neighbors=request.max_neighbors
        )
        neighbors = graph_info['neighbors']

        # Check connectivity
        connected_count = sum(1 for n in neighbors if len(n) > 0)
        print(f"[BFS Skeleton] Connected points: {connected_count}/{len(points)}")

        if connected_count < len(points) * 0.5:
            # Too many disconnected points - try to auto-adjust radius
            print(f"[BFS Skeleton] Warning: Many disconnected points. Consider increasing search_radius.")

        # Stage 3: Select root points (near base)
        root_indices = select_root_set(points, threshold=request.root_threshold)
        print(f"[BFS Skeleton] Selected {len(root_indices)} root points")

        if len(root_indices) == 0:
            return SkeletonResponse(
                success=False,
                skeleton_points=[],
                num_nodes=0,
                dominant_axis=request.dominant_axis,
                points_before_filtering=points_original_count,
                points_after_filtering=points_after_filtering,
                error="No root points found. Check root_threshold parameter."
            )

        # Stage 4: BFS label all points with distance from root
        print("[BFS Skeleton] Running BFS labeling...")
        labels = bfs_label_points(neighbors, root_indices, len(points))

        labeled_count = np.sum(labels >= 0)
        print(f"[BFS Skeleton] Labeled {labeled_count}/{len(points)} points")

        if labeled_count < len(points) * 0.3:
            return SkeletonResponse(
                success=False,
                skeleton_points=[],
                num_nodes=0,
                dominant_axis=request.dominant_axis,
                points_before_filtering=points_original_count,
                points_after_filtering=points_after_filtering,
                error=f"BFS only reached {labeled_count} points ({100*labeled_count/len(points):.1f}%). "
                      "Try increasing search_radius for better connectivity."
            )

        # Stage 5: Nonlinear quantization using sqrt scaling
        print(f"[BFS Skeleton] Quantizing labels into {request.quantization_levels} levels...")
        quantized = quantize_labels(
            labels,
            num_levels=request.quantization_levels,
            use_nonlinear=request.use_nonlinear_quantization
        )

        unique_levels = len(np.unique(quantized[quantized >= 0]))
        print(f"[BFS Skeleton] Active quantization levels: {unique_levels}")

        # Stage 6: Cluster connected points with same quantized label into blocks
        print("[BFS Skeleton] Clustering blocks...")
        block_assignments, block_info = cluster_blocks(quantized, neighbors)
        num_blocks_before = len(block_info)
        print(f"[BFS Skeleton] Created {num_blocks_before} blocks")

        if num_blocks_before == 0:
            return SkeletonResponse(
                success=False,
                skeleton_points=[],
                num_nodes=0,
                dominant_axis=request.dominant_axis,
                points_before_filtering=points_original_count,
                points_after_filtering=points_after_filtering,
                error="No blocks created. Check point cloud connectivity."
            )

        # Find block connectivity (which blocks connect to which)
        connectivity = find_block_connectivity(block_assignments, block_info, neighbors)

        # Stage 7: Filter blocks using threshold filter and proportion filter
        print(f"[BFS Skeleton] Filtering blocks (threshold={request.threshold_filter})...")
        filtered_blocks, filtered_assignments, filtered_connectivity = filter_blocks(
            block_info,
            block_assignments,
            connectivity,
            threshold_filter=request.threshold_filter,
            use_proportion_filter=request.use_proportion_filter,
            proportion_threshold=request.proportion_threshold
        )
        num_blocks_after = len(filtered_blocks)
        print(f"[BFS Skeleton] Blocks after filtering: {num_blocks_after}")

        if num_blocks_after < 2:
            return SkeletonResponse(
                success=False,
                skeleton_points=[],
                num_nodes=0,
                num_blocks_before_filter=num_blocks_before,
                num_blocks_after_filter=num_blocks_after,
                dominant_axis=request.dominant_axis,
                points_before_filtering=points_original_count,
                points_after_filtering=points_after_filtering,
                error=f"Only {num_blocks_after} blocks after filtering. "
                      "Try reducing threshold_filter or disabling proportion_filter."
            )

        # Stage 8: Compute skeleton nodes from block centroids
        print("[BFS Skeleton] Computing skeleton nodes...")
        skeleton_nodes = compute_skeleton_nodes(points, filtered_blocks)

        # Stage 9: Build skeleton edges based on block connectivity
        print("[BFS Skeleton] Building skeleton tree...")
        edges, edge_lengths = build_skeleton_tree(filtered_blocks, filtered_connectivity, skeleton_nodes)
        print(f"[BFS Skeleton] Created {len(edges)} edges")

        # Stage 10: Apply Laplace smoothing
        if request.smooth_skeleton and len(skeleton_nodes) > 2:
            print(f"[BFS Skeleton] Smoothing ({request.smoothing_iterations} iterations)...")
            skeleton_nodes = laplace_smooth_skeleton(
                skeleton_nodes,
                edges,
                iterations=request.smoothing_iterations
            )

        # Order skeleton points from root to tips and get the ordering
        ordered_points, block_id_to_idx = order_skeleton_points_with_mapping(skeleton_nodes, edges, filtered_blocks)

        # Calculate total length
        total_length = calculate_skeleton_length_from_edges(skeleton_nodes, edges)

        # Count branch points (nodes with >2 connections)
        num_branches = count_branch_points(edges)

        # Calculate branch orders (Strahler numbers)
        branch_order_by_id = calculate_branch_orders(skeleton_nodes, edges, filtered_blocks)

        # Convert branch orders to array order (matching point indices)
        # block_id_to_idx maps block_id -> array index, so we need to create array in that order
        sorted_block_ids = sorted(block_id_to_idx.keys(), key=lambda bid: block_id_to_idx[bid])
        branch_orders_list = [branch_order_by_id.get(bid, 1) for bid in sorted_block_ids]
        max_branch_order = max(branch_orders_list) if branch_orders_list else 0

        # Build block info for response
        block_response = [
            SkeletonBlock(
                block_id=b['id'],
                center=skeleton_nodes[b['id']],
                quantized_level=b['level'],
                num_points=b['size']
            )
            for b in filtered_blocks
            if b['id'] in skeleton_nodes
        ]

        # Convert edges to list format [[from_idx, to_idx], ...] using array indices
        edges_list = []
        for e in edges:
            from_id, to_id = e[0], e[1]
            if from_id in block_id_to_idx and to_id in block_id_to_idx:
                edges_list.append([block_id_to_idx[from_id], block_id_to_idx[to_id]])

        print(f"[BFS Skeleton] Complete: {len(skeleton_nodes)} nodes, {len(edges)} edges, {num_branches} branches, max order: {max_branch_order}")

        return SkeletonResponse(
            success=True,
            skeleton_points=ordered_points,
            skeleton_edges=edges_list,
            branch_orders=branch_orders_list,
            total_length=total_length,
            num_nodes=len(skeleton_nodes),
            num_edges=len(edges),
            num_branches=num_branches,
            max_branch_order=max_branch_order,
            blocks=block_response,
            points_before_filtering=points_original_count,
            points_after_filtering=points_after_filtering,
            num_blocks_before_filter=num_blocks_before,
            num_blocks_after_filter=num_blocks_after,
            dominant_axis=request.dominant_axis,
            num_slices=len(skeleton_nodes)  # Legacy compatibility
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Skeleton extraction failed: {str(e)}")


# ==================== PLANT MODEL GENERATION API ====================
# Uses pyhelios3d PlantArchitecture to generate procedural plant models

class PlantGenerationRequest(BaseModel):
    """Request for generating a plant model"""
    plant_type: str = "bean"  # Plant model name from library
    age: float = 30.0  # Age in days
    # Optional position offset (defaults to origin)
    position_x: float = 0.0
    position_y: float = 0.0
    position_z: float = 0.0
    # Advanced parameters (optional)
    random_seed: Optional[int] = None  # Random seed for reproducibility


class PlantCanopyRequest(BaseModel):
    """Request for generating a canopy of regularly spaced plants"""
    plant_type: str = "bean"  # Plant model name from library
    age: float = 30.0  # Age of all plants in days
    # Cartesian center of the canopy grid (defaults to origin)
    center_x: float = 0.0
    center_y: float = 0.0
    center_z: float = 0.0
    # Spacing between plants (meters) in the x- and y-directions
    spacing_x: float = 0.5
    spacing_y: float = 0.5
    # Number of plants in the x- and y-directions
    count_x: int = 3
    count_y: int = 3
    # Probability (0-1) that each grid position is occupied; 1.0 fills all
    germination_rate: float = 1.0
    # Advanced parameters (optional)
    random_seed: Optional[int] = None  # Random seed for reproducibility


class PlantStreamRequest(BaseModel):
    """Request for streaming plant/canopy generation with progress (SSE).

    `mode` selects single-plant vs. canopy; the relevant subset of fields is
    used for each. Single plants additionally create a retained session so the
    age slider keeps working after generation.
    """
    mode: str = "single"  # "single" | "canopy"
    plant_type: str = "bean"
    age: float = 30.0
    random_seed: Optional[int] = None
    # Single-plant position (mode == "single")
    position_x: float = 0.0
    position_y: float = 0.0
    position_z: float = 0.0
    # Canopy parameters (mode == "canopy")
    center_x: float = 0.0
    center_y: float = 0.0
    center_z: float = 0.0
    spacing_x: float = 0.5
    spacing_y: float = 0.5
    count_x: int = 3
    count_y: int = 3
    germination_rate: float = 1.0


class PlantMaterial(BaseModel):
    """Material definition for plant rendering"""
    name: str
    color: Optional[List[float]] = None  # [r, g, b] ambient/diffuse color (0-1 range)
    texture_name: Optional[str] = None  # Name of texture in textures dict
    has_alpha: bool = False  # Whether texture has alpha transparency


class PlantMaterialGroup(BaseModel):
    """Group of triangles sharing the same material"""
    material_name: str
    triangle_indices: List[int]  # Indices into the main indices array


class PlantGenerationResponse(BaseModel):
    """Response containing generated plant mesh data and Helios XML structure"""
    success: bool
    vertices: List[List[float]]  # [[x, y, z], ...]
    indices: List[List[int]]  # [[v0, v1, v2], ...] triangle indices
    normals: Optional[List[List[float]]] = None  # [[nx, ny, nz], ...]
    colors: Optional[List[List[float]]] = None  # [[r, g, b], ...] vertex colors (0-1 range)
    # Texture support (OBJ export)
    uv_coordinates: Optional[List[List[float]]] = None  # [[u, v], ...] per vertex
    materials: Optional[List[PlantMaterial]] = None  # Material definitions
    material_groups: Optional[List[PlantMaterialGroup]] = None  # Triangle-to-material mapping
    textures: Optional[Dict[str, str]] = None  # {texture_name: base64_png_data}
    # Per-triangle organ-type code (see _ORGAN_LABEL_TO_CODE / ORGAN_SCHEME),
    # parallel to `indices`. Lets a synthetic scan label hits by organ.
    organ_codes: Optional[List[int]] = None
    vertex_count: int
    triangle_count: int
    plant_type: str
    age: float
    height: Optional[float] = None
    available_models: Optional[List[str]] = None
    helios_xml: Optional[str] = None  # Plant structure XML for Helios simulation
    error: Optional[str] = None
    # Session support for age stepping
    session_id: Optional[str] = None  # Session ID for time-stepping
    # Canopy support (echoed back so the frontend can label/store the grid)
    plant_count: Optional[int] = None  # Total plants actually built (after germination)
    count_x: Optional[int] = None      # Grid columns requested
    count_y: Optional[int] = None      # Grid rows requested
    spacing_x: Optional[float] = None  # Spacing used in x (meters)
    spacing_y: Optional[float] = None  # Spacing used in y (meters)


# ==================== PLANT SESSION MANAGEMENT ====================
# Stateful session management for incremental plant growth simulation

import uuid
import threading
from dataclasses import dataclass
from typing import TYPE_CHECKING

# Global storage for active plant sessions
_plant_sessions: Dict[str, Any] = {}
_session_lock = threading.Lock()


@dataclass
class PlantSession:
    """Holds an active pyhelios session for incremental plant growth"""
    session_id: str
    plant_type: str
    plant_id: int
    context: Any  # pyhelios Context
    plantarch: Any  # pyhelios PlantArchitecture
    current_age: float
    position: tuple  # (x, y, z)
    created_at: float  # timestamp
    last_accessed: float = 0.0  # bumped on every read/advance; drives idle eviction


class PlantSessionCreateRequest(BaseModel):
    """Request to create a new plant session"""
    plant_type: str = "bean"
    initial_age: float = 1.0  # Start at age 1 day
    position_x: float = 0.0
    position_y: float = 0.0
    position_z: float = 0.0
    random_seed: Optional[int] = None


class PlantSessionCreateResponse(BaseModel):
    """Response after creating a plant session"""
    success: bool
    session_id: Optional[str] = None
    plant_type: str
    current_age: float
    height: Optional[float] = None
    helios_xml: Optional[str] = None
    error: Optional[str] = None


class PlantSessionAdvanceRequest(BaseModel):
    """Request to advance time on a plant session"""
    dt: float = 1.0  # Days to advance (must be >= 0)


class PlantSessionAdvanceResponse(BaseModel):
    """Response after advancing plant time, includes updated geometry"""
    success: bool
    session_id: str
    previous_age: float
    current_age: float
    height: Optional[float] = None
    # Geometry data
    vertices: List[List[float]]
    indices: List[List[int]]
    colors: Optional[List[List[float]]] = None
    # Texture data (Helios real UVs, V-flipped) — lets the textured renderer
    # show leaf/bark textures for session-generated plants too.
    normals: Optional[List[List[float]]] = None
    uv_coordinates: Optional[List[List[float]]] = None
    materials: Optional[List[PlantMaterial]] = None
    material_groups: Optional[List[PlantMaterialGroup]] = None
    textures: Optional[Dict[str, str]] = None
    organ_codes: Optional[List[int]] = None  # per-triangle organ code, parallel to indices
    vertex_count: int
    triangle_count: int
    error: Optional[str] = None


class PlantSessionStatusResponse(BaseModel):
    """Status of a plant session"""
    success: bool
    session_id: str
    plant_type: str
    current_age: float
    height: Optional[float] = None
    error: Optional[str] = None


# ==================== QSM (Quantitative Structure Model) ====================
# True QSM build: reconstruct a dormant TLS tree as connected cylinders with radii
# + topology, segment continuous shoots, and classify them by shoot rank (trunk=0,
# scaffolds=1, ...). All logic lives in the qsm/ package (pure, unit-tested across
# Layer-1 synthetic + Layer-2 PyHelios fixtures); this endpoint is a thin adapter.


class QSMBuildRequest(BaseModel):
    """Request for a full QSM build. Points come inline (`points`) or from a
    file/octree cloud (`source`), mirroring /api/skeleton/extract.

    For an AGGREGATE build (several pre-registered multi-view scans of ONE
    tree fused into a single QSM), pass `sources`: each is read and the points
    concatenated in world space (each source's own translation applied). This
    is the only way to fuse octree-backed clouds, whose display positions are
    empty client-side. `sources` takes precedence over `points`/`source`."""
    points: Optional[List[List[float]]] = None
    source: Optional[PointSource] = None
    sources: Optional[List[PointSource]] = None
    # Per-species measured twig radius (mm) the radius taper is anchored to as
    # growth length -> 0. Orchard cultivars are absent from rTwig's DB, so this is
    # user-supplied; default 4.23 mm is rTwig's published example.
    twig_radius_mm: float = 4.23
    # Continuation-rule weights (largest-GrowthLength by default). Exposed for
    # experimentation; the validated default is (1, 0, 0).
    w_growthlength: float = 1.0
    w_area: float = 0.0
    w_colinear: float = 0.0


class QSMCylinder(BaseModel):
    """One fitted cylinder of the woody structure."""
    cyl_id: int
    start: List[float]  # [x, y, z] meters
    end: List[float]
    radius: float  # meters
    parent_id: int  # cyl_id of parent, or -1
    shoot_id: int
    rank: int  # topological shoot rank with axis continuation (trunk = 0)
    surf_cov: Optional[float] = None  # surface coverage in [0,1]; low => one-sided
    mad: Optional[float] = None  # mean abs point-to-surface distance, meters


class QSMShoot(BaseModel):
    """A continuous botanical axis (a maximal chain of continuation cylinders)."""
    shoot_id: int
    rank: int
    cylinder_ids: List[int]  # ordered base->tip
    parent_shoot_id: int
    parent_cyl_id: int
    child_shoot_ids: List[int]


class QSMRankMetrics(BaseModel):
    rank: int
    n_shoots: int
    total_length_m: float
    mean_shoot_length_m: float
    woody_volume_m3: float
    mean_diameter_mm: float
    mean_branch_angle_deg: Optional[float] = None


class QSMMetricsResponse(BaseModel):
    tcsa_m2: float
    trunk_diameter_mm: float
    tree_height_m: float
    n_scaffolds: int
    n_shoots_total: int
    max_rank: int
    total_woody_volume_m3: float
    stem_volume_m3: float
    branch_volume_m3: float
    total_length_m: float
    canopy_width_m: float
    canopy_height_m: float
    per_rank: List[QSMRankMetrics]


class QSMBuildResponse(BaseModel):
    success: bool
    cylinders: List[QSMCylinder] = []
    shoots: List[QSMShoot] = []
    metrics: Optional[QSMMetricsResponse] = None
    n_cylinders: int = 0
    n_shoots: int = 0
    points_used: int = 0
    error: Optional[str] = None


def _do_qsm_build(request: QSMBuildRequest, progress=None) -> dict:
    """Build a true QSM from a dormant-tree point cloud, returning the dict form
    of QSMBuildResponse. `progress(fraction, message)` (optional) is called as the
    pipeline advances so the streaming endpoint can surface per-stage progress.

    Pipeline (all in the qsm/ package): geodesic level-set skeleton (B) -> segment
    tree + GrowthLength continuation + SHOOT RANK (C) -> robust IRLS cylinder fit +
    SurfCov (D) -> monotone-taper radius correction (E) -> horticultural metrics.

    The headline output is the per-shoot rank: continuous shoots classified by
    topological branching order with axis continuation (trunk=0, scaffolds=1, ...).
    """
    from qsm.skeleton import extract_skeleton
    from qsm.segments import segments_to_qsm, SegmentOptions
    from qsm.cylinders import fit_qsm_cylinders
    from qsm.radius import correct_radii, RadiusCorrectionOptions
    from qsm.metrics import compute_metrics

    def _report(frac, msg):
        if progress is not None:
            progress(frac, msg)

    try:
        _report(0.05, "Reading points")
        if request.sources:
            # Aggregate: read each source and concatenate in world space, plus any
            # inline `points` (a mixed selection of octree-backed and flat clouds
            # sends both). Each source carries its own translation, so the merged
            # cloud is already registered if the caller aligned the scans first.
            parts = [_read_points_from_source(s)[0] for s in request.sources]
            if request.points:
                parts.append(np.array(request.points, dtype=np.float64))
            parts = [p for p in parts if len(p) > 0]
            points = np.concatenate(parts, axis=0) if parts else np.empty((0, 3), dtype=np.float64)
        elif request.source is not None:
            points, _, _ = _read_points_from_source(request.source)
        else:
            points = np.array(request.points or [], dtype=np.float64)

        if len(points) < 50:
            return QSMBuildResponse(
                success=False, points_used=len(points),
                error="Need at least 50 points to build a QSM",
            ).dict()

        # B: skeleton.
        _report(0.15, "Extracting skeleton")
        graph = extract_skeleton(points)
        if len(graph) == 0:
            return QSMBuildResponse(
                success=False, points_used=len(points),
                error="Skeleton extraction produced no nodes (cloud too sparse "
                      "or disconnected)",
            ).dict()

        # C: segments + shoot rank (HEADLINE).
        _report(0.45, "Segmenting shoots")
        qsm = segments_to_qsm(graph, SegmentOptions(
            w_growthlength=request.w_growthlength,
            w_area=request.w_area,
            w_colinear=request.w_colinear,
        ))

        # D: robust cylinder fit + SurfCov/mad (replaces provisional radii).
        _report(0.65, "Fitting cylinders")
        qsm = fit_qsm_cylinders(qsm, points)

        # E: radius correction (monotone taper + parent cap + twig anchor).
        _report(0.85, "Correcting radii")
        qsm = correct_radii(qsm, RadiusCorrectionOptions(
            twig_radius=request.twig_radius_mm / 1000.0,
        ))

        _report(0.95, "Computing metrics")
        m = compute_metrics(qsm)

        _report(1.0, "Done")
        return QSMBuildResponse(
            success=True,
            points_used=len(points),
            n_cylinders=len(qsm.cylinders),
            n_shoots=len(qsm.shoots),
            cylinders=[
                QSMCylinder(
                    cyl_id=c.cyl_id, start=c.start.tolist(), end=c.end.tolist(),
                    radius=c.radius, parent_id=c.parent_id, shoot_id=c.shoot_id,
                    rank=c.rank, surf_cov=c.surf_cov, mad=c.mad,
                )
                for c in qsm.cylinders
            ],
            shoots=[
                QSMShoot(
                    shoot_id=s.shoot_id, rank=s.rank, cylinder_ids=s.cylinder_ids,
                    parent_shoot_id=s.parent_shoot_id, parent_cyl_id=s.parent_cyl_id,
                    child_shoot_ids=s.child_shoot_ids,
                )
                for s in qsm.shoots
            ],
            metrics=QSMMetricsResponse(
                tcsa_m2=m.tcsa_m2, trunk_diameter_mm=m.trunk_diameter_mm,
                tree_height_m=m.tree_height_m, n_scaffolds=m.n_scaffolds,
                n_shoots_total=m.n_shoots_total, max_rank=m.max_rank,
                total_woody_volume_m3=m.total_woody_volume_m3,
                stem_volume_m3=m.stem_volume_m3, branch_volume_m3=m.branch_volume_m3,
                total_length_m=m.total_length_m, canopy_width_m=m.canopy_width_m,
                canopy_height_m=m.canopy_height_m,
                per_rank=[
                    QSMRankMetrics(
                        rank=pr.rank, n_shoots=pr.n_shoots,
                        total_length_m=pr.total_length_m,
                        mean_shoot_length_m=pr.mean_shoot_length_m,
                        woody_volume_m3=pr.woody_volume_m3,
                        mean_diameter_mm=pr.mean_diameter_mm,
                        mean_branch_angle_deg=pr.mean_branch_angle_deg,
                    )
                    for pr in m.per_rank
                ],
            ),
        ).dict()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return QSMBuildResponse(success=False, error=f"QSM build failed: {e}").dict()


@app.post("/api/qsm/build")
async def build_qsm(request: QSMBuildRequest):
    """Build a true QSM, streaming per-stage progress as PHP1 markers ahead of the
    JSON result (mirrors triangulation / backfill). The renderer's
    fetchJsonWithProgress drains the markers and parses the trailing JSON, which is
    the same QSMBuildResponse shape _do_qsm_build returns."""
    return _bin_frame_streaming_response(
        lambda progress: json.dumps(_do_qsm_build(request, progress)).encode("utf-8")
    )


# ==================== QSM LEAF RECONSTRUCTION (Phase 1) ====================
# Procedurally add leaves to a built QSM: place leaves on terminal shoots
# following the (auto-detected) phyllotaxis, returning a textured mesh that the
# existing TexturedPlantMesh renderer draws. Pure-Python/numpy -- no PyHelios.

class QSMPhyllotaxisRequest(BaseModel):
    """Round-tripped QSM topology for phyllotaxis auto-detection. Only the
    cylinders + shoots are needed (radii/metrics are irrelevant here)."""
    cylinders: List[QSMCylinder]
    shoots: List[QSMShoot]


class QSMPhyllotaxisResponse(BaseModel):
    success: bool
    angle_deg: float = 137.5
    pattern: str = "spiral"          # opposite | spiral | decussate | alternate
    leaves_per_node: int = 1
    confidence: float = 0.0          # [0,1]; 0 when no multi-child parent exists
    n_parents_sampled: int = 0
    error: Optional[str] = None


class QSMLeavesRequest(BaseModel):
    """Add leaves to an existing QSM (cylinders + shoots round-tripped from the
    renderer). Exactly one texture source is used, in precedence order
    obj_path > texture_path > builtin_texture_name."""
    cylinders: List[QSMCylinder]
    shoots: List[QSMShoot]
    leaf_spacing: float = 0.05       # m between nodes along a shoot
    leaf_pitch_deg: float = 45.0     # leaf angle from the shoot axis
    leaf_size_m: float = 0.08        # leaf length
    phyllotaxis_deg: float = 137.5   # azimuth increment between nodes
    leaves_per_node: int = 1
    builtin_texture_name: Optional[str] = None  # e.g. "AlmondLeaf.png"
    texture_path: Optional[str] = None          # uploaded PNG path
    obj_path: Optional[str] = None              # uploaded OBJ path
    max_leaves: int = 200000


class QSMLeavesResponse(BaseModel):
    """Textured leaf mesh. Field names mirror PlantGenerationResponse so the
    frontend's plantResponseToMeshData() consumes it unchanged."""
    success: bool
    vertices: List[List[float]] = []
    indices: List[List[int]] = []
    normals: Optional[List[List[float]]] = None
    uv_coordinates: Optional[List[List[float]]] = None
    materials: Optional[List[PlantMaterial]] = None
    material_groups: Optional[List[PlantMaterialGroup]] = None
    textures: Optional[Dict[str, str]] = None
    vertex_count: int = 0
    triangle_count: int = 0
    leaf_count: int = 0
    error: Optional[str] = None


def _qsm_from_request(cylinders: List[QSMCylinder], shoots: List[QSMShoot]):
    """Rebuild a qsm.model.QSM dataclass from round-tripped request models."""
    from qsm.model import QSM as _QSM, Cylinder as _Cyl, Shoot as _Shoot

    cyls = [
        _Cyl(
            cyl_id=c.cyl_id, start=c.start, end=c.end, radius=c.radius,
            parent_id=c.parent_id, shoot_id=c.shoot_id, rank=c.rank,
            surf_cov=c.surf_cov, mad=c.mad,
        )
        for c in cylinders
    ]
    sh = [
        _Shoot(
            shoot_id=s.shoot_id, rank=s.rank, cylinder_ids=list(s.cylinder_ids),
            parent_shoot_id=s.parent_shoot_id, parent_cyl_id=s.parent_cyl_id,
            child_shoot_ids=list(s.child_shoot_ids),
        )
        for s in shoots
    ]
    return _QSM(cylinders=cyls, shoots=sh)


@app.post("/api/qsm/phyllotaxis", response_model=QSMPhyllotaxisResponse)
async def detect_qsm_phyllotaxis(request: QSMPhyllotaxisRequest):
    """Auto-detect the phyllotactic angle from the QSM's branching geometry.

    Branches follow the phyllotaxis of leaves (modulo unbroken buds), so the
    azimuths of child shoots around each parent reveal the angle. Returns a
    canonical angle + pattern + leaves-per-node + a confidence; used to pre-fill
    the Add Leaves modal."""
    from qsm.leaves import detect_phyllotaxis

    try:
        qsm = _qsm_from_request(request.cylinders, request.shoots)
        d = detect_phyllotaxis(qsm)
        return QSMPhyllotaxisResponse(success=True, **d)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return QSMPhyllotaxisResponse(success=False, error=f"Phyllotaxis detection failed: {e}")


def _resolve_leaf_texture(request):
    """Resolve a leaf request's texture/OBJ source.

    Returns ``(obj_template, texture_name, texture_b64, texture_aspect)`` where
    exactly one of obj_template / (texture_name, texture_b64) is populated.
    Precedence: obj_path > texture_path > builtin_texture_name. Raises
    ``ValueError`` if none is specified. Shared by the place-leaves and
    adjust-leaf-angles endpoints so the two never drift.
    """
    from qsm.leaves import resolve_builtin_texture, read_texture_file
    from qsm.obj_loader import load_obj_template

    if request.obj_path:
        return load_obj_template(Path(request.obj_path)), None, None, 0.6
    if request.texture_path:
        name, b64, aspect = read_texture_file(Path(request.texture_path))
        return None, name, b64, aspect
    if request.builtin_texture_name:
        name, b64, aspect = resolve_builtin_texture(request.builtin_texture_name)
        return None, name, b64, aspect
    raise ValueError("No leaf texture or OBJ specified")


def _leaf_geometry_to_response(geo, placements) -> QSMLeavesResponse:
    """Map a leaf-geometry dict to a QSMLeavesResponse (shared by both endpoints)."""
    if geo.get("error"):
        return QSMLeavesResponse(success=False, error=geo["error"])
    materials = (
        [PlantMaterial(**m) for m in geo["materials"]] if geo.get("materials") else None
    )
    material_groups = (
        [PlantMaterialGroup(**g) for g in geo["material_groups"]]
        if geo.get("material_groups") else None
    )
    return QSMLeavesResponse(
        success=True,
        vertices=geo["vertices"],
        indices=geo["indices"],
        normals=geo.get("normals"),
        uv_coordinates=geo.get("uv_coordinates"),
        materials=materials,
        material_groups=material_groups,
        textures=geo.get("textures"),
        vertex_count=len(geo["vertices"]),
        triangle_count=len(geo["indices"]),
        leaf_count=geo.get("leaf_count", len(placements)),
    )


def _build_leaf_geometry(placements, obj_template, texture_name, texture_b64):
    """Rebuild merged leaf geometry from placements (OBJ instancing or quads)."""
    from qsm.leaves import build_leaf_quad_geometry, build_leaf_obj_geometry

    if obj_template is not None:
        return build_leaf_obj_geometry(placements, obj_template)
    return build_leaf_quad_geometry(placements, texture_name, texture_b64, has_alpha=True)


@app.post("/api/qsm/leaves", response_model=QSMLeavesResponse)
async def add_qsm_leaves(request: QSMLeavesRequest):
    """Place leaves on the terminal shoots of a QSM and return a textured mesh."""
    from qsm.leaves import LeafPlacementOptions, place_leaves

    try:
        qsm = _qsm_from_request(request.cylinders, request.shoots)
        obj_template, texture_name, texture_b64, texture_aspect = _resolve_leaf_texture(request)

        opts = LeafPlacementOptions(
            leaf_spacing=request.leaf_spacing,
            leaf_pitch_deg=request.leaf_pitch_deg,
            leaf_size_m=request.leaf_size_m,
            phyllotaxis_deg=request.phyllotaxis_deg,
            leaves_per_node=request.leaves_per_node,
            texture_aspect=texture_aspect,
            max_leaves=request.max_leaves,
        )
        placements = place_leaves(qsm, opts)
        if not placements:
            return QSMLeavesResponse(
                success=False,
                error="No terminal shoots found to place leaves on",
            )

        geo = _build_leaf_geometry(placements, obj_template, texture_name, texture_b64)
        return _leaf_geometry_to_response(geo, placements)
    except ValueError as e:
        return QSMLeavesResponse(success=False, error=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return QSMLeavesResponse(success=False, error=f"Leaf placement failed: {e}")


@app.get("/api/qsm/leaf-textures")
async def get_qsm_leaf_textures():
    """List the curated built-in leaf textures available for QSM leaf placement."""
    from qsm.leaves import CURATED_LEAF_TEXTURES

    return {"textures": CURATED_LEAF_TEXTURES}


# ==================== QSM LEAF-ANGLE ADJUSTMENT (Phase 2) ====================
# Rotate procedurally-placed leaves so each voxel cell's leaf-angle distribution
# matches a target measured from a leaf-on Helios triangulation, via per-cell
# optimal assignment (Helios setPlantLeafAngleDistribution port). Pure numpy.

class QSMGrid(BaseModel):
    """The voxel grid the triangulation was built in (full extents + subdivisions)."""
    center: List[float]  # [x, y, z]
    size: List[float]    # [x, y, z] full extents
    nx: int = 1
    ny: int = 1
    nz: int = 1


class QSMTriangulationInput(BaseModel):
    """A leaf-on Helios triangulation overlapping the QSM, from the renderer."""
    vertices: List[float]            # flat x,y,z
    indices: List[int]               # flat triangle vertex indices
    triangle_cell_ids: List[int]     # per-triangle grid cell (row-major; -1/0xffffffff = outside)
    triangle_scan_ids: Optional[List[int]] = None  # per-triangle source scan (azimuth orientation)
    scan_origins: Optional[List[float]] = None      # flat per-scan x,y,z origins
    grid: QSMGrid


class QSMCellTarget(BaseModel):
    """A precomputed per-cell leaf-angle target (escape hatch / test injection)."""
    cell_id: int
    beta_mu: float
    beta_nu: float
    ecc: float
    phi0_deg: float
    n_measured: int = 0


class QSMAdjustLeafAnglesRequest(QSMLeavesRequest):
    """Re-place the QSM's leaves (same Phase-1 params) then adjust their angles to
    a measured per-cell distribution. Exactly one of `triangulation` / `cell_targets`
    must be present (with `grid` when using `cell_targets`)."""
    triangulation: Optional[QSMTriangulationInput] = None
    cell_targets: Optional[List[QSMCellTarget]] = None
    grid: Optional[QSMGrid] = None
    seed: int = 0
    max_cell_leaves: Optional[int] = None


@app.post("/api/qsm/adjust-leaf-angles", response_model=QSMLeavesResponse)
async def adjust_qsm_leaf_angles(request: QSMAdjustLeafAnglesRequest):
    """Adjust a QSM's leaf angles to match a measured per-cell distribution."""
    import numpy as np
    from qsm.leaves import LeafPlacementOptions, place_leaves
    from qsm.leaf_angles import CellTarget, adjust_placements_to_distribution
    from qsm.leaf_distribution import compute_cell_targets

    try:
        if request.triangulation is None and not request.cell_targets:
            return QSMLeavesResponse(
                success=False,
                error="Provide either a triangulation or precomputed cell_targets",
            )

        qsm = _qsm_from_request(request.cylinders, request.shoots)
        obj_template, texture_name, texture_b64, texture_aspect = _resolve_leaf_texture(request)

        # Re-place leaves identically to Phase 1 so bases/orientations match.
        opts = LeafPlacementOptions(
            leaf_spacing=request.leaf_spacing,
            leaf_pitch_deg=request.leaf_pitch_deg,
            leaf_size_m=request.leaf_size_m,
            phyllotaxis_deg=request.phyllotaxis_deg,
            leaves_per_node=request.leaves_per_node,
            texture_aspect=texture_aspect,
            max_leaves=request.max_leaves,
        )
        placements = place_leaves(qsm, opts)
        if not placements:
            return QSMLeavesResponse(
                success=False,
                error="No leaves to adjust — add leaves to this QSM first",
            )

        # Resolve the grid + per-cell targets.
        if request.triangulation is not None:
            tin = request.triangulation
            g = tin.grid
            grid = (np.array(g.center, dtype=np.float64), np.array(g.size, dtype=np.float64),
                    g.nx, g.ny, g.nz)
            verts = np.array(tin.vertices, dtype=np.float64).reshape(-1, 3)
            tris = np.array(tin.indices, dtype=np.int64).reshape(-1, 3)
            cell_ids = np.array(tin.triangle_cell_ids, dtype=np.int64)
            scan_ids = (np.array(tin.triangle_scan_ids, dtype=np.int64)
                        if tin.triangle_scan_ids else None)
            scan_origins = (np.array(tin.scan_origins, dtype=np.float64).reshape(-1, 3)
                            if tin.scan_origins else None)
            cell_targets = compute_cell_targets(
                verts, tris, cell_ids, grid,
                triangle_scan_ids=scan_ids, scan_origins=scan_origins,
            )
        else:
            if request.grid is None:
                return QSMLeavesResponse(
                    success=False, error="cell_targets requires a grid",
                )
            g = request.grid
            grid = (np.array(g.center, dtype=np.float64), np.array(g.size, dtype=np.float64),
                    g.nx, g.ny, g.nz)
            cell_targets = {
                ct.cell_id: CellTarget(
                    beta_mu=ct.beta_mu, beta_nu=ct.beta_nu, ecc=ct.ecc,
                    phi0_deg=ct.phi0_deg, n_measured=ct.n_measured,
                )
                for ct in request.cell_targets
            }

        adjusted = adjust_placements_to_distribution(
            placements, cell_targets, grid,
            seed=request.seed, max_cell_leaves=request.max_cell_leaves,
        )

        geo = _build_leaf_geometry(adjusted, obj_template, texture_name, texture_b64)
        return _leaf_geometry_to_response(geo, adjusted)
    except ValueError as e:
        return QSMLeavesResponse(success=False, error=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return QSMLeavesResponse(success=False, error=f"Leaf-angle adjustment failed: {e}")


@app.get("/api/plant/models")
async def get_available_plant_models():
    """Get list of available plant models from pyhelios library"""
    try:
        from pyhelios import Context, PlantArchitecture

        with Context() as context:
            with PlantArchitecture(context) as plantarch:
                models = plantarch.getAvailablePlantModels()
                return {"success": True, "models": models}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to get plant models: {str(e)}")


# ==================== PLANT SESSION ENDPOINTS ====================

# Organ-type labels Helios writes as the "object_label" primitive-data string on
# plant-architecture geometry, mapped to small int codes that survive the synthetic
# scan's numeric per-hit sampling (getHitDataArray is float32; a string label can't
# ride it). MIRROR of ORGAN_SCHEME in src/renderer/lib/classification.ts — keep the
# codes in sync. Unknown/absent labels collapse to 0.
_ORGAN_LABEL_TO_CODE = {
    "unknown": 0,
    "leaf": 1,
    "petiole": 2,
    "shoot": 3,       # internode tube primitives
    "peduncle": 4,
    "fruit": 5,
    "petiolule": 6,
}


def _organ_code_for_primitive(context, uuid) -> int:
    """Read a primitive's Helios ``object_label`` organ tag and map it to a small
    int code (see ``_ORGAN_LABEL_TO_CODE``). Returns 0 (unknown) when the tag is
    absent — e.g. imported meshes or organs the model didn't label."""
    try:
        if context.doesPrimitiveDataExist(uuid, "object_label"):
            return _ORGAN_LABEL_TO_CODE.get(
                context.getPrimitiveData(uuid, "object_label", str), 0
            )
    except Exception:
        pass
    return 0


def _extract_session_geometry(session: PlantSession) -> tuple:
    """
    Extract geometry from an active plant session.

    Returns (vertices, indices, colors, vertex_count, triangle_count, textures)
    where `textures` is a dict with the optional texture payload:
      {"normals", "uvs", "materials", "material_groups", "textures"}
    UVs are Helios's real per-vertex coordinates, V-flipped for three.js — the
    same convention as /api/plant/generate, so the textured renderer handles
    plants from either path identically.
    """
    import os
    import base64

    context = session.context
    plantarch = session.plantarch

    # Woody species get brown stems
    WOODY_SPECIES = {
        'almond', 'apple', 'apple_fruitingwall', 'easternredbud', 'olive',
        'pistachio', 'walnut', 'bougainvillea', 'grapevine_VSP', 'grapevine_Wye',
        'grapevine_GDC', 'grapevine_geneva_double_curtain', 'grapevine_vertical_shoot_positioned',
        'grapevine_sprawl', 'grapevine_unilateral_cordon'
    }
    is_woody = session.plant_type in WOODY_SPECIES

    # Extract geometry by object
    object_ids = context.getAllObjectIDs()

    vertices = []
    colors = []
    normals = []
    uvs = []
    faces = []
    organ_codes = []  # per-triangle organ-type code, parallel to faces
    vertex_index = 0
    triangle_index = 0

    materials_dict = {}        # mat_label -> PlantMaterial
    material_groups_dict = {}  # mat_label -> [triangle index, ...]
    texture_files = {}         # texture basename -> source path

    for obj_id in object_ids:
        try:
            prim_infos = context.getPrimitivesInfoForObject(obj_id)
        except:
            continue

        if not prim_infos:
            continue

        # Check first primitive for material/texture info
        first_prim = prim_infos[0]
        try:
            mat_label = context.getPrimitiveMaterialLabel(first_prim.uuid)
            texture_path = context.getMaterialTexture(mat_label) if mat_label else None
        except:
            mat_label = None
            texture_path = None

        # Organ type is constant within an object (a leaf object's primitives are
        # all leaves); read it once and broadcast to this object's triangles.
        organ_code = _organ_code_for_primitive(context, first_prim.uuid)

        is_textured = texture_path and len(texture_path) > 0
        texture_name_lower = texture_path.lower() if texture_path else ""
        is_leaf_texture = 'leaf' in texture_name_lower or 'Leaf' in (texture_path or "")

        if is_textured:
            texture_name = os.path.basename(texture_path)
            if mat_label not in materials_dict:
                materials_dict[mat_label] = PlantMaterial(
                    name=mat_label,
                    texture_name=texture_name,
                    has_alpha=True,
                )
                material_groups_dict[mat_label] = []
                texture_files[texture_name] = texture_path

        for prim_info in prim_infos:
            try:
                if prim_info.primitive_type.value != 1:
                    continue

                tri_verts = prim_info.vertices
                if len(tri_verts) != 3:
                    continue

                prim_color = prim_info.color
                color_rgb = [prim_color.r, prim_color.g, prim_color.b]

                # Color assignment logic (vertex-color fallback)
                if is_textured and (color_rgb[0] == 0 and color_rgb[1] == 0 and color_rgb[2] == 0):
                    if is_leaf_texture:
                        color_rgb = [0.3, 0.55, 0.2]
                    elif is_woody:
                        color_rgb = [0.45, 0.3, 0.15]
                    else:
                        color_rgb = [0.3, 0.55, 0.2]
                elif not is_textured and is_woody:
                    brightness = color_rgb[0] + color_rgb[1] + color_rgb[2]
                    if brightness < 0.5:
                        color_rgb = [0.45, 0.3, 0.15]

                prim_normal = prim_info.normal
                normal_xyz = [prim_normal.x, prim_normal.y, prim_normal.z]

                # Real Helios UVs (V-flipped). A textured primitive without UVs
                # falls back to vertex color and is left out of the group.
                prim_uvs = prim_info.texture_uv if is_textured else None
                tri_is_textured = (
                    is_textured and prim_uvs is not None and len(prim_uvs) == 3
                )

                face_indices = []
                for vi, v in enumerate(tri_verts):
                    vertices.append([v.x, v.y, v.z])
                    colors.append(color_rgb)
                    normals.append(normal_xyz)
                    if tri_is_textured:
                        uv = prim_uvs[vi]
                        uvs.append([uv.x, 1.0 - uv.y])
                    else:
                        uvs.append([0.0, 0.0])
                    face_indices.append(vertex_index)
                    vertex_index += 1

                faces.append(face_indices)
                organ_codes.append(organ_code)

                if tri_is_textured and mat_label in material_groups_dict:
                    material_groups_dict[mat_label].append(triangle_index)

                triangle_index += 1
            except:
                continue

    # Load textures as base64.
    textures_data = {}
    for tex_name, tex_path in texture_files.items():
        if os.path.exists(tex_path):
            try:
                with open(tex_path, 'rb') as f:
                    textures_data[tex_name] = base64.b64encode(f.read()).decode('utf-8')
            except Exception:
                pass

    materials_list = list(materials_dict.values()) if materials_dict else None
    material_groups_list = [
        PlantMaterialGroup(material_name=name, triangle_indices=tris)
        for name, tris in material_groups_dict.items() if tris
    ] if material_groups_dict else None

    texture_payload = {
        "normals": normals if normals else None,
        "uvs": uvs if any(u != [0.0, 0.0] for u in uvs) else None,
        "materials": materials_list,
        "material_groups": material_groups_list,
        "textures": textures_data if textures_data else None,
        "organ_codes": organ_codes if organ_codes else None,
    }

    return vertices, faces, colors, len(vertices), len(faces), texture_payload


@app.post("/api/plant/session/create", response_model=PlantSessionCreateResponse)
async def create_plant_session(request: PlantSessionCreateRequest):
    """
    Create a new plant session for incremental growth simulation.
    The session keeps the pyhelios context alive for efficient time stepping.
    """
    import time

    try:
        from pyhelios import Context, PlantArchitecture
        from pyhelios.wrappers.DataTypes import vec3

        # Create session ID
        session_id = str(uuid.uuid4())[:8]

        print(f"[Plant Session] Creating session {session_id}: type={request.plant_type}, initial_age={request.initial_age}")

        # Create context and plant architecture (NOT using 'with' - we keep them alive)
        context = Context()
        context.__enter__()

        plantarch = PlantArchitecture(context)
        plantarch.__enter__()

        # Get available models for validation
        available_models = plantarch.getAvailablePlantModels()

        if request.plant_type not in available_models:
            # Clean up on error
            plantarch.__exit__(None, None, None)
            context.__exit__(None, None, None)
            return PlantSessionCreateResponse(
                success=False,
                plant_type=request.plant_type,
                current_age=0,
                error=f"Unknown plant type '{request.plant_type}'"
            )

        # Load and build plant
        plantarch.loadPlantModelFromLibrary(request.plant_type)

        build_params = {}
        if request.random_seed is not None:
            build_params['random_seed'] = request.random_seed

        plant_id = plantarch.buildPlantInstanceFromLibrary(
            base_position=vec3(request.position_x, request.position_y, request.position_z),
            age=request.initial_age,
            build_parameters=build_params if build_params else None
        )

        height = plantarch.getPlantHeight(plant_id)
        current_age = plantarch.getPlantAge(plant_id)

        # Create and store session
        session = PlantSession(
            session_id=session_id,
            plant_type=request.plant_type,
            plant_id=plant_id,
            context=context,
            plantarch=plantarch,
            current_age=current_age,
            position=(request.position_x, request.position_y, request.position_z),
            created_at=time.time(),
            last_accessed=time.time(),
        )

        # Write plant structure XML for morph tool
        helios_xml = None
        try:
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.xml', delete=False) as tmp_xml:
                xml_path = tmp_xml.name
            plantarch.writePlantStructureXML(plant_id, xml_path)
            with open(xml_path, 'r') as f:
                helios_xml = f.read()
            import os
            os.unlink(xml_path)
        except Exception as xml_err:
            print(f"[Plant Session] Warning: failed to write XML: {xml_err}")

        _sweep_plant_sessions()
        with _session_lock:
            _plant_sessions[session_id] = session

        print(f"[Plant Session] Created session {session_id}: age={current_age:.1f}, height={height:.3f}m")

        return PlantSessionCreateResponse(
            success=True,
            session_id=session_id,
            plant_type=request.plant_type,
            current_age=current_age,
            height=height,
            helios_xml=helios_xml,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return PlantSessionCreateResponse(
            success=False,
            plant_type=request.plant_type,
            current_age=0,
            error=str(e)
        )


@app.post("/api/plant/session/{session_id}/advance", response_model=PlantSessionAdvanceResponse)
async def advance_plant_session(session_id: str, request: PlantSessionAdvanceRequest):
    """
    Advance time for a plant session and return updated geometry.
    """
    try:
        with _session_lock:
            if session_id not in _plant_sessions:
                raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
            session = _plant_sessions[session_id]
            session.last_accessed = time.time()

        if request.dt < 0:
            raise HTTPException(status_code=400, detail="dt must be >= 0")

        previous_age = session.current_age

        # Advance time
        if request.dt > 0:
            session.plantarch.advanceTime(request.dt)

        # Get updated state
        current_age = session.plantarch.getPlantAge(session.plant_id)
        height = session.plantarch.getPlantHeight(session.plant_id)
        session.current_age = current_age

        # Extract geometry (+ real Helios UVs / materials / textures)
        vertices, faces, colors, vertex_count, triangle_count, tex = _extract_session_geometry(session)

        print(f"[Plant Session] Advanced {session_id}: {previous_age:.1f} -> {current_age:.1f} days, {vertex_count} vertices")

        return PlantSessionAdvanceResponse(
            success=True,
            session_id=session_id,
            previous_age=previous_age,
            current_age=current_age,
            height=height,
            vertices=vertices,
            indices=faces,
            colors=colors,
            normals=tex["normals"],
            uv_coordinates=tex["uvs"],
            materials=tex["materials"],
            material_groups=tex["material_groups"],
            textures=tex["textures"],
            organ_codes=tex.get("organ_codes"),
            vertex_count=vertex_count,
            triangle_count=triangle_count
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        return PlantSessionAdvanceResponse(
            success=False,
            session_id=session_id,
            previous_age=0,
            current_age=0,
            vertices=[],
            indices=[],
            vertex_count=0,
            triangle_count=0,
            error=str(e)
        )


@app.get("/api/plant/session/{session_id}", response_model=PlantSessionStatusResponse)
async def get_plant_session_status(session_id: str):
    """Get the current status of a plant session."""
    try:
        with _session_lock:
            if session_id not in _plant_sessions:
                raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
            session = _plant_sessions[session_id]
            session.last_accessed = time.time()

        height = session.plantarch.getPlantHeight(session.plant_id)

        return PlantSessionStatusResponse(
            success=True,
            session_id=session_id,
            plant_type=session.plant_type,
            current_age=session.current_age,
            height=height
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        return PlantSessionStatusResponse(
            success=False,
            session_id=session_id,
            plant_type="",
            current_age=0,
            error=str(e)
        )


@app.delete("/api/plant/session/{session_id}")
async def delete_plant_session(session_id: str):
    """Delete a plant session and free resources."""
    try:
        with _session_lock:
            if session_id not in _plant_sessions:
                raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
            session = _plant_sessions.pop(session_id)

        # Clean up pyhelios resources
        try:
            session.plantarch.__exit__(None, None, None)
            session.context.__exit__(None, None, None)
        except:
            pass

        print(f"[Plant Session] Deleted session {session_id}")

        return {"success": True, "session_id": session_id}

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/plant/sessions")
async def list_plant_sessions():
    """List all active plant sessions."""
    with _session_lock:
        sessions = []
        for sid, session in _plant_sessions.items():
            sessions.append({
                "session_id": sid,
                "plant_type": session.plant_type,
                "current_age": session.current_age,
                "created_at": session.created_at
            })
        return {"success": True, "sessions": sessions, "count": len(sessions)}


# ==================== PLANT MORPH ENDPOINTS ====================
# Uses the plant structure XML roundtrip: parse existing XML to extract
# per-phytomer parameters, allow tuning, then rebuild via readPlantStructureXML.


def _parse_plant_xml(xml_string: str) -> dict:
    """
    Parse a Helios plant structure XML string into a structured dict.
    Returns { plant_age, base_position, shoots: [...] } where each shoot
    has shoot_type_label and a list of phytomers with internode/petiole/leaf params.
    """
    import xml.etree.ElementTree as ET

    root = ET.fromstring(xml_string)
    plant = root.find('plant_instance')
    if plant is None:
        raise ValueError("No <plant_instance> found in XML")

    result = {
        "plant_age": float(plant.findtext('plant_age', '0').strip()),
        "base_position": plant.findtext('base_position', '0 0 0').strip(),
        "shoots": [],
    }

    for shoot_el in plant.findall('shoot'):
        shoot = {
            "shoot_id": int(shoot_el.get('ID', '-1')),
            "shoot_type_label": (shoot_el.findtext('shoot_type_label') or 'unknown').strip(),
            "parent_shoot_id": int((shoot_el.findtext('parent_shoot_ID') or '-1').strip()),
            "parent_node_index": int((shoot_el.findtext('parent_node_index') or '0').strip()),
            "parent_petiole_index": int((shoot_el.findtext('parent_petiole_index') or '0').strip()),
            "base_rotation": (shoot_el.findtext('base_rotation') or '0 0 0').strip(),
            "phytomers": [],
        }

        for phytomer_el in shoot_el.findall('phytomer'):
            internode_el = phytomer_el.find('internode')
            if internode_el is None:
                continue

            internode = {}
            for child in internode_el:
                if child.tag == 'petiole':
                    continue  # handled separately
                internode[child.tag] = child.text.strip() if child.text else ''

            petioles = []
            for petiole_el in internode_el.findall('petiole'):
                petiole = {}
                leaves = []
                for child in petiole_el:
                    if child.tag == 'leaf':
                        leaf = {}
                        for lc in child:
                            leaf[lc.tag] = lc.text.strip() if lc.text else ''
                        leaves.append(leaf)
                    else:
                        petiole[child.tag] = child.text.strip() if child.text else ''
                petiole['leaves'] = leaves
                petioles.append(petiole)

            shoot["phytomers"].append({
                "internode": internode,
                "petioles": petioles,
            })

        result["shoots"].append(shoot)

    return result


def _build_plant_xml(parsed: dict) -> str:
    """
    Rebuild a Helios plant structure XML string from a parsed dict.
    Inverse of _parse_plant_xml.
    """
    import xml.etree.ElementTree as ET

    root = ET.Element('helios')
    plant = ET.SubElement(root, 'plant_instance', ID='0')
    ET.SubElement(plant, 'base_position').text = f' {parsed["base_position"]} '
    ET.SubElement(plant, 'plant_age').text = f' {parsed["plant_age"]} '

    for shoot in parsed['shoots']:
        shoot_el = ET.SubElement(plant, 'shoot', ID=str(shoot['shoot_id']))
        ET.SubElement(shoot_el, 'shoot_type_label').text = f' {shoot["shoot_type_label"]} '
        ET.SubElement(shoot_el, 'parent_shoot_ID').text = f' {shoot["parent_shoot_id"]} '
        ET.SubElement(shoot_el, 'parent_node_index').text = f' {shoot["parent_node_index"]} '
        ET.SubElement(shoot_el, 'parent_petiole_index').text = f' {shoot["parent_petiole_index"]} '
        ET.SubElement(shoot_el, 'base_rotation').text = f' {shoot["base_rotation"]} '

        for phytomer in shoot['phytomers']:
            phytomer_el = ET.SubElement(shoot_el, 'phytomer')
            internode_el = ET.SubElement(phytomer_el, 'internode')

            for key, val in phytomer['internode'].items():
                ET.SubElement(internode_el, key).text = val

            for petiole in phytomer['petioles']:
                petiole_el = ET.SubElement(internode_el, 'petiole')
                for key, val in petiole.items():
                    if key == 'leaves':
                        for leaf in val:
                            leaf_el = ET.SubElement(petiole_el, 'leaf')
                            for lk, lv in leaf.items():
                                ET.SubElement(leaf_el, lk).text = lv
                    else:
                        ET.SubElement(petiole_el, key).text = val

    ET.indent(root, space='\t')
    return '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(root, encoding='unicode')


def _get_distribution_params(plant_type: str, xml_string: str) -> dict:
    """
    Extract distribution-level shoot parameters from the Helios plant model library.
    Creates a disposable context, loads the model, builds a minimal instance to register
    shoot types, then calls getCurrentShootParameters for each unique shoot type label
    found in the XML.

    Returns { shoot_type_label: { param_name: { distribution, parameters } | bool } }
    Falls back to {} on any failure.
    """
    try:
        import xml.etree.ElementTree as ET
        from pyhelios import Context, PlantArchitecture

        # Extract unique shoot type labels from the XML
        root = ET.fromstring(xml_string)
        plant = root.find('plant_instance')
        if plant is None:
            return {}

        labels = set()
        for shoot_el in plant.findall('shoot'):
            label = (shoot_el.findtext('shoot_type_label') or '').strip()
            if label:
                labels.add(label)

        if not labels:
            return {}

        # Create disposable context to query parameters
        context = Context()
        context.__enter__()
        plantarch = PlantArchitecture(context)
        plantarch.__enter__()

        try:
            plantarch.loadPlantModelFromLibrary(plant_type)

            # Build a minimal instance so shoot types get registered
            from pyhelios.wrappers.DataTypes import vec3
            plantarch.buildPlantInstanceFromLibrary(
                base_position=vec3(0, 0, 0),
                age=1
            )

            result = {}
            for label in sorted(labels):
                try:
                    params = plantarch.getCurrentShootParameters(label)
                    result[label] = params
                except Exception as e:
                    print(f"[Morph] Warning: getCurrentShootParameters('{label}') failed: {e}")
                    continue

            return result
        finally:
            plantarch.__exit__(None, None, None)
            context.__exit__(None, None, None)

    except Exception as e:
        print(f"[Morph] Warning: _get_distribution_params failed: {e}")
        return {}


class PlantMorphParseRequest(BaseModel):
    """Request to parse plant XML into editable parameters"""
    helios_xml: str
    plant_type: str

class PlantMorphParseResponse(BaseModel):
    """Parsed plant structure for the morph UI"""
    success: bool
    plant_type: str
    plant_age: float = 0
    base_position: str = "0 0 0"
    shoots: List[dict] = []
    distribution_params: Dict[str, dict] = {}
    error: Optional[str] = None

class PlantMorphRequest(BaseModel):
    """Request to morph/regrow a plant from modified XML"""
    plant_type: str
    helios_xml: str  # The modified plant structure XML
    error: Optional[str] = None

class PlantMorphResponse(BaseModel):
    """Response from POST morph"""
    success: bool
    session_id: Optional[str] = None
    vertices: List[List[float]] = []
    indices: List[List[int]] = []
    colors: Optional[List[List[float]]] = None
    # Texture data so a morphed plant stays textured.
    normals: Optional[List[List[float]]] = None
    uv_coordinates: Optional[List[List[float]]] = None
    materials: Optional[List[PlantMaterial]] = None
    material_groups: Optional[List[PlantMaterialGroup]] = None
    textures: Optional[Dict[str, str]] = None
    organ_codes: Optional[List[int]] = None  # per-triangle organ code, parallel to indices
    vertex_count: int = 0
    triangle_count: int = 0
    current_age: float = 0
    height: Optional[float] = None
    helios_xml: Optional[str] = None  # Updated XML from the rebuilt plant
    error: Optional[str] = None


@app.post("/api/plant/morph/parse", response_model=PlantMorphParseResponse)
async def parse_plant_morph_parameters(request: PlantMorphParseRequest):
    """Parse a plant structure XML string into editable parameters."""
    try:
        parsed = _parse_plant_xml(request.helios_xml)

        # Get distribution-level parameters (non-fatal)
        dist_params = {}
        try:
            dist_params = _get_distribution_params(request.plant_type, request.helios_xml)
        except Exception as dp_err:
            print(f"[Morph] Warning: distribution params extraction failed: {dp_err}")

        return PlantMorphParseResponse(
            success=True,
            plant_type=request.plant_type,
            plant_age=parsed['plant_age'],
            base_position=parsed['base_position'],
            shoots=parsed['shoots'],
            distribution_params=dist_params,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return PlantMorphParseResponse(
            success=False,
            plant_type=request.plant_type,
            error=str(e),
        )


@app.post("/api/plant/morph", response_model=PlantMorphResponse)
async def morph_plant(request: PlantMorphRequest):
    """Rebuild a plant from modified structure XML."""
    import time
    import tempfile
    import os

    try:
        from pyhelios import Context, PlantArchitecture

        session_id = str(uuid.uuid4())[:8]
        print(f"[Morph] Rebuilding plant {session_id}: type={request.plant_type}")

        context = Context()
        context.__enter__()
        plantarch = PlantArchitecture(context)
        plantarch.__enter__()

        # Load the plant model library so shoot types are defined
        plantarch.loadPlantModelFromLibrary(request.plant_type)

        # Write the modified XML to a temp file and read it back
        with tempfile.NamedTemporaryFile(suffix='.xml', delete=False, mode='w') as tmp:
            tmp.write(request.helios_xml)
            xml_path = tmp.name

        try:
            plant_ids = plantarch.readPlantStructureXML(xml_path)
        finally:
            os.unlink(xml_path)

        if not plant_ids:
            plantarch.__exit__(None, None, None)
            context.__exit__(None, None, None)
            return PlantMorphResponse(
                success=False,
                error="readPlantStructureXML returned no plant IDs"
            )

        plant_id = plant_ids[0]
        current_age = plantarch.getPlantAge(plant_id)
        height = plantarch.getPlantHeight(plant_id)

        # Parse base_position from the XML for session storage
        import xml.etree.ElementTree as ET
        xml_root = ET.fromstring(request.helios_xml)
        pos_text = (xml_root.find('plant_instance').findtext('base_position') or '0 0 0').strip()
        pos_parts = [float(x) for x in pos_text.split()]
        position = (pos_parts[0] if len(pos_parts) > 0 else 0,
                     pos_parts[1] if len(pos_parts) > 1 else 0,
                     pos_parts[2] if len(pos_parts) > 2 else 0)

        # Create session
        session = PlantSession(
            session_id=session_id,
            plant_type=request.plant_type,
            plant_id=plant_id,
            context=context,
            plantarch=plantarch,
            current_age=current_age,
            position=position,
            created_at=time.time(),
            last_accessed=time.time(),
        )

        # Extract geometry (+ real Helios UVs / materials / textures)
        vertices, faces, colors, vertex_count, triangle_count, tex = _extract_session_geometry(session)

        # Write new XML from rebuilt plant (non-fatal if it fails)
        new_helios_xml = request.helios_xml  # fallback to input XML
        try:
            with tempfile.NamedTemporaryFile(suffix='.xml', delete=False, mode='r') as tmp:
                new_xml_path = tmp.name
            plantarch.writePlantStructureXML(plant_id, new_xml_path)
            with open(new_xml_path, 'r') as f:
                new_helios_xml = f.read()
            os.unlink(new_xml_path)
        except Exception as xml_err:
            print(f"[Morph] Warning: failed to write new XML: {xml_err}")

        # Store session
        with _session_lock:
            _plant_sessions[session_id] = session
        _sweep_plant_sessions()

        print(f"[Morph] Rebuilt plant {session_id}: age={current_age:.1f}, height={height:.3f}m, {vertex_count} vertices")

        return PlantMorphResponse(
            success=True,
            session_id=session_id,
            vertices=vertices,
            indices=faces,
            colors=colors,
            normals=tex["normals"],
            uv_coordinates=tex["uvs"],
            materials=tex["materials"],
            material_groups=tex["material_groups"],
            textures=tex["textures"],
            organ_codes=tex.get("organ_codes"),
            vertex_count=vertex_count,
            triangle_count=triangle_count,
            current_age=current_age,
            height=height,
            helios_xml=new_helios_xml,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            plantarch.__exit__(None, None, None)
            context.__exit__(None, None, None)
        except:
            pass
        return PlantMorphResponse(
            success=False,
            error=str(e),
        )


@app.post("/api/plant/generate", response_model=PlantGenerationResponse)
async def generate_plant_model(request: PlantGenerationRequest):
    """
    Generate a procedural plant model using pyhelios PlantArchitecture.

    Uses direct primitive extraction from pyhelios Context API to get valid geometry.
    Textured organs carry Helios's own per-vertex UV coordinates (V-flipped for
    three.js), so leaf textures sample the correct cell of the leaf atlas.
    """
    import tempfile
    import os

    # Woody species get brown stems, herbaceous species get green stems
    WOODY_SPECIES = {
        'almond', 'apple', 'apple_fruitingwall', 'easternredbud', 'olive',
        'pistachio', 'walnut', 'bougainvillea', 'grapevine_VSP', 'grapevine_Wye',
        'grapevine_GDC', 'grapevine_geneva_double_curtain', 'grapevine_vertical_shoot_positioned',
        'grapevine_sprawl', 'grapevine_unilateral_cordon'
    }

    try:
        from pyhelios import Context, PlantArchitecture
        from pyhelios.wrappers.DataTypes import vec3

        print(f"[Plant Generation] Starting: type={request.plant_type}, age={request.age}")
        is_woody = request.plant_type in WOODY_SPECIES

        # Create context and plant architecture
        with Context() as context:
            with PlantArchitecture(context) as plantarch:
                # Get available models for reference
                available_models = plantarch.getAvailablePlantModels()

                # Validate plant type
                if request.plant_type not in available_models:
                    return PlantGenerationResponse(
                        success=False,
                        vertices=[],
                        indices=[],
                        vertex_count=0,
                        triangle_count=0,
                        plant_type=request.plant_type,
                        age=request.age,
                        available_models=available_models,
                        error=f"Unknown plant type '{request.plant_type}'. Available: {', '.join(available_models[:10])}..."
                    )

                # Load plant model
                plantarch.loadPlantModelFromLibrary(request.plant_type)

                # Build plant instance with optional build_parameters
                build_params = {}
                if request.random_seed is not None:
                    build_params['random_seed'] = request.random_seed
                    print(f"[Plant Generation] Using random_seed={request.random_seed} for reproducibility")

                plant_id = plantarch.buildPlantInstanceFromLibrary(
                    base_position=vec3(request.position_x, request.position_y, request.position_z),
                    age=request.age,
                    build_parameters=build_params if build_params else None
                )

                # Get plant height
                height = plantarch.getPlantHeight(plant_id)
                print(f"[Plant Generation] Built plant: type={request.plant_type}, age={request.age}, height={height:.3f}m, seed={request.random_seed}")

                # Extract geometry by object to properly handle textures
                # Each object represents a leaf or stem section
                object_ids = context.getAllObjectIDs()
                print(f"[Plant Generation] Extracting {len(object_ids)} objects...")

                # Build mesh data
                vertices = []  # Flat list of vertex positions
                colors = []    # Per-vertex colors (RGB 0-1)
                normals = []   # Per-vertex normals
                uvs = []       # Per-vertex UV coordinates (for textured objects)
                faces = []     # Triangle indices (for compatibility)
                organ_codes = []  # Per-triangle organ-type code, parallel to faces

                # Track materials and their triangles
                materials_dict = {}  # material_name -> PlantMaterial
                material_groups_dict = {}  # material_name -> list of triangle indices
                texture_files = {}  # texture_name -> file_path

                vertex_index = 0
                triangle_index = 0

                for obj_id in object_ids:
                    # Get all primitive info for this object directly
                    try:
                        prim_infos = context.getPrimitivesInfoForObject(obj_id)
                    except:
                        continue

                    if not prim_infos:
                        continue

                    # Check first primitive for material/texture info
                    first_prim = prim_infos[0]
                    try:
                        mat_label = context.getPrimitiveMaterialLabel(first_prim.uuid)
                        texture_path = context.getMaterialTexture(mat_label) if mat_label else None
                    except:
                        mat_label = None
                        texture_path = None

                    # Organ type is constant within an object; read once, broadcast per triangle.
                    organ_code = _organ_code_for_primitive(context, first_prim.uuid)

                    # Determine if this object is textured
                    is_textured = texture_path and len(texture_path) > 0

                    # For woody species, determine if texture is for leaves vs bark/branches
                    texture_name_lower = texture_path.lower() if texture_path else ""
                    is_leaf_texture = 'leaf' in texture_name_lower or 'Leaf' in (texture_path or "")

                    # Collect all vertices for this object first (for UV computation)
                    object_vertices = []
                    object_triangles = []  # List of (prim_info, tri_verts)

                    for prim_info in prim_infos:
                        try:
                            # Only process triangles (primitive_type value 1)
                            if prim_info.primitive_type.value != 1:
                                continue

                            tri_verts = prim_info.vertices
                            if len(tri_verts) != 3:
                                continue

                            object_triangles.append((prim_info, tri_verts))
                            for v in tri_verts:
                                object_vertices.append([v.x, v.y, v.z])
                        except:
                            continue

                    if not object_triangles:
                        continue

                    # Track the material for textured objects. UVs come straight
                    # from Helios per-primitive (prim_info.texture_uv) — NOT a
                    # PCA projection — so each leaf samples the correct cell of
                    # the leaf texture atlas. They are V-flipped per triangle
                    # below for three.js (which expects V increasing upward).
                    if is_textured:
                        texture_name = os.path.basename(texture_path)

                        # Track material
                        if mat_label not in materials_dict:
                            materials_dict[mat_label] = PlantMaterial(
                                name=mat_label,
                                texture_name=texture_name,
                                has_alpha=True  # Leaf textures use alpha
                            )
                            material_groups_dict[mat_label] = []
                            texture_files[texture_name] = texture_path

                    # Add triangles to main mesh
                    for prim_info, tri_verts in object_triangles:
                        # Get color (RGBcolor has r, g, b attributes)
                        prim_color = prim_info.color
                        color_rgb = [prim_color.r, prim_color.g, prim_color.b]

                        # Vertex-color fallback. Textured organs render their
                        # real texture RGB, so these fixups only matter for the
                        # untextured fallback path (e.g. a textured leaf whose
                        # primitive lacked UVs, or non-textured stems/flowers).

                        if is_textured and (color_rgb[0] == 0 and color_rgb[1] == 0 and color_rgb[2] == 0):
                            # Textured objects often have black [0,0,0] colors from Helios
                            if is_leaf_texture:
                                # Leaf textures → green
                                color_rgb = [0.3, 0.55, 0.2]
                            elif is_woody:
                                # Bark/branch textures on woody species → brown
                                color_rgb = [0.45, 0.3, 0.15]
                            else:
                                # Default to green for other textured parts
                                color_rgb = [0.3, 0.55, 0.2]

                        elif not is_textured and is_woody:
                            # Non-textured parts on woody species
                            # Check brightness to distinguish stems (dark) from flowers (bright)
                            brightness = color_rgb[0] + color_rgb[1] + color_rgb[2]
                            if brightness < 0.5:
                                # Dark colors = stems/branches → brown
                                color_rgb = [0.45, 0.3, 0.15]
                            # else: keep original color (likely flowers)

                        # Get normal (vec3)
                        prim_normal = prim_info.normal
                        normal_xyz = [prim_normal.x, prim_normal.y, prim_normal.z]

                        # Real Helios UVs for this triangle, aligned to its
                        # vertices. A primitive may carry a texture file yet have
                        # no UVs (e.g. a flat-colored organ that shares a
                        # material); in that case fall back to degenerate UVs and
                        # don't add it to the textured material group, so it
                        # renders via vertex color instead of smearing the atlas.
                        prim_uvs = prim_info.texture_uv if is_textured else None
                        tri_is_textured = (
                            is_textured
                            and prim_uvs is not None
                            and len(prim_uvs) == 3
                        )

                        # Add vertices, colors, normals, and UVs for this triangle
                        tri_indices = []
                        for vi, v in enumerate(tri_verts):
                            vertices.append([v.x, v.y, v.z])
                            colors.append(color_rgb)
                            normals.append(normal_xyz)

                            if tri_is_textured:
                                uv = prim_uvs[vi]
                                # V-flip for three.js (flipY=false on the frontend)
                                uvs.append([uv.x, 1.0 - uv.y])
                            else:
                                uvs.append([0.0, 0.0])

                            tri_indices.append(vertex_index)
                            vertex_index += 1

                        faces.append(tri_indices)
                        organ_codes.append(organ_code)

                        # Track which material this triangle uses (only when it
                        # actually has texture coordinates to sample with)
                        if tri_is_textured and mat_label in material_groups_dict:
                            material_groups_dict[mat_label].append(triangle_index)

                        triangle_index += 1

                # Load texture files as base64
                import base64
                textures_data = {}
                for tex_name, tex_path in texture_files.items():
                    if os.path.exists(tex_path):
                        try:
                            with open(tex_path, 'rb') as f:
                                textures_data[tex_name] = base64.b64encode(f.read()).decode('utf-8')
                            print(f"[Plant Generation] Loaded texture: {tex_name}")
                        except Exception as e:
                            print(f"[Plant Generation] Failed to load texture {tex_name}: {e}")

                # Build response data structures
                materials_list = list(materials_dict.values()) if materials_dict else None
                material_groups_list = [
                    PlantMaterialGroup(material_name=mat_name, triangle_indices=tri_indices)
                    for mat_name, tri_indices in material_groups_dict.items()
                    if tri_indices
                ] if material_groups_dict else None

                print(f"[Plant Generation] Extracted {len(vertices)} vertices, {len(faces)} triangles")
                print(f"[Plant Generation] Found {len(materials_dict)} textured materials, {len(textures_data)} textures loaded")

                # Generate Helios plant structure XML
                with tempfile.NamedTemporaryFile(suffix='.xml', delete=False) as tmp_xml:
                    xml_path = tmp_xml.name

                plantarch.writePlantStructureXML(plant_id, xml_path)

                with open(xml_path, 'r') as f:
                    helios_xml = f.read()

                os.unlink(xml_path)

                print(f"[Plant Generation] Complete: {len(vertices)} vertices, {len(faces)} triangles")

                return PlantGenerationResponse(
                    success=True,
                    vertices=vertices,
                    indices=faces,
                    normals=normals,
                    colors=colors,  # Vertex colors for non-textured parts
                    uv_coordinates=uvs if uvs else None,
                    materials=materials_list,
                    material_groups=material_groups_list,
                    textures=textures_data if textures_data else None,
                    organ_codes=organ_codes if organ_codes else None,
                    vertex_count=len(vertices),
                    triangle_count=len(faces),
                    plant_type=request.plant_type,
                    age=request.age,
                    height=height,
                    available_models=available_models,
                    helios_xml=helios_xml
                )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Plant generation failed: {str(e)}")


# Woody species get brown stems, herbaceous species get green stems. Shared by
# the single-plant and canopy generation paths.
_WOODY_SPECIES = {
    'almond', 'apple', 'apple_fruitingwall', 'easternredbud', 'olive',
    'pistachio', 'walnut', 'bougainvillea', 'grapevine_VSP', 'grapevine_Wye',
    'grapevine_GDC', 'grapevine_geneva_double_curtain', 'grapevine_vertical_shoot_positioned',
    'grapevine_sprawl', 'grapevine_unilateral_cordon'
}


def _extract_context_plant_geometry(context, is_woody: bool, progress_cb=None) -> tuple:
    """
    Extract merged plant geometry from every object in a pyhelios Context.

    Walks context.getAllObjectIDs(), so a canopy of N plants (all added to the
    same context) comes back as one merged mesh. UVs are Helios's real
    per-vertex coordinates, V-flipped for three.js — identical convention to
    /api/plant/generate and _extract_session_geometry.

    If progress_cb is given, it is called with a fraction in [0, 1] as objects
    are processed, so a streaming caller can report extraction progress.

    Returns (vertices, faces, colors, normals, uvs, materials_list,
             material_groups_list, textures_data, organ_codes) where organ_codes
             is a per-triangle organ-type code (see _ORGAN_LABEL_TO_CODE).
    """
    import os
    import base64

    object_ids = context.getAllObjectIDs()
    total_objects = len(object_ids) or 1

    vertices = []
    colors = []
    normals = []
    uvs = []
    faces = []
    organ_codes = []  # per-triangle organ-type code, parallel to faces
    vertex_index = 0
    triangle_index = 0

    materials_dict = {}        # mat_label -> PlantMaterial
    material_groups_dict = {}  # mat_label -> [triangle index, ...]
    texture_files = {}         # texture basename -> source path

    for obj_idx, obj_id in enumerate(object_ids):
        if progress_cb is not None and (obj_idx % 64 == 0):
            progress_cb(obj_idx / total_objects)
        try:
            prim_infos = context.getPrimitivesInfoForObject(obj_id)
        except Exception:
            continue

        if not prim_infos:
            continue

        first_prim = prim_infos[0]
        try:
            mat_label = context.getPrimitiveMaterialLabel(first_prim.uuid)
            texture_path = context.getMaterialTexture(mat_label) if mat_label else None
        except Exception:
            mat_label = None
            texture_path = None

        # Organ type is constant within an object; read once, broadcast per triangle.
        organ_code = _organ_code_for_primitive(context, first_prim.uuid)

        is_textured = texture_path and len(texture_path) > 0
        texture_name_lower = texture_path.lower() if texture_path else ""
        is_leaf_texture = 'leaf' in texture_name_lower or 'Leaf' in (texture_path or "")

        if is_textured:
            texture_name = os.path.basename(texture_path)
            if mat_label not in materials_dict:
                materials_dict[mat_label] = PlantMaterial(
                    name=mat_label,
                    texture_name=texture_name,
                    has_alpha=True,
                )
                material_groups_dict[mat_label] = []
                texture_files[texture_name] = texture_path

        for prim_info in prim_infos:
            try:
                if prim_info.primitive_type.value != 1:
                    continue

                tri_verts = prim_info.vertices
                if len(tri_verts) != 3:
                    continue

                prim_color = prim_info.color
                color_rgb = [prim_color.r, prim_color.g, prim_color.b]

                # Vertex-color fallback (matches /api/plant/generate).
                if is_textured and (color_rgb[0] == 0 and color_rgb[1] == 0 and color_rgb[2] == 0):
                    if is_leaf_texture:
                        color_rgb = [0.3, 0.55, 0.2]
                    elif is_woody:
                        color_rgb = [0.45, 0.3, 0.15]
                    else:
                        color_rgb = [0.3, 0.55, 0.2]
                elif not is_textured and is_woody:
                    brightness = color_rgb[0] + color_rgb[1] + color_rgb[2]
                    if brightness < 0.5:
                        color_rgb = [0.45, 0.3, 0.15]

                prim_normal = prim_info.normal
                normal_xyz = [prim_normal.x, prim_normal.y, prim_normal.z]

                prim_uvs = prim_info.texture_uv if is_textured else None
                tri_is_textured = (
                    is_textured and prim_uvs is not None and len(prim_uvs) == 3
                )

                face_indices = []
                for vi, v in enumerate(tri_verts):
                    vertices.append([v.x, v.y, v.z])
                    colors.append(color_rgb)
                    normals.append(normal_xyz)
                    if tri_is_textured:
                        uv = prim_uvs[vi]
                        uvs.append([uv.x, 1.0 - uv.y])
                    else:
                        uvs.append([0.0, 0.0])
                    face_indices.append(vertex_index)
                    vertex_index += 1

                faces.append(face_indices)
                organ_codes.append(organ_code)

                if tri_is_textured and mat_label in material_groups_dict:
                    material_groups_dict[mat_label].append(triangle_index)

                triangle_index += 1
            except Exception:
                continue

    textures_data = {}
    for tex_name, tex_path in texture_files.items():
        if os.path.exists(tex_path):
            try:
                with open(tex_path, 'rb') as f:
                    textures_data[tex_name] = base64.b64encode(f.read()).decode('utf-8')
            except Exception:
                pass

    if progress_cb is not None:
        progress_cb(1.0)

    materials_list = list(materials_dict.values()) if materials_dict else None
    material_groups_list = [
        PlantMaterialGroup(material_name=name, triangle_indices=tris)
        for name, tris in material_groups_dict.items() if tris
    ] if material_groups_dict else None

    return (
        vertices, faces, colors, normals, uvs,
        materials_list, material_groups_list, textures_data, organ_codes,
    )


@app.post("/api/plant/canopy/generate", response_model=PlantGenerationResponse)
async def generate_plant_canopy(request: PlantCanopyRequest):
    """
    Generate a canopy of regularly spaced plants using pyhelios PlantArchitecture.

    Builds a grid of `count_x` x `count_y` plants of the same species, spaced by
    `spacing_x` / `spacing_y` meters and centered on the canopy center. All plants
    are added to a single context and returned as one merged mesh, matching the
    single-plant /api/plant/generate response shape so the renderer is identical.
    """
    try:
        from pyhelios import Context, PlantArchitecture
        from pyhelios.wrappers.DataTypes import vec3, vec2, int2

        # Validate canopy parameters with actionable errors.
        if request.count_x <= 0 or request.count_y <= 0:
            return PlantGenerationResponse(
                success=False, vertices=[], indices=[], vertex_count=0,
                triangle_count=0, plant_type=request.plant_type, age=request.age,
                error=f"Plant counts must be positive (got {request.count_x} x {request.count_y}).",
            )
        if request.age < 0:
            return PlantGenerationResponse(
                success=False, vertices=[], indices=[], vertex_count=0,
                triangle_count=0, plant_type=request.plant_type, age=request.age,
                error=f"Age must be non-negative (got {request.age}).",
            )
        if not (0.0 <= request.germination_rate <= 1.0):
            return PlantGenerationResponse(
                success=False, vertices=[], indices=[], vertex_count=0,
                triangle_count=0, plant_type=request.plant_type, age=request.age,
                error=f"Germination rate must be between 0 and 1 (got {request.germination_rate}).",
            )

        print(f"[Plant Canopy] Starting: type={request.plant_type}, age={request.age}, "
              f"grid={request.count_x}x{request.count_y}, spacing=({request.spacing_x},{request.spacing_y}), "
              f"germination={request.germination_rate}")
        is_woody = request.plant_type in _WOODY_SPECIES

        with Context() as context:
            with PlantArchitecture(context) as plantarch:
                available_models = plantarch.getAvailablePlantModels()

                if request.plant_type not in available_models:
                    return PlantGenerationResponse(
                        success=False, vertices=[], indices=[], vertex_count=0,
                        triangle_count=0, plant_type=request.plant_type, age=request.age,
                        available_models=available_models,
                        error=f"Unknown plant type '{request.plant_type}'. Available: {', '.join(available_models[:10])}...",
                    )

                plantarch.loadPlantModelFromLibrary(request.plant_type)

                build_params = {}
                if request.random_seed is not None:
                    build_params['random_seed'] = request.random_seed

                plant_ids = plantarch.buildPlantCanopyFromLibrary(
                    canopy_center=vec3(request.center_x, request.center_y, request.center_z),
                    plant_spacing=vec2(request.spacing_x, request.spacing_y),
                    plant_count=int2(request.count_x, request.count_y),
                    age=request.age,
                    germination_rate=request.germination_rate,
                    build_parameters=build_params if build_params else None,
                )

                if not plant_ids:
                    return PlantGenerationResponse(
                        success=False, vertices=[], indices=[], vertex_count=0,
                        triangle_count=0, plant_type=request.plant_type, age=request.age,
                        plant_count=0, count_x=request.count_x, count_y=request.count_y,
                        spacing_x=request.spacing_x, spacing_y=request.spacing_y,
                        error="No plants germinated. Try a higher germination rate.",
                    )

                # Report the first plant's height as a representative value.
                try:
                    height = plantarch.getPlantHeight(plant_ids[0])
                except Exception:
                    height = None

                print(f"[Plant Canopy] Built {len(plant_ids)} plants; extracting geometry...")

                (vertices, faces, colors, normals, uvs,
                 materials_list, material_groups_list, textures_data, organ_codes) = \
                    _extract_context_plant_geometry(context, is_woody)

                # XML export: a canopy is not a morph target, so write the first
                # plant's structure (representative) without failing the request.
                helios_xml = None
                try:
                    import tempfile
                    import os
                    with tempfile.NamedTemporaryFile(suffix='.xml', delete=False) as tmp_xml:
                        xml_path = tmp_xml.name
                    plantarch.writePlantStructureXML(plant_ids[0], xml_path)
                    with open(xml_path, 'r') as f:
                        helios_xml = f.read()
                    os.unlink(xml_path)
                except Exception as xml_err:
                    print(f"[Plant Canopy] Warning: failed to write XML: {xml_err}")

                print(f"[Plant Canopy] Complete: {len(plant_ids)} plants, "
                      f"{len(vertices)} vertices, {len(faces)} triangles")

                return PlantGenerationResponse(
                    success=True,
                    vertices=vertices,
                    indices=faces,
                    normals=normals if normals else None,
                    colors=colors,
                    uv_coordinates=uvs if uvs else None,
                    materials=materials_list,
                    material_groups=material_groups_list,
                    textures=textures_data if textures_data else None,
                    organ_codes=organ_codes if organ_codes else None,
                    vertex_count=len(vertices),
                    triangle_count=len(faces),
                    plant_type=request.plant_type,
                    age=request.age,
                    height=height,
                    available_models=available_models,
                    helios_xml=helios_xml,
                    plant_count=len(plant_ids),
                    count_x=request.count_x,
                    count_y=request.count_y,
                    spacing_x=request.spacing_x,
                    spacing_y=request.spacing_y,
                )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Plant canopy generation failed: {str(e)}")


@app.post("/api/plant/generate/stream")
async def generate_plant_stream(request: PlantStreamRequest, http_request: Request):
    """
    Generate a single plant or a canopy with Server-Sent Events progress.

    Emits:
      event: run_id    data: {"run_id": "..."}   (first, so the client can cancel)
      event: progress  data: {"progress": 0.0-1.0, "message": "..."}
      event: result    data: <PlantGenerationResponse-shaped JSON>
      event: cancelled data: {}                  (build aborted; memory freed)
      event: error     data: {"detail": "..."}

    Progress maps the C++ growth phase to 0–0.6 (via setProgressCallback),
    geometry extraction to 0.6–0.95, and JSON serialization to the final 1.0.
    Single-plant builds create a retained session (echoed as session_id) so the
    age slider keeps working; canopies are stateless.

    Cancellation: the run_id ride the first event; a client POSTing
    /api/cancel/{run_id} (or disconnecting) flips a shared ctypes flag that the
    C++ canopy/advanceTime loops poll, so a long build stops between
    plants/timesteps and its Context/PlantArchitecture are torn down promptly.
    """
    import asyncio
    import ctypes as _ctypes
    import queue as _queue
    import time

    progress_queue: "_queue.Queue" = _queue.Queue()
    is_canopy = request.mode == "canopy"
    run_id, cancel_event = _new_cancel_token()
    # Shared cancel flag the C++ build loops poll. The SSE generator flips it the
    # moment this run is cancelled (disconnect or /api/cancel); the worker also
    # checks cancel_event at stage boundaries to bail before geometry extraction.
    cancel_flag = _ctypes.c_int(0)

    def _do():
        from pyhelios import Context, PlantArchitecture
        from pyhelios.wrappers.DataTypes import vec3, vec2, int2

        is_woody = request.plant_type in _WOODY_SPECIES

        # Validate canopy params up front (cheap, before any native work).
        if is_canopy:
            if request.count_x <= 0 or request.count_y <= 0:
                progress_queue.put(("error", "Plant counts must be positive."))
                return
            if not (0.0 <= request.germination_rate <= 1.0):
                progress_queue.put(("error", "Germination rate must be between 0 and 1."))
                return
        if request.age < 0:
            progress_queue.put(("error", "Age must be non-negative."))
            return

        # Single plants retain the context/plantarch in a session (no `with`);
        # canopies build inside a `with` and discard the context afterward.
        context = Context()
        context.__enter__()
        plantarch = PlantArchitecture(context)
        plantarch.__enter__()
        keep_session = False

        try:
            available_models = plantarch.getAvailablePlantModels()
            if request.plant_type not in available_models:
                progress_queue.put(("error", f"Unknown plant type '{request.plant_type}'."))
                return

            plantarch.loadPlantModelFromLibrary(request.plant_type)

            build_params = {}
            if request.random_seed is not None:
                build_params['random_seed'] = request.random_seed

            # Growth progress (C++ advanceTime) → 0–0.6. The callback fires on a
            # native thread; just enqueue, never touch the event loop here.
            def on_progress(progress: float, message: str):
                progress_queue.put(("progress", max(0.0, min(progress, 1.0)) * 0.6, "Growing plants..."))

            plantarch.setProgressCallback(on_progress)
            # Register the cancel flag so the C++ build loops bail mid-build.
            plantarch.setCancelFlag(cancel_flag)
            progress_queue.put(("progress", 0.0, "Growing plants..."))

            if is_canopy:
                plant_ids = plantarch.buildPlantCanopyFromLibrary(
                    canopy_center=vec3(request.center_x, request.center_y, request.center_z),
                    plant_spacing=vec2(request.spacing_x, request.spacing_y),
                    plant_count=int2(request.count_x, request.count_y),
                    age=request.age,
                    germination_rate=request.germination_rate,
                    build_parameters=build_params if build_params else None,
                )
                # A cancelled canopy can return early with few/no plants — report
                # that as cancellation, not a germination failure.
                if cancel_event.is_set():
                    progress_queue.put(("cancelled", None))
                    return
                if not plant_ids:
                    progress_queue.put(("error", "No plants germinated. Try a higher germination rate."))
                    return
                primary_id = plant_ids[0]
            else:
                primary_id = plantarch.buildPlantInstanceFromLibrary(
                    base_position=vec3(request.position_x, request.position_y, request.position_z),
                    age=request.age,
                    build_parameters=build_params if build_params else None,
                )
                plant_ids = [primary_id]

            plantarch.setProgressCallback(None)

            # The build loops honor the cancel flag, so a cancelled run returns a
            # partial/empty plant. Bail here — before the expensive geometry
            # extraction + serialization — and let the finally tear everything
            # down so the C++/numpy memory is freed promptly.
            if cancel_event.is_set():
                progress_queue.put(("cancelled", None))
                return

            try:
                height = plantarch.getPlantHeight(primary_id)
            except Exception:
                height = None

            # Geometry extraction → 0.6–0.95.
            def extract_progress(frac: float):
                progress_queue.put(("progress", 0.6 + max(0.0, min(frac, 1.0)) * 0.35, "Packing geometry..."))

            (vertices, faces, colors, normals, uvs,
             materials_list, material_groups_list, textures_data, organ_codes) = \
                _extract_context_plant_geometry(context, is_woody, progress_cb=extract_progress)

            # Plant structure XML (first/only plant).
            helios_xml = None
            try:
                import tempfile
                import os as _os2
                with tempfile.NamedTemporaryFile(suffix='.xml', delete=False) as tmp_xml:
                    xml_path = tmp_xml.name
                plantarch.writePlantStructureXML(primary_id, xml_path)
                with open(xml_path, 'r') as f:
                    helios_xml = f.read()
                _os2.unlink(xml_path)
            except Exception as xml_err:
                print(f"[Plant Stream] Warning: failed to write XML: {xml_err}")

            result = {
                "success": True,
                "vertices": vertices,
                "indices": faces,
                "normals": normals if normals else None,
                "colors": colors,
                "uv_coordinates": uvs if uvs else None,
                "materials": [m.model_dump() for m in materials_list] if materials_list else None,
                "material_groups": [g.model_dump() for g in material_groups_list] if material_groups_list else None,
                "textures": textures_data if textures_data else None,
                "organ_codes": organ_codes if organ_codes else None,
                "vertex_count": len(vertices),
                "triangle_count": len(faces),
                "plant_type": request.plant_type,
                "age": request.age,
                "height": height,
                "available_models": available_models,
                "helios_xml": helios_xml,
            }

            if is_canopy:
                result.update({
                    "plant_count": len(plant_ids),
                    "count_x": request.count_x,
                    "count_y": request.count_y,
                    "spacing_x": request.spacing_x,
                    "spacing_y": request.spacing_y,
                })
            else:
                # Retain the session for age scrubbing (same shape as
                # create_plant_session). Reuse the plant_type's current age.
                session_id = str(uuid.uuid4())[:8]
                current_age = plantarch.getPlantAge(primary_id)
                session = PlantSession(
                    session_id=session_id,
                    plant_type=request.plant_type,
                    plant_id=primary_id,
                    context=context,
                    plantarch=plantarch,
                    current_age=current_age,
                    position=(request.position_x, request.position_y, request.position_z),
                    created_at=time.time(),
                    last_accessed=time.time(),
                )
                with _session_lock:
                    _plant_sessions[session_id] = session
                _sweep_plant_sessions()
                result["session_id"] = session_id
                result["age"] = current_age
                keep_session = True

            # Serialize here, in the worker thread — for a large canopy this is
            # the single most expensive uninstrumented step (millions of floats).
            # Doing it in the SSE generator instead would freeze the bar at the
            # last progress value while the event loop blocks on json.dumps.
            progress_queue.put(("progress", 0.97, "Finalizing..."))
            result_json = json.dumps(result)
            progress_queue.put(("done", result_json))
        except Exception as e:
            import traceback
            traceback.print_exc()
            progress_queue.put(("error", str(e)))
        finally:
            try:
                plantarch.setProgressCallback(None)
                # Clear the cancel flag: it's a local c_int that dies with this
                # request, but a retained single-plant session keeps `plantarch`
                # alive — leaving the pointer set would dangle on later age
                # scrubbing (advanceTime). Drop it so the session re-registers a
                # fresh flag if/when it next needs one.
                plantarch.setCancelFlag(None)
            except Exception:
                pass
            # Tear down pyhelios resources unless a session owns them now.
            if not keep_session:
                try:
                    plantarch.__exit__(None, None, None)
                    context.__exit__(None, None, None)
                except Exception:
                    pass

    async def event_generator():
        loop = asyncio.get_event_loop()
        # Emit the run_id up front so the client can cancel before heavy work.
        yield f"event: run_id\ndata: {json.dumps({'run_id': run_id})}\n\n"
        task = loop.run_in_executor(None, _do)

        try:
            while True:
                try:
                    item = await asyncio.to_thread(progress_queue.get, True, 0.25)
                except _queue.Empty:
                    # On each idle tick, propagate a cancel (from /api/cancel or a
                    # client disconnect) into the C++ cancel flag so the in-flight
                    # build bails at its next plant/timestep boundary.
                    if await http_request.is_disconnected():
                        cancel_event.set()
                    if cancel_event.is_set():
                        cancel_flag.value = 1
                    if task.done():
                        exc = task.exception()
                        if exc:
                            yield f"event: error\ndata: {json.dumps({'detail': str(exc)})}\n\n"
                        break
                    continue

                kind = item[0]
                if kind == "progress":
                    _, progress_val, message = item
                    yield f"event: progress\ndata: {json.dumps({'progress': progress_val, 'message': message})}\n\n"
                elif kind == "done":
                    # item[1] is already-serialized JSON (built in the worker thread).
                    yield f"event: result\ndata: {item[1]}\n\n"
                    break
                elif kind == "cancelled":
                    yield f"event: cancelled\ndata: {json.dumps({})}\n\n"
                    break
                elif kind == "error":
                    yield f"event: error\ndata: {json.dumps({'detail': item[1]})}\n\n"
                    break
        finally:
            _clear_run(run_id)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ==================== TEXTURED MESH IMPORT (OBJ + MTL) ====================
# Parses an OBJ (with its sibling MTL and texture images) from a disk path and
# returns geometry + real per-vertex UVs + base64 textures, in the same shape
# the renderer already consumes for plant models (so the same textured renderer
# handles both). Triangles are emitted non-indexed (3 vertices each), matching
# the plant path's expanded-geometry convention.

class MeshImportRequest(BaseModel):
    """Request to import a textured mesh from a file on disk."""
    path: str  # Absolute path to the .obj file


class MeshImportResponse(BaseModel):
    """Imported mesh geometry + textures (mirrors PlantGenerationResponse)."""
    success: bool
    vertices: List[List[float]] = []          # [[x, y, z], ...]
    indices: List[List[int]] = []             # [[v0, v1, v2], ...]
    normals: Optional[List[List[float]]] = None
    colors: Optional[List[List[float]]] = None        # per-vertex (from Kd)
    uv_coordinates: Optional[List[List[float]]] = None  # [[u, v], ...] V-flipped
    materials: Optional[List[PlantMaterial]] = None
    material_groups: Optional[List[PlantMaterialGroup]] = None
    textures: Optional[Dict[str, str]] = None         # {basename: base64 png/jpg}
    vertex_count: int = 0
    triangle_count: int = 0
    filename: Optional[str] = None
    has_textures: bool = False
    error: Optional[str] = None


def _import_ply_mesh(ply_path: Path) -> MeshImportResponse:
    """Read a polygon-mesh PLY (ASCII or binary) into MeshImportResponse geometry.

    PLY is an ambiguous container — it may hold a point cloud (vertices only) or a
    polygon mesh (vertices + faces). The caller has already decided this is a mesh
    (the frontend sniffs the header for `element face`); here we require triangles
    and reject a vertices-only PLY. open3d's read_triangle_mesh transparently
    handles ascii / binary_little_endian / binary_big_endian. PLY meshes carry no
    MTL or textures, so the textured fields stay empty."""
    import open3d as o3d

    try:
        mesh = o3d.io.read_triangle_mesh(str(ply_path))
    except Exception as e:  # noqa: BLE001 - surface a clean 400
        raise HTTPException(status_code=400, detail=f"Failed to read PLY mesh: {e}")

    triangles = np.asarray(mesh.triangles)
    if triangles.size == 0:
        raise HTTPException(
            status_code=400,
            detail="No faces found in PLY mesh (file appears to be a point cloud).",
        )

    if not mesh.has_vertex_normals():
        mesh.compute_vertex_normals()

    vertices = np.asarray(mesh.vertices, dtype=float)
    normals = np.asarray(mesh.vertex_normals, dtype=float)
    colors_arr = np.asarray(mesh.vertex_colors, dtype=float)  # 0-1, per vertex

    out_colors = colors_arr.tolist() if colors_arr.shape[0] == vertices.shape[0] else None
    out_normals = normals.tolist() if normals.shape[0] == vertices.shape[0] else None

    return MeshImportResponse(
        success=True,
        vertices=vertices.tolist(),
        indices=triangles.tolist(),
        normals=out_normals,
        colors=out_colors,
        uv_coordinates=None,
        materials=None,
        material_groups=None,
        textures=None,
        vertex_count=int(vertices.shape[0]),
        triangle_count=int(triangles.shape[0]),
        filename=ply_path.name,
        has_textures=False,
    )


@app.post("/api/mesh/import", response_model=MeshImportResponse)
async def import_textured_mesh(request: MeshImportRequest):
    """Parse an OBJ (+ MTL + texture images) from disk into textured geometry."""
    from qsm.obj_loader import load_obj_template

    obj_path = Path(request.path)
    if not obj_path.is_file():
        raise HTTPException(status_code=404, detail=f"Mesh file not found: {request.path}")
    ext = obj_path.suffix.lower()
    if ext == '.ply':
        # PLY meshes carry no MTL/textures; open3d reads ASCII + binary directly.
        return _import_ply_mesh(obj_path)
    if ext != '.obj':
        raise HTTPException(status_code=400, detail="Only .obj and .ply files are supported for mesh import")

    # Parse the OBJ (+ MTL + textures) via the shared loader. It returns expanded
    # (non-indexed) triangle geometry with V-flipped UVs and base64 textures.
    try:
        tpl = load_obj_template(obj_path)
    except ValueError as e:
        # No triangles is a 400 (bad input); anything else is a parse failure.
        if "No triangles" in str(e):
            raise HTTPException(status_code=400, detail=str(e))
        raise HTTPException(status_code=500, detail=str(e))

    out_vertices: List[List[float]] = tpl["vertices"]
    out_normals: List[List[float]] = tpl["normals"]
    out_uvs: List[List[float]] = tpl["uvs"]
    out_faces: List[List[int]] = tpl["faces"]
    tri_material: List[Optional[str]] = tpl["tri_material"]
    mtl_materials: Dict[str, dict] = tpl["mtl_materials"]
    textures_data: Dict[str, str] = tpl["textures"]
    material_texture_name: Dict[str, str] = tpl["material_texture_name"]
    has_any_normals = True  # loader always populates normals (flat-filled where absent)

    # Build per-vertex colors from each triangle's material Kd (fallback grey).
    out_colors: List[List[float]] = [[0.8, 0.8, 0.8]] * len(out_vertices)
    for ti, face in enumerate(out_faces):
        mat = tri_material[ti]
        kd = mtl_materials.get(mat, {}).get('Kd') if mat else None
        if kd:
            for i in face:
                out_colors[i] = kd

    # Material groups: only for materials that actually have a loaded texture.
    materials_list: List[PlantMaterial] = []
    groups: Dict[str, List[int]] = {}
    for ti, mat in enumerate(tri_material):
        if mat and mat in material_texture_name and material_texture_name[mat] in textures_data:
            groups.setdefault(mat, []).append(ti)
    for mat, tri_indices in groups.items():
        tex_name = material_texture_name[mat]
        kd = mtl_materials.get(mat, {}).get('Kd')
        materials_list.append(PlantMaterial(
            name=mat,
            color=kd,
            texture_name=tex_name,
            has_alpha=tex_name.lower().endswith('.png'),  # assume PNG may carry alpha
        ))
    material_groups_list = [
        PlantMaterialGroup(material_name=mat, triangle_indices=tris)
        for mat, tris in groups.items()
    ]

    has_textures = bool(textures_data) and bool(material_groups_list)

    return MeshImportResponse(
        success=True,
        vertices=out_vertices,
        indices=out_faces,
        normals=out_normals if has_any_normals else None,
        colors=out_colors,
        uv_coordinates=out_uvs if has_textures else None,
        materials=materials_list or None,
        material_groups=material_groups_list or None,
        textures=textures_data or None,
        vertex_count=len(out_vertices),
        triangle_count=len(out_faces),
        filename=obj_path.name,
        has_textures=has_textures,
    )


# ==================== POINT CLOUD LAS/LAZ IMPORT/EXPORT ====================
# Uses laspy for reading and writing LAS/LAZ files with compression support

class PointCloudExportRequest(BaseModel):
    """Request for exporting a point cloud.

    Flat clouds send inline `points` (+ optional `colors`) and `format` in
    {"las","laz"}. Octree-backed clouds send `source` instead — and may request
    any of {"las","laz","xyz","txt","csv","ply"} since the renderer has no
    positions to format text from. The backend reads the source file, applies
    pending translation, and returns base64-encoded output in all cases.
    """
    points: Optional[List[List[float]]] = None  # [[x, y, z], ...]
    colors: Optional[List[List[float]]] = None  # [[r, g, b], ...] in 0-1 range
    source: Optional[PointSource] = None         # octree-backed clouds read from disk
    format: str = "laz"  # las | laz | xyz | txt | csv | ply
    filename: Optional[str] = None  # Optional filename for the export


class PointCloudExportResponse(BaseModel):
    """Response containing the exported file data"""
    success: bool
    data: Optional[str] = None  # Base64-encoded file content
    filename: str
    point_count: int
    has_colors: bool
    format: str
    error: Optional[str] = None


class PointCloudImportResponse(BaseModel):
    """Response containing imported point cloud data"""
    success: bool
    points: Optional[List[List[float]]] = None  # [[x, y, z], ...]
    colors: Optional[List[List[float]]] = None  # [[r, g, b], ...] in 0-1 range
    point_count: int = 0
    has_colors: bool = False
    filename: Optional[str] = None
    error: Optional[str] = None


def _format_points_as_text(
    fmt: str,
    points: np.ndarray,
    colors: Optional[np.ndarray],
    intensity: Optional[np.ndarray],
) -> str:
    """Format an (N,3) point array (+ optional 0-1 colors, intensity) as XYZ /
    TXT / CSV / PLY / OBJ text. Column conventions match the renderer's flat
    text export exactly: 6-decimal positions, colors as 0-255 ints, intensity
    4-decimal.

    Vectorised with `np.savetxt` rather than a per-point Python f-string loop —
    on a multi-million-point octree cloud the old loop dominated export time and
    held the whole formatted string list in RAM. Output is byte-identical to the
    previous loop (same precision, separators, headers, and no trailing newline).
    """
    import io

    n = len(points)
    has_colors = colors is not None and len(colors) == n
    has_int = intensity is not None and len(intensity) == n
    rgb = np.clip(np.rint(colors * 255.0), 0, 255).astype(int) if has_colors else None

    def _savetxt(cols: list, fmts: list, sep: str) -> str:
        """np.savetxt the column-stacked `cols` with per-column `fmts`, joined by
        `sep`, returning the body WITHOUT the trailing newline savetxt appends."""
        buf = io.StringIO()
        np.savetxt(buf, np.column_stack(cols), fmt=sep.join(fmts), delimiter=sep)
        return buf.getvalue().rstrip("\n")

    pos_fmt = ["%.6f", "%.6f", "%.6f"]

    if fmt == "xyz":
        return _savetxt([points[:, 0], points[:, 1], points[:, 2]], pos_fmt, " ")

    if fmt in ("txt", "csv"):
        sep = "," if fmt == "csv" else " "
        head = ["X", "Y", "Z"]
        cols = [points[:, 0], points[:, 1], points[:, 2]]
        fmts = list(pos_fmt)
        if has_colors:
            head += ["R", "G", "B"]
            cols += [rgb[:, 0], rgb[:, 1], rgb[:, 2]]
            fmts += ["%d", "%d", "%d"]
        if has_int:
            head += ["Intensity"]
            cols += [np.asarray(intensity, dtype=np.float64)]
            fmts += ["%.4f"]
        body = _savetxt(cols, fmts, sep)
        # Match the old loop exactly: header only (no trailing newline) when empty.
        return sep.join(head) + ("\n" + body if body else "")

    if fmt == "ply":
        header = ["ply", "format ascii 1.0", f"element vertex {n}",
                  "property float x", "property float y", "property float z"]
        cols = [points[:, 0], points[:, 1], points[:, 2]]
        fmts = list(pos_fmt)
        if has_colors:
            header += ["property uchar red", "property uchar green", "property uchar blue"]
            cols += [rgb[:, 0], rgb[:, 1], rgb[:, 2]]
            fmts += ["%d", "%d", "%d"]
        header.append("end_header")
        body = _savetxt(cols, fmts, " ")
        return "\n".join(header) + ("\n" + body if body else "")

    if fmt == "obj":
        header = ["# Point cloud exported from Phytograph", f"# {n} points"]
        # The 'v ' prefix is the first format field (a literal column).
        body = _savetxt(
            [points[:, 0], points[:, 1], points[:, 2]],
            ["v %.6f", "%.6f", "%.6f"], " ",
        )
        return "\n".join(header) + ("\n" + body if body else "")

    raise HTTPException(status_code=400, detail=f"Unsupported text export format: {fmt}")


@app.post("/api/pointcloud/export", response_model=PointCloudExportResponse)
async def export_point_cloud_las(request: PointCloudExportRequest):
    """
    Export a point cloud.

    Flat clouds export LAS/LAZ via laspy. Octree-backed clouds send a `source`
    descriptor and may export any of LAS/LAZ/XYZ/TXT/CSV/PLY/OBJ — the backend
    reads the source file (the renderer has no positions to format). Returns
    base64-encoded file data that can be downloaded on the frontend.
    """
    import tempfile
    import base64

    fmt = request.format.lower()

    # Octree-backed export: read points (+ colors/intensity) from the source
    # file, then dispatch by format. Text formats stream out here because the
    # renderer can't iterate an empty positions buffer.
    src_colors = None
    src_intensity = None
    if request.source is not None:
        src = request.source
        src.want_colors = True
        # Export preserves whatever the user imported, sky/misses included — the
        # compute consumers drop them, but an export should round-trip the cloud.
        src.include_misses = True
        points, src_colors, src_intensity = _read_points_from_source(src)
        if fmt in ("xyz", "txt", "csv", "ply", "obj"):
            try:
                text = _format_points_as_text(fmt, points, src_colors, src_intensity)
            except HTTPException:
                raise
            except Exception as e:
                return PointCloudExportResponse(
                    success=False, data=None, filename="", point_count=0,
                    has_colors=False, format=request.format,
                    error=f"Export failed: {e}",
                )
            ext = "." + fmt
            filename = request.filename or f"pointcloud{ext}"
            if not filename.endswith(ext):
                filename = filename.rsplit(".", 1)[0] + ext
            return PointCloudExportResponse(
                success=True,
                data=base64.b64encode(text.encode("utf-8")).decode("utf-8"),
                filename=filename,
                point_count=int(len(points)),
                has_colors=src_colors is not None,
                format=request.format,
            )
        # LAS/LAZ fall through to the laspy path below using `points`/`src_colors`.

    try:
        import laspy
    except ImportError:
        return PointCloudExportResponse(
            success=False,
            data=None,
            filename="",
            point_count=0,
            has_colors=False,
            format=request.format,
            error="laspy library not installed. Run: pip install laspy[lazrs]"
        )

    try:
        if request.source is not None:
            # points already loaded above; colors come back as 0-1 (N,3) or None.
            has_colors = src_colors is not None and len(src_colors) > 0
            colors_arr = src_colors
        else:
            points = np.array(request.points or [])
            has_colors = request.colors is not None and len(request.colors) > 0
            colors_arr = np.array(request.colors) if has_colors else None

        if len(points) == 0:
            return PointCloudExportResponse(
                success=False,
                data=None,
                filename="",
                point_count=0,
                has_colors=False,
                format=request.format,
                error="No points provided"
            )

        # Choose point format: 0 = XYZ only, 2 = XYZ + RGB
        point_format = 2 if has_colors else 0

        # Create LAS header
        header = laspy.LasHeader(point_format=point_format, version="1.2")

        # Calculate offsets and scales for precision
        min_coords = points.min(axis=0)
        max_coords = points.max(axis=0)

        header.offsets = min_coords
        header.scales = [0.001, 0.001, 0.001]  # 1mm precision

        # Create LAS data
        las = laspy.LasData(header)
        las.x = points[:, 0]
        las.y = points[:, 1]
        las.z = points[:, 2]

        # Add colors if available (convert from 0-1 to 16-bit)
        if has_colors:
            # Ensure colors are in 0-1 range and convert to 16-bit
            colors = np.clip(np.asarray(colors_arr, dtype=np.float64), 0, 1)
            las.red = (colors[:, 0] * 65535).astype(np.uint16)
            las.green = (colors[:, 1] * 65535).astype(np.uint16)
            las.blue = (colors[:, 2] * 65535).astype(np.uint16)

        # Determine file extension
        ext = ".laz" if request.format.lower() == "laz" else ".las"
        filename = request.filename or f"pointcloud{ext}"
        if not filename.endswith(ext):
            filename = filename.rsplit('.', 1)[0] + ext

        # Write to temporary file
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp_path = tmp.name

        try:
            las.write(tmp_path)

            # Read back and encode as base64
            with open(tmp_path, 'rb') as f:
                file_data = base64.b64encode(f.read()).decode('utf-8')

            return PointCloudExportResponse(
                success=True,
                data=file_data,
                filename=filename,
                point_count=len(points),
                has_colors=has_colors,
                format=request.format
            )
        finally:
            import os
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return PointCloudExportResponse(
            success=False,
            data=None,
            filename="",
            point_count=0,
            has_colors=False,
            format=request.format,
            error=f"Export failed: {str(e)}"
        )


@app.post("/api/pointcloud/import", response_model=PointCloudImportResponse)
async def import_point_cloud_las(file: UploadFile = File(...)):
    """
    Import a LAS or LAZ file (multipart upload) and stream back a packed binary
    (PHX1) point-cloud response — the SAME format as `import_by_path`, decoded
    straight into Float32Arrays by the renderer.

    This is the no-disk-path fallback (a File blob with no real path can't use
    import_by_path). It previously returned `points.tolist()` as a JSON body,
    which on a large LAZ trips V8's ~512 MB max-string ceiling and triples peak
    memory; the binary stream avoids both. laspy + lazrs handle LAZ; reading is
    shared with the path-based loader via `_load_las_arrays`.
    """
    import tempfile
    import os

    filename = file.filename or "upload.las"
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ['.las', '.laz']:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: {ext}. Expected .las or .laz",
        )

    # Save the upload to a temp file so laspy (which reads from a path) can
    # decode it, then reuse the shared loader + binary packer.
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    try:
        positions, colors, intensity = _load_las_arrays(tmp_path)
        return _pack_pointcloud_response(positions, colors, intensity)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ==================== POINT CLOUD PATH-BASED IMPORT ====================
# Reads point-cloud files directly from disk rather than over HTTP. The
# renderer's TS parsers (parseXYZ, parsePLY, parsePCD) all materialise the
# whole file as a JS string and hit V8's max string size (~512 MB) on
# multi-hundred-MB scans typical of TLS surveys. The endpoint here returns a
# packed binary stream so we don't re-trip the same limit on the response.

# Helios <ASCII_format> tokens we recognise for XYZ-family files. Roles in
# DATA_ROLES populate dedicated fields (positions/colours/intensity); the
# per-pulse multi-return roles (timestamp/target_index/target_count) are carried
# as extra dimensions under their canonical slug (see `_MULTI_RETURN_SLUGS`) so
# the LAD path can recover them; any remaining known-but-unmapped role
# (deviation, …) is carried as a generic extra so pandas stays column-aligned.
_XYZ_DATA_ROLES = {
    'x', 'y', 'z',
    'r', 'g', 'b',
    'r255', 'g255', 'b255',
    'intensity', 'reflectance',
}
# Canonical per-pulse slugs Helios's ASCII loader reads from each hit's data map
# to tell single- from multi-return scans (see `_LAD_MULTI_RETURN_COLUMNS`). We
# pin these as extra-dim slugs at import so they survive edits/bake exactly like
# positions and the LAD accessor can find them deterministically — regardless of
# the source header text or how the wizard labelled the column.
_MULTI_RETURN_SLUGS = ('timestamp', 'target_index', 'target_count')
_MULTI_RETURN_LABELS = {
    'timestamp': 'Timestamp',
    'target_index': 'Target Index',
    'target_count': 'Target Count',
}
# Structured-scan raster indices: the integer (row, column) position of each
# point within the scanner's rectangular acquisition grid. Carried as extra
# dimensions under these pinned slugs so the C++ grid-based direction recovery
# (LiDAR.cpp, for gap-filling unplaceable misses) finds the raster by name —
# exactly like the E57 structured-import path emits them (see `_e57_*`). Pinning
# the slug means an ASCII export with row/column columns round-trips into the
# same recovery machinery regardless of the source header text.
_GRID_INDEX_SLUGS = ('row_index', 'column_index')
_GRID_INDEX_LABELS = {
    'row_index': 'Row Index',
    'column_index': 'Column Index',
}
# 'is_miss' is the canonical sky/miss slug (`_MISS_SLUG`, defined below). It's
# spelled literally here because `_MISS_SLUG` is declared after this set; keeping
# it in the known-roles set lets `_tokenize_ascii_format` preserve an
# `<ASCII_format>x y z is_miss</ASCII_format>` token instead of dropping it to
# 'skip' — which previously zeroed the miss column on import and broke LAD's
# Beer's-law inversion (it needs through-canopy miss rays).
_XYZ_KNOWN_ROLES = (
    _XYZ_DATA_ROLES | set(_MULTI_RETURN_SLUGS) | set(_GRID_INDEX_SLUGS)
    | {'deviation', 'is_miss'}
)

# Standard LAS point-format dimensions that are NOT carried as user-selectable
# scalar fields on import, because they're either materialised elsewhere in the
# session (x/y/z → positions, red/green/blue → colors, intensity → intensity) or
# auto-mapped to a canonical multi-return slug (`_LAS_MULTIRETURN_SRC`). Every
# OTHER standard dimension that holds non-constant data (classification,
# scan_angle, point_source_id, user_data, scanner_channel, …) IS carried so it
# shows up in the renderer's scalar picker — see `_read_las_into_arrays`. Bit-
# flag sub-fields (synthetic/key_point/withheld/overlap/...) are packed into the
# single 'classification flags' byte by LAS 1.4; they're degenerate (all-zero)
# on the clouds we ingest and noisy in the picker, so they stay on this list.
_LAS_STD_DIMS_SKIP = {
    'X', 'Y', 'Z',
    'red', 'green', 'blue',
    'intensity',
    'synthetic', 'key_point', 'withheld', 'overlap',
    'scan_direction_flag', 'edge_of_flight_line',
}
# Standard LAS dims auto-mapped to a canonical per-pulse slug in
# `_read_las_into_arrays`; carried under the slug, not their raw name, so the LAD
# path finds them — excluded from the generic standard-dim carry to avoid a
# duplicate column.
_LAS_MULTIRETURN_SRC = ('return_number', 'number_of_returns', 'gps_time')

# Per-point sky/miss flag carried as a LAS extra dimension (0.0 = hit, 1.0 =
# miss). Misses are laser pulses that returned nothing (hit the sky); Helios
# represents each as a real point placed `_MISS_GAP_DISTANCE` metres from the
# scanner along the pulse direction. They flow through the same extra-dim →
# octree → session → LAD machinery as any scalar, so deletes/bake keep them in
# lockstep with positions, the renderer can colour/hide them, and the LAD path
# reads them for free. The slug is pinned (case-insensitive aliases accepted on
# import) so the renderer and LAD find it deterministically.
_MISS_SLUG = 'is_miss'
_MISS_LABEL = 'Miss'
# Aliases a source column may use for the miss flag (PLY property / E57 field).
_MISS_ALIASES = {'is_miss', 'miss', 'sky', 'ismiss'}
# Same set with punctuation/case stripped, for matching a sanitised slug.
_MISS_ALIASES_NORMALISED = {re.sub(r'[^a-z0-9]+', '', a) for a in _MISS_ALIASES}


def _normalise_miss_alias(name: str) -> Optional[str]:
    """Return `_MISS_SLUG` when `name` is any miss-flag spelling (is_miss / miss
    / sky, case- and punctuation-insensitive), else None."""
    base = re.sub(r'[^a-z0-9]+', '', name.strip().lower())
    return _MISS_SLUG if base in _MISS_ALIASES_NORMALISED else None


# Per-pulse beam-origin triple. When all three are present these are GROUND-TRUTH
# laser emission points (one per return), in the SAME world frame as positions;
# the LAD path uses them directly and bypasses the timestamp->trajectory join
# (see `CloudSession.beam_origins` / `_do_lad_computation`). Defined HERE (rather
# than next to the LAS reader where it's also used) so the ASCII import path —
# header auto-detect and column-plan canonicalisation below — can share the exact
# same alias spellings. Origins are world/UTM coordinates needing full float64
# precision, so they are NEVER carried as float32 extra dims; the ASCII path
# captures them into a side-channel (see `_xyz_to_las` `capture_origins`) exactly
# as positions are, and the LAS path reads them as float64 in `_read_las_into_arrays`.
_ORIGIN_SLUGS = ('origin_x', 'origin_y', 'origin_z')
_ORIGIN_LABELS = {
    'origin_x': 'Beam Origin X',
    'origin_y': 'Beam Origin Y',
    'origin_z': 'Beam Origin Z',
}
# Case-insensitive ExtraBytes / header name aliases for the beam-origin triple.
# When a source carries all three of any one set, those are the emission origins.
# Matched lower-cased; the ASCII path also matches with punctuation stripped.
_BEAM_ORIGIN_ALIAS_SETS = (
    ("ox", "oy", "oz"),
    ("xorigin", "yorigin", "zorigin"),
    ("beamoriginx", "beamoriginy", "beamoriginz"),
)
# Map each alias spelling (punctuation/case stripped) to its canonical origin
# slug, preserving the x/y/z axis. Built once so `_normalise_origin_alias` is a
# dict lookup mirroring `_normalise_miss_alias`.
_ORIGIN_ALIAS_TO_SLUG = {
    re.sub(r'[^a-z0-9]+', '', alias): _ORIGIN_SLUGS[axis]
    for triple in _BEAM_ORIGIN_ALIAS_SETS
    for axis, alias in enumerate(triple)
}


def _normalise_origin_alias(name: str) -> Optional[str]:
    """Return the canonical origin slug ('origin_x'/'origin_y'/'origin_z') when
    `name` is any beam-origin spelling (ox/oy/oz, xorigin/yorigin/zorigin,
    beamoriginx/y/z — case- and punctuation-insensitive), else None. Mirrors
    `_normalise_miss_alias`, but maps to one of three axis-specific slugs."""
    base = re.sub(r'[^a-z0-9]+', '', name.strip().lower())
    return _ORIGIN_ALIAS_TO_SLUG.get(base)
# Distance (metres) at which a miss point is placed from the scanner origin along
# its pulse direction. Matches Helios's gap_distance (LiDAR.cpp gapfillMisses).
_MISS_GAP_DISTANCE = 20000.0

# Helios assigns this sentinel target_index to a sky/miss return (LiDAR.cpp:5609,
# "Special value to exclude from triangulation"). A miss-recording synthetic scan
# exported to bare ASCII (e.g. `x y z timestamp target_index target_count`) drops
# the canonical `is_miss` column but keeps this, so it's the primary signal for
# recovering the flag at import — see `_autodetect_misses`.
_MISS_TARGET_INDEX_SENTINEL = 99.0

# Default far-field distance (m) for the DISTANCE fallback when neither an
# is_miss column nor a target_index sentinel is present. Matches Helios's
# LIDAR_RAYTRACE_MISS_T (LiDAR.h:831); user-overridable via the request.
_MISS_RAYTRACE_DISTANCE = 1001.0


def _autodetect_misses(
    positions: "np.ndarray",
    extras: Dict[str, "np.ndarray"],
    extra_dims_meta: List[dict],
    origin: Optional[Sequence[float]] = None,
    distance_threshold: Optional[float] = None,
) -> int:
    """Recover the canonical `is_miss` flag for a cloud that carries Helios miss
    points but no explicit `is_miss` column, MUTATING `extras`/`extra_dims_meta`
    in place. Returns the number of misses tagged (0 if none / not applicable).

    Helios synthetic scans place sky/miss returns at a far-field distance and tag
    them `target_index == 99` (LiDAR.cpp:5607-5615). An ASCII export of such a
    scan (e.g. `x y z timestamp target_index target_count`) keeps the sentinel
    but loses `is_miss`, so the octree/overlay/LAD infrastructure — all of which
    key off `is_miss` — never sees them and the far-field points poison the
    bounding box. We synthesise the flag here, at the single file-read point, so
    every downstream consumer works unchanged.

    Precedence:
      1. `is_miss` already present with any nonzero → leave untouched (an explicit
         column always wins; matches E57 / structured-PLY recovery).
      2. `target_index` present → miss = (target_index == 99). Exact, needs no
         origin. This is the path the reported fixture hits.
      3. origin known → miss = distance-from-origin >= 0.98 * threshold (mirrors
         the C++ 0.98 band around miss_distance). `threshold` defaults to 1001 m.
      4. neither index nor origin → no signal; tag nothing (don't guess).
    """
    # 1. Honour an explicit, already-flagged column.
    existing = extras.get(_MISS_SLUG)
    if existing is not None and bool(np.any(np.asarray(existing) != 0)):
        return 0

    n = int(positions.shape[0])
    if n == 0:
        return 0

    miss: Optional[np.ndarray] = None

    # 2. Primary: the target_index == 99 sentinel.
    tindex = extras.get('target_index')
    if tindex is not None:
        miss = np.asarray(tindex) == _MISS_TARGET_INDEX_SENTINEL

    # 3. Fallback: far-field distance from the scanner origin.
    elif origin is not None:
        org = np.asarray(origin, dtype=np.float64).reshape(3)
        thr = float(distance_threshold) if distance_threshold else _MISS_RAYTRACE_DISTANCE
        dist = np.linalg.norm(positions.astype(np.float64) - org, axis=1)
        miss = dist >= 0.98 * thr

    if miss is None:
        return 0

    count = int(np.count_nonzero(miss))
    if count == 0:
        return 0

    # Overwrite any all-zero is_miss column (e.g. one synthesised by a prior step)
    # rather than appending a duplicate dim.
    extras[_MISS_SLUG] = miss.astype(np.float32)
    if not any(ed.get("slug") == _MISS_SLUG for ed in extra_dims_meta):
        extra_dims_meta.append({"slug": _MISS_SLUG, "label": _MISS_LABEL})
    return count

# Magic bytes on the wire. Renderer aborts if it sees anything else, so the
# format is implicitly versioned by this value.
_POINTCLOUD_BIN_MAGIC = b'PHX1'

# Extensions dispatched to the pandas-based ASCII path. Anything in this set
# may be accompanied by a Helios `ascii_format` hint; PLY/PCD ignore it.
_PANDAS_EXTENSIONS = {'xyz', 'txt', 'csv', 'pts', 'asc'}
_OPEN3D_EXTENSIONS = {'ply', 'pcd'}
# LAS/LAZ via laspy — served by the binary import_by_path path (was previously
# only reachable through the slow multipart /api/pointcloud/import JSON endpoint).
_LAS_EXTENSIONS = {'las', 'laz'}


class ColumnPlanEntry(BaseModel):
    """One column's import mapping, produced by the import wizard.

    `role` is a Helios-style token (x/y/z/r255/g255/b255/r/g/b/intensity/
    reflectance/skip) or the literal 'extra' for a carried scalar field. For an
    'extra' column, `slug`/`label` give the on-disk LAS extra-dim name and the
    picker label (rename), and `categorical` marks it for discrete colouring in
    the renderer. `index` is the 0-based source column position.
    """
    index: int
    role: str
    slug: Optional[str] = None
    label: Optional[str] = None
    categorical: bool = False


class ColumnPlan(BaseModel):
    """Explicit column layout for an XYZ-family file, from the import wizard.

    When attached to an import/convert request it fully overrides header/format
    auto-detection. `rgb_is_255` records whether the r/g/b columns are 0-255
    integers (True) or already 0-1 floats (False) so the LAS writer scales them
    correctly. Applies only to ASCII formats; PLY/PCD/LAS define their own
    layout and ignore it.
    """
    columns: List[ColumnPlanEntry]
    rgb_is_255: bool = True


class ImportPointCloudByPathRequest(BaseModel):
    """Path-based point-cloud import.

    `ascii_format` is a Helios <ASCII_format> string (e.g.
    'x y z r255 g255 b255 reflectance') and applies only to XYZ-family
    extensions; PLY/PCD ignore it because their column layout is in-file.
    If omitted for an XYZ-family file, columns are sniffed from the first
    non-blank, non-comment row. `column_plan`, when present, fully overrides
    both — it's the explicit layout chosen in the import wizard.
    """
    file_path: str
    ascii_format: Optional[str] = None
    column_plan: Optional[ColumnPlan] = None
    # CloudCompare-style global shift [x, y, z] SUBTRACTED from every point at
    # import (mirrors the cloud-session path). For the flat (non-octree) small-
    # cloud path the renderer keeps the resulting small coordinates as its in-RAM
    # array; the shift is echoed in the response so the renderer can persist it.
    # None / omitted = keep the original coordinates.
    world_shift: Optional[List[float]] = None


class PointCloudPreviewRequest(BaseModel):
    """Inspect a point-cloud file cheaply for the import wizard.

    Reads only the header + first `max_rows` data rows (ASCII) or the header +
    a few points (LAS/PLY/PCD) — never materialises the whole file. The optional
    `ascii_format` hint biases role detection the same way the import path does.
    """
    file_path: str
    ascii_format: Optional[str] = None
    max_rows: int = 20


class PreviewColumn(BaseModel):
    """One source column as the wizard should present it.

    `detected_role` is the auto-detected Helios role (or 'extra'/'skip');
    `suggested_slug`/`suggested_label` mirror what import would name a carried
    scalar (so the wizard's defaults match the eventual on-disk attribute).
    `type_hint` is a sniffed value shape (integer/float/categorical/empty) used
    to pre-tick the categorical checkbox. `remappable` is True only for ASCII
    formats — PLY/PCD/LAS define their own layout, so roles can't be reassigned.
    """
    index: int
    header_name: Optional[str] = None
    detected_role: str
    suggested_label: str
    suggested_slug: str
    type_hint: str
    remappable: bool


class PointCloudPreviewResponse(BaseModel):
    kind: str                       # ascii | ply | pcd | las
    delimiter: Optional[str] = None  # comma | whitespace | tab | semicolon | None
    has_header: bool
    columns: List[PreviewColumn]
    sample_rows: List[List[str]]
    warning: Optional[str] = None
    # CloudCompare-style suggested global shift [x, y, z] = floor(min) per axis,
    # populated only when the cloud's coordinates are large enough that float32
    # rendering would lose precision (any |axis min| exceeds _SHIFT_SUGGEST_THRESHOLD).
    # null otherwise. The wizard pre-fills its shift fields from this (Z defaulted
    # off, since elevation is rarely huge). Best-effort: null if the min couldn't
    # be probed cheaply.
    suggested_shift: Optional[List[float]] = None


def _tokenize_ascii_format(fmt: str) -> List[str]:
    """Map each whitespace-separated <ASCII_format> token to a column role.

    A Helios <ASCII_format> (e.g. 'row col x y z r255 g255 b255 reflectance')
    is the scan's column legend. It's the *only* source of column meaning when
    the referenced .xyz has no header comment, so we resolve each token the same
    way a real header column name would resolve (`_role_from_header_name`): that
    accepts the richer alias set ('row'→row_index, 'col'→column_index,
    'red'→r255, 'reflectivity'→reflectance, …) the bare `_XYZ_KNOWN_ROLES`
    membership test missed — so 'row col …' no longer silently drops to 'skip'.

    Resolution order per token:
      1. `_role_from_header_name` alias (covers all the spellings above);
      2. a literal token already in `_XYZ_KNOWN_ROLES` (e.g. 'deviation',
         'is_miss', the canonical multi-return/grid slugs);
      3. otherwise the token text itself, lower-cased — carried through so
         `_plan_columns` can label/slug the extra column from the legend rather
         than dropping it. Only a blank token becomes 'skip'.
    """
    roles: List[str] = []
    for tok in fmt.split():
        low = tok.lower()
        mapped = _role_from_header_name(tok)
        if mapped is not None:
            roles.append(mapped)
        elif low in _XYZ_KNOWN_ROLES:
            roles.append(low)
        elif low:
            roles.append(low)
        else:
            roles.append('skip')
    return roles


def _first_nonblank_ascii_line(file_path: str) -> Optional[tuple]:
    """Return (line_text, was_commented) for the first meaningful line of an
    ASCII point file, or None if the file is empty/all-blank.

    `line_text` has any leading comment marker ('#' / '//') and surrounding
    whitespace stripped, so the caller can inspect a *commented* header the same
    way it would a bare one. Some exporters write the column legend as a comment
    (e.g. '# x y z r255 g255 b255 row column is_miss') so pandas — which is told
    comment='#' — skips it as data while a human still sees the labels. We
    recover those labels here. `was_commented` lets callers that drive pandas's
    `skiprows` know the line is already dropped by `comment='#'` and must NOT be
    counted again.
    """
    with open(file_path) as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            if line.startswith('#'):
                return line.lstrip('#').strip(), True
            if line.startswith('//'):
                return line[2:].strip(), True
            return line, False
    return None


def _role_from_header_name(name: str) -> Optional[str]:
    """Map a header column name to a known XYZ role, or None if unrecognised.

    Recognises the common terrestrial-scanner / Helios header conventions:
    'XYZ[0][m]'/'X' → x, 'Reflectance[dB]' → reflectance, 'Intensity' → intensity,
    'Red'/'R' → r255, etc. An unrecognised name returns None so the caller can
    carry it as an extra-dimension scalar (e.g. 'Deviation[]', 'Timestamp[s]').
    """
    # Indexed position headers: 'XYZ[0][m]' → x, 'XYZ[1]' → y, 'XYZ[2]' → z.
    m = re.match(r'\s*xyz\s*\[\s*([0-2])\s*\]', name.strip(), re.IGNORECASE)
    if m:
        return ('x', 'y', 'z')[int(m.group(1))]
    base = re.sub(r'\[.*?\]', '', name).strip().lower()
    base = re.sub(r'[^a-z0-9]+', '', base)
    if base in ('x', 'easting'):
        return 'x'
    if base in ('y', 'northing'):
        return 'y'
    if base in ('z', 'height', 'elevation'):
        return 'z'
    if base in ('r', 'red', 'r255', 'red255'):
        return 'r255'
    if base in ('g', 'green', 'g255', 'green255'):
        return 'g255'
    if base in ('b', 'blue', 'b255', 'blue255'):
        return 'b255'
    if base in ('intensity',):
        return 'intensity'
    if base in ('reflectance', 'reflectivity'):
        return 'reflectance'
    # Per-pulse multi-return columns Helios's LAD path needs. Recognise the
    # canonical names plus the common LAS aliases so a header-only ASCII export
    # round-trips them under the canonical slug (see `_MULTI_RETURN_SLUGS`).
    if base in ('timestamp', 'gpstime', 'time'):
        return 'timestamp'
    if base in ('targetindex', 'returnnumber'):
        return 'target_index'
    if base in ('targetcount', 'numberofreturns', 'numreturns'):
        return 'target_count'
    # Structured-scan raster indices (see `_GRID_INDEX_SLUGS`). Recognise the
    # canonical slugs plus the common row/col header spellings so a scan export
    # carrying its grid position round-trips into the recovery path.
    if base in ('rowindex', 'row', 'scanrow', 'scanrowindex', 'rasterrow'):
        return 'row_index'
    if base in ('columnindex', 'column', 'col', 'scancolumn', 'scancol',
                'scancolumnindex', 'rastercolumn'):
        return 'column_index'
    # Sky/miss flag (see `_MISS_ALIASES`). Pinned to the canonical `is_miss`
    # slug so a header-carrying ASCII export round-trips the column the LAD path
    # reads, matching the E57/structured-PLY recovery convention.
    if base in _MISS_ALIASES_NORMALISED:
        return _MISS_SLUG
    # Per-pulse beam-origin triple (ox/oy/oz and aliases, see
    # `_BEAM_ORIGIN_ALIAS_SETS`). A headered ASCII file with `ox oy oz` columns
    # auto-maps to the canonical origin_x/y/z roles so LAD uses them directly,
    # bypassing the trajectory join — exactly like the LAS ExtraBytes path. These
    # are world/UTM coordinates, captured at full float64 precision (NOT as float32
    # extras); the column-plan/streaming path keys off the canonical slug.
    origin_slug = _normalise_origin_alias(name)
    if origin_slug is not None:
        return origin_slug
    return None


def _autodetect_xyz_columns(file_path: str) -> List[str]:
    """Pick a column layout when the caller didn't supply <ASCII_format>.

    When the file has a header row (comma- or whitespace-delimited names with
    letters), map each header name to a known role via `_role_from_header_name`
    and leave unrecognised columns as 'skip' — the LAS writer's column plan
    then carries those as extra-dimension scalars under their header name.
    This is what makes a Pistachio-style export's Reflectance/Deviation/...
    columns colourable after import.

    Without a header, fall back to the positional convention: xyz first, then
    r255/g255/b255 if there are six columns, then intensity at column seven.
    Matches legacy Helios projects shipped without a format tag.
    """
    header = _read_ascii_header_names(file_path)
    if header is not None and len(header) >= 3:
        roles = [_role_from_header_name(h) or 'skip' for h in header]
        # Only trust the header mapping if it actually pinned x/y/z; otherwise
        # the names were too exotic and the positional fallback is safer.
        if all(r in roles for r in ('x', 'y', 'z')):
            return roles

    # Sample the first chunk of data rows: we need the column count AND, for the
    # RGB assumption, the actual value ranges of the candidate colour columns.
    sample: List[List[float]] = []
    ncols = 0
    with open(file_path) as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
            toks = line.split()
            # Skip a header row — use the first data row for the count.
            if any(re.search(r'[a-zA-Z]', tok) for tok in toks):
                continue
            if ncols == 0:
                ncols = len(toks)
            try:
                sample.append([float(t) for t in toks])
            except ValueError:
                # Ragged/garbage row — keep going; the count from the first
                # clean row still drives the layout.
                pass
            if len(sample) >= 64:
                break

    if ncols < 3:
        return ['x', 'y', 'z']

    # Where do the coordinates start? Most exports lead with xyz, but some
    # terrestrial-scanner ASCII dumps prepend grid bookkeeping — e.g. a scan
    # row/column index pair (`row col x y z r g b reflectance`). Those leading
    # columns are *entirely integers* while the coordinates carry fractional
    # precision, so we step past a run of leading pure-integer columns to find
    # x — but only while a later column with decimals exists to anchor xyz. If
    # column 0 already has decimals, or no column does (an all-integer cloud),
    # xyz starts at 0 as usual, so genuine integer-valued coordinates are never
    # mis-shifted.
    max_shift = min(2, ncols - 3)  # at most two leading index columns; leave xyz room
    xyz_start = 0
    while (xyz_start < max_shift
           and _columns_are_all_integers(sample, (xyz_start,))
           and _any_column_has_decimals(sample, range(xyz_start + 1, ncols))):
        xyz_start += 1

    # Recognised structured-scan index spellings for up to two leading columns.
    index_roles = ['row_index', 'column_index']
    roles: List[str] = [index_roles[i] if i < len(index_roles) else 'skip'
                        for i in range(xyz_start)]
    roles += ['x', 'y', 'z']
    rest_start = xyz_start + 3
    rest = ncols - rest_start

    # An RGB triple — three consecutive 0-255 integer columns — typically
    # follows xyz when present. Tagging it r255/g255/b255 saves the user three
    # manual reassignments; the range check (via `_columns_look_like_rgb255`)
    # rejects timestamp / return-count columns, so we never mislabel non-colour.
    if rest >= 3 and _columns_look_like_rgb255(
            sample, (rest_start, rest_start + 1, rest_start + 2)):
        roles += ['r255', 'g255', 'b255']
        after_rgb = rest_start + 3
        trailing = ncols - after_rgb
        # One trailing column after RGB is conventionally intensity/reflectance;
        # more than one is ambiguous, so leave them reassignable.
        if trailing == 1:
            roles += ['intensity']
        else:
            roles += ['skip'] * trailing
        return roles

    # No RGB: a lone trailing column is positionally intensity; anything wider
    # is carried as reassignable extras rather than guessed.
    if rest == 1:
        roles += ['intensity']
    else:
        roles += ['skip'] * rest
    return roles


def _columns_look_like_rgb255(sample: List[List[float]], idxs) -> bool:
    """Heuristic: do the given columns plausibly hold 8-bit RGB (0-255 ints)?

    Used by the headerless positional fallback to avoid tagging a timestamp or
    return-count column as 'red'. We require, across the sampled rows, that
    every candidate value is a non-negative integer within 0-255. We do NOT
    require the columns to span the full range (a uniform grey patch is valid
    colour), only that nothing falls outside it. With no sample to inspect we
    return False — better to leave columns unassigned than to guess wrong.
    """
    if not sample:
        return False
    for row in sample:
        for i in idxs:
            if i >= len(row):
                return False
            v = row[i]
            if v < 0 or v > 255 or v != int(v):
                return False
    return True


def _columns_are_all_integers(sample: List[List[float]], idxs) -> bool:
    """Do the given columns hold only integer-valued numbers across the sample?

    A leading scan-pattern row/column index is an integer counter, whereas
    coordinate columns carry fractional precision. We require every sampled
    value in each column to be exactly integral. An out-of-range index or an
    empty sample returns False — without evidence we don't claim integrality.
    """
    if not sample:
        return False
    for row in sample:
        for i in idxs:
            if i >= len(row):
                return False
            if row[i] != int(row[i]):
                return False
    return True


def _any_column_has_decimals(sample: List[List[float]], idxs) -> bool:
    """Does any of the given columns carry fractional precision in the sample?

    Used to confirm that real (fractional) coordinates lie ahead before
    shifting xyz past a run of pure-integer leading index columns. Real LiDAR
    coordinates almost always have sub-unit decimals; an all-integer file has
    none, so we leave xyz at column 0 rather than mistake a coordinate column
    for an index. Out-of-range indices are skipped; an empty sample is False.
    """
    if not sample:
        return False
    for row in sample:
        for i in idxs:
            if i < len(row) and row[i] != int(row[i]):
                return True
    return False


def _first_data_row_has_letters(file_path: str) -> bool:
    """Detect a leading *uncommented* text header row that pandas must skip.

    Helios scan files don't have one, but plain XYZ exports from other tools
    sometimes do (e.g. 'X Y Z R G B'). We skip such a row so pandas can read
    the rest as floats without falling off the C engine.

    A *commented* header ('# X Y Z ...') is NOT counted here: the readers pass
    comment='#' to pandas, which already drops it, so adding skiprows on top
    would eat the first real data row. We still recover its labels via
    `_read_ascii_header_names`."""
    first = _first_nonblank_ascii_line(file_path)
    if first is None:
        return False
    line, was_commented = first
    if was_commented:
        return False
    return any(re.search(r'[a-zA-Z]', tok) for tok in line.split())


def _ascii_pandas_sep(file_path: str) -> str:
    """pandas `sep` for an ASCII xyz-family file, from its sniffed delimiter.

    The wizard preview splits rows with `_detect_ascii_delimiter` /
    `_split_ascii_row` (comma → tab → semicolon → whitespace), but the loaders
    historically hardcoded `sep=r'\\s+'`. That silently broke comma- /
    semicolon-delimited exports (e.g. CloudCompare's `//X,Y,Z,...` files): the
    whole row landed in column 0 and `usecols=[1, 2, ...]` raised "Usecols do
    not match columns". Resolve the same delimiter the preview used so the
    load path agrees with what the user saw.

    Returns a regex/literal suitable for pandas `sep=`. Whitespace (and the
    no-data fallback) stays `r'\\s+'` to collapse runs of spaces/tabs.
    """
    delim = _detect_ascii_delimiter(file_path)
    return {'comma': ',', 'tab': '\t', 'semicolon': ';'}.get(delim, r'\s+')


def _ascii_skiprows(file_path: str) -> int:
    """How many leading lines pandas must be told to skip explicitly.

    Every ASCII reader here passes `comment='#'` to pandas, which drops blank
    lines and any line starting with '#' on its own — so a '#'-commented header
    needs no `skiprows`. Two cases pandas WON'T handle and we must:

      1. An uncommented text header ('X Y Z R G B') — `comment='#'` doesn't
         touch it, so it would be read as a data row and crash the float cast.
      2. A '//'-commented header ('//X //Y //Z ...', CloudCompare's convention).
         pandas's comment char is a single character; we pass '#', not '//', so
         the line survives and pandas tries to parse '//X' as a float. Counting
         it here is what keeps those exports importable.

    Returns the count of such leading lines (currently 0 or 1 — these exporters
    write a single legend row). A '#'-commented header returns 0 because pandas
    already drops it; double-counting would eat the first real data row.
    """
    first = _first_nonblank_ascii_line(file_path)
    if first is None:
        return 0
    line, was_commented = first
    if was_commented:
        # '#' headers are dropped by pandas's comment='#'; only '//' ones (which
        # pandas keeps) need an explicit skip. Re-read the raw first non-blank
        # line to tell the two markers apart — `line` here is already stripped.
        with open(file_path) as f:
            for raw in f:
                stripped = raw.strip()
                if not stripped:
                    continue
                return 1 if stripped.startswith('//') else 0
        return 0
    # Bare (uncommented) line: skip it only if it's a text header, not data.
    return 1 if any(re.search(r'[a-zA-Z]', tok) for tok in line.split()) else 0


# Roles that already map to a dedicated LAS field — never carried as extra
# dimensions (they'd duplicate position/rgb/intensity in the octree).
_XYZ_RESERVED_ROLES = {
    'x', 'y', 'z', 'r', 'g', 'b', 'r255', 'g255', 'b255',
    'intensity', 'reflectance', 'skip',
}


def _sanitize_extra_dim_name(raw: str) -> str:
    """Slugify a source header name into a laspy-safe LAS extra-dimension name.

    laspy/LAS extra dimensions accept a restricted name set: we keep
    [A-Za-z0-9_], collapse every other run to a single underscore, trim
    leading/trailing underscores, and cap at the 32-char LAS limit. A name
    that sanitises to empty (e.g. all punctuation) falls back to 'field'.
    Callers dedupe collisions; this function is deterministic per input.

    Examples:
      'Reflectance[dB]' -> 'Reflectance_dB'
      'Target Index[]'  -> 'Target_Index'
      'XYZ[0][m]'       -> 'XYZ_0_m'
    """
    slug = re.sub(r'[^A-Za-z0-9_]+', '_', raw).strip('_')
    if not slug:
        slug = 'field'
    return _avoid_reserved_las_dim(slug[:32])


# LAS point format 3 (and the standard dimensions every format carries) reserve
# these names. An extra dimension may NOT reuse one — laspy's dtype build fails
# with "field '<name>' occurs more than once". A user-renamed scalar, or a
# source column literally named "Intensity"/"Classification", can sanitise onto
# one of these, so we rename the collision. Matched case-insensitively because
# laspy lower-cases extra-dim names internally.
_LAS_RESERVED_DIM_NAMES = {
    'x', 'y', 'z', 'intensity', 'bit_fields', 'raw_classification',
    'classification', 'classification_flags', 'scan_angle_rank', 'scan_angle',
    'user_data', 'point_source_id', 'gps_time', 'red', 'green', 'blue', 'nir',
    'return_number', 'number_of_returns', 'scan_direction_flag',
    'edge_of_flight_line', 'synthetic', 'key_point', 'withheld', 'overlap',
    'scanner_channel',
}


def _avoid_reserved_las_dim(slug: str) -> str:
    """Rename an extra-dimension slug that collides (case-insensitively) with a
    built-in LAS dimension, so `header.add_extra_dim` can't crash the converter.
    'Intensity' -> 'Intensity_field', 'classification' -> 'classification_field'."""
    if slug.lower() in _LAS_RESERVED_DIM_NAMES:
        suffixed = f"{slug}_field"
        return suffixed[:32]
    return slug


def _humanize_extra_dim_label(raw: str) -> str:
    """Tidy a source header into a human-readable picker label.

    Keeps unit brackets but normalises whitespace and drops empty brackets:
      'Reflectance[dB]' -> 'Reflectance [dB]'
      'Target Index[]'  -> 'Target Index'
      'Deviation[]'     -> 'Deviation'
    Falls back to the raw string (stripped) when nothing else applies.
    """
    s = raw.strip()
    # Drop empty bracket pairs entirely ('Deviation[]' -> 'Deviation').
    s = re.sub(r'\[\s*\]', '', s)
    # Space out a unit bracket that abuts a word ('Reflectance[dB]' ->
    # 'Reflectance [dB]') without touching brackets already preceded by space.
    s = re.sub(r'(?<=\S)\[', ' [', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s or raw.strip()


def _read_ascii_header_names(file_path: str) -> Optional[List[str]]:
    """Return the column display names from a leading header row, if present.

    XYZ-family files from terrestrial scanners often carry a comma-delimited
    header (e.g. 'XYZ[0][m],XYZ[1][m],...,Reflectance[dB],Deviation[]') even
    though the data rows are whitespace-delimited. We only treat the first
    non-blank, non-comment line as a header when it actually contains letters
    (mirrors `_first_data_row_has_letters`); otherwise there are no names to
    recover and we return None.

    Splits on commas when the line is comma-delimited, else on whitespace, so
    multi-word names like 'Target Index[]' survive in the comma case.

    A *commented* first line ('# x y z ...') is also accepted as a header, but
    only when its tokens actually resolve to x/y/z roles — that guard keeps a
    plain prose comment ('# exported by FooScan v2') from being mistaken for
    column names.
    """
    first = _first_nonblank_ascii_line(file_path)
    if first is None:
        return None
    line, was_commented = first
    if not any(re.search(r'[a-zA-Z]', tok) for tok in line.split()):
        return None
    parts = line.split(',') if ',' in line else line.split()
    names = [p.strip() for p in parts]
    if was_commented:
        # Be conservative with commented lines: only trust them as a header when
        # they map to a real x/y/z layout, else they're just a remark.
        roles = [_role_from_header_name(n) for n in names]
        if not all(r in roles for r in ('x', 'y', 'z')):
            return None
    return names


# A pandas `names=` list can't contain a repeated value, so a column plan that
# skips more than one column can't use bare 'skip' for each. These helpers give
# every skipped column a unique placeholder that the readers still recognise and
# drop via `usecols`. Bare 'skip' (from the auto-detect path, which never repeats
# it) is also treated as a skip for backwards compatibility.
def _skip_name(pos: int) -> str:
    return f"skip:{pos}"


def _is_skip_name(name: str) -> bool:
    return name == 'skip' or name.startswith('skip:')


# A beam-origin column rides in the pandas `names` list as `origin:<canonical>`
# (origin:origin_x/_y/_z). It is NOT a skip name (so usecols reads it) and NOT an
# `extra:` dim (origins are float64, never float32 extras) — the streaming reader
# captures these three columns into the float64 origin side-channel by name.
def _is_origin_name(name: str) -> bool:
    return name.startswith('origin:')


def _origin_slug_of(name: str) -> str:
    """Canonical origin slug ('origin_x'/'origin_y'/'origin_z') for an
    `origin:<canonical>` column token."""
    return name.split(':', 1)[1]


def _dedupe_slug(base: str, used_slugs: "set[str]") -> str:
    """Return `base` (or a numbered variant) not yet in `used_slugs`, capped at
    the 32-char LAS limit. Mutates `used_slugs` to reserve the result."""
    slug = base
    n = 2
    while slug in used_slugs:
        suffix = f"_{n}"
        slug = base[:32 - len(suffix)] + suffix
        n += 1
    used_slugs.add(slug)
    return slug


def _plan_columns(roles: List[str], header_names: Optional[List[str]]):
    """Turn a per-column role list into a pandas `names` list + extra_dims.

    Shared by `_xyz_column_plan` (the import/convert path) and the preview
    endpoint so the slugs/labels a user sees in the wizard are byte-identical to
    what import produces. A column whose role is reserved (x/y/z/rgb/intensity/
    reflectance) keeps its role token; a multi-return role (timestamp/
    target_index/target_count) becomes an 'extra:<canonical slug>' so the LAD
    path can find it deterministically; anything else ('skip', deviation, …)
    becomes an 'extra:<slug>' carried into the octree, named from the header row
    when present, else a positional 'Column N' fallback.

    Returns (names, extra_dims) where each extra dim is
    {col, slug, label, categorical} ('categorical' defaults False here; the
    structured column-plan path can override it).
    """
    names: List[str] = []
    extra_dims: List[dict] = []
    used_slugs: "set[str]" = set()
    for i, role in enumerate(roles):
        if role in _XYZ_RESERVED_ROLES and role != 'skip':
            names.append(role)
            continue
        if role in _MULTI_RETURN_SLUGS:
            # Pin the canonical slug/label regardless of header text so the LAD
            # accessor can recover these columns by name.
            slug = _dedupe_slug(role, used_slugs)
            col_id = f"extra:{slug}"
            names.append(col_id)
            extra_dims.append({"col": col_id, "slug": slug,
                               "label": _MULTI_RETURN_LABELS[role],
                               "categorical": False})
            continue
        if role in _GRID_INDEX_SLUGS:
            # Structured-scan raster index: pin the canonical slug/label so the
            # grid-based miss-recovery path finds the raster by name (same as the
            # multi-return slugs above, and what the E57 structured path emits).
            slug = _dedupe_slug(role, used_slugs)
            col_id = f"extra:{slug}"
            names.append(col_id)
            extra_dims.append({"col": col_id, "slug": slug,
                               "label": _GRID_INDEX_LABELS[role],
                               "categorical": False})
            continue
        if role == _MISS_SLUG:
            # Sky/miss flag: pin the canonical is_miss slug/label regardless of
            # the source header spelling (is_miss/miss/sky) so the LAD path and
            # renderer find it by name, matching the E57/PLY recovery convention.
            slug = _dedupe_slug(_MISS_SLUG, used_slugs)
            col_id = f"extra:{slug}"
            names.append(col_id)
            extra_dims.append({"col": col_id, "slug": slug,
                               "label": _MISS_LABEL, "categorical": False})
            continue
        if role in _ORIGIN_SLUGS:
            # Per-pulse beam-origin column (auto-detected from an ox/oy/oz header,
            # see `_role_from_header_name`). World/UTM coordinates needing full
            # float64 precision, so they are NOT carried as a float32 extra dim;
            # the `origin:<canonical>` sentinel token tells the streaming reader to
            # capture this column into the float64 origin side-channel instead (see
            # `_xyz_to_las_stream`). Deliberately NOT appended to `extra_dims`.
            names.append(f"origin:{role}")
            continue
        if header_names is not None and i < len(header_names) and header_names[i]:
            label = _humanize_extra_dim_label(header_names[i])
            base = _sanitize_extra_dim_name(header_names[i])
        elif role != 'skip':
            # No file header, but the role token carries meaning — it came from a
            # Helios <ASCII_format> legend word (e.g. 'deviation', or any custom
            # scalar name) that `_tokenize_ascii_format` passed through verbatim.
            # Use it as the column's label/slug so the legend names the column,
            # instead of a meaningless positional 'Column N'.
            label = _humanize_extra_dim_label(role)
            base = _sanitize_extra_dim_name(role)
        else:
            label = f"Column {i + 1}"
            base = f"col_{i + 1}"
        slug = _dedupe_slug(base, used_slugs)
        col_id = f"extra:{slug}"
        names.append(col_id)
        extra_dims.append({"col": col_id, "slug": slug, "label": label,
                           "categorical": False})
    return names, extra_dims


def _plan_columns_from_column_plan(column_plan: "ColumnPlan"):
    """Build (names, extra_dims) directly from a wizard-supplied ColumnPlan.

    Honours each entry's explicit role, and for role=='extra' the user's custom
    slug/label/categorical. Falls back to a sane slug when the wizard omitted
    one. Reserved roles keep their token; a multi-return role (or an extra
    slugged as one) is pinned to its canonical slug/label so the LAD path finds
    it; 'skip' (or an extra with no usable name) is dropped. Slugs are
    de-duplicated so two custom names can't collide on disk.
    """
    names: List[str] = []
    extra_dims: List[dict] = []
    used_slugs: "set[str]" = set()
    for pos, entry in enumerate(sorted(column_plan.columns, key=lambda e: e.index)):
        role = (entry.role or 'skip').lower()
        if role in _XYZ_RESERVED_ROLES and role != 'skip':
            names.append(role)
            continue
        # A column mapped to a multi-return field — either by naming the role
        # token directly or by slugging an 'extra' as one — is carried under the
        # canonical slug/label so the LAD accessor can recover it by name. But a
        # column the user explicitly marked categorical (the wizard's 'Label'
        # role) is a discrete class field by intent: multi-return per-pulse
        # values (timestamp/target index/count) are never categorical, so honour
        # the user's choice and let it fall through to the generic extra-dim path
        # (preserving its sanitised slug + categorical flag) instead of diverting
        # it into the LAD canonicalisation, which would lower-case the slug and
        # relabel it as a per-pulse field.
        mr = role if role in _MULTI_RETURN_SLUGS else (entry.slug or '').lower()
        if mr in _MULTI_RETURN_SLUGS and not entry.categorical:
            slug = _dedupe_slug(mr, used_slugs)
            col_id = f"extra:{slug}"
            names.append(col_id)
            extra_dims.append({"col": col_id, "slug": slug,
                               "label": (entry.label or '').strip() or _MULTI_RETURN_LABELS[mr],
                               "categorical": bool(entry.categorical)})
            continue
        # Structured-grid indices (scan row / column): carry under the canonical
        # `row_index`/`column_index` slug so the LAD/recovery path finds the
        # raster by name — same treatment as the multi-return slugs above, and
        # matching what the E57 structured-import path emits. These are integer
        # grid positions, never categorical class fields.
        gi = role if role in _GRID_INDEX_SLUGS else (entry.slug or '').lower()
        if gi in _GRID_INDEX_SLUGS:
            slug = _dedupe_slug(gi, used_slugs)
            col_id = f"extra:{slug}"
            names.append(col_id)
            extra_dims.append({"col": col_id, "slug": slug,
                               "label": (entry.label or '').strip() or _GRID_INDEX_LABELS[gi],
                               "categorical": False})
            continue
        # Sky/miss flag: a column whose role token or slug names a miss alias
        # (is_miss/miss/sky) carries under the canonical is_miss slug so the LAD
        # path and renderer find it by name — matching the E57/PLY convention.
        # Never categorical (it's a 0/1 flag the LAD check reads as a continuous
        # value), so honour it regardless of the wizard's categorical toggle.
        ms = role if role == _MISS_SLUG else _normalise_miss_alias(entry.slug or '')
        if ms == _MISS_SLUG:
            slug = _dedupe_slug(_MISS_SLUG, used_slugs)
            col_id = f"extra:{slug}"
            names.append(col_id)
            extra_dims.append({"col": col_id, "slug": slug,
                               "label": (entry.label or '').strip() or _MISS_LABEL,
                               "categorical": False})
            continue
        # Per-pulse beam-origin column (the wizard's Beam Origin X/Y/Z role, or an
        # 'extra' whose slug normalises to an ox/oy/oz alias). These are world/UTM
        # emission coordinates needing FULL float64 precision, so — unlike every
        # other branch above — they must NOT become a float32 extra dim (the LAS is
        # 1 mm-quantized and extras are float32; both would shatter them). Instead
        # emit an `origin:<canonical>` sentinel token so the streaming reader keeps
        # the column (it's not a skip name, so pandas/usecols read it) and captures
        # it into the float64 origin side-channel, exactly like `capture_full_xyz`
        # does for positions. Deliberately NOT appended to `extra_dims`.
        ori = role if role in _ORIGIN_SLUGS else (_normalise_origin_alias(entry.slug or '') or '')
        if ori in _ORIGIN_SLUGS:
            names.append(f"origin:{ori}")
            continue
        if role == 'skip' or (role == 'extra' and not (entry.slug or entry.label)):
            # A skipped column, or an 'extra' with no usable name, carries
            # nothing. Use a UNIQUE placeholder (not bare 'skip') so pandas
            # doesn't reject a names= list with repeated 'skip' entries; the
            # readers drop these via _is_skip_name.
            names.append(_skip_name(pos))
            continue
        # role == 'extra' (or any unreserved token) → carry as an extra dim.
        # A categorical override of a multi-return column reaches here with the
        # wizard's lower-cased canonical slug (e.g. 'target_index'); prefer the
        # human label so it slugs to a readable class-field name ('Target_Index')
        # rather than the LAD canonical it's no longer being used as.
        slug_source = entry.slug
        if entry.categorical and (entry.slug or '').lower() in _MULTI_RETURN_SLUGS:
            slug_source = entry.label or entry.slug
        base = _sanitize_extra_dim_name(slug_source or entry.label or f"col_{entry.index + 1}")
        slug = _dedupe_slug(base, used_slugs)
        label = (entry.label or "").strip() or _humanize_extra_dim_label(slug)
        col_id = f"extra:{slug}"
        names.append(col_id)
        extra_dims.append({"col": col_id, "slug": slug, "label": label,
                           "categorical": bool(entry.categorical)})
    return names, extra_dims


def _xyz_column_plan(source_path: "_Path", ascii_format: Optional[str],
                     column_plan: "Optional[ColumnPlan]" = None):
    """Resolve the column layout for an XYZ-family file, carrying unmapped
    numeric columns as octree extra dimensions.

    Returns (names, extra_dims) where:
      - `names` is the per-column identifier list for pandas `names=` — known
        roles keep their role token (x/y/z/r255/intensity/...), extras get a
        unique 'extra:<slug>' identifier, and truly droppable columns (no
        header name, beyond the recognised layout) stay 'skip'.
      - `extra_dims` is an ordered list of dicts {col, slug, label, categorical}
        for each carried extra column, where `col` is the matching entry in
        `names`.

    When `column_plan` is supplied (the import wizard's explicit choices) it
    fully determines the layout — roles, custom slugs/labels, and the
    categorical flag — bypassing header/format sniffing. When it's None,
    behaviour is exactly as before: role tokens come from
    `_tokenize_ascii_format` / `_autodetect_xyz_columns`, and extras are named
    from the header row (or a positional 'Column N' fallback).
    """
    if column_plan is not None:
        return _plan_columns_from_column_plan(column_plan)

    roles = (_tokenize_ascii_format(ascii_format)
             if ascii_format
             else _autodetect_xyz_columns(str(source_path)))
    header_names = _read_ascii_header_names(str(source_path))
    return _plan_columns(roles, header_names)


def _pack_pointcloud_response(positions: np.ndarray,
                              colors: Optional[np.ndarray],
                              intensity: Optional[np.ndarray]) -> StreamingResponse:
    """Stream a packed binary response matching the renderer decoder.

    Response layout (little-endian):
      offset  size   field
      0       4      magic 'PHX1'
      4       4      uint32 point_count
      8       1      uint8  has_colors (0/1)
      9       1      uint8  has_intensity (0/1)
      10      22     reserved (zero)
      32      pc*12  float32 positions [x0,y0,z0,x1,y1,z1,...]
      ...     pc*12  float32 colors    [r0,g0,b0,...]      (if has_colors)
      ...     pc*4   float32 intensity [i0,i1,...]         (if has_intensity)
    """
    import struct

    point_count = int(positions.shape[0])
    header = struct.pack(
        '<4sIBB22x', _POINTCLOUD_BIN_MAGIC, point_count,
        1 if colors is not None else 0,
        1 if intensity is not None else 0,
    )

    def chunks():
        yield header
        yield np.ascontiguousarray(positions, dtype=np.float32).tobytes()
        if colors is not None:
            yield np.ascontiguousarray(colors, dtype=np.float32).tobytes()
        if intensity is not None:
            yield np.ascontiguousarray(intensity, dtype=np.float32).tobytes()

    return StreamingResponse(chunks(), media_type='application/octet-stream')


# ==================== GENERIC BINARY FRAME (PHB1) ====================
# Large array responses (meshes, scans, voxel grids) are sent as a binary frame
# instead of JSON: ~3-4x smaller, no per-number text parsing, and no V8 ~512 MB
# string-length ceiling (which broke big triangulations). Layout (little-endian):
#   0            4         magic 'PHB1'
#   4            4         uint32 header_len
#   8            header_len JSON header, space-padded to a multiple of 4:
#                            {"meta": {...small scalars/lists...},
#                             "buffers": [{"name","dtype":"f32"|"u32","length"}]}
#   8+header_len ...        the buffers, concatenated in header order
# magic+len (8) + padded header keep every buffer at a 4-byte-aligned offset, so
# the renderer makes zero-copy Float32Array/Uint32Array views. `meta` carries
# everything that isn't a big array (counts, estimate, grid, warnings, errors).
_BIN_FRAME_MAGIC = b"PHB1"


def _json_safe(o):
    """json.dumps default= for stray numpy scalars in `meta`."""
    if hasattr(o, "item"):
        return o.item()
    if isinstance(o, np.ndarray):
        return o.tolist()
    raise TypeError(f"not JSON-serializable: {type(o)}")


def _bin_frame_bytes(meta: dict, buffers: "list[tuple]") -> bytes:
    """Pack a PHB1 frame. `buffers` is a list of (name, array, dtype) where dtype
    is 'f32' or 'u32'; arrays are raveled to little-endian contiguous bytes."""
    import struct
    descs = []
    payloads = []
    for name, arr, dtype in buffers:
        np_dtype = np.float32 if dtype == "f32" else np.uint32
        a = np.ascontiguousarray(arr, dtype=np_dtype).ravel()
        payloads.append(a.tobytes())
        descs.append({"name": name, "dtype": dtype, "length": int(a.size)})
    header = json.dumps({"meta": meta, "buffers": descs}, default=_json_safe).encode("utf-8")
    header += b" " * ((-len(header)) % 4)  # pad so buffers land on 4-byte offsets
    out = bytearray()
    out += _BIN_FRAME_MAGIC
    out += struct.pack("<I", len(header))
    out += header
    for p in payloads:
        out += p
    return bytes(out)


# Progress markers ride the same octet stream as the keepalives, ahead of the
# PHB1 frame. Layout (little-endian):
#   0  4              magic 'PHP1'
#   4  4              uint32 json_len
#   8  json_len       JSON {"progress": <float|null>, "message": "<str>"}
#                     padded with spaces so the whole marker is a 4-byte multiple
# The renderer skips/parses leading PHP1 markers (and whitespace keepalives)
# before the PHB1 magic, so the frame's buffers stay 4-byte aligned for
# zero-copy decode.
_PROGRESS_MARKER_MAGIC = b"PHP1"


def _pack_progress_marker(progress, message: str, *, run_id=None, cancelled=False) -> bytes:
    """Pack one PHP1 progress marker. `progress` is a 0..1 fraction or None.

    `run_id` (when set) rides the first marker so the renderer learns the
    cancellation token before any heavy work starts; `cancelled` rides the
    terminal marker emitted in place of a frame when a run is cancelled."""
    import struct
    obj = {"progress": progress, "message": message}
    if run_id is not None:
        obj["run_id"] = run_id
    if cancelled:
        obj["cancelled"] = True
    payload = json.dumps(obj).encode("utf-8")
    payload += b" " * ((-len(payload)) % 4)  # keep total marker a 4-byte multiple
    return _PROGRESS_MARKER_MAGIC + struct.pack("<I", len(payload)) + payload


# ---- Cancellation registry -------------------------------------------------
# Long-running streaming ops (synthetic scan, triangulation, LAD inversion) run
# off-thread in an executor; the heavy C++/Open3D primitives are monolithic, so
# cancellation is cooperative — the worker polls a per-run threading.Event at
# every stage boundary and raises ScanCancelled, which unwinds its `with
# Context()/LiDARCloud()` blocks and frees the multi-GB C++/numpy memory
# promptly. A run is cancelled either by the client POSTing /api/cancel/{run_id}
# or by the client disconnecting (detected in _bin_frame_streaming_response).
class ScanCancelled(Exception):
    """Raised inside a streaming worker when its run has been cancelled."""


_CANCEL_REGISTRY: "dict[str, threading.Event]" = {}
_CANCEL_REGISTRY_LOCK = threading.Lock()


def _cancel_checkpoint(progress) -> None:
    """Raise ScanCancelled if `progress` is a reporter whose run was cancelled.

    Workers call this at stage boundaries; raising unwinds their `with Context()/
    LiDARCloud()` blocks so the C++/numpy memory is freed promptly. A no-op when
    progress is None (e.g. unit tests calling a worker directly)."""
    if progress is not None and getattr(progress, "should_cancel", None) and progress.should_cancel():
        raise ScanCancelled()


def _new_cancel_token() -> "tuple[str, threading.Event]":
    """Mint a run_id + its cancel Event and register them."""
    run_id = uuid.uuid4().hex
    event = threading.Event()
    with _CANCEL_REGISTRY_LOCK:
        _CANCEL_REGISTRY[run_id] = event
    return run_id, event


def _cancel_run(run_id: str) -> bool:
    """Signal cancellation for run_id. Returns True if the run was known."""
    with _CANCEL_REGISTRY_LOCK:
        event = _CANCEL_REGISTRY.get(run_id)
    if event is not None:
        event.set()
        return True
    return False


def _clear_run(run_id: str) -> None:
    """Drop a run from the registry once its stream has finished."""
    with _CANCEL_REGISTRY_LOCK:
        _CANCEL_REGISTRY.pop(run_id, None)


class ClientDisconnected(Exception):
    """Raised by `_run_blocking_until_disconnect` when the HTTP client goes away
    (panel closed, fetch AbortController timeout) before the blocking work
    returns."""


async def _run_blocking_until_disconnect(fn, http_request: "Optional[Request]" = None,
                                         poll: float = 0.25):
    """Run a blocking callable off the event loop and stop awaiting it the moment
    the client disconnects.

    For a long, CPU-bound, *non-streaming* JSON endpoint (e.g. TreeIso tree
    segmentation), calling the worker directly inside the `async def` pins a CPU
    core AND blocks the event loop, so the server can't service other requests
    and a closed panel / fetch-timeout leaves the request hanging until the work
    finishes. This runs `fn` in the default thread-pool executor instead and
    polls `http_request.is_disconnected()`; on disconnect it raises
    `ClientDisconnected` so the caller returns at once.

    Honest limitation: the orphaned thread keeps running to completion (TreeIso's
    cut-pursuit C-extension exposes no cancel hook and Python threads can't be
    force-killed); its result is simply discarded. What this *does* buy is an
    unblocked event loop and a request that no longer hangs after the client is
    gone — the worker is no longer wedged onto the live request. The auto-scaled
    decimation already bounds the runtime, so the orphaned thread is short-lived."""
    import asyncio

    loop = asyncio.get_event_loop()
    fut = loop.run_in_executor(None, fn)
    if http_request is None:
        return await fut
    while not fut.done():
        if await http_request.is_disconnected():
            # Detach: the executor thread finishes on its own and its result is
            # GC'd. We surface the disconnect so the endpoint returns promptly.
            raise ClientDisconnected()
        await asyncio.sleep(poll)
    return await fut


class _ProgressReporter:
    """Callable progress reporter handed to streaming workers.

    Calling it (`progress(fraction, message)`) queues a PHP1 marker for the
    stream. `should_cancel()` / `cancelled` let the worker poll its run's cancel
    Event at stage boundaries and raise ScanCancelled to unwind promptly.

    `cancel_int` is an optional ctypes c_int shared with the C++ ray loop: a
    worker registers one via `bind_cancel_int()` and the stream loop flips it to
    1 the instant the cancel Event fires, so the in-flight C++ trace (which holds
    no GIL and can't poll the Python Event) sees the cancel and bails its loop."""

    def __init__(self, progress_queue, cancel_event):
        self._queue = progress_queue
        self._cancel_event = cancel_event
        self._cancel_int = None  # ctypes.c_int, set by bind_cancel_int()

    def __call__(self, fraction, message: str) -> None:
        self._queue.put((fraction, message))

    def should_cancel(self) -> bool:
        return self._cancel_event is not None and self._cancel_event.is_set()

    @property
    def cancelled(self) -> bool:
        return self.should_cancel()

    def raise_if_cancelled(self) -> None:
        if self.should_cancel():
            raise ScanCancelled()

    def bind_cancel_int(self, cancel_int) -> None:
        """Register the ctypes c_int the C++ ray loop polls for this run."""
        self._cancel_int = cancel_int

    def propagate_cancel(self) -> None:
        """Mirror a set Event into the shared c_int (called from stream loop)."""
        if self._cancel_int is not None and self.should_cancel():
            self._cancel_int.value = 1


def _bin_frame_streaming_response(
    build_frame,
    *,
    request: "Optional[Request]" = None,
    cancel_event: "Optional[threading.Event]" = None,
    run_id: "Optional[str]" = None,
) -> StreamingResponse:
    """Run build_frame() off-thread, emitting keepalives until it's done (so
    WebKit's ~60s stall timeout doesn't fire on long computations), then the PHB1
    frame. Keepalives are 4-byte whitespace chunks so the leading padding stays a
    multiple of 4; the renderer skips them and the frame's buffers remain 4-byte
    aligned for zero-copy decode.

    `build_frame` may be either a zero-arg callable (returns PHB1 bytes) or a
    one-arg callable taking a `progress` reporter (a _ProgressReporter). When it
    takes the reporter, queued progress updates are flushed as PHB1-compatible
    PHP1 markers in place of blank keepalives, so the renderer can surface real
    per-stage progress.

    Cancellation: when `cancel_event`/`run_id` are supplied, the run_id rides the
    first PHP1 marker (so the client can target /api/cancel/{run_id}), the stream
    loop also sets the event if the client disconnects, and a worker that raises
    ScanCancelled yields a terminal `cancelled` marker instead of a frame."""
    import asyncio
    import inspect
    import queue as _queue

    wants_progress = len(inspect.signature(build_frame).parameters) >= 1
    progress_queue: "_queue.Queue" = _queue.Queue()
    reporter = _ProgressReporter(progress_queue, cancel_event)

    def _run():
        return build_frame(reporter) if wants_progress else build_frame()

    # Poll the executor future frequently in BOTH modes so the finished frame is
    # flushed as soon as the off-thread build returns — `poll` is the re-check
    # granularity, NOT the keepalive cadence. (Keepalives are gated separately by
    # `silence_keepalive_after`.) In progress mode the fine poll also lets each
    # stage flush as its own chunk for a smooth bar. The earlier non-progress
    # value of 5.0 conflated the two: a build that finished mid-`sleep(5.0)` —
    # e.g. the misses overlay, whose frame is prebuilt before the wrapper even
    # runs — sat idle for the rest of the interval, adding a ~5 s stall to every
    # fast binary-frame response. A 0.1 s poll while waiting costs nothing.
    poll = 0.02 if wants_progress else 0.1
    silence_keepalive_after = 5.0

    async def stream():
        loop = asyncio.get_event_loop()
        # Emit the run_id up front so the client can cancel before heavy work.
        if run_id is not None:
            yield _pack_progress_marker(None, "", run_id=run_id)
        fut = loop.run_in_executor(None, _run)
        silent_for = 0.0
        disconnect_checked = 0.0
        try:
            while not fut.done():
                emitted = False
                try:
                    while True:
                        fraction, message = progress_queue.get_nowait()
                        yield _pack_progress_marker(fraction, message)
                        emitted = True
                except _queue.Empty:
                    pass
                # Detect client disconnect (~every 0.25s) and flip the cancel
                # event so the worker unwinds and frees C++/numpy memory.
                if request is not None and cancel_event is not None:
                    disconnect_checked += poll
                    if disconnect_checked >= 0.25:
                        disconnect_checked = 0.0
                        if await request.is_disconnected():
                            cancel_event.set()
                # When the cancel Event is set (here, or via /api/cancel on
                # another request), mirror it into the C++ shared cancel int so
                # the in-flight ray loop — which can't poll the Python Event —
                # bails. Cheap and idempotent; runs every poll tick.
                if cancel_event is not None and cancel_event.is_set():
                    reporter.propagate_cancel()
                if emitted:
                    silent_for = 0.0
                else:
                    silent_for += poll
                    if silent_for >= silence_keepalive_after:
                        yield b"    "
                        silent_for = 0.0
                await asyncio.sleep(poll)
            # Drain any progress queued in the final tick before the frame.
            try:
                while True:
                    fraction, message = progress_queue.get_nowait()
                    yield _pack_progress_marker(fraction, message)
            except _queue.Empty:
                pass
            try:
                yield await fut
            except ScanCancelled:
                # Cooperative cancel landed: the worker unwound its Context/
                # LiDARCloud (memory freed). Tell the client instead of a frame.
                yield _pack_progress_marker(None, "Cancelled", cancelled=True)
        finally:
            if run_id is not None:
                _clear_run(run_id)

    return StreamingResponse(stream(), media_type="application/octet-stream")


def _load_xyz_arrays(
    file_path: str, ascii_format: Optional[str],
    column_plan: "Optional[ColumnPlan]" = None,
) -> tuple[np.ndarray, Optional[np.ndarray], Optional[np.ndarray]]:
    """Parse an ASCII xyz-family file via pandas and return numpy arrays.

    Returns (positions[N,3] float32, colors[N,3] float32 | None,
    intensity[N] float32 | None). Separated from the response-packing step
    so the crop endpoint can reuse the loader and apply a boolean mask
    before responding.

    When `column_plan` is supplied (import wizard) it determines the column
    roles and whether r/g/b are 0-255 ints or 0-1 floats; otherwise roles come
    from the Helios `ascii_format` hint or are auto-detected.
    """
    if column_plan is not None:
        columns = [(e.role or 'skip').lower()
                   for e in sorted(column_plan.columns, key=lambda e: e.index)]
        rgb_is_255 = column_plan.rgb_is_255
    else:
        columns = (_tokenize_ascii_format(ascii_format)
                   if ascii_format
                   else _autodetect_xyz_columns(file_path))
        rgb_is_255 = True

    # role -> column index. Keep first occurrence on duplicates so pandas
    # doesn't see a repeated column name.
    role_to_idx: Dict[str, int] = {}
    for idx, role in enumerate(columns):
        if role in _XYZ_DATA_ROLES and role not in role_to_idx:
            role_to_idx[role] = idx

    if not all(r in role_to_idx for r in ('x', 'y', 'z')):
        raise HTTPException(
            status_code=400,
            detail="ASCII format must include x, y, and z columns."
        )

    sorted_roles = sorted(role_to_idx.items(), key=lambda kv: kv[1])
    usecols = [idx for _, idx in sorted_roles]
    names = [role for role, _ in sorted_roles]

    skiprows = _ascii_skiprows(file_path)
    sep = _ascii_pandas_sep(file_path)

    try:
        df = pd.read_csv(
            file_path,
            sep=sep, header=None, comment='#',
            usecols=usecols, names=names,
            dtype=np.float32, engine='c', skip_blank_lines=True,
            skiprows=skiprows, on_bad_lines='skip',
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to parse {file_path}: {e}")

    if df.empty:
        raise HTTPException(status_code=400, detail=f"No data rows found in {file_path}")

    df = df.dropna(subset=['x', 'y', 'z'])
    if df.empty:
        raise HTTPException(status_code=400, detail=f"No valid xyz rows in {file_path}")

    positions = np.column_stack([
        df['x'].to_numpy(dtype=np.float32, copy=False),
        df['y'].to_numpy(dtype=np.float32, copy=False),
        df['z'].to_numpy(dtype=np.float32, copy=False),
    ]).astype(np.float32, copy=False)

    colors = None
    if all(c in df.columns for c in ('r255', 'g255', 'b255')):
        colors = np.column_stack([
            df['r255'].to_numpy(dtype=np.float32, copy=False) / 255.0,
            df['g255'].to_numpy(dtype=np.float32, copy=False) / 255.0,
            df['b255'].to_numpy(dtype=np.float32, copy=False) / 255.0,
        ]).astype(np.float32, copy=False)
    elif all(c in df.columns for c in ('r', 'g', 'b')):
        colors = np.column_stack([
            df['r'].to_numpy(dtype=np.float32, copy=False),
            df['g'].to_numpy(dtype=np.float32, copy=False),
            df['b'].to_numpy(dtype=np.float32, copy=False),
        ]).astype(np.float32, copy=False)

    intensity = None
    for role in ('intensity', 'reflectance'):
        if role in df.columns:
            intensity = df[role].to_numpy(dtype=np.float32, copy=False)
            break

    return positions, colors, intensity


def _load_ply_pcd_arrays(
    file_path: str,
) -> tuple[np.ndarray, Optional[np.ndarray], Optional[np.ndarray]]:
    """Parse PLY or PCD via open3d. Handles both ASCII and binary variants
    (open3d picks based on the file header), so the renderer is freed from
    streaming binary PLY/PCD too — same code path either way.

    Open3D loses scalar fields (intensity, reflectance, …) on read for PLY
    and PCD. If a downstream consumer needs them we'd need to parse the
    formats directly; none for now."""
    try:
        import open3d as o3d
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="open3d is not available; PLY/PCD by-path import is disabled.",
        )

    try:
        cloud = o3d.io.read_point_cloud(file_path)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to parse {file_path}: {e}")

    points = np.asarray(cloud.points)
    if points.size == 0:
        raise HTTPException(status_code=400, detail=f"No points found in {file_path}")

    positions = points.astype(np.float32, copy=False)
    colors = np.asarray(cloud.colors).astype(np.float32, copy=False) if cloud.has_colors() else None
    intensity = None
    return positions, colors, intensity


def _load_pointcloud_arrays(
    file_path: str, ascii_format: Optional[str],
    column_plan: "Optional[ColumnPlan]" = None,
) -> tuple[np.ndarray, Optional[np.ndarray], Optional[np.ndarray]]:
    """Dispatch a path-based point-cloud load to the right backend by
    extension and return the raw numpy arrays. Shared entry point for the
    import and crop endpoints — keeps file-IO, format detection, and the
    PLY/PCD vs ASCII dispatch in one place.

    `column_plan` (import wizard) applies only to the XYZ-family branch.

    Raises HTTPException on missing file or unsupported extension.
    """
    import os

    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail=f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower().lstrip('.')
    if ext in _PANDAS_EXTENSIONS:
        return _load_xyz_arrays(file_path, ascii_format, column_plan)
    if ext in _OPEN3D_EXTENSIONS:
        return _load_ply_pcd_arrays(file_path)
    if ext in _LAS_EXTENSIONS:
        return _load_las_arrays(file_path)

    raise HTTPException(
        status_code=400,
        detail=(
            f"Unsupported extension for path-based import: .{ext}. "
            f"Supported: {sorted(_PANDAS_EXTENSIONS | _OPEN3D_EXTENSIONS | _LAS_EXTENSIONS)}"
        ),
    )


def _load_las_arrays(
    file_path: str,
) -> tuple[np.ndarray, Optional[np.ndarray], Optional[np.ndarray]]:
    """Read a LAS/LAZ file into (positions[N,3] float64, colors[N,3] float32 in
    0-1 | None, intensity[N] float32 | None). Same laspy read as the legacy
    multipart `/api/pointcloud/import` endpoint, but returns the arrays for the
    binary `import_by_path` stream instead of JSON-serialising them — so a large
    LAZ no longer round-trips through `points.tolist()` + a 100s-of-MB JSON body.
    """
    try:
        import laspy
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="laspy library not installed. Run: pip install laspy[lazrs]",
        )

    las = laspy.read(file_path)
    positions = np.column_stack([las.x, las.y, las.z]).astype(np.float64, copy=False)

    colors = None
    if hasattr(las, 'red') and hasattr(las, 'green') and hasattr(las, 'blue'):
        try:
            red = np.asarray(las.red)
            green = np.asarray(las.green)
            blue = np.asarray(las.blue)
            # Only treat RGB as present if actually set (not an all-zero block).
            if red.max() > 0 or green.max() > 0 or blue.max() > 0:
                colors = (np.column_stack([red, green, blue]) / 65535.0).astype(
                    np.float32, copy=False
                )
        except Exception:
            colors = None

    intensity = None
    if hasattr(las, 'intensity'):
        try:
            inten = np.asarray(las.intensity, dtype=np.float32)
            if inten.size == len(positions) and inten.max() > 0:
                intensity = inten
        except Exception:
            intensity = None

    return positions, colors, intensity


def _read_points_from_source(
    src: PointSource,
) -> tuple[np.ndarray, Optional[np.ndarray], Optional[np.ndarray]]:
    """Resolve a PointSource to (positions[N,3] float64, colors[N,3] float32 in
    0-1 | None, intensity[N] float32 | None).

    Reuses `_load_pointcloud_arrays` (which dispatches XYZ via pandas and
    PLY/PCD via open3d, and 404s on a missing file), then applies the optional
    stride-downsample and translation. Positions come back as float64 because
    every consumer (open3d KD-trees, skeleton BFS, raycasting) wants double
    precision; colors/intensity keep their loader dtype.

    When `src.session_id` is set, points come from the live cloud session's
    in-RAM array with its per-point deletions already applied (the Family-1
    source of truth), so downstream ops honor unbaked deletions without a
    rebuild. The compute consumers of this path want positions only, so colours
    and intensity resolve to None here (the session DOES hold them — see
    `_read_las_into_arrays` — they're simply not surfaced for these ops).

    Sky/miss points (`is_miss != 0`) are also dropped here, exactly as the
    hits-only octree does (see `_session_to_las(exclude_misses=True)`). A miss is
    a ray that hit nothing, projected ~1 km out — it is not a surface point, so
    every compute consumer of this chokepoint (triangulate, skeleton, hits-only
    LAD, export) must skip it. Leaving them in not only meshes a phantom shell a
    kilometre away but, for ball-pivoting, makes BPA explode combinatorially and
    hang. The misses-overlay endpoint reads `sess.extras` directly, not through
    this function, so it still sees them.
    """
    if src.session_id is not None:
        sess = _get_cloud_session(src.session_id)
        with _cloud_session_lock:
            keep = ~sess.deleted
            if not src.include_misses and _MISS_SLUG in sess.extras:
                keep = keep & (sess.extras[_MISS_SLUG] == 0)
            positions = sess.positions[keep].copy()
            session_world_shift = sess.world_shift
        # Restore true world coordinates: the session stored points with the
        # import-time global shift SUBTRACTED, so add it back here — the single
        # chokepoint every downstream op (triangulate/skeleton/LAD/export) reads
        # through. Done before the explicit `src.translation` so a per-op
        # translation still composes on top of world coords.
        if session_world_shift is not None:
            positions = positions.astype(np.float64, copy=False) + session_world_shift
        colors = None
        intensity = None
    else:
        if not src.source_path:
            raise HTTPException(
                status_code=400,
                detail=("Point source has neither a session_id nor a source_path. "
                        "A session-backed cloud must send its session_id; a "
                        "file-backed cloud must send source_path."),
            )
        positions, colors, intensity = _load_pointcloud_arrays(
            src.source_path, src.ascii_format
        )

    if src.max_points is not None and src.max_points > 0 and len(positions) > src.max_points:
        stride = int(math.ceil(len(positions) / src.max_points))
        positions = positions[::stride]
        colors = colors[::stride] if colors is not None else None
        intensity = intensity[::stride] if intensity is not None else None

    positions = positions.astype(np.float64, copy=False)
    if src.translation is not None:
        t = np.asarray(src.translation, dtype=np.float64)
        if t.shape != (3,):
            raise HTTPException(
                status_code=400,
                detail=f"translation must be [tx, ty, tz]; got {src.translation!r}",
            )
        positions = positions + t

    if not src.want_colors:
        colors = None

    return positions, colors, intensity


@app.post("/api/pointcloud/import_by_path")
async def import_pointcloud_by_path(request: ImportPointCloudByPathRequest):
    """Parse a point-cloud file from disk and stream back a packed binary
    representation. Dispatches by extension:

    * `.xyz` / `.txt` / `.csv` / `.pts` / `.asc` → pandas + optional
      Helios `ascii_format` hint.
    * `.ply` / `.pcd` → open3d (handles ASCII and binary).
    """
    positions, colors, intensity = _load_pointcloud_arrays(
        request.file_path, request.ascii_format, request.column_plan
    )
    # CloudCompare-style global shift: subtract the requested offset so the small
    # cloud's in-RAM array (kept as-is by the flat renderer path) holds small,
    # precision-friendly coordinates. The renderer persists the shift it sent on
    # the cloud for world-coord readouts/exports. None/zero = keep coordinates.
    if request.world_shift is not None:
        ws = np.asarray(request.world_shift, dtype=np.float64)
        if ws.shape != (3,):
            raise HTTPException(
                status_code=400,
                detail=f"world_shift must be [x, y, z]; got {request.world_shift!r}",
            )
        if np.any(ws != 0.0):
            positions = positions.astype(np.float64, copy=False) - ws
    return _pack_pointcloud_response(positions, colors, intensity)


# ============================================================================
# Point cloud → Potree 2.0 octree pipeline (added in 0.3.0)
# ============================================================================
# Flat-array import paths above OOM the renderer on multi-GB clouds. This
# section converts source files into Potree 2.0 octrees (metadata.json +
# hierarchy.bin + octree.bin) which the renderer streams via potree-core.
#
# PotreeConverter 2.x accepts LAS/LAZ only, so XYZ-family inputs are
# pre-converted via laspy in chunks. Both phases combined run at ~3 M pts/sec
# on M-series Macs with NVMe.

import hashlib as _hashlib
import os as _os
import shutil as _shutil
import subprocess as _subprocess
import sys as _sys
from pathlib import Path as _Path


_OCTREE_CACHE_VERSION = 1  # bump if cache layout changes (forces re-conversion)

# Per-cache-key build locks. Octrees are keyed by the LAS bytes' hash, so a
# batch import of N scans whose hits-LAS bytes collide (identical/empty/synthetic
# scans) would otherwise run N concurrent builds against the SAME staging dir —
# and one request's rmtree (initial-clean or except-cleanup) deletes another's
# in-flight staging dir, crashing the second with FileNotFoundError while it
# writes attribute_labels.json. Serializing per key makes the first build win and
# every later one a cache hit; distinct keys still build in parallel. The registry
# lock guards only the dict lookup, never the (slow) build itself.
_octree_build_locks: dict[str, threading.Lock] = {}
_octree_build_locks_guard = threading.Lock()


def _octree_build_lock(cache_key: str) -> threading.Lock:
    """Return the process-wide build lock for a given octree cache key."""
    with _octree_build_locks_guard:
        lock = _octree_build_locks.get(cache_key)
        if lock is None:
            lock = threading.Lock()
            _octree_build_locks[cache_key] = lock
        return lock

# Default cache cap. Overridable via PHYTOGRAPH_OCTREE_CACHE_MAX_BYTES.
# A typical 28 M-point cloud is ~340 MB on disk; 20 GB holds ~60 such clouds.
_DEFAULT_OCTREE_CACHE_MAX_BYTES = 20 * 1024 * 1024 * 1024


def _octree_cache_root() -> _Path:
    """OS-appropriate user-data directory for cached octrees.

    macOS: ~/Library/Application Support/Phytograph/cache/octrees
    Linux: $XDG_CACHE_HOME/Phytograph/octrees (or ~/.cache/Phytograph/octrees)
    Windows: %LOCALAPPDATA%/Phytograph/cache/octrees
    Overridable via PHYTOGRAPH_OCTREE_CACHE_ROOT for tests."""
    override = _os.environ.get("PHYTOGRAPH_OCTREE_CACHE_ROOT")
    if override:
        return _Path(override)
    if _sys.platform == "darwin":
        base = _Path.home() / "Library" / "Application Support" / "Phytograph" / "cache"
    elif _sys.platform.startswith("win"):
        base = _Path(_os.environ.get("LOCALAPPDATA", _Path.home() / "AppData" / "Local")) / "Phytograph" / "cache"
    else:
        base = _Path(_os.environ.get("XDG_CACHE_HOME", _Path.home() / ".cache")) / "Phytograph"
    return base / "octrees"


def _canonical_ascii_format(ascii_format: Optional[str]) -> str:
    """Normalise whitespace and case so equivalent format strings hash the same."""
    if not ascii_format:
        return ""
    return " ".join(ascii_format.split()).lower()


def _canonical_column_plan(column_plan: "Optional[ColumnPlan]") -> str:
    """Stable JSON form of a ColumnPlan for hashing. Empty string when None so a
    plan-less import keeps the same cache identity it had before this field
    existed (no cache churn for existing users)."""
    if column_plan is None:
        return ""
    payload = {
        "rgb_is_255": column_plan.rgb_is_255,
        "columns": [
            {"index": e.index, "role": (e.role or "").lower(),
             "slug": e.slug or "", "label": e.label or "",
             "categorical": bool(e.categorical)}
            for e in sorted(column_plan.columns, key=lambda e: e.index)
        ],
    }
    return json.dumps(payload, sort_keys=True, separators=(",", ":"))


def _octree_cache_key(source_path: str, ascii_format: Optional[str],
                      column_plan: "Optional[ColumnPlan]" = None) -> str:
    """Stable cache key for (source file, format, column plan). Includes mtime
    so edits to the source XYZ invalidate the cached octree. Two different
    wizard column plans on the same file get distinct cache entries."""
    p = _Path(source_path).resolve()
    try:
        mtime_ns = p.stat().st_mtime_ns
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Source file not found: {source_path}")
    h = _hashlib.sha1()
    h.update(str(_OCTREE_CACHE_VERSION).encode())
    h.update(b"\x00")
    h.update(str(p).encode())
    h.update(b"\x00")
    h.update(str(mtime_ns).encode())
    h.update(b"\x00")
    h.update(_canonical_ascii_format(ascii_format).encode())
    h.update(b"\x00")
    h.update(_canonical_column_plan(column_plan).encode())
    return h.hexdigest()


def _octree_cache_dir(source_path: str, ascii_format: Optional[str],
                      column_plan: "Optional[ColumnPlan]" = None) -> _Path:
    """Path where this (source, format, plan) tuple's octree lives. May not exist yet."""
    return _octree_cache_root() / _octree_cache_key(source_path, ascii_format, column_plan)


def _resolve_potree_converter_path() -> _Path:
    """Locate the PotreeConverter binary.

    Resolution order:
      1. $PHYTOGRAPH_POTREECONVERTER (explicit override, used by tests / CI).
      2. $PHYTOGRAPH_RESOURCES/potree_converter/<platform>/PotreeConverter
         (set by Electron main process for packaged builds).
      3. <repo root>/resources/potree_converter/<platform>/PotreeConverter
         (CI-built binary, dev fallback).
      4. <repo root>/tmp/potree-converter-src/build/PotreeConverter
         (locally built spike binary; dev convenience only).
    """
    override = _os.environ.get("PHYTOGRAPH_POTREECONVERTER")
    if override:
        p = _Path(override)
        if p.is_file():
            return p
        raise HTTPException(
            status_code=503,
            detail=f"PHYTOGRAPH_POTREECONVERTER points to a missing file: {override}",
        )

    if _sys.platform == "darwin":
        platform_dir = "darwin-arm64" if _os.uname().machine == "arm64" else "darwin-x64"
        bin_name = "PotreeConverter"
    elif _sys.platform.startswith("win"):
        platform_dir = "win-x64"
        bin_name = "PotreeConverter.exe"
    else:
        platform_dir = "linux-x64"
        bin_name = "PotreeConverter"

    candidates = []
    resources_env = _os.environ.get("PHYTOGRAPH_RESOURCES")
    if resources_env:
        candidates.append(_Path(resources_env) / "potree_converter" / platform_dir / bin_name)

    repo_root = _Path(__file__).resolve().parent.parent
    candidates.append(repo_root / "resources" / "potree_converter" / platform_dir / bin_name)
    candidates.append(repo_root / "tmp" / "potree-converter-src" / "build" / bin_name)

    for c in candidates:
        if c.is_file():
            return c

    raise HTTPException(
        status_code=503,
        detail=(
            f"PotreeConverter binary not found. Looked in: "
            f"{[str(c) for c in candidates]}. "
            f"Run `npm run build:potree-converter` or set PHYTOGRAPH_POTREECONVERTER."
        ),
    )


def _intensity_to_las_uint16(values: "np.ndarray",
                             lo: "Optional[float]" = None,
                             hi: "Optional[float]" = None) -> "np.ndarray":
    """Map an intensity/reflectance column to the LAS uint16 field by normalising
    a finite range to 0..65535.

    The LAS intensity field is uint16, so the source values must land in
    [0, 65535] as non-negative integers. We can't assume a fixed source scale:
    Helios reflectance is reported in dB (negative, e.g. -14..0), other exports
    use [0, 1] floats, the legacy convention was [0, 255], and Helios's raw
    "intensity" is a signed beam·normal dot product. A fixed `* 256` + clip(0, …)
    crushes every all-negative column (dB, signed dot-products) to a uniform 0 —
    losing the field entirely. Normalising from the column's range instead
    preserves the gradient for any scale; the renderer colours by the resulting
    range either way (see OctreePointCloud `attributeRanges.intensity`). Mirrors
    the E57 path's per-scan normalisation.

    Pass `lo`/`hi` to normalise against a precomputed GLOBAL range (the chunked
    ASCII path does this so the gradient is consistent across chunk seams); omit
    them to use this array's own finite min/max. Non-finite values map to 0. A
    zero-width range maps to all-0 (no gradient, and avoids a divide-by-zero).
    """
    v = np.asarray(values, dtype=np.float64)
    finite = np.isfinite(v)
    if lo is None or hi is None:
        if not finite.any():
            return np.zeros(len(v), dtype=np.uint16)
        lo = float(v[finite].min())
        hi = float(v[finite].max())
    span = hi - lo
    if span <= 1e-12:
        return np.zeros(len(v), dtype=np.uint16)
    scaled = np.where(finite, (v - lo) / span * 65535.0, 0.0)
    return np.clip(scaled, 0, 65535).astype(np.uint16)


def _xyz_to_las(source_path: _Path, ascii_format: Optional[str], out_las: _Path,
                column_plan: "Optional[ColumnPlan]" = None,
                capture_full_xyz: bool = False,
                capture_origins: bool = False,
                ) -> "tuple[int, List[dict], Optional[np.ndarray], Optional[np.ndarray]]":
    """Stream an XYZ-family ASCII file into a LAS file via laspy in chunks.

    PotreeConverter 2.x accepts only LAS/LAZ; XYZ goes through here first.
    Streaming keeps peak memory bounded by `chunk_rows` * (cols × 8B), not
    by total point count.

    Column layout uses the same `ascii_format` convention as
    `_load_xyz_arrays` — roles are tokenised, with x/y/z mandatory and
    r255/g255/b255/intensity (or reflectance, mapped to intensity) optional.
    Any remaining numeric columns are carried into the octree as LAS extra
    dimensions (float32) so the renderer can colour by them later.

    Returns (total_points, extra_dims, full_xyz, origins), where extra_dims is
    the [{slug, label}, ...] list of carried scalar attributes (for the cache's
    slug→label sidecar). `full_xyz` is the (N,3) float64 source-precision
    coordinate array when `capture_full_xyz` is set, else None — the LAS itself
    is 1 mm-quantized and must not be the session's source of truth, so the
    session reads positions from this instead (see _xyz_to_las_stream).

    `origins` is the (N,3) float64 per-pulse beam-origin array when
    `capture_origins` is set AND the column plan carried an ox/oy/oz triple (the
    `origin:*` tokens), else None. Like `full_xyz` it is captured directly from
    the source columns — origins are world/UTM coordinates and must keep full
    precision, so they bypass the 1 mm LAS / float32 extras entirely and feed the
    LAD path's ground-truth-origin shortcut (see CloudSession.beam_origins).
    """
    import laspy  # local: only when this code path runs

    names, extra_dims = _xyz_column_plan(source_path, ascii_format, column_plan)

    has_xyz = all(role in names for role in ("x", "y", "z"))
    if not has_xyz:
        raise HTTPException(
            status_code=400,
            detail=f"ASCII format must include x/y/z. Got columns: {names}",
        )

    # RGB may arrive as 0-255 ints (r255/g255/b255) or 0-1 floats (r/g/b). The
    # wizard's rgb_is_255 flag, when a column_plan is present, is authoritative;
    # otherwise infer from which role tokens the plan produced.
    rgb255_cols = ("r255", "g255", "b255")
    rgb01_cols = ("r", "g", "b")
    has_rgb255 = all(role in names for role in rgb255_cols)
    has_rgb01 = all(role in names for role in rgb01_cols)
    if column_plan is not None and (has_rgb255 or has_rgb01):
        rgb_is_255 = column_plan.rgb_is_255
    else:
        rgb_is_255 = has_rgb255
    rgb_cols = rgb255_cols if has_rgb255 else (rgb01_cols if has_rgb01 else None)
    intensity_role = next((r for r in ("intensity", "reflectance") if r in names), None)

    # A scan with BOTH intensity and reflectance columns can only put one in the
    # LAS intensity field. The other is a reserved role (so `_plan_columns`
    # didn't carry it) yet pandas still reads it — without rescue it's silently
    # dropped. Carry it as an extra dim under its own slug so it's colourable /
    # filterable like any scalar, instead of vanishing.
    secondary_intensity = None
    if intensity_role is not None:
        other = "reflectance" if intensity_role == "intensity" else "intensity"
        if other in names:
            secondary_intensity = {
                "col": other,
                "slug": _dedupe_slug(other, {e["slug"] for e in extra_dims}),
                "label": other.capitalize(),
            }
            extra_dims = extra_dims + [secondary_intensity]

    # LAS point format 3 carries XYZ + intensity + RGB. Extra numeric columns
    # are added as float32 extra dimensions; PotreeConverter writes the full
    # LAS schema (no --attributes filter), so they survive into the octree's
    # metadata.json attributes list and decode into named buffers in the
    # potree-core 2.0 loader.
    header = laspy.LasHeader(point_format=3, version="1.4")
    header.scales = np.array([0.001, 0.001, 0.001], dtype=np.float64)
    for ed in extra_dims:
        header.add_extra_dim(laspy.ExtraBytesParams(name=ed["slug"], type=np.float32))

    # Per-pulse beam origins ride in `names` as `origin:origin_x/_y/_z` tokens
    # (see `_plan_columns*`). They are deliberately NOT extra dims — origins are
    # world/UTM coordinates and a float32 LAS extra would shatter their precision —
    # so we capture them straight from the source into a float64 side-channel,
    # exactly like `capture_full_xyz` does for positions. Map each canonical slug
    # to its `names` token so the streaming loop can stack the three columns in
    # x,y,z order regardless of source column order; capture only when all three
    # are present (a partial triple is not a usable origin).
    origin_cols: Optional[Dict[str, str]] = None
    if capture_origins:
        present = {_origin_slug_of(c): c for c in names if _is_origin_name(c)}
        if all(s in present for s in _ORIGIN_SLUGS):
            origin_cols = present

    skiprows = _ascii_skiprows(str(source_path))
    sep = _ascii_pandas_sep(str(source_path))

    chunk_rows = 2_000_000

    # Normalise the intensity/reflectance column to the LAS uint16 field by its
    # GLOBAL finite range so the gradient is consistent across chunk boundaries.
    # A per-chunk min/max would rescale each 2M-row chunk independently, banding
    # the cloud at chunk seams. Scan just that one column first (cheap); skip the
    # pass when there's no intensity column to map.
    intensity_lo = intensity_hi = None
    if intensity_role is not None:
        # Read just that column by POSITION (not via names=, which may be
        # narrower than the file's field count for a partial column plan).
        intensity_pos = names.index(intensity_role)
        gmin, gmax = np.inf, -np.inf
        for col_chunk in pd.read_csv(
            source_path, sep=sep, header=None,
            usecols=[intensity_pos], comment="#",
            skiprows=skiprows, chunksize=chunk_rows, engine="c",
        ):
            vals = col_chunk.iloc[:, 0].to_numpy(dtype=np.float64)
            vals = vals[np.isfinite(vals)]
            if vals.size:
                gmin = min(gmin, float(vals.min()))
                gmax = max(gmax, float(vals.max()))
        if np.isfinite(gmin) and np.isfinite(gmax):
            intensity_lo, intensity_hi = gmin, gmax
    try:
        # Offset the LAS to the data's MIN coordinate so projected clouds (UTM
        # northings ~5.4e6 m) fit the 32-bit int range — with offset 0 and 1 mm
        # scale only ±2.1 km is representable. Cheap pre-pass over just x/y/z
        # (laspy re-applies the offset on read, so coordinates are unchanged
        # downstream). Inside the try so a column-format mismatch here raises the
        # same actionable 400 as the main stream below, not a raw ValueError.
        xyz_pos = [names.index(r) for r in ("x", "y", "z")]
        order = np.argsort(np.argsort(xyz_pos))  # usecols sorts by position; map back to x,y,z
        xyz_min = np.array([np.inf, np.inf, np.inf])
        for col_chunk in pd.read_csv(
            source_path, sep=sep, header=None,
            usecols=xyz_pos, comment="#",
            skiprows=skiprows, chunksize=chunk_rows, engine="c",
        ):
            arr = col_chunk.to_numpy(dtype=np.float64)[:, order]
            arr = arr[np.isfinite(arr).all(axis=1)]
            if arr.size:
                xyz_min = np.minimum(xyz_min, arr.min(axis=0))
        header.offsets = np.floor(np.where(np.isfinite(xyz_min), xyz_min, 0.0))

        full_xyz_chunks: "Optional[list]" = [] if capture_full_xyz else None
        origin_chunks: "Optional[list]" = [] if origin_cols is not None else None
        total_points = _xyz_to_las_stream(
            source_path, out_las, header, names, skiprows, sep, chunk_rows,
            rgb_cols, rgb_is_255, intensity_role, intensity_lo, intensity_hi,
            extra_dims, full_xyz_out=full_xyz_chunks,
            origin_cols=origin_cols, origins_out=origin_chunks,
        )
    except (ValueError, KeyError) as e:
        # A non-numeric value reaching the x/y/z/RGB float cast almost always
        # means the chosen column format doesn't match THIS file — e.g. a bulk
        # import applied one file's layout to another with different columns or
        # a stray header row. Turn the cryptic pandas/numpy error into an
        # actionable 400 naming the file and the chosen columns, instead of an
        # uncaught 500 the UI can only report as "Internal Server Error".
        raise HTTPException(
            status_code=400,
            detail=(
                f"Could not read {source_path.name} with the selected column "
                f"format ({', '.join(names)}): {e}. The file's actual columns "
                f"likely differ — re-import it on its own and pick a matching "
                f"format."
            ),
        ) from e
    full_xyz = None
    if full_xyz_chunks is not None:
        full_xyz = (np.concatenate(full_xyz_chunks, axis=0)
                    if full_xyz_chunks else np.empty((0, 3), dtype=np.float64))
    origins = None
    if origin_chunks is not None:
        origins = (np.concatenate(origin_chunks, axis=0)
                   if origin_chunks else np.empty((0, 3), dtype=np.float64))
    return (total_points,
            [{"slug": ed["slug"], "label": ed["label"]} for ed in extra_dims],
            full_xyz, origins)


def _xyz_to_las_stream(source_path, out_las, header, names, skiprows, sep,
                       chunk_rows, rgb_cols, rgb_is_255, intensity_role,
                       intensity_lo, intensity_hi, extra_dims,
                       full_xyz_out: "Optional[list]" = None,
                       origin_cols: "Optional[Dict[str, str]]" = None,
                       origins_out: "Optional[list]" = None) -> int:
    """Inner streaming loop for `_xyz_to_las`, split out so the caller can wrap
    column-mismatch errors (raised here from the float casts) into a clean 400.
    Returns the total points written.

    The LAS this writes is quantized to the header's 1 mm scale — fine as the
    octree's input (the octree is a display cache) but NOT precise enough to be
    the session's source-of-truth array (1 mm shatters precision-sensitive ops
    like triangulation; see CloudSession.positions). When `full_xyz_out` is a
    list, the FULL-PRECISION float64 xyz of each written chunk (already
    NaN-filtered, in LAS point order) is appended to it, so the caller can
    populate the session positions directly from the source instead of reading
    them back from the quantized LAS. The chunks are filtered identically to the
    LAS write here, so the concatenation aligns point-for-point with the LAS-
    derived colors/intensity/extras.

    When `origins_out` is a list, the per-pulse beam origins are captured the same
    way: `origin_cols` maps each canonical slug ('origin_x'/'_y'/'_z') to its
    column name in `names`, and the three columns are stacked in x,y,z order as
    float64 into `origins_out`, aligned point-for-point with the same NaN-filtered
    rows. Origins are world/UTM coordinates kept OUT of the quantized LAS and the
    float32 extras so LAD's ground-truth-origin shortcut sees full precision."""
    import laspy
    total_points = 0
    with laspy.open(str(out_las), mode="w", header=header) as writer:
        reader = pd.read_csv(
            source_path,
            sep=sep,
            header=None,
            names=names,
            usecols=[i for i, c in enumerate(names) if not _is_skip_name(c)],
            comment="#",
            skiprows=skiprows,
            chunksize=chunk_rows,
            engine="c",
        )
        for chunk in reader:
            # Drop rows with non-numeric / missing x/y/z. A .pts file leads
            # with a bare point-count line ('12345') that pandas pads to a
            # 1-field row (xyz NaN); without this it would otherwise be cast
            # to a garbage (0,0,0)-ish point in the octree. Mirrors the flat
            # loader's `df.dropna(subset=['x','y','z'])`.
            chunk = chunk.dropna(subset=["x", "y", "z"])
            n = len(chunk)
            if n == 0:
                continue
            record = laspy.ScaleAwarePointRecord.zeros(n, header=header)
            cx = chunk["x"].to_numpy(dtype=np.float64)
            cy = chunk["y"].to_numpy(dtype=np.float64)
            cz = chunk["z"].to_numpy(dtype=np.float64)
            record.x = cx
            record.y = cy
            record.z = cz
            if full_xyz_out is not None:
                # Stash the unquantized xyz of this chunk for the session array.
                full_xyz_out.append(np.column_stack([cx, cy, cz]))
            if origins_out is not None and origin_cols is not None:
                # Stash the full-precision beam origins of this chunk, stacked in
                # x,y,z order (regardless of source column order) and filtered to
                # the exact same rows as xyz above, so the concatenation aligns
                # point-for-point with positions/colors/extras. Origins stay
                # float64 — never routed through the 1 mm LAS or float32 extras.
                origins_out.append(np.column_stack([
                    chunk[origin_cols['origin_x']].to_numpy(dtype=np.float64),
                    chunk[origin_cols['origin_y']].to_numpy(dtype=np.float64),
                    chunk[origin_cols['origin_z']].to_numpy(dtype=np.float64),
                ]))
            if rgb_cols is not None:
                # LAS RGB is uint16 (16-bit per channel). 0-255 source scales by
                # 256 (preserves perceptual brightness; renderer right-shifts to
                # recover 8-bit). 0-1 float source scales straight to the full
                # 16-bit range (×65535).
                rc, gc, bc = rgb_cols
                if rgb_is_255:
                    record.red = chunk[rc].to_numpy(dtype=np.uint16) * 256
                    record.green = chunk[gc].to_numpy(dtype=np.uint16) * 256
                    record.blue = chunk[bc].to_numpy(dtype=np.uint16) * 256
                else:
                    record.red = np.clip(chunk[rc].to_numpy(dtype=np.float32) * 65535.0, 0, 65535).astype(np.uint16)
                    record.green = np.clip(chunk[gc].to_numpy(dtype=np.float32) * 65535.0, 0, 65535).astype(np.uint16)
                    record.blue = np.clip(chunk[bc].to_numpy(dtype=np.float32) * 65535.0, 0, 65535).astype(np.uint16)
            if intensity_role is not None:
                # Map intensity/reflectance to the LAS uint16 field, normalising
                # the column's GLOBAL finite range (computed above) — handles dB,
                # [0,1], [0,255], and signed dot-product scales uniformly.
                record.intensity = _intensity_to_las_uint16(
                    chunk[intensity_role].to_numpy(dtype=np.float64),
                    intensity_lo, intensity_hi,
                )
            for ed in extra_dims:
                # Extra dims carry RAW values (the renderer normalises by the
                # attribute's own range) — including a rescued secondary
                # intensity/reflectance column, so its true dB/float values stay
                # available for colour-by and filtering.
                record[ed["slug"]] = chunk[ed["col"]].to_numpy(dtype=np.float32)
            writer.write_points(record)
            total_points += n
    return total_points


def _ply_to_las(source_path: _Path, out_las: _Path) -> tuple[int, List[dict]]:
    """Convert a PLY to LAS via plyfile so it can feed the PotreeConverter
    octree pipeline, preserving scalar fields as LAS extra dimensions.

    open3d (used on the flat path) drops every vertex property except position
    and RGB, so we parse the PLY directly. Recognised roles: x/y/z (required),
    red/green/blue or r/g/b (RGB), and the first of intensity/scalar_intensity/
    reflectance (mapped to LAS intensity). Every remaining numeric vertex
    property is carried as a float32 extra dimension so the renderer can colour
    by it — the same mechanism `_xyz_to_las` uses for unmapped ASCII columns.

    Sky/miss handling (for LAD): a property aliased to `is_miss` (is_miss/miss/
    sky) is normalised to the canonical `is_miss` extra dim. Rows with non-finite
    (NaN/Inf) coordinates are also treated as misses — a structured/organized PLY
    marks empty grid cells that way. A generic PLY carries no scanner origin, so
    a NaN-coord miss has no recoverable beam direction; those rows are dropped
    (they can't be placed as a far-field point) while any miss that DOES have
    finite far-field coords (e.g. exported from Helios) is kept and tagged.

    Returns (total_points, extra_dims), where extra_dims is the
    [{slug, label}, ...] list for the slug→label sidecar; `is_miss` is included
    whenever the PLY carried miss information.
    """
    import laspy  # local: only when this code path runs
    from plyfile import PlyData

    ply = PlyData.read(str(source_path))
    try:
        vertex = ply["vertex"].data
    except KeyError:
        raise HTTPException(
            status_code=400,
            detail="PLY has no 'vertex' element; cannot import as a point cloud.",
        )
    names = list(vertex.dtype.names or ())

    if not all(role in names for role in ("x", "y", "z")):
        raise HTTPException(
            status_code=400,
            detail=f"PLY missing x/y/z vertex properties. Got: {names}",
        )

    # RGB: PLY conventionally uses red/green/blue; accept short r/g/b too.
    if all(c in names for c in ("red", "green", "blue")):
        rgb_cols = ("red", "green", "blue")
    elif all(c in names for c in ("r", "g", "b")):
        rgb_cols = ("r", "g", "b")
    else:
        rgb_cols = None

    intensity_col = next(
        (c for c in ("intensity", "scalar_intensity", "reflectance") if c in names),
        None,
    )

    # An explicit miss-flag property (is_miss/miss/sky) maps to the canonical
    # `is_miss` slug, not a generic positional extra dim.
    miss_col = next((c for c in names if c.lower() in _MISS_ALIASES), None)

    reserved = {"x", "y", "z"}
    if rgb_cols:
        reserved.update(rgb_cols)
    if intensity_col:
        reserved.add(intensity_col)
    if miss_col:
        reserved.add(miss_col)

    # Carry every other numeric property as a float32 extra dim. Dedupe slugs
    # the same way the ASCII path does, since two headers can sanitise alike.
    extra_dims: List[dict] = []
    used_slugs: set[str] = set()
    for col in names:
        if col in reserved or not np.issubdtype(vertex[col].dtype, np.number):
            continue
        slug = _sanitize_extra_dim_name(col)
        base, i = slug, 1
        while slug in used_slugs:
            slug = f"{base[:29]}_{i}"
            i += 1
        used_slugs.add(slug)
        extra_dims.append({"col": col, "slug": slug, "label": _humanize_extra_dim_label(col)})

    x = vertex["x"].astype(np.float64)
    y = vertex["y"].astype(np.float64)
    z = vertex["z"].astype(np.float64)

    # Build the per-point miss flag from the explicit column and/or non-finite
    # coordinates (an organized PLY's empty grid cells).
    is_miss = np.zeros(len(vertex), dtype=np.float32)
    has_miss_info = False
    if miss_col is not None:
        is_miss = (np.asarray(vertex[miss_col]).astype(np.float64) != 0).astype(np.float32)
        has_miss_info = True
    nonfinite = ~(np.isfinite(x) & np.isfinite(y) & np.isfinite(z))
    if nonfinite.any():
        is_miss[nonfinite] = 1.0
        has_miss_info = True

    # Drop misses we can't place: a generic PLY has no scanner origin, so a
    # NaN-coord miss has no direction to project onto. A miss that still carries
    # finite (far-field) coords — e.g. a Helios export — is kept and tagged.
    keep = np.ones(len(vertex), dtype=bool)
    if nonfinite.any():
        keep = ~nonfinite

    if has_miss_info:
        # `is_miss` rides along as a normal extra dim through octree/session/LAD.
        extra_dims.append({"col": None, "slug": _MISS_SLUG, "label": _MISS_LABEL})

    n = int(keep.sum())
    header = laspy.LasHeader(point_format=3, version="1.4")
    header.scales = np.array([0.001, 0.001, 0.001], dtype=np.float64)
    # Offset to the data min so projected (e.g. UTM) coordinates fit the LAS
    # 32-bit int range — see _session_to_las for the full rationale.
    header.offsets = (np.floor([x[keep].min(), y[keep].min(), z[keep].min()]) if n else np.zeros(3))
    for ed in extra_dims:
        header.add_extra_dim(laspy.ExtraBytesParams(name=ed["slug"], type=np.float32))

    record = laspy.ScaleAwarePointRecord.zeros(n, header=header)
    record.x = x[keep]
    record.y = y[keep]
    record.z = z[keep]
    if rgb_cols:
        # PLY RGB is 0-255 (uint8); LAS RGB is uint16. *256 keeps perceptual
        # brightness and lets the renderer right-shift to recover 8-bit, the
        # same convention as `_xyz_to_las`.
        record.red = vertex[rgb_cols[0]][keep].astype(np.uint16) * 256
        record.green = vertex[rgb_cols[1]][keep].astype(np.uint16) * 256
        record.blue = vertex[rgb_cols[2]][keep].astype(np.uint16) * 256
    if intensity_col is not None:
        # Normalise the intensity/reflectance property's range to the LAS uint16
        # field — handles dB / [0,1] / [0,255] / signed scales uniformly, instead
        # of a fixed `* 256` that crushes all-negative columns to 0.
        record.intensity = _intensity_to_las_uint16(vertex[intensity_col][keep])
    for ed in extra_dims:
        if ed["slug"] == _MISS_SLUG and ed.get("col") is None:
            record[_MISS_SLUG] = is_miss[keep]
        else:
            record[ed["slug"]] = vertex[ed["col"]][keep].astype(np.float32)

    with laspy.open(str(out_las), mode="w", header=header) as writer:
        writer.write_points(record)

    return n, [{"slug": ed["slug"], "label": ed["label"]} for ed in extra_dims]


def _pcd_viewpoint_origin(source_path: _Path) -> Optional[list]:
    """Read the PCD header `VIEWPOINT tx ty tz qw qx qy qz` field and return the
    scanner translation `[tx, ty, tz]` when it is meaningfully non-identity.

    PCD is the only other supported format (besides E57) that records a sensor
    pose, but it carries ONLY a translation + orientation quaternion — no angular
    sweep or grid — and the overwhelming majority of files leave it at the
    identity default (`0 0 0 1 0 0 0`). We surface just the origin, and only when
    it differs from the default, so we don't fabricate a (0,0,0) scan origin for
    every PCD. The header is plain ASCII even in binary PCDs, so a small text
    scan of the first lines suffices. Returns None when absent/identity/unreadable.
    """
    try:
        with open(source_path, "r", encoding="ascii", errors="ignore") as fh:
            for _ in range(64):  # VIEWPOINT lives in the small fixed header
                line = fh.readline()
                if not line:
                    break
                if line.startswith("VIEWPOINT"):
                    parts = line.split()[1:]
                    if len(parts) < 3:
                        return None
                    tx, ty, tz = (float(parts[0]), float(parts[1]), float(parts[2]))
                    if abs(tx) < 1e-9 and abs(ty) < 1e-9 and abs(tz) < 1e-9:
                        return None  # identity default — nothing to populate
                    return [tx, ty, tz]
                if line.startswith("DATA"):
                    break  # header ended without a VIEWPOINT
    except Exception:
        return None
    return None


def _pcd_to_las(source_path: _Path, out_las: _Path) -> tuple[int, List[dict]]:
    """Convert a PCD to LAS so it can feed the PotreeConverter octree pipeline.

    PCD goes through open3d (`_load_ply_pcd_arrays`), which carries position and
    RGB only — PCD's ascii/binary/binary_compressed variants make robust scalar
    parsing more than this is worth for now, so scalar fields are not preserved.
    A non-identity PCD `VIEWPOINT` translation is stashed (keyed by the output
    LAS path) so create_cloud_session can auto-populate the scan origin, the same
    channel E57 uses. Returns (total_points, []).
    """
    import laspy  # local: only when this code path runs

    positions, colors, _ = _load_ply_pcd_arrays(str(source_path))

    n = len(positions)
    header = laspy.LasHeader(point_format=3, version="1.4")
    header.scales = np.array([0.001, 0.001, 0.001], dtype=np.float64)
    # Offset to the data min so projected (e.g. UTM) coordinates fit the LAS
    # 32-bit int range — see _session_to_las for the full rationale.
    header.offsets = (np.floor(positions[:, :3].min(axis=0)) if n else np.zeros(3))

    record = laspy.ScaleAwarePointRecord.zeros(n, header=header)
    record.x = positions[:, 0].astype(np.float64)
    record.y = positions[:, 1].astype(np.float64)
    record.z = positions[:, 2].astype(np.float64)
    if colors is not None:
        # open3d colors are 0-1 floats; LAS RGB is uint16. Scale to 8-bit then
        # *256 to match the `_xyz_to_las` / `_ply_to_las` convention.
        rgb8 = np.clip(colors * 255.0, 0, 255).astype(np.uint16)
        record.red = rgb8[:, 0] * 256
        record.green = rgb8[:, 1] * 256
        record.blue = rgb8[:, 2] * 256

    with laspy.open(str(out_las), mode="w", header=header) as writer:
        writer.write_points(record)

    # Surface a non-identity sensor origin from the PCD VIEWPOINT, if any, via
    # the same per-output-LAS channel E57 uses. Only origin is recoverable from
    # PCD (no angular sweep / grid), so the rest of ScanParameters stays default.
    vp = _pcd_viewpoint_origin(source_path)
    if vp is not None:
        _e57_scan_meta[str(out_las.resolve())] = {
            "origin": vp,
            "scan_params": {"origin": vp},
            "has_misses": False,
            "miss_count": 0,
            "unplaceable_miss_count": 0,
        }

    return n, []


# Scanner origin recovered from the most recent E57 conversion, keyed by the
# absolute output LAS path. `_e57_to_las` writes it; `create_cloud_session` pops
# it right after conversion to surface the origin (and a hasMisses flag) to the
# renderer so it can place the scan's `ScanParameters.origin` and relocate miss
# points onto the bounding sphere for display. Threaded this way (rather than
# widening `_source_to_las`'s tuple) so the many other callers stay untouched.
_e57_scan_meta: Dict[str, dict] = {}


def _e57_scan_params(header, has_grid: bool) -> dict:
    """Recover scanner scan-pattern parameters from an E57 scan header so a
    non-XML E57 import can auto-populate a Scan's `ScanParameters` (origin +
    angular sweep + grid resolution), the same fields the Helios XML carries.

    E57 stores these in two optional header sub-structures:
      - `sphericalBounds`: azimuthStart/azimuthEnd and elevationMinimum/Maximum
        (radians). E57 elevation is measured from the XY plane (+up, range
        -pi/2..pi/2); Phytograph/Helios uses zenith (polar angle from +Z), so
        `zenith = 90 - elevation_deg`. Azimuth maps directly to phi (degrees).
      - `indexBounds`: row/column min/max — the structured-grid dimensions,
        which give the sample counts (n_theta = rows, n_phi = columns).

    `has_grid` says whether the SCAN DATA actually carried row/column indices
    (a true raster). indexBounds is only trusted for sample counts when it does:
    pye57 writes a degenerate indexBounds even for a flat, unstructured cloud
    (rowMaximum = point_count - 1), which would otherwise masquerade as a zenith
    resolution. Angular bounds (sphericalBounds) are independent of the grid and
    are read whenever present.

    Both sub-structures are optional in the spec; many files omit one or both.
    Every field is guarded independently — whatever isn't present is simply left
    out so the renderer falls back to its default (XML-parity behaviour: blank
    stays blank). Returns a dict with only the recoverable keys (may be empty).
    """
    params: dict = {}

    # Angular sweep from sphericalBounds (radians -> degrees, elevation->zenith).
    try:
        sb = header["sphericalBounds"]
    except Exception:
        sb = None
    if sb is not None:
        def _node_val(name):
            try:
                return float(sb[name].value())
            except Exception:
                return None
        az_start = _node_val("azimuthStart")
        az_end = _node_val("azimuthEnd")
        el_min = _node_val("elevationMinimum")
        el_max = _node_val("elevationMaximum")
        if az_start is not None and az_end is not None:
            params["phi_min"] = math.degrees(az_start)
            params["phi_max"] = math.degrees(az_end)
        if el_min is not None and el_max is not None:
            # zenith = 90 - elevation; min/max swap under the negation.
            params["theta_min"] = 90.0 - math.degrees(el_max)
            params["theta_max"] = 90.0 - math.degrees(el_min)

    # Grid resolution from indexBounds (row/column counts = sample counts), but
    # only when the scan is genuinely a raster (carried row/column indices) —
    # otherwise pye57's degenerate indexBounds would fake a zenith resolution.
    try:
        ib = header["indexBounds"] if has_grid else None
    except Exception:
        ib = None
    if ib is not None:
        def _ib_count(min_name, max_name):
            try:
                lo = int(ib[min_name].value())
                hi = int(ib[max_name].value())
            except Exception:
                return None
            n = hi - lo + 1
            return n if n > 1 else None
        n_rows = _ib_count("rowMinimum", "rowMaximum")
        n_cols = _ib_count("columnMinimum", "columnMaximum")
        # E57 grids are row=elevation/theta, column=azimuth/phi.
        if n_rows is not None:
            params["n_theta"] = n_rows
        if n_cols is not None:
            params["n_phi"] = n_cols

    return params


def _e57_to_las(source_path: _Path, out_las: _Path) -> tuple[int, List[dict]]:
    """Convert an E57 to LAS, recovering sky/miss points from the structured
    grid so the LAD inversion can account for beams that returned nothing.

    E57 carries each scan as a structured grid; a cell with no return is flagged
    via `cartesianInvalidState` (or `sphericalInvalidState`). We read the RAW
    scan (all cells, in the scanner-local frame) so misses survive, then:
      - valid cells -> real world point (pose-transformed), `is_miss = 0`;
      - invalid cells (misses) -> `is_miss = 1`. When a direction is recoverable
        (spherical angles, or a non-zero cartesian vector) the miss is PLACED at
        a far-field point `origin + dir * gap_distance` along the pulse direction,
        matching how Helios stores a miss. When it is NOT recoverable here (the
        common cartesian-grid case where invalid cells are zeroed and there are
        no spherical angles), the miss is KEPT and flagged but left at the scanner
        origin: direction recovery from the row/column raster (with tilt/noise
        handling and edge extrapolation) is done downstream in Helios C++. We
        preserve `row_index`/`column_index` so that recovery has the grid to work
        from. Unplaceable misses are counted and surfaced, never silently dropped.

    Each scan is transformed by ITS OWN pose (rotation + translation); a
    multi-scan E57 merges into one cloud with every scan's misses placed relative
    to its own origin. The first scan's origin is stashed in `_e57_scan_meta`
    (plus per-scan origins + the unplaceable-miss count) for the create endpoint.
    Intensity is normalised per scan from its observed valid range (E57 intensity
    is often 0..1 float); RGB colour (colorRed/Green/Blue) is carried into the LAS
    when present (misses get black). Returns (n, extra_dims) with `is_miss` always
    present and `row_index`/`column_index` present when the source carried a grid.
    """
    import laspy  # local: only when this code path runs
    import pye57

    e = pye57.E57(str(source_path))
    try:
        n_scans = e.scan_count
        if n_scans == 0:
            raise HTTPException(status_code=400, detail="E57 file has no scans.")

        all_xyz: List[np.ndarray] = []
        all_miss: List[np.ndarray] = []
        all_intensity: List[np.ndarray] = []
        all_rgb: List[np.ndarray] = []
        all_row: List[np.ndarray] = []
        all_col: List[np.ndarray] = []
        any_intensity = False
        any_color = False
        any_grid = False
        scan_origins: List[list] = []
        scan_params_list: List[dict] = []
        unplaceable_misses = 0

        for si in range(n_scans):
            header = e.get_header(si)
            try:
                rot = np.asarray(header.rotation_matrix, dtype=np.float64).reshape(3, 3)
            except Exception:
                rot = np.eye(3)
            try:
                trans = np.asarray(header.translation, dtype=np.float64).reshape(3)
            except Exception:
                trans = np.zeros(3)
            scan_origins.append(trans.tolist())

            raw = e.read_scan_raw(si)
            keys = set(raw.keys())

            # Local-frame cartesian for every cell (misses may be zeroed).
            if {"cartesianX", "cartesianY", "cartesianZ"} <= keys:
                local = np.column_stack([
                    np.asarray(raw["cartesianX"], dtype=np.float64),
                    np.asarray(raw["cartesianY"], dtype=np.float64),
                    np.asarray(raw["cartesianZ"], dtype=np.float64),
                ])
            elif {"sphericalRange", "sphericalAzimuth", "sphericalElevation"} <= keys:
                rng = np.asarray(raw["sphericalRange"], dtype=np.float64)
                az = np.asarray(raw["sphericalAzimuth"], dtype=np.float64)
                el = np.asarray(raw["sphericalElevation"], dtype=np.float64)
                local = np.column_stack([
                    rng * np.cos(el) * np.cos(az),
                    rng * np.cos(el) * np.sin(az),
                    rng * np.sin(el),
                ])
            else:
                raise HTTPException(
                    status_code=400,
                    detail="E57 scan has neither cartesian nor spherical point data.",
                )

            n_cell = local.shape[0]
            # Recover this scan's scan-pattern parameters (angular sweep + grid
            # resolution) so a non-XML E57 import auto-populates ScanParameters,
            # the same way an XML import does. Whatever the file omits is left
            # out and the renderer falls back to its default. Grid resolution is
            # only trusted when the scan is a true raster (carried row/column
            # indices) — see _e57_scan_params.
            cell_has_grid = {"rowIndex", "columnIndex"} <= keys
            sp = _e57_scan_params(header, cell_has_grid)
            sp["origin"] = trans.tolist()
            scan_params_list.append(sp)

            # Miss flag: prefer the explicit invalid-state field. invalidState
            # is 0 for a good return, non-zero for a miss/invalid cell.
            miss = np.zeros(n_cell, dtype=bool)
            for inv_key in ("cartesianInvalidState", "sphericalInvalidState"):
                if inv_key in keys:
                    miss = np.asarray(raw[inv_key]).astype(np.int64) != 0
                    break

            # Per-cell pulse direction (local frame). For misses with zeroed
            # cartesian, recover direction from spherical angles when present.
            local_dir = local.copy()
            if miss.any() and {"sphericalAzimuth", "sphericalElevation"} <= keys:
                az = np.asarray(raw["sphericalAzimuth"], dtype=np.float64)
                el = np.asarray(raw["sphericalElevation"], dtype=np.float64)
                sdir = np.column_stack([
                    np.cos(el) * np.cos(az),
                    np.cos(el) * np.sin(az),
                    np.sin(el),
                ])
                local_dir[miss] = sdir[miss]

            dir_norm = np.linalg.norm(local_dir, axis=1)
            # A miss we can place here has a usable direction; one we can't
            # (zeroed cartesian, no spherical) is KEPT and flagged for Helios to
            # recover from the grid, not dropped.
            placeable_miss = miss & (dir_norm >= 1e-9)
            unplaceable_miss = miss & (dir_norm < 1e-9)
            unplaceable_misses += int(unplaceable_miss.sum())

            # World coords: pose-transform real hits; far-field place the misses
            # we can; leave unplaceable misses at the scanner origin (flagged).
            world = (rot @ local.T).T + trans
            if placeable_miss.any():
                unit = local_dir[placeable_miss] / dir_norm[placeable_miss][:, None]
                world_dir = (rot @ unit.T).T
                world[placeable_miss] = trans + world_dir * _MISS_GAP_DISTANCE
            if unplaceable_miss.any():
                world[unplaceable_miss] = trans  # at origin until C++ recovery

            all_xyz.append(world)
            all_miss.append(miss.astype(np.float32))

            # Preserve the structured-grid indices so downstream (Helios C++) can
            # recover unplaceable-miss directions from the raster.
            if {"rowIndex", "columnIndex"} <= keys:
                all_row.append(np.asarray(raw["rowIndex"], dtype=np.float32))
                all_col.append(np.asarray(raw["columnIndex"], dtype=np.float32))
                any_grid = True
            else:
                all_row.append(np.full(n_cell, -1.0, dtype=np.float32))
                all_col.append(np.full(n_cell, -1.0, dtype=np.float32))

            if "intensity" in keys:
                inten = np.asarray(raw["intensity"], dtype=np.float64)
                # Normalise from the VALID cells' observed range to LAS uint16.
                # E57 intensity is commonly 0..1 float; a flat clip(0,65535) would
                # crush it to ~0. Misses have no real return -> 0.
                valid = ~miss
                if valid.any():
                    lo = float(np.nanmin(inten[valid]))
                    hi = float(np.nanmax(inten[valid]))
                else:
                    lo, hi = 0.0, 1.0
                span = hi - lo
                if span > 1e-12:
                    scaled = (inten - lo) / span * 65535.0
                else:
                    scaled = np.zeros_like(inten)
                scaled[miss] = 0.0
                all_intensity.append(np.clip(scaled, 0, 65535).astype(np.float32))
                any_intensity = True
            else:
                all_intensity.append(np.zeros(n_cell, dtype=np.float32))

            # RGB colour, when the scan carries it. E57 stores per-channel
            # colorRed/Green/Blue, usually uint8 0..255 but the spec allows other
            # integer ranges (declared in the file's colorLimits). Normalise each
            # channel from its declared/observed max to 8-bit, then store as the
            # LAS uint16 convention (8-bit << 8) used by the PLY/PCD paths. Misses
            # get black — they have no real return.
            if {"colorRed", "colorGreen", "colorBlue"} <= keys:
                cr = np.asarray(raw["colorRed"], dtype=np.float64)
                cg = np.asarray(raw["colorGreen"], dtype=np.float64)
                cb = np.asarray(raw["colorBlue"], dtype=np.float64)
                cmax = float(max(cr.max(initial=0.0), cg.max(initial=0.0),
                                 cb.max(initial=0.0)))
                # Scale so the brightest channel value maps to 255. Files that are
                # already 0..255 (the common case) pass through unchanged; a
                # 0..1 or 0..65535 file is brought into 8-bit range.
                scale = (255.0 / cmax) if cmax > 255.0 or (0.0 < cmax <= 1.0) else 1.0
                rgb8 = np.clip(
                    np.column_stack([cr, cg, cb]) * scale, 0, 255
                ).astype(np.uint16)
                rgb8[miss] = 0
                all_rgb.append(rgb8)
                any_color = True
            else:
                all_rgb.append(np.zeros((n_cell, 3), dtype=np.uint16))
    finally:
        e.close()

    xyz = np.concatenate(all_xyz, axis=0) if all_xyz else np.empty((0, 3), np.float64)
    is_miss = np.concatenate(all_miss, axis=0) if all_miss else np.empty((0,), np.float32)
    intensity = (np.concatenate(all_intensity, axis=0) if all_intensity
                 else np.empty((0,), np.float32))
    rgb = (np.concatenate(all_rgb, axis=0) if all_rgb
           else np.empty((0, 3), np.uint16))
    row_idx = np.concatenate(all_row, axis=0) if all_row else np.empty((0,), np.float32)
    col_idx = np.concatenate(all_col, axis=0) if all_col else np.empty((0,), np.float32)
    n = int(xyz.shape[0])

    extra_dims = [{"slug": _MISS_SLUG, "label": _MISS_LABEL}]
    if any_grid:
        extra_dims.append({"slug": "row_index", "label": "Row Index"})
        extra_dims.append({"slug": "column_index", "label": "Column Index"})

    header = laspy.LasHeader(point_format=3, version="1.4")
    header.scales = np.array([0.001, 0.001, 0.001], dtype=np.float64)
    # Offset to the data min so projected (e.g. UTM) coordinates fit the LAS
    # 32-bit int range — see _session_to_las for the full rationale.
    header.offsets = (np.floor(xyz[:, :3].min(axis=0)) if n > 0 else np.zeros(3))
    for ed in extra_dims:
        header.add_extra_dim(laspy.ExtraBytesParams(name=ed["slug"], type=np.float32))

    record = laspy.ScaleAwarePointRecord.zeros(n, header=header)
    if n > 0:
        record.x = xyz[:, 0]
        record.y = xyz[:, 1]
        record.z = xyz[:, 2]
        if any_intensity:
            record.intensity = np.clip(intensity, 0, 65535).astype(np.uint16)
        if any_color:
            # Point format 3 carries RGB; *256 lifts 8-bit into the 16-bit LAS
            # channel, matching _ply_to_las / _pcd_to_las so colours render
            # identically regardless of source format.
            record.red = rgb[:, 0] * 256
            record.green = rgb[:, 1] * 256
            record.blue = rgb[:, 2] * 256
        record[_MISS_SLUG] = is_miss
        if any_grid:
            record["row_index"] = row_idx
            record["column_index"] = col_idx

    with laspy.open(str(out_las), mode="w", header=header) as writer:
        writer.write_points(record)

    _e57_scan_meta[str(out_las.resolve())] = {
        "origin": (scan_origins[0] if scan_origins else [0.0, 0.0, 0.0]),
        "scan_origins": scan_origins,
        # Full scan-pattern parameters of the first scan (origin + whatever
        # angular sweep / grid resolution the file carried). create_cloud_session
        # forwards this so a lone-E57 import auto-creates a Scan with populated
        # ScanParameters, mirroring the Helios-XML import path.
        "scan_params": (scan_params_list[0] if scan_params_list else {"origin": [0.0, 0.0, 0.0]}),
        "has_misses": bool(is_miss.any()),
        "miss_count": int(is_miss.sum()),
        "unplaceable_miss_count": unplaceable_misses,
    }
    return n, extra_dims


def _las_extra_dim_labels(source_path: _Path) -> List[dict]:
    """Read a LAS/LAZ file's extra-dimension names (header only) so a passed-
    through octree still gets a slug→label sidecar.

    LAS/LAZ feed PotreeConverter unchanged, so their native extra dimensions
    already survive into the octree — but without a sidecar the renderer shows
    raw slugs. The LAS field name is already canonical, so slug == label.
    Reading just the header is cheap (no point data is decoded).
    """
    import laspy  # local: only when this code path runs

    try:
        with laspy.open(str(source_path)) as reader:
            dims = [d.name for d in reader.header.point_format.extra_dimensions]
    except Exception:
        # A malformed/unreadable header shouldn't block conversion; the file
        # is handed to PotreeConverter regardless, and the renderer falls back
        # to raw slugs when the sidecar is absent.
        return []
    # Drop a complete beam-origin triple: `_read_las_into_arrays` consumes it into
    # the float64 `beam_origins` array, so it's not a renderer scalar field and
    # shouldn't get a slug→label sidecar entry (mirrors `_preview_las`).
    dims_lower = {d.lower() for d in dims}
    origin_cols: "set[str]" = set()
    for triple in _BEAM_ORIGIN_ALIAS_SETS:
        if all(ax in dims_lower for ax in triple):
            origin_cols = set(triple)
            break
    return [{"slug": d, "label": d} for d in dims if d.lower() not in origin_cols]


def _source_to_las(source_path: _Path, ascii_format: Optional[str], work_dir: _Path,
                   column_plan: "Optional[ColumnPlan]" = None,
                   ) -> "tuple[_Path, bool, List[dict], Optional[np.ndarray], Optional[np.ndarray]]":
    """Get a LAS file path for `source_path`, converting from another format
    if needed.

    Returns (las_path, is_temp, extra_dims, full_xyz, beam_origins) — caller
    deletes the file if is_temp. `extra_dims` is the [{slug, label}, ...] list of
    carried scalar attributes (read from the header for LAS/LAZ; derived during
    conversion for XYZ/PLY; empty for PCD, which carries position + RGB only).

    `full_xyz` is the (N,3) float64 SOURCE-PRECISION coordinate array for the
    XYZ-family branch, where the LAS we synthesise is 1 mm-quantized and so must
    not be the session's source of truth (1 mm shatters triangulation). It is
    None for every other branch: LAS/LAZ keep their own header scale (the user's
    original precision), and PLY/PCD/E57 already read positions losslessly into
    their LAS. Callers that need the session array use `full_xyz` when present
    and otherwise fall back to reading the LAS.

    `beam_origins` is the (N,3) float64 per-pulse origin array for the XYZ-family
    branch when the column plan carried an ox/oy/oz triple, else None — captured
    at full precision alongside `full_xyz` (origins are world/UTM coords that must
    not pass through the 1 mm LAS / float32 extras). For LAS/LAZ it is None here:
    those origins live in ExtraBytes and are read later by `_read_las_into_arrays`.
    PLY/PCD/E57 do not carry an ASCII origin triple, so None.

    `column_plan` (import wizard) applies only to the XYZ-family branch; PLY/PCD/
    LAS define their own layout and ignore it.
    """
    ext = source_path.suffix.lower().lstrip(".")
    if ext in ("las", "laz"):
        return source_path, False, _las_extra_dim_labels(source_path), None, None
    if ext in _PANDAS_EXTENSIONS:
        out = work_dir / (source_path.stem + ".las")
        _, extra_dims, full_xyz, beam_origins = _xyz_to_las(
            source_path, ascii_format, out, column_plan,
            capture_full_xyz=True, capture_origins=True)
        return out, True, extra_dims, full_xyz, beam_origins
    if ext == "ply":
        out = work_dir / (source_path.stem + ".las")
        _, extra_dims = _ply_to_las(source_path, out)
        return out, True, extra_dims, None, None
    if ext == "pcd":
        out = work_dir / (source_path.stem + ".las")
        _, extra_dims = _pcd_to_las(source_path, out)
        return out, True, extra_dims, None, None
    if ext == "e57":
        out = work_dir / (source_path.stem + ".las")
        _, extra_dims = _e57_to_las(source_path, out)
        return out, True, extra_dims, None, None
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported source extension for octree conversion: .{ext}",
    )


def _run_potree_converter(input_las: _Path, out_dir: _Path) -> None:
    """Invoke PotreeConverter on input_las, writing to out_dir.

    PyInstaller-bundled Pythons inject DYLD_LIBRARY_PATH / LD_LIBRARY_PATH
    pointing at the bundle's libs. Those collide with PotreeConverter's
    expectation of system libs, so we scrub them before spawning the child.
    """
    converter = _resolve_potree_converter_path()
    out_dir.mkdir(parents=True, exist_ok=True)

    env = _os.environ.copy()
    for var in ("DYLD_LIBRARY_PATH", "DYLD_FALLBACK_LIBRARY_PATH", "LD_LIBRARY_PATH"):
        env.pop(var, None)

    # No --attributes filter: PotreeConverter's default writes the full
    # LAS attribute schema (37 bytes/point). The renderer's potree-core
    # decoder is laid out for that exact byte ordering — filtering with
    # `--attributes position rgb intensity` produces a 20-byte stride but
    # the metadata's two `position` entries (Potree 2.0's morton-encoded
    # double-uint32 position format) make the worker expect 16 bytes per
    # point for position alone, which gives every-other-point garbage on
    # the filtered layout. Trading ~17 bytes/point of cache size for a
    # correct render.
    cmd = [
        str(converter),
        str(input_las),
        "-o", str(out_dir),
    ]
    result = _subprocess.run(cmd, capture_output=True, text=True, env=env)
    if result.returncode != 0:
        # Surface the converter's stderr tail directly so failures are debuggable.
        tail = (result.stderr or result.stdout or "")[-1500:]
        raise HTTPException(
            status_code=500,
            detail=f"PotreeConverter failed (exit {result.returncode}): {tail}",
        )


# Sidecar JSON mapping LAS extra-dimension slugs to human-readable labels.
# Written alongside metadata.json at conversion time and read back by
# _read_octree_metadata so the renderer can show clean picker labels
# (e.g. slug 'Reflectance_dB' → label 'Reflectance [dB]') even on cache hits.
_OCTREE_LABELS_FILENAME = "attribute_labels.json"


def _write_octree_labels(octree_dir: _Path, extra_dims: List[dict]) -> None:
    """Persist the slug→label map for an octree's extra dimensions.

    No-op when there are no extra dims, so LAS/LAZ-sourced octrees don't grow
    an empty sidecar. `extra_dims` is the list returned by `_xyz_to_las`.
    """
    if not extra_dims:
        return
    mapping = {ed["slug"]: ed["label"] for ed in extra_dims}
    (octree_dir / _OCTREE_LABELS_FILENAME).write_text(json.dumps(mapping))


def _read_octree_labels(octree_dir: _Path) -> dict:
    """Load the slug→label sidecar, or {} if absent/unreadable.

    Octrees built before this feature (or from LAS/LAZ) have no sidecar; the
    renderer falls back to showing the raw slug in that case.
    """
    p = octree_dir / _OCTREE_LABELS_FILENAME
    if not p.is_file():
        return {}
    try:
        data = json.loads(p.read_text())
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _read_octree_metadata(octree_dir: _Path) -> dict:
    """Load metadata.json and return a renderer-friendly subset.

    Two quirks of PotreeConverter 2.x output we work around here:
      1. The JSON contains lowercase `inf`/`-inf` literals on uninitialised
         min/max fields for one of the attribute entries. Standard JSON
         rejects these, so we rewrite them to `null` before parsing.
      2. The attribute list often contains a duplicate `position` entry —
         a placeholder that never gets updated. We keep only the first
         occurrence of each attribute name.

    Extra-dimension attributes (carried from unmapped numeric source columns)
    get a human-readable `label` attached from the slug→label sidecar so the
    renderer's scalar picker can show clean names.
    """
    labels = _read_octree_labels(octree_dir)
    meta_path = octree_dir / "metadata.json"
    if not meta_path.is_file():
        raise HTTPException(
            status_code=500,
            detail=f"metadata.json missing from octree dir: {octree_dir}",
        )
    raw_text = meta_path.read_text()
    raw_text = re.sub(r'(?<![\w\.])-?inf(?![\w\.])', 'null', raw_text)
    raw_text = re.sub(r'(?<![\w\.])nan(?![\w\.])', 'null', raw_text)
    raw = json.loads(raw_text)
    bbox = raw.get("boundingBox", {})

    # PotreeConverter writes the FULL position attribute schema as two
    # entries: a primary "position" with the true min/max bounds, and a
    # second "position" with morton-encoded extension bits (uninitialised
    # min/max, hence the `inf` literals). The first occurrence is the one
    # we want for tight bounds; subsequent ones get dropped here.
    seen_attrs: set[str] = set()
    deduped = []
    tight_bounds = None
    for a in raw.get("attributes", []):
        name = a.get("name")
        if not name or name in seen_attrs:
            continue
        seen_attrs.add(name)
        # Preserve per-attribute min/max when present. The renderer needs
        # these for the intensity / height shaders' uniform ranges
        # (intensityRange, heightMin/Max) — without them the shader maps
        # every point to the same gradient sample and the cloud renders
        # as a solid colour.
        amin = a.get("min")
        amax = a.get("max")
        entry: dict = {
            "name": name,
            "size": int(a.get("size", 0)),
            "type": a.get("type"),
            "num_elements": int(a.get("numElements", 0)),
        }
        if name in labels:
            entry["label"] = labels[name]
        if isinstance(amin, list) and isinstance(amax, list):
            # Filter out None entries (came from `inf` rewrite).
            if all(v is not None for v in amin) and all(v is not None for v in amax):
                entry["min"] = [float(v) for v in amin]
                entry["max"] = [float(v) for v in amax]
        deduped.append(entry)
        if name == "position" and tight_bounds is None:
            mn = a.get("min")
            mx = a.get("max")
            # Skip if the values were rewritten to None by the `inf` sub.
            if (isinstance(mn, list) and isinstance(mx, list)
                and len(mn) == 3 and len(mx) == 3
                and all(v is not None for v in mn + mx)):
                tight_bounds = {"min": [float(v) for v in mn],
                                "max": [float(v) for v in mx]}

    return {
        "version": raw.get("version", "2.0"),
        "point_count": int(raw.get("points", 0)),
        "spacing": float(raw.get("spacing", 0.0)),
        "scale": list(raw.get("scale", [1.0, 1.0, 1.0])),
        "offset": list(raw.get("offset", [0.0, 0.0, 0.0])),
        # `bounds` is PotreeConverter's cube-padded octree extent (used by
        # the loader for LOD math). `tight_bounds` is the actual data
        # extent — what the UI should use for camera framing and crop-box
        # initialisation. Fall back to the padded box if the tight values
        # were missing.
        "bounds": {
            "min": list(bbox.get("min", [0.0, 0.0, 0.0])),
            "max": list(bbox.get("max", [0.0, 0.0, 0.0])),
        },
        "tight_bounds": tight_bounds or {
            "min": list(bbox.get("min", [0.0, 0.0, 0.0])),
            "max": list(bbox.get("max", [0.0, 0.0, 0.0])),
        },
        "attributes": deduped,
    }


def _dir_total_size(p: _Path) -> int:
    """Sum of sizes of regular files at any depth under p. Symlinks ignored."""
    total = 0
    for child in p.rglob("*"):
        try:
            if child.is_file() and not child.is_symlink():
                total += child.stat().st_size
        except (FileNotFoundError, PermissionError):
            # File can disappear between iterdir and stat (concurrent eviction).
            continue
    return total


def _evict_octree_cache(max_bytes: int,
                        keep: "Optional[Union[_Path, Iterable[_Path]]]" = None) -> List[str]:
    """Trim the octree cache to at most `max_bytes` of regular file content,
    removing oldest-accessed cache directories first. Returns the cache_ids
    that were evicted.

    `keep`, if provided, is one path OR an iterable of paths that are never
    evicted — pass the cache dir(s) we just wrote (e.g. a hits octree AND its
    sibling miss octree) so a fresh convert doesn't immediately drop itself when
    the cache is at the limit.
    """
    root = _octree_cache_root()
    if not root.is_dir():
        return []

    # Normalise `keep` to a set of resolved paths so callers can pass one path
    # or several (the hits + miss octrees built in the same bake).
    if keep is None:
        keep_set: set = set()
    elif isinstance(keep, _Path):
        keep_set = {keep.resolve()}
    else:
        keep_set = {p.resolve() for p in keep}

    # Skip the .staging dirs and any non-sha1 entries (defensive against
    # files dropped in here by other tools).
    entries: list[tuple[float, _Path]] = []
    for child in root.iterdir():
        if not child.is_dir():
            continue
        if child.name.endswith(".staging"):
            continue
        if len(child.name) != 40 or not all(c in "0123456789abcdef" for c in child.name):
            continue
        try:
            atime = child.stat().st_atime
        except FileNotFoundError:
            continue
        entries.append((atime, child))

    # Oldest first.
    entries.sort(key=lambda e: e[0])

    total = sum(_dir_total_size(p) for _, p in entries)
    evicted: List[str] = []
    if total <= max_bytes:
        return evicted

    for _, candidate in entries:
        if total <= max_bytes:
            break
        if candidate.resolve() in keep_set:
            continue
        size = _dir_total_size(candidate)
        try:
            _shutil.rmtree(candidate)
        except (FileNotFoundError, PermissionError):
            continue
        evicted.append(candidate.name)
        total -= size

    return evicted


# ---------------------------------------------------------------------------
# Import-wizard preview
# ---------------------------------------------------------------------------

# How a value column reads, used to pre-tick the wizard's "categorical" box.
_CATEGORICAL_MAX_DISTINCT = 32


def _detect_ascii_delimiter(file_path: str) -> Optional[str]:
    """Sniff the delimiter of the first data row. Mirrors the renderer's
    detectDelimiter precedence (comma → tab → semicolon → whitespace). Returns a
    human label ('comma'/'tab'/'semicolon'/'whitespace') or None if no data."""
    with open(file_path) as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
            # Skip a leading text header row — we want a data row's delimiter.
            if any(re.search(r'[a-zA-Z]', tok) for tok in line.split()):
                continue
            if ',' in line:
                return 'comma'
            if '\t' in line:
                return 'tab'
            if ';' in line:
                return 'semicolon'
            return 'whitespace'
    return None


def _split_ascii_row(line: str, delimiter: Optional[str]) -> List[str]:
    if delimiter == 'comma':
        return [t.strip() for t in line.split(',')]
    if delimiter == 'tab':
        return [t.strip() for t in line.split('\t')]
    if delimiter == 'semicolon':
        return [t.strip() for t in line.split(';')]
    return line.split()


def _read_ascii_sample_rows(file_path: str, delimiter: Optional[str],
                            skip_header: bool, max_rows: int) -> List[List[str]]:
    """Return up to `max_rows` data rows as raw string tokens. Reads at most
    that many lines — never the whole file.

    `skip_header` drops a single leading header row. A *commented* header
    ('# x y z ...') is already dropped by the comment-line skip below, so we
    must NOT also consume the first data row — otherwise the wizard preview
    would silently lose row one. Detect that case up front."""
    # A commented header is removed by the '#'/'//' filter, so there is no
    # uncommented header row left to skip.
    first = _first_nonblank_ascii_line(file_path)
    header_is_commented = first is not None and first[1]
    rows: List[List[str]] = []
    header_skipped = (not skip_header) or header_is_commented
    with open(file_path) as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
            if not header_skipped:
                header_skipped = True
                continue
            rows.append(_split_ascii_row(line, delimiter))
            if len(rows) >= max_rows:
                break
    return rows


def _column_type_hint(values: List[str]) -> str:
    """Sniff a value shape from sampled string tokens. 'categorical' when the
    column is small non-negative integers with few distinct values (a class /
    label column); else integer/float/empty."""
    nonblank = [v for v in values if v != '']
    if not nonblank:
        return 'empty'
    all_int = all(re.fullmatch(r'-?\d+', v) is not None for v in nonblank)
    if all_int:
        ints = [int(v) for v in nonblank]
        distinct = set(ints)
        if all(i >= 0 for i in ints) and len(distinct) <= _CATEGORICAL_MAX_DISTINCT:
            return 'categorical'
        return 'integer'
    all_float = True
    for v in nonblank:
        try:
            float(v)
        except ValueError:
            all_float = False
            break
    return 'float' if all_float else 'categorical'


def _preview_ascii(file_path: str, ascii_format: Optional[str],
                   max_rows: int) -> PointCloudPreviewResponse:
    header_names = _read_ascii_header_names(file_path)
    has_header = header_names is not None
    delimiter = _detect_ascii_delimiter(file_path)
    sample_rows = _read_ascii_sample_rows(file_path, delimiter, has_header, max_rows)

    roles = (_tokenize_ascii_format(ascii_format)
             if ascii_format
             else _autodetect_xyz_columns(file_path))

    # The name shown as each column's heading in the wizard. Prefer a real file
    # header; otherwise, if the XML supplied an <ASCII_format> legend, use its
    # raw tokens so the heading reads the legend word (e.g. 'row', 'reflectance')
    # exactly as it would for a commented '# x y z ...' header. The file genuinely
    # has no header row, so `has_header` and the sample-row skipping above stay
    # bound to the real file — this only names the columns for display.
    display_names = header_names
    if display_names is None and ascii_format:
        display_names = ascii_format.split()
    # Determine column count from the widest of (roles, header, first row).
    ncols = len(roles)
    if header_names:
        ncols = max(ncols, len(header_names))
    if sample_rows:
        ncols = max(ncols, max(len(r) for r in sample_rows))

    columns: List[PreviewColumn] = []
    for i in range(ncols):
        role = roles[i] if i < len(roles) else 'skip'
        # Re-derive the same default slug/label _plan_columns would assign for
        # THIS position, so the wizard's suggestion matches import before edits.
        if role in _XYZ_RESERVED_ROLES and role != 'skip':
            detected_role = role
            suggested_slug = ''
            suggested_label = (header_names[i] if header_names and i < len(header_names)
                               and header_names[i] else role)
        elif role in _MULTI_RETURN_SLUGS:
            # A per-pulse multi-return column (timestamp/target_index/
            # target_count): pin the canonical slug/label regardless of header
            # text, exactly as `_plan_columns` does, so the wizard's column plan
            # carries it under the name the LAD accessor recovers it by. Without
            # this the column defaulted to a positional 'col_N' slug, the LAD
            # path failed to find the three multi-return columns, and the
            # full-waveform inversion ran on zeroed data.
            detected_role = 'extra'
            suggested_slug = role
            suggested_label = _MULTI_RETURN_LABELS[role]
        elif role in _GRID_INDEX_SLUGS:
            # A structured-scan raster index (row_index/column_index): the wizard
            # exposes these as dedicated dropdown roles, so report the role token
            # directly (not 'extra') to pre-select the matching option. The slug
            # is pinned to the canonical name so the recovery path finds the grid.
            detected_role = role
            suggested_slug = role
            suggested_label = _GRID_INDEX_LABELS[role]
        elif role == _MISS_SLUG:
            # A sky/miss flag: the wizard exposes a dedicated 'Miss Flag' role, so
            # report the canonical is_miss role token directly (not 'extra') to
            # pre-select it — mirroring the grid-index roles above. The slug/label
            # are pinned to the canonical name (regardless of the source spelling
            # is_miss/miss/sky) so the LAD path and the renderer's fixed Hit/Miss
            # colour scheme find it by name, matching `_plan_columns`.
            detected_role = _MISS_SLUG
            suggested_slug = _MISS_SLUG
            suggested_label = _MISS_LABEL
        elif role in _ORIGIN_SLUGS:
            # A per-pulse beam-origin column (auto-detected from an ox/oy/oz
            # header): the wizard exposes dedicated 'Beam Origin X/Y/Z' roles, so
            # report the canonical origin role token directly (not 'extra') to
            # pre-select the matching option — mirroring the grid-index/miss roles
            # above. These are captured as full-precision float64 origins (see
            # `_plan_columns`), never as a float32 extra, so there's no extra-dim
            # slug to suggest; the role token is what the column plan carries.
            detected_role = role
            suggested_slug = role
            suggested_label = _ORIGIN_LABELS[role]
        else:
            detected_role = 'extra' if role != 'skip' else 'skip'
            if header_names is not None and i < len(header_names) and header_names[i]:
                suggested_label = _humanize_extra_dim_label(header_names[i])
                suggested_slug = _sanitize_extra_dim_name(header_names[i])
            elif role != 'skip':
                # No file header, but the role token is a passed-through
                # <ASCII_format> legend word — label/slug the column from it so
                # the wizard's suggestion matches import (see `_plan_columns`).
                suggested_label = _humanize_extra_dim_label(role)
                suggested_slug = _sanitize_extra_dim_name(role)
            else:
                suggested_label = f"Column {i + 1}"
                suggested_slug = f"col_{i + 1}"
        col_values = [r[i] for r in sample_rows if i < len(r)]
        columns.append(PreviewColumn(
            index=i,
            header_name=(display_names[i] if display_names and i < len(display_names) else None),
            detected_role=detected_role,
            suggested_label=suggested_label,
            suggested_slug=suggested_slug,
            type_hint=_column_type_hint(col_values),
            remappable=True,
        ))

    return PointCloudPreviewResponse(
        kind='ascii', delimiter=delimiter, has_header=has_header,
        columns=columns, sample_rows=sample_rows,
    )


def _ply_header_properties(file_path: str) -> tuple[List[str], bool, List[List[str]], int]:
    """Parse a PLY header without decoding the body. Returns
    (vertex_property_names, is_ascii, sample_rows, vertex_count). sample_rows is
    empty for binary PLY (we don't decode it for preview)."""
    props: List[str] = []
    fmt = 'ascii'
    vertex_count = 0
    in_vertex = False
    header_lines = 0
    with open(file_path, 'rb') as f:
        for raw in f:
            header_lines += 1
            line = raw.decode('latin-1', errors='replace').strip()
            low = line.lower()
            if low.startswith('format'):
                fmt = low.split()[1] if len(low.split()) > 1 else 'ascii'
            elif low.startswith('element '):
                parts = line.split()
                in_vertex = len(parts) >= 2 and parts[1].lower() == 'vertex'
                if in_vertex and len(parts) >= 3:
                    try:
                        vertex_count = int(parts[2])
                    except ValueError:
                        vertex_count = 0
            elif low.startswith('property') and in_vertex:
                props.append(line.split()[-1])
            elif low == 'end_header':
                break
            if header_lines > 1000:  # malformed / not really a PLY
                break

    sample_rows: List[List[str]] = []
    if fmt == 'ascii':
        with open(file_path) as f:
            past_header = False
            taken = 0
            for raw in f:
                s = raw.strip()
                if not past_header:
                    if s.lower() == 'end_header':
                        past_header = True
                    continue
                if not s:
                    continue
                sample_rows.append(s.split())
                taken += 1
                if taken >= 20:
                    break
    return props, fmt == 'ascii', sample_rows, vertex_count


def _ply_role_for(name: str) -> str:
    n = name.lower()
    if n in ('x', 'y', 'z'):
        return n
    if n in ('red', 'green', 'blue', 'r', 'g', 'b'):
        return {'red': 'r255', 'green': 'g255', 'blue': 'b255',
                'r': 'r255', 'g': 'g255', 'b': 'b255'}[n]
    if n in ('intensity', 'scalar_intensity', 'reflectance'):
        return 'intensity'
    return 'extra'


def _preview_ply(file_path: str) -> PointCloudPreviewResponse:
    props, is_ascii, sample_rows, _ = _ply_header_properties(file_path)
    columns: List[PreviewColumn] = []
    for i, name in enumerate(props):
        # `is_miss`/`miss`/`sky` is a system-managed sky/miss flag, not a
        # user-mappable column — keep it out of the wizard.
        if name.lower() in _MISS_ALIASES:
            continue
        role = _ply_role_for(name)
        is_extra = role == 'extra'
        col_values = [r[i] for r in sample_rows if i < len(r)]
        columns.append(PreviewColumn(
            index=i, header_name=name, detected_role=role,
            suggested_label=_humanize_extra_dim_label(name) if is_extra else name,
            suggested_slug=_sanitize_extra_dim_name(name) if is_extra else '',
            type_hint=_column_type_hint(col_values) if sample_rows else ('float' if is_extra else 'float'),
            remappable=False,
        ))
    warning = None if is_ascii else "Binary PLY: preview rows unavailable (fields shown from header)."
    return PointCloudPreviewResponse(
        kind='ply', delimiter=None, has_header=True,
        columns=columns, sample_rows=sample_rows, warning=warning,
    )


def _preview_pcd(file_path: str) -> PointCloudPreviewResponse:
    fields: List[str] = []
    with open(file_path, 'rb') as f:
        for raw in f:
            line = raw.decode('latin-1', errors='replace').strip()
            low = line.lower()
            if low.startswith('fields'):
                fields = line.split()[1:]
            elif low.startswith('data'):
                break
    columns: List[PreviewColumn] = []
    for i, name in enumerate(fields):
        n = name.lower()
        role = ('x' if n == 'x' else 'y' if n == 'y' else 'z' if n == 'z'
                else 'r255' if n in ('r', 'red') else 'g255' if n in ('g', 'green')
                else 'b255' if n in ('b', 'blue') else 'r255' if n == 'rgb' else 'extra')
        columns.append(PreviewColumn(
            index=i, header_name=name, detected_role=role,
            suggested_label=name, suggested_slug='', type_hint='float',
            remappable=False,
        ))
    return PointCloudPreviewResponse(
        kind='pcd', delimiter=None, has_header=True, columns=columns,
        sample_rows=[],
        warning="PCD import preserves position and RGB only; other scalar fields are dropped.",
    )


def _preview_las(file_path: str, max_rows: int) -> PointCloudPreviewResponse:
    import laspy
    columns: List[PreviewColumn] = []
    sample_rows: List[List[str]] = []
    with laspy.open(file_path) as reader:
        header = reader.header
        std = ['X', 'Y', 'Z']
        pf = header.point_format
        names = [d.name for d in pf.dimensions]
        # Present standard + extra dims; mark x/y/z and rgb/intensity roles.
        def role_for(n: str) -> str:
            ln = n.lower()
            if ln in ('x', 'y', 'z'):
                return ln
            if ln == 'red':
                return 'r255'
            if ln == 'green':
                return 'g255'
            if ln == 'blue':
                return 'b255'
            if ln == 'intensity':
                return 'intensity'
            return 'extra'
        extra_names = {d.name for d in pf.extra_dimensions}
        # Beam-origin ExtraBytes (ox/oy/oz or an alias set) are auto-consumed by
        # `_read_las_into_arrays` into the float64 `beam_origins` array — exactly
        # like x/y/z become positions — so they are NOT user-mappable scalars and
        # must be hidden from the wizard. Only suppress them when a COMPLETE triple
        # is present (matching the reader): a lone `ox` with no `oy`/`oz` isn't an
        # origin set, so it stays a normal scalar column.
        _extra_lower = {nm.lower() for nm in extra_names}
        _origin_cols = set()
        for _triple in _BEAM_ORIGIN_ALIAS_SETS:
            if all(ax in _extra_lower for ax in _triple):
                _origin_cols = set(_triple)
                break
        for i, n in enumerate(names):
            # `is_miss` is a system-managed sky/miss flag, not a user-mappable
            # column — never present it in the wizard (and don't let it be
            # renamed off the canonical slug the renderer/LAD depend on).
            if n.lower() in _MISS_ALIASES:
                continue
            # Beam-origin triple: auto-consumed as float64 `beam_origins`, not a
            # mappable scalar (see comment above) — hide like x/y/z.
            if n.lower() in _origin_cols:
                continue
            role = role_for(n)
            is_extra = n in extra_names or role == 'extra'
            columns.append(PreviewColumn(
                index=i, header_name=n, detected_role=role,
                suggested_label=n, suggested_slug=n if is_extra else '',
                type_hint='categorical' if n.lower() == 'classification' else 'float',
                remappable=False,
            ))
        # Cheap value sample: read a small batch of points.
        try:
            pts = reader.read_points(min(max_rows, 20))
            for k in range(len(pts)):
                row = []
                for n in names:
                    try:
                        row.append(str(pts[n][k]))
                    except Exception:
                        row.append('')
                sample_rows.append(row)
        except Exception:
            sample_rows = []
    return PointCloudPreviewResponse(
        kind='las', delimiter=None, has_header=True, columns=columns,
        sample_rows=sample_rows,
    )


def _preview_e57(file_path: str) -> PointCloudPreviewResponse:
    """Summarise an E57's structure for the import wizard without decoding the
    full point data. Reports the scan count, total points, and which attributes
    the file carries (read cheaply from the first scan's declared `point_fields`,
    no point decode). Columns are fixed (E57 defines its own layout), so they're
    not remappable — we surface position plus whatever of intensity / RGB the
    file actually contains, so the user can see (and colour by) them.

    `is_miss` is deliberately omitted: it's a system-managed flag the converter
    populates, not a user column — showing it as an editable "scalar" would imply
    a rename the import ignores, and the flag is excluded from the octree anyway.
    """
    import pye57
    e = pye57.E57(file_path)
    try:
        n_scans = e.scan_count
        total = 0
        fields: set[str] = set()
        for si in range(n_scans):
            try:
                h = e.get_header(si)
                total += int(h.point_count)
                # Declared field names — available without decoding any points.
                fields.update(getattr(h, "point_fields", []) or [])
            except Exception:
                pass

        columns = [
            PreviewColumn(index=0, header_name='x', detected_role='x',
                          suggested_label='x', suggested_slug='', type_hint='float',
                          remappable=False),
            PreviewColumn(index=1, header_name='y', detected_role='y',
                          suggested_label='y', suggested_slug='', type_hint='float',
                          remappable=False),
            PreviewColumn(index=2, header_name='z', detected_role='z',
                          suggested_label='z', suggested_slug='', type_hint='float',
                          remappable=False),
        ]
        idx = 3
        # Surface the real scalars the converter carries through, so the wizard
        # shows the user what they'll be able to colour by (intensity / RGB).
        if 'intensity' in fields:
            columns.append(PreviewColumn(
                index=idx, header_name='intensity', detected_role='intensity',
                suggested_label='intensity', suggested_slug='', type_hint='float',
                remappable=False))
            idx += 1
        if {'colorRed', 'colorGreen', 'colorBlue'} <= fields:
            for ch, role in (('red', 'r255'), ('green', 'g255'), ('blue', 'b255')):
                columns.append(PreviewColumn(
                    index=idx, header_name=ch, detected_role=role,
                    suggested_label=ch, suggested_slug='', type_hint='float',
                    remappable=False))
                idx += 1

        warn = (f"E57 with {n_scans} scan(s), ~{total} points. Sky/miss points "
                "are recovered from the structured grid and tagged for LAD "
                "(hidden by default; use 'Show sky/miss points' to view).")
    finally:
        e.close()
    return PointCloudPreviewResponse(
        kind='e57', delimiter=None, has_header=True, columns=columns,
        sample_rows=[], warning=warn,
    )


# Any coordinate whose magnitude exceeds this triggers a suggested global shift:
# at ~1e4 m, float32's ~1e-7 relative precision already gives ~1 mm error, and it
# degrades linearly past that (UTM ~1e6 → ~0.1 m). Below it, raw coordinates render
# cleanly and a shift would only be a nuisance.
_SHIFT_SUGGEST_THRESHOLD = 1.0e4


def _suggest_global_shift(file_path: str,
                          response: PointCloudPreviewResponse) -> Optional[List[float]]:
    """Best-effort suggested CloudCompare-style global shift for `file_path`,
    or None when the coordinates are small enough to render cleanly (or the min
    can't be probed cheaply). The suggestion is floor(min) per axis when any
    axis min exceeds `_SHIFT_SUGGEST_THRESHOLD`.

    ASCII: a capped pandas read of just the x/y/z columns (roles come from the
    already-computed preview `columns`). LAS/LAZ: the header mins (no body read).
    Other formats: skipped (None) — the viewer's render-offset still keeps them
    artifact-free; only the explicit import shift is unavailable there."""
    try:
        ext = _Path(file_path).suffix.lower().lstrip('.')
        mins: Optional[np.ndarray] = None
        if ext in _PANDAS_EXTENSIONS:
            role_idx = {c.detected_role: c.index for c in response.columns
                        if c.detected_role in ('x', 'y', 'z')}
            if not all(r in role_idx for r in ('x', 'y', 'z')):
                return None
            usecols = [role_idx['x'], role_idx['y'], role_idx['z']]
            df = pd.read_csv(
                file_path, sep=_ascii_pandas_sep(file_path), header=None,
                comment='#', usecols=usecols, names=['x', 'y', 'z'],
                dtype=np.float64, engine='c', skip_blank_lines=True,
                skiprows=_ascii_skiprows(file_path), on_bad_lines='skip',
                nrows=200_000,  # cap: enough for a stable min without a full read
            ).dropna()
            if df.empty:
                return None
            mins = df.to_numpy().min(axis=0)
        elif ext in ('las', 'laz'):
            import laspy
            with laspy.open(file_path) as reader:
                mins = np.asarray(reader.header.mins, dtype=np.float64)
        if mins is None or mins.shape != (3,) or not np.all(np.isfinite(mins)):
            return None
        if not np.any(np.abs(mins) > _SHIFT_SUGGEST_THRESHOLD):
            return None
        return [float(v) for v in np.floor(mins)]
    except Exception:
        return None  # never block preview on a shift probe


@app.post("/api/pointcloud/preview")
async def preview_pointcloud(request: PointCloudPreviewRequest) -> PointCloudPreviewResponse:
    """Cheaply inspect a point-cloud file for the import wizard.

    Reads only enough of the file to show the wizard what was auto-detected and
    a handful of sample rows. Never 500s on a parse problem — returns a 200 with
    a `warning` and best-effort columns so the wizard can still offer
    "import with auto-detect"."""
    source = _Path(request.file_path).expanduser()
    if not source.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {request.file_path}")
    max_rows = max(1, min(int(request.max_rows or 20), 100))
    ext = source.suffix.lower().lstrip('.')
    try:
        resp: Optional[PointCloudPreviewResponse] = None
        if ext in _PANDAS_EXTENSIONS:
            resp = _preview_ascii(str(source), request.ascii_format, max_rows)
        elif ext == 'ply':
            resp = _preview_ply(str(source))
        elif ext == 'pcd':
            resp = _preview_pcd(str(source))
        elif ext == 'e57':
            resp = _preview_e57(str(source))
        elif ext in ('las', 'laz'):
            resp = _preview_las(str(source), max_rows)
        if resp is not None:
            # Probe coordinate magnitude and pre-fill a suggested global shift for
            # large (e.g. UTM) clouds, so the wizard can offer it on by default.
            resp.suggested_shift = _suggest_global_shift(str(source), resp)
            return resp
    except Exception as e:  # never block import on a preview failure
        return PointCloudPreviewResponse(
            kind=ext or 'ascii', delimiter=None, has_header=False,
            columns=[], sample_rows=[],
            warning=f"Could not preview this file ({e}). You can still import with auto-detect.",
        )
    return PointCloudPreviewResponse(
        kind=ext or 'ascii', delimiter=None, has_header=False,
        columns=[], sample_rows=[],
        warning=f"Unsupported extension for preview: .{ext}.",
    )


def _canonical_translation(translation: Optional[List[float]]) -> str:
    """Stable string form for hashing. None and (0,0,0) collide on purpose —
    a no-op translation has the same cache identity as no translation."""
    if translation is None:
        return ""
    if len(translation) != 3:
        raise HTTPException(
            status_code=400,
            detail="translation, if provided, must be a 3-element [x, y, z] array.",
        )
    if all(float(v) == 0.0 for v in translation):
        return ""
    return ",".join(f"{float(v):.9g}" for v in translation)


def _canonical_region(region: dict) -> str:
    """Stable string form of the crop region for cache keying. Same shape →
    same string; same string → same octree. Polygon points and matrices are
    serialised verbatim because re-cropping with even slightly different
    camera framing produces a different filter mask."""
    kind = region.get("kind")
    if kind == "box":
        mn = region.get("min", [])
        mx = region.get("max", [])
        invert = bool(region.get("invert", False))
        if len(mn) != 3 or len(mx) != 3:
            raise HTTPException(
                status_code=400,
                detail="region.min and region.max must each be 3-element arrays.",
            )
        return "box|{}|{}|{}".format(
            ",".join(f"{float(v):.9g}" for v in mn),
            ",".join(f"{float(v):.9g}" for v in mx),
            "1" if invert else "0",
        )
    if kind == "polygon":
        pts = region.get("points", [])
        proj = region.get("projection", [])
        view = region.get("view", [])
        canvas = region.get("canvas", {})
        invert = bool(region.get("invert", False))
        if not isinstance(pts, list) or len(pts) < 3:
            raise HTTPException(
                status_code=400,
                detail="region.points must have at least 3 [x, y] entries.",
            )
        if len(proj) != 16 or len(view) != 16:
            raise HTTPException(
                status_code=400,
                detail="region.projection and region.view must each be 16-element matrices.",
            )
        w = canvas.get("width")
        h = canvas.get("height")
        if not isinstance(w, (int, float)) or not isinstance(h, (int, float)) or w <= 0 or h <= 0:
            raise HTTPException(
                status_code=400,
                detail="region.canvas must have positive width and height.",
            )
        pts_s = ";".join(f"{float(p[0]):.6g},{float(p[1]):.6g}" for p in pts)
        return "polygon|{}|{}|{}|{}x{}|{}".format(
            pts_s,
            ",".join(f"{float(v):.6g}" for v in proj),
            ",".join(f"{float(v):.6g}" for v in view),
            int(w), int(h),
            "1" if invert else "0",
        )
    if kind == "squares_union":
        centers = region.get("centers", [])
        half_sizes = region.get("half_sizes", [])
        proj = region.get("projection", [])
        view = region.get("view", [])
        canvas = region.get("canvas", {})
        invert = bool(region.get("invert", False))
        if not isinstance(centers, list) or not isinstance(half_sizes, list):
            raise HTTPException(
                status_code=400,
                detail="region.centers and region.half_sizes must be arrays.",
            )
        if len(centers) == 0 or len(centers) != len(half_sizes):
            raise HTTPException(
                status_code=400,
                detail="region.centers and region.half_sizes must be non-empty and the same length.",
            )
        for c in centers:
            if len(c) != 2:
                raise HTTPException(
                    status_code=400,
                    detail="each region.centers entry must be a 2-element [px, py] array.",
                )
        if any(float(h) <= 0 for h in half_sizes):
            raise HTTPException(
                status_code=400,
                detail="every region.half_sizes entry must be positive.",
            )
        if len(proj) != 16 or len(view) != 16:
            raise HTTPException(
                status_code=400,
                detail="region.projection and region.view must each be 16-element matrices.",
            )
        w = canvas.get("width")
        h = canvas.get("height")
        if not isinstance(w, (int, float)) or not isinstance(h, (int, float)) or w <= 0 or h <= 0:
            raise HTTPException(
                status_code=400,
                detail="region.canvas must have positive width and height.",
            )
        squares_s = ";".join(
            "{},{}".format(
                ",".join(f"{float(v):.6g}" for v in c),
                f"{float(hs):.6g}",
            )
            for c, hs in zip(centers, half_sizes)
        )
        return "squares_union|{}|{}|{}|{}x{}|{}".format(
            squares_s,
            ",".join(f"{float(v):.6g}" for v in proj),
            ",".join(f"{float(v):.6g}" for v in view),
            int(w), int(h),
            "1" if invert else "0",
        )
    raise HTTPException(
        status_code=400,
        detail=f"region.kind must be 'box', 'polygon', or 'squares_union'. Got: {kind!r}",
    )


def _canonical_scalar_filters(filters: Optional[List[dict]]) -> str:
    """Stable, order-independent string form of the scalar filters for cache
    keying. Sorted so [{a},{b}] and [{b},{a}] collide on purpose — filter
    order does not change which points survive. Empty/None → "" so a
    scalar-free crop keeps its prior cache identity."""
    if not filters:
        return ""

    def _one(f: dict) -> str:
        vals = f.get("values")
        if vals:
            # Categorical: sorted unique class ids; independent of min/max.
            ids = sorted({int(round(float(v))) for v in vals})
            return "{}:set:{}".format(f["slug"], ",".join(str(i) for i in ids))
        return "{}:{:.9g}:{:.9g}".format(
            f["slug"], float(f["min"]), float(f["max"]),
        )

    parts = sorted(_one(f) for f in filters)
    return "scalar|" + ";".join(parts)


def _resolve_scalar_filter(f: dict) -> tuple:
    """Normalise one scalar-filter spec into `(lo, hi, value_set)`.

    `value_set` is a Python set of int class ids for a categorical filter, or
    None for a continuous [lo, hi] range. Continuous fields keep their float
    bounds; categorical fields ignore them. Shared by both filter functions so
    the membership semantics stay identical across ASCII and LAS sources."""
    vals = f.get("values")
    if vals:
        return float("-inf"), float("inf"), {int(round(float(v))) for v in vals}
    return float(f.get("min", float("-inf"))), float(f.get("max", float("inf"))), None


def _scalar_filter_mask(vals: "np.ndarray", lo: float, hi: float,
                        value_set: Optional[set]) -> "np.ndarray":
    """Boolean keep-mask for one resolved scalar filter over a value array.

    Categorical (value_set given): keep iff round(value) ∈ value_set. Continuous:
    keep iff lo <= value <= hi. Extra dims are float32 on disk, so rounding makes
    integer class labels robust to the float round-trip."""
    if value_set is not None:
        rounded = np.rint(vals).astype(np.int64)
        return np.isin(rounded, list(value_set))
    return (vals >= lo) & (vals <= hi)


def _squares_union_mask(
    pixels: "np.ndarray", centers: "np.ndarray", half_sizes: "np.ndarray",
) -> "np.ndarray":
    """Boolean mask: True for points whose canvas pixel falls inside ANY of the
    screen-space square stamps (the union).

    `pixels` is (n, 2) canvas pixels (from _project_world_to_pixel), `centers`
    is (k, 2) pixel centers, `half_sizes` is (k,) pixel half-extents. The test
    is axis-aligned in screen space: |px - cx| <= h AND |py - cy| <= h. Because
    the membership is purely 2D, a square removes points at every depth behind
    it — a stamp that extrudes through the whole cloud, like the polygon path.
    The erase brush keeps the complement (region invert=True). Shared by both
    filter functions so ASCII and LAS sources erase identically."""
    px = pixels[:, 0]
    py = pixels[:, 1]
    mask = np.zeros(pixels.shape[0], dtype=bool)
    for i in range(centers.shape[0]):
        h = half_sizes[i]
        mask |= (np.abs(px - centers[i, 0]) <= h) & (np.abs(py - centers[i, 1]) <= h)
    return mask


def _region_mask(positions: "np.ndarray", region: Optional[dict]) -> "np.ndarray":
    """Spatial keep-mask for already-materialised Nx3 world positions.

    Supports box / polygon / squares_union regions over a whole positions
    array at once, so the cloud session can mask its in-memory points.
    `region["invert"]` flips the spatial mask. `region` is None → keep all
    (all-True).

    Positions are world-space (translation already applied by the caller). For
    polygon/squares_union the camera matrices and canvas come from `region`,
    matching the renderer's frozen-camera preview. Returns a bool ndarray of
    length N.
    """
    n = positions.shape[0]
    if region is None:
        return np.ones(n, dtype=bool)
    if n == 0:
        return np.zeros(0, dtype=bool)

    kind = region.get("kind")
    if kind == "box":
        cmin = np.asarray(region["min"], dtype=np.float64)
        cmax = np.asarray(region["max"], dtype=np.float64)
        xs, ys, zs = positions[:, 0], positions[:, 1], positions[:, 2]
        mask = (
            (xs >= cmin[0]) & (xs <= cmax[0]) &
            (ys >= cmin[1]) & (ys <= cmax[1]) &
            (zs >= cmin[2]) & (zs <= cmax[2])
        )
    elif kind == "polygon":
        proj = np.asarray(region["projection"], dtype=np.float64)
        view = np.asarray(region["view"], dtype=np.float64)
        canvas = region["canvas"]
        polygon = np.asarray(region["points"], dtype=np.float64)
        if polygon.ndim != 2 or polygon.shape[1] != 2 or polygon.shape[0] < 3:
            raise HTTPException(
                status_code=400,
                detail="region.points must be at least 3 [x, y] entries.",
            )
        pixels = _project_world_to_pixel(
            positions, proj, view, int(canvas["width"]), int(canvas["height"]),
        )
        mask = _points_in_polygon_mask(pixels, polygon)
    elif kind == "squares_union":
        proj = np.asarray(region["projection"], dtype=np.float64)
        view = np.asarray(region["view"], dtype=np.float64)
        canvas = region["canvas"]
        centers = np.asarray(region["centers"], dtype=np.float64)
        half_sizes = np.asarray(region["half_sizes"], dtype=np.float64)
        pixels = _project_world_to_pixel(
            positions, proj, view, int(canvas["width"]), int(canvas["height"]),
        )
        mask = _squares_union_mask(pixels, centers, half_sizes)
    else:
        raise HTTPException(status_code=400, detail=f"Unknown region.kind: {kind!r}")

    if bool(region.get("invert", False)):
        mask = ~mask
    return mask


def _project_world_to_pixel(
    positions: np.ndarray,
    projection: np.ndarray,
    view: np.ndarray,
    canvas_w: int,
    canvas_h: int,
) -> np.ndarray:
    """Project Nx3 world positions to Nx2 canvas pixels via the renderer's
    frozen camera matrices. Matches `projectWorldToCanvasPixel` in
    src/renderer/lib/cropGeometry.ts so the polygon test produces identical
    in/out membership to the renderer's preview.

    THREE.js stores matrices in column-major order, so the renderer's
    `Matrix4.elements` array applied as `m * v` is `elements.T @ v` in NumPy.
    """
    n = positions.shape[0]
    if n == 0:
        return np.zeros((0, 2), dtype=np.float64)

    # column-major flat → 4x4 row-major: transpose
    P = projection.reshape(4, 4, order="F").astype(np.float64)
    V = view.reshape(4, 4, order="F").astype(np.float64)
    M = P @ V  # 4x4

    # Homogeneous Nx4
    ones = np.ones((n, 1), dtype=np.float64)
    homo = np.concatenate([positions.astype(np.float64), ones], axis=1)  # Nx4
    clip = homo @ M.T  # Nx4 = (4x4 @ 4xN).T -> Nx4

    w = clip[:, 3]
    # Avoid division by zero — these points are behind the camera plane;
    # mark them as outside polygon space.
    safe = np.abs(w) > 1e-12
    ndc = np.zeros((n, 2), dtype=np.float64)
    ndc[safe, 0] = clip[safe, 0] / w[safe]
    ndc[safe, 1] = clip[safe, 1] / w[safe]
    # Points with w<=0 are behind the camera; push them well outside canvas.
    ndc[~safe, 0] = -1e9
    ndc[~safe, 1] = -1e9

    px = (ndc[:, 0] + 1.0) * 0.5 * canvas_w
    # Canvas Y flipped vs NDC.
    py = (1.0 - (ndc[:, 1] + 1.0) * 0.5) * canvas_h
    return np.stack([px, py], axis=1)


def _points_in_polygon_mask(pixels: np.ndarray, polygon: np.ndarray) -> np.ndarray:
    """Vectorised ray-cast point-in-polygon for Nx2 pixels against an Mx2
    polygon. Returns a bool ndarray of length N.

    Matches `pointInPolygon` in src/renderer/lib/cropGeometry.ts (winding
    via crossing count: a point is inside iff the horizontal ray from it
    crosses an odd number of edges)."""
    n = pixels.shape[0]
    m = polygon.shape[0]
    if n == 0 or m < 3:
        return np.zeros(n, dtype=bool)

    inside = np.zeros(n, dtype=bool)
    px = pixels[:, 0]
    py = pixels[:, 1]
    # Edges: (polygon[i], polygon[j]) with j = i-1 wrap-around.
    j = m - 1
    for i in range(m):
        xi, yi = polygon[i, 0], polygon[i, 1]
        xj, yj = polygon[j, 0], polygon[j, 1]
        # Edge straddles the horizontal ray from the point (yi-py and yj-py
        # have opposite signs). Compute intersection x; flip `inside` when
        # the test point is to the left of that intersection.
        cond1 = (yi > py) != (yj > py)
        # Guard against zero-length vertical span (parallel edge): if (yj-yi)
        # is 0, cond1 is False for all points anyway.
        denom = (yj - yi)
        # Where denom is 0, intersect_x is irrelevant (cond1 is False).
        with np.errstate(divide="ignore", invalid="ignore"):
            intersect_x = np.where(denom != 0, (xj - xi) * (py - yi) / denom + xi, 0.0)
        cond2 = px < intersect_x
        inside = np.where(cond1 & cond2, ~inside, inside)
        j = i
    return inside


class CropOctreeRegion(BaseModel):
    """Box, polygon, or sphere-union spatial region for the cloud-session edit
    endpoints (delete_region / filter / split / extract). See _canonical_region
    for validation rules — the handlers delegate to that helper."""
    kind: str
    # Box fields
    min: Optional[List[float]] = None
    max: Optional[List[float]] = None
    # Polygon fields (screen-space, frozen camera matrices)
    points: Optional[List[List[float]]] = None
    projection: Optional[List[float]] = None
    view: Optional[List[float]] = None
    canvas: Optional[dict] = None
    # Squares-union fields (the erase brush). The brush paints a list of
    # screen-space square stamps under one frozen camera; each point is
    # projected to canvas pixels (via the same projection/view/canvas as the
    # polygon path) and is "inside" the region if it falls within ANY square —
    # a depth-independent stamp that extrudes through the whole cloud, exactly
    # like the polygon crop. The erase tool sends invert=True to keep the
    # complement (every point outside all painted squares). `centers` are
    # [px, py] pixel positions and `half_sizes` are the squares' half-extents
    # in pixels (axis-aligned in screen space).
    centers: Optional[List[List[float]]] = None
    half_sizes: Optional[List[float]] = None
    invert: bool = False


class ScalarFilter(BaseModel):
    """Keep only points whose imported scalar attribute `slug` matches.

    Two modes:
      - Continuous (default): keep points in the inclusive range [min, max].
      - Categorical: when `values` is set, keep points whose value rounds to one
        of the listed class ids (an OR within the field — `min`/`max` ignored).
        Used by the filter UI's class-checkbox path for integer-valued labels
        like `ground_class` / `tree_instance`, where a value such as `2` means a
        discrete class, not a position on a continuum, and a multi-select keep
        need not be contiguous.

    `slug` is the on-disk extra-dimension name (matches a key in the octree's
    `attributeRanges` / the `extra_dims` slugs produced by `_xyz_column_plan`).
    """
    slug: str
    # Optional so a categorical (`values`) filter can omit them. Defaulted to the
    # widest range so a malformed continuous filter keeps everything rather than
    # silently dropping points.
    min: float = float("-inf")
    max: float = float("inf")
    # When non-empty, switches this filter to categorical membership: keep iff
    # round(value) ∈ {round(v) for v in values}.
    values: Optional[List[float]] = None


# ===================== MUTABLE CLOUD SESSIONS (Family-1) =====================
#
# A cloud session holds an imported point cloud's positions in RAM as the
# mutable source of truth (the "Family-1" in-core model). Deletions accumulate
# as an exact per-point boolean mask plus the ordered list of delete regions
# that produced it. The Potree octree becomes a *derived* on-disk cache: the
# first one is built at session create, and a fresh one is rebuilt only on an
# explicit "bake" (the one deliberately-slow step). Downstream compute ops read
# `positions[~deleted]` straight from the session array instead of re-reading
# the source file — see `_read_points_from_source`.
#
# Why store regions (not just the mask) for bake: the regions are the auditable
# record of what was removed. The survivor LAS is written straight from the
# in-RAM arrays by `_session_to_las` (colours + scalar extra-dims intact, no
# second full copy held in RAM). The in-RAM `deleted` mask and the region replay
# are kept in lock-step (both go through the shared `_region_mask`) so the array
# the compute ops see and the baked octree always agree.

_cloud_sessions: Dict[str, "CloudSession"] = {}
_cloud_session_lock = threading.Lock()


# ---- In-RAM session eviction ------------------------------------------------
# Cloud and plant sessions each hold the full source-of-truth in RAM (a cloud
# session is ~30-60 bytes/point; a plant session pins a live pyhelios context).
# They are reclaimed only by an explicit DELETE, so a renderer reload/crash that
# never issues one leaks them until the backend dies. We bound that lazily (no
# background thread): every session create/read/mutate bumps `last_accessed` and
# opportunistically sweeps idle + over-cap sessions, mirroring the LRU-by-time
# policy of `_evict_octree_cache`. All three limits are env-overridable.
try:
    _SESSION_IDLE_TTL_SECONDS = float(
        os.environ.get("PHYTOGRAPH_SESSION_IDLE_TTL_SECONDS", str(30 * 60))
    )
except (ValueError, TypeError):
    _SESSION_IDLE_TTL_SECONDS = 30 * 60.0
try:
    _MAX_CLOUD_SESSIONS = int(os.environ.get("PHYTOGRAPH_MAX_CLOUD_SESSIONS", "8"))
except (ValueError, TypeError):
    _MAX_CLOUD_SESSIONS = 8
try:
    _MAX_PLANT_SESSIONS = int(os.environ.get("PHYTOGRAPH_MAX_PLANT_SESSIONS", "8"))
except (ValueError, TypeError):
    _MAX_PLANT_SESSIONS = 8
try:
    _MAX_DELETED_HISTORY = int(os.environ.get("PHYTOGRAPH_MAX_DELETED_HISTORY", "50"))
except (ValueError, TypeError):
    _MAX_DELETED_HISTORY = 50


def _evict_session_ids(sessions: Dict[str, Any], max_count: int, now: float) -> List[str]:
    """Return the ids to evict from `sessions` to satisfy the idle-TTL and the
    count cap, oldest `last_accessed` first. Pure (no mutation) so the caller can
    do teardown under whatever lock it already holds. Sessions are evicted when
    untouched for longer than the TTL, then — if still over `max_count` — the
    least-recently-accessed survivors are dropped until at the cap."""
    if not sessions:
        return []
    # (last_accessed, id), oldest first. Fall back to created_at if unset.
    ranked = sorted(
        sessions.items(),
        key=lambda kv: getattr(kv[1], "last_accessed", 0.0) or getattr(kv[1], "created_at", 0.0),
    )
    evict: List[str] = []
    ttl = _SESSION_IDLE_TTL_SECONDS
    for sid, sess in ranked:
        last = getattr(sess, "last_accessed", 0.0) or getattr(sess, "created_at", 0.0)
        if ttl > 0 and (now - last) > ttl:
            evict.append(sid)
    survivors = [sid for sid, _ in ranked if sid not in evict]
    overflow = len(survivors) - max(0, max_count)
    if overflow > 0:
        evict.extend(survivors[:overflow])  # oldest survivors first
    return evict


def _sweep_cloud_sessions() -> None:
    """Lazily drop idle / over-cap cloud sessions (frees their RAM arrays)."""
    now = time.time()
    with _cloud_session_lock:
        for sid in _evict_session_ids(_cloud_sessions, _MAX_CLOUD_SESSIONS, now):
            sess = _cloud_sessions.pop(sid, None)
            if sess is not None:
                print(f"[Cloud Session] Evicted idle/over-cap session {sid}")


def _sweep_plant_sessions() -> None:
    """Lazily drop idle / over-cap plant sessions, tearing down pyhelios."""
    now = time.time()
    with _session_lock:
        for sid in _evict_session_ids(_plant_sessions, _MAX_PLANT_SESSIONS, now):
            sess = _plant_sessions.pop(sid, None)
            if sess is None:
                continue
            # Same teardown as the DELETE endpoint so native resources are freed.
            try:
                sess.plantarch.__exit__(None, None, None)
                sess.context.__exit__(None, None, None)
            except Exception:
                pass
            print(f"[Plant Session] Evicted idle/over-cap session {sid}")


@dataclass
class CloudSession:
    """An imported point cloud held in RAM as the COMPLETE source of truth.

    The full attribute set — positions, colours, intensity, and every scalar
    extra-dimension — lives in these arrays. The source FILE is read exactly
    once (at create); after that every operation (delete/crop/erase, filter,
    ground/tree segment, bake, downstream compute) reads or mutates these arrays
    and never touches the file again. The Potree octree is a derived cache
    rebuilt from the arrays on bake.
    """
    session_id: str
    source_path: str                 # provenance only — never re-read after create
    ascii_format: Optional[str]
    column_plan: Optional[Any]       # ColumnPlan | None (wizard layout, honored once)
    positions: np.ndarray            # (N,3) float64 — FULL source resolution.
    # For ASCII/XYZ imports this comes straight from the source via
    # `_xyz_to_las(capture_full_xyz=True)`, NOT from reading back the 1 mm-scale
    # octree LAS — that quantization shatters precision-sensitive ops like Helios
    # triangulation (it projects to spherical angles from the scan origin, where
    # 1 mm jitter blows up edge lengths). LAS/LAZ inputs keep their own header
    # precision; PLY/PCD/E57 read losslessly into the LAS already.
    colors: Optional[np.ndarray]     # (N,3) uint16 (0-65535 LAS scale) | None
    intensity: Optional[np.ndarray]  # (N,) uint16 | None
    extras: Dict[str, np.ndarray]    # slug -> (N,) float32 scalar extra-dim columns
    extra_dims_meta: List[dict]      # ordered [{slug, label}] for the octree sidecar
    deleted: np.ndarray              # (N,) bool — True == deleted (hidden)
    deleted_history: List[np.ndarray]  # mask snapshots, one per committed delete (undo)
    octree_cache_id: Optional[str]   # currently-built derived octree, or None if stale
    created_at: float
    last_accessed: float = 0.0       # bumped on every read/mutate; drives idle eviction
    # CloudCompare-style global shift applied at import: the per-axis offset that
    # was SUBTRACTED from `positions` at create, so `world = stored + world_shift`.
    # None means the session holds true world coordinates (no shift). Stored here
    # so `_read_points_from_source` can restore world coords for every downstream
    # op (triangulate/skeleton/LAD/export) at the single read chokepoint, keeping
    # the in-RAM array (small, precision-friendly) the source of truth. Defaulted
    # so direct constructions that predate this field (tests, internal helpers)
    # need no change — a session with no shift behaves exactly as before.
    world_shift: Optional[np.ndarray] = None  # (3,) float64 | None
    # Explicitly recovered sky/miss points (see POST .../backfill-misses). A
    # SEPARATE lightweight buffer rather than rows interleaved into `positions`/
    # `extras`, so a sparse scan that is 30-50% sky doesn't pad every scalar
    # column with millions of miss rows. Populated only when the scan had NO real
    # misses but carried the data to reconstruct them (timestamp and/or row/column
    # grid). None until Backfill Misses runs. Stored in the SAME frame as
    # `positions` (no world-shift applied here — _session_to_lad_arrays keeps that
    # frame too). Keys:
    #   positions  (M,3) float64 — far-field miss coordinates
    #   directions (M,3) float32 — [radius, elevation, azimuth] for LAD beams
    #   timestamp  (M,)  float64 — when the source carried it (optional)
    #   origins    (M,3) float64 — per-pulse emission points (moving scans only)
    # `is_miss` is implicitly 1 for every row, so no flag array is stored.
    backfilled_misses: Optional[dict] = None
    # Derived projected-miss octree built alongside the hits octree (create / bake
    # / backfill). Streamed via app://octree/<id>/ exactly like the hits octree but
    # rendered flat-orange. None when the scan has no placeable misses, or stale
    # until the next bake. See `_build_miss_octree`. Defaulted so the existing
    # direct CloudSession(...) constructions need no change.
    miss_octree_cache_id: Optional[str] = None
    # Scanner origin (in the session's frame, world-shift already subtracted) used
    # to project the misses onto the display sphere. Captured at create so a later
    # bake can reproject the surviving misses without re-deriving it. None for a
    # scan with no origin (the miss octree is then built at true coordinates).
    miss_octree_origin: Optional[List[float]] = None
    # True when the separate `backfilled_misses` buffer was computed BEFORE a later
    # crop, so it reflects pre-crop geometry and would skew LAD's hit/miss ratio.
    # Set by `delete_region` when a crop touches a cloud that has backfilled misses;
    # cleared when Backfill Misses recomputes them. Surfaced to the user at crop
    # time (toast) and at LAD time (warning). Only the backfilled buffer can go
    # stale — interleaved is_miss points live in `extras` and are subset by bake.
    backfilled_misses_stale: bool = False
    # Per-point GPS/relative time, the moving-platform LAD join key. Kept as a
    # dedicated float64 column (NOT in the float32 `extras` dict) because GPS
    # Adjusted-Standard time is a huge double (~3.5e8 s) and a float32 cast has
    # only ~7 significant digits → ~32 s of resolution, which collapses every
    # return within tens of seconds onto a single trajectory pose. The LAD path
    # reads this via `_session_to_lad_arrays`'s `_get('timestamp')` chokepoint,
    # which prefers this field over any `extras['timestamp']`. None for clouds
    # with no usable time column (plain XYZ, point-format 0/2 LAS). Defaulted so
    # existing direct CloudSession(...) constructions need no change.
    timestamps: Optional[np.ndarray] = None  # (N,) float64 | None
    # How the LAS `gps_time` column is encoded, from global_encoding bit 0:
    #   'adjusted_standard' — Adjusted-Standard GPS seconds (Standard − 1e9), an
    #                         ABSOLUTE clock reconcilable with a survey trajectory.
    #   'gps_week'          — GPS Week Time (seconds-into-week, 0–604800): no
    #                         absolute epoch, so it CANNOT align to an absolute
    #                         trajectory clock without an external week reference.
    #                         The moving join refuses it loudly rather than treating
    #                         week-seconds as adjusted-standard.
    #   None                — no usable gps_time (plain XYZ, point-format 0/2 LAS,
    #                         or an ASCII-derived timestamp that carries no encoding).
    gps_time_encoding: Optional[str] = None
    # Per-pulse emission points read from LAS ExtraBytes (ox/oy/oz aliases), in the
    # SAME frame as `positions` (world_shift already subtracted at create). When
    # present these are GROUND TRUTH origins and the LAD path uses them directly,
    # bypassing the timestamp -> trajectory join (see `_do_lad_computation`). None
    # when the LAS carried no origin triple. Defaulted so existing direct
    # CloudSession(...) constructions need no change.
    beam_origins: Optional[np.ndarray] = None  # (N,3) float64 | None
    # EPSG code of the source file's coordinate reference system, parsed from the
    # LAS/LAZ CRS VLRs at create (None for ASCII/PLY or a file with no CRS). Used
    # to georeference DEM raster exports (GeoTIFF). Defaulted so existing direct
    # CloudSession(...) constructions need no change.
    crs_epsg: Optional[int] = None


def _epsg_from_wkt_vlr(header) -> Optional[int]:
    """Pull the top-level EPSG code straight from a LAS WKT CRS VLR's raw string.

    pyproj's `to_epsg()` refuses a WKT whose body differs from the canonical EPSG
    definition — e.g. a `TOWGS84[0,…]` clause, common in real survey LAS files —
    even when the WKT's own top-level `AUTHORITY` clearly names the code. WKT1's
    top-level object closes with its own `AUTHORITY["EPSG","<code>"]` (so the LAST
    authority in the string is the CRS code, after any child datum/unit codes);
    WKT2 uses `ID["EPSG",<code>]`. Each candidate is validated as a real CRS so a
    stray unit/datum/ellipsoid authority can't be mistaken for the CRS."""
    import re
    wkt = None
    for v in getattr(header, "vlrs", []):
        s = getattr(v, "string", None)
        if s and any(tok in s for tok in ("PROJCS", "GEOGCS", "PROJCRS", "GEOGCRS")):
            wkt = s
            break
    if not wkt:
        return None
    codes = [int(a or b) for a, b in
             re.findall(r'(?:AUTHORITY\[\s*"EPSG"\s*,\s*"?(\d+)"?\]|ID\[\s*"EPSG"\s*,\s*(\d+))', wkt)]
    try:
        import pyproj
    except Exception:
        return None
    for code in reversed(codes):   # top-level CRS authority is last in WKT1
        try:
            crs = pyproj.CRS.from_epsg(code)
            if crs.is_projected or crs.is_geographic:
                return code
        except Exception:
            continue
    return None


def _read_las_crs_epsg(path: "_Path") -> Optional[int]:
    """Best-effort EPSG code from a LAS/LAZ file's CRS VLRs. Tries pyproj's clean
    database match first, then falls back to the EPSG code embedded in the file's
    own WKT CRS VLR (real survey files often carry a CRS pyproj won't match to an
    EPSG, but whose WKT names the code). Returns None for a non-LAS source or no
    recoverable CRS. Reads only the header (no point data)."""
    if path.suffix.lower() not in (".las", ".laz"):
        return None
    try:
        import laspy
        with laspy.open(str(path)) as reader:
            header = reader.header
        crs = header.parse_crs()
        if crs is not None:
            epsg = crs.to_epsg()
            if epsg is not None:
                return int(epsg)
        return _epsg_from_wkt_vlr(header)
    except Exception:
        return None


def _get_cloud_session(session_id: str) -> "CloudSession":
    with _cloud_session_lock:
        sess = _cloud_sessions.get(session_id)
        if sess is not None:
            sess.last_accessed = time.time()
    if sess is None:
        raise HTTPException(status_code=404, detail=f"Cloud session not found: {session_id}")
    return sess


# `_BEAM_ORIGIN_ALIAS_SETS` (the ExtraBytes name aliases for a per-pulse
# beam-origin triple, used by `_read_las_into_arrays` below) is defined earlier,
# next to `_normalise_origin_alias`, so the ASCII import path can share it.


@dataclass
class LasReadResult:
    """Everything `_read_las_into_arrays` materialises from a normalised LAS.

    A dataclass (not a tuple) because the set has grown past readable positional
    unpacking and carries several Optionals with subtle precision contracts.
      positions        (N,3) float64 — full-resolution coordinates.
      colors           (N,3) uint16 | None — kept in LAS scale for byte round-trip.
      intensity        (N,)  uint16 | None.
      extras           {slug: (N,) float32} — scalar extra-dim columns.
      extra_dims_meta  ordered [{slug, label}] for the octree sidecar.
      timestamps       (N,) float64 | None — gps_time, kept OUT of float32 extras
                       (the LAD trajectory-join key; see CloudSession.timestamps).
      gps_time_encoding 'adjusted_standard' | 'gps_week' | None.
      beam_origins     (N,3) float64 | None — per-pulse emission points read from
                       ExtraBytes (ox/oy/oz aliases); when present LAD uses them
                       directly and skips the trajectory join.
    """
    positions: np.ndarray
    colors: Optional[np.ndarray]
    intensity: Optional[np.ndarray]
    extras: Dict[str, np.ndarray]
    extra_dims_meta: List[dict]
    timestamps: Optional[np.ndarray] = None
    gps_time_encoding: Optional[str] = None
    beam_origins: Optional[np.ndarray] = None


def _read_las_into_arrays(las_path: _Path) -> "LasReadResult":
    """Read a LAS file fully into RAM as the session's source-of-truth arrays.

    Returns a `LasReadResult`. RGB/intensity are kept in LAS uint16 scale so the
    bake writer can round-trip them byte-for-byte; extras are the float32 extra-
    dimension columns. `gps_time` is returned SEPARATELY as float64 `timestamps`
    (not in `extras`) — it is the moving-platform LAD join key and a float32 cast
    would destroy its precision (see CloudSession.timestamps). Per-pulse beam-origin
    ExtraBytes are read as float64 `beam_origins` and likewise kept out of `extras`.
    This is the ONE point where a normalised LAS is materialised into the session —
    used by create after `_source_to_las` converts whatever the source format was.
    """
    import laspy
    with laspy.open(str(las_path)) as reader:
        las = reader.read()
    positions = np.stack([np.asarray(las.x), np.asarray(las.y), np.asarray(las.z)], axis=1).astype(np.float64)
    dim_names = set(las.point_format.dimension_names)
    colors = None
    if {"red", "green", "blue"} <= dim_names:
        colors = np.stack([
            np.asarray(las.red, dtype=np.uint16),
            np.asarray(las.green, dtype=np.uint16),
            np.asarray(las.blue, dtype=np.uint16),
        ], axis=1)
    intensity = None
    if "intensity" in dim_names and np.any(np.asarray(las.intensity)):
        intensity = np.asarray(las.intensity, dtype=np.uint16)

    # Detect per-pulse beam-origin ExtraBytes FIRST so their three columns are
    # skipped by the generic extra-dim loop below (they must not also become lossy
    # float32 display scalars). Read as float64 — these are world/UTM coordinates.
    extra_by_lower = {d.name.lower(): d.name for d in las.point_format.extra_dimensions}
    beam_origins: Optional[np.ndarray] = None
    _origin_skip: set = set()
    for ax, ay, az in _BEAM_ORIGIN_ALIAS_SETS:
        if ax in extra_by_lower and ay in extra_by_lower and az in extra_by_lower:
            nx, ny, nz = extra_by_lower[ax], extra_by_lower[ay], extra_by_lower[az]
            beam_origins = np.stack([
                np.asarray(las[nx], dtype=np.float64),
                np.asarray(las[ny], dtype=np.float64),
                np.asarray(las[nz], dtype=np.float64),
            ], axis=1)
            _origin_skip = {nx, ny, nz}
            break

    extras: Dict[str, np.ndarray] = {}
    extra_dims_meta: List[dict] = []
    for d in las.point_format.extra_dimensions:
        name = d.name
        if name in _origin_skip:
            continue  # carried as float64 beam_origins, not a float32 scalar
        extras[name] = np.asarray(las[name], dtype=np.float32)
        extra_dims_meta.append({"slug": name, "label": name})

    # Auto-map LAS native per-pulse multi-return dimensions to the canonical
    # slugs Helios's LAD path reads (see `_MULTI_RETURN_SLUGS`). return_number
    # is carried verbatim as target_index — Helios auto-detects 0- vs 1-based
    # indexing, so no base conversion. Only mapped when the source dim is
    # present and a same-named extra-dim wasn't already captured above.
    #
    # CRUCIALLY, only map a dim that actually carries data. return_number,
    # number_of_returns and gps_time are STANDARD LAS dimensions present in every
    # point-format-3 record, so they exist (all-zero) even on a plain XYZ/E57
    # import that never had per-pulse data. Mapping those zeros would make the LAD
    # path see phantom multi-return columns — flipping it to the full-waveform
    # algorithm and running gapfillMisses() on garbage timestamps. Require a
    # non-degenerate column (some non-zero value, and for indices not all-equal)
    # so only genuinely populated airborne-style LAS triggers multi-return.
    _las_multireturn = (
        ("return_number", "target_index"),
        ("number_of_returns", "target_count"),
    )
    for src_dim, slug in _las_multireturn:
        if src_dim in dim_names and slug not in extras:
            vals = np.asarray(las[src_dim])
            if vals.size and np.any(vals != vals.flat[0]):  # not constant/all-zero
                extras[slug] = vals.astype(np.float32)
                extra_dims_meta.append({"slug": slug, "label": _MULTI_RETURN_LABELS[slug]})

    # gps_time is the LAD trajectory-join key — route it to a dedicated float64
    # array, NOT the float32 `extras` dict (a float32 cast at adjusted-standard
    # magnitude has ~32 s resolution and collapses every return onto one pose).
    # Same non-degenerate guard as the multi-return dims: gps_time is a STANDARD
    # dimension present (all-zero) on plain XYZ/E57 imports, so only carry it when
    # it actually varies. The non-constant check runs on the original float64
    # values, before any cast.
    timestamps: Optional[np.ndarray] = None
    gps_time_encoding: Optional[str] = None
    if "gps_time" in dim_names:
        _gps = np.asarray(las["gps_time"], dtype=np.float64)
        if _gps.size and np.any(_gps != _gps.flat[0]):  # not constant/all-zero
            timestamps = _gps
            # global_encoding bit 0 (laspy GpsTimeType): 1 = Adjusted-Standard GPS
            # seconds (absolute clock), 0 = GPS Week Time (seconds-into-week, no
            # epoch). Recorded so the moving-platform join can compare like-for-like
            # and refuse a week clock it cannot align. Keep the RAW values either
            # way — week-time is a different clock, not a subtractable offset.
            try:
                gps_time_encoding = (
                    'adjusted_standard'
                    if bool(las.header.global_encoding.gps_time_type)
                    else 'gps_week'
                )
            except Exception:
                # Older/odd headers: fall back to the raw bit.
                gps_time_encoding = (
                    'adjusted_standard'
                    if (int(las.header.global_encoding.value) & 1)
                    else 'gps_week'
                )

    # Carry the remaining STANDARD LAS dimensions (classification, scan_angle,
    # point_source_id, user_data, scanner_channel, …) as user-selectable scalar
    # fields. Without this they'd be dropped here — the octree is rebuilt from
    # these session arrays (`_session_to_las` → PotreeConverter), NOT from the
    # original file, so anything not in `extras` never reaches the scalar picker.
    # Skip the dims handled elsewhere (positions/colors/intensity, the multi-
    # return sources mapped above, bit-flags) and any all-constant column (an
    # all-zero classification etc. is noise in the picker, not signal).
    #
    # CRUCIAL: the slug is PREFIXED ('las_classification', …), not the bare LAS
    # name. `_session_to_las` re-adds every extra-dim slug to the rebuilt LAS via
    # add_extra_dim — and a slug that collides with a reserved standard-dim name
    # (e.g. 'classification') makes laspy try to bit-pack the float column into
    # the classification-flags byte and hard-crash the process. The 'las_' prefix
    # keeps the slug clear of the standard schema; the label stays the clean name.
    for d in las.point_format.standard_dimensions:
        name = d.name
        if (name in _LAS_STD_DIMS_SKIP or name in _LAS_MULTIRETURN_SRC):
            continue
        slug = f"las_{name}"
        if slug in extras or name in extras:
            continue
        vals = np.asarray(las[name])
        if vals.size and np.any(vals != vals.flat[0]):  # not constant
            extras[slug] = vals.astype(np.float32)
            extra_dims_meta.append({"slug": slug, "label": name})
    return LasReadResult(
        positions=positions,
        colors=colors,
        intensity=intensity,
        extras=extras,
        extra_dims_meta=extra_dims_meta,
        timestamps=timestamps,
        gps_time_encoding=gps_time_encoding,
        beam_origins=beam_origins,
    )


def _trajectory_wire_from_beam_origins(beam_origins, timestamps, max_poses=3000):
    """Reconstruct a decimated platform-trajectory wire dict from per-pulse beam
    origins + their timestamps (e.g. a LAS carrying ox/oy/oz ExtraBytes).

    The origins ARE the platform path sampled once per pulse — far too dense to use
    as poses directly (millions of near-duplicates). Sort by time, even-stride down
    to <= `max_poses` (always keeping first + last so the span is preserved), and
    emit one identity-attitude pose per kept sample. Attitude is unknown from
    positions alone, so it is identity — the EXACT per-pulse origins are still used
    for the LAD inversion (this trajectory is for display + as the moving-scan flag,
    not for re-deriving origins). Returns the canonical PoseStream wire dict, or None
    when there is no usable origin/time data. Drops non-finite and non-monotonic
    samples so the resulting pose times strictly increase.
    """
    import numpy as np

    if beam_origins is None or timestamps is None:
        return None
    bo = np.asarray(beam_origins, dtype=np.float64)
    ts = np.asarray(timestamps, dtype=np.float64)
    if bo.ndim != 2 or bo.shape[0] != ts.shape[0] or bo.shape[0] == 0:
        return None

    finite = np.isfinite(ts) & np.all(np.isfinite(bo), axis=1)
    bo, ts = bo[finite], ts[finite]
    if ts.shape[0] == 0:
        return None

    order = np.argsort(ts, kind="stable")
    ts, bo = ts[order], bo[order]
    # Keep strictly-increasing times (collapse exact-duplicate timestamps, keeping
    # the first), so SLERP/interp on the renderer + backend never sees a flat step.
    keep_t = np.concatenate(([True], np.diff(ts) > 0))
    ts, bo = ts[keep_t], bo[keep_t]
    n = ts.shape[0]
    if n == 0:
        return None

    if n > max_poses:
        idx = np.arange(0, n, int(np.ceil(n / max_poses)))
        if idx[-1] != n - 1:
            idx = np.append(idx, n - 1)
    else:
        idx = np.arange(n)

    poses = [
        {"t": float(ts[i]),
         "x": float(bo[i, 0]), "y": float(bo[i, 1]), "z": float(bo[i, 2]),
         "qx": 0.0, "qy": 0.0, "qz": 0.0, "qw": 1.0}
        for i in idx
    ]
    if len(poses) < 2:
        return None  # a single pose is a static scan, not a trajectory
    return {
        "poses": poses,
        "frame": {"crs": None, "up_axis": "z", "body_convention": "FLU", "time_ref": "gps"},
        "lever_arm": [0.0, 0.0, 0.0],
        "boresight_rpy": [0.0, 0.0, 0.0],
        "source_format": "las_extrabytes",
    }


# Row block size for the in-RAM → LAS writers. Caps the transient laspy record to
# one block instead of one record for all N survivors (see `_session_to_las`).
# 2M matches `_xyz_to_las`'s streaming chunk.
_LAS_WRITE_CHUNK = 2_000_000


def _session_to_las(sess: "CloudSession", out_las: _Path,
                    exclude_misses: bool = False) -> int:
    """Write the session's SURVIVING points (positions[~deleted] + all
    attributes) to a LAS, entirely from the in-RAM arrays — no source file read.
    Mirrors `_xyz_to_las`'s header/record layout so PotreeConverter ingests it
    identically. Returns the survivor count.

    `exclude_misses=True` additionally drops sky/miss points (is_miss != 0).
    Misses are real points placed ~20 km away, so leaving them in poisons the
    octree's bounding box (and thus camera framing). The octree is built
    hits-only; misses live in the session for LAD + the on-demand overlay. A bake
    (which round-trips the LAS back into the session) must NOT exclude them, or
    the misses would be lost — only the octree build passes True.
    """
    import laspy
    keep = ~sess.deleted
    if exclude_misses and _MISS_SLUG in sess.extras:
        keep = keep & (sess.extras[_MISS_SLUG] == 0)
    n = int(keep.sum())

    pos = sess.positions[keep]
    header = laspy.LasHeader(point_format=3, version="1.4")
    header.scales = np.array([0.001, 0.001, 0.001], dtype=np.float64)
    # Offset to the data minimum, NOT [0,0,0]: a LAS coordinate is stored as a
    # 32-bit int (value - offset) / scale, so with offset 0 and 1 mm scale the
    # representable range is only ±2.1 km. Real-world projected clouds (UTM
    # northings ~5.4e6 m) overflow that. Subtracting the per-axis minimum keeps
    # the stored ints small for any CRS; laspy re-applies offset on read, so
    # downstream coordinates are unchanged.
    header.offsets = (np.floor(pos.min(axis=0)) if n else np.zeros(3, dtype=np.float64))
    for ed in sess.extra_dims_meta:
        header.add_extra_dim(laspy.ExtraBytesParams(name=ed["slug"], type=np.float32))

    # bbox over the full selection (cheap — just the f64 xyz), computed before the
    # chunked write so the header mins/maxs cover every chunk.
    if n > 0:
        pos_min = pos.min(axis=0)
        pos_max = pos.max(axis=0)

    # Write the LAS in row chunks rather than materialising ONE laspy record for
    # all N survivors. `ScaleAwarePointRecord.zeros(n, header)` allocates a full
    # structured array (point-format-3 + every float32 extra dim ≈ tens of bytes
    # per point); on a multi-million-point synthetic scan that's a multi-GB
    # transient stacked on top of the session arrays. Chunking caps the record to
    # one block (`_LAS_WRITE_CHUNK` rows) at a time — same bytes on disk, a
    # fraction of the peak RAM. Mirrors `_xyz_to_las_stream`'s chunked writer.
    idx = np.flatnonzero(keep)
    with laspy.open(str(out_las), mode="w", header=header) as writer:
        for start in range(0, n, _LAS_WRITE_CHUNK):
            block = idx[start:start + _LAS_WRITE_CHUNK]
            m = block.shape[0]
            record = laspy.ScaleAwarePointRecord.zeros(m, header=header)
            bpos = sess.positions[block]
            record.x = bpos[:, 0]
            record.y = bpos[:, 1]
            record.z = bpos[:, 2]
            if sess.colors is not None:
                c = sess.colors[block]
                record.red, record.green, record.blue = c[:, 0], c[:, 1], c[:, 2]
            if sess.intensity is not None:
                record.intensity = sess.intensity[block]
            for ed in sess.extra_dims_meta:
                record[ed["slug"]] = sess.extras[ed["slug"]][block]
            writer.write_points(record)
            del record, bpos
        if n > 0:
            pad = 0.001  # matches header.scales — keeps boundary points in-bbox
            writer.header.mins = (pos_min - pad).tolist()
            writer.header.maxs = (pos_max + pad).tolist()
    return n


def _gather_miss_positions(sess: "CloudSession",
                           origin: Optional[List[float]]) -> tuple[np.ndarray, float]:
    """Surviving sky/miss positions for the miss octree, in display form.

    Gathers the session's interleaved misses (`extras[is_miss] != 0`, honoring
    `~deleted`) UNION the separate `backfilled_misses` buffer — the same union
    the old miss overlay drew. Then:

    - `origin` is None → return the positions verbatim (true far-field coords),
      radius 0. The miss octree is built at true coordinates; its own bbox keeps
      the ~20 km extent from ever touching the hits octree's framing.
    - `origin` supplied → project each placeable miss onto a sphere centred on
      the origin at `radius = far + max(0.05*depth, 0.05*far, 0.05)` (far/near are
      the max/min HIT distance from the origin, depth = far − near), a THIN halo
      hugging the cloud. Misses sitting AT the origin (no beam direction yet) are
      dropped.

    This is the projection math factored out of the former `get_cloud_misses`
    endpoint, MINUS the _MISS_OVERLAY_CAP stride — the octree streams via LOD, so
    there is no cap and no aliasing. Returns (positions (M,3) float64, radius).
    """
    with _cloud_session_lock:
        keep = ~sess.deleted
        miss_arr = sess.extras.get(_MISS_SLUG)
        if miss_arr is not None:
            is_miss = (miss_arr != 0) & keep
            hits = (~(miss_arr != 0)) & keep
            miss_pos = np.ascontiguousarray(sess.positions[is_miss], dtype=np.float64)
        else:
            hits = keep.copy()
            miss_pos = np.empty((0, 3), np.float64)
        hit_pos = np.ascontiguousarray(sess.positions[hits], dtype=np.float64)
        backfilled = sess.backfilled_misses
        if backfilled is not None and backfilled.get("positions") is not None:
            bf = np.ascontiguousarray(backfilled["positions"], dtype=np.float64)
            if bf.shape[0] > 0:
                miss_pos = np.vstack([miss_pos, bf]) if miss_pos.shape[0] else bf

    if miss_pos.shape[0] == 0:
        return np.empty((0, 3), np.float64), 0.0

    if origin is None:
        return miss_pos, 0.0

    origin_arr = np.asarray(origin, dtype=np.float64)
    if hit_pos.shape[0] > 0:
        hit_dists = np.linalg.norm(hit_pos - origin_arr, axis=1)
        far = float(np.max(hit_dists))
        near = float(np.min(hit_dists))
        # The radius clears the FARTHEST hit by only a SMALL, bounded margin so the
        # miss shell reads as a thin halo HUGGING the cloud — not a band parked far
        # outside it. The far*1.4 shell was the bug: the miss octree is LOD-streamed
        # and frustum-culled like any cloud, so a shell beyond the camera's framing
        # of the hits renders nothing. The margin scales with the cloud's own radial
        # depth (far − near) so it adapts to scene size, with a small multiplicative
        # floor (5% of far) and a small absolute floor for near-zero-depth clouds.
        depth = max(far - near, 0.0)
        radius = far + max(0.05 * depth, 0.05 * far, 0.05)
    else:
        radius = 1.0
    if radius <= 0:
        radius = 1.0

    d = miss_pos - origin_arr
    n = np.linalg.norm(d, axis=1)
    drawable = n > 1e-9  # drop misses at the origin (no beam direction yet)
    d = d[drawable]
    n = n[drawable]
    relocated = origin_arr + (d / n[:, None]) * radius
    return np.ascontiguousarray(relocated, dtype=np.float64), float(radius)


def _miss_positions_to_las(positions: np.ndarray, out_las: _Path) -> int:
    """Write position-only miss points to a LAS for PotreeConverter. The miss
    octree renders flat-orange, so it carries no colours/intensity/extra-dims.
    Mirrors `_session_to_las`'s header (pt fmt 3, 1 mm scale, offset = floor(min)
    so far-field coords don't overflow the LAS int32). Returns the count."""
    import laspy
    n = int(positions.shape[0])
    header = laspy.LasHeader(point_format=3, version="1.4")
    header.scales = np.array([0.001, 0.001, 0.001], dtype=np.float64)
    header.offsets = (np.floor(positions.min(axis=0)) if n else np.zeros(3, dtype=np.float64))
    # Chunk the write so the laspy record is one block, not all M misses at once —
    # a sparse scan can be 30-50% sky, so the miss set is itself multi-million-row.
    # See `_session_to_las` for the rationale.
    with laspy.open(str(out_las), mode="w", header=header) as writer:
        for start in range(0, n, _LAS_WRITE_CHUNK):
            block = positions[start:start + _LAS_WRITE_CHUNK]
            record = laspy.ScaleAwarePointRecord.zeros(block.shape[0], header=header)
            record.x = block[:, 0]
            record.y = block[:, 1]
            record.z = block[:, 2]
            writer.write_points(record)
            del record
        if n > 0:
            pad = 0.001
            writer.header.mins = (positions.min(axis=0) - pad).tolist()
            writer.header.maxs = (positions.max(axis=0) + pad).tolist()
    return n


def _build_miss_octree(sess: "CloudSession",
                       origin: Optional[List[float]]) -> Optional[str]:
    """Build a projected-miss octree for the session and return its cache id, or
    None when the session has no placeable misses. Failures are logged and
    swallowed (return None) — a miss-octree build must never abort the caller
    (session create, bake, or backfill); the hits octree and LAD are the
    priority. Eager but cheap: when there are no misses, no PotreeConverter runs."""
    import tempfile
    try:
        positions, _radius = _gather_miss_positions(sess, origin)
        if positions.shape[0] == 0:
            return None
        with tempfile.TemporaryDirectory() as td:
            miss_las = _Path(td) / "octree_misses.las"
            _miss_positions_to_las(positions, miss_las)
            cache_key, _cache_dir, _meta = _build_octree_from_las(miss_las, [])
        return cache_key
    except Exception:
        logger.exception("Miss octree build failed for session %s", sess.session_id)
        return None


def _build_octree_from_las(las_path: _Path, extra_dims_meta: List[dict]) -> tuple[str, _Path, dict]:
    """Run PotreeConverter on a LAS and atomically install it into the octree
    cache keyed by the LAS bytes' hash. Returns (cache_key, cache_dir, meta).
    Shared by create (initial) and bake (post-edit) — both now feed a LAS that
    was produced WITHOUT re-reading the source file at edit time."""
    h = _hashlib.sha1()
    with open(las_path, "rb") as f:
        for block in iter(lambda: f.read(1 << 20), b""):
            h.update(block)
    cache_key = h.hexdigest()
    cache_dir = _octree_cache_root() / cache_key

    # Serialize builds of THIS key so a batch import of identical-hash scans
    # can't trample a shared staging dir (see _octree_build_lock). The cache is
    # re-checked inside the lock so the first builder wins and the rest no-op.
    with _octree_build_lock(cache_key):
        if not (cache_dir / "metadata.json").is_file():
            cache_dir.parent.mkdir(parents=True, exist_ok=True)
            staging_dir = cache_dir.parent / (cache_key + ".staging")
            if staging_dir.exists():
                _shutil.rmtree(staging_dir)
            staging_dir.mkdir(parents=True)
            try:
                _run_potree_converter(las_path, staging_dir)
                _write_octree_labels(staging_dir, extra_dims_meta)
                if cache_dir.exists():
                    _shutil.rmtree(cache_dir)
                staging_dir.rename(cache_dir)
            except Exception:
                try:
                    _shutil.rmtree(staging_dir)
                except (FileNotFoundError, OSError):
                    pass
                raise

    meta = _read_octree_metadata(cache_dir)
    return cache_key, cache_dir, meta


class CloudSessionCreateRequest(BaseModel):
    """Create a mutable cloud session from a source file and build its first
    (derived) octree. `column_plan` is the import wizard's explicit layout and
    is honored ONCE here, then carried for the life of the session so edits
    never re-auto-detect columns (the import-wizard option-loss fix)."""
    source_path: str
    ascii_format: Optional[str] = None
    column_plan: Optional[ColumnPlan] = None
    # CloudCompare-style global shift chosen in the import wizard: [x, y, z]
    # SUBTRACTED from every point at create so the in-RAM array holds small,
    # precision-friendly coordinates. The shift is stored on the session and added
    # back at the read chokepoint, so downstream ops/exports recover true world
    # coords. None / omitted = keep the original (possibly large) coordinates.
    world_shift: Optional[List[float]] = None
    # Far-field distance (m) used by miss auto-detection's DISTANCE fallback only
    # — when a scan carries no `is_miss` column AND no `target_index` sentinel, a
    # point this far (>=0.98x) from the scanner origin is treated as a sky/miss.
    # Mirrors Helios's LIDAR_RAYTRACE_MISS_T = 1001 m. None → 1001. The primary
    # signal (target_index == 99) needs neither this nor an origin.
    miss_distance_threshold: Optional[float] = None


class DeleteRegionRequest(BaseModel):
    """Mark points inside `region` as deleted on a cloud session. Instant: sets
    the in-RAM mask; does NOT rebuild the octree. The renderer mirrors the
    deletion on the GPU via its clip-volume stack, so the viewport updates
    immediately."""
    region: CropOctreeRegion


class BackfillMissesRequest(BaseModel):
    """Explicitly recover a session's sky/miss points and persist them (see
    POST .../backfill-misses).

    Mirrors the LAD-relevant subset of HeliosScanEntry so the backfill cloud is
    built on the exact same array + addScan path as `/api/lad/compute`. `origin`
    is the scanner position (per-beam directions are reconstructed from it); the
    optional angular raster (n_theta/n_phi/theta_*/phi_*) sets the scan grid the
    gapfiller reconstructs misses over, falling back to a count-based estimate
    when omitted. `trajectory` marks a moving-platform scan (per-pulse origins
    joined by timestamp), which forces the timestamp gapfill path."""
    origin: List[float]                 # [x, y, z] scanner position
    n_theta: Optional[int] = None
    n_phi: Optional[int] = None
    theta_min: Optional[float] = None   # degrees
    theta_max: Optional[float] = None
    phi_min: Optional[float] = None
    phi_max: Optional[float] = None
    beam_exit_diameter: Optional[float] = None   # meters
    beam_divergence: Optional[float] = None       # milliradians
    trajectory: Optional[PoseStream] = None


@app.post("/api/cloud/session/create")
async def create_cloud_session(request: CloudSessionCreateRequest):
    """Load a source cloud FULLY into an in-RAM session (the complete source of
    truth — positions + colours + intensity + every scalar extra-dim), build its
    first octree, and return `{session_id, ...octree metadata}`. This is the ONLY
    point the source FILE is read; every later edit/bake/op works on the arrays.

    The source is normalised to a LAS once via `_source_to_las` (handling
    XYZ/PLY/PCD/LAS/LAZ + the wizard column_plan uniformly), that LAS is read
    into the session arrays AND fed to PotreeConverter for the first octree."""
    source_path = _Path(request.source_path).expanduser()
    if not source_path.is_file():
        raise HTTPException(status_code=404, detail=f"Source file not found: {request.source_path}")

    import time
    session_id = uuid.uuid4().hex[:8]

    # Normalise to a LAS in a temp dir, read it fully into RAM, then build the
    # octree from that same LAS. After this the file is never touched again.
    import tempfile
    with tempfile.TemporaryDirectory() as _tmp:
        tmp_dir = _Path(_tmp)
        las_path, las_is_temp, source_extra_dims, full_xyz, source_origins = _source_to_las(
            source_path, request.ascii_format, tmp_dir, request.column_plan,
        )
        _las = _read_las_into_arrays(las_path)
        # The session array is the source of truth and must hold FULL precision
        # (it is never re-read from the file). For ASCII/XYZ imports the LAS we
        # synthesised is 1 mm-quantized — coarse enough to shatter precision-
        # sensitive ops like triangulation — so prefer the source-precision xyz
        # `_source_to_las` captured during conversion. Colors/intensity/extras
        # still come from the LAS read (it filters NaN-xyz rows identically, so
        # the arrays stay point-for-point aligned). LAS/LAZ inputs keep their own
        # header precision (full_xyz is None) and need no override.
        if full_xyz is not None:
            if full_xyz.shape[0] != _las.positions.shape[0]:
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "Internal error importing cloud: source-precision point "
                        f"count ({full_xyz.shape[0]}) disagrees with the converted "
                        f"LAS ({_las.positions.shape[0]}). Please report this file."
                    ),
                )
            positions = full_xyz
        else:
            positions = _las.positions
        colors = _las.colors
        intensity = _las.intensity
        extras = _las.extras
        extra_dims_meta = _las.extra_dims_meta
        timestamps = _las.timestamps
        gps_time_encoding = _las.gps_time_encoding
        # Beam origins: for LAS/LAZ they're read from ExtraBytes by
        # `_read_las_into_arrays`. For the ASCII/XYZ path the 1 mm LAS can't carry
        # them (origins are world/UTM coords needing full precision), so prefer the
        # float64 triple `_source_to_las` captured from the source columns. Like
        # `full_xyz`, guard the row count so a malformed plan can't desync them
        # from positions; the LAD path then uses these directly, bypassing the
        # timestamp->trajectory join (see CloudSession.beam_origins).
        beam_origins = _las.beam_origins
        if source_origins is not None:
            if source_origins.shape[0] != _las.positions.shape[0]:
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "Internal error importing cloud: beam-origin point count "
                        f"({source_origins.shape[0]}) disagrees with the converted "
                        f"LAS ({_las.positions.shape[0]}). Please report this file."
                    ),
                )
            beam_origins = source_origins
        # CloudCompare-style global shift: subtract the requested offset so the
        # in-RAM array (the source of truth) — and the octree built from it below —
        # hold small, precision-friendly coordinates. The shift is stored on the
        # session and added back in `_read_points_from_source`, so triangulate /
        # skeleton / LAD / export all see true world coords. A near-zero or absent
        # shift means the cloud keeps its original coordinates.
        world_shift_arr: Optional[np.ndarray] = None
        if request.world_shift is not None:
            ws = np.asarray(request.world_shift, dtype=np.float64)
            if ws.shape != (3,):
                raise HTTPException(
                    status_code=400,
                    detail=f"world_shift must be [x, y, z]; got {request.world_shift!r}",
                )
            if np.any(ws != 0.0):
                positions = positions - ws  # new array; positions is float64 here
                # beam_origins are in the SAME world frame as positions — shift them
                # in lockstep so per-beam origins stay consistent with the points.
                if beam_origins is not None:
                    beam_origins = beam_origins - ws
                world_shift_arr = ws
        # `_read_las_into_arrays` sets label==slug from the LAS header, which
        # loses the wizard's custom labels. `_source_to_las` returns the proper
        # [{slug, label}] (honoring the column_plan), so overlay those labels.
        _label_by_slug = {ed["slug"]: ed.get("label", ed["slug"]) for ed in (source_extra_dims or [])}
        for ed in extra_dims_meta:
            ed["label"] = _label_by_slug.get(ed["slug"], ed["label"])
        # E57 stashes the scanner origin + miss summary keyed by the temp LAS
        # path; pop it so we can surface them in the response (renderer uses the
        # origin for the scan params and miss-point display relocation).
        scan_meta = _e57_scan_meta.pop(str(las_path.resolve()), None)
        if las_is_temp:
            try:
                las_path.unlink()
            except FileNotFoundError:
                pass

        # Auto-detect Helios sky/miss points that arrived without an explicit
        # `is_miss` column (e.g. an ASCII export carrying only target_index): tag
        # them so they're excluded from the hits-only octree / bbox below and feed
        # the overlay + LAD. Runs at this single read point — never re-reads the
        # file (see CLAUDE.md "array is source of truth"). The DISTANCE fallback
        # compares against the scanner origin in the SAME frame as `positions`, so
        # shift the (true-world) origin by world_shift when one was applied.
        _miss_origin = scan_meta["origin"] if (scan_meta and scan_meta.get("origin") is not None) else None
        if _miss_origin is not None and world_shift_arr is not None:
            _miss_origin = (np.asarray(_miss_origin, dtype=np.float64) - world_shift_arr).tolist()
        autodetected_misses = _autodetect_misses(
            positions, extras, extra_dims_meta,
            origin=_miss_origin,
            distance_threshold=request.miss_distance_threshold,
        )

        n = int(len(positions))
        sess = CloudSession(
            session_id=session_id,
            source_path=str(source_path),
            ascii_format=request.ascii_format,
            column_plan=request.column_plan,
            positions=positions,
            colors=colors,
            intensity=intensity,
            extras=extras,
            extra_dims_meta=extra_dims_meta,
            world_shift=world_shift_arr,
            timestamps=timestamps,  # float64 GPS/relative time, not in float32 extras
            gps_time_encoding=gps_time_encoding,
            beam_origins=beam_origins,  # float64 ExtraBytes origins; bypass the join
            crs_epsg=_read_las_crs_epsg(source_path),  # for DEM raster georeferencing
            deleted=np.zeros(n, dtype=bool),
            deleted_history=[],
            octree_cache_id=None,
            created_at=time.time(),
            last_accessed=time.time(),
        )
        # Build the octree from a HITS-ONLY LAS so far-field misses (~20 km) don't
        # poison its bounding box / camera framing. Misses stay in the session
        # (is_miss + true coords) for LAD and the on-demand miss overlay.
        hits_las = tmp_dir / "octree_hits.las"
        _session_to_las(sess, hits_las, exclude_misses=True)
        cache_key, cache_dir, meta = _build_octree_from_las(hits_las, extra_dims_meta)
        sess.octree_cache_id = cache_key
        # Build a SECOND octree from the projected (or true-coord, when no origin)
        # misses, so the renderer streams them with LOD just like the hits —
        # replacing the flat, stride-subsampled overlay (no slowdown, no Moiré).
        # `_miss_origin` is already in the session's frame (world_shift subtracted).
        # The projection radius is baked in here; bake reprojects from the stored
        # origin. None when the scan has no placeable misses.
        sess.miss_octree_origin = _miss_origin
        sess.miss_octree_cache_id = _build_miss_octree(sess, _miss_origin)

    _sweep_cloud_sessions()
    with _cloud_session_lock:
        _cloud_sessions[session_id] = sess
    meta = {"cache_id": cache_key, "cache_dir": str(cache_dir), **meta}

    # Surface sky/miss info so the renderer can hide misses by default, colour
    # them distinctly, and relocate them onto the bounding sphere for display.
    # `has_misses` is derived from the actual data (any source that carried an
    # is_miss extra dim), not just E57. The scanner origin (when known, e.g. from
    # E57 pose) lets the renderer place the scan params + the display relocation.
    miss_arr = extras.get(_MISS_SLUG)
    has_misses = bool(miss_arr is not None and np.any(miss_arr != 0))
    miss_info: dict = {"has_misses": has_misses, "miss_slug": _MISS_SLUG,
                       "miss_octree_cache_id": sess.miss_octree_cache_id}
    if has_misses:
        miss_info["miss_count"] = int(np.count_nonzero(miss_arr != 0))
    # When misses were recovered (no explicit is_miss column), report it so the
    # detection is never silent — the user sees that far-field points were
    # reclassified as sky and pulled out of the bounding box / octree.
    if autodetected_misses > 0:
        miss_info["autodetected_misses"] = autodetected_misses
        _by = ("their target_index == 99 sentinel"
               if extras.get('target_index') is not None
               else "their far-field distance from the scanner")
        miss_info.setdefault("warnings", []).append(
            f"Detected {autodetected_misses} sky/miss point(s) by {_by} and tagged "
            "them (the scan carried no explicit is_miss column). They are excluded "
            "from the displayed cloud and used for the miss overlay + LAD."
        )
    if scan_meta and scan_meta.get("origin") is not None:
        miss_info["scan_origin"] = scan_meta["origin"]
    # Full scan-pattern params (origin + angular sweep + grid resolution) when
    # the source carried them (E57, and PCD VIEWPOINT for origin). The renderer
    # uses these to auto-create a Scan with populated ScanParameters, mirroring
    # the Helios-XML import path. Only the recoverable fields are present.
    if scan_meta and scan_meta.get("scan_params"):
        miss_info["scan_params"] = scan_meta["scan_params"]
    # A LAS carrying per-pulse beam-origin ExtraBytes is a MOVING-platform scan:
    # reconstruct a decimated platform trajectory from the origins + timestamps so
    # the renderer auto-creates a moving scan (path drawn, LAD takes the per-beam
    # path) instead of a plain static cloud. The origins were recentered by
    # world_shift at read, so the trajectory is in the SAME frame as the points.
    if beam_origins is not None:
        _traj_wire = _trajectory_wire_from_beam_origins(beam_origins, timestamps)
        if _traj_wire is not None:
            _first = _traj_wire["poses"][0]
            sp = miss_info.get("scan_params") or {}
            sp.setdefault("origin", [_first["x"], _first["y"], _first["z"]])
            sp["trajectory"] = _traj_wire
            miss_info["scan_params"] = sp
        else:
            miss_info.setdefault("warnings", []).append(
                "This cloud carries per-pulse beam-origin columns but no usable "
                "timestamp to order them into a trajectory; the origins are still "
                "used directly for moving-platform LAD.")
    # Unplaceable misses: flagged miss cells whose beam direction couldn't be
    # recovered at import (zeroed cartesian, no spherical) and so sit at the
    # scanner origin until Helios recovers them from the row/column grid. Surface
    # the count + a warning so this is never silent — a scan that imported "with
    # misses" but whose misses are all unplaceable would otherwise look fine while
    # the LAD geometry is incomplete until the C++ grid recovery runs.
    if scan_meta and scan_meta.get("unplaceable_miss_count"):
        upc = int(scan_meta["unplaceable_miss_count"])
        miss_info["unplaceable_miss_count"] = upc
        miss_info.setdefault("warnings", []).append(
            f"{upc} sky/miss point(s) were flagged but could not be placed at "
            "import (the scanner zeroed invalid-cell coordinates and the file "
            "carries no scan angles). They are kept and tagged; their beam "
            "directions are recovered from the scan grid during LAD."
        )

    # Surface the LAS GPS-time encoding so the renderer can record it on the cloud
    # and the user knows which clock the per-point time uses. A GPS Week Time clock
    # carries no absolute epoch, so a moving-platform trajectory join (attached
    # later) needs a matching relative/week clock — flag that now rather than
    # letting the join silently fail to overlap.
    if gps_time_encoding is not None:
        miss_info["gps_time_encoding"] = gps_time_encoding
        if gps_time_encoding == "gps_week":
            miss_info.setdefault("warnings", []).append(
                "This cloud's per-point GPS time uses GPS Week Time (seconds-into-"
                "week, no absolute epoch). A moving-platform trajectory join needs a "
                "matching clock — attach a trajectory on the same week/relative time, "
                "or re-export the cloud with Adjusted-Standard GPS time."
            )

    # Echo the applied global shift so the renderer persists it on the cloud's
    # OctreeRef (provenance + world-coord readouts). null when no shift was applied.
    world_shift_out = world_shift_arr.tolist() if world_shift_arr is not None else None
    return {"session_id": session_id, "point_count": n, "world_shift": world_shift_out,
            **miss_info, **meta}



def _do_backfill_misses(sess, request, xyz, dirs, labels, vals, flags, progress=None) -> dict:
    """Build an ephemeral PyHelios cloud from a session's hits, gap-fill the misses,
    and persist them in `sess.backfilled_misses`. Returns the result dict.

    The heavy worker behind the backfill endpoint, factored out so it can run
    off-thread under `_bin_frame_streaming_response` and report per-stage progress
    via `progress(fraction, message)` (mirroring `_do_lad_computation`'s `_report`).
    The caller has already resolved the session, assembled the hit arrays, and
    checked eligibility — this only does the expensive build/gapfill/extract.

    On a Helios reconstruction failure (sparse/degenerate grid) it returns an
    `{error: ...}` dict in place of raising, so the streaming JSON tail carries a
    clean, actionable message the renderer can surface as a toast.
    """
    import math
    import numpy as np

    def _report(fraction, message):
        if progress is not None:
            progress(fraction, message)

    origin = request.origin

    # Moving-platform: reconstruct per-beam emission origins so the timestamp
    # gapfill path (the only one valid for moving scans) groups beams correctly.
    moving = request.trajectory is not None
    origins = None
    if moving:
        # The join key is a per-return 'timestamp'; without it _apply_trajectory_origins
        # raises. A moving scan virtually always carries one, but eligibility only
        # requires timestamp OR grid — so a grid-only moving scan would otherwise break
        # the stream with a raw 500. Return a clean, actionable error in the JSON tail
        # instead (mirrors the gapfill-failure / PyHelios-unavailable handling below).
        try:
            dirs, labels, vals, origins = _apply_trajectory_origins(
                xyz, dirs, labels, vals, request.trajectory)
        except ValueError as exc:
            return {"backfilled": 0, "miss_count": 0, "has_misses": False,
                    "scan_origin": list(origin), "already_had_misses": False,
                    "error": (f"{exc} A moving-platform scan needs a per-return "
                              "timestamp to join returns to its trajectory; re-import "
                              "a format that retains per-point timestamps.")}

    # Choose the reconstruction path by what we feed the cloud. The C++ dispatcher
    # auto-selects row/column whenever BOTH 'row'/'column' and 'timestamp' exist,
    # but the row/column path is brittle — it errors when the grid is too sparse
    # ("too few populated scan rows"). The timestamp path reconstructs the grid
    # from per-hit times and is far more robust, so PREFER it:
    #   - has_timestamp        -> feed timestamp only; DROP the grid columns so the
    #                             dispatcher takes the timestamp path.
    #   - grid only (no time)  -> relabel row_index/column_index -> the bare
    #                             'row'/'column' keys the dispatcher probes, so the
    #                             row/column path triggers (its only option here).
    # addHitPointsWithData passes labels through verbatim as the C++ data-map keys.
    prefer_timestamp = flags["has_timestamp"]
    cloud_labels: List[str] = []
    keep_cols: List[int] = []
    for i, l in enumerate(labels):
        if l in ('row_index', 'column_index'):
            if prefer_timestamp:
                continue  # drop the grid column; timestamp path will be used
            cloud_labels.append('row' if l == 'row_index' else 'column')
            keep_cols.append(i)
        else:
            cloud_labels.append(l)
            keep_cols.append(i)
    cloud_vals = (vals[:, keep_cols] if vals is not None and len(keep_cols) < vals.shape[1]
                  else vals)

    # Resolve the scan's angular raster: prefer the supplied values, else estimate
    # from the point count over a default full-hemisphere sweep (mirrors
    # _do_lad_computation's _resolution). The raster sets the grid gapfillMisses()
    # reconstructs over, so it must reflect the scanner, not the surviving points.
    theta_min = request.theta_min if request.theta_min is not None else 0.0
    theta_max = request.theta_max if request.theta_max is not None else 180.0
    phi_min = request.phi_min if request.phi_min is not None else 0.0
    phi_max = request.phi_max if request.phi_max is not None else 360.0
    n_pts = xyz.shape[0]
    if request.n_theta and request.n_phi:
        n_theta, n_phi = int(request.n_theta), int(request.n_phi)
    else:
        aspect = (theta_max - theta_min) / max(phi_max - phi_min, 1e-10)
        n_phi = max(int(math.sqrt(n_pts / max(aspect, 0.01))), 10)
        n_theta = max(int(n_pts / n_phi), 10)

    try:
        from pyhelios import LiDARCloud
    except Exception as exc:  # pragma: no cover - native import guarded at startup
        return {"backfilled": 0, "miss_count": 0, "has_misses": False,
                "scan_origin": list(origin), "already_had_misses": False,
                "error": f"PyHelios unavailable: {exc}"}

    _report(0.05, "Reading scan")
    cloud = LiDARCloud()
    cloud.disableMessages()
    sid = cloud.addScan(
        origin=list(origin),
        Ntheta=n_theta,
        theta_range=(math.radians(theta_min), math.radians(theta_max)),
        Nphi=n_phi,
        phi_range=(math.radians(phi_min), math.radians(phi_max)),
        exit_diameter=(request.beam_exit_diameter or 0.0),
        beam_divergence=((request.beam_divergence or 0.0) / 1000.0),
    )
    if n_pts > 0:
        _report(0.15, f"Building cloud ({n_pts:,} points)")
        cloud.addHitPointsWithData(sid, xyz, dirs, cloud_labels, cloud_vals)

    # The gapfill itself is one opaque C++ call — report an INDETERMINATE stage
    # (None fraction → pulsing bar, like LAD's ray-trace step) so the UI shows
    # activity while it runs.
    _report(None, "Reconstructing misses")
    try:
        synth_xyz, count = _run_gapfill_extract(cloud)
    except Exception as exc:
        # Helios raises (HeliosRuntimeError) when it can't reconstruct the scan
        # grid — e.g. a sparse row/column raster ("too few populated scan rows"),
        # or returns/timestamps too degenerate to group. Return a clean, actionable
        # error in the JSON tail rather than breaking the stream with a raw 500.
        msg = str(exc).split("ERROR (")[-1].rstrip(") ") or str(exc)
        return {"backfilled": 0, "miss_count": 0, "has_misses": False,
                "scan_origin": list(origin), "already_had_misses": False,
                "error": ("Could not reconstruct sky/miss points for this scan: "
                          f"{msg}. The scan grid is too sparse or irregular to "
                          "gap-fill. If it carries a timestamp, ensure each return "
                          "keeps it; otherwise re-import a format that retains "
                          "misses (E57 / structured PLY).")}

    _report(0.9, "Storing misses")
    # Reconstruct per-beam directions for the synthesised misses so LAD's beam
    # path can re-emit them. For a moving scan we lack each synthetic miss's own
    # origin, so fall back to the scan origin (the timestamp grouping already
    # placed the miss in the right direction relative to it).
    synth_dirs = _directions_from_origin(synth_xyz, origin)

    buffer = {
        "positions": synth_xyz,
        "directions": synth_dirs,
    }
    if moving and origins is not None and synth_xyz.shape[0] > 0:
        # Per-pulse origins aren't recoverable per synthetic miss here; store the
        # scan-origin broadcast so the LAD reader has an origin column. Refined
        # only if a future bulk getter exposes per-miss origins.
        buffer["origins"] = np.tile(np.asarray(origin, dtype=np.float64),
                                    (synth_xyz.shape[0], 1))

    with _cloud_session_lock:
        sess.backfilled_misses = buffer if count > 0 else None
        # Freshly computed against the current hits — no longer stale, and the
        # scan origin is now known for the display projection.
        sess.backfilled_misses_stale = False
        if count > 0:
            sess.miss_octree_origin = list(origin)
        sess.last_accessed = time.time()

    # Rebuild the miss octree so the newly recovered sky points stream in (fresh
    # sha1 → the renderer remounts on the changed cache id). None when count == 0.
    miss_cache_id = _build_miss_octree(sess, list(origin)) if count > 0 else None
    with _cloud_session_lock:
        sess.miss_octree_cache_id = miss_cache_id

    _report(1.0, "Done")
    return {
        "backfilled": count,
        "miss_count": count,
        "has_misses": count > 0,
        "scan_origin": list(origin),
        "already_had_misses": False,
        "miss_octree_cache_id": miss_cache_id,
    }


@app.post("/api/cloud/session/{session_id}/backfill-misses")
async def backfill_cloud_misses(session_id: str, request: BackfillMissesRequest):
    """Explicitly recover sky/miss points for a session and persist them.

    LAD needs miss points (beams that returned nothing) for the Beer's-law
    transmission denominator. Some formats retain them (E57 / structured PLY);
    others don't but carry the data to RECONSTRUCT them — a per-hit timestamp
    and/or scan-grid row/column indices. This endpoint builds an ephemeral
    PyHelios cloud from the session's surviving points, runs gapfillMisses()
    (which auto-selects the row/column or timestamp path in C++), extracts the
    synthesised misses, and stores them in a lightweight per-session buffer
    (`sess.backfilled_misses`) — leaving the hit arrays untouched.

    Session resolve + array assembly + eligibility run up front (so a bad request
    is a clean 404/400); the heavy build/gapfill/extract streams PHP1 progress
    markers ahead of the JSON tail (see `_bin_frame_streaming_response`) so the
    renderer shows a per-stage progress bar. The JSON tail is
    {backfilled, miss_count, has_misses, scan_origin, already_had_misses} on
    success, or carries `error` when reconstruction failed.
    """
    sess = _get_cloud_session(session_id)
    origin = request.origin
    if len(origin) != 3:
        raise HTTPException(status_code=400, detail="origin must have 3 elements")

    with _cloud_session_lock:
        xyz, dirs, labels, vals, flags = _session_to_lad_arrays(
            sess, origin, include_backfilled=False)

    # Already carries real misses (E57 / structured PLY): nothing to recover. A
    # trivial no-op — return a plain JSON response (no streaming needed).
    if flags["has_misses"]:
        return JSONResponse({
            "backfilled": 0,
            "miss_count": 0,
            "has_misses": True,
            "scan_origin": list(origin),
            "already_had_misses": True,
        })

    # Eligibility: gapfillMisses() reconstructs miss directions from EITHER a
    # per-hit timestamp OR the native scan-grid row/column indices. With neither,
    # there's no way to recover them — 400 with an actionable message instead of
    # letting Helios raise.
    if not (flags["has_timestamp"] or flags["has_grid"]):
        raise HTTPException(
            status_code=400,
            detail=("This scan has no sky/miss points and no way to reconstruct "
                    "them: it carries neither a per-pulse timestamp nor scan-grid "
                    "row/column indices. Re-import a scan that retains misses "
                    "(E57 / structured PLY) or one of those columns."),
        )

    # Stream the heavy build/gapfill/extract with per-stage progress markers.
    return _bin_frame_streaming_response(
        lambda progress: json.dumps(
            _do_backfill_misses(sess, request, xyz, dirs, labels, vals, flags,
                                progress=progress)).encode("utf-8"))


@app.post("/api/cloud/session/{session_id}/delete_region")
async def delete_cloud_region(session_id: str, request: DeleteRegionRequest):
    """Set the per-point deleted mask for points inside `region`. No rebuild."""
    sess = _get_cloud_session(session_id)
    region_dict = request.region.model_dump()
    _canonical_region(region_dict)  # validate shape (raises 400)

    # `_region_mask` returns the spatial keep-mask (region invert already
    # applied). The points to DELETE are exactly those the region selects —
    # i.e. the True entries of the (un-inverted) selection. We OR them into the
    # session's deleted mask so deletions accumulate.
    with _cloud_session_lock:
        select = _region_mask(sess.positions, region_dict)
        # NEVER crop sky/miss points. A crop box is drawn around hits; a miss
        # point falling inside it (its far-field/projected coords landing in the
        # box) is a coordinate accident, not user intent — and deleting it would
        # silently corrupt LAD's beam set. Exclude is_miss points from the
        # selection so a crop only ever deletes hits. (Bake already subsets the
        # is_miss extras column; this stops crop from SELECTING misses at all.)
        miss_arr = sess.extras.get(_MISS_SLUG)
        if miss_arr is not None:
            select = select & (miss_arr == 0)
        sess.deleted |= select
        # Record the post-delete mask snapshot so undo can pop back to it. We
        # store the boolean mask per applied delete (cheap: 1 bit/point) rather
        # than replaying regions, so undo is exact regardless of edit kind.
        sess.deleted_history.append(sess.deleted.copy())
        # Bound the undo stack: each snapshot is a full (N,) bool mask, so an
        # unbounded history grows RAM by ~1 byte/point per erase click. Keep only
        # the most recent _MAX_DELETED_HISTORY snapshots (older undos are dropped).
        if len(sess.deleted_history) > _MAX_DELETED_HISTORY:
            sess.deleted_history = sess.deleted_history[-_MAX_DELETED_HISTORY:]
        sess.octree_cache_id = None  # derived octree is now stale until bake
        # A crop that actually removed hits invalidates any SEPARATELY backfilled
        # misses: they were gap-filled against the pre-crop hits, so their
        # hit/miss ratio no longer matches and would skew LAD. Per the product
        # decision we KEEP them (don't discard the user's compute) but flag them
        # stale so the renderer warns at crop time and LAD warns at compute time.
        # Interleaved is_miss points (in extras) aren't affected — only the
        # separate backfilled buffer can drift out of sync.
        cropped_hits = bool(select.any())
        if (cropped_hits and sess.backfilled_misses is not None
                and sess.backfilled_misses.get("positions") is not None):
            sess.backfilled_misses_stale = True
        deleted_count = int(sess.deleted.sum())
        total = int(len(sess.positions))
        backfilled_misses_stale = sess.backfilled_misses_stale

    return {
        "session_id": session_id,
        "deleted_count": deleted_count,
        "remaining_count": total - deleted_count,
        "total_count": total,
        "backfilled_misses_stale": backfilled_misses_stale,
    }


class ResetCloudEditsRequest(BaseModel):
    """Undo. `edit_count` = how many committed deletes to KEEP; the mask is
    restored to that snapshot in the history and later ones are discarded.
    Omit to clear all deletions (edit_count = 0)."""
    edit_count: Optional[int] = None


@app.post("/api/cloud/session/{session_id}/reset_edits")
async def reset_cloud_edits(session_id: str, request: ResetCloudEditsRequest):
    """Restore the deleted mask to an earlier snapshot (undo)."""
    sess = _get_cloud_session(session_id)
    with _cloud_session_lock:
        k = 0 if request.edit_count is None else max(0, int(request.edit_count))
        k = min(k, len(sess.deleted_history))
        sess.deleted_history = sess.deleted_history[:k]
        sess.deleted = (
            sess.deleted_history[-1].copy() if k > 0
            else np.zeros(len(sess.positions), dtype=bool)
        )
        sess.octree_cache_id = None
        deleted_count = int(sess.deleted.sum())
        total = int(len(sess.positions))
    return {
        "session_id": session_id,
        "deleted_count": deleted_count,
        "remaining_count": total - deleted_count,
        "total_count": total,
    }


@app.post("/api/cloud/session/{session_id}/bake")
async def bake_cloud_session(session_id: str):
    """Permanently apply deletions by rebuilding the octree FROM THE IN-RAM
    ARRAYS — the survivors (positions[~deleted] + colours + intensity + every
    scalar extra-dim) are written to a LAS via `_session_to_las` and fed to
    PotreeConverter. The source file is NOT read. Then the in-RAM arrays are
    compacted to the survivors and the mask cleared. Returns the new octree
    metadata. The deliberately-slow step (the PotreeConverter run).

    No deletions → returns the current octree without rebuilding.
    """
    sess = _get_cloud_session(session_id)
    with _cloud_session_lock:
        has_deletions = bool(sess.deleted.any())
        survivors = int((~sess.deleted).sum())

    # Everything deleted → nothing to bake. Don't feed a 0-point LAS to
    # PotreeConverter (it exits non-zero). Report empty; the renderer raises a
    # delete-confirmation. The mask is left intact (the cloud is still "all
    # deleted" until the user removes it).
    if survivors == 0:
        return {"session_id": session_id, "point_count": 0, "baked": False}

    if not has_deletions:
        cache_dir = _octree_cache_root() / (sess.octree_cache_id or "")
        if sess.octree_cache_id and (cache_dir / "metadata.json").is_file():
            meta = _read_octree_metadata(cache_dir)
            return {
                "session_id": session_id, "point_count": int(len(sess.positions)),
                "baked": False, "cache_id": sess.octree_cache_id,
                "cache_dir": str(cache_dir), **meta,
            }

    import tempfile
    with tempfile.TemporaryDirectory() as _tmp:
        las_path = _Path(_tmp) / "baked.las"
        with _cloud_session_lock:
            # Octree is hits-only (misses stay in the session for LAD/overlay).
            _session_to_las(sess, las_path, exclude_misses=True)
            extra_dims_meta = list(sess.extra_dims_meta)
        cache_key, cache_dir, meta = _build_octree_from_las(las_path, extra_dims_meta)

    # Compact the in-RAM arrays to the survivors and clear the mask + history, so
    # the session's source of truth matches the baked octree and further edits
    # start from the reduced set.
    with _cloud_session_lock:
        keep = ~sess.deleted
        sess.positions = sess.positions[keep]
        if sess.colors is not None:
            sess.colors = sess.colors[keep]
        if sess.intensity is not None:
            sess.intensity = sess.intensity[keep]
        for slug in list(sess.extras.keys()):
            sess.extras[slug] = sess.extras[slug][keep]
        if sess.timestamps is not None:
            sess.timestamps = sess.timestamps[keep]
        sess.deleted = np.zeros(len(sess.positions), dtype=bool)
        sess.deleted_history = []
        sess.octree_cache_id = cache_key
        remaining = int(len(sess.positions))

    # Rebuild the miss octree from the surviving misses so the displayed shell
    # tracks the baked cloud (a crop that removed hits also reprojects the misses
    # against the new far hit). Reprojected from the stored origin; None if no
    # misses survive. Note the backfilled-miss DATA can still be flagged stale
    # (see delete_region) — this only keeps the rendered shell current.
    miss_cache_id = _build_miss_octree(sess, sess.miss_octree_origin)
    with _cloud_session_lock:
        sess.miss_octree_cache_id = miss_cache_id

    try:
        max_bytes = int(_os.environ.get(
            "PHYTOGRAPH_OCTREE_CACHE_MAX_BYTES", _DEFAULT_OCTREE_CACHE_MAX_BYTES,
        ))
    except ValueError:
        max_bytes = _DEFAULT_OCTREE_CACHE_MAX_BYTES
    keep_dirs = [cache_dir]
    if miss_cache_id:
        keep_dirs.append(_octree_cache_root() / miss_cache_id)
    _evict_octree_cache(max_bytes, keep=keep_dirs)

    return {
        "session_id": session_id,
        "point_count": remaining,
        "baked": True,
        "cache_id": cache_key,
        "cache_dir": str(cache_dir),
        "miss_octree_cache_id": miss_cache_id,
        **meta,
    }


def _session_add_extra_column(sess: "CloudSession", slug: str, label: str, values: np.ndarray) -> None:
    """Append (or replace) a per-point scalar extra-dim column on the session
    array. `values` is aligned to the SURVIVING points (positions[~deleted]);
    it's scattered back to a full-length (N,) column with 0 for deleted rows so
    every session array stays the same length. Caller holds the lock."""
    full = np.zeros(len(sess.positions), dtype=np.float32)
    full[~sess.deleted] = values.astype(np.float32)
    sess.extras[slug] = full
    if slug not in {ed["slug"] for ed in sess.extra_dims_meta}:
        sess.extra_dims_meta.append({"slug": slug, "label": label})


def _session_rebuild(sess: "CloudSession") -> tuple[str, _Path, dict]:
    """Rebuild the session's derived octree FROM THE IN-RAM ARRAYS (survivors +
    all attributes), update octree_cache_id, and return (cache_key, dir, meta).
    No source file read. Caller must NOT hold the lock (PotreeConverter is slow);
    this snapshots under the lock then converts outside it."""
    import tempfile
    with tempfile.TemporaryDirectory() as _tmp:
        las_path = _Path(_tmp) / "rebuilt.las"
        with _cloud_session_lock:
            # Octree is hits-only (misses stay in the session for LAD/overlay).
            _session_to_las(sess, las_path, exclude_misses=True)
            extra_dims_meta = list(sess.extra_dims_meta)
        cache_key, cache_dir, meta = _build_octree_from_las(las_path, extra_dims_meta)
    with _cloud_session_lock:
        sess.octree_cache_id = cache_key
    return cache_key, cache_dir, meta


def _session_subset_locked(sess: "CloudSession", keep: np.ndarray) -> "CloudSession":
    """Build a NEW session from the `keep` subset of `sess`'s SURVIVING points
    (positions[~deleted][keep] + aligned attributes) and register it. The CALLER
    MUST HOLD `_cloud_session_lock` — this reads `sess`'s arrays and inserts the
    new session into the registry without taking the lock itself, so it can be
    composed inside a larger locked critical section (e.g. split's commit). The
    octree is NOT built here (caller rebuilds, outside the lock)."""
    import time
    surv = ~sess.deleted
    new_id = uuid.uuid4().hex[:8]
    new_sess = CloudSession(
        session_id=new_id,
        source_path=sess.source_path,  # provenance label only
        ascii_format=sess.ascii_format,
        column_plan=sess.column_plan,
        positions=sess.positions[surv][keep].copy(),
        colors=sess.colors[surv][keep].copy() if sess.colors is not None else None,
        intensity=sess.intensity[surv][keep].copy() if sess.intensity is not None else None,
        extras={k: v[surv][keep].copy() for k, v in sess.extras.items()},
        extra_dims_meta=list(sess.extra_dims_meta),
        # Carry the float64 timestamps onto the subset so the moving-platform LAD
        # join key survives a split/crop with full precision (and stays aligned).
        timestamps=(np.asarray(sess.timestamps)[surv][keep].copy()
                    if sess.timestamps is not None else None),
        # Inherit the parent's import shift so the subset's stored positions stay
        # in the same (shifted) space and still restore to world coords on read.
        world_shift=(sess.world_shift.copy() if sess.world_shift is not None else None),
        crs_epsg=sess.crs_epsg,  # subsets keep the parent's CRS for DEM georeferencing
        deleted=np.zeros(int(keep.sum()), dtype=bool),
        deleted_history=[],
        octree_cache_id=None,
        created_at=time.time(),
        last_accessed=time.time(),
    )
    _cloud_sessions[new_id] = new_sess
    return new_sess


def _session_subset(sess: "CloudSession", keep: np.ndarray) -> "CloudSession":
    """Lock-acquiring wrapper around `_session_subset_locked` for callers that do
    NOT already hold the lock (e.g. `extract`)."""
    with _cloud_session_lock:
        return _session_subset_locked(sess, keep)


class SessionSplitRequest(BaseModel):
    """Split a session into the points a filter KEEPS (stay on this session) and
    the points it EXCLUDES (a NEW leftover session). Operates entirely on the
    in-RAM arrays — no source file read. Same predicate shape as the filter."""
    region: Optional[CropOctreeRegion] = None
    scalar_filters: Optional[List[ScalarFilter]] = None


@app.post("/api/cloud/session/{session_id}/split")
async def session_split(session_id: str, request: SessionSplitRequest):
    """Keep the filter-passing points on this session; move the excluded points
    to a NEW leftover session. Rebuilds both octrees from arrays. Returns
    {kept: {...octree}, leftover: {session_id, ...octree}} (leftover null if
    empty). No file read."""
    sess = _get_cloud_session(session_id)
    region_dict = request.region.model_dump() if request.region else None
    if region_dict is not None:
        _canonical_region(region_dict)
    if region_dict is None and not request.scalar_filters:
        raise HTTPException(status_code=400, detail="split requires `region` or `scalar_filters`.")

    # Compute the keep-mask AND commit the leftover deletion under ONE lock, so
    # the survivor snapshot used for the index scatter can't go stale against a
    # concurrent edit on the same session (the leftover subset is built from the
    # same snapshot before the lock is released for the slow rebuilds).
    with _cloud_session_lock:
        surv = ~sess.deleted
        pos = sess.positions[surv]
        keep = _region_mask(pos, region_dict) if region_dict is not None else np.ones(len(pos), dtype=bool)
        for f in (request.scalar_filters or []):
            if f.slug not in sess.extras:
                raise HTTPException(status_code=400, detail=f"Unknown scalar attribute: {f.slug!r}.")
            lo, hi, value_set = _resolve_scalar_filter(f.model_dump())
            keep &= _scalar_filter_mask(sess.extras[f.slug][surv], lo, hi, value_set)

        kept_count = int(keep.sum())
        # Kept side empty → the filter would leave nothing on the parent. Don't
        # commit or rebuild (0-point PotreeConverter would 500); report empty so
        # the renderer raises a delete-confirmation. The session is untouched.
        if kept_count == 0:
            return {
                "session_id": session_id,
                "kept": {"point_count": 0, "cache_id": None, "cache_dir": None},
                "leftover": None,
            }

        leftover_mask = ~keep
        leftover = _session_subset_locked(sess, leftover_mask) if bool(leftover_mask.any()) else None
        # Commit the leftover deletion on the parent using the SAME snapshot.
        # Split is a non-undoable commit (renderer clears its erase undo stack),
        # so reset the undo history to keep both sides in lock-step.
        idx_surv = np.where(surv)[0]
        sess.deleted[idx_surv[leftover_mask]] = True
        sess.deleted_history = []
        sess.octree_cache_id = None

    leftover_meta = None
    if leftover is not None:
        lk, lcd, lmeta = _session_rebuild(leftover)
        leftover_meta = {"session_id": leftover.session_id, "point_count": int(len(leftover.positions)),
                         "cache_id": lk, "cache_dir": str(lcd), **lmeta}
    kk, kcd, kmeta = _session_rebuild(sess)

    return {
        "session_id": session_id,
        "kept": {"point_count": kept_count, "cache_id": kk, "cache_dir": str(kcd), **kmeta},
        "leftover": leftover_meta,
    }


class SessionExtractRequest(BaseModel):
    """Extract the points a spatial+scalar filter SELECTS into a NEW child
    session, leaving the parent UNCHANGED. Operates on the in-RAM arrays — no
    source file read. Used by 'split into clouds' workflows that keep the
    classified parent and spin off per-class child clouds."""
    region: Optional[CropOctreeRegion] = None
    scalar_filters: Optional[List[ScalarFilter]] = None


@app.post("/api/cloud/session/{session_id}/extract")
async def session_extract(session_id: str, request: SessionExtractRequest):
    """Create a NEW child session from the filter-selected points (parent
    untouched). Returns {session_id, ...octree} or null if the selection is
    empty. No source file read."""
    sess = _get_cloud_session(session_id)
    region_dict = request.region.model_dump() if request.region else None
    if region_dict is not None:
        _canonical_region(region_dict)
    if region_dict is None and not request.scalar_filters:
        raise HTTPException(status_code=400, detail="extract requires `region` or `scalar_filters`.")

    with _cloud_session_lock:
        surv = ~sess.deleted
        pos = sess.positions[surv]
        sel = _region_mask(pos, region_dict) if region_dict is not None else np.ones(len(pos), dtype=bool)
        for f in (request.scalar_filters or []):
            if f.slug not in sess.extras:
                raise HTTPException(status_code=400, detail=f"Unknown scalar attribute: {f.slug!r}.")
            lo, hi, value_set = _resolve_scalar_filter(f.model_dump())
            sel &= _scalar_filter_mask(sess.extras[f.slug][surv], lo, hi, value_set)

    if not bool(sel.any()):
        return {"session_id": session_id, "extracted": None}
    child = _session_subset(sess, sel)
    ck, ccd, cmeta = _session_rebuild(child)
    return {
        "session_id": session_id,
        "extracted": {"session_id": child.session_id, "point_count": int(len(child.positions)),
                      "cache_id": ck, "cache_dir": str(ccd), **cmeta},
    }


@app.post("/api/cloud/session/{session_id}/duplicate")
async def session_duplicate(session_id: str):
    """Copy a session's SURVIVING points into a NEW independent session (parent
    untouched) and build its octree. This is the keep-everything degenerate case
    of `extract`: a pure array copy via `_session_subset` — NO source file read,
    so every wizard customization (column plan, custom labels, categorical
    slugs, dropped/renamed extras) is preserved on the copy. The new session is
    fully independent: later edits to either side don't affect the other.
    Returns {session_id, duplicate: {session_id, point_count, ...octree}}."""
    sess = _get_cloud_session(session_id)
    with _cloud_session_lock:
        # `_session_subset_locked` keeps `surv[keep]`; an all-true keep over the
        # survivors copies the whole surviving cloud.
        keep = np.ones(int((~sess.deleted).sum()), dtype=bool)
        child = _session_subset_locked(sess, keep)
    ck, ccd, cmeta = _session_rebuild(child)
    return {
        "session_id": session_id,
        "duplicate": {"session_id": child.session_id, "point_count": int(len(child.positions)),
                      "cache_id": ck, "cache_dir": str(ccd), **cmeta},
    }


class SessionGroundSegmentRequest(BaseModel):
    """Run CSF ground segmentation on the session's in-RAM points and append a
    `ground_class` scalar column (1=ground, 2=plant). No source file read."""
    cloth_resolution: float = 0.05
    rigidness: int = 3
    class_threshold: float = 0.02
    iterations: int = 500
    slope_smooth: bool = False


@app.post("/api/cloud/session/{session_id}/segment_ground")
async def session_segment_ground(session_id: str, request: SessionGroundSegmentRequest):
    """CSF on the in-RAM survivors → append `ground_class` → rebuild octree from
    the arrays. No source file read."""
    sess = _get_cloud_session(session_id)
    with _cloud_session_lock:
        pts = sess.positions[~sess.deleted].copy()
    if len(pts) < 10:
        raise HTTPException(status_code=400, detail="Need at least 10 points for ground segmentation.")
    try:
        labels = segment_ground(
            pts,
            cloth_resolution=request.cloth_resolution,
            rigidness=request.rigidness,
            class_threshold=request.class_threshold,
            iterations=request.iterations,
            slope_smooth=request.slope_smooth,
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="CSF (cloth-simulation-filter) not installed.")
    with _cloud_session_lock:
        _session_add_extra_column(sess, GROUND_CLASS_SLUG, GROUND_CLASS_LABEL, labels)
    cache_key, cache_dir, meta = _session_rebuild(sess)
    return {"session_id": session_id, "point_count": int(len(pts)), "cache_id": cache_key, "cache_dir": str(cache_dir), **meta}


class SessionDemRequest(BaseModel):
    """Generate a DEM from a session's in-RAM survivors. Ground-aware: a prior
    `ground_class` column restricts gridding to ground points; else CSF is run
    when `auto_segment_ground` is set, else all points are used. Optionally write
    a `height_above_ground` (CHM) scalar back onto the cloud and rebuild."""
    auto_segment_ground: bool = True
    cloth_resolution: float = 0.05
    rigidness: int = 3
    class_threshold: float = 0.02
    cell_size: Optional[float] = None
    bbox: Optional[List[float]] = None
    method: str = "tin"
    ground_percentile: float = 5.0
    fill_voids: bool = False
    add_height_column: bool = False


def _do_session_dem(sess: "CloudSession", request: "SessionDemRequest", progress=None) -> dict:
    """Session DEM worker. The mesh is built in the session's (world_shift-
    subtracted) display coordinates so it aligns with the rendered octree; the
    `world_shift` rides in the result so raster export can recover true-world
    coordinates. When `add_height_column`, samples the DEM at every survivor and
    appends a `height_above_ground` column, then rebuilds the octree."""
    try:
        with _cloud_session_lock:
            keep = ~sess.deleted
            pts = sess.positions[keep].copy()
            ground_col = sess.extras.get(GROUND_CLASS_SLUG)
            is_ground = (ground_col[keep] == GROUND_CLASS_GROUND
                         if ground_col is not None else None)
            world_shift = sess.world_shift
            crs_epsg = sess.crs_epsg
        if len(pts) < 3:
            return {"success": False, "method_used": f"dem_{request.method}", "num_triangles": 0,
                    "num_vertices": 0, "error": "Need at least 3 points to build a DEM."}

        warning = None
        if is_ground is not None and int(is_ground.sum()) >= 3:
            ground_source = "column"
            ground_pts = pts[is_ground]
        elif request.auto_segment_ground:
            try:
                if progress is not None:
                    progress(0.05, "Extracting ground points (CSF)")
                csf = _auto_csf_params(pts)   # extent-scaled — never the 5 cm plant default on a big tile
                labels = segment_ground(pts, **csf)
                gmask = labels == GROUND_CLASS_GROUND
                if int(gmask.sum()) >= 3:
                    ground_source = "csf_auto"
                    ground_pts = pts[gmask]
                else:
                    ground_source = "all_points"; ground_pts = pts
                    warning = "CSF found too few ground points; used all points."
            except ImportError:
                ground_source = "all_points"; ground_pts = pts
                warning = "CSF not installed; used all points (lowest returns)."
        else:
            ground_source = "all_points"; ground_pts = pts
            warning = "No ground classification; used all points (lowest returns)."

        # When writing height-above-ground, hand _compute_dem every survivor's XY
        # so it returns a gap-free ground elevation under each point (linear in the
        # ground hull, nearest-ground outside) — never a 0-height artifact for
        # canopy past the ground footprint or over a ground gap.
        result = _compute_dem(ground_pts, cell_size=request.cell_size, bbox=request.bbox,
                              method=request.method, ground_percentile=request.ground_percentile,
                              fill_voids=request.fill_voids,
                              sample_xy=(pts[:, :2] if request.add_height_column else None),
                              progress=progress)
        if not result.get("success"):
            return result
        sample_ground_z = result.pop("sample_ground_z", None)   # not framed
        result["ground_source"] = ground_source
        result["world_shift"] = [float(v) for v in (world_shift if world_shift is not None else (0.0, 0.0, 0.0))]
        if crs_epsg is not None:
            result["crs_epsg"] = int(crs_epsg)
        if warning:
            result["warning"] = warning

        if request.add_height_column and sample_ground_z is not None:
            hag = pts[:, 2] - sample_ground_z
            hag[~np.isfinite(hag)] = 0.0
            with _cloud_session_lock:
                _session_add_extra_column(sess, HEIGHT_ABOVE_GROUND_SLUG, HEIGHT_ABOVE_GROUND_LABEL, hag)
            cache_key, cache_dir, meta = _session_rebuild(sess)
            result["cache_id"] = cache_key
            result["cache_dir"] = str(cache_dir)
            result["point_count"] = int(len(pts))
            result.update(meta)
        return result
    except ValueError as e:
        return {"success": False, "method_used": f"dem_{request.method}", "num_triangles": 0,
                "num_vertices": 0, "error": str(e)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "method_used": f"dem_{request.method}", "num_triangles": 0,
                "num_vertices": 0, "error": f"DEM generation failed: {e}"}


@app.post("/api/cloud/session/{session_id}/dem")
async def session_generate_dem(session_id: str, request: SessionDemRequest, http_request: Request):
    """DEM from a session's in-RAM survivors (ground-aware). Returns a PHB1 frame
    (heightmap mesh + grid). When `add_height_column`, also appends a
    `height_above_ground` scalar and rebuilds the octree (cache_id in meta)."""
    sess = _get_cloud_session(session_id)
    run_id, cancel_event = _new_cancel_token()
    return _bin_frame_streaming_response(
        lambda progress: _pack_dem_frame(_do_session_dem(sess, request, progress=progress)),
        request=http_request, cancel_event=cancel_event, run_id=run_id)


class SessionWoodSegmentRequest(WoodSegmentationRequest):
    """Run wood/leaf segmentation on the session's in-RAM points and append a
    `wood_class` column (1=wood, 2=leaf). Inherits the segment_wood tuning
    fields; `points`/`source` are ignored (the session's arrays are the source
    of truth)."""
    pass


def _session_reflectance_scalar(sess, scalar_slug, keep):
    """Resolve an optional per-point reflectance/intensity scalar from a session,
    masked to surviving points (`keep`), as float64 aligned 1:1 with the points,
    or None when nothing usable is present.

    Lookup order: an explicit `scalar_slug` (case-insensitive) in the extra-dim
    columns → the common reflectance/intensity slugs → the LAS `intensity` field.
    Slug matching is case-insensitive because the Riegl extra-dim is 'Reflectance'
    while ASCII roles are lowercase 'reflectance'/'intensity'.
    """
    extras = sess.extras or {}
    lower = {k.lower(): k for k in extras}

    def _take(arr):
        a = np.asarray(arr, dtype=np.float64)
        return a[keep] if a.shape[0] == keep.shape[0] else None

    if scalar_slug:
        key = lower.get(scalar_slug.lower())
        if key is not None:
            return _take(extras[key])
    for cand in ("reflectance", "intensity"):
        key = lower.get(cand)
        if key is not None:
            return _take(extras[key])
    if sess.intensity is not None:
        return _take(sess.intensity)
    return None


@app.post("/api/cloud/session/{session_id}/segment_wood")
async def session_segment_wood(session_id: str, request: SessionWoodSegmentRequest):
    """Wood/leaf segmentation on the in-RAM survivors → append `wood_class` →
    rebuild octree from the arrays. No source file read."""
    sess = _get_cloud_session(session_id)
    with _cloud_session_lock:
        keep = ~sess.deleted
        pts = sess.positions[keep].copy()
        # Resolve an optional per-point reflectance scalar from the session,
        # aligned 1:1 with the surviving points. `scalar_slug` picks a specific
        # extra-dim (e.g. 'Reflectance'); default tries the common slugs then the
        # LAS intensity field. None ⇒ pure geometry. The reflectance_weight_max
        # request field still gates whether the assist runs at all.
        reflectance = None
        if request.reflectance_weight_max and request.reflectance_weight_max > 0:
            reflectance = _session_reflectance_scalar(sess, request.scalar_slug, keep)
    if len(pts) < 3:
        raise HTTPException(status_code=400, detail="Need at least 3 points for wood/leaf segmentation.")
    warns: List[str] = []
    labels = segment_wood(pts, reflectance=reflectance, warnings=warns, **_wood_segment_kwargs(request))
    with _cloud_session_lock:
        _session_add_extra_column(sess, WOOD_CLASS_SLUG, WOOD_CLASS_LABEL, labels)
    cache_key, cache_dir, meta = _session_rebuild(sess)
    return {"session_id": session_id, "point_count": int(len(pts)), "cache_id": cache_key, "cache_dir": str(cache_dir), "warnings": warns, **meta}


class SessionTreeSegmentRequest(TreeSegmentationRequest):
    """Run TreeIso on the session's in-RAM points and append a `tree_instance`
    column. Inherits the TreeIso tuning fields; `points`/`source` are ignored."""
    pass


@app.post("/api/cloud/session/{session_id}/segment_trees")
async def session_segment_trees(session_id: str, request: SessionTreeSegmentRequest,
                                http_request: Request):
    """TreeIso on the in-RAM survivors → append `tree_instance` → rebuild octree
    from the arrays. No source file read.

    The TreeIso pipeline runs off the event loop (`_run_blocking_until_disconnect`)
    so the server stays responsive during the tens-of-seconds compute and a client
    disconnect (panel closed / fetch timeout) returns promptly instead of holding
    the request open.

    If the cloud carries a `ground_class` column (from a prior ground
    segmentation that was *labeled* but not removed), the ground points are
    excluded from TreeIso and assigned tree id 0 (unassigned) — TreeIso only
    sees the plant points, so ground never gets clustered into a "tree"."""
    sess = _get_cloud_session(session_id)
    with _cloud_session_lock:
        keep = ~sess.deleted
        pts = sess.positions[keep].copy()
        # Survivor-aligned ground mask: True where a prior ground segmentation
        # tagged the point as ground. Absent column ⇒ no exclusion.
        ground_col = sess.extras.get(GROUND_CLASS_SLUG)
        is_ground = (
            ground_col[keep] == GROUND_CLASS_GROUND
            if ground_col is not None else np.zeros(len(pts), dtype=bool)
        )
    plant_mask = ~is_ground
    n_plant = int(plant_mask.sum())
    if n_plant > _TREEISO_MAX_POINTS:
        raise HTTPException(
            status_code=400,
            detail=f"Tree segmentation is capped at {_TREEISO_MAX_POINTS:,} points; this cloud has {n_plant:,} non-ground points. Crop or downsample first.",
        )
    if n_plant < 10:
        raise HTTPException(
            status_code=400,
            detail=f"Need at least 10 non-ground points for tree segmentation; found {n_plant}.",
        )
    seeds = (
        np.asarray(request.seed_points, dtype=np.float64)
        if request.seed_points else None
    )
    plant_pts = pts[plant_mask]

    def _segment():
        ti_params = _treeiso_params(request)
        _auto_treeiso_decimation(plant_pts, ti_params)   # ~0.8s probe — keep off-loop too
        return segment_trees(plant_pts, ti_params, seeds=seeds)

    try:
        plant_labels = await _run_blocking_until_disconnect(_segment, http_request)
    except ClientDisconnected:
        # Client gave up before the octree rebuild; skip it and leave the session
        # untouched (no tree_instance column written).
        raise HTTPException(status_code=499, detail="Tree segmentation cancelled (client disconnected).")
    # Scatter the plant tree ids back onto all survivors; ground stays 0.
    labels = np.zeros(len(pts), dtype=np.int64)
    labels[plant_mask] = np.asarray(plant_labels)
    with _cloud_session_lock:
        _session_add_extra_column(sess, TREE_INSTANCE_SLUG, TREE_INSTANCE_LABEL, labels)
    cache_key, cache_dir, meta = _session_rebuild(sess)
    return {"session_id": session_id, "point_count": int(len(pts)), "cache_id": cache_key, "cache_dir": str(cache_dir), **meta}


class SessionFilterRequest(BaseModel):
    """Apply a spatial + scalar filter to the session by DELETING the points the
    filter excludes (sets the deleted mask). Operates entirely on the in-RAM
    arrays — no source file read. `region` keeps points inside it (invert to
    flip); `scalar_filters` keep points whose attribute is in range/class. A
    point survives iff it passes the region AND every scalar filter."""
    region: Optional[CropOctreeRegion] = None
    scalar_filters: Optional[List[ScalarFilter]] = None
    # Rebuild the octree now (the filter is "applied permanently"). When False,
    # only the mask is set (instant; bake later).
    rebuild: bool = True


@app.post("/api/cloud/session/{session_id}/filter")
async def session_filter(session_id: str, request: SessionFilterRequest):
    """Delete the points a spatial+scalar filter excludes, on the in-RAM arrays."""
    sess = _get_cloud_session(session_id)
    region_dict = request.region.model_dump() if request.region else None
    if region_dict is not None:
        _canonical_region(region_dict)
    if region_dict is None and not request.scalar_filters:
        raise HTTPException(status_code=400, detail="filter requires `region` or `scalar_filters`.")

    with _cloud_session_lock:
        # keep-mask over the SURVIVING points: region (defaults all-True) AND
        # every scalar filter. Computed on survivors so a second filter composes
        # on the current point set, not the original.
        surv = ~sess.deleted
        pos = sess.positions[surv]
        keep = _region_mask(pos, region_dict) if region_dict is not None else np.ones(len(pos), dtype=bool)
        for f in (request.scalar_filters or []):
            slug = f.slug
            if slug not in sess.extras:
                raise HTTPException(status_code=400, detail=f"Unknown scalar attribute: {slug!r}. Available: {sorted(sess.extras)}")
            lo, hi, value_set = _resolve_scalar_filter(f.model_dump())
            keep &= _scalar_filter_mask(sess.extras[slug][surv], lo, hi, value_set)

        kept_count = int(keep.sum())
        total = int(len(sess.positions))
        # Empty result: do NOT commit the deletion or rebuild (PotreeConverter
        # can't ingest 0 points). The renderer raises a delete-confirmation.
        if kept_count == 0:
            return {"session_id": session_id, "point_count": 0, "rebuilt": False,
                    "remaining_count": int(surv.sum()), "total_count": total}

        # Commit: delete the filter-excluded survivors. Filter is presented as a
        # non-undoable commit (the renderer clears its erase undo stack), so we
        # reset the undo history to keep both sides in lock-step — a later
        # erase-undo must not reach back across this filter.
        idx_surv = np.where(surv)[0]
        sess.deleted[idx_surv[~keep]] = True
        sess.deleted_history = []
        sess.octree_cache_id = None
        remaining = int((~sess.deleted).sum())

    if not request.rebuild:
        return {"session_id": session_id, "remaining_count": remaining, "deleted_count": total - remaining, "total_count": total, "rebuilt": False}
    cache_key, cache_dir, meta = _session_rebuild(sess)
    return {"session_id": session_id, "point_count": remaining, "rebuilt": True, "cache_id": cache_key, "cache_dir": str(cache_dir), **meta}


@app.delete("/api/cloud/session/{session_id}")
async def delete_cloud_session(session_id: str):
    """Free a cloud session's in-RAM arrays."""
    with _cloud_session_lock:
        existed = _cloud_sessions.pop(session_id, None) is not None
    return {"session_id": session_id, "deleted": existed}

GROUND_CLASS_SLUG = "ground_class"
GROUND_CLASS_LABEL = "Ground Class"

# Height-above-ground (DEM-normalized elevation; canopy-height-model precursor).
# Written onto a cloud by the session DEM endpoint when add_height_column is set.
HEIGHT_ABOVE_GROUND_SLUG = "height_above_ground"
HEIGHT_ABOVE_GROUND_LABEL = "Height Above Ground"

def _load_cloud_for_segmentation(
    source_path: _Path, ascii_format: Optional[str],
) -> tuple[np.ndarray, dict, List[dict]]:
    """Load a cloud for tree segmentation, format-agnostically. Returns
    (xyz[N,3] float64, scalars, extra_dims) where:
      - `scalars` maps each carried extra-dim slug -> float32 array aligned to xyz
        (RGB/intensity are folded in as ordinary scalars here; the LAS writer
        re-maps r255/g255/b255 to RGB and intensity to the LAS intensity field).
      - `extra_dims` is the [{slug,label}, ...] list for the slug->label sidecar
        (matches the shape `_xyz_column_plan` / `_ply_to_las` produce).

    XYZ-family is read via pandas (honouring the Helios ascii_format); PLY via
    plyfile (carries every numeric vertex property — incl. benchmark `instance`/
    `semantic` labels); PCD via open3d (points/RGB only). Keeping this in one
    place lets segment_trees/apply accept the same formats the importer does.
    """
    ext = source_path.suffix.lower().lstrip(".")

    if ext in _PANDAS_EXTENSIONS:
        names, extra_dims = _xyz_column_plan(source_path, ascii_format)
        if not all(role in names for role in ("x", "y", "z")):
            raise HTTPException(
                status_code=400,
                detail=f"ASCII format must include x/y/z. Got columns: {names}",
            )
        skiprows = _ascii_skiprows(str(source_path))
        sep = _ascii_pandas_sep(str(source_path))
        df = pd.read_csv(
            source_path, sep=sep, header=None, names=names,
            usecols=[i for i, c in enumerate(names) if not _is_skip_name(c)],
            comment="#", skiprows=skiprows, engine="c",
        )
        xyz = np.column_stack([
            df["x"].to_numpy(dtype=np.float64),
            df["y"].to_numpy(dtype=np.float64),
            df["z"].to_numpy(dtype=np.float64),
        ])
        scalars: dict = {}
        if all(c in names for c in ("r255", "g255", "b255")):
            for c in ("r255", "g255", "b255"):
                scalars[c] = df[c].to_numpy(dtype=np.float32)
        ir = next((r for r in ("intensity", "reflectance") if r in names), None)
        if ir is not None:
            scalars["intensity"] = df[ir].to_numpy(dtype=np.float32)
        for ed in extra_dims:
            scalars[ed["slug"]] = df[ed["col"]].to_numpy(dtype=np.float32)
        return xyz, scalars, extra_dims

    if ext == "ply":
        from plyfile import PlyData
        try:
            vertex = PlyData.read(str(source_path))["vertex"].data
        except KeyError:
            raise HTTPException(
                status_code=400,
                detail="PLY has no 'vertex' element; cannot import as a point cloud.",
            )
        names = list(vertex.dtype.names or ())
        if not all(c in names for c in ("x", "y", "z")):
            raise HTTPException(
                status_code=400, detail=f"PLY missing x/y/z vertex properties. Got: {names}",
            )
        xyz = np.column_stack([
            vertex["x"].astype(np.float64),
            vertex["y"].astype(np.float64),
            vertex["z"].astype(np.float64),
        ])
        rgb = (("red", "green", "blue") if all(c in names for c in ("red", "green", "blue"))
               else ("r", "g", "b") if all(c in names for c in ("r", "g", "b")) else None)
        intensity_col = next(
            (c for c in ("intensity", "scalar_intensity", "reflectance") if c in names), None)
        reserved = {"x", "y", "z"}
        if rgb:
            reserved.update(rgb)
        if intensity_col:
            reserved.add(intensity_col)
        scalars = {}
        extra_dims = []
        used_slugs: set[str] = set()
        if rgb:
            scalars["r255"] = vertex[rgb[0]].astype(np.float32)
            scalars["g255"] = vertex[rgb[1]].astype(np.float32)
            scalars["b255"] = vertex[rgb[2]].astype(np.float32)
        if intensity_col is not None:
            scalars["intensity"] = vertex[intensity_col].astype(np.float32)
        for col in names:
            if col in reserved or not np.issubdtype(vertex[col].dtype, np.number):
                continue
            slug = _sanitize_extra_dim_name(col)
            base, i = slug, 1
            while slug in used_slugs:
                slug = f"{base[:29]}_{i}"; i += 1
            used_slugs.add(slug)
            extra_dims.append({"col": col, "slug": slug, "label": _humanize_extra_dim_label(col)})
            scalars[slug] = vertex[col].astype(np.float32)
        return xyz, scalars, extra_dims

    if ext == "pcd":
        positions, colors, _ = _load_ply_pcd_arrays(str(source_path))
        xyz = positions.astype(np.float64, copy=False)
        scalars = {}
        if colors is not None:
            # open3d colors are 0-1; store as 0-255 to match the r255 convention.
            scalars["r255"] = (colors[:, 0] * 255.0).astype(np.float32)
            scalars["g255"] = (colors[:, 1] * 255.0).astype(np.float32)
            scalars["b255"] = (colors[:, 2] * 255.0).astype(np.float32)
        return xyz, scalars, []

    raise HTTPException(
        status_code=400,
        detail=(f"segment_trees/apply: unsupported source extension .{ext}. "
                f"Supported: {sorted(_PANDAS_EXTENSIONS | _OPEN3D_EXTENSIONS)}"),
    )


# ==================== Cloud-to-Mesh Distance Comparison ====================

class C2MDistanceRequest(BaseModel):
    """Request for computing cloud-to-mesh distance statistics."""
    # Point cloud data (flat array: x,y,z,x,y,z,...). Optional because
    # octree-backed clouds send `source` instead.
    points: Optional[List[float]] = None
    # Read the cloud from a file on disk (octree-backed clouds).
    source: Optional[PointSource] = None
    # Mesh vertices (flat array: x,y,z,x,y,z,...)
    mesh_vertices: List[float]
    # Mesh triangle indices (flat array: i0,i1,i2,i0,i1,i2,...)
    mesh_indices: List[int]


class C2MDistanceResponse(BaseModel):
    """Response with cloud-to-mesh distance statistics."""
    success: bool
    error: Optional[str] = None
    # Statistics
    mean_distance: Optional[float] = None
    rmse: Optional[float] = None  # Root Mean Square Error
    std_deviation: Optional[float] = None
    min_distance: Optional[float] = None
    max_distance: Optional[float] = None
    median_distance: Optional[float] = None
    percentile_90: Optional[float] = None
    percentile_95: Optional[float] = None
    percentile_99: Optional[float] = None
    # Coverage metrics
    points_within_1mm: Optional[float] = None  # Percentage of points within 1mm
    points_within_5mm: Optional[float] = None  # Percentage of points within 5mm
    points_within_10mm: Optional[float] = None  # Percentage of points within 10mm
    # Raw data for visualization (optional, can be large)
    point_count: Optional[int] = None


@app.post("/api/c2m/distance", response_model=C2MDistanceResponse)
async def compute_c2m_distance(request: C2MDistanceRequest):
    """
    Compute Cloud-to-Mesh (C2M) distance statistics.

    Uses Open3D's RaycastingScene for efficient point-to-mesh distance computation.
    Returns comprehensive statistics about how well the mesh fits the point cloud.
    """
    try:
        import open3d as o3d
        import numpy as np

        # Convert flat arrays to numpy arrays. The source reader already returns
        # (N,3) — only the inline flat array needs reshaping.
        if request.source is not None:
            points, _, _ = _read_points_from_source(request.source)
        else:
            points = np.array(request.points or [], dtype=np.float64).reshape(-1, 3)
        vertices = np.array(request.mesh_vertices, dtype=np.float64).reshape(-1, 3)
        triangles = np.array(request.mesh_indices, dtype=np.int32).reshape(-1, 3)

        if len(points) == 0:
            return C2MDistanceResponse(
                success=False,
                error="No points provided"
            )

        if len(vertices) == 0 or len(triangles) == 0:
            return C2MDistanceResponse(
                success=False,
                error="No mesh data provided"
            )

        # Create Open3D triangle mesh
        mesh = o3d.t.geometry.TriangleMesh()
        mesh.vertex.positions = o3d.core.Tensor(vertices, dtype=o3d.core.float32)
        mesh.triangle.indices = o3d.core.Tensor(triangles, dtype=o3d.core.int32)

        # Create raycasting scene for efficient distance queries
        scene = o3d.t.geometry.RaycastingScene()
        scene.add_triangles(mesh)

        # Compute unsigned distances from each point to the mesh
        query_points = o3d.core.Tensor(points, dtype=o3d.core.float32)
        distances = scene.compute_distance(query_points).numpy()

        # Compute statistics
        mean_dist = float(np.mean(distances))
        rmse = float(np.sqrt(np.mean(distances ** 2)))
        std_dev = float(np.std(distances))
        min_dist = float(np.min(distances))
        max_dist = float(np.max(distances))
        median_dist = float(np.median(distances))
        p90 = float(np.percentile(distances, 90))
        p95 = float(np.percentile(distances, 95))
        p99 = float(np.percentile(distances, 99))

        # Coverage metrics (using adaptive thresholds based on data scale)
        # Use percentages of the bounding box diagonal as thresholds
        bbox_diag = np.linalg.norm(points.max(axis=0) - points.min(axis=0))
        thresh_1mm = bbox_diag * 0.001  # 0.1% of diagonal
        thresh_5mm = bbox_diag * 0.005  # 0.5% of diagonal
        thresh_10mm = bbox_diag * 0.01  # 1% of diagonal

        within_1mm = float(np.sum(distances <= thresh_1mm) / len(distances) * 100)
        within_5mm = float(np.sum(distances <= thresh_5mm) / len(distances) * 100)
        within_10mm = float(np.sum(distances <= thresh_10mm) / len(distances) * 100)

        return C2MDistanceResponse(
            success=True,
            mean_distance=mean_dist,
            rmse=rmse,
            std_deviation=std_dev,
            min_distance=min_dist,
            max_distance=max_dist,
            median_distance=median_dist,
            percentile_90=p90,
            percentile_95=p95,
            percentile_99=p99,
            points_within_1mm=within_1mm,
            points_within_5mm=within_5mm,
            points_within_10mm=within_10mm,
            point_count=len(points)
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return C2MDistanceResponse(
            success=False,
            error=f"C2M distance computation failed: {str(e)}"
        )


# ==================== ICP Registration (Snap to Fit) ====================

class ICPRegistrationRequest(BaseModel):
    """Request for ICP registration to align mesh to point cloud."""
    # Point cloud data (flat array: x,y,z,x,y,z,...) - the TARGET (stays fixed).
    # Optional because octree-backed clouds send `source` instead.
    points: Optional[List[float]] = None
    # Read the target cloud from a file on disk (octree-backed clouds).
    source: Optional[PointSource] = None
    # Mesh vertices (flat array: x,y,z,x,y,z,...) - the SOURCE (to be moved)
    mesh_vertices: List[float]
    # Mesh triangle indices (flat array: i0,i1,i2,i0,i1,i2,...)
    mesh_indices: List[int]
    # Maximum correspondence distance threshold
    max_correspondence_distance: Optional[float] = None
    # Maximum iterations per ICP round
    max_iterations: int = 100
    # RMSE convergence threshold (stop when RMSE improvement < this)
    rmse_threshold: float = 1e-6


class ICPRegistrationResponse(BaseModel):
    """Response with ICP registration transformation."""
    success: bool
    error: Optional[str] = None
    # Translation to apply to the mesh (dx, dy, dz) - includes center alignment + ICP
    translation: Optional[List[float]] = None
    # Full 4x4 transformation matrix (row-major, for advanced usage)
    transformation_matrix: Optional[List[float]] = None
    # Fitness score (0-1, higher is better)
    fitness: Optional[float] = None
    # RMSE after alignment
    rmse: Optional[float] = None
    # Number of ICP iterations performed
    iterations: Optional[int] = None


def run_icp_until_convergence(source_pcd, target_pcd, max_corr_dist, init_transform, max_iterations=100, rmse_threshold=1e-6):
    """
    Run ICP iteratively until RMSE plateaus (convergence).
    Returns the final transformation and metrics.
    """
    import open3d as o3d
    import numpy as np

    current_transform = init_transform.copy()
    prev_rmse = float('inf')
    total_iterations = 0

    # Run ICP in batches until convergence
    batch_size = 20  # iterations per batch
    max_batches = max_iterations // batch_size

    for batch in range(max_batches):
        reg_result = o3d.pipelines.registration.registration_icp(
            source_pcd,
            target_pcd,
            max_corr_dist,
            current_transform,
            o3d.pipelines.registration.TransformationEstimationPointToPlane(),
            o3d.pipelines.registration.ICPConvergenceCriteria(max_iteration=batch_size)
        )

        current_transform = reg_result.transformation
        current_rmse = reg_result.inlier_rmse
        total_iterations += batch_size

        # Check for convergence (RMSE plateau)
        rmse_improvement = prev_rmse - current_rmse
        if rmse_improvement < rmse_threshold and rmse_improvement >= 0:
            print(f"ICP converged after {total_iterations} iterations. RMSE: {current_rmse:.6f}")
            break

        prev_rmse = current_rmse

    return reg_result, total_iterations


@app.post("/api/c2m/icp-register", response_model=ICPRegistrationResponse)
async def icp_register_mesh_to_cloud(request: ICPRegistrationRequest):
    """
    Perform ICP (Iterative Closest Point) registration to align a mesh to a point cloud.

    The point cloud is the TARGET (stays fixed), the mesh is the SOURCE (will be transformed).
    Pre-aligns by moving source center to target center, then runs ICP until convergence.
    """
    try:
        import open3d as o3d
        import numpy as np

        # Convert flat arrays to numpy arrays. The source reader already returns
        # (N,3) — only the inline flat array needs reshaping.
        if request.source is not None:
            points, _, _ = _read_points_from_source(request.source)
        else:
            points = np.array(request.points or [], dtype=np.float64).reshape(-1, 3)
        vertices = np.array(request.mesh_vertices, dtype=np.float64).reshape(-1, 3)
        triangles = np.array(request.mesh_indices, dtype=np.int32).reshape(-1, 3)

        if len(points) == 0:
            return ICPRegistrationResponse(
                success=False,
                error="No points provided"
            )

        if len(vertices) == 0 or len(triangles) == 0:
            return ICPRegistrationResponse(
                success=False,
                error="No mesh data provided"
            )

        # Create point cloud (TARGET - stays fixed)
        target_pcd = o3d.geometry.PointCloud()
        target_pcd.points = o3d.utility.Vector3dVector(points)

        # Create mesh and sample points from it (SOURCE - to be moved)
        mesh = o3d.geometry.TriangleMesh()
        mesh.vertices = o3d.utility.Vector3dVector(vertices)
        mesh.triangles = o3d.utility.Vector3iVector(triangles)

        # Sample points from mesh surface for ICP
        num_samples = min(len(points), 50000)
        source_pcd = mesh.sample_points_uniformly(number_of_points=num_samples)

        # --- STEP 1: Center alignment (move source centroid to target centroid) ---
        target_center = target_pcd.get_center()
        source_center = source_pcd.get_center()
        center_offset = target_center - source_center

        # Apply center alignment to source
        source_pcd.translate(center_offset)

        print(f"Center alignment: moved source by [{center_offset[0]:.4f}, {center_offset[1]:.4f}, {center_offset[2]:.4f}]")

        # Estimate normals for point-to-plane ICP (more robust)
        # Use adaptive radius based on point cloud density
        bbox = target_pcd.get_axis_aligned_bounding_box()
        diagonal = np.linalg.norm(bbox.max_bound - bbox.min_bound)
        normal_radius = diagonal * 0.02  # 2% of diagonal

        target_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=normal_radius, max_nn=30))
        source_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=normal_radius, max_nn=30))

        # Determine max correspondence distance if not provided
        if request.max_correspondence_distance is None:
            max_corr_dist = diagonal * 0.05  # Start with 5% of diagonal for robustness
        else:
            max_corr_dist = request.max_correspondence_distance

        # --- STEP 2: Run ICP until convergence ---
        init_transform = np.eye(4)
        reg_result, iterations = run_icp_until_convergence(
            source_pcd, target_pcd, max_corr_dist, init_transform,
            max_iterations=request.max_iterations, rmse_threshold=request.rmse_threshold
        )

        # Combine center alignment with ICP transformation using proper matrix composition
        # The sequence is: first translate source to target center (T_center), then apply ICP (T_icp)
        # Combined transform = T_icp @ T_center
        icp_transform = reg_result.transformation

        # Create center alignment matrix
        T_center = np.eye(4)
        T_center[0, 3] = center_offset[0]
        T_center[1, 3] = center_offset[1]
        T_center[2, 3] = center_offset[2]

        # Compose transformations: T_icp @ T_center
        # This correctly handles rotation - ICP transform is applied AFTER center alignment
        combined_transform = icp_transform @ T_center

        # Extract translation for convenience (this is the total translation to apply to original source)
        translation = [float(combined_transform[0, 3]), float(combined_transform[1, 3]), float(combined_transform[2, 3])]
        transform_flat = combined_transform.flatten().tolist()

        print(f"ICP complete - fitness: {reg_result.fitness:.4f}, RMSE: {reg_result.inlier_rmse:.6f}, iterations: {iterations}")

        return ICPRegistrationResponse(
            success=True,
            translation=translation,
            transformation_matrix=transform_flat,
            fitness=float(reg_result.fitness),
            rmse=float(reg_result.inlier_rmse),
            iterations=iterations
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return ICPRegistrationResponse(
            success=False,
            error=f"ICP registration failed: {str(e)}"
        )


class CloudToCloudICPRequest(BaseModel):
    """Request for ICP registration to align one point cloud to another."""
    # Target point cloud (flat array: x,y,z,x,y,z,...) - stays fixed. Optional
    # because an octree-backed target sends `target_source` instead.
    target_points: Optional[List[float]] = None
    # Source point cloud (flat array: x,y,z,x,y,z,...) - will be transformed.
    # Optional because an octree-backed source sends `source_source` instead.
    source_points: Optional[List[float]] = None
    # Read either side from a file on disk (octree-backed clouds). Each side is
    # resolved independently, so a flat cloud and an octree cloud can be mixed.
    target_source: Optional[PointSource] = None
    source_source: Optional[PointSource] = None
    # Maximum correspondence distance threshold
    max_correspondence_distance: Optional[float] = None
    # Maximum iterations per ICP round
    max_iterations: int = 100
    # RMSE convergence threshold
    rmse_threshold: float = 1e-6


@app.post("/api/c2c/icp-register", response_model=ICPRegistrationResponse)
async def icp_register_cloud_to_cloud(request: CloudToCloudICPRequest):
    """
    Perform ICP (Iterative Closest Point) registration to align one point cloud to another.

    The target cloud stays fixed, the source cloud will be transformed.
    Pre-aligns by moving source center to target center, then runs ICP until convergence.
    """
    try:
        import open3d as o3d
        import numpy as np

        # Resolve each side independently — either may be a flat inline array
        # or an octree source descriptor. The source reader returns (N,3);
        # inline flat arrays need reshaping.
        if request.target_source is not None:
            target_points, _, _ = _read_points_from_source(request.target_source)
        else:
            target_points = np.array(request.target_points or [], dtype=np.float64).reshape(-1, 3)
        if request.source_source is not None:
            source_points, _, _ = _read_points_from_source(request.source_source)
        else:
            source_points = np.array(request.source_points or [], dtype=np.float64).reshape(-1, 3)

        if len(target_points) == 0:
            return ICPRegistrationResponse(
                success=False,
                error="No target points provided"
            )

        if len(source_points) == 0:
            return ICPRegistrationResponse(
                success=False,
                error="No source points provided"
            )

        # Create target point cloud (stays fixed)
        target_pcd = o3d.geometry.PointCloud()
        target_pcd.points = o3d.utility.Vector3dVector(target_points)

        # Create source point cloud (will be transformed)
        source_pcd = o3d.geometry.PointCloud()
        source_pcd.points = o3d.utility.Vector3dVector(source_points)

        # Optionally downsample if clouds are very large (for performance)
        max_points = 100000
        if len(target_points) > max_points:
            target_pcd = target_pcd.uniform_down_sample(every_k_points=max(1, len(target_points) // max_points))
        if len(source_points) > max_points:
            source_pcd = source_pcd.uniform_down_sample(every_k_points=max(1, len(source_points) // max_points))

        # --- STEP 1: Center alignment (move source centroid to target centroid) ---
        target_center = target_pcd.get_center()
        source_center = source_pcd.get_center()
        center_offset = target_center - source_center

        # Apply center alignment to source
        source_pcd.translate(center_offset)

        print(f"Center alignment: moved source by [{center_offset[0]:.4f}, {center_offset[1]:.4f}, {center_offset[2]:.4f}]")

        # Estimate normals for point-to-plane ICP (more robust)
        bbox = target_pcd.get_axis_aligned_bounding_box()
        diagonal = np.linalg.norm(bbox.max_bound - bbox.min_bound)
        normal_radius = diagonal * 0.02

        target_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=normal_radius, max_nn=30))
        source_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=normal_radius, max_nn=30))

        # Determine max correspondence distance if not provided
        if request.max_correspondence_distance is None:
            max_corr_dist = diagonal * 0.05
        else:
            max_corr_dist = request.max_correspondence_distance

        # --- STEP 2: Run ICP until convergence ---
        init_transform = np.eye(4)
        reg_result, iterations = run_icp_until_convergence(
            source_pcd, target_pcd, max_corr_dist, init_transform,
            max_iterations=request.max_iterations, rmse_threshold=request.rmse_threshold
        )

        # Combine center alignment with ICP transformation using proper matrix composition
        # The sequence is: first translate source to target center (T_center), then apply ICP (T_icp)
        # Combined transform = T_icp @ T_center
        icp_transform = reg_result.transformation

        # Create center alignment matrix
        T_center = np.eye(4)
        T_center[0, 3] = center_offset[0]
        T_center[1, 3] = center_offset[1]
        T_center[2, 3] = center_offset[2]

        # Compose transformations: T_icp @ T_center
        # This correctly handles rotation - ICP transform is applied AFTER center alignment
        combined_transform = icp_transform @ T_center

        # Extract translation for convenience (this is the total translation to apply to original source)
        translation = [float(combined_transform[0, 3]), float(combined_transform[1, 3]), float(combined_transform[2, 3])]
        transform_flat = combined_transform.flatten().tolist()

        print(f"Cloud-to-cloud ICP complete - fitness: {reg_result.fitness:.4f}, RMSE: {reg_result.inlier_rmse:.6f}, iterations: {iterations}")

        return ICPRegistrationResponse(
            success=True,
            translation=translation,
            transformation_matrix=transform_flat,
            fitness=float(reg_result.fitness),
            rmse=float(reg_result.inlier_rmse),
            iterations=iterations
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return ICPRegistrationResponse(
            success=False,
            error=f"Cloud-to-cloud ICP registration failed: {str(e)}"
        )


class MeshToMeshICPRequest(BaseModel):
    """Request for ICP registration to align one mesh to another."""
    # Target mesh (stays fixed)
    target_vertices: List[float]  # flat array: x,y,z,x,y,z,...
    target_indices: List[int]     # triangle indices: i,j,k,i,j,k,...
    # Source mesh (will be transformed)
    source_vertices: List[float]  # flat array: x,y,z,x,y,z,...
    source_indices: List[int]     # triangle indices: i,j,k,i,j,k,...
    # Maximum correspondence distance threshold
    max_correspondence_distance: Optional[float] = None
    # Maximum iterations per ICP round
    max_iterations: int = 100
    # RMSE convergence threshold
    rmse_threshold: float = 1e-6


@app.post("/api/m2m/icp-register", response_model=ICPRegistrationResponse)
async def icp_register_mesh_to_mesh(request: MeshToMeshICPRequest):
    """
    Perform ICP (Iterative Closest Point) registration to align one mesh to another.

    The target mesh stays fixed, the source mesh will be transformed.
    Pre-aligns by moving source center to target center, then runs ICP until convergence.
    """
    try:
        import open3d as o3d
        import numpy as np

        # Convert flat arrays to numpy arrays
        target_verts = np.array(request.target_vertices, dtype=np.float64).reshape(-1, 3)
        target_tris = np.array(request.target_indices, dtype=np.int32).reshape(-1, 3)
        source_verts = np.array(request.source_vertices, dtype=np.float64).reshape(-1, 3)
        source_tris = np.array(request.source_indices, dtype=np.int32).reshape(-1, 3)

        if len(target_verts) == 0 or len(target_tris) == 0:
            return ICPRegistrationResponse(
                success=False,
                error="No target mesh data provided"
            )

        if len(source_verts) == 0 or len(source_tris) == 0:
            return ICPRegistrationResponse(
                success=False,
                error="No source mesh data provided"
            )

        # Create target mesh and sample points
        target_mesh = o3d.geometry.TriangleMesh()
        target_mesh.vertices = o3d.utility.Vector3dVector(target_verts)
        target_mesh.triangles = o3d.utility.Vector3iVector(target_tris)

        # Create source mesh and sample points
        source_mesh = o3d.geometry.TriangleMesh()
        source_mesh.vertices = o3d.utility.Vector3dVector(source_verts)
        source_mesh.triangles = o3d.utility.Vector3iVector(source_tris)

        # Sample points from mesh surfaces for ICP
        num_samples = min(50000, max(len(target_verts), len(source_verts)) * 10)
        target_pcd = target_mesh.sample_points_uniformly(number_of_points=num_samples)
        source_pcd = source_mesh.sample_points_uniformly(number_of_points=num_samples)

        # --- STEP 1: Center alignment (move source centroid to target centroid) ---
        target_center = target_pcd.get_center()
        source_center = source_pcd.get_center()
        center_offset = target_center - source_center

        # Apply center alignment to source
        source_pcd.translate(center_offset)

        print(f"Mesh-to-mesh center alignment: moved source by [{center_offset[0]:.4f}, {center_offset[1]:.4f}, {center_offset[2]:.4f}]")

        # Estimate normals for point-to-plane ICP (more robust)
        bbox = target_pcd.get_axis_aligned_bounding_box()
        diagonal = np.linalg.norm(bbox.max_bound - bbox.min_bound)
        normal_radius = diagonal * 0.02

        target_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=normal_radius, max_nn=30))
        source_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=normal_radius, max_nn=30))

        # Determine max correspondence distance if not provided
        if request.max_correspondence_distance is None:
            max_corr_dist = diagonal * 0.05
        else:
            max_corr_dist = request.max_correspondence_distance

        # --- STEP 2: Run ICP until convergence ---
        init_transform = np.eye(4)
        reg_result, iterations = run_icp_until_convergence(
            source_pcd, target_pcd, max_corr_dist, init_transform,
            max_iterations=request.max_iterations, rmse_threshold=request.rmse_threshold
        )

        # Combine center alignment with ICP transformation using proper matrix composition
        # The sequence is: first translate source to target center (T_center), then apply ICP (T_icp)
        # Combined transform = T_icp @ T_center
        icp_transform = reg_result.transformation

        # Create center alignment matrix
        T_center = np.eye(4)
        T_center[0, 3] = center_offset[0]
        T_center[1, 3] = center_offset[1]
        T_center[2, 3] = center_offset[2]

        # Compose transformations: T_icp @ T_center
        # This correctly handles rotation - ICP transform is applied AFTER center alignment
        combined_transform = icp_transform @ T_center

        # Extract translation for convenience (this is the total translation to apply to original source)
        translation = [float(combined_transform[0, 3]), float(combined_transform[1, 3]), float(combined_transform[2, 3])]
        transform_flat = combined_transform.flatten().tolist()

        print(f"Mesh-to-mesh ICP complete - fitness: {reg_result.fitness:.4f}, RMSE: {reg_result.inlier_rmse:.6f}, iterations: {iterations}")

        return ICPRegistrationResponse(
            success=True,
            translation=translation,
            transformation_matrix=transform_flat,
            fitness=float(reg_result.fitness),
            rmse=float(reg_result.inlier_rmse),
            iterations=iterations
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return ICPRegistrationResponse(
            success=False,
            error=f"Mesh-to-mesh ICP registration failed: {str(e)}"
        )
